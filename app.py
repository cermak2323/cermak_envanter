# -*- coding: utf-8 -*-
from flask import Flask, render_template, request, jsonify, send_file, session, redirect, url_for

from flask_socketio import SocketIO, emit

from werkzeug.utils import secure_filename

from functools import wraps, lru_cache

import time

from collections import defaultdict

import sqlite3

import pandas as pd

import qrcode

from PIL import Image, ImageDraw, ImageFont

from io import BytesIO

import base64

import os

import uuid

from datetime import datetime, timedelta

import zipfile

import re

import hashlib

import secrets

import random

import string

from dotenv import load_dotenv

from contextlib import contextmanager

import unicodedata

import logging

# Takeuchi Modülü
from takeuchi_module import TakeuchiOrderManager

import threading

import json

# Yazıcı Entegrasyonu (Linux/Ubuntu için)
try:
    import sys
    if 'linux' in sys.platform.lower() or platform.system().lower() == 'linux':
        from printer_integration import get_printer_manager
        PRINTER_ENABLED = True
    else:
        PRINTER_ENABLED = False
except ImportError:
    PRINTER_ENABLED = False
except Exception:
    PRINTER_ENABLED = False


# SQLAlchemy ve Models - Veritabanı ORM

from models import db, PartCode, QRCode, CountSession, ScannedQR, User, CountPassword

from db_config import DevelopmentConfig, ProductionConfig



# Logging Configuration

from logging.handlers import RotatingFileHandler
import os
import sys

# Frozen exe için log yolunu belirle
if getattr(sys, 'frozen', False):
    # Running as exe - use AppData for user-specific logs
    LOG_DIR = os.path.join(os.environ.get('APPDATA', os.path.expanduser('~')), 'Cermak-Envanter', 'logs')
    os.makedirs(LOG_DIR, exist_ok=True)
else:
    # Running as script
    LOG_DIR = 'logs'
    os.makedirs(LOG_DIR, exist_ok=True)

# Loglama ayarları - RotatingFileHandler kullan (Windows'ta daha stabil)
app_log_handler = RotatingFileHandler(
    os.path.join(LOG_DIR, 'app.log'), 
    maxBytes=5*1024*1024,  # 5MB
    backupCount=3, 
    encoding='utf-8',
    delay=True  # Dosyayı hemen açma
)

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
    handlers=[
        app_log_handler,
        logging.StreamHandler()
    ]
)

# === SHARED FOLDER CONFIGURATION ===
import platform
import os
from dotenv import load_dotenv

# .env dosyasını yükle (SESSION_SECRET için)
load_dotenv()

if platform.system() == 'Linux':
    # NFS mount: sudo mount -t nfs 192.168.0.57:/tahsinortak/CermakDepo/CermakEnvanter/static /mnt/ortakdepo
    SHARED_BASE = '/mnt/ortakdepo'
else:
    SHARED_BASE = r'\\DCSRV\tahsinortak\CermakDepo\CermakEnvanter\static'

STATIC_DIR = SHARED_BASE

print(f"[INFO] Paylaşımlı klasör yolu: {STATIC_DIR}")

# Flask app'i oluştur – AMA static_folder olmadan!
app = Flask(__name__)

# SECRET_KEY'i HEMEN tanımla (.env'den al veya fallback)
app.config['SECRET_KEY'] = os.environ['SESSION_SECRET']

# Diğer config'ler
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024
app.config['SESSION_COOKIE_HTTPONLY'] = True
app.config['SESSION_COOKIE_SECURE'] = False
app.config['SESSION_COOKIE_SAMESITE'] = 'Lax'
app.config['PERMANENT_SESSION_LIFETIME'] = 365 * 24 * 60 * 60  # 1 yıl

# Static folder'ı sonradan ayarla
app.static_folder = STATIC_DIR
print(f"[FLASK] Static folder ayarlandı: {STATIC_DIR}")

# Yazıcı Yöneticisini Başlat (Linux/Ubuntu için)
_printer_manager = None
if PRINTER_ENABLED:
    try:
        _printer_manager = get_printer_manager()
        printer_status = _printer_manager.get_status()
        if printer_status.get('connected'):
            print(f"[PRINTER] ✓ USB Yazıcı Hazır - {printer_status.get('device', '/dev/usb/lp0')}")
        else:
            print(f"[PRINTER] ⚠️  USB Yazıcı Bağlanamadı - Yazıcı devre dışı")
    except Exception as e:
        print(f"[PRINTER] ⚠️  Yazıcı başlatılamadı: {e}")
        _printer_manager = None
else:
    print("[PRINTER] Yazıcı devre dışı (Windows ortamında çalışıyor)")

# Ek API blueprint (güncelleme kontrolü + login basitleştirilmiş API)
try:
    from backend.api_blueprint import bp as api_ext_bp  # backend/api_blueprint.py
    app.register_blueprint(api_ext_bp)
    print("[API] Extended blueprint yüklendi (/api/*)")
except Exception as _bp_err:
    print(f"[API] Blueprint yüklenemedi: {_bp_err}")

# APScheduler entegrasyonu (periyodik senkronizasyon)
try:
    from backend.scheduler import init_scheduler
    scheduler = init_scheduler(app)
except Exception as _sched_err:
    print(f"[SCHED] Scheduler başlatılamadı: {_sched_err}")

# Paylaşımlı klasör senkronizasyonu
def get_sync_service():
    """Paylaşımlı klasör senkronizasyon yöneticisini döndürür veya None."""
    try:
        from shared_folder_sync import SharedFolderSyncManager
        return SharedFolderSyncManager()
    except Exception:
        return None

# Manuel senkron tetik endpoint (herkes tetikleyebilir, istersen admin-only yapabilirsin)
from flask import jsonify
@app.route('/api/sync/trigger', methods=['POST'])
def manual_sync():
    try:
        from shared_folder_sync import SharedFolderSyncManager
        mgr = SharedFolderSyncManager()
        stats = mgr.full_sync('both')
        return jsonify({'success': True, 'stats': stats})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500



#  MYSQL SİSTEM - Şirket içi MySQL Sunucusu

IS_PRODUCTION = False

IS_LOCAL = True

USE_MYSQL = True



print(f"\n[MYSQL] MYSQL SISTEM")

print(f"[DB] Database: MySQL (192.168.0.57:3306)")

print(f"[STORAGE] Storage: Shared Folder (\\\\DCSRV\\tahsinortak\\CermakDepo\\CermakEnvanter\\static)")

print(f"[DATA] Data: MySQL veritabaninda")



# db_config.py kullan

app.config.from_object(DevelopmentConfig)



print("\n" + "="*60)

print(" MYSQL MODU")

print("="*60)

print(" Tüm veriler MySQL veritabanında")

if platform.system() == 'Linux':
    print(" QR kodlar shared folder'da (/mnt/ortakdepo/qr_codes)")
else:
    print(" QR kodlar shared folder'da (\\\\DCSRV\\...\\qr_codes)")

print(" Excel/Raporlar shared folder'da (\\\\DCSRV\\...\\reports)")

print("="*60)

print()



# SQLAlchemy'yi app'e bala

db.init_app(app)



# Uygulama balang zaman (uptime iin)

app.config['START_TIME'] = time.time()



# ? ULTRA FAST QR SCANNING OPTIMIZATION
# In-memory cache for instant QR lookup (milliseconds)
QR_LOOKUP_CACHE = {}  # {qr_id: {'part_code': str, 'part_name': str, 'part_code_id': int}}
PART_CODE_CACHE = {}  # {part_code: {'part_name': str, 'id': int}}
CACHE_LOADED = False
CACHE_LOCK = threading.Lock()
CACHE_LOAD_TIME = None

# Session stats cache (güncellemeleri hızlandırır)
SESSION_STATS_CACHE = {}  # {session_id: {'total': int, 'last_qr': str}}

# 🚀 ACTIVE SESSION CACHE (3845 parçalık JSON'u her 3 saniyede parse etmemek için!)
ACTIVE_SESSION_CACHE = {
    'session_id': None,
    'expected_parts': [],
    'cache_time': None,
    'ttl': None  # ⭐ SINIRSIZ - Sayım bitene kadar cache'de kalır, bittiğinde temizlenir
}

# Batch processing için buffer
SCAN_BUFFER = {}  # {session_id: [qr_ids]}
BUFFER_MAX_SIZE = 10
BUFFER_TIMEOUT = 0.5  # 500ms



# Static dosya sktrma iin
app.config['SEND_FILE_MAX_AGE_DEFAULT'] = 31536000  # 1 yl cache

# SocketIO - Canlı takip için gerekli
socketio = SocketIO(
    app, 
    cors_allowed_origins="*",
    async_mode='threading',
    ping_timeout=60,
    ping_interval=25,
    logger=False,
    engineio_logger=False
)

# ======================
# PERFORMANS OPTIMIZASYONLARI
# ======================



@app.before_request

def qr_folder_security():

    """QR klasör güvenlik kontrolü - SHARED FOLDER kullanıldığı için artık gerekli değil"""
    # NOTE: Dosyalar artık shared folder'da
    # Windows: \\DCSRV\tahsinortak\CermakDepo\CermakEnvanter\static\qr_codes
    # Linux: /mnt/ortakdepo/qr_codes
    # Flask tarafından /static/qr_codes/ URL'si üzerinden erişilebilir
    pass



@app.after_request

def add_performance_headers(response):

    """Performans iin header'lar ekle"""

    # Static dosyalar iin cache

    if request.endpoint == 'static':

        response.cache_control.max_age = 31536000  # 1 yl

        response.cache_control.public = True



    # Dier dosyalar iin

    else:

        response.cache_control.no_cache = True

        response.cache_control.must_revalidate = True



    # Sktrma header'

    if response.status_code == 200 and response.content_length and response.content_length > 1024:

        response.headers['Vary'] = 'Accept-Encoding'



    return response



# ======================

# PERFORMANS CACHE SISTEMI

# ======================



# Bellek tabanl cache (production'da Redis kullanlmal)

cache_store = {}

cache_lock = threading.Lock()

CACHE_TTL = 300  # 5 dakika cache sresi



def cache_get(key):

    """Cache'den veri al"""

    with cache_lock:

        if key in cache_store:

            data, timestamp = cache_store[key]

            if time.time() - timestamp < CACHE_TTL:

                return data

            else:

                del cache_store[key]

    return None



def cache_set(key, value):

    """Cache'e veri kaydet"""

    with cache_lock:

        cache_store[key] = (value, time.time())



def cache_delete(key):

    """Cache'den veri sil"""

    with cache_lock:

        if key in cache_store:

            del cache_store[key]



def cache_clear():

    """Tm cache'i temizle"""

    with cache_lock:

        cache_store.clear()



#  QR Image Memory Cache - LRU ile disk I/O azaltma

@lru_cache(maxsize=1000)

def generate_qr_pil_image(qr_id, box_size=8, border=2):

    """

    QR kod PIL Image olutur - Barkod makinesi iin optimize

    

    Scanner Specs:

    - Minimum znrlk 2D: 8.7mil (0.22mm)

    - Okuma Mesafesi: 55-350mm

    - Sensr: 640x480 piksel

    

    QR Boyutlar:

    - box_size=8 -> ~240x240px QR (8.7mil znrlk iin ideal)

    - border=2 -> Minimal ereve (tarayc quiet zone iin)

    - Toplam: ~255x275px (ereve + text ile)

    

    Args:

        qr_id: QR kod ID

        box_size: QR kutucuk boyutu (8 = barkod makinesi iin optimize)

        border: Quiet zone (2 = minimal, tarayc iin yeterli)

    

    Returns:

        PIL.Image: QR kod grseli (krmz ereveli)

    """

    qr = qrcode.QRCode(

        version=1,

        box_size=box_size,  # 8px - barkod makinesi iin optimize

        border=border,      # 2px quiet zone - tarayc iin minimum

        error_correction=qrcode.constants.ERROR_CORRECT_M  # M seviyesi (15% hata tolerans)

    )

    qr.add_data(qr_id)

    qr.make(fit=True)

    qr_img = qr.make_image(fill_color="black", back_color="white")

    qr_img = qr_img.convert('RGB')

    

    # QR boyutlar

    qr_width, qr_height = qr_img.size

    text_height = 20  # Text alan (14pt font iin)

    border_width = 3  # Krmz ereve

    

    # Krmz ereveli grsel

    final_width = qr_width + (border_width * 2)

    final_height = qr_height + text_height + (border_width * 2)

    final_img = Image.new('RGB', (final_width, final_height), '#dc2626')

    

    # Beyaz i alan

    white_bg = Image.new('RGB', (qr_width, qr_height + text_height), 'white')

    white_bg.paste(qr_img, (0, 0))

    final_img.paste(white_bg, (border_width, border_width))

    

    # Text ekle

    draw = ImageDraw.Draw(final_img)

    try:

        font = ImageFont.truetype("arialbd.ttf", 14)

    except:

        try:

            font = ImageFont.truetype("arial.ttf", 14)

        except:

            font = ImageFont.load_default()

    

    qr_text = qr_id

    text_width = len(qr_text) * 8

    x_position = max(border_width, (final_width - text_width) // 2)

    draw.text((x_position, qr_height + border_width + 2), qr_text, fill='black', font=font)

    

    return final_img



# Cache temizleme thread'i

def cache_cleanup():

    """Eski cache verilerini temizle"""

    while True:

        time.sleep(60)  # Her dakika kontrol et

        current_time = time.time()

        with cache_lock:

            expired_keys = [

                key for key, (_, timestamp) in cache_store.items()

                if current_time - timestamp > CACHE_TTL

            ]

            for key in expired_keys:

                del cache_store[key]



# Cache temizleme thread'ini balat

cleanup_thread = threading.Thread(target=cache_cleanup, daemon=True)

cleanup_thread.start()



# Rate limiting iin IP tabanl takip

login_attempts = defaultdict(list)



def add_security_headers(response):

    """Gvenlik header'larn ekle"""

    response.headers['X-Content-Type-Options'] = 'nosniff'

    response.headers['X-Frame-Options'] = 'DENY'

    response.headers['X-XSS-Protection'] = '1; mode=block'

    response.headers['Strict-Transport-Security'] = 'max-age=31536000; includeSubDomains'

    response.headers['Content-Security-Policy'] = "default-src 'self'; script-src 'self' 'unsafe-inline' https://cdn.jsdelivr.net https://cdn.socket.io https://unpkg.com; style-src 'self' 'unsafe-inline' https://cdn.jsdelivr.net https://fonts.googleapis.com; font-src 'self' https://cdn.jsdelivr.net https://fonts.gstatic.com; img-src 'self' data: blob:; media-src 'self' data: blob:; connect-src 'self' ws: wss: https://cdn.jsdelivr.net;"

    return response



@app.after_request

def security_headers(response):

    return add_security_headers(response)



def rate_limit_login(f):

    """Login denemelerini snrla"""

    @wraps(f)

    def decorated_function(*args, **kwargs):

        client_ip = request.environ.get('HTTP_X_FORWARDED_FOR', request.environ.get('REMOTE_ADDR', '127.0.0.1'))

        current_time = time.time()



        # Son 15 dakikadaki denemeleri filtrele

        login_attempts[client_ip] = [t for t in login_attempts[client_ip] if current_time - t < 900]



        # 15 dakikada 50'den fazla deneme varsa engelle (test için artırıldı)

        if len(login_attempts[client_ip]) >= 50:

            return jsonify({'error': 'ok fazla login denemesi. 15 dakika bekleyin.'}), 429



        # Denemeyi kaydet

        login_attempts[client_ip].append(current_time)



        return f(*args, **kwargs)

    return decorated_function



# ----------------------

# Configuration

# ----------------------



# Project root

BASE_DIR = os.path.abspath(os.path.dirname(__file__))



# ======================

# VERTABANI YAPILANDI - SADECE MySQL

# ======================



# MySQL database

DATABASE_URL = app.config['SQLALCHEMY_DATABASE_URI']

print(f"[MYSQL] MySQL Sunucu: {DATABASE_URL}")



# ========== SHARED FOLDER PATHS ==========
# Tüm dosyalar paylaşımlı klasöre kaydedilir - FALLBACK YOK!
# SHARED_BASE yukarıda (satır ~107) tanımlı
REPORTS_DIR = os.path.join(SHARED_BASE, 'reports')
QR_CODES_DIR = os.path.join(SHARED_BASE, 'qr_codes')
PART_PHOTOS_DIR = os.path.join(SHARED_BASE, 'part_photos')
# ==========================================


def verify_shared_folder_access():
    """
    Paylaşımlı klasöre erişimi kontrol et ve izinleri doğrula
    Linux'ta /mnt/ortakdepo erişilebilir ve yazılabilir olmalı
    """
    try:
        if not os.path.exists(SHARED_BASE):
            print(f"[ERROR] Paylaşımlı klasör bulunamadı: {SHARED_BASE}")
            print(f"[WARN] Linux: '/mnt/ortakdepo' mount edilmiş olmalı")
            return False
        
        # Yazma testi
        test_file = os.path.join(SHARED_BASE, '.write_test')
        try:
            with open(test_file, 'w') as f:
                f.write('test')
            os.remove(test_file)
            print(f"[OK] Paylaşımlı klasör erişilebilir: {SHARED_BASE}")
            return True
        except PermissionError:
            print(f"[ERROR] Paylaşımlı klasöre yazma izni yok: {SHARED_BASE}")
            print(f"[WARN] Linux'ta: sudo mount -t nfs 192.168.0.57:/tahsinortak/CermakDepo/CermakEnvanter/static /mnt/ortakdepo")
            return False
    except Exception as e:
        print(f"[ERROR] Paylaşımlı klasör kontrol hatası: {e}")
        return False




def generate_strong_password():

    """Gl parola olutur (8 karakter: byk harf, kk harf, rakam, zel karakter)"""

    characters = string.ascii_uppercase + string.ascii_lowercase + string.digits + "!@#$%^&*"

    # En az 1 byk harf, 1 kk harf, 1 rakam, 1 zel karakter olacak ekilde

    password = [

        random.choice(string.ascii_uppercase),  # Byk harf

        random.choice(string.ascii_lowercase),  # Kk harf

        random.choice(string.digits),           # Rakam

        random.choice("!@#$%^&*")              # zel karakter

    ]

    # Kalan 4 karakteri rastgele se

    for _ in range(4):

        password.append(random.choice(characters))



    # Kartr

    random.shuffle(password)

    return ''.join(password)



def generate_count_password():

    """Saym iin parola olutur (6 haneli sadece say) - Basit ve hzl giri iin"""

    return ''.join([str(random.randint(0, 9)) for _ in range(6)])



def save_qr_code_to_file(part_code, qr_id, qr_number):

    """

    QR kodunu fiziksel dosya olarak kaydet

    - Her para iin ayr klasr oluturur (rn: static/qr_codes/Y129513-14532/)

    - QR kodlarn numaral kaydeder (rn: Y129513-14532_1.png)

    - QR kod altna kod ve para ad yazar

    """

    try:

        import qrcode

        import platform as sys_platform

        

        # Para adn database'den al (ORM)
        part_obj = db.session.query(PartCode).filter_by(part_code=part_code).first()
        part_name = part_obj.part_name if part_obj else ""
        
        # QR klasörü ana dizini - SHARED FOLDER KULLAN
        qrcodes_base_dir = QR_CODES_DIR
        
        # Parça için klasör oluştur (örn: \\DCSRV\...\qr_codes\Y129513-14532/)
        part_dir = os.path.join(qrcodes_base_dir, part_code)
        os.makedirs(part_dir, exist_ok=True)

        

        # Dosya ad: part_code_number.png (rn: Y129513-14532_1.png)

        filename = f"{part_code}_{qr_number}.png"

        filepath = os.path.join(part_dir, filename)
        
        # FONT BULUCU - Windows ve Linux'ta aynı kalınlıkta font bulsun
        def get_font(size=32):
            """Windows ve Linux'ta uyumlu kalın font bul"""
            if sys_platform.system() == 'Windows':
                font_paths = [
                    "arialblk.ttf",      # Arial Black (en kalın)
                    "arialbd.ttf",       # Arial Bold
                    "arial.ttf",         # Arial
                ]
            else:  # Linux
                font_paths = [
                    "/usr/share/fonts/truetype/liberation/LiberationSans-Bold.ttf",
                    "/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf",
                    "/usr/share/fonts/truetype/ubuntu/Ubuntu-B.ttf",  # Ubuntu font
                    "arialblk.ttf",      # fallback
                ]
            
            for font_path in font_paths:
                try:
                    return ImageFont.truetype(font_path, size)
                except:
                    continue
            
            return ImageFont.load_default()

        

        # QR kod olutur - Barkod makinesi iin optimize (8.7mil znrlk)

        qr = qrcode.QRCode(

            version=1,

            error_correction=qrcode.constants.ERROR_CORRECT_M,  # M seviyesi (15% hata tolerans)

            box_size=8,  # 8px - barkod makinesi iin ideal (~240x240px)

            border=2,    # 2px quiet zone - tarayc minimum gereksinimi

        )

        qr.add_data(qr_id)

        qr.make(fit=True)

        

        # QR kodunu olutur

        qr_img = qr.make_image(fill_color="black", back_color="white")

        qr_img = qr_img.convert('RGB')  # PIL Image'a dntr

        

        # QR kod boyutlarn al

        qr_width, qr_height = qr_img.size

        

        # Alanlar hesapla

        logo_height = 40  # CERMAK yazısı iin st alan

        text_height = 35  # Alt yaz (para numaras) iin alan

        

        # Krmz ereve iin padding

        border_width = 3  # 3px krmz ereve

        

        # Yeni grsel olutur (CERMAK + QR + text alan + ereve)

        final_width = qr_width + (border_width * 2)

        final_height = logo_height + qr_height + text_height + (border_width * 2)

        final_img = Image.new('RGB', (final_width, final_height), '#dc2626')  # Krmz arka plan (ereve)

        

        # Beyaz i alan olutur (CERMAK + QR + text)

        white_bg = Image.new('RGB', (qr_width, logo_height + qr_height + text_height), 'white')

        

        # CERMAK yazısı ekle - st ortasna

        try:

            cermak_font = get_font(32)  # 32pt font CERMAK için
            
            draw_temp = ImageDraw.Draw(white_bg)

            cermak_text = "CERMAK"

            # Gerçek text genişliğini ölç

            bbox = draw_temp.textbbox((0, 0), cermak_text, font=cermak_font)

            text_width = bbox[2] - bbox[0]

            x_pos = (qr_width - text_width) // 2

            y_pos = 5

            draw_temp.text((x_pos, y_pos), cermak_text, fill='black', font=cermak_font)

        except Exception as e:

            print(f"CERMAK yazısı ekleme hatas: {e}")

        

        # QR kodu beyaz alana yaptr (logo'nun altna)

        white_bg.paste(qr_img, (0, logo_height))

        

        # Beyaz alan krmz erevenin iine yaptr

        final_img.paste(white_bg, (border_width, border_width))

        

        # Text ekleme iin draw nesnesi

        draw = ImageDraw.Draw(final_img)

        

        # Font (kaln ve byk)

        font = get_font(24)  # 24pt font parça kodu için

        

        # QR ID yazs (Y129150-49811_1) - Sadece bu

        qr_text = f"{part_code}_{qr_number}"

        

        # Gerçek text genişliğini hesapla (textbbox ile)

        bbox = draw.textbbox((0, 0), qr_text, font=font)

        text_width = bbox[2] - bbox[0]

        

        # QR ID'yi ortala ve yaz (border_width offset ekle)

        x_position = (final_width - text_width) // 2

        draw.text((x_position, logo_height + qr_height + border_width + 5), qr_text, fill='black', font=font)

        

        # Kaydet

        final_img.save(filepath)

        

        print(f" [OK] QR kod kaydedildi: {filepath}")

        return filepath

        

    except Exception as e:

        print(f" QR kod kaydetme hatas: {e}")

        return None



# Admin saym ifresi

ADMIN_COUNT_PASSWORD = "@R9t$L7e!xP2w"

print(f"DEBUG: ADMIN_COUNT_PASSWORD = '{ADMIN_COUNT_PASSWORD}'")  # DEBUG



os.makedirs(REPORTS_DIR, exist_ok=True)



# Dosya upload iin ayarlar
# UNC path - paylaşımlı klasör (FALLBACK YOK!)
UPLOAD_FOLDER = PART_PHOTOS_DIR
PROFILE_PHOTOS_DIR = os.path.join(UPLOAD_FOLDER, 'profiles')
os.makedirs(PROFILE_PHOTOS_DIR, exist_ok=True)
ALLOWED_EXTENSIONS = {'png', 'jpg', 'jpeg', 'gif', 'webp'}
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024  # 16MB limit



def allowed_file(filename):
    """Dosya uzantsn kontrol et"""
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

def get_static_path(subpath):
    """Static dosya yolu döndürür - HER ZAMAN UNC PATH KULLANIR"""
    return os.path.join(SHARED_BASE, subpath)

def get_db_placeholder():

    """Database'e gre doru placeholder dndr - SQLite iin  kullanr"""

    return ''



def execute_query(cursor, query, params=None):
    """
    Execute SQL query via pymysql (MySQL)
    
    Uses %s placeholders for MySQL parameterized queries
    """
    try:
        cursor.execute(query, params or ())
        return cursor
    
    except Exception as e:
        logging.error(f"MySQL execute_query error: {e}")
        
        class EmptyCursor:
            def fetchone(self):
                return None
            def fetchall(self):
                return []
        return EmptyCursor()



def get_db():
    """Get MySQL database connection from connection pool"""
    import pymysql
    import pymysql.cursors
    import os
    from dotenv import load_dotenv
    
    load_dotenv()
    
    # Şirket içi MySQL bağlantı bilgileri
    mysql_config = {
        'host': '192.168.0.57',
        'user': 'flaskuser',
        'password': 'FlaskSifre123!',
        'database': 'flaskdb',
        'port': 3306,
        'charset': 'utf8mb4',
        'cursorclass': pymysql.cursors.Cursor,
        'autocommit': False
    }
    
    try:
        # Yeni MySQL bağlantısı oluştur
        conn = pymysql.connect(**mysql_config)
        
        return conn
    except Exception as e:
        logging.error(f"MySQL connection error: {e}")
        import pymysql
        return pymysql.connect(**mysql_config)


def close_db(conn):
    """Close MySQL connection"""
    if conn:
        try:
            conn.close()
        except Exception as e:
            logging.error(f"MySQL close error: {e}")
            logging.debug(f"Error closing PostgreSQL connection: {e}")


def load_qr_cache_to_memory():
    """
    ? CRITICAL OPTIMIZATION: Tüm QR kodlarını belleğe yükle
    Bu fonksiyon app startup'ta çalışır - QR lookup'ı 50ms'den 0.1ms'ye düşürür
    """
    global QR_LOOKUP_CACHE, PART_CODE_CACHE, CACHE_LOADED, CACHE_LOAD_TIME
    
    if CACHE_LOADED:
        return True
    
    try:
        with CACHE_LOCK:
            start_time = time.time()
            conn = get_db()
            cursor = conn.cursor()
            
            # Tüm QR kodlarını tek sorguda çek
            cursor.execute('''
                SELECT qc.id, pc.part_code, pc.part_name, pc.id
                FROM qr_codes qc
                INNER JOIN part_codes pc ON qc.part_code_id = pc.id
            ''')
            
            QR_LOOKUP_CACHE.clear()
            for row in cursor.fetchall():
                qr_id, part_code, part_name, part_code_id = row
                QR_LOOKUP_CACHE[qr_id] = {
                    'part_code': part_code,
                    'part_name': part_name,
                    'part_code_id': part_code_id
                }
            
            # Part codes da cache'e al
            cursor.execute('SELECT part_code, part_name, id FROM part_codes')
            PART_CODE_CACHE.clear()
            for row in cursor.fetchall():
                part_code, part_name, pc_id = row
                PART_CODE_CACHE[part_code] = {
                    'part_name': part_name,
                    'id': pc_id
                }
            
            close_db(conn)
            
            load_time = (time.time() - start_time) * 1000
            CACHE_LOADED = True
            CACHE_LOAD_TIME = datetime.now()
            
            logging.info(f"[CACHE] QR Cache loaded: {len(QR_LOOKUP_CACHE)} QR codes, {len(PART_CODE_CACHE)} parts in {load_time:.1f}ms")
            return True
            
    except Exception as e:
        logging.error(f"Cache loading error: {e}")
        CACHE_LOADED = False
        return False


def get_qr_from_cache(qr_id):
    """? Bellekten ultra hızlı QR bilgisi - ~0.1ms"""
    if not CACHE_LOADED:
        load_qr_cache_to_memory()
    
    # Direkt cache lookup
    if qr_id in QR_LOOKUP_CACHE:
        return QR_LOOKUP_CACHE[qr_id]
    
    # Part code olarak dene (QR = part_code durumu)
    if qr_id in PART_CODE_CACHE:
        return {
            'part_code': qr_id,
            'part_name': PART_CODE_CACHE[qr_id]['part_name'],
            'part_code_id': PART_CODE_CACHE[qr_id]['id']
        }
    
    return None


def reload_cache():
    """Cache'i yeniden yükle (yeni QR eklendiğinde)"""
    global CACHE_LOADED
    CACHE_LOADED = False
    return load_qr_cache_to_memory()


def fix_sequence(table_name, id_column='id'):
    """
    PostgreSQL sequence'ını düzelt - id%slarını önler
    """
    try:
        conn = get_db()
        cursor = conn.cursor()
        
        # Sequence adını bul (genelde tablename_id_seq)
        sequence_name = f"{table_name}_{id_column}_seq"
        
        # En yüksek id'yi al
        cursor.execute(f"SELECT MAX({id_column}) FROM {table_name}")
        max_id = cursor.fetchone()[0]
        
        if max_id is not None:
            # Sequence'ı max_id + 1'e set et
            cursor.execute(f"SELECT setval('{sequence_name}', %s, true)", (max_id,))
            conn.commit()
            logging.info(f"[OK] {table_name} sequence fixed: set to {max_id + 1}")
        
        close_db(conn)
        return True
    except Exception as e:
        logging.error(f"Sequence fix error for {table_name}: {e}")
        return False


@contextmanager

def db_connection():

    """

    Database connection context manager - Otomatik cleanup

    

    Kullanm:

        with db_connection() as conn:

            cursor = conn.cursor()

            execute_query(cursor, 'SELECT ...')

            conn.commit()

        # conn otomatik kapanr, hata olsa bile

    """

    conn = get_db()

    try:

        yield conn

    except sqlite3.Error as e:

        logging.error(f"Database error: {e}")

        if conn:

            conn.rollback()

        raise

    except Exception as e:

        logging.error(f"Unexpected error in database operation: {e}")

        if conn:

            conn.rollback()

        raise

    finally:

        close_db(conn)



def create_performance_indexes(cursor):

    """Performans iin kritik indexleri olutur"""

    indexes = [

        # QR Codes indexleri

        "CREATE INDEX IF NOT EXISTS idx_qr_codes_qr_id ON qr_codes(qr_id);",

        "CREATE INDEX IF NOT EXISTS idx_qr_codes_part_code ON qr_codes(part_code_id);",

        "CREATE INDEX IF NOT EXISTS idx_qr_codes_is_used ON qr_codes(is_used);",



        # Count Sessions indexleri

        "CREATE INDEX IF NOT EXISTS idx_count_sessions_is_active ON count_sessions(is_active);",

        "CREATE INDEX IF NOT EXISTS idx_count_sessions_created_by ON count_sessions(created_by);",

        "CREATE INDEX IF NOT EXISTS idx_count_sessions_created_at ON count_sessions(created_at);",



        # Scanned QR indexleri - COMPOSITE INDEX N NEML

        "CREATE INDEX IF NOT EXISTS idx_scanned_qr_session_id ON scanned_qr(session_id);",

        "CREATE INDEX IF NOT EXISTS idx_scanned_qr_qr_id ON scanned_qr(qr_id);",

        "CREATE INDEX IF NOT EXISTS idx_scanned_qr_scanned_by ON scanned_qr(scanned_by);",

        "CREATE INDEX IF NOT EXISTS idx_scanned_qr_scanned_at ON scanned_qr(scanned_at);",

        

        #  COMPOSITE INDEX - Duplicate check iin kritik

        "CREATE INDEX IF NOT EXISTS idx_scanned_qr_session_qr ON scanned_qr(session_id, qr_id);",



        # Part Codes indexleri

        "CREATE INDEX IF NOT EXISTS idx_part_codes_part_code ON part_codes(part_code);",



        # Kullanc indexleri

        "CREATE INDEX IF NOT EXISTS idx_envanter_users_username ON envanter_users(username);",

        "CREATE INDEX IF NOT EXISTS idx_envanter_users_is_active ON envanter_users(is_active_user);",

    ]



    try:

        for index_sql in indexes:

            execute_query(cursor, index_sql)

        print(" Performance indexes created/verified")

    except Exception as e:

        print(f" Warning: Could not create some indexes: {e}")



def init_db_part_details():

    """Part codes tablosuna detay kolonlar ekle"""

    try:

        conn = get_db()

        cursor = conn.cursor()

        

        # Yeni kolonlar - Geriye uyumlu (ALTER TABLE)

        new_columns = [

            ('photo_path', 'TEXT'),           # Para fotoraf

            ('catalog_image', 'TEXT'),        # Katalog grseli

            ('description', 'TEXT'),          # Aklama

            ('used_in_machines', 'TEXT'),     # JSON: Kullanld makineler

            ('specifications', 'TEXT'),       # JSON: Teknik zellikler

            ('stock_location', 'TEXT'),       # Stok konumu

            ('supplier', 'TEXT'),             # Tedariki

            ('unit_price', 'REAL'),          # Birim fiyat

            ('notes', 'TEXT')                # Notlar

        ]

        

        for col_name, col_type in new_columns:

            try:

                cursor.execute(f'ALTER TABLE part_codes ADD COLUMN {col_name} {col_type}')

                print(f" Kolon eklendi: {col_name}")

            except sqlite3.OperationalError as e:

                if 'duplicate column name' in str(e).lower():

                    print(f" Kolon zaten var: {col_name}")

                else:

                    raise

        

        conn.commit()

        close_db(conn)

        print(" Part details schema gncellendi")

        

    except Exception as e:

        print(f" Part details schema hatas: {e}")

        try:

            if 'conn' in locals():

                close_db(conn)

        except:

            pass


def init_parts_info_table():
    """Yedek Parça Bilgi Sistemi için ayrı tablo oluştur"""
    try:
        conn = get_db()
        cursor = conn.cursor()
        
        # Parts Info tablosu - Ana sistemden tamamen bağımsız
        create_table_sql = '''
        CREATE TABLE IF NOT EXISTS parts_info (
            id INT AUTO_INCREMENT PRIMARY KEY,
            part_code VARCHAR(100) UNIQUE NOT NULL,
            part_name VARCHAR(255) NOT NULL,
            stock INT DEFAULT 0,
            critical_stock INT DEFAULT 0,
            expected_stock INT DEFAULT 0,
            supplier VARCHAR(255),
            purchase_price_eur DECIMAL(10, 2) DEFAULT 0.00,
            sale_price_eur DECIMAL(10, 2) DEFAULT 0.00,
            description TEXT,
            photo_path VARCHAR(500),
            stock_location VARCHAR(255),
            machines TEXT,
            specifications TEXT,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP ON UPDATE CURRENT_TIMESTAMP,
            INDEX idx_part_code (part_code),
            INDEX idx_part_name (part_name)
        ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4 COLLATE=utf8mb4_unicode_ci
        '''
        
        cursor = execute_query(cursor, create_table_sql)
        conn.commit()
        print("✓ Parts Info tablosu oluşturuldu/kontrol edildi")
        
        close_db(conn)
        
    except Exception as e:
        print(f"✗ Parts Info tablo hatası: {e}")
        try:
            if 'conn' in locals():
                close_db(conn)
        except:
            pass


def update_parts_info_columns():
    """Parts Info tablosuna yeni kolonları ekle"""
    try:
        conn = get_db()
        cursor = conn.cursor()
        
        # Eklenecek yeni kolonlar
        new_columns = [
            ('stock_location', 'VARCHAR(255)'),
            ('machines', 'TEXT'),
            ('specifications', 'TEXT'),
            ('replacement_code', 'VARCHAR(100)'),  # Değişen parça kodu
            ('build_out', 'TINYINT(1) DEFAULT 0')  # BUILD OUT bayrağı
        ]
        
        for col_name, col_type in new_columns:
            try:
                cursor.execute(f'ALTER TABLE parts_info ADD COLUMN {col_name} {col_type}')
                print(f"✓ Parts Info kolonu eklendi: {col_name}")
            except Exception as e:
                if 'Duplicate column name' in str(e):
                    # Kolon zaten var
                    pass
                else:
                    print(f"✗ Parts Info kolon eklenemedi {col_name}: {e}")
        
        conn.commit()
        close_db(conn)
        print("✓ Parts Info tablo kolonları güncellendi")
        
    except Exception as e:
        print(f"✗ Parts Info kolon güncelleme hatası: {e}")
        try:
            if 'conn' in locals():
                close_db(conn)
        except:
            pass


def init_user_permissions():
    """Kullanıcı izinleri için yeni kolonlar ekle"""
    try:
        conn = get_db()
        cursor = conn.cursor()
        
        # Yeni izin kolonları
        new_columns = [
            ('can_access_inventory', 'TINYINT(1) DEFAULT 1'),      # Envanter sistemine erişim
            ('can_access_order_system', 'TINYINT(1) DEFAULT 0'),   # Sipariş sistemine erişim
            ('can_access_parts_info', 'TINYINT(1) DEFAULT 0'),     # Yedek parça sistemine erişim
            ('can_view_purchase_price', 'TINYINT(1) DEFAULT 0'),   # Geliş fiyatını görme yetkisi
            ('can_edit_parts_info', 'TINYINT(1) DEFAULT 0')        # Parça bilgilerini düzenleme yetkisi
        ]
        
        for col_name, col_type in new_columns:
            try:
                # Her kolon için yeni cursor al
                if cursor is None or not hasattr(cursor, 'execute'):
                    cursor = conn.cursor()
                    
                cursor.execute(f'ALTER TABLE envanter_users ADD COLUMN {col_name} {col_type}')
                print(f"✓ Kolon eklendi: {col_name}")
            except Exception as e:
                if 'Duplicate column name' in str(e):
                    # Kolon zaten var, hata değil
                    pass
                else:
                    print(f"✗ Kolon eklenemedi {col_name}: {e}")
        
        conn.commit()
        close_db(conn)
        print("✓ Kullanıcı izin şeması güncellendi")
        
    except Exception as e:
        print(f"✗ Kullanıcı izin şeması hatası: {e}")
        try:
            if 'conn' in locals():
                close_db(conn)
        except:
            pass


def update_admin_permissions():
    """Admin kullanıcılarına tüm yetkileri ver"""
    try:
        conn = get_db()
        cursor = conn.cursor()
        
        # Admin rolündeki tüm kullanıcıları güncelle
        cursor.execute('''
            UPDATE envanter_users 
            SET can_access_inventory = 1,
                can_access_order_system = 1,
                can_access_parts_info = 1,
                can_view_purchase_price = 1,
                can_edit_parts_info = 1
            WHERE role = 'admin'
        ''')
        
        updated = cursor.rowcount
        conn.commit()
        close_db(conn)
        
        if updated > 0:
            print(f"✓ {updated} admin kullanıcısının yetkileri güncellendi")
        
    except Exception as e:
        print(f"✗ Admin yetki güncelleme hatası: {e}")
        try:
            if 'conn' in locals():
                close_db(conn)
        except:
            pass


def init_db():

    """Initialize database tables using SQLAlchemy ORM



    Lokal: SQLite tables olutur

    Production: PostgreSQL tables kontrol et

    """

    try:

        with app.app_context():

            if USE_MYSQL:

                # PRODUCTION: MySQL - SQLAlchemy ile

                print("\n" + "="*70)

                print("[INIT] INITIALIZING MYSQL TABLES")

                print("="*70)



                inspector = db.inspect(db.engine)

                existing_tables = inspector.get_table_names()

                print(f"Existing tables: {existing_tables}")



                required_tables = ['envanter_users', 'part_codes', 'qr_codes', 'count_sessions', 'count_passwords', 'scanned_qr']



                missing_tables = [t for t in required_tables if t not in existing_tables]



                if missing_tables:

                    print(f" Creating missing MySQL tables: {', '.join(missing_tables)}")

                    db.create_all()

                    print(" MySQL tables created successfully")

                else:

                    print(" All MySQL tables already exist")



                #  DATABASE MIGRATION: Add finished_by column to count_sessions

                try:

                    conn = get_db()

                    cursor = conn.cursor()



                    # Check if finished_by column exists

                    execute_query(cursor, """

                        SELECT column_name 

                        FROM information_schema.columns 

                        WHERE table_name='count_sessions' AND column_name='finished_by'

                    """)



                    if not cursor.fetchone():

                        print(" Adding finished_by column to count_sessions table...")

                        execute_query(cursor, """

                            ALTER TABLE count_sessions 

                            ADD COLUMN finished_by INTEGER REFERENCES envanter_users(id)

                        """)

                        conn.commit()

                        print(" finished_by column added successfully")

                    else:

                        print(" finished_by column already exists")



                    close_db(conn)

                except Exception as e:

                    print(f" Migration warning: {e}")



                #  DATABASE MIGRATION: Add total_expected and total_scanned columns to count_sessions

                try:

                    conn = get_db()

                    cursor = conn.cursor()



                    # Check total_expected

                    execute_query(cursor, """

                        SELECT column_name FROM information_schema.columns

                        WHERE table_name='count_sessions' AND column_name='total_expected'

                    """)

                    if not cursor.fetchone():

                        print(" Adding total_expected column to count_sessions table...")

                        execute_query(cursor, """

                            ALTER TABLE count_sessions

                            ADD COLUMN total_expected INTEGER DEFAULT 0

                        """)

                        conn.commit()

                        print(" total_expected column added successfully")

                    else:

                        print(" total_expected column already exists")



                    # Check total_scanned

                    execute_query(cursor, """

                        SELECT column_name FROM information_schema.columns

                        WHERE table_name='count_sessions' AND column_name='total_scanned'

                    """)

                    if not cursor.fetchone():

                        print(" Adding total_scanned column to count_sessions table...")

                        execute_query(cursor, """

                            ALTER TABLE count_sessions

                            ADD COLUMN total_scanned INTEGER DEFAULT 0

                        """)

                        conn.commit()

                        print(" total_scanned column added successfully")

                    else:

                        print(" total_scanned column already exists")



                    #  DATABASE MIGRATION: Add report_file_path column to count_sessions

                    cursor = conn.cursor()

                    # PostgreSQL: Check if column exists
                    cursor.execute("""

                        SELECT COUNT(*) FROM information_schema.columns 

                        WHERE table_name='count_sessions' AND column_name='report_file_path'

                    """)

                    if cursor.fetchone()[0] == 0:

                        print(" Adding report_file_path column to count_sessions table...")

                        cursor.execute("""

                            ALTER TABLE count_sessions

                            ADD COLUMN report_file_path TEXT

                        """)

                        conn.commit()

                        print(" report_file_path column added successfully")

                    else:

                        print(" report_file_path column already exists")



                    close_db(conn)

                except Exception as e:

                    print(f" Migration warning (total_*): {e}")



                # Verify scanned_qr table specifically (critical for duplicate detection)

                if 'scanned_qr' in existing_tables or 'scanned_qr' not in missing_tables:

                    print(" scanned_qr table verified - duplicate detection will work")
                    
                    # 🔒 MULTI-SCANNER SAFETY: Add unique constraint if not exists
                    try:
                        conn = get_db()
                        cursor = conn.cursor()
                        
                        # Check if unique constraint already exists
                        cursor.execute("""
                            SELECT COUNT(*) FROM information_schema.table_constraints 
                            WHERE table_name='scanned_qr' 
                            AND constraint_type='UNIQUE'
                            AND constraint_name='unique_qr_session'
                        """)
                        
                        if cursor.fetchone()[0] == 0:
                            print(" [CONSTRAINT] Adding UNIQUE constraint for multi-scanner safety...")
                            cursor.execute("""
                                ALTER TABLE scanned_qr 
                                ADD CONSTRAINT unique_qr_session UNIQUE (qr_id, session_id)
                            """)
                            conn.commit()
                            print(" [OK] UNIQUE constraint added - multi-scanner duplicate protection enabled")
                        else:
                            print(" [OK] UNIQUE constraint already exists - multi-scanner safe")
                        
                        close_db(conn)
                    except Exception as constraint_error:
                        print(f" [INFO] Constraint check: {str(constraint_error)[:100]}")
                        if 'conn' in locals():
                            close_db(conn)

                else:

                    print(" WARNING: scanned_qr table not found - duplicate detection may fail!")

                    print("   Creating scanned_qr table manually...")

                    try:

                        # Create scanned_qr table manually if SQLAlchemy fails

                        conn = get_db()

                        cursor = conn.cursor()

                        execute_query(cursor, '''

                            CREATE TABLE IF NOT EXISTS scanned_qr (

                                id SERIAL PRIMARY KEY,

                                session_id VARCHAR(255) NOT NULL,

                                qr_id VARCHAR(255) NOT NULL,

                                part_code VARCHAR(255) NOT NULL,

                                scanned_by INTEGER,

                                scanned_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                                UNIQUE (qr_id, session_id)

                            )

                        ''')

                        conn.commit()

                        close_db(conn)

                        print(" scanned_qr table created successfully with UNIQUE constraint")

                    except Exception as e:

                        print(f" Failed to create scanned_qr: {e}")



                # 🔧 Fix description column size for large JSON data
                try:
                    conn = get_db()
                    cursor = conn.cursor()
                    
                    # Check if description column needs to be expanded
                    cursor.execute("""
                        SELECT DATA_TYPE, CHARACTER_MAXIMUM_LENGTH 
                        FROM information_schema.COLUMNS 
                        WHERE TABLE_NAME='count_sessions' AND COLUMN_NAME='description'
                    """)
                    col_info = cursor.fetchone()
                    
                    if col_info and col_info[0] != 'longtext':
                        print(" 🔧 Expanding description column to LONGTEXT for large JSON...")
                        cursor.execute("""
                            ALTER TABLE count_sessions 
                            MODIFY COLUMN description LONGTEXT
                        """)
                        conn.commit()
                        print("[OK] description column expanded successfully")
                    else:
                        print("[OK] description column already LONGTEXT")
                    
                    close_db(conn)
                except Exception as desc_error:
                    print(f"[INFO] Description column check: {str(desc_error)[:100]}")
                    if 'conn' in locals():
                        close_db(conn)

                # PostgreSQL admin user kontrolü

                conn = get_db()
                cursor = conn.cursor()
                
                execute_query(cursor, 'SELECT id FROM envanter_users WHERE username = %s', ('admin',))
                admin_user = cursor.fetchone()

                if not admin_user:

                    print(" Creating default PostgreSQL admin user...")

                    from werkzeug.security import generate_password_hash

                    admin_password = generate_password_hash("@R9t$L7e!xP2w")

                    execute_query(cursor, '''

                        INSERT INTO envanter_users (username, password_hash, full_name, role, created_at, is_active_user)

                        VALUES (%s, %s, %s, %s, %s, %s)

                    ''', ('admin', admin_password, 'Administrator', 'admin', datetime.now(), True))

                    conn.commit()

                    print(" MySQL admin user created (admin/@R9t$L7e!xP2w)")

                else:

                    print(" MySQL admin user already exists")

                
                close_db(conn)

                print(" MySQL database initialized successfully")

            

            # Part details kolonlarn ekle

            init_db_part_details()



            return True



    except Exception as e:

        print(f" Database initialization error: {str(e)}")

        import traceback

        traceback.print_exc()

        return False



def init_db_part_details():

    """Part codes tablosuna detay kolonlar ekle"""

    try:

        conn = get_db()

        cursor = conn.cursor()

        

        new_columns = [

            ('photo_path', 'TEXT'),

            ('catalog_image', 'TEXT'),

            ('description', 'TEXT'),

            ('used_in_machines', 'TEXT'),  # JSON array

            ('specifications', 'TEXT'),     # JSON object

            ('stock_location', 'TEXT'),

            ('supplier', 'TEXT'),

            ('unit_price', 'REAL'),

            ('critical_stock_level', 'INTEGER'),

            ('notes', 'TEXT'),

            ('last_updated', 'TIMESTAMP'),

            ('updated_by', 'INTEGER')

        ]

        

        for col_name, col_type in new_columns:

            try:

                cursor.execute(f'ALTER TABLE part_codes ADD COLUMN {col_name} {col_type}')

                conn.commit()  # Commit after each column addition

                print(f" Kolon eklendi: {col_name}")

            except Exception as e:

                conn.rollback()  # Rollback on error

                error_msg = str(e).lower()

                if 'already exists' in error_msg or 'duplicate' in error_msg:

                    print(f" Kolon zaten var: {col_name}")

                else:

                    print(f" Column error (ignored): {col_name} - {e}")

        

        conn.commit()

        close_db(conn)

        print(" Part details schema updated")

    except Exception as e:

        print(f" Part details schema error: {e}")



def login_required(f):

    from functools import wraps

    @wraps(f)

    def decorated_function(*args, **kwargs):

        if 'user_id' not in session:

            return jsonify({'error': 'Giri yapmanz gerekiyor'}), 401

        return f(*args, **kwargs)

    return decorated_function



def admin_required(f):

    from functools import wraps

    @wraps(f)

    def decorated_function(*args, **kwargs):

        if 'user_id' not in session or session.get('role') != 'admin':

            return redirect('/admin')

        return f(*args, **kwargs)

    return decorated_function



def admin_required_decorator(f):

    """Sadece admin eriimi iin decorator"""

    from functools import wraps

    @wraps(f)

    def decorated_function(*args, **kwargs):

        if 'user_id' not in session:

            return jsonify({'error': 'Giri yapmanz gerekiyor'}), 401

        if session.get('role') != 'admin':

            return jsonify({'error': 'Bu ilem iin admin yetkisi gerekli'}), 403

        return f(*args, **kwargs)

    return decorated_function



@app.route('/favicon.ico')

def favicon():

    return '', 204



# ================================

# PART DETAILS API ENDPOINTS

# ================================



@app.route('/api/upload_part_photo/<part_code>', methods=['POST'])

@login_required

def upload_part_photo(part_code):

    """Para fotoraf ykle"""

    if 'photo' not in request.files:

        return jsonify({'error': 'Dosya bulunamad'}), 400

    

    file = request.files['photo']

    if file.filename == '':

        return jsonify({'error': 'Dosya seilmedi'}), 400

    

    if file and allowed_file(file.filename):

        # HER ZAMAN UNC PATH - paylaşımlı klasöre kaydet
        upload_dir = UPLOAD_FOLDER
        os.makedirs(upload_dir, exist_ok=True)

        

        # Gvenli dosya ad

        filename = secure_filename(file.filename)

        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')

        new_filename = f"{part_code}_{timestamp}_{filename}"

        filepath = os.path.join(upload_dir, new_filename)

        

        file.save(filepath)

        

        # Database'e kaydet

        conn = get_db()

        cursor = conn.cursor()

        execute_query(cursor, f'''

            UPDATE part_codes 

            SET photo_path = %s, 

                last_updated = %s,

                updated_by = %s

            WHERE part_code = %s

        ''', (f'part_photos/{new_filename}', datetime.now(), session['user_id'], part_code))

        

        conn.commit()

        close_db(conn)

        

        return jsonify({

            'success': True,

            'photo_path': f'part_photos/{new_filename}'

        })

    

    return jsonify({'error': 'Geersiz dosya format (PNG, JPG, GIF, WEBP)'}), 400



@app.route('/api/upload_catalog_image/<part_code>', methods=['POST'])

@login_required

def upload_catalog_image(part_code):

    """Katalog grnts ykle"""

    if 'catalog' not in request.files:

        return jsonify({'error': 'Dosya bulunamad'}), 400

    

    file = request.files['catalog']

    if file.filename == '':

        return jsonify({'error': 'Dosya seilmedi'}), 400

    

    if file and allowed_file(file.filename):

        os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)

        

        filename = secure_filename(file.filename)

        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')

        new_filename = f"{part_code}_catalog_{timestamp}_{filename}"

        filepath = os.path.join(app.config['UPLOAD_FOLDER'], new_filename)

        

        file.save(filepath)

        

        conn = get_db()

        cursor = conn.cursor()

        execute_query(cursor, f'''

            UPDATE part_codes 

            SET catalog_image = %s, 

                last_updated = %s,

                updated_by = %s

            WHERE part_code = %s

        ''', (f'part_photos/{new_filename}', datetime.now(), session['user_id'], part_code))

        

        conn.commit()

        close_db(conn)

        

        return jsonify({

            'success': True,

            'catalog_image': f'part_photos/{new_filename}'

        })

    

    return jsonify({'error': 'Geersiz dosya format'}), 400



@app.route('/api/parts', methods=['GET'])

@login_required

def get_all_parts():

    """Tm paralar listele (admin panel iin)"""

    conn = get_db()

    cursor = conn.cursor()

    

    execute_query(cursor, '''

        SELECT part_code, description, photo_path, catalog_image

        FROM part_codes

        ORDER BY part_code

    ''')

    

    parts = cursor.fetchall()

    close_db(conn)

    

    result = []

    for part in parts:

        result.append({

            'part_code': part[0],

            'description': part[1],

            'photo_path': part[2],

            'catalog_image': part[3]

        })

    

    return jsonify(result)



@app.route('/api/part_details/<part_code>', methods=['GET'])

def get_part_details(part_code):

    """Parça detaylarını getir (herkes görebilir)"""

    try:
        conn = get_db()
        cursor = conn.cursor()
        
        # Önce parts_info tablosundan kontrol et
        execute_query(cursor, '''
            SELECT id, part_code, part_name, description, photo_path, 
                   stock_location, supplier, created_at, updated_at
            FROM parts_info
            WHERE part_code = %s
        ''', (part_code,))
        
        row = cursor.fetchone()
        
        # parts_info'dan geldiyse basit format (9 alan)
        if row:
            close_db(conn)
            return jsonify({
                'id': row[0],
                'part_code': row[1],
                'part_name': row[2],  # part_name artık SELECT'te
                'description': row[3],
                'photo_path': row[4],
                'catalog_image': None,
                'stock_location': row[5],
                'supplier': row[6],
                'created_at': row[7].isoformat() if row[7] else None,
                'updated_at': row[8].isoformat() if row[8] else None,
                'used_in_machines': [],
                'specifications': {},
                'user_role': session.get('role', None)  # Yetki kontrolü için
            })
        
        # parts_info'da yoksa part_codes'dan bak
        execute_query(cursor, '''
            SELECT id, part_code, part_name, description, is_package, package_items, created_at, updated_at,
                   photo_path, catalog_image, stock_location, supplier, notes, used_in_machines, specifications
            FROM part_codes
            WHERE part_code = %s
        ''', (part_code,))
        
        row = cursor.fetchone()
        
        close_db(conn)
        
        if not row:
            return jsonify({'error': 'Parça bulunamadı'}), 404
        
        import json
        
        # part_codes'dan geldiyse eski format
        # used_in_machines ve specifications JSON olabilir veya string
        used_in_machines = row[13]
        if isinstance(used_in_machines, str):
            try:
                used_in_machines = json.loads(used_in_machines)
            except:
                used_in_machines = []
        if not used_in_machines:
            used_in_machines = []
        
        specifications = row[14]
        if isinstance(specifications, str):
            try:
                specifications = json.loads(specifications)
            except:
                specifications = {}
        if not specifications:
            specifications = {}
        
        return jsonify({
            'id': row[0],
            'part_code': row[1],
            'part_name': row[2],
            'description': row[3],
            'is_package': row[4],
            'package_items': row[5],
            'created_at': row[6].isoformat() if row[6] else None,
            'updated_at': row[7].isoformat() if row[7] else None,
            'photo_path': row[8],
            'catalog_image': row[9],
            'stock_location': row[10],
            'supplier': row[11],
            'notes': row[12],
            'used_in_machines': used_in_machines,
            'specifications': specifications,
            'user_role': session.get('role', None)  # Yetki kontrolü için
        })
    except Exception as e:
        logging.error(f"Error getting part details: {e}")
        return jsonify({'error': str(e)}), 500



@app.route('/api/update_part_details/<part_code>', methods=['POST'])

@login_required

def update_part_details(part_code):

    """Para detaylarn gncelle (admin only)"""

    if session.get('role') != 'admin':

        return jsonify({'error': 'Admin yetkisi gerekli'}), 403

    

    data = request.json
    
    try:
        conn = get_db()
        cursor = conn.cursor()
        
        # Check if part exists
        execute_query(cursor, 'SELECT id FROM part_codes WHERE part_code = %s', (part_code,))
        if not cursor.fetchone():
            close_db(conn)
            return jsonify({'error': 'Para bulunamad'}), 404
        
        # Update fields
        updates = []
        params = []
        
        if 'part_name' in data:
            updates.append('part_name = %s')
            params.append(data['part_name'])
        if 'description' in data:
            updates.append('description = %s')
            params.append(data['description'])
        if 'is_package' in data:
            updates.append('is_package = %s')
            params.append(data['is_package'])
        if 'package_items' in data:
            updates.append('package_items = %s')
            params.append(data['package_items'])
        if 'stock_location' in data:
            updates.append('stock_location = %s')
            params.append(data['stock_location'])
        if 'supplier' in data:
            updates.append('supplier = %s')
            params.append(data['supplier'])
        if 'notes' in data:
            updates.append('notes = %s')
            params.append(data['notes'])
        if 'used_in_machines' in data:
            updates.append('used_in_machines = %s')
            params.append(json.dumps(data['used_in_machines']))
        if 'specifications' in data:
            updates.append('specifications = %s')
            params.append(json.dumps(data['specifications']))
        
        if updates:
            updates.append('updated_at = %s')
            params.append(datetime.now())
            params.append(part_code)
            
            sql = f"UPDATE part_codes SET {', '.join(updates)} WHERE part_code = %s"
            execute_query(cursor, sql, tuple(params))
            conn.commit()
        
        close_db(conn)
        return jsonify({'success': True})
    
    except Exception as e:
        logging.error(f"Error updating part details: {e}")
        return jsonify({'error': str(e)}), 500



@app.route('/qr/<qr_id>')

def qr_redirect(qr_id):

    """QR Code ynlendirme - Aktif saym varsa scanner, yoksa part info"""

    conn = get_db()

    cursor = conn.cursor()

    # Aktif saym var m

    execute_query(cursor, f'SELECT id FROM count_sessions WHERE is_active = %s', (True,))

    active_session = cursor.fetchone()

    

    if active_session:

        # Aktif saym var -> Scanner sayfas

        close_db(conn)

        return redirect(url_for('scanner'))

    else:

        # Aktif saym yok -> Part detay sayfas

        # QR'dan part_code bul

        execute_query(cursor, f'''

            SELECT pc.part_code

            FROM qr_codes qr

            JOIN part_codes pc ON qr.part_code_id = pc.id

            WHERE qr.qr_id = %s

        ''', (qr_id,))

        

        result = cursor.fetchone()

        close_db(conn)

        

        if result:

            return redirect(url_for('part_info', part_code=result[0]))

        else:

            return "QR Code bulunamad", 404



@app.route('/part_info/<part_code>')

def part_info(part_code):

    """Para bilgi sayfas"""

    return render_template('part_info.html', part_code=part_code)



@app.route('/edit_part/<part_code>')

@admin_required

def edit_part(part_code):

    """Para dzenleme sayfas (admin)"""

    return render_template('edit_part.html', part_code=part_code)



@app.route('/part_search')
def part_search():
    """Parça arama sayfası"""
    return render_template('part_search.html')

@app.route('/part_advanced_search')
def part_advanced_search():
    """Gelişmiş parça sorgulama sayfası - Kod ve ad ile arama"""
    return render_template('part_advanced_search.html')

@app.route('/api/search_parts')
def search_parts():
    """Parça arama API'si - TR Klavye uyumlu"""
    # NUL karakterlerini temizle (PostgreSQL hatası önleme)
    raw_query = request.args.get('q', '')
    if '\x00' in raw_query or '\0' in raw_query:
        raw_query = raw_query.replace('\x00', '').replace('\0', '')
    
    query = raw_query.strip()
    
    # Boş query'yi hemen döndür
    if not query or len(query) < 1:
        return jsonify([])
    
    try:
        # Türkçe karakterleri normalize et
        def normalize_turkish(text):
            """Türkçe › İngilizce karakter + scanner düzeltme"""
            if not text:
                return text
            
            # 1. Büyük harfe çevir
            text = text.upper()
            
            # 2. Türkçe › İngilizce
            tr_map = {
                'Ç': 'C', 'Ğ': 'G', 'İ': 'I', 'Ö': 'O', 'Ş': 'S', 'Ü': 'U',
                'ç': 'C', 'ğ': 'G', 'ı': 'I', 'i': 'I', 'ö': 'O', 'ş': 'S', 'ü': 'U'
            }
            for tr_char, en_char in tr_map.items():
                text = text.replace(tr_char, en_char)
            
            # 3. Scanner hataları: * ? _ › -
            text = text.replace('*', '-').replace('?', '-').replace('_', '-')
            
            # 4. Sondaki -X kaldır (QR kopya no)
            import re
            text = re.sub(r'-\d+$', '', text)
            
            return text.strip()
        
        normalized_query = normalize_turkish(query)
        original_upper = query.upper().strip()
        
        # Normalize sonrası boşsa hemen döndür
        if not normalized_query:
            return jsonify([])
        
        conn = get_db()
        cursor = conn.cursor()
        
        # 1. TAM EŞLEŞME - En yüksek öncelik
        cursor = execute_query(cursor, '''
            SELECT part_code, part_name
            FROM part_codes
            WHERE part_code = %s OR part_code = %s
            LIMIT 1
        ''', (normalized_query, original_upper))
        
        exact_match = cursor.fetchall()
        
        # 2. BAŞLANGIÇ EŞLEŞME - İkinci öncelik (19111-01342 ile başlayanlar)
        starts_with_results = []
        if not exact_match:
            normalized_pattern_start = f'{normalized_query}%'
            original_pattern_start = f'{original_upper}%'
            cursor = execute_query(cursor, '''
                SELECT part_code, part_name
                FROM part_codes
                WHERE part_code LIKE %s 
                   OR part_code LIKE %s
                ORDER BY LENGTH(part_code) ASC
                LIMIT 10
            ''', (normalized_pattern_start, original_pattern_start))
            starts_with_results = cursor.fetchall()
        
        # 3. İÇERİK EŞLEŞME - Üçüncü öncelik (19111-01342 içerenler)
        contains_results = []
        if not exact_match and not starts_with_results:
            normalized_pattern = f'%{normalized_query}%'
            original_pattern = f'%{original_upper}%'
            cursor = execute_query(cursor, '''
                SELECT part_code, part_name
                FROM part_codes
                WHERE (part_code LIKE %s OR part_code LIKE %s
                   OR part_name LIKE %s OR part_name LIKE %s)
                   AND part_code NOT LIKE %s
                   AND part_code NOT LIKE %s
                ORDER BY LENGTH(part_code) ASC
                LIMIT 20
            ''', (normalized_pattern, original_pattern, normalized_pattern, original_pattern,
                  normalized_pattern_start, original_pattern_start))
            contains_results = cursor.fetchall()
        
        # Sonuçları birleştir: Önce tam eşleşme, sonra başlangıç, sonra içerik
        results = exact_match or starts_with_results or contains_results
        
        # 3. Hala yoksa fuzzy match (Türkçe karakter eksikliği için)
        # ANTF03 › ANTİF03 gibi durumlar (QR okuyucu İ, Ğ, Ş gibi karakterleri atlıyor)
        if not results and len(original_upper) >= 3:
            print(f'[Search API] Fuzzy matching başlatılıyor...')
            
            # ASCII normalizasyon (Türkçe › İngilizce mapping)
            def ascii_normalize(text):
                """Türkçe karakterleri İngilizce'ye çevir"""
                if not text:
                    return ''
                text = text.upper()
                replacements = {
                    'Ç': 'C', 'Ğ': 'G', 'İ': 'I', 'Ö': 'O', 'Ş': 'S', 'Ü': 'U',
                    'ç': 'C', 'ğ': 'G', 'ı': 'I', 'i': 'I', 'ö': 'O', 'ş': 'S', 'ü': 'U'
                }
                for tr, en in replacements.items():
                    text = text.replace(tr, en)
                return text
            
            query_ascii = ascii_normalize(original_upper)
            print(f'[Search API] Query ASCII: "{query_ascii}"')
            
            # Benzer uzunluktaki tüm part_code'ları al
            cursor = execute_query(cursor, '''
                SELECT part_code, part_name
                FROM part_codes
                WHERE LENGTH(part_code) BETWEEN %s AND %s
                LIMIT 1000
            ''', (len(query_ascii) - 2, len(query_ascii) + 2))
            
            all_parts = cursor.fetchall()
            print(f'[Search API] Fuzzy için {len(all_parts)} parça alındı')
            
            # Her part_code'u ASCII normalize edip karşılaştır
            matches = []
            for part_code, part_name in all_parts:
                part_ascii = ascii_normalize(part_code)
                
                # Exact match (ASCII versiyonlar)
                if part_ascii == query_ascii:
                    matches.append((100, part_code, part_name))
                    print(f'[Search API] EXACT MATCH: {part_code} (ASCII: {part_ascii})')
                # Substring match
                elif query_ascii in part_ascii:
                    score = int((len(query_ascii) / len(part_ascii)) * 90)
                    matches.append((score, part_code, part_name))
                    print(f'[Search API] CONTAINS: {part_code} › Score: {score}%')
                elif part_ascii in query_ascii:
                    score = int((len(part_ascii) / len(query_ascii)) * 85)
                    matches.append((score, part_code, part_name))
                    print(f'[Search API] REVERSE CONTAINS: {part_code} › Score: {score}%')
                # Levenshtein benzeri (basit karakter farkı)
                else:
                    # Tek karakter farkı varsa eşleştir
                    if abs(len(query_ascii) - len(part_ascii)) <= 1:
                        diff_count = sum(1 for a, b in zip(query_ascii, part_ascii) if a != b)
                        if diff_count <= 2:  # En fazla 2 karakter farkı
                            score = int((1 - (diff_count / max(len(query_ascii), len(part_ascii)))) * 75)
                            matches.append((score, part_code, part_name))
                            print(f'[Search API] SIMILAR: {part_code} (diff: {diff_count}) › Score: {score}%')
            
            # En yüksek skorlu 5 sonuç
            matches.sort(reverse=True, key=lambda x: x[0])
            results = [(code, name) for _, code, name in matches[:5]]
            
            if results:
                print(f'[Search API] Fuzzy match bulundu: {len(results)} sonuç')
            else:
                print(f'[Search API] Fuzzy match sonuç bulamadı')
        
        close_db(conn)
        
        print(f'[Search API] Original: "{query}" | Normalized: "{normalized_query}" | Found: {len(results)}')
        
        return jsonify([{
            'part_code': row[0],
            'part_name': row[1]
        } for row in results])
    except Exception as e:
        print(f'Search error: {e}')
        return jsonify([])

@app.route('/api/search_parts_advanced')
def search_parts_advanced():
    """Akıllı parça arama API'si - Kod veya ad otomatik algılama"""
    query = request.args.get('q', '').strip()
    
    if not query or len(query) < 2:
        return jsonify([])
    
    try:
        conn = get_db()
        cursor = conn.cursor()
        
        # Akıllı mod algılama: 
        # - Eğer query rakam ve tire içeriyorsa (örn: Y129-123, 12345) -> Kod araması
        # - Eğer sadece harf içeriyorsa -> İsim araması
        # - Karışık ise -> Her ikisinde de ara
        
        query_upper = query.upper()
        has_digits = any(c.isdigit() for c in query)
        has_letters = any(c.isalpha() for c in query)
        
        all_results = []
        
        # PARÇA KODU ARAMASI (rakam içeriyorsa veya tire varsa)
        if has_digits or '-' in query:
            # 1. Tam eşleşme - önce parts_info'dan ara
            cursor = execute_query(cursor, '''
                SELECT part_code, part_name, stock_location, 0, 0
                FROM parts_info
                WHERE UPPER(part_code) = %s
                LIMIT 1
            ''', (query_upper,))
            exact_results = cursor.fetchall()
            
            # parts_info'da yoksa part_codes'dan ara
            if not exact_results:
                cursor = execute_query(cursor, '''
                    SELECT pc.part_code, pc.part_name, pc.stock_location, pc.quantity, pc.critical_stock_level
                    FROM part_codes pc
                    WHERE UPPER(pc.part_code) = %s
                    LIMIT 1
                ''', (query_upper,))
                exact_results = cursor.fetchall()
            
            all_results.extend(exact_results)
            
            # 2. Başlangıç eşleşme
            if not exact_results:
                # parts_info'dan
                cursor = execute_query(cursor, '''
                    SELECT part_code, part_name, stock_location, 0, 0
                    FROM parts_info
                    WHERE UPPER(part_code) LIKE %s
                    ORDER BY LENGTH(part_code) ASC
                    LIMIT 25
                ''', (f'{query_upper}%',))
                start_results = cursor.fetchall()
                all_results.extend(start_results)
                
                # part_codes'dan
                cursor = execute_query(cursor, '''
                    SELECT pc.part_code, pc.part_name, pc.stock_location, pc.quantity, pc.critical_stock_level
                    FROM part_codes pc
                    WHERE UPPER(pc.part_code) LIKE %s
                    ORDER BY LENGTH(pc.part_code) ASC
                    LIMIT 25
                ''', (f'{query_upper}%',))
                start_results = cursor.fetchall()
                all_results.extend(start_results)
            
            # 3. İçerik eşleşme (hala sonuç yoksa)
            if not all_results:
                # parts_info'dan
                cursor = execute_query(cursor, '''
                    SELECT part_code, part_name, stock_location, 0, 0
                    FROM parts_info
                    WHERE UPPER(part_code) LIKE %s
                    ORDER BY LENGTH(part_code) ASC
                    LIMIT 25
                ''', (f'%{query_upper}%',))
                contains_results = cursor.fetchall()
                all_results.extend(contains_results)
                
                # part_codes'dan
                cursor = execute_query(cursor, '''
                    SELECT pc.part_code, pc.part_name, pc.stock_location, pc.quantity, pc.critical_stock_level
                    FROM part_codes pc
                    WHERE UPPER(pc.part_code) LIKE %s
                    ORDER BY LENGTH(pc.part_code) ASC
                    LIMIT 25
                ''', (f'%{query_upper}%',))
                contains_results = cursor.fetchall()
                all_results.extend(contains_results)
        
        # PARÇA ADI ARAMASI (harf içeriyorsa veya kod araması sonuç vermediyse)
        if (has_letters and not has_digits) or (not all_results and has_letters):
            cursor = execute_query(cursor, '''
                SELECT pc.part_code, pc.part_name, pc.stock_location, pc.quantity, pc.critical_stock_level
                FROM part_codes pc
                WHERE UPPER(pc.part_name) LIKE %s
                ORDER BY 
                    CASE 
                        WHEN UPPER(pc.part_name) LIKE %s THEN 1
                        ELSE 2
                    END,
                    pc.part_name ASC
                LIMIT 50
            ''', (f'%{query_upper}%', f'{query_upper}%'))
            name_results = cursor.fetchall()
            
            # Duplicate kontrolü - aynı part_code varsa ekleme
            existing_codes = {r[0] for r in all_results}
            for row in name_results:
                if row[0] not in existing_codes:
                    all_results.append(row)
                    existing_codes.add(row[0])
        
        close_db(conn)
        
        return jsonify([{
            'part_code': row[0],
            'part_name': row[1],
            'location': row[2],
            'quantity': row[3],
            'min_stock': row[4]
        } for row in all_results[:50]])  # Maksimum 50 sonuç
    
    except Exception as e:
        print(f'Smart search error: {e}')
        import traceback
        traceback.print_exc()
        return jsonify([])



# ================================

# EXISTING ROUTES BELOW

# ================================



@app.route('/')

def index():

    """Ana sayfa - Sistem Seçimi"""

    if 'user_id' not in session:

        return redirect('/login')

    

    return render_template('system_selection.html')


@app.route('/inventory')
@login_required
def inventory_dashboard():
    """Envanter Sistemi Ana Sayfası"""
    return render_template('index.html')


@app.route('/parts_info/')
@login_required
def parts_info_system():
    """Yedek Parça Bilgi Sistemi Ana Sayfası"""
    return render_template('parts_info/main.html')

@app.route('/parts_info/detail/<path:part_code>')
@login_required
def parts_info_detail(part_code):
    """Parça detay sayfası"""
    # URL decode için unquote kullan (Türkçe karakter desteği)
    from urllib.parse import unquote
    part_code = unquote(part_code)
    return render_template('parts_info/detail.html', part_code=part_code)

@app.route('/api/parts_info/detail/<path:part_code>')
@login_required
def get_part_detail(part_code):
    """Parça detaylarını getir"""
    try:
        # URL decode için unquote kullan (Türkçe karakter desteği)
        from urllib.parse import unquote
        part_code = unquote(part_code)
        conn = get_db()
        cursor = conn.cursor()
        
        # Kullanıcının geliş fiyatı görme yetkisini kontrol et
        user_id = session.get('user_id')
        cursor = execute_query(cursor, 
            'SELECT can_view_purchase_price, can_edit_parts_info, role, username FROM envanter_users WHERE id = %s', 
            (user_id,))
        user_data = cursor.fetchone()
        can_view_purchase = user_data[0] if user_data else False
        can_edit = user_data[1] if user_data else False
        
        # Debug log
        if user_data:
            print(f"[PARTS_DETAIL] User: {user_data[2]}, Role: {user_data[1]}, Can View Purchase: {can_view_purchase}")
        
        # Parça bilgilerini getir
        cursor = execute_query(cursor, '''
            SELECT 
                part_code, part_name, stock, critical_stock, expected_stock,
                supplier, purchase_price_eur, description, sale_price_eur, photo_path, stock_location, machines, specifications,
                replacement_code, build_out
            FROM parts_info
            WHERE part_code = %s
        ''', (part_code,))
        
        row = cursor.fetchone()
        close_db(conn)
        
        if not row:
            return jsonify({'error': 'Parça bulunamadı'}), 404
        
        # EUR -> TRY dönüşümü için kur al
        try:
            exchange_rate = get_tcmb_exchange_rate()
        except:
            exchange_rate = 35.0
        
        part = {
            'part_code': row[0],
            'part_name': row[1],
            'stock': row[2] or 0,
            'critical_stock': row[3] or 0,
            'expected_stock': row[4] or 0,
            'supplier': row[5] or '',
            'purchase_price_eur': float(row[6]) if (row[6] and can_view_purchase) else None,
            'description': row[7] or '',
            'sale_price_eur': float(row[8]) if row[8] else 0,
            'photo_path': row[9] or '',
            'stock_location': row[10] or '',
            'machines': row[11] or '',
            'specifications': row[12] or '',
            'replacement_code': row[13] or '',  # Değişen parça kodu
            'build_out': bool(row[14]) if row[14] else False,  # BUILD OUT bayrağı
            'exchange_rate': exchange_rate,
            'purchase_price_try': round(float(row[6] or 0) * exchange_rate, 2) if can_view_purchase else None,
            'sale_price_try': round(float(row[8] or 0) * exchange_rate, 2),
            'can_view_purchase': can_view_purchase,
            'can_edit': can_edit
        }
        
        # Debug log
        print(f"[PARTS_DETAIL] Part: {part['part_code']}, Can View: {can_view_purchase}, Can Edit: {can_edit}")
        
        return jsonify(part)
        
    except Exception as e:
        print(f"Part detail error: {e}")
        return jsonify({'error': str(e)}), 500

@app.route('/api/parts_info/update/<path:part_code>', methods=['PUT'])
@login_required
def update_part_info(part_code):
    """Parça açıklamasını güncelle - sadece description"""
    try:
        # URL decode için unquote kullan (Türkçe karakter desteği)
        from urllib.parse import unquote
        part_code = unquote(part_code)
        
        # Kullanıcının düzenleme yetkisini kontrol et
        user_id = session.get('user_id')
        conn = get_db()
        cursor = conn.cursor()
        
        cursor = execute_query(cursor, 
            'SELECT can_edit_parts_info FROM envanter_users WHERE id = %s', 
            (user_id,))
        user_data = cursor.fetchone()
        
        if not user_data or not user_data[0]:
            close_db(conn)
            return jsonify({'error': 'Bu işlem için yetkiniz yok'}), 403
        
        data = request.get_json()
        
        cursor = execute_query(cursor, '''
            UPDATE parts_info
            SET description = %s,
                machines = %s,
                specifications = %s
            WHERE part_code = %s
        ''', (
            data.get('description', ''),
            data.get('machines', ''),
            data.get('specifications', ''),
            part_code
        ))
        
        conn.commit()
        close_db(conn)
        
        return jsonify({'success': True, 'message': 'Parça bilgileri güncellendi'})
        
    except Exception as e:
        print(f"Part update error: {e}")
        return jsonify({'error': str(e)}), 500

@app.route('/api/parts_info/upload_photo', methods=['POST'])
@login_required
def parts_info_upload_photo():
    """Parça fotoğrafı yükle - Parts Info sistemi için - Ortak klasör kullanır"""
    try:
        if 'photo' not in request.files:
            return jsonify({'error': 'Fotoğraf bulunamadı'}), 400
        
        photo = request.files['photo']
        part_code = request.form.get('part_code')
        
        if not photo or not part_code:
            return jsonify({'error': 'Eksik bilgi'}), 400
        
        # Dosya uzantısını kontrol et
        allowed_extensions = {'png', 'jpg', 'jpeg', 'gif', 'webp'}
        file_ext = photo.filename.rsplit('.', 1)[1].lower() if '.' in photo.filename else ''
        
        if file_ext not in allowed_extensions:
            return jsonify({'error': 'Geçersiz dosya tipi'}), 400
        
        # Dosyayı ortak klasöre kaydet (PART_PHOTOS_DIR)
        import os
        from werkzeug.utils import secure_filename
        
        os.makedirs(PART_PHOTOS_DIR, exist_ok=True)
        
        filename = f"{secure_filename(part_code)}_{int(time.time())}.{file_ext}"
        filepath = os.path.join(PART_PHOTOS_DIR, filename)
        photo.save(filepath)
        
        # Veritabanını güncelle - sadece filename (part_photos klasörü baz alınacak)
        photo_path = f"part_photos/{filename}"
        
        conn = get_db()
        cursor = conn.cursor()
        
        cursor = execute_query(cursor, '''
            UPDATE parts_info
            SET photo_path = %s
            WHERE part_code = %s
        ''', (photo_path, part_code))
        
        conn.commit()
        close_db(conn)
        
        return jsonify({'success': True, 'photo_path': photo_path})
        
    except Exception as e:
        print(f"Photo upload error: {e}")
        return jsonify({'error': str(e)}), 500

@app.route('/api/parts_info/upload_photo/<path:part_code>', methods=['POST'])
@login_required
def parts_info_upload_photo_by_code(part_code):
    """Parça fotoğrafı yükle - URL'den part_code alır"""
    try:
        from urllib.parse import unquote
        part_code = unquote(part_code)
        
        if 'photo' not in request.files:
            return jsonify({'error': 'Fotoğraf bulunamadı'}), 400
        
        photo = request.files['photo']
        
        if not photo:
            return jsonify({'error': 'Fotoğraf yüklenemedi'}), 400
        
        # Dosya uzantısını kontrol et
        allowed_extensions = {'png', 'jpg', 'jpeg', 'gif', 'webp'}
        file_ext = photo.filename.rsplit('.', 1)[1].lower() if '.' in photo.filename else ''
        
        if file_ext not in allowed_extensions:
            return jsonify({'error': 'Geçersiz dosya tipi. Sadece resim dosyaları yüklenebilir.'}), 400
        
        # Dosyayı ortak klasöre kaydet (PART_PHOTOS_DIR)
        import os
        from werkzeug.utils import secure_filename
        
        os.makedirs(PART_PHOTOS_DIR, exist_ok=True)
        
        filename = f"{secure_filename(part_code)}_{int(time.time())}.{file_ext}"
        filepath = os.path.join(PART_PHOTOS_DIR, filename)
        photo.save(filepath)
        
        # Veritabanını güncelle - sadece filename (part_photos klasörü baz alınacak)
        photo_path = f"part_photos/{filename}"
        
        conn = get_db()
        cursor = conn.cursor()
        
        cursor = execute_query(cursor, '''
            UPDATE parts_info
            SET photo_path = %s,
                updated_at = NOW()
            WHERE part_code = %s
        ''', (photo_path, part_code))
        
        conn.commit()
        close_db(conn)
        
        return jsonify({'success': True, 'photo_path': photo_path})
        
    except Exception as e:
        print(f"Photo upload error: {e}")
        return jsonify({'error': str(e)}), 500

@app.route('/api/parts_info/update_field', methods=['POST'])
@login_required
def update_parts_info_field():
    """Parça bilgilerini güncelle - sadece admin yetkisi olanlar"""
    try:
        # Yetki kontrolü
        if session.get('role') != 'admin':
            return jsonify({'error': 'Yetkiniz yok'}), 403
        
        data = request.json
        part_code = data.get('part_code')
        field = data.get('field')  # 'description' veya 'supplier'
        value = data.get('value', '')
        
        if not part_code or not field:
            return jsonify({'error': 'Eksik parametre'}), 400
        
        if field not in ['description', 'supplier']:
            return jsonify({'error': 'Geçersiz alan'}), 400
        
        conn = get_db()
        cursor = conn.cursor()
        
        # parts_info tablosunda var mı kontrol et
        cursor = execute_query(cursor, 'SELECT id FROM parts_info WHERE part_code = %s', (part_code,))
        existing = cursor.fetchone()
        
        if existing:
            # Güncelle
            cursor = execute_query(cursor, f'''
                UPDATE parts_info 
                SET {field} = %s,
                    updated_at = NOW()
                WHERE part_code = %s
            ''', (value, part_code))
        else:
            # Yeni kayıt oluştur
            cursor = execute_query(cursor, 'SELECT part_name FROM part_codes WHERE UPPER(part_code) = %s', (part_code.upper(),))
            part_name_row = cursor.fetchone()
            part_name = part_name_row[0] if part_name_row else ''
            
            if field == 'description':
                cursor = execute_query(cursor, '''
                    INSERT INTO parts_info (part_code, part_name, description, created_at, updated_at)
                    VALUES (%s, %s, %s, NOW(), NOW())
                ''', (part_code, part_name, value))
            else:  # supplier
                cursor = execute_query(cursor, '''
                    INSERT INTO parts_info (part_code, part_name, supplier, created_at, updated_at)
                    VALUES (%s, %s, %s, NOW(), NOW())
                ''', (part_code, part_name, value))
        
        conn.commit()
        close_db(conn)
        
        return jsonify({'success': True, 'message': 'Güncellendi'})
        
    except Exception as e:
        print(f"Update error: {e}")
        return jsonify({'error': str(e)}), 500

@app.route('/api/parts_info/get_exchange_rate')
@login_required
def get_exchange_rate_api():
    """Güncel kur bilgisini getir"""
    try:
        rate = get_tcmb_exchange_rate()
        source = getattr(get_tcmb_exchange_rate, 'last_source', 'Yapıkredi API')
        return jsonify({
            'rate': rate, 
            'success': True,
            'source': source
        })
    except Exception as e:
        return jsonify({
            'rate': 35.0, 
            'success': False, 
            'error': str(e),
            'source': 'Varsayılan'
        })

@app.route('/api/parts_info/search')
@login_required
def parts_info_search():
    """Parça arama API - Türkçe karakter destekli detaylı arama"""
    query = request.args.get('q', '').strip()
    
    if not query:
        return jsonify([])
    
    try:
        # Türkçe karakterleri normalize et
        import unicodedata
        
        def normalize_turkish(text):
            """Türkçe karakterleri normalize eder - büyük/küçük harf duyarsız"""
            if not text:
                return ''
            # Türkçe karakterleri koruyarak lowercase
            tr_map = str.maketrans('ıİ', 'iI')
            text = text.translate(tr_map).lower()
            return text
        
        # Kullanıcının geliş fiyatı görme yetkisini kontrol et
        conn = get_db()
        cursor = conn.cursor()
        
        user_id = session.get('user_id')
        cursor = execute_query(cursor, 
            'SELECT can_view_purchase_price FROM envanter_users WHERE id = %s', 
            (user_id,))
        user_data = cursor.fetchone()
        can_view_purchase = user_data[0] if user_data else False
        
        # Arama terimlerini kelimelerine ayır
        search_terms = normalize_turkish(query).split()
        
        # Her kelime için ayrı LIKE koşulu oluştur
        conditions = []
        params = []
        
        for term in search_terms:
            # Her terimi hem part_code hem part_name hem description'da ara
            conditions.append('''
                (LOWER(part_code) LIKE %s OR 
                 LOWER(part_name) LIKE %s OR 
                 LOWER(description) LIKE %s OR
                 LOWER(supplier) LIKE %s)
            ''')
            search_pattern = f'%{term}%'
            params.extend([search_pattern, search_pattern, search_pattern, search_pattern])
        
        where_clause = ' AND '.join(conditions) if conditions else '1=1'
        
        cursor = execute_query(cursor, f'''
            SELECT 
                part_code,
                part_name,
                stock,
                critical_stock,
                expected_stock,
                supplier,
                purchase_price_eur,
                description,
                sale_price_eur,
                photo_path,
                stock_location,
                machines,
                specifications,
                replacement_code,
                build_out
            FROM parts_info
            WHERE {where_clause}
            ORDER BY part_code
            LIMIT 100
        ''', tuple(params))
        
        results = cursor.fetchall()
        print(f"[PARTS_SEARCH] Sorgu sonucu: {len(results)} parça bulundu")
        if results and len(results) > 0:
            print(f"[PARTS_SEARCH] İlk satırda {len(results[0])} kolon var")
        
        # EUR -> TRY dönüşümü için kur al
        try:
            exchange_rate = get_tcmb_exchange_rate()
        except:
            exchange_rate = 35.0  # Varsayılan kur
        
        parts = []
        for row in results:
            try:
                part = {
                    'part_code': row[0],
                    'part_name': row[1],
                    'stock': row[2] or 0,
                    'critical_stock': row[3] or 0,
                    'expected_stock': row[4] or 0,
                    'supplier': row[5] or '',
                    'purchase_price_eur': float(row[6]) if (row[6] and can_view_purchase) else None,
                    'description': row[7] or '',
                    'sale_price_eur': float(row[8]) if row[8] else 0,
                    'photo_path': row[9] or '',
                    'stock_location': row[10] or '',
                    'machines': row[11] or '',
                    'specifications': row[12] or '',
                    'replacement_code': row[13] or '' if len(row) > 13 else '',
                    'build_out': bool(row[14]) if (len(row) > 14 and row[14]) else False,
                    'exchange_rate': exchange_rate,
                    'purchase_price_try': round(float(row[6] or 0) * exchange_rate, 2) if can_view_purchase else None,
                    'sale_price_try': round(float(row[8] or 0) * exchange_rate, 2),
                    'can_view_purchase': can_view_purchase
                }
                parts.append(part)
            except IndexError as ie:
                print(f"[PARTS_SEARCH] IndexError: {ie} - row length: {len(row)}")
                raise
        
        print(f"[PARTS_SEARCH] {len(parts)} parça işlendi ve JSON dönüldü")
        return jsonify(parts)
        
    except Exception as e:
        print(f"Parts info search error: {e}")
        return jsonify({'error': str(e)}), 500

@app.route('/api/parts_info/upload_excel', methods=['POST'])
@login_required
def upload_parts_info_excel():
    """Excel dosyasından parça bilgilerini yükle"""
    if 'file' not in request.files:
        return jsonify({'error': 'Dosya bulunamadı'}), 400
    
    file = request.files['file']
    if file.filename == '':
        return jsonify({'error': 'Dosya seçilmedi'}), 400
    
    if not file.filename.endswith(('.xlsx', '.xls')):
        return jsonify({'error': 'Sadece Excel dosyaları yüklenebilir'}), 400
    
    try:
        import pandas as pd
        import io
        
        # Excel dosyasını oku
        df = pd.read_excel(io.BytesIO(file.read()))
        
        # Beklenen sütunlar
        required_columns = ['Parça Kodu', 'Parça Adı', 'Stok', 'Tedarikçi', 'Geliş (Euro)', 'Tanım', 'Satış Fiyatı (EUR)']
        
        # Sütun kontrolü
        missing_columns = [col for col in required_columns if col not in df.columns]
        if missing_columns:
            return jsonify({'error': f'Eksik sütunlar: {", ".join(missing_columns)}'}), 400
        
        conn = get_db()
        cursor = conn.cursor()
        
        inserted_count = 0
        updated_count = 0
        
        for _, row in df.iterrows():
            part_code = str(row['Parça Kodu']).strip()
            part_name = str(row['Parça Adı']).strip()
            stock = int(row['Stok']) if pd.notna(row['Stok']) else 0
            supplier = str(row['Tedarikçi']).strip() if pd.notna(row['Tedarikçi']) else ''
            purchase_price = float(row['Geliş (Euro)']) if pd.notna(row['Geliş (Euro)']) else 0.0
            description = str(row['Tanım']).strip() if pd.notna(row['Tanım']) else ''
            sale_price = float(row['Satış Fiyatı (EUR)']) if pd.notna(row['Satış Fiyatı (EUR)']) else 0.0
            
            # Opsiyonel sütunlar
            critical_stock = int(row.get('Kritik stok', 0)) if pd.notna(row.get('Kritik stok')) else 0
            expected_stock = int(row.get('Beklenen stok', 0)) if pd.notna(row.get('Beklenen stok')) else 0
            
            # Var olan kaydı kontrol et
            cursor = execute_query(cursor, 'SELECT id FROM parts_info WHERE part_code = %s', (part_code,))
            existing = cursor.fetchone()
            
            if existing:
                # Güncelle
                cursor = execute_query(cursor, '''
                    UPDATE parts_info 
                    SET part_name = %s,
                        stock = %s,
                        critical_stock = %s,
                        expected_stock = %s,
                        supplier = %s,
                        purchase_price_eur = %s,
                        description = %s,
                        sale_price_eur = %s,
                        updated_at = NOW()
                    WHERE part_code = %s
                ''', (part_name, stock, critical_stock, expected_stock, supplier, 
                      purchase_price, description, sale_price, part_code))
                updated_count += 1
            else:
                # Yeni kayıt ekle
                cursor = execute_query(cursor, '''
                    INSERT INTO parts_info 
                    (part_code, part_name, stock, critical_stock, expected_stock, 
                     supplier, purchase_price_eur, description, sale_price_eur, created_at)
                    VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, NOW())
                ''', (part_code, part_name, stock, critical_stock, expected_stock,
                      supplier, purchase_price, description, sale_price))
                inserted_count += 1
        
        conn.commit()
        
        return jsonify({
            'success': True,
            'inserted': inserted_count,
            'updated': updated_count,
            'total': inserted_count + updated_count
        })
        
    except Exception as e:
        print(f"Excel upload error: {e}")
        return jsonify({'error': f'Excel yükleme hatası: {str(e)}'}), 500

@app.route('/api/parts_info/get_all')
@login_required
def get_all_parts_info():
    """Tüm parça bilgilerini getir"""
    try:
        # Kullanıcının geliş fiyatı görme yetkisini kontrol et
        conn = get_db()
        cursor = conn.cursor()
        
        user_id = session.get('user_id')
        cursor = execute_query(cursor, 
            'SELECT can_view_purchase_price FROM envanter_users WHERE id = %s', 
            (user_id,))
        user_data = cursor.fetchone()
        can_view_purchase = user_data[0] if user_data else False
        
        cursor = execute_query(cursor, '''
            SELECT 
                part_code,
                part_name,
                stock,
                critical_stock,
                expected_stock,
                supplier,
                purchase_price_eur,
                description,
                sale_price_eur,
                photo_path,
                replacement_code,
                build_out
            FROM parts_info
            ORDER BY part_code
        ''')
        
        results = cursor.fetchall()
        
        # EUR -> TRY kur bilgisi
        try:
            exchange_rate = get_tcmb_exchange_rate()
        except:
            exchange_rate = 35.0
        
        parts = []
        for row in results:
            part = {
                'part_code': row[0],
                'part_name': row[1],
                'stock': row[2] or 0,
                'critical_stock': row[3] or 0,
                'expected_stock': row[4] or 0,
                'supplier': row[5] or '',
                'purchase_price_eur': float(row[6]) if (row[6] and can_view_purchase) else None,
                'description': row[7] or '',
                'sale_price_eur': float(row[8]) if row[8] else 0,
                'photo_path': row[9] or '',
                'replacement_code': row[10] or '',
                'build_out': bool(row[11]) if row[11] else False,
                'exchange_rate': exchange_rate,
                'purchase_price_try': round(float(row[6] or 0) * exchange_rate, 2) if can_view_purchase else None,
                'sale_price_try': round(float(row[8] or 0) * exchange_rate, 2),
                'can_view_purchase': can_view_purchase
            }
            parts.append(part)
        
        return jsonify(parts)
        
    except Exception as e:
        print(f"Get all parts info error: {e}")
        return jsonify({'error': str(e)}), 500

@app.route('/api/parts_info/debug/schema')
@login_required
def debug_parts_schema():
    """DEBUG: Parts Info tablo şemasını göster"""
    try:
        conn = get_db()
        cursor = conn.cursor()
        
        cursor = execute_query(cursor, 'DESCRIBE parts_info')
        columns = cursor.fetchall()
        
        schema = []
        for col in columns:
            schema.append({
                'name': col[0],
                'type': col[1],
                'null': col[2],
                'key': col[3],
                'default': col[4],
                'extra': col[5]
            })
        
        print(f"[DEBUG] Parts Info schema: {schema}")
        return jsonify({
            'status': 'ok',
            'columns': schema,
            'column_count': len(schema)
        })
        
    except Exception as e:
        print(f"[DEBUG] Schema sorgu hatası: {e}")
        return jsonify({'error': str(e)}), 500

@app.route('/api/parts_info/debug/test_search')
@login_required
def debug_test_search():
    """DEBUG: İlk satırı görmek için test ara"""
    try:
        conn = get_db()
        cursor = conn.cursor()
        
        cursor = execute_query(cursor, '''
            SELECT * FROM parts_info LIMIT 1
        ''')
        
        result = cursor.fetchone()
        if result:
            print(f"[DEBUG] Test row length: {len(result)}")
            print(f"[DEBUG] Test row: {result}")
            return jsonify({
                'status': 'ok',
                'row_length': len(result),
                'row_data': str(result)[:500]  # İlk 500 karakter
            })
        else:
            return jsonify({'status': 'ok', 'message': 'No data in parts_info'})
        
    except Exception as e:
        print(f"[DEBUG] Test search hatası: {e}")
        return jsonify({'error': str(e)}), 500

@app.route('/api/parts_info/upload_location', methods=['POST'])
@login_required
def upload_parts_location():
    """Excel dosyasından parça stok lokasyonlarını güncelle"""
    try:
        if 'file' not in request.files:
            return jsonify({'error': 'Dosya bulunamadı'}), 400
        
        file = request.files['file']
        if file.filename == '':
            return jsonify({'error': 'Dosya seçilmedi'}), 400
        
        if not file.filename.endswith(('.xlsx', '.xls')):
            return jsonify({'error': 'Sadece Excel dosyaları yüklenebilir'}), 400
        
        import pandas as pd
        import io
        
        # Excel dosyasını oku
        df = pd.read_excel(io.BytesIO(file.read()))
        
        # Sütun isimlerini temizle ve normalize et
        df.columns = df.columns.str.strip()
        
        # Beklenen sütunlar (zorunlu ve isteğe bağlı)
        required_columns = ['Parça Kodu']  # Sadece Parça Kodu zorunlu
        optional_columns = ['Stok Lokasyonu', 'Build Out', 'Replacement Kodu', 'Replacement Code', 'Değişen Parça Kodu']
        
        # Sütun kontrolü - case insensitive mapping oluştur
        column_mapping = {}
        df_columns_lower = {col.lower(): col for col in df.columns}
        
        # Zorunlu sütunları kontrol et
        missing_columns = []
        for req_col in required_columns:
            req_col_lower = req_col.lower()
            if req_col_lower in df_columns_lower:
                column_mapping[df_columns_lower[req_col_lower]] = req_col
            else:
                missing_columns.append(req_col)
        
        if missing_columns:
            available_cols = ', '.join(df.columns.tolist())
            return jsonify({
                'error': f'Eksik sütunlar: {", ".join(missing_columns)}. Mevcut sütunlar: {available_cols}'
            }), 400
        
        # İsteğe bağlı sütunları map et
        for opt_col in optional_columns:
            opt_col_lower = opt_col.lower()
            if opt_col_lower in df_columns_lower:
                # Standardize et
                if 'replacement' in opt_col_lower or 'değişen' in opt_col_lower:
                    if 'replacement code' not in [k.lower() for k in column_mapping.values()]:
                        column_mapping[df_columns_lower[opt_col_lower]] = 'Replacement Kodu'
                elif 'build out' in opt_col_lower:
                    if 'build out' not in [k.lower() for k in column_mapping.values()]:
                        column_mapping[df_columns_lower[opt_col_lower]] = 'Build Out'
                elif 'lokasyon' in opt_col_lower or 'location' in opt_col_lower:
                    if 'stok lokasyonu' not in [k.lower() for k in column_mapping.values()]:
                        column_mapping[df_columns_lower[opt_col_lower]] = 'Stok Lokasyonu'
        
        # Sütunları standart isimlere dönüştür
        df = df.rename(columns=column_mapping)
        
        conn = get_db()
        cursor = conn.cursor()
        
        updated_count = 0
        not_found_count = 0
        not_found_parts = []
        
        for _, row in df.iterrows():
            part_code = str(row['Parça Kodu']).strip().upper()  # Büyük harfe çevir
            stock_location = str(row.get('Stok Lokasyonu', '')).strip() if pd.notna(row.get('Stok Lokasyonu', '')) else ''
            
            # Build Out kontrolü - "Build Out" text varsa 1 yap
            build_out = 0
            if 'Build Out' in row and pd.notna(row['Build Out']):
                build_out_value = str(row['Build Out']).strip().lower()
                if build_out_value in ['build out', 'yes', 'evet', '1', 'true']:
                    build_out = 1
            
            # Replacement Code
            replacement_code = str(row.get('Replacement Kodu', '')).strip() if pd.notna(row.get('Replacement Kodu', '')) else ''
            
            # Boş satırları atla
            if not part_code or part_code == 'NAN':
                continue
            
            print(f"[EXCEL UPLOAD] Processing: {part_code} | Build Out: {build_out} | Replacement: {replacement_code}")
            
            # parts_info tablosunda ara
            cursor = execute_query(cursor, 'SELECT id, part_code FROM parts_info WHERE UPPER(part_code) = %s', (part_code,))
            existing = cursor.fetchone()
            
            if existing:
                # Güncelle - part_name'i de güncelle (part_codes'dan al)
                actual_part_code = existing[1]
                print(f"[EXCEL UPLOAD] Found in parts_info: {actual_part_code} (ID: {existing[0]})")
                
                # part_codes tablosundan part_name al (varsa)
                cursor = execute_query(cursor, 'SELECT part_name FROM part_codes WHERE UPPER(part_code) = %s', (part_code,))
                part_name_row = cursor.fetchone()
                part_name = part_name_row[0] if part_name_row else ''
                
                # UPDATE - tüm sütunları güncelle
                cursor = execute_query(cursor, '''
                    UPDATE parts_info 
                    SET stock_location = %s,
                        part_name = %s,
                        replacement_code = %s,
                        build_out = %s,
                        updated_at = NOW()
                    WHERE part_code = %s
                ''', (stock_location, part_name, replacement_code if replacement_code else None, build_out, actual_part_code))
                print(f"[EXCEL UPLOAD] Updated: {actual_part_code} | Name: {part_name} | BuildOut: {build_out} | Repl: {replacement_code}")
                updated_count += 1
            else:
                # Yeni kayıt ekle - Önce part_codes'da var mı kontrol et
                print(f"[EXCEL UPLOAD] Creating new entry for: {part_code}")
                
                # part_codes tablosundan part_name al (varsa)
                cursor = execute_query(cursor, 'SELECT part_name FROM part_codes WHERE UPPER(part_code) = %s', (part_code,))
                part_name_row = cursor.fetchone()
                part_name = part_name_row[0] if part_name_row else ''
                
                cursor = execute_query(cursor, '''
                    INSERT INTO parts_info (part_code, part_name, stock_location, replacement_code, build_out, created_at, updated_at)
                    VALUES (%s, %s, %s, %s, %s, NOW(), NOW())
                ''', (part_code, part_name, stock_location, replacement_code if replacement_code else None, build_out))
                print(f"[EXCEL UPLOAD] Created: {part_code} | Name: {part_name} | BuildOut: {build_out} | Repl: {replacement_code}")
                updated_count += 1
        
        conn.commit()
        close_db(conn)
        
        return jsonify({
            'success': True,
            'updated': updated_count,
            'not_found': not_found_count,
            'not_found_parts': not_found_parts[:10]  # İlk 10 bulunamayan parçayı göster
        })
        
    except Exception as e:
        import traceback
        error_details = traceback.format_exc()
        print(f"Location upload error: {error_details}")
        return jsonify({'error': f'Lokasyon yükleme hatası: {str(e)}'}), 500

def get_access_token():
    """Yapıkredi OAuth2 token al"""
    try:
        import requests
        
        client_id = os.environ.get('YAPIKREDI_CLIENT_ID')
        client_secret = os.environ.get('YAPIKREDI_CLIENT_SECRET')
        
        if not client_id or not client_secret or client_id == 'your_client_id_here':
            print("[EXCHANGE RATE] ⚠️ Yapıkredi credentials .env dosyasında ayarlanmamış")
            raise Exception("Missing Yapıkredi credentials")
        
        api_url = os.environ.get('YAPIKREDI_API_URL', 'https://api.yapikredi.com.tr')
        
        token_url = f"{api_url}/oauth/v1/token"
        
        print(f"[EXCHANGE RATE] 🔑 Token isteniyor: {token_url}")
        
        payload = {
            'grant_type': 'client_credentials',
            'client_id': client_id,
            'client_secret': client_secret
        }
        
        headers = {
            'Content-Type': 'application/x-www-form-urlencoded'
        }
        
        response = requests.post(token_url, data=payload, headers=headers, timeout=10)
        
        print(f"[EXCHANGE RATE] Token Status: {response.status_code}")
        
        if response.status_code == 200:
            data = response.json()
            access_token = data.get('access_token')
            if access_token:
                print("[EXCHANGE RATE] ✅ Token başarıyla alındı")
                return access_token
            else:
                raise Exception("Token response'da access_token bulunamadı")
        else:
            print(f"[EXCHANGE RATE] ❌ Token hatası: {response.text}")
            raise Exception(f"Token request failed: {response.status_code}")
    
    except Exception as e:
        print(f"[EXCHANGE RATE] ❌ Token alma hatası: {e}")
        raise


def get_tcmb_exchange_rate():
    """Yapıkredi API'den EUR/TRY kurunu çek - OAuth2 ile"""
    try:
        import requests
        
        print("[EXCHANGE RATE] 🔄 Yapıkredi API'ye bağlanılıyor...")
        
        # 3 kez deneme (retry logic)
        for attempt in range(1, 4):
            try:
                # OAuth token al
                token = get_access_token()
                
                # API endpoint
                api_url = os.environ.get('YAPIKREDI_API_URL', 'https://api.yapikredi.com.tr')
                url = f"{api_url}/foreign-currency/v1/currency-rates"
                
                headers = {
                    "Authorization": f"Bearer {token}",
                    "Accept": "application/json",
                    "User-Agent": "CermakEnvanterQR/2.0"
                }
                
                print(f"[EXCHANGE RATE] Deneme {attempt}/3 - URL: {url}")
                
                response = requests.get(url, headers=headers, timeout=10)
                
                print(f"[EXCHANGE RATE] Status Code: {response.status_code}")
                print(f"[EXCHANGE RATE] Raw response: {response.text[:500]}")  # İlk 500 char göster
                
                response.raise_for_status()
                data = response.json()
                
                # EUR kuru ara
                for rate in data.get("exchangeRates", []):
                    if rate.get("currencyCode") == "EUR":
                        selling_rate = float(rate.get("sellingRate"))
                        print(f"[EXCHANGE RATE] ✅✅✅ YAPIKREDI'DEN EUR/TRY ALINDI: {selling_rate} TRY/EUR")
                        print(f"[EXCHANGE RATE] Tam Response: {rate}")
                        
                        # Kaynağı kaydet
                        get_tcmb_exchange_rate.last_source = '✅ Yapıkredi API (OAuth2)'
                        
                        return selling_rate
                
                # EUR bulunamazsa
                print(f"[EXCHANGE RATE] ❌ EUR kuru response içinde bulunamadı")
                print(f"[EXCHANGE RATE] Available rates: {[r.get('currencyCode') for r in data.get('exchangeRates', [])]}")
                
            except requests.exceptions.RequestException as e:
                print(f"[EXCHANGE RATE] ❌ Deneme {attempt}/3 başarısız: {str(e)}")
                if attempt < 3:
                    import time
                    time.sleep(2)
                    continue
            except Exception as e:
                print(f"[EXCHANGE RATE] ❌ Deneme {attempt}/3 hatası: {str(e)}")
                if attempt < 3:
                    import time
                    time.sleep(2)
                    continue
        
        # 3 deneme başarısız
        print("[EXCHANGE RATE] ❌❌❌ UYARI: Yapıkredi API'den kur alınamadı!")
        print("[EXCHANGE RATE] Lütfen:")
        print("  1. Yapıkredi credentials'ını .env dosyasında kontrol edin")
        print("  2. API endpoint'ini doğrulayın")
        print("  3. Network bağlantısını kontrol edin")
        get_tcmb_exchange_rate.last_source = '❌ API Hatası'
        raise Exception("Yapıkredi API başarısız")
    
    except Exception as e:
        print(f"[EXCHANGE RATE] ❌ HATA: {str(e)}")
        print("[EXCHANGE RATE] ⚠️ VARSAYILAN KUR KULLANILIYOR: 35.0 TRY/EUR")
        get_tcmb_exchange_rate.last_source = '⚠️ Varsayılan'
        return 35.0


# Global veri saklamak için
get_tcmb_exchange_rate.last_source = 'Yapıkredi API (OAuth2)'

@app.route('/api/parts_info/sync_photos', methods=['POST'])
@login_required
def sync_parts_info_photos():
    """Admin sistemindeki fotoğrafları parts_info'ya eşitle"""
    try:
        conn = get_db()
        cursor = conn.cursor()
        
        # Admin part_codes tablosundan fotoğraf bilgilerini al
        cursor = execute_query(cursor, '''
            SELECT part_code, photo_path, catalog_image
            FROM part_codes
            WHERE photo_path IS NOT NULL OR catalog_image IS NOT NULL
        ''')
        
        admin_photos = cursor.fetchall()
        synced_count = 0
        
        for row in admin_photos:
            part_code = row[0]
            photo_path = row[1] or row[2]  # photo_path öncelikli, yoksa catalog_image
            
            if photo_path:
                # parts_info tablosunda bu parça varsa fotoğrafı güncelle
                cursor = execute_query(cursor, '''
                    UPDATE parts_info
                    SET photo_path = %s
                    WHERE part_code = %s
                ''', (photo_path, part_code))
                
                if cursor.rowcount > 0:
                    synced_count += 1
        
        conn.commit()
        
        return jsonify({
            'success': True,
            'synced': synced_count,
            'message': f'{synced_count} parçanın fotoğrafı eşitlendi'
        })
        
    except Exception as e:
        print(f"Photo sync error: {e}")
        return jsonify({'error': str(e)}), 500


@app.route('/dashboard')

@login_required

def dashboard():

    """Dashboard - Ana sayfaya ynlendir"""

    return redirect('/')





@app.route('/clear_cache.html')

def clear_cache_page():

    """Cache temizleme yardmc sayfas"""

    with open('clear_cache.html', 'r', encoding='utf-8') as f:

        return f.read()





# Note: a more comprehensive /health endpoint is defined later in this file.

# The detailed health check includes DB and storage checks and will be used by

# load balancers and Render readiness probes.



@app.route('/api/dashboard/stats')

@login_required

def dashboard_stats():

    """Cache'li dashboard istatistikleri"""

    cache_key = 'dashboard_stats'



    # Cache'den kontrol et

    cached_data = cache_get(cache_key)

    if cached_data:

        return jsonify(cached_data)



    # Cache yoksa veritabanndan al - ORM implementation

    try:

        total_qr_codes = db.session.query(QRCode).count()

        total_scans = db.session.query(ScannedQR).count()

        active_sessions = db.session.query(CountSession).filter_by(is_active=True).count()

        last_scan_obj = db.session.query(ScannedQR).order_by(ScannedQR.scanned_at.desc()).first()
        last_scan = last_scan_obj.scanned_at if last_scan_obj else None

        from datetime import date
        today_start = datetime.combine(date.today(), datetime.min.time())
        today_scans = db.session.query(ScannedQR).filter(ScannedQR.scanned_at >= today_start).count()

        from datetime import timedelta
        week_ago = datetime.now() - timedelta(days=7)
        week_scans = db.session.query(ScannedQR).filter(ScannedQR.scanned_at >= week_ago).count()

        stats = {

            'total_qr_codes': total_qr_codes,

            'total_scans': total_scans,

            'active_sessions': active_sessions,

            'last_scan': last_scan.isoformat() if last_scan else None,

            'today_scans': today_scans,

            'week_scans': week_scans,

            'cache_time': datetime.now().isoformat()

        }



        # Cache'e kaydet

        cache_set(cache_key, stats)



        return jsonify(stats)



    finally:

        pass



# ========================================

#  QR SCANNER SAYIM SSTEM (Sadece Desktop Scanner)

# ========================================



@app.route('/scanner')

@login_required

def scanner_page():

    """Desktop QR Scanner Saym Sayfas"""

    return render_template('scanner.html')



@app.route('/live_dashboard')

@login_required

def live_dashboard_page():

    """Canl Dashboard Sayfas"""

    return render_template('live_dashboard.html')



@app.route('/api/get_active_count_session')

def get_active_count_session():

    """Aktif saym seansn kontrol et - Public endpoint (landing page iin)"""

    try:

        conn = get_db()

        cursor = conn.cursor()

        

        cursor.execute('''

            SELECT id, session_name, created_at, started_at, total_expected, total_scanned, description

            FROM count_sessions 

            WHERE is_active = %s

            ORDER BY created_at DESC

            LIMIT 1

        ''', (True,))

        

        result = cursor.fetchone()

        close_db(conn)

        

        if result:

            session_id, name, created_at, started_at, expected, scanned, description = result

            

            # Sre hesapla - PostgreSQL datetime'ı zaten Python datetime nesnesi

            # SQLite'de string idi ama psycopg2 otomatik convert ediyor

            if isinstance(started_at, str):

                start_time = datetime.fromisoformat(started_at) if started_at else datetime.fromisoformat(created_at)

            else:

                # Already a datetime object from PostgreSQL

                start_time = started_at if started_at else created_at

            duration_seconds = (datetime.now() - start_time).total_seconds()

            

            # Yzde hesapla

            percentage = (scanned / expected * 100) if expected > 0 else 0

            

            # 🚀 PERFORMANCE: Beklenen parçaları cache'den al (3845 parçalık JSON'u sürekli parse etmeyelim!)
            expected_parts = []
            
            # Cache kontrolü - TTL yok, sadece session_id eşleşmesi kontrol et
            if (ACTIVE_SESSION_CACHE['session_id'] == session_id and 
                ACTIVE_SESSION_CACHE['expected_parts']):
                # ✅ Cache HIT - Parse etmeye gerek yok!
                expected_parts = ACTIVE_SESSION_CACHE['expected_parts']
                logging.debug(f"✅ Cache HIT for session {session_id}")
            else:
                # ❌ Cache MISS - Parse et ve cache'e al
                if description:
                    try:
                        expected_parts = json.loads(description)
                        logging.info(f"🔄 Caching session {session_id} - Expected parts: {len(expected_parts)} items (cache until session ends)")
                        
                        # Cache'e kaydet (sayım bitene kadar geçerli!)
                        ACTIVE_SESSION_CACHE['session_id'] = session_id
                        ACTIVE_SESSION_CACHE['expected_parts'] = expected_parts
                        ACTIVE_SESSION_CACHE['cache_time'] = time.time()  # Bilgi amaçlı
                    except Exception as e:
                        logging.error(f"Description parse error in active session: {e}")
                        pass
            
            return jsonify({
                'success': True,
                'has_active': True,
                'session': {
                    'id': session_id,
                    'name': name,
                    'created_at': created_at,
                    'started_at': started_at,
                    'duration_seconds': int(duration_seconds),
                    'total_expected': expected or 0,
                    'total_scanned': scanned or 0,
                    'percentage': round(percentage, 2),
                    'expected_parts': expected_parts
                }
            })

        else:

            return jsonify({

                'success': True,

                'has_active': False

            })

        

    except Exception as e:

        print(f" Get active session error: {e}")

        return jsonify({'success': False, 'error': str(e)}), 500





@app.route('/api/get_session_report/<session_id>')

@login_required

def get_session_report(session_id):

    """Saym raporu getir - beklenen vs saylan karlatrmas"""

    try:

        conn = get_db()

        cursor = conn.cursor()

        

        # Session bilgilerini al

        execute_query(cursor, '''

            SELECT session_name, description, total_expected, total_scanned

            FROM count_sessions 

            WHERE id = %s

        ''', (session_id,))

        

        session_info = cursor.fetchone()

        

        if not session_info:

            close_db(conn)

            return jsonify({'success': False, 'error': 'Session bulunamad'}), 404

        

        session_name, description, total_expected, total_scanned = session_info

        

        # Beklenen paralar parse et

        expected_parts = []

        if description:

            try:

                expected_parts = json.loads(description)

                logging.info(f"Session {session_id} - Expected parts parsed: {len(expected_parts)} items")

                if expected_parts and len(expected_parts) > 0:

                    logging.info(f"First item structure: {expected_parts[0]}")

            except Exception as e:

                logging.error(f"Description parse error: {e}")

                pass

        

        # Sayılan parçaları al - direkt scanned_qr tablosundan (PAKETLERİ HARİÇ TUT)

        execute_query(cursor, '''

            SELECT sq.part_code, COUNT(*) as scanned_count

            FROM scanned_qr sq

            LEFT JOIN part_codes pc ON sq.part_code = pc.part_code

            WHERE sq.session_id = %s

              AND (pc.is_package IS NULL OR pc.is_package = FALSE)

            GROUP BY sq.part_code

        ''', (session_id,))

        

        scanned_parts = {}

        for row in cursor.fetchall():

            part_code, count = row

            scanned_parts[part_code] = count

        

        logging.info(f"Scanned parts: {len(scanned_parts)} unique codes")

        

        # Part name'leri part_codes tablosundan al

        part_names = {}

        if expected_parts:

            # Tm olası field isimleri ile part_code al

            part_codes_list = []

            for item in expected_parts:

                # FIX: Correct field name is 'Parça Kodu' (with ç) not 'Para Kodu'

                pc = item.get('Parça Kodu') or item.get('Para Kodu') or item.get('part_code')

                if pc:

                    part_codes_list.append(pc)

            

            if part_codes_list:

                placeholders = ','.join(['%s' for _ in part_codes_list])

                execute_query(cursor, f'SELECT part_code, part_name FROM part_codes WHERE part_code IN ({placeholders})', part_codes_list)

                for row in cursor.fetchall():

                    part_names[row[0]] = row[1]

        

        close_db(conn)

        

        # Raporu olutur

        report_items = []

        for expected in expected_parts:

            # Tm olası field isimlerini kontrol et - FIX: Parça not Para

            part_code = expected.get('Parça Kodu') or expected.get('Para Kodu') or expected.get('part_code')

            part_name = expected.get('Parça Adı') or expected.get('Para Ad') or expected.get('part_name') or part_names.get(part_code, '')

            

            # Tm olası quantity field isimlerini kontrol et

            expected_qty = (expected.get('expected_quantity') or 

                          expected.get('quantity') or 

                          expected.get('Beklenen Adet') or 

                          expected.get('Adet') or 0)

            

            # part_code veya expected_qty yoksa atla

            if not part_code:

                logging.warning(f"Skipping item with no part_code: {expected}")

                continue

            

            try:

                expected_qty = int(expected_qty)

            except (ValueError, TypeError):

                logging.warning(f"Invalid quantity for {part_code}: {expected_qty}")

                expected_qty = 0

            

            scanned_qty = scanned_parts.get(part_code, 0)

            

            difference = scanned_qty - expected_qty

            

            if difference == 0:

                status = ' Tam'

            elif difference > 0:

                status = ' Fazla'

            else:

                status = ' Eksik'

            

            report_items.append({

                'part_code': part_code,

                'part_name': part_name,

                'expected': expected_qty,

                'scanned': scanned_qty,

                'difference': difference,

                'status': status

            })

        

        return jsonify({

            'success': True,

            'session_name': session_name,

            'total_expected': total_expected,

            'total_scanned': total_scanned,

            'report': report_items

        })

        

    except Exception as e:

        print(f" Get session report error: {e}")

        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/get_recent_scans/<session_id>')
@login_required
def get_recent_scans(session_id):
    """Son taramaları getir - live dashboard için"""
    try:
        conn = get_db()
        cursor = conn.cursor()
        
        execute_query(cursor, '''
            SELECT sq.qr_code, sq.part_code, sq.scanned_at,
                   pc.part_name
            FROM scanned_qr sq
            LEFT JOIN part_codes pc ON sq.part_code = pc.part_code
            WHERE sq.session_id = %s
            ORDER BY sq.scanned_at DESC
            LIMIT 50
        ''', (session_id,))
        
        scans = []
        for row in cursor.fetchall():
            qr_code, part_code, scanned_at, part_name = row
            scans.append({
                'qr_code': qr_code or '',
                'part_code': part_code or '',
                'part_name': part_name or '',
                'scanned_at': scanned_at.strftime('%H:%M:%S') if scanned_at else '',
                'is_duplicate': False
            })
        
        close_db(conn)
        return jsonify({'success': True, 'scans': scans})
        
    except Exception as e:
        logging.error(f"Get recent scans error: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500



@app.route('/api/start_count_session', methods=['POST'])

@login_required

def start_count_session():

    """Yeni saym seans balat"""

    try:

        data = request.get_json()

        expected_parts = data.get('expected_parts', [])  # [{part_code, expected_quantity}]

        

        logging.info(f"Starting count session with {len(expected_parts)} expected parts")

        if expected_parts and len(expected_parts) > 0:

            logging.info(f"First expected part structure: {expected_parts[0]}")

        

        conn = get_db()

        cursor = conn.cursor()

        

        # nce aktif saym var m kontrol et

        execute_query(cursor, '''
            SELECT id FROM count_sessions WHERE is_active = %s LIMIT 1
        ''', (True,))

        if cursor.fetchone():

            close_db(conn)

            return jsonify({

                'success': False,

                'error': 'Zaten aktif bir saym var! nce onu bitirin.'

            }), 400

        

        # Yeni saym seans olutur

        session_name = f"Saym {datetime.now().strftime('%Y%m%d_%H%M%S')}"

        

        # Tm olası field isimlerinden total_expected hesapla

        total_expected = 0

        for item in expected_parts:

            qty = (item.get('expected_quantity') or 

                  item.get('quantity') or 

                  item.get('Beklenen Adet') or 

                  item.get('Adet') or 0)

            try:

                total_expected += int(qty)

            except (ValueError, TypeError):

                logging.warning(f"Invalid quantity in item: {item}")

        

        logging.info(f"Total expected: {total_expected}")

        
        # MySQL INSERT (RETURNING id yerine LAST_INSERT_ID kullan)
        now = datetime.now()

        execute_query(cursor, '''
            INSERT INTO count_sessions 
            (session_name, created_by, is_active, created_at, started_at, total_expected, total_scanned, description)
            VALUES (%s, %s, TRUE, %s, %s, %s, 0, %s)
        ''', (session_name, session.get('user_id', 1), now, now, total_expected, json.dumps(expected_parts) if expected_parts else None))

        # MySQL'de son eklenen ID'yi al
        execute_query(cursor, 'SELECT LAST_INSERT_ID()')
        row = cursor.fetchone()

        session_id = row[0] if row else None

        

        conn.commit()

        close_db(conn)

        

        if not session_id:

            return jsonify({'success': False, 'error': 'Oturum kimliği oluşturulamadı'}), 500

        

        return jsonify({

            'success': True,

            'session_id': session_id,

            'session_name': session_name,

            'message': f' Saym balatld: {session_name}'

        })

        

    except Exception as e:

        logging.error(f"Start count session error: {e}")

        return jsonify({'success': False, 'error': str(e)}), 500





@app.route('/api/finish_count', methods=['POST'])

@login_required

def finish_count():

    """Saym bitir, Excel raporunu kaydet ve rapor hazrla"""

    try:

        data = request.get_json()

        session_id = data.get('session_id', '1')

        

        conn = get_db()

        cursor = conn.cursor()

        

        # Toplam tarama saysn hesapla

        cursor.execute('''
            SELECT COUNT(*) FROM scanned_qr WHERE session_id = %s
        ''', (str(session_id),))

        total_scanned = cursor.fetchone()[0]

        

        # Session bilgilerini al

        execute_query(cursor, '''

            SELECT session_name, created_at, started_at, description, total_expected

            FROM count_sessions 

            WHERE id = %s

        ''', (str(session_id),))

        

        session_info = cursor.fetchone()

        

        if not session_info:

            close_db(conn)

            return jsonify({'success': False, 'error': 'Saym bulunamad'}), 404

        

        session_name, created_at, started_at, description, total_expected = session_info

        

        # Excel raporu olutur

        report_filename = None

        try:

            # Beklenen paralar parse et

            expected_parts = {}

            if description:

                try:

                    expected_list = json.loads(description)

                    for item in expected_list:

                        # FIX: Correct field name is 'Parça Kodu' (with ç) not 'Para Kodu'

                        part_code = item.get('Parça Kodu') or item.get('Para Kodu') or item.get('part_code')

                        expected_qty = item.get('Beklenen Adet') or item.get('expected_quantity') or item.get('quantity') or item.get('Adet') or 0

                        if part_code:

                            expected_parts[part_code] = int(expected_qty)

                except:

                    pass

            

            # Taranan parçaları al (PAKETLERİ HARİÇ TUT)

            execute_query(cursor, '''

                SELECT sq.part_code, pc.part_name, COUNT(*) as scanned_count

                FROM scanned_qr sq

                LEFT JOIN part_codes pc ON sq.part_code = pc.part_code

                WHERE sq.session_id = %s

                  AND (pc.is_package IS NULL OR pc.is_package = FALSE)

                GROUP BY sq.part_code, pc.part_name

                ORDER BY pc.part_name

            ''', (str(session_id),))

            

            scanned_results = cursor.fetchall()

            

            # Beklenen paralar iin part_name'leri ek

            part_names = {}

            if expected_parts:

                part_codes_list = list(expected_parts.keys())

                if part_codes_list:

                    placeholders = ','.join(['%s' for _ in part_codes_list])

                    execute_query(cursor, f'SELECT part_code, part_name FROM part_codes WHERE part_code IN ({placeholders})', part_codes_list)

                    for row in cursor.fetchall():

                        part_names[row[0]] = row[1]

            

            # Tm paralar birletir (beklenen + taranan)

            all_parts = {}

            

            # Beklenen paralar ekle

            for part_code, expected_qty in expected_parts.items():

                all_parts[part_code] = {

                    'part_code': part_code,

                    'part_name': part_names.get(part_code, ''),

                    'expected': expected_qty,

                    'scanned': 0

                }

            

            # Taranan paralar ekle/gncelle

            for part_code, part_name, scanned_count in scanned_results:

                if part_code in all_parts:

                    all_parts[part_code]['part_name'] = part_name or ''

                    all_parts[part_code]['scanned'] = scanned_count

                else:

                    all_parts[part_code] = {

                        'part_code': part_code,

                        'part_name': part_name or '',

                        'expected': 0,

                        'scanned': scanned_count

                    }

            

            # DataFrame olutur

            rows = []

            for part in all_parts.values():

                difference = part['scanned'] - part['expected']

                

                if difference == 0:

                    status = ' Tam'

                elif difference > 0:

                    status = ' Fazla'

                else:

                    status = ' Eksik'

                

                rows.append({

                    'Para Kodu': part['part_code'],

                    'Para Ad': part['part_name'],

                    'Beklenen Adet': part['expected'],

                    'Saylan Adet': part['scanned'],

                    'Fark': difference,

                    'Durum': status

                })

            

            df = pd.DataFrame(rows)

            

            # Excel dosyasn kaydet - HER ZAMAN UNC PATH

            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")

            report_filename = f'sayim_raporu_{session_name}_{timestamp}.xlsx'

            os.makedirs(REPORTS_DIR, exist_ok=True)
            report_path = os.path.join(REPORTS_DIR, report_filename)

            

            with pd.ExcelWriter(report_path, engine='openpyxl') as writer:

                df.to_excel(writer, sheet_name='Saym Raporu', index=False)

                

                worksheet = writer.sheets['Saym Raporu']

                

                # Kolon geniliklerini ayarla

                column_widths = {

                    'A': 18,  # Para Kodu

                    'B': 30,  # Para Ad

                    'C': 15,  # Beklenen Adet

                    'D': 15,  # Saylan Adet

                    'E': 12,  # Fark

                    'F': 15   # Durum

                }

                

                for col, width in column_widths.items():

                    worksheet.column_dimensions[col].width = width

                

                # Balk satrn formatla

                from openpyxl.styles import Font, Alignment, PatternFill

                header_fill = PatternFill(start_color='DC2626', end_color='DC2626', fill_type='solid')

                header_font = Font(bold=True, color='FFFFFF', size=12)

                

                for cell in worksheet[1]:

                    cell.fill = header_fill

                    cell.font = header_font

                    cell.alignment = Alignment(horizontal='center', vertical='center')

                

                # Durum stununu renklendir

                for row_idx in range(2, len(rows) + 2):

                    cell = worksheet[f'F{row_idx}']

                    fark_cell = worksheet[f'E{row_idx}']

                    

                    fark_value = fark_cell.value

                    if fark_value == 0:

                        cell.fill = PatternFill(start_color='D4EDDA', end_color='D4EDDA', fill_type='solid')  # Yeil

                    elif fark_value < 0:

                        cell.fill = PatternFill(start_color='FFF3CD', end_color='FFF3CD', fill_type='solid')  # Sar

                    else:

                        cell.fill = PatternFill(start_color='D1ECF1', end_color='D1ECF1', fill_type='solid')  # Mavi

            

            print(f" Excel raporu kaydedildi: {report_path}")

            

        except Exception as e:

            print(f" Excel kayt hatas: {e}")

            import traceback

            traceback.print_exc()

        

        # Session' bitir ve rapor yolunu kaydet

        execute_query(cursor, '''

            UPDATE count_sessions 

            SET is_active = FALSE, ended_at = %s, total_scanned = %s, report_file_path = %s

            WHERE id = %s

        ''', (datetime.now(), total_scanned, report_filename, session_id))

        

        # 🧹 CACHE CLEANUP: Clear active session cache
        ACTIVE_SESSION_CACHE['session_id'] = None
        ACTIVE_SESSION_CACHE['expected_parts'] = []
        ACTIVE_SESSION_CACHE['cache_time'] = None
        print(f"🧹 Cache cleared for finished session {session_id}")

        

        conn.commit()

        close_db(conn)

        

        return jsonify({

            'success': True,

            'message': f' Saym tamamland! {total_scanned} adet tarama kaydedildi. Rapor oluturuldu.',

            'report_filename': report_filename

        })

        

    except Exception as e:

        print(f" Finish count error: {e}")

        import traceback

        traceback.print_exc()

        return jsonify({'success': False, 'error': str(e)}), 500





@app.route('/reports')

@login_required

def reports_page():

    """Tamamlanm saym raporlarn listele"""

    return render_template('reports.html')





@app.route('/api/get_saved_reports')

@login_required

def get_saved_reports():

    """Kaydedilmi raporlar JSON olarak dndr"""
    max_retries = 3
    
    for attempt in range(max_retries):
        conn = None
        try:
            conn = get_db()
            cursor = conn.cursor()
            
            # Tamamlanm saymlar al (is_active = FALSE ve report_file_path dolu olanlar)
            cursor.execute('''
                SELECT id, session_name, created_at, ended_at, 
                       total_expected, total_scanned, report_file_path
                FROM count_sessions 
                WHERE is_active = %s AND report_file_path IS NOT NULL
                ORDER BY ended_at DESC
            ''', (False,))
            
            sessions = cursor.fetchall()
            close_db(conn)
            
            # Tuple'lar dict'e evir
            sessions_list = []
            for s in sessions:
                sessions_list.append({
                    'id': s[0],
                    'session_name': s[1],
                    'created_at': s[2],
                    'ended_at': s[3],
                    'total_expected': s[4] or 0,
                    'total_scanned': s[5] or 0,
                    'report_file_path': s[6]
                })
            
            return jsonify(sessions_list)
            
        except Exception as e:
            if conn:
                try:
                    close_db(conn)
                except:
                    pass
            
            error_str = str(e).lower()
            if 'ssl' in error_str or 'connection' in error_str or 'closed' in error_str:
                if attempt < max_retries - 1:
                    import time
                    time.sleep(0.5)
                    continue
            
            app.logger.error(f"Get saved reports error: {e}")
            return jsonify({'success': False, 'error': str(e)}), 500
    
    return jsonify({'success': False, 'error': 'Bağlantı hatası, lütfen tekrar deneyin'}), 500





@app.route('/api/download_count_excel/<session_id>')

@login_required

def download_count_excel(session_id):

    """Saym sonularn detayl Excel olarak indir (beklenen vs taranan)"""

    try:

        conn = get_db()

        cursor = conn.cursor()

        

        # Saym bilgilerini al

        execute_query(cursor, '''

            SELECT session_name, created_at, started_at, ended_at, 

                   description, total_expected, total_scanned

            FROM count_sessions 

            WHERE id = %s

        ''', (str(session_id),))

        

        session_info = cursor.fetchone()

        

        if not session_info:

            close_db(conn)

            return jsonify({'success': False, 'error': 'Saym bulunamad'}), 404

        

        session_name, created_at, started_at, ended_at, description, total_expected, total_scanned = session_info

        

        # Beklenen paralar parse et

        expected_parts = {}

        if description:

            try:

                expected_list = json.loads(description)

                for item in expected_list:

                    # FIX: Correct field name is 'Parça Kodu' (with ç) not 'Para Kodu'

                    part_code = item.get('Parça Kodu') or item.get('Para Kodu') or item.get('part_code')

                    # Excel'den 'quantity', API'den 'expected_quantity', eski format 'Beklenen Adet'

                    expected_qty = item.get('Beklenen Adet') or item.get('expected_quantity') or item.get('quantity') or item.get('Adet') or 0

                    if part_code:

                        expected_parts[part_code] = int(expected_qty)

            except:

                pass

        

        # Taranan parçaları al (PAKETLERİ HARİÇ TUT - sadece gerçek parçalar)

        execute_query(cursor, '''

            SELECT sq.part_code, pc.part_name, COUNT(*) as scanned_count

            FROM scanned_qr sq

            LEFT JOIN part_codes pc ON sq.part_code = pc.part_code

            WHERE sq.session_id = %s 

              AND (pc.is_package IS NULL OR pc.is_package = FALSE)

            GROUP BY sq.part_code, pc.part_name

            ORDER BY pc.part_name

        ''', (str(session_id),))

        

        scanned_results = cursor.fetchall()

        

        # Beklenen paralar iin part_name'leri ek

        part_names = {}

        if expected_parts:

            part_codes_list = list(expected_parts.keys())

            if part_codes_list:

                placeholders = ','.join(['%s' for _ in part_codes_list])

                execute_query(cursor, f'SELECT part_code, part_name FROM part_codes WHERE part_code IN ({placeholders})', part_codes_list)

                for row in cursor.fetchall():

                    part_names[row[0]] = row[1]

        

        close_db(conn)

        

        # Tm paralar birletir (beklenen + taranan)

        all_parts = {}

        

        # Beklenen paralar ekle

        for part_code, expected_qty in expected_parts.items():

            all_parts[part_code] = {

                'part_code': part_code,

                'part_name': part_names.get(part_code, ''),  # DB'den ekilen part_name

                'expected': expected_qty,

                'scanned': 0

            }

        

        # Taranan paralar ekle/gncelle

        for part_code, part_name, scanned_count in scanned_results:

            if part_code in all_parts:

                all_parts[part_code]['part_name'] = part_name or ''

                all_parts[part_code]['scanned'] = scanned_count

            else:

                all_parts[part_code] = {

                    'part_code': part_code,

                    'part_name': part_name or '',

                    'expected': 0,

                    'scanned': scanned_count

                }

        

        # DataFrame olutur - Sadece para detaylar

        rows = []

        for part in all_parts.values():

            difference = part['scanned'] - part['expected']

            

            # Durum emoji

            if difference == 0:

                status = ' Tam'

            elif difference > 0:

                status = ' Fazla'

            else:

                status = ' Eksik'

            

            rows.append({

                'Para Kodu': part['part_code'],

                'Para Ad': part['part_name'],

                'Beklenen Adet': part['expected'],

                'Saylan Adet': part['scanned'],

                'Fark': difference,

                'Durum': status

            })

        

        df = pd.DataFrame(rows)

        

        # Excel buffer

        output = BytesIO()

        with pd.ExcelWriter(output, engine='openpyxl') as writer:

            # Tek sayfa - Detayl Rapor

            df.to_excel(writer, sheet_name='Saym Raporu', index=False)

            

            # Formatla

            worksheet = writer.sheets['Saym Raporu']

            

            # Kolon geniliklerini ayarla

            column_widths = {

                'A': 18,  # Para Kodu

                'B': 30,  # Para Ad

                'C': 15,  # Beklenen Adet

                'D': 15,  # Saylan Adet

                'E': 12,  # Fark

                'F': 15   # Durum

            }

            

            for col, width in column_widths.items():

                worksheet.column_dimensions[col].width = width

            

            # Balk satrn kalnlatr

            from openpyxl.styles import Font, Alignment, PatternFill

            header_fill = PatternFill(start_color='DC2626', end_color='DC2626', fill_type='solid')

            header_font = Font(bold=True, color='FFFFFF', size=12)

            

            for cell in worksheet[1]:

                cell.fill = header_fill

                cell.font = header_font

                cell.alignment = Alignment(horizontal='center', vertical='center')

            

            # Durum stununu renklendir

            for row_idx in range(2, len(rows) + 2):

                cell = worksheet[f'F{row_idx}']

                fark_cell = worksheet[f'E{row_idx}']

                

                fark_value = fark_cell.value

                if fark_value == 0:

                    cell.fill = PatternFill(start_color='D4EDDA', end_color='D4EDDA', fill_type='solid')  # Yeil

                elif fark_value < 0:

                    cell.fill = PatternFill(start_color='FFF3CD', end_color='FFF3CD', fill_type='solid')  # Sar

                else:

                    cell.fill = PatternFill(start_color='D1ECF1', end_color='D1ECF1', fill_type='solid')  # Mavi

        

        output.seek(0)

        

        return send_file(

            output,

            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',

            as_attachment=True,

            download_name=f'sayim_raporu_{session_name}_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'

        )

        

    except Exception as e:

        print(f" Download Excel error: {e}")

        import traceback

        traceback.print_exc()

        return jsonify({'success': False, 'error': str(e)}), 500





@app.route('/login', methods=['GET', 'POST'])
@rate_limit_login
def login():
    """Login sayfasını göster veya giriş işlemini yap"""
    if request.method == 'GET':
        if 'user_id' in session:
            return redirect('/')
        return render_template('login.html')
    
    # POST - Giriş işlemini yap
    from werkzeug.security import check_password_hash

    data = request.get_json()

    username = data.get('username')

    password = data.get('password')



    if not username or not password:

        return jsonify({'error': 'Kullanc ad ve ifre gerekli'}), 400



    conn = None
    try:
        # Doğrudan PostgreSQL sorgusu
        conn = get_db()
        cursor = conn.cursor()
        cursor.execute('''
            SELECT id, username, password_hash, full_name, role 
            FROM envanter_users 
            WHERE username = %s
        ''', (username,))
        user = cursor.fetchone()
        close_db(conn)
        
        # DEBUG LOG
        logging.info(f"[LOGIN DEBUG] Username: {username}")
        logging.info(f"[LOGIN DEBUG] User found: {user is not None}")
        if user:
            logging.info(f"[LOGIN DEBUG] User ID: {user[0]}, Role: {user[4]}")
            pw_check = check_password_hash(user[2], password)
            logging.info(f"[LOGIN DEBUG] Password check result: {pw_check}")
        
        if user and check_password_hash(user[2], password):
            session.permanent = True  # Session kalıcı yap - kapatana kadar
            session['user_id'] = user[0]
            session['username'] = user[1]
            session['full_name'] = user[3]
            session['role'] = user[4]
            if user[4] == 'admin':
                session['admin_authenticated'] = True
            return jsonify({'success': True, 'role': user[4]})
        else:
            logging.warning(f"[LOGIN DEBUG] Login failed for: {username}")
            return jsonify({'error': 'Kullanc ad veya ifre hatal'}), 401

    except Exception as e:
        if conn:
            try:
                close_db(conn)
            except:
                pass
        logging.exception(f"Login error: {e}")
        return jsonify({'error': 'Veritaban hatas'}), 500



@app.route('/logout', methods=['GET', 'POST'])

def logout():

    session.clear()

    return redirect('/login')



@app.route('/profile')
@login_required
def profile_page():
    """Kullanıcı profil sayfası"""
    return render_template('profile.html')



@app.route('/api/user/profile', methods=['GET', 'POST'])
@login_required
def user_profile():
    """Profil verilerini getir veya güncelle"""
    try:
        user_id = session.get('user_id')
        if not user_id:
            return jsonify({'success': False, 'error': 'Oturum sona erdi'}), 401
        
        user = User.query.get(user_id)
        
        if not user:
            return jsonify({'success': False, 'error': 'Kullanıcı bulunamadı'}), 404
        
        if request.method == 'GET':
            # Profil verilerini döndür
            return jsonify({
                'success': True,
                'user': {
                    'id': user.id,
                    'username': user.username,
                    'full_name': user.full_name,
                    'title': user.title or '',
                    'email': user.email or '',
                    'role': user.role,
                    'profile_image_path': user.profile_image_path or '',
                    'created_at': user.created_at.isoformat() if user.created_at else ''
                }
            })
        
        # POST - Profili güncelle
        data = request.get_json()
        
        # Şifre kontrolü varsa
        if 'current_password' in data and data.get('current_password'):
            from werkzeug.security import check_password_hash
            if not check_password_hash(user.password_hash, data['current_password']):
                return jsonify({'success': False, 'error': 'Mevcut şifre hatalı'}), 401
            
            # Yeni şifre ayarla
            if data.get('new_password'):
                from werkzeug.security import generate_password_hash
                user.password_hash = generate_password_hash(data['new_password'])
                user.password = data['new_password']
        
        # Diğer alanları güncelle
        if 'full_name' in data:
            user.full_name = data['full_name']
        if 'title' in data:
            user.title = data['title']
        if 'email' in data:
            user.email = data['email']
        
        user.updated_at = datetime.utcnow()
        db.session.commit()
        
        # Session'ı güncelle
        session['full_name'] = user.full_name
        
        return jsonify({'success': True, 'message': 'Profil başarıyla güncellendi'})
    
    except Exception as e:
        db.session.rollback()
        logging.error(f'Profile update error: {e}')
        return jsonify({'success': False, 'error': f'Profil güncellenirken hata oluştu: {str(e)}'}), 500



@app.route('/api/user/profile/photo', methods=['POST'])
@login_required
def upload_profile_photo():
    """Profil fotoğrafı yükle"""
    try:
        user_id = session.get('user_id')
        if not user_id:
            return jsonify({'success': False, 'error': 'Oturum sona erdi'}), 401
        
        user = User.query.get(user_id)
        
        if not user:
            return jsonify({'success': False, 'error': 'Kullanıcı bulunamadı'}), 404
        
        if 'photo' not in request.files:
            return jsonify({'success': False, 'error': 'Dosya seçilmedi'}), 400
        
        file = request.files['photo']
        
        if file.filename == '':
            return jsonify({'success': False, 'error': 'Dosya seçilmedi'}), 400
        
        # Dosya türü kontrolü
        allowed_extensions = {'png', 'jpg', 'jpeg', 'gif', 'webp'}
        if not ('.' in file.filename and file.filename.rsplit('.', 1)[1].lower() in allowed_extensions):
            return jsonify({'success': False, 'error': 'Sadece resim dosyaları yüklenebilir'}), 400
        
        # Dosya boyutu kontrolü (5MB)
        if len(file.read()) > 5 * 1024 * 1024:
            return jsonify({'success': False, 'error': 'Dosya boyutu 5MB\'dan küçük olmalıdır'}), 400
        
        file.seek(0)
        
        # Dosya adını oluştur
        filename = secure_filename(f'profile_{user.id}_{int(time.time())}.{file.filename.rsplit(".", 1)[1].lower()}')
        filepath = os.path.join(PROFILE_PHOTOS_DIR, filename)
        
        # Eski fotoğrafı sil
        if user.profile_image_path:
            old_filepath = os.path.join(PROFILE_PHOTOS_DIR, os.path.basename(user.profile_image_path))
            if os.path.exists(old_filepath):
                try:
                    os.remove(old_filepath)
                except:
                    pass
        
        # Dosyayı kaydet
        os.makedirs(PROFILE_PHOTOS_DIR, exist_ok=True)
        file.save(filepath)
        
        # Veritabanını güncelle
        user.profile_image_path = f'/uploads/profiles/{filename}'
        user.updated_at = datetime.utcnow()
        db.session.commit()
        
        return jsonify({
            'success': True,
            'message': 'Fotoğraf başarıyla yüklendi',
            'photo_path': user.profile_image_path
        })
    
    except Exception as e:
        db.session.rollback()
        logging.error(f'Photo upload error: {e}')
        return jsonify({'success': False, 'error': f'Fotoğraf yüklenirken hata oluştu: {str(e)}'}), 500



@app.route('/check_auth')

def check_auth():

    if 'user_id' in session:

        return jsonify({

            'authenticated': True,

            'username': session.get('username', None),

            'full_name': session.get('full_name', None),

            'role': session.get('role', None)

        })

    return jsonify({'authenticated': False})



# ============================================================================

#  HEALTH CHECK & DEBUG ENDPOINTS

# ============================================================================



@app.route('/health', methods=['GET'])
@app.route('/api/health', methods=['GET'])
def health_check():

    """Sistem salk kontrol - monitoring iin"""

    try:

        # DB balants test et

        db.session.execute(db.text('SELECT 1'))

        db_status = 'healthy'

        db_error = None

    except Exception as e:

        db_status = 'unhealthy'

        db_error = str(e)

    

    # Session says

    try:

        active_sessions = CountSession.query.filter_by(is_active=True).count()

    except:

        active_sessions = 0

    

    # Disk kullanm

    try:

        import shutil

        disk_usage = shutil.disk_usage('.')

        disk_free_gb = round(disk_usage.free / (1024**3), 2)

        disk_total_gb = round(disk_usage.total / (1024**3), 2)

        disk_used_percent = round((disk_usage.used / disk_usage.total) * 100, 1)

    except:

        disk_free_gb = 0

        disk_total_gb = 0

        disk_used_percent = 0

    

    # DB dosya boyutu

    try:

        db_size_mb = round(os.path.getsize('instance/envanter_local.db') / (1024**2), 2)

    except:

        db_size_mb = 0

    

    # Uptime

    uptime_seconds = int(time.time() - app.config.get('START_TIME', time.time()))

    uptime_hours = round(uptime_seconds / 3600, 1)

    

    health_data = {

        'status': 'ok' if db_status == 'healthy' else 'degraded',

        'timestamp': datetime.utcnow().isoformat(),

        'database': {

            'status': db_status,

            'error': db_error,

            'size_mb': db_size_mb,

            'type': 'SQLite'

        },

        'environment': {

            'mode': 'LOCAL' if IS_LOCAL else 'PRODUCTION',

            'storage': 'Local Files',

            'python_version': f"{os.sys.version_info.major}.{os.sys.version_info.minor}.{os.sys.version_info.micro}"

        },

        'sessions': {

            'active_count': active_sessions

        },

        'disk': {

            'free_gb': disk_free_gb,

            'total_gb': disk_total_gb,

            'used_percent': disk_used_percent

        },

        'uptime': {

            'seconds': uptime_seconds,

            'hours': uptime_hours

        }

    }

    

    return jsonify(health_data)


@app.route('/api/video/portal', methods=['GET'])
def serve_portal_video():
    """Serve portal video for splash screen"""
    import os
    from flask import send_file

    # Önce yerel geliştirme klasöründe ara
    local_path = os.path.join(os.path.dirname(__file__), 'static', 'portal_video.mp4')
    if os.path.exists(local_path):
        return send_file(local_path, mimetype='video/mp4')

    # Sonra shared/static klasöründe ara
    shared_path = os.path.join(STATIC_DIR, 'portal_video.mp4')
    if os.path.exists(shared_path):
        return send_file(shared_path, mimetype='video/mp4')

    # Bulunamazsa hata döndür
    return {'error': 'Portal video bulunamadı'}, 404


@app.route('/api/sync/status', methods=['GET'])
def api_sync_status():
    """Electron için sync durumu endpoint'i - paylaşımlı klasör sync durumunu döndürür"""
    try:
        from shared_folder_sync import sync_progress
        
        state = sync_progress.get_state()
        
        # Sync tamamlandı mı kontrol et
        is_ready = state['status'] in ['completed', 'idle', 'failed']
        
        return jsonify({
            'status': 'ok',
            'sync_enabled': True,
            'sync_status': state['status'],  # idle, syncing, completed, failed
            'is_ready': is_ready,  # True olunca login açılabilir
            'progress_percent': state.get('progress_percent', 0),
            'current_category': state.get('current_category'),
            'message': state.get('message', ''),
            'downloaded': state.get('downloaded', 0),
            'total_files': state.get('total_files', 0),
            'processed_files': state.get('processed_files', 0),
            'errors': state.get('errors', 0),
            'timestamp': datetime.utcnow().isoformat()
        })
    except Exception as e:
        # Sync modülü yüklenemezse varsayılan olarak hazır döndür
        return jsonify({
            'status': 'ok',
            'sync_enabled': False,
            'is_ready': True,
            'message': f'Sync modülü yüklenemedi: {str(e)}',
            'timestamp': datetime.utcnow().isoformat()
        })


@app.route('/log_frontend_error', methods=['POST'])

def log_frontend_error():

    """Frontend hatalarn logla"""

    try:

        data = request.json

        

        # Log dosyasna yaz

        app.logger.error(f"FRONTEND ERROR: {json.dumps(data, ensure_ascii=False)}")

        

        # Ayr dosyaya da yaz

        # os.makedirs('logs', exist_ok=True) # COMMENTED - handled by startup_orchestrator

        with open('logs/frontend_errors.log', 'a', encoding='utf-8') as f:

            error_line = {

                'timestamp': datetime.utcnow().isoformat(),

                'error': data.get('error', {}),

                'context': data.get('context', {}),

                'user': {

                    'username': session.get('username'),

                    'user_id': session.get('user_id')

                }

            }

            f.write(json.dumps(error_line, ensure_ascii=False) + '\n')

        

        return jsonify({'success': True})

    except Exception as e:

        app.logger.error(f'Frontend error logging hatas: {e}')

        return jsonify({'success': False, 'error': str(e)}), 500



@app.route('/admin/frontend_errors_log', methods=['GET'])

@login_required

def get_frontend_errors_log():

    """Frontend error log dosyasn oku - sadece admin"""

    # Admin kontrol

    if session.get('role') != 'admin':

        return jsonify({'error': 'Admin eriimi gerekli'}), 403

    

    try:

        errors = []

        log_path = 'logs/frontend_errors.log'

        

        if os.path.exists(log_path):

            with open(log_path, 'r', encoding='utf-8') as f:

                lines = f.readlines()

                # Son 50 satr

                for line in lines[-50:]:

                    try:

                        errors.append(json.loads(line.strip()))

                    except:

                        pass

        

        return jsonify({

            'success': True,

            'errors': errors,

            'count': len(errors)

        })

    except Exception as e:

        app.logger.error(f'Frontend errors log okuma hatas: {e}')

        return jsonify({'success': False, 'error': str(e)}), 500



        return jsonify({'success': True, 'message': 'Hata kaydedildi'})

    except Exception as e:

        app.logger.error(f'Frontend error logging failed: {e}')

        return jsonify({'success': False, 'error': str(e)}), 500



@app.route('/upload_parts', methods=['POST'])

@login_required

def upload_parts():

    """

    Yeni Para Ykleme Sistemi:

    - Excel'den sadece para bilgileri (part_code, part_name) yklenir

    - QR kod retimi yaplmaz (manuel olarak para detay sayfasndan retilir)

    - Mevcut paralar gncellenir, yeni paralar eklenir

    """

    if 'file' not in request.files:

        return jsonify({'error': 'Dosya bulunamad'}), 400



    file = request.files['file']

    if file.filename == '':

        return jsonify({'error': 'Dosya seilmedi'}), 400



    if not (file.filename and file.filename.endswith(('.xlsx', '.xls'))):

        return jsonify({'error': 'Sadece Excel dosyalar yklenebilir'}), 400



    try:

        df = pd.read_excel(file)



        # Sadece part_code ve part_name gerekli

        required_columns = ['part_code', 'part_name']

        if not all(col in df.columns for col in required_columns):

            return jsonify({'error': 'Excel dosyas "part_code" ve "part_name" stunlarn iermelidir'}), 400



        conn = get_db()

        cursor = conn.cursor()

        # Mevcut paralar al

        execute_query(cursor, 'SELECT id, part_code, part_name FROM part_codes')

        existing_parts = {}  # {part_code: (id, part_name)}

        for row in cursor.fetchall():

            existing_parts[row[1]] = (row[0], row[2])



        new_parts = []

        updated_parts = []

        processing_summary = {

            'new_parts': 0,

            'updated_parts': 0

        }



        print(f"\n PARA YKLEME SSTEM")

        print(f" Excel'den gelen para says: {len(df)}")

        print("="*50)



        for _, row in df.iterrows():

            part_code = str(row['part_code']).strip()

            part_name = str(row['part_name']).strip()



            if not part_code or not part_name:

                continue  # Bo satrlar atla



            if part_code in existing_parts:

                # MEVCUT PARA - Sadece ismi gncelle

                part_code_id, old_part_name = existing_parts[part_code]

                if old_part_name != part_name:

                    execute_query(cursor, f'UPDATE part_codes SET part_name = %s WHERE part_code = %s',

                                 (part_name, part_code))

                    updated_parts.append(part_code)

                    processing_summary['updated_parts'] += 1

                    print(f" {part_code}: '{old_part_name}'  '{part_name}'")

                else:

                    print(f" {part_code}: Zaten gncel")

            else:

                # YEN PARA - Sadece part_codes'a ekle (QR kod retilmez)

                execute_query(cursor, f'INSERT INTO part_codes (part_code, part_number, part_name, description, created_at) VALUES (%s, %s, %s, %s, %s)',

                             (part_code, part_code, part_name, '', datetime.now()))

                new_parts.append(part_code)

                processing_summary['new_parts'] += 1

                print(f" {part_code}: Yeni para eklendi - '{part_name}'")



        conn.commit()

        close_db(conn)



        print(f"\n LEM TAMAMLANDI")

        print(f" Yeni paralar: {processing_summary['new_parts']}")

        print(f" Gncellenen paralar: {processing_summary['updated_parts']}")

        print(" QR kodlar para detay sayfasndan retilebilir")

        print("="*50)



        return jsonify({

            'success': True,

            'message': f'lem tamamland! {processing_summary["new_parts"]} yeni para eklendi, {processing_summary["updated_parts"]} para gncellendi.',

            'summary': {

                'new_parts': processing_summary['new_parts'],

                'updated_parts': processing_summary['updated_parts']

            }

        })



    except Exception as e:

        logging.exception(f"Error in smart QR upload system: {e}")

        try:

            close_db(conn)

        except Exception:

            pass

        return jsonify({'error': f'lem srasnda hata olutu: {str(e)}'}), 500



@app.route('/get_qr_codes')

@login_required

def get_qr_codes():

    search = request.args.get('search', '').strip()

    page = int(request.args.get('page', 1))

    limit = int(request.args.get('limit', 100))

    offset = (page - 1) * limit



    conn = get_db()

    cursor = conn.cursor()



    if search:

        # nce tam eleme ara (JOIN ile part_code ve part_name al)

        execute_query(cursor, """

            SELECT qc.id, pc.part_code, pc.part_name, qc.is_used, 

                   CASE WHEN qc.blob_url IS NOT NULL THEN 1 ELSE 0 END as is_downloaded

            FROM qr_codes qc

            JOIN part_codes pc ON qc.part_code_id = pc.id

            WHERE pc.part_code = %s OR pc.part_name = %s

            ORDER BY pc.part_code, qc.id

            LIMIT %s OFFSET %s

        """, (search, search, limit, offset))

        exact_matches = cursor.fetchall()



        if exact_matches:

            qr_codes = []

            for row in exact_matches:

                qr_codes.append({

                    'qr_id': row[0],

                    'part_code': row[1],

                    'part_name': row[2],

                    'is_used': row[3],

                    'is_downloaded': row[4]

                })

        else:

            # Tam eleme bulunamazsa ksmi eleme ara

            execute_query(cursor, """

                SELECT qc.id, pc.part_code, pc.part_name, qc.is_used,

                       CASE WHEN qc.blob_url IS NOT NULL THEN 1 ELSE 0 END as is_downloaded

                FROM qr_codes qc

                JOIN part_codes pc ON qc.part_code_id = pc.id

                WHERE pc.part_code LIKE  OR pc.part_name LIKE 

                ORDER BY pc.part_code, qc.id

                LIMIT  OFFSET 

            """, (f'%{search}%', f'%{search}%', limit, offset))

            rows = cursor.fetchall()

            qr_codes = []

            for row in rows:

                qr_codes.append({

                    'qr_id': row[0],

                    'part_code': row[1],

                    'part_name': row[2],

                    'is_used': row[3],

                    'is_downloaded': row[4]

                })

    else:

        # Arama terimi yoksa tm QR kodlar getir (sayfalama ile)

        execute_query(cursor, """

            SELECT qc.id, pc.part_code, pc.part_name, qc.is_used,

                   CASE WHEN qc.blob_url IS NOT NULL THEN 1 ELSE 0 END as is_downloaded

            FROM qr_codes qc

            JOIN part_codes pc ON qc.part_code_id = pc.id

            ORDER BY pc.part_code, qc.id

            LIMIT  OFFSET 

        """, (limit, offset))

        rows = cursor.fetchall()

        qr_codes = []

        for row in rows:

            qr_codes.append({

                'qr_id': row[0],

                'part_code': row[1],

                'part_name': row[2],

                'is_used': row[3],

                'is_downloaded': row[4]

            })



    # Toplam sayy al (JOIN ile)

    if search:

        execute_query(cursor, """

            SELECT COUNT(*) 

            FROM qr_codes qc

            JOIN part_codes pc ON qc.part_code_id = pc.id

            WHERE pc.part_code LIKE  OR pc.part_name LIKE 

        """, (f'%{search}%', f'%{search}%'))

    else:

        execute_query(cursor, "SELECT COUNT(*) FROM qr_codes")



    total_count = cursor.fetchone()[0]



    close_db(conn)



    return jsonify({

        'qr_codes': qr_codes,

        'total_count': total_count,

        'current_page': page,

        'total_pages': (total_count + limit - 1) // limit,

        'has_more': (page * limit) < total_count

    })



@app.route('/clear_all_qrs', methods=['POST'])

@login_required

def clear_all_qrs():

    """Tm QR kodlarn temizle (aktif saym oturumu yoksa)"""

    try:

        conn = get_db()

        cursor = conn.cursor()



        # Aktif saym oturumu kontrol

        execute_query(cursor, 'SELECT COUNT(*) FROM count_sessions WHERE is_active = %s', (True,))

        active_session = cursor.fetchone()[0]



        if active_session > 0:

            close_db(conn)

            return jsonify({'error': 'Aktif bir saym oturumu var. QR kodlar silinemez.'}), 400



        #  SHARED FOLDER: QR klasörünü temizle
        if os.path.exists(QR_CODES_DIR):
            for part_folder in os.listdir(QR_CODES_DIR):
                part_folder_path = os.path.join(QR_CODES_DIR, part_folder)
                if os.path.isdir(part_folder_path):
                    for file in os.listdir(part_folder_path):
                        try:
                            os.remove(os.path.join(part_folder_path, file))
                        except Exception as e:
                            logging.error(f"Shared folder dosya {file} silinirken hata: {e}")



        # QR kodlarn ve paralar sil

        execute_query(cursor, 'DELETE FROM qr_codes')

        execute_query(cursor, 'DELETE FROM part_codes')



        conn.commit()

        close_db(conn)



        # Cache'i temizle

        cache_clear()



        logging.info("Tm QR kodlar temizlendi")

        return jsonify({

            'success': True,

            'message': 'Tm QR kodlar baaryla silindi'

        })

    except Exception as e:

        try:

            close_db(conn)

        except:

            pass

        logging.error(f"QR kodlar silinirken hata: {e}", exc_info=True)

        return jsonify({'error': f'QR kodlar silinirken hata: {str(e)}'}), 500



@app.route('/generate_qr_image/<qr_id>')

@login_required

def generate_qr_image(qr_id):

    """QR kod oluturma - QR altna kod ve para ad ekler"""

    try:

        # Cache'den kontrol et

        cache_key = f'qr_image_{qr_id}'

        cached_image = cache_get(cache_key)



        if cached_image:

            buf = BytesIO(cached_image)

            return send_file(buf, mimetype='image/png')



        # QR ID'den parça kodunu ve numarayı çıkar (Y129150-49811_1)
        parts = qr_id.rsplit('_', 1)
        part_code = parts[0] if len(parts) > 0 else qr_id
        qr_number = parts[1] if len(parts) > 1 else "1"

        #  SHARED FOLDER: Static klasörden kontrol et
        static_path = os.path.join(QR_CODES_DIR, part_code, f'{qr_id}.png')
        if os.path.exists(static_path):
            with open(static_path, 'rb') as f:
                file_content = f.read()
            cache_set(cache_key, file_content)
            buf = BytesIO(file_content)
            return send_file(buf, mimetype='image/png')

        

        # Para adn database'den al

        conn = get_db()

        cursor = conn.cursor()

        execute_query(cursor, 'SELECT part_name FROM part_codes WHERE part_code = %s', (part_code,))

        result = cursor.fetchone()

        part_name = result[0] if result else ""

        close_db(conn)



        # QR kod olutur - Barkod makinesi iin optimize

        qr = qrcode.QRCode(

            version=1, 

            box_size=8,  # 8px - barkod makinesi iin ideal

            border=2,    # 2px quiet zone

            error_correction=qrcode.constants.ERROR_CORRECT_M

        )

        qr.add_data(qr_id)

        qr.make(fit=True)

        qr_img = qr.make_image(fill_color="black", back_color="white")

        qr_img = qr_img.convert('RGB')  # PIL Image'a dntr



        # QR kod boyutlarn al

        qr_width, qr_height = qr_img.size

        

        # Alanlar hesapla

        logo_height = 40  # Logo iin st alan

        text_height = 35  # Alt yaz (para numaras) iin alan

        

        # Krmz ereve iin padding

        border_width = 3  # 3px krmz ereve

        

        # Yeni grsel olutur (logo + QR + text alan + ereve)

        final_width = qr_width + (border_width * 2)

        final_height = logo_height + qr_height + text_height + (border_width * 2)

        final_img = Image.new('RGB', (final_width, final_height), '#dc2626')  # Krmz arka plan (ereve)

        

        # Beyaz i alan olutur (logo + QR + text)

        white_bg = Image.new('RGB', (qr_width, logo_height + qr_height + text_height), 'white')

        

        # Logo ekle (varsa) - st ortasna

        try:

            logo_path = os.path.join(os.path.dirname(__file__), 'cermak-logo.png')

            if os.path.exists(logo_path):

                logo_img = Image.open(logo_path).convert('RGBA')

                # Logo boyutunu alan yksekliine gre ayarla

                logo_width = 150

                logo_height_logo = 40

                try:

                    logo_img = logo_img.resize((logo_width, logo_height_logo), Image.Resampling.LANCZOS)

                except AttributeError:

                    logo_img = logo_img.resize((logo_width, logo_height_logo), Image.LANCZOS)

                

                # Logo'yu ortala

                logo_x = (qr_width - logo_width) // 2

                logo_y = 5  # stten 5px boluk

                

                # RGBA logo'yu blend et

                if logo_img.mode == 'RGBA':

                    alpha = logo_img.split()[3]

                    logo_img = logo_img.convert('RGB')

                    white_bg.paste(logo_img, (logo_x, logo_y), alpha)

                else:

                    white_bg.paste(logo_img, (logo_x, logo_y))

        except Exception as e:

            print(f"Logo ekleme hatas: {e}")

        

        # QR kodu beyaz alana yaptr (logo'nun altna)

        white_bg.paste(qr_img, (0, logo_height))

        

        # Beyaz alan krmz erevenin iine yaptr

        final_img.paste(white_bg, (border_width, border_width))

        

        # Text ekleme iin draw nesnesi

        draw = ImageDraw.Draw(final_img)

        

        # Font (kaln ve byk)

        try:

            font = ImageFont.truetype("arialblk.ttf", 24)

        except:

            try:

                font = ImageFont.truetype("arialbd.ttf", 24)

            except:

                try:

                    font = ImageFont.truetype("arial.ttf", 24)

                except:

                    font = ImageFont.load_default()

        

        # QR ID yazs - Sadece bu

        qr_text = qr_id

        

        # Text geniliini hesapla (24pt font iin)

        text_width = len(qr_text) * 14

        

        # QR ID'yi ortala ve yaz (border_width offset ekle)

        x_position = max(border_width, (final_width - text_width) // 2)

        draw.text((x_position, logo_height + qr_height + border_width + 5), qr_text, fill='black', font=font)



        buf = BytesIO()

        final_img.save(buf, format='PNG', optimize=True)

        buf.seek(0)



        # Cache'e kaydet

        img_data = buf.getvalue()

        cache_set(cache_key, img_data)



        #  SHARED FOLDER: static klasörüne kaydet (part_code subfolder ile)
        qr_dir = os.path.join(QR_CODES_DIR, part_code)
        os.makedirs(qr_dir, exist_ok=True)
        with open(static_path, 'wb') as f:
            f.write(img_data)
        
        # Dosyayı read-only yap (shared folder'da hata olabilir)
        try:
            os.chmod(static_path, 0o444)
        except Exception as chmod_e:
            logging.warning(f"chmod hatası (shared folder): {chmod_e}")

        # Dosyayı döndür
        buf.seek(0)
        return send_file(buf, mimetype='image/png')

    except Exception as e:
        logging.error(f"Error generating QR image for {qr_id}: {e}")

        # Hata durumunda basit QR olutur

        qr = qrcode.QRCode(version=1, box_size=4, border=1)

        qr.add_data(qr_id)

        qr.make(fit=True)

        img = qr.make_image(fill_color="black", back_color="white")



        buf = BytesIO()

        img.save(buf, format='PNG')

        buf.seek(0)



        return send_file(buf, mimetype='image/png')



@app.route('/download_single_qr/<qr_id>')

@login_required

def download_single_qr(qr_id):

    conn = get_db()

    cursor = conn.cursor()

    execute_query(cursor, '''

        SELECT pc.part_code 

        FROM qr_codes qc

        JOIN part_codes pc ON qc.part_code_id = pc.id

        WHERE qc.id = %s

    ''', (qr_id,))

    result = cursor.fetchone()



    if not result:

        close_db(conn)

        return jsonify({'error': 'QR kod bulunamad'}), 404



    # is_downloaded kolonu yok, sadece blob_url kontrol yeterli

    close_db(conn)

    
    # QR ID'den parça kodunu çıkar
    parts = qr_id.rsplit('_', 1)
    part_code_from_id = parts[0] if len(parts) > 0 else qr_id

    try:

        # SHARED FOLDER: Static dosyadan kontrol et
        static_path = os.path.join(QR_CODES_DIR, part_code_from_id, f'{qr_id}.png')

        if os.path.exists(static_path):

            return send_file(static_path, mimetype='image/png', as_attachment=True, download_name=f'{qr_id}.png')



        # QR kod yoksa oluştur

        qr = qrcode.QRCode(version=1, box_size=10, border=4)

        qr.add_data(qr_id)

        qr.make(fit=True)

        img = qr.make_image(fill_color="black", back_color="white")



        buf = BytesIO()

        img.save(buf, format='PNG')

        buf.seek(0)



        img_data = buf.getvalue()



        # SHARED FOLDER: Static klasörüne kaydet (part_code subfolder ile)

        qr_dir = os.path.join(QR_CODES_DIR, part_code_from_id)

        os.makedirs(qr_dir, exist_ok=True)

        local_path = os.path.join(qr_dir, f'{qr_id}.png')

        with open(local_path, 'wb') as f:

            f.write(img_data)



        # Dosyay dndr

        buf.seek(0)

        return send_file(buf, mimetype='image/png', as_attachment=True, download_name=f'{qr_id}.png')



    except Exception as e:

        logging.error(f"Error downloading QR image for {qr_id}: {e}")

        # Hata durumunda geleneksel yntemle olutur

        qr = qrcode.QRCode(version=1, box_size=10, border=4)

        qr.add_data(qr_id)

        qr.make(fit=True)

        img = qr.make_image(fill_color="black", back_color="white")



        buf = BytesIO()

        img.save(buf, format='PNG')

        buf.seek(0)



        return send_file(buf, mimetype='image/png', as_attachment=True, download_name=f'{qr_id}.png')



# ========================================================================
# SOCKET.IO HANDLER DEVRE DIŞI - Tek PC + Barkod makinesi kullanıldığı için
# Tarama işlemleri /api/scan_qr HTTP endpoint üzerinden yapılıyor
# ========================================================================
# @socketio.on('scan_qr_radical')  # DEVRE DIŞI
def handle_scan_radical_DISABLED(data):

    """ Ultra reliable QR scanning with enhanced features - SOCKET.IO DEVRE DIŞI"""

    print("\n" + ""*50)

    print("ULTRA QR SCAN RECEIVED")

    print(""*50)



    try:

        qr_id = data.get('qr_id', '').strip()

        session_id = int(data.get('session_id', 1))  # INT olarak al

        user_id = session.get('user_id', 1)



        logging.debug(f" QR ID: {qr_id}, Session: {session_id}, User: {user_id}")



        if not qr_id:

            logging.warning("QR ID missing in scan request")

            # emit('scan_result', {'success': False, 'message': ' QR ID eksik!'})  # DEVRE DIŞI

            return



        with db_connection() as conn:

            cursor = conn.cursor()



            # Check QR exists (JOIN ile part bilgisi al)

            execute_query(cursor, '''

                SELECT pc.part_code, pc.part_name 

                FROM qr_codes qc

                JOIN part_codes pc ON qc.part_code_id = pc.id

                WHERE qc.id = %s

            ''', (qr_id,))

            qr_data = cursor.fetchone()



            if not qr_data:

                logging.warning(f"QR not found: {qr_id}")

                emit('scan_result', {'success': False, 'message': f' QR kod bulunamad: {qr_id}'})

                return



            part_code, part_name = qr_data

            logging.info(f"QR found: {part_code} - {part_name}")



            # Ensure session exists

            execute_query(cursor, 'SELECT COUNT(*) FROM count_sessions WHERE id = ', (session_id,))

            if cursor.fetchone()[0] == 0:

                logging.info(f"Creating new session {session_id}")

                execute_query(cursor, 

                    'INSERT INTO count_sessions (id, session_name, is_active, created_at) VALUES (?, ?, ?, ?)',

                    (session_id, f'Session_{session_id}', 1, datetime.now()))

                conn.commit()



            #  COMPOSITE INDEX kullanr - ok hzl + AKILLI DUPLICATE

            execute_query(cursor, '''

                SELECT scanned_by, scanned_at 

                FROM scanned_qr 

                WHERE session_id = ? AND qr_id = ?

            ''', (session_id, qr_id))

            duplicate_record = cursor.fetchone()

            

            if duplicate_record:

                # Duplicate bulundu - Akll mesaj olutur

                prev_user_id, prev_scanned_at = duplicate_record

                

                # nceki kullanc bilgisi

                execute_query(cursor, 'SELECT full_name FROM envanter_users WHERE id = %s', (prev_user_id,))

                prev_user_result = cursor.fetchone()

                prev_user_name = prev_user_result[0] if prev_user_result else 'Bilinmeyen'

                

                # Zaman fark hesapla

                prev_time = datetime.fromisoformat(prev_scanned_at) if isinstance(prev_scanned_at, str) else prev_scanned_at

                time_diff = datetime.now() - prev_time

                

                # Zaman format

                if time_diff.total_seconds() < 60:

                    time_str = f"{int(time_diff.total_seconds())} saniye nce"

                    is_suspicious = time_diff.total_seconds() < 30  # 30 saniyeden ksa ise pheli

                elif time_diff.total_seconds() < 3600:

                    time_str = f"{int(time_diff.total_seconds() / 60)} dakika nce"

                    is_suspicious = False

                else:

                    time_str = f"{int(time_diff.total_seconds() / 3600)} saat nce"

                    is_suspicious = False

                

                # Ayn kullanc m kontrol et

                is_same_user = (prev_user_id == user_id)

                

                # Akll mesaj olutur

                if is_suspicious:

                    duplicate_msg = f" PHEL! {part_name} {time_str} tarand ({prev_user_name})"

                elif is_same_user:

                    duplicate_msg = f" {part_name} zaten taradnz! ({time_str})"

                else:

                    duplicate_msg = f" {part_name} zaten sayld! {prev_user_name} tarafndan {time_str}"

                

                logging.warning(f"Duplicate scan: {qr_id} - {duplicate_msg}")

                emit('scan_result', {

                    'success': False, 

                    'message': duplicate_msg,

                    'duplicate': True,

                    'previous_user': prev_user_name,

                    'time_ago': time_str,

                    'is_suspicious': is_suspicious,

                    'same_user': is_same_user

                }, broadcast=True)

                return



            #  TRANSACTION - Atomik ilem

            cursor.execute('BEGIN TRANSACTION')

            try:

                # Insert scan record

                execute_query(cursor, 

                    'INSERT INTO scanned_qr (session_id, qr_id, qr_code, part_code, scanned_by, scanned_at) VALUES (%s, %s, %s, %s, %s, %s)',

                    (session_id, qr_id, qr_id, part_code, user_id, datetime.now()))



                # Mark QR as used

                execute_query(cursor, 'UPDATE qr_codes SET is_used = %s, used_at = %s WHERE qr_id = %s',

                             (True, datetime.now(), qr_id))



                # Update session stats

                execute_query(cursor, '''

                    UPDATE count_sessions 

                    SET total_scanned = (SELECT COUNT(*) FROM scanned_qr WHERE session_id = %s)

                    WHERE id = %s

                ''', (session_id, session_id))



                conn.commit()

                logging.info(f"SUCCESS: {part_name} scanned")

            except Exception as e:

                conn.rollback()

                logging.error(f"Transaction failed: {e}")

                raise



        # Get user name (ayr connection)

        with db_connection() as conn2:

            cursor2 = conn2.cursor()

            execute_query(cursor2, 'SELECT full_name FROM envanter_users WHERE id = %s', (user_id,))

            user_result = cursor2.fetchone()

            user_name = user_result[0] if user_result else 'Kullanc'



        #  TRIPLE BROADCAST

        success_data = {

            'success': True,

            'message': f' {part_name} sayld!',

            'qr_code': qr_id,

            'part_code': part_code,

            'part_name': part_name,

            'session_id': session_id,

            'scanned_by': user_name,

            'scanned_at': datetime.now().strftime('%H:%M:%S')

        }



        # Total scan count (session stats'dan al)

        with db_connection() as conn3:

            cursor3 = conn3.cursor()

            execute_query(cursor3, 'SELECT total_scanned FROM count_sessions WHERE id = ', (session_id,))

            result = cursor3.fetchone()

            success_data['total_scans'] = result[0] if result else 0



        # Canlı takip için broadcast
        socketio.emit('scan_result', success_data, broadcast=True)
        socketio.emit('qr_scanned', success_data, broadcast=True)
        socketio.emit('activity_update', success_data, broadcast=True)

        logging.info(f"ULTRA SUCCESS - {part_name} scanned by {user_name}")



    except sqlite3.Error as e:

        logging.error(f"Database error in scan: {e}", exc_info=True)

        # emit('scan_result', {'success': False, 'message': f' Veritaban hatas: {e}'})  # DEVRE DIŞI

    except ValueError as e:

        logging.error(f"Value error in scan: {e}", exc_info=True)

        # emit('scan_result', {'success': False, 'message': f' Geersiz veri: {e}'})  # DEVRE DIŞI

    except Exception as e:

        logging.error(f"Unexpected error in scan: {e}", exc_info=True)

        # emit('scan_result', {'success': False, 'message': f' Sistem hatas: {e}'})  # DEVRE DIŞI



#  ULTRA MODERN API ENDPOINTS

@app.route('/api/scan_qr', methods=['POST'])

def api_scan_qr_ultra():
    """? ULTRA FAST QR scanning - Optimized for speed"""
    try:
        data = request.get_json()
        if not data:
            return jsonify({'success': False, 'message': 'No data provided'}), 400

        qr_id = data.get('qr_id', '').strip()
        session_id = data.get('session_id', '1')

        if not qr_id:
            return jsonify({'success': False, 'message': 'QR ID required'}), 400

        # ? QR TRANSFORMATION - Scanner ? ve * okur, biz düzeltiriz
        # Scanner okur: ANTF03?6  -> Gerçek: ANTF03_6
        # Scanner okur: Y129648*01780 -> Gerçek: Y129648-01780
        qr_id_original = qr_id
        qr_id = qr_id.replace('?', '_').replace('*', '-')
        
        if qr_id != qr_id_original:
            logging.info(f"QR FIX: '{qr_id_original}' -> '{qr_id}'")

        # ? CACHE'DEN KONTROL - 50ms -> 0.1ms
        start_time = time.time()
        qr_info = get_qr_from_cache(qr_id)
        
        if not qr_info:
            # Cache miss
            conn = get_db()
            cursor = conn.cursor()
            
            cursor.execute('''
                SELECT pc.part_code, pc.part_name, pc.id
                FROM qr_codes qc
                INNER JOIN part_codes pc ON qc.part_code_id = pc.id
                WHERE qc.qr_id = %s
            ''', (qr_id,))
            
            row = cursor.fetchone()
            close_db(conn)
            
            if not row:
                return jsonify({'success': False, 'message': f'QR bulunamadı'}), 404
            
            qr_info = {'part_code': row[0], 'part_name': row[1], 'part_code_id': row[2]}
            QR_LOOKUP_CACHE[qr_id] = qr_info
        
        # Process scan
        result = process_qr_scan_ultra_fast(qr_id, qr_info, session_id)
        
        # Performance monitoring
        elapsed = (time.time() - start_time) * 1000
        if elapsed > 50:
            logging.warning(f"[SLOW] Slow scan: {elapsed:.0f}ms")

        socketio.emit('scan_result', result)
        return jsonify(result)

    except Exception as e:
        logging.error(f"Scan error: {e}")
        return jsonify({'success': False, 'message': str(e)}), 500


def process_qr_scan_ultra_fast(qr_id, qr_info, session_id):
    """? Ultra fast processing - <30ms target"""
    conn = None
    try:
        conn = get_db()
        cursor = conn.cursor()
        
        part_code = qr_info['part_code']
        part_name = qr_info['part_name']
        
        # ========================================
        # PAKET KONTROL - Önce paketin olup olmadığını kontrol et
        # ========================================
        cursor.execute('''
            SELECT is_package, package_items, part_name 
            FROM part_codes 
            WHERE part_code = %s
        ''', (qr_id,))
        
        package_check = cursor.fetchone()
        
        app.logger.info(f'[PAKET CHECK] QR: {qr_id}, package_check: {package_check}')
        
        # HOTFIX: JCB/ATAŞMAN paketleri is_package=FALSE olarak kaydedilmiş ama package_items var
        # Bunları düzeltmek için is_package değerini package_items var mı diye kontrol et
        is_package = package_check[0] if package_check else False
        if package_check and not is_package and package_check[1]:
            # Eğer is_package FALSE ama package_items varsa, bu paket olmalı!
            is_package = True
            app.logger.warning(f'[HOTFIX] {qr_id} is_package was FALSE but has items - forcing TRUE')
        
        # Eğer bu bir paketse, içindeki tüm parçaları tek tek tara
        if package_check and is_package:  # is_package = True
            app.logger.info(f'[PAKET DETECTED] {qr_id} is a package!')
            try:
                package_items = json.loads(package_check[1]) if package_check[1] else []
                package_name = package_check[2]
                
                if not package_items:
                    close_db(conn)
                    return {
                        'success': False,
                        'message': f'%s {package_name} paketi boş!',
                        'item_name': package_name
                    }
                
                # 🔒 Paket taraması duplicate kontrol (TRANSACTION LOCK)
                cursor.execute('START TRANSACTION')
                
                cursor.execute('''
                    SELECT COUNT(*) FROM scanned_qr 
                    WHERE qr_id = %s AND session_id = %s
                    FOR UPDATE
                ''', (qr_id, str(session_id)))
                
                if cursor.fetchone()[0] > 0:
                    conn.rollback()
                    close_db(conn)
                    return {
                        'success': False,
                        'message': f'❌ {package_name} paketi zaten tarandı!',
                        'item_name': package_name,
                        'duplicate': True
                    }
                
                # Paket içindeki her parçayı otomatik tara
                total_items = 0
                app.logger.info(f'[PAKET TARAMA] {package_name} - package_items: {package_items}')
                
                for idx, item in enumerate(package_items):
                    item_part_code = item.get('part_code')
                    quantity = item.get('quantity', 1)
                    
                    app.logger.info(f'[PAKET TARAMA] Item {idx}: part_code={item_part_code}, quantity={quantity}')
                    
                    if not item_part_code:
                        app.logger.warning(f'[PAKET TARAMA] Item {idx} has no part_code, skipping')
                        continue
                    
                    # Her bir parça için quantity kadar tarama kaydı oluştur
                    for qty_idx in range(quantity):
                        virtual_qr = f"{qr_id}_PKG_{item_part_code}_{idx}_{qty_idx}"
                        cursor.execute('''
                            INSERT INTO scanned_qr (qr_id, qr_code, session_id, part_code, scanned_by, scanned_at)
                            VALUES (%s, %s, %s, %s, %s, %s)
                        ''', (virtual_qr, virtual_qr, str(session_id), item_part_code, session.get('user_id', 1), datetime.now()))
                        total_items += 1
                
                conn.commit()
                
                # count_sessions tablosunu güncelle (İLERLEME ÇUBUĞU İÇİN ÖNEMLİ!)
                cursor.execute('''
                    UPDATE count_sessions 
                    SET total_scanned = total_scanned + %s
                    WHERE id = %s
                ''', (total_items, session_id))
                conn.commit()
                
                # Cache güncelle
                if session_id in SESSION_STATS_CACHE:
                    SESSION_STATS_CACHE[session_id]['total'] += total_items
                    new_total = SESSION_STATS_CACHE[session_id]['total']
                else:
                    cursor.execute('SELECT COUNT(*) FROM scanned_qr WHERE session_id = %s', (str(session_id),))
                    new_total = cursor.fetchone()[0]
                    SESSION_STATS_CACHE[session_id] = {'total': new_total}
                
                close_db(conn)
                
                app.logger.info(f'[PAKET COMPLETE] {package_name}: {total_items} items scanned, new total: {new_total}')
                
                return {
                    'success': True,
                    'message': f'%s {package_name} paketi tarandı! {total_items} parça eklendi. Toplam: {new_total}',
                    'item_name': package_name,
                    'part_code': qr_id,
                    'total_scans': new_total,
                    'qr_id': qr_id,
                    'is_package': True,
                    'package_items_count': total_items
                }
                
            except Exception as pkg_error:
                conn.rollback()
                close_db(conn)
                app.logger.error(f'[PAKET ERROR] {qr_id}: {pkg_error}')
                return {
                    'success': False,
                    'message': f'Paket tarama hatası: {str(pkg_error)}'
                }
        
        # Normal parça taraması (paket değilse)
        # 🔒 MULTI-SCANNER SAFE: Transaction lock ile duplicate check
        try:
            # START TRANSACTION ile atomik işlem
            cursor.execute('START TRANSACTION')
            
            # 🔒 ROW LOCK (FOR UPDATE) - Aynı QR'ı aynı anda okuyan diğer scanner'ları bekletir
            cursor.execute('''
                SELECT 1 FROM scanned_qr 
                WHERE qr_id = %s AND session_id = %s 
                FOR UPDATE
            ''', (qr_id, str(session_id)))
            
            if cursor.fetchone():
                conn.rollback()
                close_db(conn)
                return {'success': False, 'message': f'❌ {part_name} zaten tarandı!', 'duplicate': True}

            # INSERT - Artık %100 güvenli
            now = datetime.now()
            cursor.execute('''
                INSERT INTO scanned_qr (qr_id, qr_code, session_id, part_code, scanned_by, scanned_at)
                VALUES (%s, %s, %s, %s, %s, %s)
            ''', (qr_id, qr_id, str(session_id), part_code, session.get('user_id', 1), now))
            
            # ✅ COMMIT - Atomik olarak kaydet
            conn.commit()
            
        except Exception as lock_error:
            conn.rollback()
            close_db(conn)
            logging.error(f"Lock error: {lock_error}")
            return {'success': False, 'message': f'Tarama hatası: {str(lock_error)}'}
        
        # count_sessions tablosunu güncelle (İLERLEME ÇUBUĞU İÇİN ÖNEMLİ!)
        cursor.execute('''
            UPDATE count_sessions 
            SET total_scanned = total_scanned + 1
            WHERE id = %s
        ''', (session_id,))
        conn.commit()

        # ASYNC QR update (non-blocking)
        def update_qr_async():
            try:
                c = get_db()
                cur = c.cursor()
                cur.execute('UPDATE qr_codes SET is_used=TRUE, used_at=%s, used_count=COALESCE(used_count,0)+1 WHERE qr_id=%s', 
                           (now, qr_id))
                c.commit()
                close_db(c)
            except:
                pass
        threading.Thread(target=update_qr_async, daemon=True).start()

        # Cached stats
        if session_id in SESSION_STATS_CACHE:
            SESSION_STATS_CACHE[session_id]['total'] += 1
            total = SESSION_STATS_CACHE[session_id]['total']
        else:
            cursor.execute('SELECT COUNT(*) FROM scanned_qr WHERE session_id = %s', (str(session_id),))
            total = cursor.fetchone()[0]
            SESSION_STATS_CACHE[session_id] = {'total': total}

        close_db(conn)

        return {
            'success': True,
            'message': f'{part_name} başarıyla tarandı! (#{total})',
            'item_name': part_name,
            'part_code': part_code,
            'total_scans': total,
            'qr_id': qr_id
        }

    except Exception as e:
        if conn:
            conn.rollback()
            close_db(conn)
        logging.error(f"Process error: {e}")
        return {'success': False, 'message': f'Hata: {str(e)}'}


def process_qr_scan_ultra(qr_id, session_id):
    """Ultra reliable QR processing - ESKI FONKSİYON (yedek)"""
    # Bu fonksiyon artık kullanılmıyor, process_qr_scan_ultra_fast kullanılıyor
    return process_qr_scan_ultra_fast(qr_id, {'part_code': qr_id, 'part_name': 'Unknown'}, session_id)


def process_qr_scan_ultra_legacy(qr_id, session_id):

    """ Ultra reliable QR processing with enhanced features"""

    conn = None

    try:

        conn = get_db()

        cursor = conn.cursor()

        

        # Ultra session management - ensure session exists with better naming

        session_name = f"Tarama Seans {session_id}"



        # Defensive: check if session row exists, then insert using DB-agnostic columns

        # count_sessions kontrol artk gereksiz - session ID yoksa oluturma

        # Session ID artk auto-increment, manuel oluturmaya gerek yok

        

        # ========================================

        #  PAKET KONTROL - nce paketin olup olmadn kontrol et

        # ========================================

        cursor.execute('''

            SELECT is_package, package_items, part_name 

            FROM part_codes 
            WHERE part_code = %s
        ''', (qr_id,))

        

        package_check = cursor.fetchone()

        

        app.logger.info(f'[PAKET CHECK] QR: {qr_id}, package_check: {package_check}')

        

        # Eer bu bir paketse, iindeki tm paralar tek tek tara

        if package_check and package_check[0]:  # is_package = True

            app.logger.info(f'[PAKET DETECTED] {qr_id} is a package!')

            try:

                package_items = json.loads(package_check[1]) if package_check[1] else []

                package_name = package_check[2]

                

                if not package_items:

                    return {

                        'success': False,

                        'message': f' {package_name} paketi bo!',

                        'item_name': package_name

                    }

                

                # Paket taramas duplicate kontrol

                cursor.execute('''

                    SELECT COUNT(*) FROM scanned_qr 

                    WHERE qr_id = %s AND session_id = %s
                ''', (qr_id, str(session_id)))

                

                if cursor.fetchone()[0] > 0:

                    return {

                        'success': False,

                        'message': f' {package_name} paketi zaten tarand!',

                        'item_name': package_name,

                        'duplicate': True

                    }

                

                #  PAKET KENDISINI TARA DEL - SADECE ERDEK PARALARI TARA

                # Paket, sadece tracker olarak kaydediliyor, toplam sayya EKLENMYOR

                

                # Paket içindeki her parçayı otomatik tara

                total_items = 0

                app.logger.info(f'[PAKET TARAMA] {package_name} - package_items: {package_items}')

                

                for idx, item in enumerate(package_items):

                    part_code = item.get('part_code')

                    quantity = item.get('quantity', 1)

                    

                    app.logger.info(f'[PAKET TARAMA] Item {idx}: part_code={part_code}, quantity={quantity}')

                    

                    if not part_code:

                        app.logger.warning(f'[PAKET TARAMA] Item {idx} has no part_code, skipping')

                        continue

                    

                    # Her bir parça için quantity kadar tarama kaydı oluştur

                    for qty_idx in range(quantity):

                        virtual_qr = f"{qr_id}_PKG_{part_code}_{idx}_{qty_idx}"

                        cursor.execute('''

                            INSERT INTO scanned_qr (qr_id, qr_code, session_id, part_code, scanned_by, scanned_at)

                            VALUES (%s, %s, %s, %s, %s, %s)

                        ''', (virtual_qr, virtual_qr, str(session_id), part_code, session.get('user_id', 1), datetime.now()))

                        total_items += 1

                

                # statistikleri gncelle (SADECE ERDEK PARALARI SAY, PAKET SAYMA!)

                cursor.execute('''

                    SELECT COUNT(*) as total_scans

                    FROM scanned_qr 

                    WHERE session_id = %s
                ''', (str(session_id),))

                

                total_scans = cursor.fetchone()[0]

                

                try:
                    cursor.execute('UPDATE count_sessions SET total_scanned = %s WHERE id = %s', 
                                 (total_scans, str(session_id)))

                except Exception as e:

                    print(f" Package total_scanned gncelleme hatas: {e}")

                    pass

                

                conn.commit()

                

                app.logger.info(f'[PAKET TARAMA] {package_name} - {total_items} para otomatik tarand')

                

                return {

                    'success': True,

                    'message': f'  {package_name}\n{total_items} para otomatik sayld!',

                    'item_name': package_name,

                    'total_scans': total_scans,

                    'is_package': True,

                    'package_items': total_items

                }

                

            except Exception as pkg_err:

                app.logger.error(f'[PAKET HATASI] {qr_id}: {pkg_err}')

                # Paket hatas olursa normal QR gibi ile

                pass

        

        # ========================================

        #  NORMAL QR LEME - Standart para tarama

        # ========================================

        

        # Check if QR exists with enhanced data retrieval

        qr_data = None

        try:

            # nce tam eleme dene - INNER JOIN daha hızlı

            execute_query(cursor, """

                SELECT qc.id, pc.part_code, pc.part_name, qc.is_used

                FROM qr_codes qc

                INNER JOIN part_codes pc ON qc.part_code_id = pc.id

                WHERE qc.id = %s

            """, (qr_id,))

            qr_data = cursor.fetchone()

            

            # Hala bulunamadysa, part_code olarak ara (QR = part_code durumu)

            if not qr_data:

                execute_query(cursor, """

                    SELECT part_code, part_code, part_name, FALSE

                    FROM part_codes

                    WHERE part_code = %s

                """, (qr_id,))

                qr_data = cursor.fetchone()

                

        except Exception as e:

            # Schema mismatch or missing column in older DBs - fall back to defaults

            print(f" QR lookup failed (schema mismatch): {e}")

            qr_data = None



        if not qr_data:

            # Do NOT attempt to modify the schema here. Use a safe fallback QR record in-memory.

            unknown_name = f"Bilinmeyen rn ({qr_id[:15]})"

            print(f" QR BULUNAMADI: {qr_id}")  # Debug log

            qr_data = (qr_id, qr_id[:10], unknown_name, False)



        qr_id_db, part_code, part_name, is_used = qr_data



        #  KALICI DUPLICATE KONTROL - Bir session'da bir QR sadece 1 kez okunabilir

        execute_query(cursor, """

            SELECT 1 FROM scanned_qr 

            WHERE qr_id = %s AND session_id = %s

            LIMIT 1

        """, (qr_id, str(session_id)))



        if cursor.fetchone():

            # Bu QR zaten bu session'da taranm - asla tekrar taranmamal

            return {

                'success': False,

                'message': f' {part_name} zaten tarand!',

                'item_name': part_name,

                'duplicate': True

            }



        # Insert ultra scan record with enhanced data

        try:

            execute_query(cursor, """

                INSERT INTO scanned_qr (qr_id, qr_code, session_id, part_code, scanned_by, scanned_at)

                VALUES (%s, %s, %s, %s, %s, %s)

            """, (qr_id, qr_id, str(session_id), part_code, session.get('user_id', 1), datetime.now()))

        except Exception as insert_err:

            # Hata olursa transaction rollback yap

            conn.rollback()

            logging.error(f"scanned_qr INSERT error: {insert_err}")

            

            # Sequence hatası ise düzelt ve tekrar dene

            if "duplicate key" in str(insert_err).lower():

                logging.warning("Duplicate key error - fixing sequence...")

                fix_sequence('scanned_qr', 'id')

                

                # Yeni connection ile tekrar dene

                try:

                    execute_query(cursor, """

                        INSERT INTO scanned_qr (qr_id, qr_code, session_id, part_code, scanned_by, scanned_at)

                        VALUES (%s, %s, %s, %s, %s, %s)

                    """, (qr_id, qr_id, str(session_id), part_code, session.get('user_id', 1), datetime.now()))

                except Exception as retry_err:

                    logging.error(f"Retry INSERT failed: {retry_err}")

                    conn.rollback()

                    raise

            else:

                # Fallback for older schemas without part_code column

                try:

                    execute_query(cursor, """

                        INSERT INTO scanned_qr (qr_id, qr_code, session_id, scanned_by, scanned_at)

                        VALUES (%s, %s, %s, %s, %s)

                    """, (qr_id, qr_id, str(session_id), session.get('user_id', 1), datetime.now()))

                except Exception as ie:

                    conn.rollback()

                    print(f" Failed to insert scanned_qr record: {ie}")

                    raise



        # CRITICAL: scanned_qr INSERT başarılı - hemen commit et

        conn.commit()



        # Mark QR as used - AYRI bir işlem olarak (başarısız olsa bile scanned_qr kaydı korunur)

        try:

            now = datetime.now()

            cursor.execute("""

                UPDATE qr_codes 

                SET is_used = %s,

                    used_at = %s,

                    used_count = COALESCE(used_count, 0) + 1,

                    last_used_at = %s,

                    first_used_at = COALESCE(first_used_at, %s)

                WHERE qr_id = %s

            """, (True, now, now, now, qr_id))

            conn.commit()

        except Exception as qr_update_err:

            # QR codes update hatası - ignore et (scanned_qr zaten kaydedildi)

            logging.warning(f"QR update failed (ignored): {qr_update_err}")

            conn.rollback()  # Sadece bu UPDATE'i geri al

            pass



        # Get enhanced session statistics

        try:

            execute_query(cursor, """

                SELECT 

                    COUNT(*) as total_scans,

                    COUNT(DISTINCT qr_id) as unique_items

                FROM scanned_qr 

                WHERE session_id = %s

            """, (str(session_id),))



            stats = cursor.fetchone()

            total_scans = stats[0] if stats else 0

            unique_items = stats[1] if stats else 0

        except Exception as stats_err:

            logging.error(f"Stats query failed: {stats_err}")

            total_scans = 0

            unique_items = 0



        # Update count_sessions.total_scanned for dashboard

        try:

            execute_query(cursor, 'UPDATE count_sessions SET total_scanned = %s WHERE id = %s', (total_scans, str(session_id)))

            conn.commit()

        except Exception as e:

            logging.error(f"total_scanned update error: {e}")

            conn.rollback()

            pass



        # Get user info for enhanced feedback

        user_id = session.get('user_id', 1)

        execute_query(cursor, 'SELECT full_name FROM envanter_users WHERE id = %s', (user_id,))

        user_result = cursor.fetchone()

        user_name = user_result[0] if user_result else 'Kullanc'



        # Ultra success response

        return {

            'success': True,

            'message': f'{part_name} başarıyla tarandı! (#{total_scans})',

            'item_name': part_name,

            'part_code': part_code,

            'total_scans': total_scans,

            'unique_items': unique_items,

            'qr_id': qr_id,

            'session_id': session_id,

            'scanned_by': user_name,

            'scan_time': datetime.now().isoformat(),

            'was_used_before': is_used

        }



    except Exception as e:

        if conn:

            conn.rollback()

        print(f" ULTRA Process QR error: {e}")

        return {

            'success': False,

            'message': f' Sistem hatas: {str(e)}',

            'error_detail': str(e),

            'qr_id': qr_id

        }

    finally:

        if conn:

            close_db(conn)



@app.route('/api/session/<session_id>/stats', methods=['GET'])

def get_ultra_session_stats(session_id):

    """ Get ultra detailed session statistics"""

    try:

        conn = get_db()

        cursor = conn.cursor()



        # Get session info

        execute_query(cursor, """

            SELECT 

                CASE 

                    WHEN is_active::integer = 1 THEN 'active'

                    WHEN is_active::integer = 0 THEN 'completed'

                    ELSE 'unknown'

                END as status, 

                started_at, 

                ended_at 

            FROM count_sessions 

            WHERE session_id = %s

        """, (str(session_id),))



        session_info = cursor.fetchone()



        if not session_info:

            return jsonify({'error': 'Session not found'}), 404



        # Get ultra scan statistics

        execute_query(cursor, """

            SELECT 

                COUNT(*) as total_scans,

                COUNT(DISTINCT qr_id) as unique_items,

                MIN(scanned_at) as first_scan,

                MAX(scanned_at) as last_scan

            FROM scanned_qr 

            WHERE session_id = %s

        """, (str(session_id),))



        stats = cursor.fetchone()



        # Get recent scans with enhanced data

        execute_query(cursor, """

            SELECT sq.qr_id, qc.part_name, sq.scanned_at, u.full_name

            FROM scanned_qr sq

            LEFT JOIN qr_codes qc ON sq.qr_id = qc.id

            LEFT JOIN envanter_users u ON sq.scanned_by = u.id

            WHERE sq.session_id = %s

            ORDER BY sq.scanned_at DESC

            LIMIT 20

        """, (str(session_id),))



        recent_scans = cursor.fetchall()



        # Calculate session duration

        start_time = session_info[1]

        end_time = session_info[2] or (stats[3] if stats[3] else datetime.now())

        duration_seconds = (end_time - start_time).total_seconds() if start_time else 0



        return jsonify({

            'session_id': session_id,

            'status': session_info[0],

            'started_at': start_time.isoformat() if start_time else None,

            'ended_at': session_info[2].isoformat() if session_info[2] else None,

            'duration_seconds': duration_seconds,

            'total_scans': stats[0] or 0,

            'unique_items': stats[1] or 0,

            'first_scan': stats[2].isoformat() if stats[2] else None,

            'last_scan': stats[3].isoformat() if stats[3] else None,

            'scans_per_minute': round((stats[0] or 0) / max(duration_seconds / 60, 1), 2),

            'recent_scans': [

                {

                    'qr_id': scan[0],

                    'item_name': scan[1] or f'Unknown ({scan[0]})',

                    'scan_time': scan[2].isoformat() if scan[2] else None,

                    'scanned_by': scan[3] or 'Bilinmeyen'

                }

                for scan in recent_scans

            ]

        })



    except Exception as e:

        print(f" ULTRA Stats error: {e}")

        return jsonify({'error': str(e)}), 500

    finally:

        if conn:

            close_db(conn)



# ========================================

#  TELEFON QR & PARA BLGLER SSTEM KALDIRILDI

# Artk telefon saym YOK - sadece desktop QR scanner kullanlyor

# Kaldrlan zellikler:

# - /qr-info (mobil QR bilgi sayfas)

# - /api/qr_info/<qr_id> (QR bilgi API)

# - /api/part_details/<qr_id> (Para detaylar)

# - /api/update_part_details (Para gncelleme)

# - /api/upload_part_photo (Fotoraf ykleme)

# - /api/qr/<qr_id>/info (Detayl QR bilgisi)

# ========================================



#  RAPOR YNETM (Admin Count - ifre sistemi kaldrld)

    try:

        #  ULTRA SECURITY: Sadece admin kullancs saym bitirebilir

        current_user_id = session.get('user_id')



        if not current_user_id:

            return jsonify({'success': False, 'error': 'Oturum bulunamad - Ltfen tekrar giri yapn'}), 401



        # Admin kontrol - kullanc bilgilerini al

        conn = get_db()

        cursor = conn.cursor()



        execute_query(cursor, "SELECT username, role FROM envanter_users WHERE id = %s", (current_user_id,))

        user_result = cursor.fetchone()



        if not user_result:

            close_db(conn)

            return jsonify({'success': False, 'error': 'Kullanc bulunamad'}), 404



        username, user_role = user_result



        # DEBUG LOG

        print(f"[FINISH_COUNT DEBUG] User: {username}, Role: {user_role}, ID: {current_user_id}")

        print(f"[FINISH_COUNT DEBUG] Role type: {type(user_role)}, Value: [{user_role}]")

        print(f"[FINISH_COUNT DEBUG] Role == 'admin': {user_role == 'admin'}")

        print(f"[FINISH_COUNT DEBUG] Role.lower() == 'admin': {user_role.lower() == 'admin' if user_role else 'NULL'}")



        #  SADECE ADMIN YETKS - Role tabanl kontrol (daha gvenli)

        if not user_role or user_role.lower() != 'admin':

            close_db(conn)

            security_logger.warning(f'USER {username} (Role: {user_role}, ID: {current_user_id}) tried to finish count session - PERMISSION DENIED')

            return jsonify({

                'success': False, 

                'error': f'YETKISIZ ERM: Sadece admin yetkisine sahip kullanclar saym bitirebilir (Your role: {user_role})',

                'permission_required': 'admin',

                'current_role': user_role

            }), 403



        # Saym eriim kontrol (secondary check) - ARTIK GEREKL DEL

        # if not session.get('count_access'):

        #     close_db(conn)

        #     return jsonify({'success': False, 'error': 'Saym eriimi iin ifre gerekli'}), 403



        # Aktif saym oturumunu kontrol et

        execute_query(cursor, "SELECT id, is_active, created_by FROM count_sessions WHERE is_active = %s LIMIT 1", (True,))

        session_result = cursor.fetchone()



        if not session_result:

            close_db(conn)

            return jsonify({'success': False, 'error': 'Aktif saym oturumu bulunamad'}), 400



        session_id, is_active_status, created_by = session_result



        # Log admin action

        security_logger.info(f'ADMIN {username} finishing count session {session_id}')



        # ift ilem kontrol - eer bu oturum zaten tamamlandysa

        if not is_active_status:

            close_db(conn)

            return jsonify({'success': False, 'error': 'Bu saym oturumu zaten tamamlanm'}), 400



        # Saym oturumunu sonlandir (admin yetkisiyle)

        execute_query(cursor, "UPDATE count_sessions SET is_active = %s, ended_at = %s WHERE id = %s",

                     (False, datetime.now(), session_id))



        # Rapor verilerini topla - BEKLENEN ADETLERLE KARILATIR

        # Her QR tek bir part_code'a ait, o yzden JOIN ile part_code ekiyoruz

        execute_query(cursor, '''

            SELECT 

                COALESCE(pc.part_code, sq.part_code) as part_code,

                COALESCE(pc.part_name, 'Bilinmeyen Para') as part_name,

                COUNT(*) as sayilan_adet

            FROM scanned_qr sq

            LEFT JOIN qr_codes qc ON sq.qr_id = qc.id

            LEFT JOIN part_codes pc ON qc.part_code_id = pc.id

            WHERE sq.session_id = %s

            GROUP BY COALESCE(pc.part_code, sq.part_code), COALESCE(pc.part_name, 'Bilinmeyen Para')

            ORDER BY part_code

        ''', (session_id,))



        scanned_parts = {}  # {part_code: (part_name, sayilan_adet)}

        for row in cursor.fetchall():

            part_code = row[0]

            part_name = row[1]

            sayilan_adet = row[2]

            scanned_parts[part_code] = (part_name, sayilan_adet)

            print(f"[RAPOR DEBUG] {part_code}: {part_name} - {sayilan_adet} adet okundu")



        # Tm part_codes'dan beklenen adetleri al (yklenmi Excel'den)

        execute_query(cursor, '''

            SELECT 

                pc.part_code,

                pc.part_name,

                COUNT(qc.id) as beklenen_adet

            FROM part_codes pc

            LEFT JOIN qr_codes qc ON qc.part_code_id = pc.id

            GROUP BY pc.part_code, pc.part_name

        ''')



        # Rapor verilerini hazrla - TM PARALAR (okutulan + okutulmayan)

        report_data = []

        total_scanned = 0

        total_expected = 0



        for row in cursor.fetchall():

            part_code = row[0]

            part_name = row[1]

            beklenen_adet = row[2] or 0

            

            # Bu para okutulan paralar arasnda var m

            if part_code in scanned_parts:

                sayilan_adet = scanned_parts[part_code][1]

            else:

                sayilan_adet = 0  # Hi okutulmam

            

            fark = sayilan_adet - beklenen_adet

            

            part_data = {

                'Para Kodu': part_code or 'Bilinmiyor',

                'Para Ad': part_name,

                'Beklenen Adet': beklenen_adet,

                'Saylan Adet': sayilan_adet,

                'Fark': fark,

                'Durum': ' Tamam' if fark == 0 else (' Eksik' if fark < 0 else '+ Fazla')

            }

            report_data.append(part_data)

            total_scanned += sayilan_adet

            total_expected += beklenen_adet



        # Excel raporu olutur

        df = pd.DataFrame(report_data)



        output = BytesIO()

        with pd.ExcelWriter(output, engine='openpyxl') as writer:

            df.to_excel(writer, index=False, sheet_name='Saym Raporu')

            

            # Worksheet'i al ve formatla

            workbook = writer.book

            worksheet = writer.sheets['Saym Raporu']

            

            # Stun geniliklerini ayarla

            worksheet.column_dimensions['A'].width = 15  # Para Kodu

            worksheet.column_dimensions['B'].width = 30  # Para Ad

            worksheet.column_dimensions['C'].width = 15  # Beklenen Adet

            worksheet.column_dimensions['D'].width = 15  # Saylan Adet

            worksheet.column_dimensions['E'].width = 10  # Fark

            worksheet.column_dimensions['F'].width = 12  # Durum

            

            # Header stilini ayarla

            from openpyxl.styles import Font, PatternFill, Alignment

            header_fill = PatternFill(start_color='DC3545', end_color='DC3545', fill_type='solid')

            header_font = Font(bold=True, color='FFFFFF')

            

            for cell in worksheet[1]:

                cell.fill = header_fill

                cell.font = header_font

                cell.alignment = Alignment(horizontal='center', vertical='center')

        

        output.seek(0)



        # Rapor dosyasn kaydet

        report_filename = f'sayim_raporu_{session_id}_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'

        report_path = os.path.join(REPORTS_DIR, report_filename)



        with open(report_path, 'wb') as f:

            f.write(output.getvalue())



        # Doruluk orann hesapla

        accuracy_rate = (total_scanned / total_expected * 100) if total_expected > 0 else 0.0



        # Raporu count_reports table'na kaydet (varsa)

        report_title = f"Saym Raporu - {datetime.now().strftime('%d.%m.%Y %H:%M')}"

        try:

            execute_query(cursor, '''

                INSERT INTO count_reports (session_id, report_filename, report_title, 

                                         total_expected, total_scanned, accuracy_rate, created_at)

                VALUES (%s, %s, %s, %s, %s, %s, %s)

            ''', (session_id, report_filename, report_title, total_expected, total_scanned, accuracy_rate, datetime.now()))

        except Exception as e:

            # count_reports tablosu yoksa sadece dosya kaydet

            logging.warning(f"count_reports tablosu yok, sadece dosya kaydedildi: {e}")



        # Database ilemini commit et

        conn.commit()

        close_db(conn)



        # WebSocket ile saym bittii bilgisini gnder
        try:
            socketio.emit('count_finished', {'session_id': session_id})
        except Exception as ws_error:
            logging.warning(f"WebSocket notification failed: {str(ws_error)}")

        # Session'dan saym bilgilerini temizle

        session.pop('count_access', None)

        session.pop('current_session', None)



        return jsonify({

            'success': True,

            'message': 'Saym baaryla tamamland',

            'report_file': report_filename,

            'session_id': session_id,

            'total_expected': total_expected,

            'total_scanned': total_scanned,

            'accuracy_rate': round(accuracy_rate, 2)

        })



    except Exception as e:

        # Hata durumunda database balantsn kapatmay garanti et

        try:

            if 'conn' in locals():

                close_db(conn)

        except:

            pass



        error_msg = f"Saym tamamlama hatas: {str(e)}"

        logging.error(error_msg, exc_info=True)



        return jsonify({

            'success': False,

            'error': 'Saym tamamlanamad - sistem hatas',

            'debug': str(e) if app.debug else None

        }), 500



@app.route('/stop_all_counts', methods=['POST'])

def stop_all_counts():

    """Tm aktif saymlar durdur - ACIL DURUMFONKSYONU"""

    # Admin authentication check

    admin_password = request.json.get('admin_password')

    if admin_password != ADMIN_COUNT_PASSWORD:

        return jsonify({'success': False, 'error': 'Yetki gerekli - yanl admin ifresi'}), 403



    conn = get_db()

    cursor = conn.cursor()



    try:

        # Tm aktif saymlar bul

        execute_query(cursor, "SELECT id FROM count_sessions WHERE is_active = %s", (True,))

        active_sessions = cursor.fetchall()



        if not active_sessions:

            close_db(conn)

            return jsonify({'success': True, 'message': 'Durdurulacak aktif saym bulunamad'})



        # Tm aktif saymlar "completed" olarak iaretle

        stopped_count = 0

        for session_tuple in active_sessions:

            session_id = session_tuple[0]

            execute_query(cursor, 'UPDATE count_sessions SET is_active = %s, ended_at = %s WHERE id = %s',

                         (False, datetime.now(), session_id))

            stopped_count += 1



        # Session'lar temizle

        session.pop('count_access', None)

        session.pop('count_authenticated', None) 

        session.pop('current_session', None)



        conn.commit()

        close_db(conn)



        # WebSocket ile tm kullanclara saymlarn durdurulduunu bildir
        socketio.emit('all_counts_stopped', {
            'message': f'{stopped_count} aktif saym durduruldu',
            'stopped_sessions': [s[0] for s in active_sessions]
        })



        return jsonify({

            'success': True,

            'message': f'{stopped_count} aktif saym baaryla durduruldu',

            'stopped_sessions': [s[0] for s in active_sessions]

        })



    except Exception as e:

        conn.rollback()

        close_db(conn)

        return jsonify({'success': False, 'error': f'Sistem hatas: {str(e)}'}), 500



@app.route('/qr_codes')

def qr_codes_page():

    if 'user_id' not in session:

        return render_template('login.html')

    return render_template('qr_codes.html')



@app.route('/parts')

@login_required

def parts_list():

    """Tüm parçaları listele - QR sayıları ile"""

    # ORM ile parça ve QR kod sayılarını getir (TÜM parçalar, QR olup olmadığına bakılmaksızın)

    from sqlalchemy import func

    parts_data = db.session.query(

        PartCode.id,

        PartCode.part_code,

        PartCode.part_name,

        PartCode.description,

        PartCode.created_at,

        func.count(QRCode.qr_id).label('qr_count')

    ).outerjoin(

        QRCode, PartCode.id == QRCode.part_code_id

    ).group_by(

        PartCode.id, PartCode.part_code, PartCode.part_name, 

        PartCode.description, PartCode.created_at

    ).order_by(

        PartCode.created_at.desc()

    ).all()

    

    parts = []

    for row in parts_data:

        parts.append({

            'id': row[0],

            'part_code': row[1],

            'part_name': row[2],

            'description': row[3] or '',

            'created_at': row[4],

            'qr_count': row[5]

        })

    

    return render_template('parts.html', parts=parts)



@app.route('/parts/<part_code>')

@login_required

def part_detail(part_code):

    """Para detay sayfas - QR kod retme"""

    # ORM ile parça bilgilerini getir

    part_obj = db.session.query(PartCode).filter_by(part_code=part_code).first()

    

    if not part_obj:

        return "Para bulunamad", 404

    

    part = {

        'id': part_obj.id,

        'part_code': part_obj.part_code,

        'part_name': part_obj.part_name,

        'description': part_obj.description or '',

        'created_at': part_obj.created_at

    }

    

    # Bu paraya ait QR kodlarn getir (SADECE KULLANILMAYANLAR)

    unused_qrs = db.session.query(QRCode).filter(

        QRCode.part_code_id == part_obj.id,

        QRCode.is_used == False

    ).order_by(QRCode.qr_id.desc()).all()

    

    qr_codes = []

    for qr in unused_qrs:

        qr_codes.append({

            'qr_id': qr.qr_id,

            'created_at': qr.created_at,

            'is_used': qr.is_used,

            'is_downloaded': 0

        })

    

    # Toplam QR saysn al (kullanlan + kullanlmayan)

    total_qr_count = db.session.query(QRCode).filter(

        QRCode.part_code_id == part_obj.id

    ).count()

    

    # Kullanlan QR saysn al

    used_qr_count = db.session.query(QRCode).filter(

        QRCode.part_code_id == part_obj.id,

        QRCode.is_used == True

    ).count()

    

    return render_template('part_detail.html', part=part, qr_codes=qr_codes, total_qr_count=total_qr_count, used_qr_count=used_qr_count)



@app.route('/generate_qr/<part_code>', methods=['POST'])

@login_required

def generate_qr_codes(part_code):

    """Belirtilen para iin birden fazla QR kod ret (quantity parametresi ile)"""

    try:

        req = request.get_json(silent=True) or {}

        quantity = int(req.get('quantity', 1) or 1)

        if quantity < 1:

            quantity = 1

        max_qty = 500

        if quantity > max_qty:

            return jsonify({'success': False, 'error': f'Quantity too large (max {max_qty})'}), 400



        # ORM ile parça bilgilerini al

        part_obj = db.session.query(PartCode).filter_by(part_code=part_code).first()

        

        if not part_obj:

            return jsonify({'error': 'Para bulunamad'}), 404



        part_code_id = part_obj.id

        part_name = part_obj.part_name



        # Mevcut QR kod saysn ren (tm QR'lar, kullanlm olanlar dahil)

        current_count = db.session.query(QRCode).filter_by(part_code_id=part_code_id).count()



        generated = []

        file_paths = []



        for i in range(quantity):

            qr_number = current_count + i + 1

            qr_id = f"{part_code}_{qr_number}"



            try:

                # ORM ile yeni QR kod oluştur

                new_qr = QRCode(

                    qr_id=qr_id,

                    part_code_id=part_code_id,

                    created_at=datetime.now(),

                    is_used=False

                )

                db.session.add(new_qr)

                db.session.flush()  # ID'yi almak için flush et



                save_qr_code_to_file(part_code, qr_id, qr_number)



                generated.append({

                    'qr_id': qr_id,

                    'qr_number': qr_number,

                    'qr_image_url': f'/qr_image/{part_code}/{qr_id}'

                })



                qr_file_path = os.path.join(QR_CODES_DIR, part_code, f"{qr_id}.png")

                file_paths.append(qr_file_path)

                print(f"DEBUG: QR {qr_id} created successfully")

            except Exception as e:

                # Duplicate olabilir - skip et

                db.session.rollback()

                print(f"DEBUG: QR {qr_id} error: {e}")

                logging.debug(f"QR {qr_id} skip (duplicate): {e}")

                continue



        db.session.commit()



        zip_url = None

        try:

            if len(file_paths) > 1:

                zip_dir = os.path.join(QR_CODES_DIR, part_code)

                os.makedirs(zip_dir, exist_ok=True)

                zip_name = f'bulk_{datetime.now().strftime("%Y%m%d_%H%M%S")}.zip'

                zip_path = os.path.join(zip_dir, zip_name)

                import zipfile

                with zipfile.ZipFile(zip_path, 'w', compression=zipfile.ZIP_DEFLATED) as zf:

                    for fp in file_paths:

                        if os.path.exists(fp):

                            zf.write(fp, arcname=os.path.basename(fp))

                zip_url = f'/download_qr_zip/{part_code}/{zip_name}'

        except Exception:

            logging.exception('ZIP oluturulamad')



        print(f" {part_code} iin {quantity} QR kod retildi. Balang: {current_count + 1}")



        return jsonify({

            'success': True,

            'message': f'{quantity} adet QR kod retildi',

            'generated': generated,

            'zip_url': zip_url

        })



    except Exception as e:

        logging.exception(f"QR kod retme hatas: {e}")

        db.session.rollback()

        return jsonify({'error': f'Hata: {str(e)}'}), 500


@app.route('/download_qr_zip/<part_code>/<zip_name>')
@login_required
def download_qr_zip(part_code, zip_name):
    """QR ZIP dosyasını shared folder'dan serve et"""
    try:
        zip_path = os.path.join(QR_CODES_DIR, part_code, zip_name)
        if os.path.exists(zip_path):
            return send_file(zip_path, mimetype='application/zip', as_attachment=True, download_name=zip_name)
        return jsonify({'error': 'ZIP dosyası bulunamadı'}), 404
    except Exception as e:
        logging.error(f"ZIP indirme hatası: {e}")
        return jsonify({'error': str(e)}), 500


@app.route('/static/part_photos/<path:filename>')
def serve_part_photo(filename):
    """Parça fotoğraflarını shared folder'dan serve et"""
    try:
        # URL decode - %2F gibi karakterleri düzelt
        from urllib.parse import unquote
        filename = unquote(filename)
        
        file_path = os.path.join(PART_PHOTOS_DIR, filename)
        logging.info(f"[PHOTO] Serving: {file_path}")
        
        if os.path.exists(file_path):
            # Dosya uzantısına göre mimetype belirle
            ext = filename.lower().split('.')[-1]
            mimetypes = {
                'jpg': 'image/jpeg',
                'jpeg': 'image/jpeg',
                'png': 'image/png',
                'gif': 'image/gif',
                'webp': 'image/webp'
            }
            mimetype = mimetypes.get(ext, 'image/jpeg')
            return send_file(file_path, mimetype=mimetype)
        else:
            logging.warning(f"[PHOTO] Not found: {file_path}")
            return jsonify({'error': 'Fotoğraf bulunamadı'}), 404
    except Exception as e:
        logging.error(f"Fotoğraf serve hatası: {e}")
        return jsonify({'error': str(e)}), 500


@app.route('/qr_image/<part_code>/<qr_id>')

@login_required

def serve_qr_image(part_code, qr_id):

    """QR kod grselini serve et (text ile birlikte)"""

    try:

        # SHARED FOLDER: Önce statik dosyayı kontrol et
        qr_file = os.path.join(QR_CODES_DIR, part_code, f"{qr_id}.png")
        
        if os.path.exists(qr_file):
            return send_file(qr_file, mimetype='image/png')

        

        # Dosya yoksa dinamik olutur (generate_qr_image gibi)

        # QR ID'den para kodunu ve numaray kar

        parts = qr_id.rsplit('_', 1)

        qr_number = parts[1] if len(parts) > 1 else "1"

        

        # Para adn database'den al

        conn = get_db()

        cursor = conn.cursor()

        execute_query(cursor, 'SELECT part_name FROM part_codes WHERE part_code = %s', (part_code,))

        result = cursor.fetchone()

        part_name = result[0] if result else ""

        close_db(conn)



        # QR kod olutur - Barkod makinesi iin optimize

        qr = qrcode.QRCode(

            version=1, 

            box_size=8,  # 8px - barkod makinesi iin ideal

            border=2,    # 2px quiet zone

            error_correction=qrcode.constants.ERROR_CORRECT_M

        )

        qr.add_data(qr_id)

        qr.make(fit=True)

        qr_img = qr.make_image(fill_color="black", back_color="white")

        

        # PIL Image'a dntr

        qr_img = qr_img.convert('RGB')



        # QR boyutlar

        qr_width, qr_height = qr_img.size

        

        # Alanlar hesapla

        logo_height = 40  # Logo iin st alan

        text_height = 35  # 1 satr - sadece QR ID

        

        # Krmz ereve iin padding

        border_width = 3  # 3px krmz ereve

        

        # Final grsel (ereveli)

        final_width = qr_width + (border_width * 2)

        final_height = logo_height + qr_height + text_height + (border_width * 2)

        final_img = Image.new('RGB', (final_width, final_height), '#dc2626')  # Krmz arka plan (ereve)

        

        # Beyaz i alan olutur (logo + QR + text)

        white_bg = Image.new('RGB', (qr_width, logo_height + qr_height + text_height), 'white')

        

        # Logo ekle (varsa) - st ortasna

        try:

            logo_path = os.path.join(os.path.dirname(__file__), 'cermak-logo.png')

            if os.path.exists(logo_path):

                logo_img = Image.open(logo_path).convert('RGBA')

                # Logo boyutunu alan yksekliine gre ayarla

                logo_width = 150

                logo_height_logo = 40

                try:

                    logo_img = logo_img.resize((logo_width, logo_height_logo), Image.Resampling.LANCZOS)

                except AttributeError:

                    logo_img = logo_img.resize((logo_width, logo_height_logo), Image.LANCZOS)

                

                # Logo'yu ortala

                logo_x = (qr_width - logo_width) // 2

                logo_y = 5  # stten 5px boluk

                

                # RGBA logo'yu blend et

                if logo_img.mode == 'RGBA':

                    alpha = logo_img.split()[3]

                    logo_img = logo_img.convert('RGB')

                    white_bg.paste(logo_img, (logo_x, logo_y), alpha)

                else:

                    white_bg.paste(logo_img, (logo_x, logo_y))

        except Exception as e:

            print(f"Logo ekleme hatas: {e}")

        

        # QR kodu beyaz alana yaptr (logo'nun altna)

        white_bg.paste(qr_img, (0, logo_height))

        

        # Beyaz alan krmz erevenin iine yaptr

        final_img.paste(white_bg, (border_width, border_width))

        

        # Text ekle

        draw = ImageDraw.Draw(final_img)

        try:

            font = ImageFont.truetype("arialblk.ttf", 24)

        except:

            try:

                font = ImageFont.truetype("arialbd.ttf", 24)

            except:

                try:

                    font = ImageFont.truetype("arial.ttf", 24)

                except:

                    font = ImageFont.load_default()

        

        # Sadece QR ID

        qr_text = qr_id

        text_width = len(qr_text) * 14

        x_position = max(border_width, (final_width - text_width) // 2)

        draw.text((x_position, logo_height + qr_height + border_width + 5), qr_text, fill='black', font=font)



        # BytesIO'ya kaydet

        buf = BytesIO()

        final_img.save(buf, format='PNG', optimize=True)

        buf.seek(0)

        

        return send_file(buf, mimetype='image/png')

        

    except Exception as e:

        logging.exception(f"QR grsel hatas: {e}")

        return f"Hata: {str(e)}", 500



@app.route('/mark_qr_used/<qr_id>', methods=['POST'])

@login_required

def mark_qr_used(qr_id):

    """QR kodu manuel olarak 'kullanld' iaretle"""

    try:

        conn = get_db()

        cursor = conn.cursor()

        # QR kodunu kontrol et

        execute_query(cursor, f'SELECT qr_id, is_used FROM qr_codes WHERE qr_id = %s', (qr_id,))

        qr = cursor.fetchone()

        

        if not qr:

            close_db(conn)

            return jsonify({'error': 'QR kod bulunamad'}), 404

        

        if qr[1]:  # is_used

            close_db(conn)

            return jsonify({'error': 'Bu QR kod zaten kullanlm'}), 400

        

        # is_used = True yap

        execute_query(cursor, f'UPDATE qr_codes SET is_used = %s WHERE qr_id = %s',

                     (True, qr_id))

        conn.commit()

        close_db(conn)

        

        print(f" QR kod manuel kullanld iaretlendi: {qr_id}")

        

        return jsonify({

            'success': True,

            'message': f'QR kod kullanld olarak iaretlendi: {qr_id}'

        })

        

    except Exception as e:

        logging.exception(f"QR iaretleme hatas: {e}")

        try:

            close_db(conn)

        except:

            pass

        return jsonify({'error': f'Hata: {str(e)}'}), 500



@app.route('/check_existing_qrs')

@login_required

def check_existing_qrs():

    conn = get_db()

    cursor = conn.cursor()

    execute_query(cursor, 'SELECT COUNT(*) FROM qr_codes')

    count = cursor.fetchone()[0]

    close_db(conn)



    return jsonify({

        'hasQRs': count > 0,

        'count': count

    })



@app.route('/qr_management', methods=['GET'])

@login_required

def qr_management():

    """QR Ynetim Paneli - Gvenli QR ilemleri"""

    if not session.get('admin_authenticated'):

        return redirect('/admin')

    return render_template('qr_management.html')


@app.route('/api/check_qr_files', methods=['POST'])
@login_required
def check_qr_files():
    """QR dosyaları artık doğrudan shared folder'da - kontrol gereksiz"""
    return jsonify({
        'success': True,
        'message': 'QR dosyaları shared folder üzerinden erişilebilir'
    })


@app.route('/get_reports')

@login_required

def get_reports():


    conn = None

    try:

        # Veritabanndan raporlar ek

        conn = get_db()

        cursor = conn.cursor()

        

        execute_query(cursor, '''

            SELECT 

                cr.id,

                cr.session_id,

                cr.report_filename,

                cr.report_title,

                cr.total_expected,

                cr.total_scanned,

                cr.accuracy_rate,

                cr.created_at,

                u.username as created_by

            FROM count_reports cr

            LEFT JOIN count_sessions cs ON cr.session_id = cs.id

            LEFT JOIN envanter_users u ON cs.created_by = u.id

            ORDER BY cr.created_at DESC

        ''')

        

        rows = cursor.fetchall()

        reports = []

        

        for row in rows:

            total_difference = (row[5] or 0) - (row[4] or 0)  # scanned - expected

            

            reports.append({

                'id': row[0],

                'session_id': row[1],

                'filename': row[2],

                'title': row[3] or f"Saym Raporu #{row[1]}",

                'created_at': row[7],

                'total_expected': row[4] or 0,

                'total_scanned': row[5] or 0,

                'accuracy_rate': round(row[6] or 0.0, 2),

                'session_name': f"Oturum #{row[1]}",

                'created_by': row[8] or 'Sistem',

                'total_difference': total_difference

            })

        

        return jsonify(reports)



    except Exception as e:

        logging.exception(f"Error in get_reports: {e}")

        return jsonify({'error': str(e)}), 500

    finally:

        try:

            if conn:

                close_db(conn)

        except Exception:

            pass



@app.route('/download_report/<filename>')

@login_required

def download_report(filename):

    #  Gvenlik: Filename formatn kontrol et (session_id integer olabilir)

    # Format: sayim_raporu_{session_id}_{timestamp}.xlsx

    if not re.match(r'^sayim_raporu_\d+_\d{8}_\d{6}\.xlsx$', filename):

        print(f" Invalid filename format: {filename}")

        return jsonify({'error': 'Geersiz dosya ad format'}), 400



    safe_filename = secure_filename(filename)

    report_path = os.path.join(REPORTS_DIR, safe_filename)



    if not os.path.exists(report_path):

        print(f" Report file not found: {report_path}")

        return jsonify({'error': 'Rapor dosyas bulunamad'}), 404



    #  Path traversal attack'e kar gvenlik kontrol

    real_path = os.path.realpath(report_path)

    reports_real_path = os.path.realpath(REPORTS_DIR)



    if not real_path.startswith(reports_real_path):

        print(f" Invalid path (security): {real_path}")

        return jsonify({'error': 'Geersiz dosya yolu'}), 403



    print(f" Sending report file: {report_path}")

    return send_file(report_path, as_attachment=True, download_name=filename)



#  SAYIM YNETM (Admin Count - ifre sistemi kaldrld)

@app.route('/admin_count')

@login_required

@admin_required_decorator

def admin_count_page():

    """Admin saym kontrol sayfas - Excel ykleme ile saym balatma"""

    return render_template('admin_count.html')



@app.route('/admin_count/start_count', methods=['POST'])

@login_required

@admin_required_decorator

def admin_start_count():

    """Admin saym balatma endpoint'i"""

    print("DEBUG: admin_start_count arld")

    

    try:

        # Aktif saym var m kontrol et

        active_session = db.session.query(CountSession).filter_by(is_active=True).first()

        

        if active_session:

            return jsonify({

                'success': False,

                'error': 'Zaten aktif bir saym oturumu var!'

            }), 400

        

        # 🧹 Yeni sayım başlatıldığında önceki cache'i temizle

        ACTIVE_SESSION_CACHE['session_id'] = None

        ACTIVE_SESSION_CACHE['expected_parts'] = []

        ACTIVE_SESSION_CACHE['cache_time'] = None

        print("🧹 Previous session cache cleared for new count session")

        

        # Yeni saym oturumu olutur (PAROLA YOK ARTIK)

        current_user_id = session.get('user_id')

        

        # Toplam beklenen adet (tüm QR kodlar)

        total_expected = db.session.query(QRCode).count()

        

        count_session = CountSession(

            is_active=True,

            started_at=datetime.now(),

            created_by=current_user_id,

            created_at=datetime.now(),

            total_expected=total_expected,

            total_scanned=0

        )

        

        db.session.add(count_session)

        db.session.commit()

        

        session_id = count_session.id

        

        print(f" Saym oturumu balatld: {session_id}")

        

        #  SOCKET.IO BLDRM: Tm clientlara yeni saym baladn syle
        try:
            socketio.emit('session_reset', {
                'session_id': session_id,
                'total_expected': total_expected,
                'message': ' Yeni saym balatld - sayfa sfrlanyor...'
            }, broadcast=True)
            print(f" Socket bildirimi gnderildi: session_reset")
        except Exception as socket_err:

            print(f" Socket bildirimi gnderilemedi: {socket_err}")

        

        return jsonify({

            'success': True,

            'message': 'Saym oturumu balatld! PSC Scanner ile QR okutmaya balayabilirsiniz.',

            'session_id': session_id,

            'total_expected': total_expected,

            'redirect': '/count-scanner'  # PSC Scanner sayfasna ynlendir

        })

        

    except Exception as e:

        print(f" Saym balatma hatas: {e}")

        import traceback

        traceback.print_exc()

        

        db.session.rollback()

        

        return jsonify({

            'success': False,

            'error': f'Saym balatlamad: {str(e)}'

        }), 500



# API Endpoints for Dashboard Statistics

@app.route('/api/qr_codes')

@login_required

def api_get_qr_codes():

    """QR kodlar listesi - istatistik iin"""

    conn = get_db()

    cursor = conn.cursor()



    execute_query(cursor, '''

        SELECT qc.id, pc.part_code, pc.part_name, qc.is_used, 

               CASE WHEN qc.blob_url IS NOT NULL THEN 1 ELSE 0 END as is_downloaded,

               qc.created_at

        FROM qr_codes qc

        LEFT JOIN part_codes pc ON qc.part_code_id = pc.id

        ORDER BY qc.created_at DESC

    ''')



    qr_codes = []

    for row in cursor.fetchall():

        qr_codes.append({

            'qr_id': row[0],

            'part_code': row[1] or '',

            'part_name': row[2] or 'Bilinmeyen',

            'is_used': bool(row[3]),

            'is_downloaded': bool(row[4]),

            'created_at': row[5]

        })



    close_db(conn)

    return jsonify(qr_codes)


# ======== YAZICI ENTEGRASYONU ENDPOINTS ========

@app.route('/api/printer/status', methods=['GET'])
@login_required
def api_printer_status():
    """Yazıcı durumunu kontrol et"""
    if not PRINTER_ENABLED or _printer_manager is None:
        return jsonify({
            'success': False,
            'enabled': False,
            'status': 'Yazıcı devre dışı'
        }), 200
    
    try:
        status = _printer_manager.get_status()
        return jsonify({
            'success': True,
            'enabled': True,
            'connected': status.get('connected', False),
            'status': status.get('status', 'Bilinmiyor'),
            'device': status.get('device', '/dev/usb/lp0')
        })
    except Exception as e:
        app.logger.error(f"[PRINTER] Durum kontrolü hatası: {e}")
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500


@app.route('/api/printer/print-qr', methods=['POST'])
@login_required
def api_printer_print_qr():
    """QR kodlu etiket yazdır"""
    if not PRINTER_ENABLED or _printer_manager is None:
        return jsonify({
            'success': False,
            'error': 'Yazıcı devre dışı'
        }), 400
    
    try:
        data = request.get_json()
        qr_data = data.get('qr_data', '')
        label_text = data.get('label_text', '')
        quantity = int(data.get('quantity', 1))
        
        if not qr_data:
            return jsonify({
                'success': False,
                'error': 'QR kodu gerekli'
            }), 400
        
        quantity = max(1, min(100, quantity))  # 1-100 arası
        
        success = _printer_manager.print_qr_label(
            qr_data=qr_data,
            label_text=label_text,
            quantity=quantity,
            auto_print=True
        )
        
        if success:
            app.logger.info(f"[PRINTER] QR etiketi yazdırıldı: {qr_data[:20]} x{quantity}")
            return jsonify({
                'success': True,
                'message': f'QR etiketi başarıyla yazdırıldı ({quantity} adet)'
            })
        else:
            return jsonify({
                'success': False,
                'error': 'QR etiketi yazdırılamadı'
            }), 500
    
    except Exception as e:
        app.logger.error(f"[PRINTER] QR yazdırma hatası: {e}")
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500


@app.route('/api/printer/print-barcode', methods=['POST'])
@login_required
def api_printer_print_barcode():
    """Barkodlu etiket yazdır"""
    if not PRINTER_ENABLED or _printer_manager is None:
        return jsonify({
            'success': False,
            'error': 'Yazıcı devre dışı'
        }), 400
    
    try:
        data = request.get_json()
        barcode_data = data.get('barcode_data', '')
        label_text = data.get('label_text', '')
        quantity = int(data.get('quantity', 1))
        barcode_type = data.get('barcode_type', 'CODE128')
        
        if not barcode_data:
            return jsonify({
                'success': False,
                'error': 'Barkod verileri gerekli'
            }), 400
        
        quantity = max(1, min(100, quantity))
        
        success = _printer_manager.print_barcode_label(
            barcode_data=barcode_data,
            barcode_type=barcode_type,
            label_text=label_text,
            quantity=quantity,
            auto_print=True
        )
        
        if success:
            app.logger.info(f"[PRINTER] Barkod etiketi yazdırıldı: {barcode_data} x{quantity}")
            return jsonify({
                'success': True,
                'message': f'Barkod etiketi başarıyla yazdırıldı ({quantity} adet)'
            })
        else:
            return jsonify({
                'success': False,
                'error': 'Barkod etiketi yazdırılamadı'
            }), 500
    
    except Exception as e:
        app.logger.error(f"[PRINTER] Barkod yazdırma hatası: {e}")
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500


@app.route('/api/printer/print-combined', methods=['POST'])
@login_required
def api_printer_print_combined():
    """QR + Barkod kombinasyonu yazdır"""
    if not PRINTER_ENABLED or _printer_manager is None:
        return jsonify({
            'success': False,
            'error': 'Yazıcı devre dışı'
        }), 400
    
    try:
        data = request.get_json()
        qr_data = data.get('qr_data', '')
        barcode_data = data.get('barcode_data', '')
        title = data.get('title', '')
        quantity = int(data.get('quantity', 1))
        
        if not qr_data:
            return jsonify({
                'success': False,
                'error': 'QR kodu gerekli'
            }), 400
        
        quantity = max(1, min(100, quantity))
        
        success = _printer_manager.print_combined_label(
            qr_data=qr_data,
            barcode_data=barcode_data,
            title=title,
            quantity=quantity,
            auto_print=True
        )
        
        if success:
            app.logger.info(f"[PRINTER] Kombinasyon etiketi yazdırıldı: {qr_data[:20]} x{quantity}")
            return jsonify({
                'success': True,
                'message': f'Kombinasyon etiketi başarıyla yazdırıldı ({quantity} adet)'
            })
        else:
            return jsonify({
                'success': False,
                'error': 'Kombinasyon etiketi yazdırılamadı'
            }), 500
    
    except Exception as e:
        app.logger.error(f"[PRINTER] Kombinasyon yazdırma hatası: {e}")
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500


@app.route('/api/printer/test', methods=['POST'])
@admin_required
def api_printer_test():
    """Test etiketi yazdır (Admin)"""
    if not PRINTER_ENABLED or _printer_manager is None:
        return jsonify({
            'success': False,
            'error': 'Yazıcı devre dışı'
        }), 400
    
    try:
        success = _printer_manager.test_print()
        
        if success:
            app.logger.info("[PRINTER] Test etiketi yazdırıldı")
            return jsonify({
                'success': True,
                'message': 'Test etiketi başarıyla yazdırıldı'
            })
        else:
            return jsonify({
                'success': False,
                'error': 'Test etiketi yazdırılamadı'
            }), 500
    
    except Exception as e:
        app.logger.error(f"[PRINTER] Test yazdırma hatası: {e}")
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500


@app.route('/api/reports')

@login_required

def api_get_reports():

    """Raporlar listesi - istatistik iin"""

    conn = get_db()

    cursor = conn.cursor()



    execute_query(cursor, '''

        SELECT id, session_id, report_name, file_path, created_at, 

               total_expected, total_scanned, accuracy_rate

        FROM count_reports

        ORDER BY created_at DESC

    ''')



    reports = []

    for row in cursor.fetchall():

        reports.append({

            'id': row[0],

            'session_id': row[1],

            'report_name': row[2],

            'file_path': row[3],

            'created_at': row[4],

            'total_expected': row[5],

            'total_scanned': row[6],

            'accuracy_rate': row[7]

        })





    close_db(conn)

    return jsonify(reports)



@app.route('/api/dashboard_stats')

def api_dashboard_stats():



    """Dashboard iin genel istatistikler"""

    print("DEBUG: /api/dashboard_stats endpoint arld")  # DEBUG









    # QR kodlar says

    total_qr_codes = db.session.query(QRCode).count()



    # Raporlar says

    total_reports = db.session.query(CountSession).count()



    # Aktif sayım oturumları
    active_counts = db.session.query(CountSession).filter_by(is_active=True).count()
    
    # Tamamlanmış sayım oturumları
    completed_counts = db.session.query(CountSession).filter_by(is_active=False).count()
    
    # Son sayım tarihi
    last_count = db.session.query(CountSession).order_by(CountSession.ended_at.desc()).first()
    last_count_date = last_count.ended_at.isoformat() if last_count and last_count.ended_at else None



    stats = {

        'total_qr_codes': total_qr_codes,

        'total_reports': total_reports,

        'active_counts': active_counts,

        'completed_counts': completed_counts,

        'last_count_date': last_count_date

    }

    print(f"DEBUG: Gnderilen stats: {stats}")  # DEBUG

    return jsonify(stats)



# Eksik endpoint'ler

@app.route('/get_session_stats')

@login_required

def get_session_stats():

    """Saym session istatistikleri"""

    try:

        # URL parametresinden session_id al

        requested_session_id = request.args.get('session_id')

        

        conn = get_db()

        cursor = conn.cursor()

        

        #  Eer session_id verilmise, o session' kullan

        if requested_session_id:

            session_id = requested_session_id

            

            # Test mode iin expected=3

            expected = 3 if requested_session_id.startswith('test-') else 0

            

            # count_sessions tablosundan expected deerini al (varsa)

            try:

                execute_query(cursor, '''

                    SELECT total_expected

                    FROM count_sessions

                    WHERE id = ?

                ''', (session_id,))

                row = cursor.fetchone()

                if row and row[0]:

                    expected = row[0]

            except:

                pass

            

        else:

            # Session ID verilmemise, aktif session bul

            execute_query(cursor, '''

                SELECT id, total_expected, total_scanned

                FROM count_sessions

                WHERE is_active = %s

                ORDER BY started_at DESC

                LIMIT 1

            ''', (True,))



            row = cursor.fetchone()



            if not row:

                close_db(conn)

                return jsonify({'success': False, 'message': 'No active session', 'scanned': 0, 'expected': 0, 'scanned_qrs': []})



            session_id = row[0]

            expected = row[1] if row[1] is not None else 0

        

        # scanned_qr tablosundan gerek scan saysn al
        execute_query(cursor, '''
            SELECT COUNT(DISTINCT qr_id)
            FROM scanned_qr
            WHERE session_id = %s
        ''', (session_id,))

        

        count_row = cursor.fetchone()

        scanned_count = count_row[0] if count_row else 0

        

        # Scanned QR listesi (en son taranan en stte)

        execute_query(cursor, '''

            SELECT qr_id, MAX(scanned_at) as last_scan

            FROM scanned_qr

            WHERE session_id = %s

            GROUP BY qr_id

            ORDER BY last_scan DESC

            LIMIT 500

        ''', (session_id,))

        

        scanned_rows = cursor.fetchall()

        scanned_qrs = [r[0] for r in scanned_rows] if scanned_rows else []



        close_db(conn)



        return jsonify({

            'success': True,

            'session_id': session_id,

            'scanned': scanned_count,

            'expected': expected,

            'scanned_qrs': scanned_qrs

        })

        

    except Exception as e:

        logging.error(f"Error in get_session_stats: {e}")

        return jsonify({'error': str(e)}), 500



@app.route('/get_recent_activities')

@login_required

def get_recent_activities():

    """Son QR tarama aktiviteleri - kullanc adlaryla"""

    try:

        conn = get_db()

        cursor = conn.cursor()

        

        # Get recent scanned QRs with user names

        execute_query(cursor, '''

            SELECT 

                sq.qr_id, 

                sq.scanned_by, 

                sq.scanned_at, 

                sq.part_code,

                eu.full_name,

                eu.username,

                pc.part_name

            FROM scanned_qr sq

            LEFT JOIN envanter_users eu ON sq.scanned_by = eu.id

            LEFT JOIN qr_codes qc ON sq.qr_id = qc.id

            LEFT JOIN part_codes pc ON qc.part_code_id = pc.id

            ORDER BY sq.scanned_at DESC

            LIMIT 10

        ''')

        

        activities = []

        for row in cursor.fetchall():

            user_name = row[4] if row[4] else (row[5] if row[5] else f'Kullanc #{row[1]}')

            part_name = row[6] if row[6] else 'Bilinmeyen rn'

            

            activities.append({

                'qr_code': row[0],

                'scanned_by': user_name,

                'scanned_by_id': row[1],

                'scanned_at': row[2],

                'part_code': row[3] if row[3] else None,

                'part_name': part_name

            })

        

        close_db(conn)

        return jsonify(activities)

    except Exception as e:

        logging.error(f"Error in get_recent_activities: {e}")

        return jsonify([])  # Return empty array instead of error object



@app.route('/get_live_count_status')

@login_required

def get_live_count_status():

    """Aktif saym srasnda para baznda anlk durum"""

    try:

        conn = get_db()

        cursor = conn.cursor()

        # Aktif saym var m

        execute_query(cursor, f'SELECT id FROM count_sessions WHERE is_active = %s', (True,))

        active_session = cursor.fetchone()

        

        if not active_session:

            close_db(conn)

            return jsonify({

                'active': False,

                'message': 'Aktif saym bulunamad'

            })

        

        session_id = active_session[0]

        

        # Para baznda saym durumu

        execute_query(cursor, '''

            SELECT 

                pc.part_code,

                pc.part_name,

                COUNT(DISTINCT qc.id) as beklenen_adet,

                COUNT(DISTINCT sq.qr_id) as sayilan_adet

            FROM part_codes pc

            LEFT JOIN qr_codes qc ON pc.id = qc.part_code_id

            LEFT JOIN scanned_qr sq ON qc.qr_id = sq.qr_id AND sq.session_id = %s

            GROUP BY pc.id, pc.part_code, pc.part_name

            HAVING beklenen_adet > 0

            ORDER BY pc.part_name

        ''', (str(session_id),))

        

        parts = []

        for row in cursor.fetchall():

            part_code = row[0]

            part_name = row[1]

            beklenen = row[2]

            sayilan = row[3]

            kalan = beklenen - sayilan

            durum = 'Tamamland' if kalan == 0 else f'{kalan} eksik'

            yuzd = round((sayilan / beklenen * 100) if beklenen > 0 else 0, 1)

            

            parts.append({

                'part_code': part_code,

                'part_name': part_name,

                'beklenen_adet': beklenen,

                'sayilan_adet': sayilan,

                'kalan_adet': kalan,

                'durum': durum,

                'tamamlanma_yuzdesi': yuzd

            })

        

        close_db(conn)

        return jsonify({

            'active': True,

            'session_id': session_id,

            'parts': parts,

            'total_parts': len(parts),

            'completed_parts': len([p for p in parts if p['kalan_adet'] == 0])

        })

        

    except Exception as e:

        logging.error(f"Error in get_live_count_status: {e}")

        return jsonify({'active': False, 'error': str(e)})



@app.route('/export_live_count')

@login_required

def export_live_count():

    """Anlk saym durumunu Excel olarak indir"""

    try:

        from openpyxl import Workbook

        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

        from io import BytesIO

        

        conn = get_db()

        cursor = conn.cursor()

        # Aktif saym var m

        execute_query(cursor, f'SELECT id FROM count_sessions WHERE is_active = %s', (True,))

        active_session = cursor.fetchone()

        

        if not active_session:

            close_db(conn)

            return jsonify({'error': 'Aktif saym bulunamad'}), 404

        

        session_id = active_session[0]

        

        # Para baznda saym durumu

        execute_query(cursor, '''

            SELECT 

                pc.part_code,

                pc.part_name,

                COUNT(DISTINCT qc.id) as beklenen_adet,

                COUNT(DISTINCT sq.qr_id) as sayilan_adet

            FROM part_codes pc

            LEFT JOIN qr_codes qc ON pc.id = qc.part_code_id

            LEFT JOIN scanned_qr sq ON qc.qr_id = sq.qr_id AND sq.session_id = %s

            GROUP BY pc.id, pc.part_code, pc.part_name

            HAVING beklenen_adet > 0

            ORDER BY pc.part_name

        ''', (str(session_id),))

        

        # Excel olutur

        wb = Workbook()

        ws = wb.active

        ws.title = "Canl Saym Durumu"

        

        # Header

        headers = ['Para Kodu', 'Para Ad', 'Beklenen Adet', 'Saylan Adet', 'Kalan Adet', 'Tamamlanma %', 'Durum']

        ws.append(headers)

        

        # Header stil

        header_fill = PatternFill(start_color="DC3545", end_color="DC3545", fill_type="solid")

        header_font = Font(bold=True, color="FFFFFF", size=12)

        

        for cell in ws[1]:

            cell.fill = header_fill

            cell.font = header_font

            cell.alignment = Alignment(horizontal='center', vertical='center')

        

        # Veri satrlar

        for row in cursor.fetchall():

            beklenen = row[2]

            sayilan = row[3]

            kalan = beklenen - sayilan

            yuzde = round((sayilan / beklenen * 100) if beklenen > 0 else 0, 1)

            durum = ' Tamamland' if kalan == 0 else f' {kalan} eksik'

            

            ws.append([

                row[0],  # part_code

                row[1],  # part_name

                beklenen,

                sayilan,

                kalan,

                f"{yuzde}%",

                durum

            ])

        

        # Kolon genilikleri

        ws.column_dimensions['A'].width = 20

        ws.column_dimensions['B'].width = 35

        ws.column_dimensions['C'].width = 15

        ws.column_dimensions['D'].width = 15

        ws.column_dimensions['E'].width = 15

        ws.column_dimensions['F'].width = 15

        ws.column_dimensions['G'].width = 20

        

        close_db(conn)

        

        # Excel dosyasn BytesIO'ya kaydet

        excel_file = BytesIO()

        wb.save(excel_file)

        excel_file.seek(0)

        

        filename = f"Canli_Sayim_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

        

        return send_file(

            excel_file,

            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',

            as_attachment=True,

            download_name=filename

        )

        

    except Exception as e:

        logging.exception(f"Error in export_live_count: {e}")

        try:

            close_db(conn)

        except:

            pass

        return jsonify({'error': str(e)}), 500



@app.route('/export_qr_activities')

@login_required

def export_qr_activities():

    """QR tarama hareketlerini Excel olarak indir"""

    try:

        from openpyxl import Workbook

        from openpyxl.styles import Font, Alignment, PatternFill

        from io import BytesIO

        

        conn = get_db()

        cursor = conn.cursor()

        

        # Tm QR tarama hareketlerini ek

        execute_query(cursor, '''

            SELECT 

                sq.qr_id, 

                sq.scanned_by, 

                sq.scanned_at, 

                sq.part_code,

                sq.session_id,

                eu.full_name,

                eu.username,

                pc.part_name,

                pc.part_code as part_code_full

            FROM scanned_qr sq

            LEFT JOIN envanter_users eu ON sq.scanned_by = eu.id

            LEFT JOIN qr_codes qc ON sq.qr_id = qc.id

            LEFT JOIN part_codes pc ON qc.part_code_id = pc.id

            ORDER BY sq.scanned_at DESC

        ''')

        

        activities = cursor.fetchall()

        close_db(conn)

        

        # Excel olutur

        wb = Workbook()

        ws = wb.active

        ws.title = "QR Hareketleri"

        

        # Balk satr

        headers = ['QR Kodu', 'Para Kodu', 'Para Ad', 'Okuyan Kullanc', 

                  'Kullanc Ad', 'Okuma Tarihi', 'Seans ID']

        ws.append(headers)

        

        # Balk stilini ayarla

        header_fill = PatternFill(start_color="DC3545", end_color="DC3545", fill_type="solid")

        header_font = Font(bold=True, color="FFFFFF", size=12)

        

        for cell in ws[1]:

            cell.fill = header_fill

            cell.font = header_font

            cell.alignment = Alignment(horizontal='center', vertical='center')

        

        # Veri satrlar

        for row in activities:

            qr_id = row[0]

            scanned_by_id = row[1]

            scanned_at = row[2]

            part_code = row[3] or row[8] or ''  # sq.part_code veya pc.part_code

            session_id = row[4]

            full_name = row[5] or ''

            username = row[6] or f'user_{scanned_by_id}'

            part_name = row[7] or 'Bilinmeyen rn'

            

            ws.append([

                qr_id,

                part_code,

                part_name,

                full_name,

                username,

                scanned_at,

                session_id

            ])

        

        # Kolon geniliklerini ayarla

        ws.column_dimensions['A'].width = 25  # QR Kodu

        ws.column_dimensions['B'].width = 15  # Para Kodu

        ws.column_dimensions['C'].width = 30  # Para Ad

        ws.column_dimensions['D'].width = 20  # Okuyan Kullanc

        ws.column_dimensions['E'].width = 15  # Kullanc Ad

        ws.column_dimensions['F'].width = 20  # Okuma Tarihi

        ws.column_dimensions['G'].width = 10  # Seans ID

        

        # Excel'i memory'ye kaydet

        excel_file = BytesIO()

        wb.save(excel_file)

        excel_file.seek(0)

        

        # Dosya ad

        filename = f"QR_Hareketleri_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

        

        return send_file(

            excel_file,

            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',

            as_attachment=True,

            download_name=filename

        )

        

    except Exception as e:

        logging.error(f"Error in export_qr_activities: {e}")

        return jsonify({'error': str(e)}), 500



@app.route('/get_count_status')

@login_required

def get_count_status():

    """Aktif saym durumu"""

    try:

        conn = get_db()

        cursor = conn.cursor()

        

        execute_query(cursor, '''

            SELECT id, started_at, total_expected, total_scanned, is_active

            FROM count_sessions

            WHERE is_active = %s

            ORDER BY started_at DESC

            LIMIT 1

        ''', (True,))

        

        row = cursor.fetchone()

        close_db(conn)

        

        if row:

            return jsonify({

                'active': True,

                'session_id': row[0],

                'started_at': row[1],

                'total_expected': row[2],

                'total_scanned': row[3]

            })

        else:

            return jsonify({'active': False})

    except Exception as e:

        logging.error(f"Error in get_count_status: {e}")

        return jsonify({'error': str(e)}), 500



#  KULLANICI YNETM (Sadece Admin)

@app.route('/api/users', methods=['GET'])

@login_required

def get_users():

    """Tm kullanclar listele - Sadece Admin"""

    if session.get('role') != 'admin':

        return jsonify({'error': 'Admin eriimi gerekli'}), 403

    

    try:

        conn = get_db()

        cursor = conn.cursor()

        execute_query(cursor, '''

            SELECT id, username, full_name, role, is_active_user, can_mark_used

            FROM envanter_users

            ORDER BY id

        ''')

        users = cursor.fetchall()

        close_db(conn)

        

        user_list = []

        for user in users:

            user_list.append({

                'id': user[0],

                'username': user[1],

                'full_name': user[2],

                'role': user[3],

                'is_active': user[4],

                'can_mark_used': user[5]

            })

        

        return jsonify({'success': True, 'users': user_list})

    except Exception as e:

        logging.error(f"Error in get_users: {e}")

        return jsonify({'error': str(e)}), 500



@app.route('/api/users', methods=['POST'])

@login_required

def create_user():

    """Yeni kullanc olutur - Sadece Admin"""

    if session.get('role') != 'admin':

        return jsonify({'error': 'Admin eriimi gerekli'}), 403

    

    from werkzeug.security import generate_password_hash

    

    try:

        data = request.get_json()

        username = data.get('username')

        password = data.get('password')

        full_name = data.get('full_name')

        role = data.get('role', 'user')  # Varsaylan: user (sadece QR okutabilir)
        
        # Portal erişim izinleri
        can_access_inventory = data.get('can_access_inventory', True)  # Varsayılan: Envanter erişimi var
        can_access_order_system = data.get('can_access_order_system', False)
        can_access_parts_info = data.get('can_access_parts_info', False)
        can_view_purchase_price = data.get('can_view_purchase_price', False)  # Geliş fiyatını görme

        

        if not username or not password or not full_name:

            return jsonify({'error': 'Kullanc ad, ifre ve tam ad gerekli'}), 400

        

        # Sadece 'user' ve 'admin' rolleri kabul et

        if role not in ['user', 'admin']:

            role = 'user'

        

        # Admin saysn kontrol et - Sadece 1 admin olmal

        conn = get_db()

        cursor = conn.cursor()

        

        if role == 'admin':

            execute_query(cursor, 'SELECT COUNT(*) FROM envanter_users WHERE role = %s', ('admin',))

            admin_count = cursor.fetchone()[0]

            if admin_count > 0:

                close_db(conn)

                return jsonify({'error': 'Sistemde zaten bir admin var. Sadece 1 admin olabilir.'}), 400

        

        # Kullanc ad kontrol

        execute_query(cursor, 'SELECT COUNT(*) FROM envanter_users WHERE username = %s', (username,))

        if cursor.fetchone()[0] > 0:

            close_db(conn)

            return jsonify({'error': 'Bu kullanc ad zaten kullanlyor'}), 400

        

        # ifreyi hashle

        password_hash = generate_password_hash(password)

        

        # Kullanc rolne gre izinler

        can_mark_used = (role == 'admin')  # Sadece admin QR'lar kullanlm iaretleyebilir
        
        # Admin ise tüm portallara erişim ver
        if role == 'admin':
            can_access_inventory = True
            can_access_order_system = True
            can_access_parts_info = True
            can_view_purchase_price = True

        

        # Kullancy ekle

        execute_query(cursor, '''

            INSERT INTO envanter_users 

            (username, password, password_hash, full_name, role, is_active_user, can_mark_used,
             can_access_inventory, can_access_order_system, can_access_parts_info, can_view_purchase_price)

            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)

        ''', (username, password, password_hash, full_name, role, True, can_mark_used,
              can_access_inventory, can_access_order_system, can_access_parts_info, can_view_purchase_price))

        

        conn.commit()

        close_db(conn)

        

        return jsonify({

            'success': True,

            'message': f'Kullanc {username} baaryla oluturuldu',

            'user': {

                'username': username,

                'full_name': full_name,

                'role': role,

                'can_mark_used': can_mark_used,
                'can_access_inventory': can_access_inventory,
                'can_access_order_system': can_access_order_system,
                'can_access_parts_info': can_access_parts_info,
                'can_view_purchase_price': can_view_purchase_price

            }

        })

    except Exception as e:

        logging.error(f"Error in create_user: {e}")

        return jsonify({'error': str(e)}), 500



@app.route('/api/users/<int:user_id>', methods=['DELETE'])

@login_required

def delete_user(user_id):

    """Kullanc sil - Sadece Admin"""

    if session.get('role') != 'admin':

        return jsonify({'error': 'Admin eriimi gerekli'}), 403

    

    try:

        conn = get_db()

        cursor = conn.cursor()

        

        # Admin kendini silemez

        if user_id == session.get('user_id'):

            close_db(conn)

            return jsonify({'error': 'Kendi hesabnz silemezsiniz'}), 400

        

        # Kullanc kontrol

        execute_query(cursor, 'SELECT username, role FROM envanter_users WHERE id = ', (user_id,))

        user = cursor.fetchone()

        

        if not user:

            close_db(conn)

            return jsonify({'error': 'Kullanc bulunamad'}), 404

        

        # Admin silinemez

        if user[1] == 'admin':

            close_db(conn)

            return jsonify({'error': 'Admin kullancs silinemez'}), 400

        

        # Sil

        execute_query(cursor, 'DELETE FROM envanter_users WHERE id = ', (user_id,))

        conn.commit()

        close_db(conn)

        

        return jsonify({'success': True, 'message': f'Kullanc {user[0]} silindi'})

    except Exception as e:

        logging.error(f"Error in delete_user: {e}")

        return jsonify({'error': str(e)}), 500



@app.route('/admin')

@login_required

def admin_panel():

    """Admin panel - ifre korumal"""

    # Admin kontrol

    if session.get('role') != 'admin':

        return jsonify({'error': 'Admin eriimi gerekli'}), 403

    

    # Admin panel ifresi kontrol (session'da tutulur)

    if not session.get('admin_panel_unlocked'):

        # ifre girilmemi, ifre sayfasn gster

        return render_template('admin_login.html')

    

    # ifre doru, admin paneli gster

    return render_template('admin.html')



@app.route('/admin/verify', methods=['POST'])

@login_required

def verify_admin_password():

    """Admin panel ifre dorulama"""

    if session.get('role') != 'admin':

        return jsonify({'error': 'Admin eriimi gerekli'}), 403

    

    data = request.get_json()

    password = data.get('password')

    

    # Admin saym ifresi

    ADMIN_PASSWORD = '@R9t$L7e!xP2w'

    

    if password == ADMIN_PASSWORD:

        # ifreyi session'a kaydet

        session['admin_panel_unlocked'] = True

        return jsonify({'success': True, 'message': 'Admin paneline hogeldiniz'})

    else:

        return jsonify({'success': False, 'error': 'Hatal ifre'}), 401



@app.route('/admin/logout', methods=['GET', 'POST'])

@login_required

def admin_panel_logout():

    """Admin panel k (sadece panel kilidini kaldr)"""

    session.pop('admin_panel_unlocked', None)

    return jsonify({'success': True})



@app.route('/admin/users')

@login_required

def admin_users_page():

    """Kullanc ynetim sayfas veya API - GET request"""

    # Admin kontrol

    if session.get('role') != 'admin':

        return jsonify({'error': 'Admin eriimi gerekli'}), 403

    

    # Accept header kontrol - JSON mu HTML mi isteniyor

    if request.accept_mimetypes.best_match(['application/json', 'text/html']) == 'application/json':

        # API ars - Kullanc listesini dndr

        try:

            conn = get_db()

            cursor = conn.cursor()

            

            execute_query(cursor, '''

                SELECT id, username, full_name, role, created_at, is_active_user

                FROM envanter_users

                ORDER BY created_at DESC

            ''')

            

            users = []

            for row in cursor.fetchall():

                users.append({

                    'id': row[0],

                    'username': row[1],

                    'fullname': row[2] or '',

                    'role': row[3],

                    'created_at': row[4],

                    'is_active': row[5] if len(row) > 5 else True

                })

            

            close_db(conn)

            return jsonify(users)

            

        except Exception as e:

            logging.error(f"Error in get users API: {e}")

            return jsonify({'error': str(e)}), 500

    else:

        # HTML sayfas - Kullanc ynetim arayz

        # Admin panel kilidini kontrol et

        if not session.get('admin_panel_unlocked'):

            return redirect('/admin')

        

        # Kullanclar veritabanndan ek

        try:

            conn = get_db()

            cursor = conn.cursor()

            

            execute_query(cursor, '''

                SELECT id, username, full_name, role, created_at, is_active_user

                FROM envanter_users

                ORDER BY created_at DESC

            ''')

            

            users = []

            for row in cursor.fetchall():

                users.append({

                    'id': row[0],

                    'username': row[1],

                    'full_name': row[2] or 'simsiz',

                    'role': row[3],

                    'created_at': row[4],

                    'is_active': row[5] if len(row) > 5 else True

                })

            

            close_db(conn)

            

            # Kullanclar template'e gnder

            return render_template('admin_users.html', users=users)

            

        except Exception as e:

            logging.error(f"Error loading users page: {e}")

            return render_template('admin_users.html', users=[])



@app.route('/admin/users', methods=['POST'])

@login_required

def admin_create_user():

    """Yeni kullanc olutur - Admin Users sayfas"""

    from werkzeug.security import generate_password_hash

    

    # Admin kontrol

    if session.get('role') != 'admin':

        return jsonify({'error': 'Admin eriimi gerekli'}), 403

    

    try:

        # Debug: Request bilgilerini logla

        logging.info(f"Content-Type: {request.content_type}")

        logging.info(f"Request data: {request.data}")

        

        # JSON verisini al - force=True ile Content-Type kontroln bypass et

        data = request.get_json(force=True)

        if not data:

            logging.error("JSON data is None or empty")

            return jsonify({'error': 'Geersiz JSON verisi'}), 400

        

        logging.info(f"Parsed JSON: {data}")

        

        username = str(data.get('username', '')).strip()

        password = str(data.get('password', '')).strip()

        fullname = str(data.get('fullname', '')).strip()

        role = str(data.get('role', 'user')).strip()

        

        # Validasyonlar

        if not username or not password:

            return jsonify({'error': 'Kullanc ad ve ifre gerekli'}), 400

        

        if len(username) < 3:

            return jsonify({'error': 'Kullanc ad en az 3 karakter olmal'}), 400

        

        if len(password) < 4:

            return jsonify({'error': 'ifre en az 4 karakter olmal'}), 400

        

        if role not in ['admin', 'user']:

            return jsonify({'error': 'Geersiz rol. admin veya user olmal'}), 400

        

        # Kullanc zaten var m

        conn = get_db()

        cursor = conn.cursor()

        

        execute_query(cursor, 'SELECT id FROM envanter_users WHERE username = %s', (username,))

        if cursor.fetchone():

            close_db(conn)

            return jsonify({'error': 'Bu kullanc ad zaten kullanlyor'}), 400

        

        # ifreyi hashle

        hashed_password = generate_password_hash(password)

        

        # Yeni kullanıcı ekle (MySQL için RETURNING id kullanma)

        execute_query(cursor, '''

            INSERT INTO envanter_users (username, password_hash, full_name, role, created_at, is_active_user)

            VALUES (%s, %s, %s, %s, %s, %s)

        ''', (username, hashed_password, fullname or '', role, datetime.now(), True))

        

        conn.commit()

        close_db(conn)

        

        return jsonify({'success': True, 'message': f'Kullanc {username} oluturuldu'})

        

    except Exception as e:

        logging.error(f"Error in admin_create_user: {e}")

        return jsonify({'error': str(e)}), 500



@app.route('/admin/users/<int:user_id>', methods=['GET'])
@login_required
def admin_get_user(user_id):
    """Kullanıcı detaylarını getir - yetki bilgileriyle birlikte"""
    # Admin kontrolü
    if session.get('role') != 'admin':
        return jsonify({'error': 'Admin erişimi gerekli'}), 403
    
    try:
        conn = get_db()
        cursor = conn.cursor()
        
        execute_query(cursor, '''
            SELECT id, username, full_name, role, 
                   can_access_inventory, can_access_order_system, 
                   can_access_parts_info, can_view_purchase_price
            FROM envanter_users 
            WHERE id = %s
        ''', (user_id,))
        
        row = cursor.fetchone()
        close_db(conn)
        
        if not row:
            return jsonify({'error': 'Kullanıcı bulunamadı'}), 404
        
        user = {
            'id': row[0],
            'username': row[1],
            'full_name': row[2],
            'role': row[3],
            'can_access_inventory': row[4],
            'can_access_order_system': row[5],
            'can_access_parts_info': row[6],
            'can_view_purchase_price': row[7]
        }
        
        return jsonify({'success': True, 'user': user})
        
    except Exception as e:
        logging.error(f"Error in admin_get_user: {e}")
        return jsonify({'error': str(e)}), 500


@app.route('/admin/users/<int:user_id>', methods=['PUT'])

@login_required

def admin_update_user(user_id):

    """Kullanıcı bilgilerini güncelle - yetkileriyle birlikte"""

    # Admin kontrolü

    if session.get('role') != 'admin':

        return jsonify({'error': 'Admin erişimi gerekli'}), 403

    

    try:

        data = request.get_json()

        if not data:

            return jsonify({'error': 'Geçersiz JSON verisi'}), 400

        

        fullname = str(data.get('fullname', '')).strip()

        role = str(data.get('role', 'user')).strip()

        

        # Yetkileri al (varsayılan değerlerle)
        can_access_inventory = int(data.get('can_access_inventory', 1))
        can_access_order_system = int(data.get('can_access_order_system', 0))
        can_access_parts_info = int(data.get('can_access_parts_info', 0))
        can_view_purchase_price = int(data.get('can_view_purchase_price', 0))
        
        if role not in ['admin', 'user']:

            return jsonify({'error': 'Geçersiz rol'}), 400

        

        conn = get_db()

        cursor = conn.cursor()

        

        # Kullanıcı var mı kontrol et

        execute_query(cursor, 'SELECT id FROM envanter_users WHERE id = %s', (user_id,))

        if not cursor.fetchone():

            close_db(conn)

            return jsonify({'error': 'Kullanıcı bulunamadı'}), 404

        

        # Kullanıcıyı güncelle - yetkileri de ekle

        execute_query(cursor, '''

            UPDATE envanter_users 

            SET full_name = %s, role = %s, updated_at = %s,
                can_access_inventory = %s,
                can_access_order_system = %s,
                can_access_parts_info = %s,
                can_view_purchase_price = %s

            WHERE id = %s

        ''', (fullname, role, datetime.now(),
              can_access_inventory, can_access_order_system,
              can_access_parts_info, can_view_purchase_price,
              user_id))

        

        conn.commit()

        close_db(conn)

        

        return jsonify({'success': True, 'message': 'Kullanıcı güncellendi'})

        

    except Exception as e:

        logging.error(f"Error in admin_update_user: {e}")

        return jsonify({'error': str(e)}), 500



@app.route('/admin/users/<int:user_id>', methods=['DELETE'])

@login_required

def admin_delete_user(user_id):

    """Kullancy sil - Admin Users sayfas"""

    # Admin kontrol

    if session.get('role') != 'admin':

        return jsonify({'error': 'Admin eriimi gerekli'}), 403

    

    try:

        # Kendi hesabn silemesin

        if user_id == session.get('user_id'):

            return jsonify({'error': 'Kendi hesabnz silemezsiniz'}), 400

        

        conn = get_db()

        cursor = conn.cursor()

        

        # Kullanc bilgilerini al

        execute_query(cursor, 'SELECT username FROM envanter_users WHERE id = ', (user_id,))

        user = cursor.fetchone()

        

        if not user:

            close_db(conn)

            return jsonify({'error': 'Kullanc bulunamad'}), 404

        

        # Kullancy sil

        execute_query(cursor, 'DELETE FROM envanter_users WHERE id = ', (user_id,))

        conn.commit()

        close_db(conn)

        

        return jsonify({'success': True, 'message': f'Kullanc {user[0]} silindi'})

        

    except Exception as e:

        logging.error(f"Error in admin_delete_user: {e}")

        return jsonify({'error': str(e)}), 500



@app.route('/admin/users/<int:user_id>/change_password', methods=['POST'])

@login_required

def admin_change_password(user_id):

    """Kullanc ifresini deitir - Admin Users sayfas"""

    from werkzeug.security import generate_password_hash

    

    # Admin kontrol

    if session.get('role') != 'admin':

        return jsonify({'error': 'Admin eriimi gerekli'}), 403

    

    try:

        data = request.get_json()

        if not data:

            return jsonify({'error': 'Geersiz JSON verisi'}), 400

        

        new_password = str(data.get('new_password', '')).strip()

        

        if not new_password:

            return jsonify({'error': 'Yeni ifre gerekli'}), 400

        

        if len(new_password) < 4:

            return jsonify({'error': 'ifre en az 4 karakter olmal'}), 400

        

        conn = get_db()

        cursor = conn.cursor()

        

        # Kullanc var m kontrol et

        execute_query(cursor, 'SELECT username FROM envanter_users WHERE id = ', (user_id,))

        user = cursor.fetchone()

        if not user:

            close_db(conn)

            return jsonify({'error': 'Kullanc bulunamad'}), 404

        

        # ifreyi hashle ve gncelle

        hashed_password = generate_password_hash(new_password)

        execute_query(cursor, '''

            UPDATE envanter_users 

            SET password_hash = %s, last_password_change = %s

            WHERE id = %s

        ''', (hashed_password, datetime.now(), user_id))

        

        conn.commit()

        close_db(conn)

        

        return jsonify({'success': True, 'message': 'ifre deitirildi'})

        

    except Exception as e:

        logging.error(f"Error in admin_change_password: {e}")

        return jsonify({'error': str(e)}), 500



@app.route('/unlock_qrcodes', methods=['POST'])

def unlock_qrcodes():

    """QR klasr kilidini a (admin paneli ifresi ile)"""

    password = request.form.get('password', '').strip()

    

    # Admin paneli ifresi ile kontrol et

    if password == ADMIN_COUNT_PASSWORD:

        session['qrcodes_unlocked'] = True

        # Kullancy qrcodes klasrne ynlendir

        return '''

        <html>

        <head>

            <title>Eriim zni Verildi</title>

            <meta http-equiv="refresh" content="2;url=/static/qrcodes/">

            <style>

                body { font-family: Arial; text-align: center; padding: 50px; background: #f5f5f5; }

                .success { background: white; padding: 30px; border-radius: 10px; max-width: 400px; margin: auto; box-shadow: 0 2px 10px rgba(0,0,0,0.1); }

                h2 { color: #4caf50; }

            </style>

        </head>

        <body>

            <div class="success">

                <h2> Eriim zni Verildi</h2>

                <p>QR klasrne ynlendiriliyorsunuz...</p>

            </div>

        </body>

        </html>

        '''

    else:

        return '''

        <html>

        <head>

            <title>Hatal ifre</title>

            <style>

                body { font-family: Arial; text-align: center; padding: 50px; background: #f5f5f5; }

                .error { background: white; padding: 30px; border-radius: 10px; max-width: 400px; margin: auto; box-shadow: 0 2px 10px rgba(0,0,0,0.1); }

                h2 { color: #d32f2f; }

                a { color: #1976d2; text-decoration: none; }

            </style>

        </head>

        <body>

            <div class="error">

                <h2> Hatal ifre</h2>

                <p>Admin paneli ifresini doru girdiinizden emin olun.</p>

                <p><a href="/static/qrcodes/">Tekrar Dene</a></p>

            </div>

        </body>

        </html>

        ''', 403



@app.route('/admin/reset_active_sessions', methods=['POST'])

@login_required

@admin_required_decorator

def reset_active_sessions():

    """Tm aktif saym oturumlarn sfrla"""

    try:

        conn = get_db()

        cursor = conn.cursor()

        

        # Tüm aktif saymlar pasif yap

        execute_query(cursor, 'UPDATE count_sessions SET is_active = %s WHERE is_active = %s', (False, True))

        

        conn.commit()

        affected_rows = cursor.rowcount

        close_db(conn)

        

        return jsonify({

            'success': True,

            'message': f' {affected_rows} aktif saym oturumu sfrland'

        })

        

    except Exception as e:

        logging.exception(f"Reset active sessions error: {e}")

        return jsonify({

            'success': False,

            'message': f' Hata: {str(e)}'

        }), 500



@app.route('/metrics')

def metrics():

    """Sistem metrikleri"""

    try:

        conn = get_db()

        cursor = conn.cursor()



        # statistikler

        execute_query(cursor, 'SELECT COUNT(*) FROM qr_codes')

        total_qr = cursor.fetchone()[0]



        execute_query(cursor, 'SELECT COUNT(*) FROM qr_codes WHERE is_used = 1')

        used_qr = cursor.fetchone()[0]



        execute_query(cursor, 'SELECT COUNT(*) FROM envanter_users')

        total_users = cursor.fetchone()[0]



        execute_query(cursor, "SELECT COUNT(*) FROM count_sessions WHERE is_active = %s", (True,))

        active_sessions = cursor.fetchone()[0]



        close_db(conn)



        return jsonify({

            'qr_codes': {

                'total': total_qr,

                'used': used_qr,

                'remaining': total_qr - used_qr

            },

            'users': total_users,

            'active_sessions': active_sessions,

            'timestamp': datetime.now().isoformat()

        })

    except Exception as e:

        return jsonify({'error': str(e)}), 500



# ============================================================================
# PORT VE DEPLOYMENT FONKSİYONLARI
# ============================================================================

def get_port():
    """Port numarasını al"""
    return int(os.environ.get('PORT', 5002))


# ============================================================================
# LOKAL SİSTEM - QR Admin modülü kaldırıldı
# ============================================================================



# ==================== CLOUDFLARE TUNNEL YNETM ====================

import subprocess

import threading

import re



# Global deikenler

tunnel_process = None

tunnel_url = None

tunnel_running = False



@app.route('/tunnel/start', methods=['POST'])

@login_required

def start_tunnel():

    """Cloudflare Tunnel balat"""

    global tunnel_process, tunnel_url, tunnel_running

    

    try:

        if tunnel_running and tunnel_process:

            return jsonify({

                'success': True,

                'message': 'Tunnel zaten alyor',

                'url': tunnel_url

            })

        

        # Cloudflared.exe yolu

        cloudflared_path = os.path.join(os.path.dirname(__file__), 'cloudflared.exe')

        

        if not os.path.exists(cloudflared_path):

            return jsonify({

                'success': False,

                'message': f'cloudflared.exe bulunamad: {cloudflared_path}'

            }), 404

        

        # Tunnel balat

        tunnel_process = subprocess.Popen(

            [cloudflared_path, 'tunnel', '--url', 'http://localhost:5002'],

            stdout=subprocess.PIPE,

            stderr=subprocess.PIPE,

            text=True,

            bufsize=1,

            creationflags=subprocess.CREATE_NO_WINDOW if os.name == 'nt' else 0

        )

        

        # URL'yi yakalamak iin arka planda thread balat

        def capture_url():

            global tunnel_url, tunnel_running

            try:

                # Hem stdout hem stderr'i kontrol et

                while True:

                    # stdout'u kontrol et

                    line = tunnel_process.stderr.readline()

                    if not line:

                        break

                    

                    print(f"[Cloudflare] {line.strip()}")

                    

                    # trycloudflare.com URL'sini yakala

                    if 'trycloudflare.com' in line:

                        match = re.search(r'https://[\w-]+\.trycloudflare\.com', line)

                        if match:

                            tunnel_url = match.group(0)

                            tunnel_running = True

                            print(f" Cloudflare Tunnel balatld: {tunnel_url}")

                            break

            except Exception as e:

                print(f" URL yakalama hatas: {e}")

        

        threading.Thread(target=capture_url, daemon=True).start()

        

        # URL'nin yakalanmas iin biraz daha bekle

        import time

        max_wait = 10  # maksimum 10 saniye bekle

        waited = 0

        while not tunnel_url and waited < max_wait:

            time.sleep(0.5)

            waited += 0.5

        

        if tunnel_url:

            return jsonify({

                'success': True,

                'message': 'Tunnel baaryla balatld',

                'url': tunnel_url

            })

        else:

            # Yine de baarl say, arka planda yakalanacak

            tunnel_running = True

            return jsonify({

                'success': True,

                'message': 'Tunnel balatld, URL yaknda hazr olacak',

                'url': None

            })

            

    except Exception as e:

        logging.exception("Tunnel balatma hatas")

        tunnel_running = False

        tunnel_url = None

        return jsonify({

            'success': False,

            'message': f'Hata: {str(e)}'

        }), 500



@app.route('/tunnel/stop', methods=['POST'])

@login_required

def stop_tunnel():

    """Cloudflare Tunnel durdur"""

    global tunnel_process, tunnel_url, tunnel_running

    

    try:

        if not tunnel_running or not tunnel_process:

            return jsonify({

                'success': False,

                'message': 'Tunnel zaten durdurulmu'

            })

        

        # Process'i durdur

        tunnel_process.terminate()

        tunnel_process.wait(timeout=5)

        

        tunnel_process = None

        tunnel_url = None

        tunnel_running = False

        

        print(" Cloudflare Tunnel durduruldu")

        

        return jsonify({

            'success': True,

            'message': 'Tunnel baaryla durduruldu'

        })

        

    except Exception as e:

        logging.exception("Tunnel durdurma hatas")

        return jsonify({

            'success': False,

            'message': f'Hata: {str(e)}'

        }), 500



@app.route('/tunnel/status')

@login_required

def tunnel_status():

    """Tunnel durumunu kontrol et"""

    global tunnel_running, tunnel_url, tunnel_process

    

    # Process hala alyor mu kontrol et

    process_alive = tunnel_process is not None and tunnel_process.poll() is None

    

    return jsonify({

        'running': tunnel_running,

        'url': tunnel_url,

        'process_alive': process_alive,

        'debug': {

            'tunnel_running_flag': tunnel_running,

            'tunnel_url': tunnel_url,

            'process_exists': tunnel_process is not None,

            'process_alive': process_alive

        }

    })



@app.route('/tunnel/logs')

@login_required

def tunnel_logs():

    """Tunnel loglarn gster (debug iin)"""

    global tunnel_process

    

    if not tunnel_process:

        return jsonify({'error': 'Tunnel almyor'}), 404

    

    # Son birka satr oku

    try:

        # stderr'den oku (cloudflared loglar oraya yazyor)

        logs = []

        return jsonify({'logs': logs, 'message': 'Log okuma aktif deil'})

    except Exception as e:

        return jsonify({'error': str(e)}), 500



# ============================================================================

#  DATABASE BACKUP SSTEM

# ============================================================================



def backup_database():

    """SQLite veritabanini yedekle - Veri kaybi korumas"""

    try:

        db_path = 'instance/envanter_local.db'

        if not os.path.exists(db_path):

            app.logger.warning(f'[ERROR] Database dosyas bulunamad: {db_path}')

            return None

            

        backup_dir = 'backups'

        os.makedirs(backup_dir, exist_ok=True)

        

        # Timestamp ile backup ad (Yl-Ay-Gn_Saat-Dakika-Saniye)

        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')

        backup_path = f'{backup_dir}/envanter_backup_{timestamp}.db'

        

        # Orijinal dosya boyutu

        original_size = os.path.getsize(db_path)

        

        # Kopyala

        import shutil

        shutil.copy2(db_path, backup_path)

        

        # Backup boyutu kontrol et

        backup_size = os.path.getsize(backup_path)

        

        # Btnlk kontrol

        if backup_size != original_size:

            app.logger.error(f'[ERROR] Backup btnl baarsz! Orijinal: {original_size}, Backup: {backup_size}')

            os.remove(backup_path)

            return None

        

        # Baarl

        size_mb = round(original_size / (1024**2), 2)

        app.logger.info(f'[OK] Database backup oluturuldu:')

        app.logger.info(f'   Konum: {os.path.abspath(backup_path)}')

        app.logger.info(f'   Boyut: {size_mb} MB')

        app.logger.info(f'   Zaman: {timestamp}')

        

        # Eski backup'lar temizle (son 30'u tut)

        cleanup_old_backups(backup_dir, keep=30)

        

        return backup_path

    except Exception as e:

        app.logger.error(f'[ERROR] Backup hatas: {e}')

        import traceback

        app.logger.error(traceback.format_exc())

        return None



def verify_backup_integrity():

    """Backup dosyalarnn btnln kontrol et"""

    try:

        backup_dir = 'backups'

        if not os.path.exists(backup_dir):

            return

        

        backups = [

            os.path.join(backup_dir, f) 

            for f in os.listdir(backup_dir) 

            if f.startswith('envanter_backup_') and f.endswith('.db')

        ]

        

        if not backups:

            return

        

        # En son backup' kontrol et

        latest_backup = max(backups, key=os.path.getmtime)

        

        try:

            # SQLite database'i a ve kontrol et

            import sqlite3

            conn = sqlite3.connect(latest_backup)

            cursor = conn.cursor()

            cursor.execute("SELECT name FROM sqlite_master WHERE type='table';")

            tables = cursor.fetchall()

            conn.close()

            

            if tables:

                app.logger.info(f'[OK] Backup btnl OK: {os.path.basename(latest_backup)} ({len(tables)} tablo)')

            else:

                app.logger.warning(f'[WARNING] Backup bo grnyor: {os.path.basename(latest_backup)}')

        except Exception as e:

            app.logger.warning(f'[WARNING] Backup kontrol hatas: {os.path.basename(latest_backup)} - {e}')

            

    except Exception as e:

        app.logger.error(f'[ERROR] Backup integrity check hatas: {e}')



def cleanup_old_backups(backup_dir, keep=30):

    """Eski backup'lar temizle - Disk alan tasarrufu"""

    try:

        if not os.path.exists(backup_dir):

            return

            

        backups = sorted([

            os.path.join(backup_dir, f) 

            for f in os.listdir(backup_dir) 

            if f.startswith('envanter_backup_') and f.endswith('.db')

        ], key=os.path.getmtime, reverse=True)

        

        deleted_count = 0

        for old_backup in backups[keep:]:

            try:

                size_mb = round(os.path.getsize(old_backup) / (1024**2), 2)

                os.remove(old_backup)

                deleted_count += 1

                app.logger.info(f'[OK] Eski backup silindi: {os.path.basename(old_backup)} ({size_mb} MB)')

            except Exception as e:

                app.logger.error(f'[ERROR] Backup silme hatas: {old_backup} - {e}')

        

        if deleted_count > 0:

            app.logger.info(f'[OK] {deleted_count} eski backup silindi (Son {keep} tutuldu)')

            

    except Exception as e:

        app.logger.error(f'[ERROR] Backup temizleme hatas: {e}')



def get_backup_list():

    """Mevcut backup'lar listele"""

    try:

        backup_dir = 'backups'

        if not os.path.exists(backup_dir):

            return []

            

        backups = []

        for f in os.listdir(backup_dir):

            if f.startswith('envanter_backup_') and f.endswith('.db'):

                full_path = os.path.join(backup_dir, f)

                backups.append({

                    'filename': f,

                    'path': full_path,

                    'size_mb': round(os.path.getsize(full_path) / (1024**2), 2),

                    'created_at': datetime.fromtimestamp(os.path.getmtime(full_path)).isoformat()

                })

        

        # En yeniden eskiye srala

        backups.sort(key=lambda x: x['created_at'], reverse=True)

        return backups

    except Exception as e:

        app.logger.error(f'[ERROR] Backup listesi hatas: {e}')

        return []



def restore_database(backup_filename):

    """Backup'tan veritabann geri ykle"""

    try:

        backup_path = f'backups/{backup_filename}'

        db_path = 'instance/envanter_local.db'

        

        # Dosya varl kontrol et

        if not os.path.exists(backup_path):

            app.logger.error(f'[ERROR] Backup dosyas bulunamad: {backup_path}')

            return False, 'Backup dosyas bulunamad'

        

        # Mevcut db'nin yedeini al

        import shutil

        if os.path.exists(db_path):

            recovery_backup = f'backups/emergency_recovery_{datetime.now().strftime("%Y%m%d_%H%M%S")}.db'

            shutil.copy2(db_path, recovery_backup)

            app.logger.info(f'[RECOVERY] Kurtarma backup oluturuldu: {recovery_backup}')

        

        # Backup' geri ykle

        shutil.copy2(backup_path, db_path)

        

        app.logger.warning(f'[RESTORE] Database geri yklendi:')

        app.logger.warning(f'   [SRC] Kaynak: {backup_path}')

        app.logger.warning(f'   [DST] Hedef: {db_path}')

        

        return True, f'Database baaryla geri yklendi: {backup_filename}'

        

    except Exception as e:

        app.logger.error(f'[ERROR] Restore hatas: {e}')

        import traceback

        app.logger.error(traceback.format_exc())

        return False, f'Restore hatas: {str(e)}'



@app.route('/admin/backup_now', methods=['POST'])

@login_required

def backup_now():

    """Manuel backup tetikle - sadece admin"""

    # Admin kontrol

    if session.get('role') != 'admin':

        return jsonify({'error': 'Admin eriimi gerekli'}), 403

    

    try:

        app.logger.info('[ADMIN] Admin manuel backup balatt')

        backup_path = backup_database()

        if backup_path:

            return jsonify({

                'success': True, 

                'backup_path': backup_path,

                'message': '[OK] Backup baaryla oluturuldu'

            })

        else:

            return jsonify({

                'success': False, 

                'error': '[ERROR] Backup oluturulamad'

            }), 500

    except Exception as e:

        app.logger.error(f'[ERROR] Backup endpoint hatas: {e}')

        return jsonify({'success': False, 'error': str(e)}), 500



@app.route('/admin/backups', methods=['GET'])

@login_required

def list_backups():

    """Backup listesini getir - sadece admin"""

    # Admin kontrol

    if session.get('role') != 'admin':

        return jsonify({'error': 'Admin eriimi gerekli'}), 403

    

    try:

        backups = get_backup_list()

        backup_dir = os.path.abspath('backups')

        

        return jsonify({

            'success': True,

            'backups': backups,

            'count': len(backups),

            'backup_dir': backup_dir,

            'total_size_mb': sum(b['size_mb'] for b in backups)

        })

    except Exception as e:

        app.logger.error(f'[ERROR] Backup listesi endpoint hatas: {e}')

        return jsonify({'success': False, 'error': str(e)}), 500



@app.route('/admin/restore_backup/<filename>', methods=['POST'])

@login_required

def restore_backup(filename):

    """Backup'tan geri ykle - sadece admin - OK DKKATL!"""

    # Admin kontrol

    if session.get('role') != 'admin':

        return jsonify({'error': 'Admin eriimi gerekli'}), 403

    

    try:

        # Gvenlik: Yalnzca backup dosyalar

        if not filename.startswith('envanter_backup_') or not filename.endswith('.db'):

            return jsonify({'success': False, 'error': 'Geersiz backup dosyas'}), 400

        

        success, message = restore_database(filename)

        

        if success:

            app.logger.warning(f'[RESTORE-ADMIN] Admin database restore yapt: {filename}')

            return jsonify({

                'success': True,

                'message': f'[OK] {message}'

            })

        else:

            return jsonify({

                'success': False,

                'error': f'[ERROR] {message}'

            }), 400

            

    except Exception as e:

        app.logger.error(f'[ERROR] Restore endpoint hatas: {e}')

        return jsonify({'success': False, 'error': str(e)}), 500



@app.route('/admin/backup_status', methods=['GET'])

@login_required

def backup_status():

    """Backup sistem durumu - sadece admin"""

    if session.get('role') != 'admin':

        return jsonify({'error': 'Admin eriimi gerekli'}), 403

    

    try:

        backups = get_backup_list()

        db_path = 'instance/envanter_local.db'

        db_size = 0

        

        if os.path.exists(db_path):

            db_size = round(os.path.getsize(db_path) / (1024**2), 2)

        

        last_backup = None

        if backups:

            last_backup = backups[0]

        

        return jsonify({

            'success': True,

            'current_db_size_mb': db_size,

            'total_backups': len(backups),

            'total_backup_size_mb': sum(b['size_mb'] for b in backups),

            'last_backup': last_backup,

            'backup_dir': os.path.abspath('backups'),

            'schedule': {

                'daily_backup': 'Her gn 02:00\'de',

                'hourly_check': 'Her saat banda'

            }

        })

    except Exception as e:

        app.logger.error(f'[ERROR] Backup status endpoint hatas: {e}')

        return jsonify({'success': False, 'error': str(e)}), 500

        return jsonify({'error': 'Admin eriimi gerekli'}), 403

    

    try:

        # Gvenlik: sadece backup klasrndeki dosyalar

        backup_path = os.path.join('backups', secure_filename(filename))

        

        if not os.path.exists(backup_path):

            return jsonify({'success': False, 'error': 'Backup dosyas bulunamad'}), 404

        

        if not filename.startswith('envanter_backup_'):

            return jsonify({'success': False, 'error': 'Geersiz backup dosyas'}), 400

        

        # Mevcut DB'yi yedekle (geri dn iin)

        current_backup = backup_database()

        

        # Backup' geri ykle

        db_path = 'instance/envanter_local.db'

        import shutil

        shutil.copy2(backup_path, db_path)

        

        app.logger.warning(f'[WARNING] DATABASE RESTORE: {filename} by {session.get("username")}')

        

        return jsonify({

            'success': True,

            'message': 'Backup geri yklendi. Ltfen uygulamay yeniden balatn.',

            'restored_from': filename,

            'safety_backup': current_backup

        })

    except Exception as e:

        app.logger.error(f'Restore hatas: {e}')

        return jsonify({'success': False, 'error': str(e)}), 500



# ========================================

#  PAKET YNETM API ENDPOINT'LER

# ========================================



@app.route('/api/create_package', methods=['POST'])

def create_package():

    """Yeni paket oluturma - Tek QR ile birden fazla para"""

    try:

        data = request.get_json()

        package_name = data.get('package_name', '').strip()

        package_desc = data.get('package_desc', '').strip()

        items = data.get('items', [])

        

        app.logger.info(f'[PAKET OLUŞTUR] Package: {package_name}, Items: {items}')

        

        if not package_name:

            return jsonify({'success': False, 'error': 'Paket ad gerekli'}), 400

        

        if not items or len(items) == 0:

            return jsonify({'success': False, 'error': 'En az 1 para ekleyin'}), 400

        

        # Paket kodunun zaten var olup olmadn kontrol et (sadece paketler için kontrol et)

        conn = get_db()

        cursor = conn.cursor()

        

        execute_query(cursor, 'SELECT part_code FROM part_codes WHERE part_code = %s AND is_package = TRUE', (package_name,))

        if cursor.fetchone():

            close_db(conn)

            return jsonify({'success': False, 'error': 'Bu paket ad zaten kullanmda'}), 400

        

        # JSON formatnda paket ieriini hazrla

        package_items_json = json.dumps(items)

        

        # QR kod resmi olutur (CERMAK + QR + Paket Ad formatlı)

        qr = qrcode.QRCode(

            version=1,

            box_size=10,

            border=2,

            error_correction=qrcode.constants.ERROR_CORRECT_M

        )

        qr.add_data(package_name)

        qr.make(fit=True)

        

        qr_img = qr.make_image(fill_color="black", back_color="white")

        qr_img = qr_img.convert('RGB')

        

        # QR kod boyutlarn al

        qr_width, qr_height = qr_img.size

        

        # Alanlar hesapla

        logo_height = 40  # CERMAK yazısı için üst alan

        text_height = 35  # Alt yazı (paket adı) için alan

        border_width = 3  # 3px kırmızı çerçeve

        

        # Yeni görsel oluştur (CERMAK + QR + text alanı + çerçeve)

        final_width = qr_width + (border_width * 2)

        final_height = logo_height + qr_height + text_height + (border_width * 2)

        final_img = Image.new('RGB', (final_width, final_height), '#dc2626')  # Kırmızı arka plan (çerçeve)

        

        # Beyaz iç alan oluştur (CERMAK + QR + text)

        white_bg = Image.new('RGB', (qr_width, logo_height + qr_height + text_height), 'white')

        

        # CERMAK yazısı ekle - üst ortasına

        try:

            # Font (kalın ve büyük) - CERMAK yazısı için

            try:

                cermak_font = ImageFont.truetype("arialblk.ttf", 32)

            except:

                try:

                    cermak_font = ImageFont.truetype("arialbd.ttf", 32)

                except:

                    try:

                        cermak_font = ImageFont.truetype("arial.ttf", 32)

                    except:

                        cermak_font = ImageFont.load_default()

            

            draw_temp = ImageDraw.Draw(white_bg)

            cermak_text = "CERMAK"

            # Gerçek text genişliğini ölç

            bbox = draw_temp.textbbox((0, 0), cermak_text, font=cermak_font)

            text_width = bbox[2] - bbox[0]

            x_pos = (qr_width - text_width) // 2

            y_pos = 5

            draw_temp.text((x_pos, y_pos), cermak_text, fill='black', font=cermak_font)

        except Exception as e:

            print(f"CERMAK yazısı ekleme hatası: {e}")

        

        # QR kodu beyaz alana yapıştır (logo'nun altına)

        white_bg.paste(qr_img, (0, logo_height))

        

        # Beyaz alanı kırmızı çerçevenin içine yapıştır

        final_img.paste(white_bg, (border_width, border_width))

        

        # Alt text ekleme için draw nesnesi

        draw = ImageDraw.Draw(final_img)

        

        # Font (kalın ve büyük) - Paket adı için

        try:

            font = ImageFont.truetype("arialblk.ttf", 24)

        except:

            try:

                font = ImageFont.truetype("arialbd.ttf", 24)

            except:

                try:

                    font = ImageFont.truetype("arial.ttf", 24)

                except:

                    font = ImageFont.load_default()

        

        # Paket adını (örn: JCB1) QR kodunun altına ekle

        bbox = draw.textbbox((0, 0), package_name, font=font)

        text_width = bbox[2] - bbox[0]

        x_position = border_width + (qr_width - text_width) // 2

        y_position = border_width + logo_height + qr_height + 5

        draw.text((x_position, y_position), package_name, fill='black', font=font)

        

        # Base64'e çevir

        buffer = BytesIO()

        final_img.save(buffer, 'PNG')

        qr_base64 = base64.b64encode(buffer.getvalue()).decode()

        

        # Paketi veritabanna ekle

        execute_query(cursor, '''

            INSERT INTO part_codes (part_code, part_number, part_name, is_package, package_items)

            VALUES (%s, %s, %s, %s, %s)

        ''', (package_name, package_name, package_desc if package_desc else f'Paket: {package_name}', True, package_items_json))

        

        # package_id'yi getir

        execute_query(cursor, 'SELECT id FROM part_codes WHERE part_code = %s', (package_name,))

        result = cursor.fetchone()

        package_id = result[0] if result else None

        

        if not package_id:

            close_db(conn)

            return jsonify({'success': False, 'error': 'Paket oluturulurken hata olutu'}), 500

        

        # QR codes tablosuna da ekle (tek bir QR)

        execute_query(cursor, '''

            INSERT INTO qr_codes (qr_id, part_code_id, is_used, created_at)

            VALUES (%s, %s, %s, %s)

        ''', (package_name, package_id, False, datetime.now()))

        

        conn.commit()

        close_db(conn)

        

        app.logger.info(f'[PAKET] Yeni paket oluturuldu: {package_name} ({len(items)} para)')

        

        return jsonify({

            'success': True,

            'message': f'Paket "{package_name}" oluturuldu',

            'package_code': package_name,

            'qr_image': qr_base64,  # Base64 QR kodu

            'items_count': len(items),

            'total_quantity': sum(item.get('quantity', 1) for item in items)

        })

        

    except Exception as e:

        app.logger.error(f'[HATA] Paket oluturma hatas: {e}')

        return jsonify({'success': False, 'error': str(e)}), 500



@app.route('/api/get_packages', methods=['GET'])

def get_packages():

    """Tm paketleri listele"""

    try:

        conn = get_db()

        cursor = conn.cursor()

        

        cursor.execute('''

            SELECT part_code, part_name, package_items

            FROM part_codes

            WHERE is_package = TRUE

            ORDER BY part_code

        ''')

        

        packages = []

        for row in cursor.fetchall():

            try:

                items = json.loads(row[2]) if row[2] else []

            except:

                items = []

            

            # Her paket iin QR kod olutur

            qr = qrcode.QRCode(version=1, box_size=10, border=5)

            qr.add_data(row[0])  # part_code

            qr.make(fit=True)

            

            img = qr.make_image(fill_color="black", back_color="white")

            buffer = BytesIO()

            img.save(buffer, format='PNG')

            qr_base64 = base64.b64encode(buffer.getvalue()).decode()

            

            packages.append({

                'part_code': row[0],

                'part_name': row[1],

                'items': items,

                'items_count': len(items),

                'total_quantity': sum(item.get('quantity', 1) for item in items),

                'qr_image': qr_base64  # QR kod base64

            })

        

        conn.close()

        

        return jsonify({

            'success': True,

            'packages': packages,

            'count': len(packages)

        })

        

    except Exception as e:

        app.logger.error(f'[HATA] Paket listeleme hatas: {e}')

        return jsonify({'success': False, 'error': str(e)}), 500



@app.route('/api/delete_package', methods=['POST'])

def delete_package():

    """Paket silme"""

    try:

        data = request.get_json()

        part_code = data.get('part_code', '').strip()

        

        if not part_code:

            return jsonify({'success': False, 'error': 'Paket kodu gerekli'}), 400

        

        conn = get_db()

        cursor = conn.cursor()

        

        # Paketin varlığını kontrol et

        cursor.execute('SELECT is_package FROM part_codes WHERE part_code = %s', (part_code,))

        row = cursor.fetchone()

        

        if not row:

            return jsonify({'success': False, 'error': 'Paket bulunamadı'}), 404

        

        if not row[0]:

            return jsonify({'success': False, 'error': 'Bu bir paket değil'}), 400

        

        # Önce qr_codes tablosundan sil (foreign key constraint)

        cursor.execute('DELETE FROM qr_codes WHERE qr_id = %s', (part_code,))

        

        # Sonra paketi part_codes'dan sil

        cursor.execute('DELETE FROM part_codes WHERE part_code = %s', (part_code,))

        conn.commit()

        conn.close()

        

        app.logger.info(f'[PAKET] Paket silindi: {part_code}')

        

        return jsonify({

            'success': True,

            'message': f'Paket "{part_code}" silindi'

        })

        

    except Exception as e:

        app.logger.error(f'[HATA] Paket silme hatas: {e}')

        return jsonify({'success': False, 'error': str(e)}), 500



#  CANLI DASHBOARD API - Gerek zamanl istatistikler

@app.route('/api/live_dashboard/<int:session_id>')

@login_required

def live_dashboard_stats(session_id):

    """

    Canl dashboard iin tm istatistikler

    

    Returns:

        - Tarama hz (son 5dk, son 1 saat)

        - En ok taranan 10 para

        - Kullanc performans

        - Saym biti tahmini

        - Saatlik tarama dalm

    """

    # Session'dan kullanc adn al

    username = session.get('username', 'Bilinmeyen')

    logging.info(f"Dashboard API called - Session: {session_id}, User: {username}")

    

    try:

        with db_connection() as conn:

            cursor = conn.cursor()

            

            # 1. Tarama Hz (Scans per Minute)

            execute_query(cursor, '''

                SELECT 

                    COUNT(*) as total_scans,

                    MIN(scanned_at) as first_scan,

                    MAX(scanned_at) as last_scan

                FROM scanned_qr
                WHERE session_id = %s
            ''', (session_id,))

            scan_stats = cursor.fetchone()

            

            total_scans = scan_stats[0] if scan_stats else 0

            first_scan = scan_stats[1] if scan_stats else None

            last_scan = scan_stats[2] if scan_stats else None

            

            logging.debug(f"Total scans: {total_scans}, First: {first_scan}, Last: {last_scan}")

            

            # Tarama hz hesapla

            scan_rate_5min = 0

            scan_rate_1hour = 0

            

            if last_scan:

                # Son 5 dakika

                execute_query(cursor, '''

                    SELECT COUNT(*) FROM scanned_qr

                    WHERE session_id = %s AND scanned_at >= NOW() - INTERVAL 5 MINUTE

                ''', (session_id,))

                result = cursor.fetchone()

                scans_5min = result[0] if result else 0

                scan_rate_5min = round(scans_5min / 5, 1)  # tarama/dakika

                

                # Son 1 saat

                execute_query(cursor, '''

                    SELECT COUNT(*) FROM scanned_qr

                    WHERE session_id = %s AND scanned_at >= NOW() - INTERVAL 1 HOUR

                ''', (session_id,))

                result = cursor.fetchone()

                scans_1hour = result[0] if result else 0

                scan_rate_1hour = round(scans_1hour / 60, 1)  # tarama/dakika

            

            # 2. En ok Taranan 10 Para

            execute_query(cursor, '''

                SELECT 

                    pc.part_name,

                    pc.part_code,

                    COUNT(*) as scan_count

                FROM scanned_qr sq

                JOIN part_codes pc ON sq.part_code = pc.part_code

                WHERE sq.session_id = %s

                GROUP BY pc.part_code, pc.part_name

                ORDER BY scan_count DESC

                LIMIT 10

            ''', (session_id,))

            top_parts = [{'name': row[0], 'code': row[1], 'count': row[2]} for row in cursor.fetchall()]

            

            # 3. Kullanc Performans

            execute_query(cursor, '''

                SELECT 

                    u.full_name,

                    u.username,

                    COUNT(*) as scan_count,

                    MIN(sq.scanned_at) as first_scan,

                    MAX(sq.scanned_at) as last_scan

                FROM scanned_qr sq

                JOIN envanter_users u ON sq.scanned_by = u.id

                WHERE sq.session_id = %s

                GROUP BY u.id, u.full_name, u.username

                ORDER BY scan_count DESC

            ''', (session_id,))

            

            user_performance = []

            for row in cursor.fetchall():

                # PostgreSQL can return datetime objects or strings
                if isinstance(row[3], str):
                    user_first = datetime.fromisoformat(row[3]) if row[3] else None
                else:
                    user_first = row[3]

                if isinstance(row[4], str):
                    user_last = datetime.fromisoformat(row[4]) if row[4] else None
                else:
                    user_last = row[4]

                work_time_minutes = 0

                personal_rate = 0

                

                if user_first and user_last:

                    work_time_seconds = (user_last - user_first).total_seconds()

                    work_time_minutes = round(work_time_seconds / 60, 1)

                    if work_time_minutes > 0:

                        personal_rate = round(row[2] / work_time_minutes, 1)

                

                user_performance.append({

                    'name': row[0],

                    'username': row[1],

                    'scan_count': row[2],

                    'work_time_minutes': work_time_minutes,

                    'scans_per_minute': personal_rate

                })

            

            # 4. Saym Biti Tahmini

            execute_query(cursor, '''

                SELECT COUNT(DISTINCT qr_id) FROM qr_codes

            ''')

            total_qr_codes = cursor.fetchone()[0]

            remaining = total_qr_codes - total_scans

            

            eta_minutes = 0

            eta_text = "Hesaplanyor..."

            

            if scan_rate_1hour > 0 and remaining > 0:

                eta_minutes = round(remaining / scan_rate_1hour)

                if eta_minutes < 60:

                    eta_text = f"{eta_minutes} dakika"

                else:

                    eta_hours = round(eta_minutes / 60, 1)

                    eta_text = f"{eta_hours} saat"

            

            # 5. Saatlik Tarama Dalm (Son 24 saat)

            execute_query(cursor, '''

                SELECT 

                    HOUR(scanned_at) as hour,

                    COUNT(*) as count

                FROM scanned_qr

                WHERE session_id = %s AND scanned_at >= NOW() - INTERVAL 24 HOUR

                GROUP BY HOUR(scanned_at)

                ORDER BY HOUR(scanned_at)

            ''', (session_id,))

            

            hourly_distribution = {}

            for row in cursor.fetchall():

                hourly_distribution[row[0]] = row[1]

            

            result = {

                'success': True,

                'session_id': session_id,

                'total_scans': total_scans,

                'remaining': remaining,

                'progress_percent': round((total_scans / total_qr_codes * 100) if total_qr_codes > 0 else 0, 1),

                'scan_rate': {

                    'last_5min': scan_rate_5min,

                    'last_1hour': scan_rate_1hour,

                    'unit': 'tarama/dakika'

                },

                'top_parts': top_parts,

                'user_performance': user_performance,

                'eta': {

                    'minutes': eta_minutes,

                    'text': eta_text

                },

                'hourly_distribution': hourly_distribution,

                'first_scan': first_scan,

                'last_scan': last_scan

            }

            

            logging.info(f"Dashboard API success - {total_scans} scans, {len(user_performance)} users")

            return jsonify(result)

            

    except Exception as e:

        logging.error(f"Live dashboard error: {e}", exc_info=True)

        return jsonify({'success': False, 'error': str(e)}), 500



# Otomatik backup scheduler iin (APScheduler kullanarak)

try:

    from apscheduler.schedulers.background import BackgroundScheduler

    

    backup_scheduler = BackgroundScheduler()

    

    #  GNLK BACKUP: Her gn 02:00'de yaplr

    backup_scheduler.add_job(

        func=backup_database,

        trigger="cron",

        hour=2,

        minute=0,

        id='daily_auto_backup',

        name='Gnlk Otomatik Database Backup',

        replace_existing=True

    )

    

    #  SAATLK KONTROL: Her saat banda backup kontrol

    backup_scheduler.add_job(

        func=verify_backup_integrity,

        trigger="cron",

        minute=0,

        id='hourly_backup_check',

        name='Saatlik Backup Btnl Kontrol',

        replace_existing=True

    )
    
    # QR Sync artık shared folder üzerinden çalışıyor, scheduler job'a gerek yok

    backup_scheduler.start()

    app.logger.info('[OK] Backup Scheduler Balatld:')

    app.logger.info('   [DAILY] Gnlk Backup: Her gn 02:00\'de')

    app.logger.info('   [HOURLY] Saatlik Kontrol: Her saat banda')

    app.logger.info(f'   [PATH] Backup Klasr: {os.path.abspath("backups")}')

except ImportError:

    app.logger.warning('[!] APScheduler ykl deil. Otomatik backup devre d.')

    app.logger.warning('   Yklemek iin: pip install apscheduler')

except Exception as e:

    app.logger.error(f'Backup scheduler hatas: {e}')



# NOTE: This is handled by render_startup_alt.py for Render.com

# DO NOT call socketio.run() here to avoid port binding conflicts

# ============================================================================
# APPLICATION STARTUP
# ============================================================================

if __name__ == '__main__':
    # Initialize database on startup
    try:
        init_db()
        init_parts_info_table()  # Yedek Parça Bilgi Sistemi tablosu
        update_parts_info_columns()  # Parts Info yeni kolonları ekle
        init_user_permissions()  # Kullanıcı portal erişim izinleri
        update_admin_permissions()  # Admin kullanıcılarına tüm yetkileri ver
        

        # QR CACHE YÜKLEME - Ultra hızlı tarama için
        print("[CACHE] Loading QR cache into memory...")
        if load_qr_cache_to_memory():
            print(f"[OK] QR Cache loaded: {len(QR_LOOKUP_CACHE)} codes ready!")
        else:
            print("[WARN] Cache loading failed - will load on first request")
        
        # MySQL AUTO_INCREMENT değerlerini kontrol et

        print("[DB] Checking MySQL AUTO_INCREMENT values...")

        try:
            conn = get_db()
            cursor = conn.cursor()
            
            tables = ['scanned_qr', 'count_sessions', 'qr_codes', 'part_codes', 'envanter_users']
            for table in tables:
                cursor.execute(f"SELECT MAX(id) FROM {table}")
                max_id = cursor.fetchone()[0]
                if max_id:
                    cursor.execute(f"ALTER TABLE {table} AUTO_INCREMENT = {max_id + 1}")
                    logging.info(f"[OK] {table} AUTO_INCREMENT set to {max_id + 1}")
            
            conn.commit()
            close_db(conn)
            print("[OK] AUTO_INCREMENT values checked!")
        except Exception as e:
            logging.error(f"AUTO_INCREMENT check error: {e}")

        

    except Exception as e:

        print(f" Failed to initialize database: {e}")


# ============================================================================
# SHARED FOLDER SYNC API ENDPOINTS
# ============================================================================

@app.route('/api/sync_status')
@login_required
def sync_status():
    """Paylaşımlı klasör senkronizasyon durumunu kontrol et"""
    try:
        from shared_folder_sync import SharedFolderSyncManager
        
        manager = SharedFolderSyncManager()
        
        # Paylaşımlı klasör erişilebilir mi?
        import os
        shared_accessible = os.path.exists(manager.SHARED_FOLDER_PATH)
        
        status = {
            'connected': shared_accessible,
            'storage_type': 'shared_folder',
            'shared_folder': manager.SHARED_FOLDER_PATH,
            'categories': {}
        }
        
        if shared_accessible:
            # Her kategori için durum
            for category, paths in manager.SYNC_CATEGORIES.items():
                local_files = manager.get_local_files(category)
                shared_files = manager.get_shared_files(category)
                
                local_paths = {f['path'] for f in local_files}
                shared_paths = {f['path'] for f in shared_files}
                
                status['categories'][category] = {
                    'local_count': len(local_files),
                    'shared_count': len(shared_files),
                    'missing_in_shared': len(local_paths - shared_paths),
                    'missing_locally': len(shared_paths - local_paths),
                    'synced': len(local_paths & shared_paths)
                }
                
                app.logger.info(f"[SYNC STATUS] {category}: local={len(local_files)}, shared={len(shared_files)}")
        
        app.logger.info(f"[SYNC STATUS] Returning status: {status}")
        return jsonify({'success': True, 'status': status})
        
    except Exception as e:
        app.logger.error(f"Sync status error: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/sync_category/<category>', methods=['POST'])
@login_required
def sync_category(category):
    """Belirli bir kategoriyi senkronize et"""
    try:
        from shared_folder_sync import SharedFolderSyncManager
        
        data = request.get_json() or {}
        direction = data.get('direction', 'both')  # 'upload', 'download', 'both'
        
        manager = SharedFolderSyncManager()
        
        if category not in manager.SYNC_CATEGORIES:
            return jsonify({'success': False, 'error': 'Invalid category'}), 400
        
        stats = manager.sync_category(category, direction)
        
        return jsonify({
            'success': True,
            'category': category,
            'direction': direction,
            'stats': stats
        })
        
    except Exception as e:
        app.logger.error(f"Sync category error: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/sync_all', methods=['POST'])
@login_required
def sync_all():
    """Tüm kategorileri senkronize et"""
    try:
        from shared_folder_sync import SharedFolderSyncManager
        
        data = request.get_json() or {}
        direction = data.get('direction', 'both')
        
        manager = SharedFolderSyncManager()
        stats = manager.full_sync(direction)
        
        return jsonify({
            'success': True,
            'direction': direction,
            'stats': stats
        })
        
    except Exception as e:
        app.logger.error(f"Sync all error: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500


# ============================================================================
# PARÇA SİPARİŞ SİSTEMİ MOD ÜLÜ - ENVANTER SİSTEMİNDEN BAĞIMSIZ
# ============================================================================
from order_system import register_order_system
register_order_system(app)

# === ENVANTER SİSTEMİ İZOLASYONU ===
from inventory_isolation import protect_inventory_tables, verify_system_isolation
protect_inventory_tables()
verify_system_isolation()


# ==================== ADMIN/SYSTEM ROUTES ====================

@app.route('/api/fix_package_flags', methods=['POST'])
def fix_package_flags():
    """Fix packages with is_package=FALSE but have package_items data"""
    try:
        conn = get_db()
        cursor = conn.cursor()
        
        # Find packages with is_package=FALSE but have package_items
        cursor.execute('''
            SELECT part_code, part_name, package_items 
            FROM part_codes 
            WHERE is_package = FALSE AND package_items IS NOT NULL AND package_items != ''
        ''')
        
        broken_packages = cursor.fetchall()
        
        if not broken_packages:
            return jsonify({'success': True, 'message': 'Tüm paketler OK', 'count': 0}), 200
        
        # Fix them by setting is_package=TRUE
        fixed_count = 0
        for pkg in broken_packages:
            part_code = pkg[0]
            try:
                cursor.execute('UPDATE part_codes SET is_package = TRUE WHERE part_code = %s', (part_code,))
                app.logger.warning(f'[FIX] Fixed package {part_code} - set is_package=TRUE')
                fixed_count += 1
            except Exception as e:
                app.logger.error(f'[FIX ERROR] Failed to fix {part_code}: {e}')
        
        conn.commit()
        close_db(conn)
        
        return jsonify({
            'success': True, 
            'message': f'{fixed_count} paket düzeltildi',
            'count': fixed_count,
            'packages': [pkg[0] for pkg in broken_packages]
        }), 200
        
    except Exception as e:
        app.logger.error(f'[FIX ERROR] {e}')
        return jsonify({'success': False, 'error': str(e)}), 500


# ============================================================
# TAKEUCHI PARÇA SİPARİŞ MODÜLÜ - ROUTES
# ============================================================

@app.route('/takeuchi', methods=['GET'])
@login_required
def takeuchi_main():
    """Takeuchi Ana Sayfası - Basit Menü"""
    return render_template('takeuchi/main.html')


@app.route('/takeuchi/add', methods=['GET'])
@login_required
def takeuchi_add_page():
    """Parça Ekle Sayfası"""
    return render_template('takeuchi/add_part.html')


@app.route('/takeuchi/check', methods=['GET'])
@login_required
def takeuchi_check_page():
    """Parça Kontrol Et Sayfası"""
    return render_template('takeuchi/check_part.html')


@app.route('/takeuchi/admin', methods=['GET'])
@admin_required
def takeuchi_admin_page():
    """Admin Panel - Geçici Siparişleri Yönet"""
    return render_template('takeuchi/admin.html')


# ============================================================
# TAKEUCHI API ENDPOINTS
# ============================================================

@app.route('/api/takeuchi/init-session', methods=['POST'])
@login_required
def api_takeuchi_init_session():
    """Kullanıcı için yeni geçici sipariş oturumu başlat"""
    try:
        user_id = session.get('user_id')
        result = TakeuchiOrderManager.create_temp_order_session(user_id)
        return jsonify(result)
    except Exception as e:
        app.logger.error(f'[TAKEUCHI] Error init session: {e}')
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/takeuchi/part-info', methods=['POST'])
@login_required
def api_takeuchi_part_info():
    """Parça bilgileri ve geçmişini al"""
    try:
        data = request.get_json()
        part_code = data.get('part_code', '').strip().upper()
        
        if not part_code:
            return jsonify({'success': False, 'error': 'Parça kodu gerekli'}), 400
        
        # Parça bilgisini al
        part = db.session.query(PartCode).filter_by(part_code=part_code).first()
        if not part:
            return jsonify({
                'success': False,
                'error': f'Parça kodu bulunamadı: {part_code}'
            }), 404
        
        # Geçmişi al
        history_result = TakeuchiOrderManager.get_part_history(part_code)
        
        return jsonify({
            'success': True,
            'part_code': part.part_code,
            'part_name': part.part_name,
            'description': part.description,
            'history': history_result.get('history', []) if history_result.get('success') else []
        })
    except Exception as e:
        app.logger.error(f'[TAKEUCHI] Error getting part info: {e}')
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/takeuchi/search-part', methods=['GET'])
@login_required
def api_takeuchi_search_part():
    """Parça kodu ile Takeuchi parçasını ara ve bilgilerini döndür"""
    try:
        part_code = request.args.get('code', '').strip()
        
        if not part_code:
            return jsonify({'success': False, 'error': 'Parça kodu gerekli'}), 400
        
        # TakeuchiPart tablosunda ara
        part = TakeuchiPart.query.filter_by(part_code=part_code).first()
        
        if not part:
            return jsonify({'success': False, 'error': 'Parça bulunamadı'}), 404
        
        return jsonify({
            'success': True,
            'part': {
                'id': part.id,
                'part_code': part.part_code,
                'part_name': part.part_name,
                'build_out': part.build_out or '',
                'alternative_code': part.alternative_code or '',
                'description': part.description or '',
                'unit_price': float(part.unit_price) if part.unit_price else 0
            }
        })
    except Exception as e:
        app.logger.error(f'[TAKEUCHI] Search part error: {e}')
        return jsonify({'success': False, 'error': f'Arama hatası: {str(e)}'}), 500


@app.route('/api/takeuchi/add-part', methods=['POST'])
@login_required
def api_takeuchi_add_part():
    """Geçici siparişe parça ekle"""
    try:
        data = request.get_json()
        session_id = data.get('session_id')
        part_code = data.get('part_code', '').strip().upper()
        quantity = data.get('quantity', 1)
        user_id = session.get('user_id')
        
        if not session_id or not part_code:
            return jsonify({'success': False, 'error': 'Gerekli parametreler eksik'}), 400
        
        try:
            quantity = int(quantity)
            if quantity <= 0:
                return jsonify({'success': False, 'error': 'Miktar 0\'dan büyük olmalı'}), 400
        except (ValueError, TypeError):
            return jsonify({'success': False, 'error': 'Miktar geçersiz'}), 400
        
        result = TakeuchiOrderManager.add_part_to_temp_order(
            session_id, 
            part_code, 
            quantity,
            user_id
        )
        
        if result.get('warning'):
            return jsonify(result), 409  # Conflict - aktif sipariş var
        
        return jsonify(result)
        
    except Exception as e:
        app.logger.error(f'[TAKEUCHI] Error adding part: {e}')
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/takeuchi/temp-order/<session_id>', methods=['GET'])
@login_required
def api_takeuchi_temp_order(session_id):
    """Geçici siparişin kalemleri al"""
    try:
        result = TakeuchiOrderManager.get_temp_order_items(session_id)
        return jsonify(result)
    except Exception as e:
        app.logger.error(f'[TAKEUCHI] Error getting temp order: {e}')
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/takeuchi/remove-item/<int:item_id>', methods=['DELETE'])
@login_required
def api_takeuchi_remove_item(item_id):
    """Geçici siparişten parça kaldır"""
    try:
        result = TakeuchiOrderManager.remove_temp_order_item(item_id)
        return jsonify(result)
    except Exception as e:
        app.logger.error(f'[TAKEUCHI] Error removing item: {e}')
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/takeuchi/orders', methods=['GET'])
@login_required
def api_takeuchi_orders():
    """Tüm Takeuchi siparişlerini listele"""
    try:
        result = TakeuchiOrderManager.get_all_orders()
        return jsonify(result)
    except Exception as e:
        app.logger.error(f'[TAKEUCHI] Error getting orders: {e}')
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/takeuchi/mark-received', methods=['POST'])
@login_required
def api_takeuchi_mark_received():
    """Sipariş kalemini teslim alındı olarak işaretle"""
    try:
        data = request.get_json()
        item_id = data.get('item_id')
        received_quantity = data.get('received_quantity')
        
        if item_id is None or received_quantity is None:
            return jsonify({'success': False, 'error': 'Gerekli parametreler eksik'}), 400
        
        try:
            received_quantity = int(received_quantity)
            if received_quantity < 0:
                return jsonify({'success': False, 'error': 'Miktar negatif olamaz'}), 400
        except (ValueError, TypeError):
            return jsonify({'success': False, 'error': 'Miktar geçersiz'}), 400
        
        result = TakeuchiOrderManager.mark_item_received(item_id, received_quantity)
        return jsonify(result)
        
    except Exception as e:
        app.logger.error(f'[TAKEUCHI] Error marking received: {e}')
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/takeuchi/admin/temp-orders', methods=['GET'])
@admin_required
def api_takeuchi_admin_temp_orders():
    """Admin: Tüm geçici siparişleri listele"""
    try:
        result = TakeuchiOrderManager.get_temp_orders_for_admin()
        return jsonify(result)
    except Exception as e:
        app.logger.error(f'[TAKEUCHI] Error getting temp orders: {e}')
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/takeuchi/admin/create-order', methods=['POST'])
@admin_required
def api_takeuchi_admin_create_order():
    """Admin: Geçici siparişi resmi sipariş haline dönüştür"""
    try:
        data = request.get_json()
        temp_order_id = data.get('temp_order_id')
        order_name = data.get('order_name', '')
        user_id = session.get('user_id')
        
        if not temp_order_id:
            return jsonify({'success': False, 'error': 'Geçici sipariş ID gerekli'}), 400
        
        try:
            temp_order_id = int(temp_order_id)
        except (ValueError, TypeError):
            return jsonify({'success': False, 'error': 'Geçersiz sipariş ID'}), 400
        
        result = TakeuchiOrderManager.create_official_order(
            temp_order_id,
            order_name,
            user_id
        )
        return jsonify(result)
        
    except Exception as e:
        app.logger.error(f'[TAKEUCHI] Error creating order: {e}')
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/takeuchi/admin/upload-parts', methods=['POST'])
@admin_required
def api_takeuchi_admin_upload_parts():
    """Admin: Excel dosyasından Takeuchi parçalarını yükle"""
    try:
        if 'file' not in request.files:
            return jsonify({'success': False, 'error': 'Dosya yüklenmedi'}), 400
        
        file = request.files['file']
        
        if file.filename == '':
            return jsonify({'success': False, 'error': 'Dosya seçilmedi'}), 400
        
        # Excel formatı kontrol et
        if not file.filename.lower().endswith(('.xlsx', '.xls')):
            return jsonify({'success': False, 'error': 'Yalnızca Excel dosyaları kabul edilir (.xlsx, .xls)'}), 400
        
        # Dosya içeriğini oku
        file_content = file.read()
        user_id = session.get('user_id')
        
        # Excel'i işle
        result = TakeuchiOrderManager.import_parts_from_excel(file_content, user_id)
        
        if result.get('success'):
            app.logger.info(f"[TAKEUCHI] {result['imported_count']} parça başarıyla yüklendi (User: {user_id})")
            return jsonify(result), 200
        else:
            return jsonify(result), 400
            
    except Exception as e:
        app.logger.error(f'[TAKEUCHI] Error uploading parts: {e}')
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/takeuchi/admin/parts-list', methods=['GET'])
@admin_required
def api_takeuchi_admin_parts_list():
    """Admin: Yüklenen tüm Takeuchi parçalarını listele"""
    try:
        result = TakeuchiOrderManager.get_all_takeuchi_parts()
        return jsonify(result)
        
    except Exception as e:
        app.logger.error(f'[TAKEUCHI] Error getting parts list: {e}')
        return jsonify({'success': False, 'error': str(e)}), 500


if __name__ == '__main__':
    # Paylaşımlı klasör erişimini kontrol et
    if not verify_shared_folder_access():
        print("\n[CRITICAL] Paylaşımlı klasöre erişim sağlanamamıştır!")
        print("Ubuntu'da NFS mount yapılmamış olabilir.")
        print("Lütfen kontrol edip uygulamayı yeniden başlatın.\n")
        exit(1)
    
    port = 5002
    print("\n" + "="*70)
    print(" CERMAK ENVANTER QR SİSTEMİ v2.0")
    print("="*70)
    print(f" Dashboard:      http://localhost:{port}")
    print(f" Admin Panel:    http://localhost:{port}/admin")
    print(f" Scanner:        http://localhost:{port}/scanner")
    print(f" Reports:        http://localhost:{port}/reports")
    print(f" Health Check:   http://localhost:{port}/health")
    print(f" Metrics:        http://localhost:{port}/metrics")
    print("-"*70)
    if platform.system() == 'Linux':
        print(" Storage:        Shared Folder (/mnt/ortakdepo)")
    else:
        print(" Storage:        Shared Folder (\\\\DCSRV\\tahsinortak\\CermakDepo)")
    print(" Database:       MySQL (192.168.0.57:3306)")
    print(" Security:       Headers + Rate Limiting Active")
    print(" Network:        WiFi/LAN Access (0.0.0.0)")
    print(" Live:           Socket.IO Aktif (Canlı Takip)")
    if PRINTER_ENABLED and _printer_manager:
        printer_status = _printer_manager.get_status()
        if printer_status.get('connected'):
            print(f" Printer:        USB TSPL Ready ({printer_status.get('device')})")
        else:
            print(f" Printer:        USB TSPL (Not Connected)")
    print("="*70 + "\n")
    
    # Network erişimi için host ayarı
    socketio.run(app, host='0.0.0.0', port=port, debug=False, allow_unsafe_werkzeug=True)



