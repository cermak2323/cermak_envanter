from flask import Flask, render_template, request, jsonify, send_file, session, redirect, url_for
from flask_socketio import SocketIO, emit
from werkzeug.utils import secure_filename
from functools import wraps, lru_cache
import time
from collections import defaultdict
import sqlite3
import pandas as pd
import qrcode
from PIL import Image, ImageDraw, ImageFont
from io import BytesIO
import base64
import os
import uuid
from datetime import datetime, timedelta
import zipfile
import re
import hashlib
import secrets
import random
import string
from dotenv import load_dotenv
from contextlib import contextmanager
import unicodedata
import logging
import threading
import json

# Load environment variables
load_dotenv()

# SQLAlchemy ve Models - VeritabanÃƒÂ„Ã‚Â± ORM
from models import db, PartCode, QRCode, CountSession, ScannedQR, User, CountPassword
from db_config import DevelopmentConfig, ProductionConfig

# Logging Configuration
from logging.handlers import RotatingFileHandler
import os

# Log klasÃƒÂƒÃ‚Â¶rÃƒÂƒÃ‚Â¼ oluÃƒÂ…Ã…Â¸tur
os.makedirs('logs', exist_ok=True)

# Loglama ayarlarÃƒÂ„Ã‚Â±
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
    handlers=[
        RotatingFileHandler('logs/app.log', maxBytes=10*1024*1024, backupCount=5),  # 10MB, 5 backup
        RotatingFileHandler('logs/security.log', maxBytes=5*1024*1024, backupCount=3),  # Security events
        logging.StreamHandler()
    ]
)

# Security logger
security_logger = logging.getLogger('security')
security_handler = RotatingFileHandler('logs/security.log', maxBytes=5*1024*1024, backupCount=3)
security_handler.setFormatter(logging.Formatter('%(asctime)s - SECURITY - %(levelname)s - %(message)s'))
security_logger.addHandler(security_handler)
security_logger.setLevel(logging.WARNING)

app = Flask(__name__)
app.config['SECRET_KEY'] = os.environ.get('SESSION_SECRET', 'dev-secret-key-change-in-production')
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024

#  LOKAL SÃƒÂ„Ã‚Â°STEM - Her zaman SQLite + Local Storage
IS_PRODUCTION = False
IS_LOCAL = True
USE_B2_STORAGE = False
USE_POSTGRESQL = False

print(f"\n LOKAL SISTEM")
print(f" Database: SQLite (instance/envanter_local.db)")
print(f"ÃƒÂ¯Ã‚Â¿Ã‚Â½ Storage: Local Files (static/qrcodes/)")
print(f" Data: Yerel dosya sisteminde")

# db_config.py kullan
app.config.from_object(DevelopmentConfig)

print("\n" + "="*60)
print(" LOKAL DEPOLAMA MODU")
print("="*60)
print(" TÃƒÂƒÃ‚Â¼m veriler yerel SQLite veritabanÃƒÂ„Ã‚Â±nda")
print(" QR kodlarÃƒÂ„Ã‚Â± static/qrcodes/ klasÃƒÂƒÃ‚Â¶rÃƒÂƒÃ‚Â¼nde")
print(" Excel dosyalarÃƒÂ„Ã‚Â± static/exports/ klasÃƒÂƒÃ‚Â¶rÃƒÂƒÃ‚Â¼nde")
print("="*60)
print()

# SQLAlchemy'yi app'e baÃƒÂ„Ã…Â¸la
db.init_app(app)

# Uygulama baÃƒÂ…Ã…Â¸langÃƒÂ„Ã‚Â±ÃƒÂƒÃ‚Â§ zamanÃƒÂ„Ã‚Â± (uptime iÃƒÂƒÃ‚Â§in)
app.config['START_TIME'] = time.time()

# Static dosya sÃƒÂ„Ã‚Â±kÃƒÂ„Ã‚Â±ÃƒÂ…Ã…Â¸tÃƒÂ„Ã‚Â±rma iÃƒÂƒÃ‚Â§in
app.config['SEND_FILE_MAX_AGE_DEFAULT'] = 31536000  # 1 yÃƒÂ„Ã‚Â±l cache

# SocketIO - Lokal sistem iÃƒÂƒÃ‚Â§in
socketio = SocketIO(
    app, 
    cors_allowed_origins="*",
    async_mode='eventlet',
    ping_timeout=60,
    ping_interval=25,
    max_http_buffer_size=1000000
)

# ======================
# MOBIL PERFORMANS OPTIMIZASYONLARI
# ======================

@app.before_request
def qr_folder_security():
    """QR klasÃƒÂƒÃ‚Â¶rÃƒÂƒÃ‚Â¼ gÃƒÂƒÃ‚Â¼venlik kontrolÃƒÂƒÃ‚Â¼"""
    #  QR KLASÃƒÂƒÃ¢Â€Â“RÃƒÂƒÃ…Â“ GÃƒÂƒÃ…Â“VENLÃƒÂ„Ã‚Â°ÃƒÂ„Ã‚ÂžÃƒÂ„Ã‚Â° - ÃƒÂ…Ã‚Âžifre ile korumalÃƒÂ„Ã‚Â± (admin paneli ÃƒÂ…Ã…Â¸ifresi ile aynÃƒÂ„Ã‚Â±)
    if request.path.startswith('/static/qrcodes/'):
        # Session'da qrcodes_unlocked anahtarÃƒÂ„Ã‚Â± yoksa veya False ise eriÃƒÂ…Ã…Â¸imi engelle
        if not session.get('qrcodes_unlocked', False):
            return '''
            <html>
            <head>
                <title>QR KlasÃƒÂƒÃ‚Â¶rÃƒÂƒÃ‚Â¼ - EriÃƒÂ…Ã…Â¸im Engellendi</title>
                <style>
                    body { font-family: Arial; text-align: center; padding: 50px; background: #f5f5f5; }
                    .container { background: white; padding: 30px; border-radius: 10px; max-width: 400px; margin: auto; box-shadow: 0 2px 10px rgba(0,0,0,0.1); }
                    h2 { color: #d32f2f; }
                    input { padding: 10px; width: 100%; margin: 10px 0; border: 1px solid #ddd; border-radius: 5px; }
                    button { padding: 10px 20px; background: #1976d2; color: white; border: none; border-radius: 5px; cursor: pointer; width: 100%; }
                    button:hover { background: #1565c0; }
                </style>
            </head>
            <body>
                <div class="container">
                    <h2> QR KlasÃƒÂƒÃ‚Â¶rÃƒÂƒÃ‚Â¼ KorumalÃƒÂ„Ã‚Â±</h2>
                    <p>Bu klasÃƒÂƒÃ‚Â¶re eriÃƒÂ…Ã…Â¸mek iÃƒÂƒÃ‚Â§in admin paneli ÃƒÂ…Ã…Â¸ifresini girin:</p>
                    <form method="POST" action="/unlock_qrcodes">
                        <input type="password" name="password" placeholder="Admin Paneli ÃƒÂ…Ã‚Âžifresi" required>
                        <button type="submit">Kilidi AÃƒÂƒÃ‚Â§</button>
                    </form>
                </div>
            </body>
            </html>
            ''', 403

@app.after_request
def add_performance_headers(response):
    """Performans iÃƒÂƒÃ‚Â§in header'lar ekle"""
    # Static dosyalar iÃƒÂƒÃ‚Â§in cache
    if request.endpoint == 'static':
        response.cache_control.max_age = 31536000  # 1 yÃƒÂ„Ã‚Â±l
        response.cache_control.public = True

    # DiÃƒÂ„Ã…Â¸er dosyalar iÃƒÂƒÃ‚Â§in
    else:
        response.cache_control.no_cache = True
        response.cache_control.must_revalidate = True

    # SÃƒÂ„Ã‚Â±kÃƒÂ„Ã‚Â±ÃƒÂ…Ã…Â¸tÃƒÂ„Ã‚Â±rma header'ÃƒÂ„Ã‚Â±
    if response.status_code == 200 and response.content_length and response.content_length > 1024:
        response.headers['Vary'] = 'Accept-Encoding'

    return response

# ======================
# PERFORMANS CACHE SISTEMI
# ======================

# Bellek tabanlÃƒÂ„Ã‚Â± cache (production'da Redis kullanÃƒÂ„Ã‚Â±lmalÃƒÂ„Ã‚Â±)
cache_store = {}
cache_lock = threading.Lock()
CACHE_TTL = 300  # 5 dakika cache sÃƒÂƒÃ‚Â¼resi

def cache_get(key):
    """Cache'den veri al"""
    with cache_lock:
        if key in cache_store:
            data, timestamp = cache_store[key]
            if time.time() - timestamp < CACHE_TTL:
                return data
            else:
                del cache_store[key]
    return None

def cache_set(key, value):
    """Cache'e veri kaydet"""
    with cache_lock:
        cache_store[key] = (value, time.time())

def cache_delete(key):
    """Cache'den veri sil"""
    with cache_lock:
        if key in cache_store:
            del cache_store[key]

def cache_clear():
    """TÃƒÂƒÃ‚Â¼m cache'i temizle"""
    with cache_lock:
        cache_store.clear()

# ÃƒÂ¢Ã…Â¡Ã‚Â¡ QR Image Memory Cache - LRU ile disk I/O azaltma
@lru_cache(maxsize=1000)
def generate_qr_pil_image(qr_id, box_size=8, border=2):
    """
    QR kod PIL Image oluÃƒÂ…Ã…Â¸tur - Barkod makinesi iÃƒÂƒÃ‚Â§in optimize
    
    Scanner Specs:
    - Minimum ÃƒÂƒÃ¢Â€Â¡ÃƒÂƒÃ‚Â¶zÃƒÂƒÃ‚Â¼nÃƒÂƒÃ‚Â¼rlÃƒÂƒÃ‚Â¼k 2D: ÃƒÂ¢Ã¢Â€Â°Ã‚Â¥8.7mil (0.22mm)
    - Okuma Mesafesi: 55-350mm
    - SensÃƒÂƒÃ‚Â¶r: 640x480 piksel
    
    QR BoyutlarÃƒÂ„Ã‚Â±:
    - box_size=8 -> ~240x240px QR (8.7mil ÃƒÂƒÃ‚Â§ÃƒÂƒÃ‚Â¶zÃƒÂƒÃ‚Â¼nÃƒÂƒÃ‚Â¼rlÃƒÂƒÃ‚Â¼k iÃƒÂƒÃ‚Â§in ideal)
    - border=2 -> Minimal ÃƒÂƒÃ‚Â§erÃƒÂƒÃ‚Â§eve (tarayÃƒÂ„Ã‚Â±cÃƒÂ„Ã‚Â± quiet zone iÃƒÂƒÃ‚Â§in)
    - Toplam: ~255x275px (ÃƒÂƒÃ‚Â§erÃƒÂƒÃ‚Â§eve + text ile)
    
    Args:
        qr_id: QR kod ID
        box_size: QR kutucuk boyutu (8 = barkod makinesi iÃƒÂƒÃ‚Â§in optimize)
        border: Quiet zone (2 = minimal, tarayÃƒÂ„Ã‚Â±cÃƒÂ„Ã‚Â± iÃƒÂƒÃ‚Â§in yeterli)
    
    Returns:
        PIL.Image: QR kod gÃƒÂƒÃ‚Â¶rseli (kÃƒÂ„Ã‚Â±rmÃƒÂ„Ã‚Â±zÃƒÂ„Ã‚Â± ÃƒÂƒÃ‚Â§erÃƒÂƒÃ‚Â§eveli)
    """
    qr = qrcode.QRCode(
        version=1,
        box_size=box_size,  # 8px - barkod makinesi iÃƒÂƒÃ‚Â§in optimize
        border=border,      # 2px quiet zone - tarayÃƒÂ„Ã‚Â±cÃƒÂ„Ã‚Â± iÃƒÂƒÃ‚Â§in minimum
        error_correction=qrcode.constants.ERROR_CORRECT_M  # M seviyesi (15% hata toleransÃƒÂ„Ã‚Â±)
    )
    qr.add_data(qr_id)
    qr.make(fit=True)
    qr_img = qr.make_image(fill_color="black", back_color="white")
    qr_img = qr_img.convert('RGB')
    
    # QR boyutlarÃƒÂ„Ã‚Â±
    qr_width, qr_height = qr_img.size
    text_height = 20  # Text alanÃƒÂ„Ã‚Â± (14pt font iÃƒÂƒÃ‚Â§in)
    border_width = 3  # KÃƒÂ„Ã‚Â±rmÃƒÂ„Ã‚Â±zÃƒÂ„Ã‚Â± ÃƒÂƒÃ‚Â§erÃƒÂƒÃ‚Â§eve
    
    # KÃƒÂ„Ã‚Â±rmÃƒÂ„Ã‚Â±zÃƒÂ„Ã‚Â± ÃƒÂƒÃ‚Â§erÃƒÂƒÃ‚Â§eveli gÃƒÂƒÃ‚Â¶rsel
    final_width = qr_width + (border_width * 2)
    final_height = qr_height + text_height + (border_width * 2)
    final_img = Image.new('RGB', (final_width, final_height), '#dc2626')
    
    # Beyaz iÃƒÂƒÃ‚Â§ alan
    white_bg = Image.new('RGB', (qr_width, qr_height + text_height), 'white')
    white_bg.paste(qr_img, (0, 0))
    final_img.paste(white_bg, (border_width, border_width))
    
    # Text ekle
    draw = ImageDraw.Draw(final_img)
    try:
        font = ImageFont.truetype("arialbd.ttf", 14)
    except:
        try:
            font = ImageFont.truetype("arial.ttf", 14)
        except:
            font = ImageFont.load_default()
    
    qr_text = qr_id
    text_width = len(qr_text) * 8
    x_position = max(border_width, (final_width - text_width) // 2)
    draw.text((x_position, qr_height + border_width + 2), qr_text, fill='black', font=font)
    
    return final_img

# Cache temizleme thread'i
def cache_cleanup():
    """Eski cache verilerini temizle"""
    while True:
        time.sleep(60)  # Her dakika kontrol et
        current_time = time.time()
        with cache_lock:
            expired_keys = [
                key for key, (_, timestamp) in cache_store.items()
                if current_time - timestamp > CACHE_TTL
            ]
            for key in expired_keys:
                del cache_store[key]

# Cache temizleme thread'ini baÃƒÂ…Ã…Â¸lat
cleanup_thread = threading.Thread(target=cache_cleanup, daemon=True)
cleanup_thread.start()

# Rate limiting iÃƒÂƒÃ‚Â§in IP tabanlÃƒÂ„Ã‚Â± takip
login_attempts = defaultdict(list)

def add_security_headers(response):
    """GÃƒÂƒÃ‚Â¼venlik header'larÃƒÂ„Ã‚Â±nÃƒÂ„Ã‚Â± ekle"""
    response.headers['X-Content-Type-Options'] = 'nosniff'
    response.headers['X-Frame-Options'] = 'DENY'
    response.headers['X-XSS-Protection'] = '1; mode=block'
    response.headers['Strict-Transport-Security'] = 'max-age=31536000; includeSubDomains'
    response.headers['Content-Security-Policy'] = "default-src 'self'; script-src 'self' 'unsafe-inline' https://cdn.jsdelivr.net https://cdn.socket.io https://unpkg.com; style-src 'self' 'unsafe-inline' https://cdn.jsdelivr.net https://fonts.googleapis.com; font-src 'self' https://cdn.jsdelivr.net https://fonts.gstatic.com; img-src 'self' data:;"
    return response

@app.after_request
def security_headers(response):
    return add_security_headers(response)

def rate_limit_login(f):
    """Login denemelerini sÃƒÂ„Ã‚Â±nÃƒÂ„Ã‚Â±rla"""
    @wraps(f)
    def decorated_function(*args, **kwargs):
        client_ip = request.environ.get('HTTP_X_FORWARDED_FOR', request.environ.get('REMOTE_ADDR', '127.0.0.1'))
        current_time = time.time()

        # Son 15 dakikadaki denemeleri filtrele
        login_attempts[client_ip] = [t for t in login_attempts[client_ip] if current_time - t < 900]

        # 15 dakikada 5'ten fazla deneme varsa engelle
        if len(login_attempts[client_ip]) >= 5:
            return jsonify({'error': 'ÃƒÂƒÃ¢Â€Â¡ok fazla login denemesi. 15 dakika bekleyin.'}), 429

        # Denemeyi kaydet
        login_attempts[client_ip].append(current_time)

        return f(*args, **kwargs)
    return decorated_function

# ----------------------
# Configuration
# ----------------------

# Project root
BASE_DIR = os.path.abspath(os.path.dirname(__file__))

# ======================
# VERÃƒÂ„Ã‚Â°TABANI YAPILANDI - SADECE SQLite
# ======================

# Lokal SQLite database
DATABASE_URL = app.config['SQLALCHEMY_DATABASE_URI']
print(f" Local SQLite: {DATABASE_URL}")

REPORTS_DIR = 'reports'

def generate_strong_password():
    """GÃƒÂƒÃ‚Â¼ÃƒÂƒÃ‚Â§lÃƒÂƒÃ‚Â¼ parola oluÃƒÂ…Ã…Â¸tur (8 karakter: bÃƒÂƒÃ‚Â¼yÃƒÂƒÃ‚Â¼k harf, kÃƒÂƒÃ‚Â¼ÃƒÂƒÃ‚Â§ÃƒÂƒÃ‚Â¼k harf, rakam, ÃƒÂƒÃ‚Â¶zel karakter)"""
    characters = string.ascii_uppercase + string.ascii_lowercase + string.digits + "!@#$%^&*"
    # En az 1 bÃƒÂƒÃ‚Â¼yÃƒÂƒÃ‚Â¼k harf, 1 kÃƒÂƒÃ‚Â¼ÃƒÂƒÃ‚Â§ÃƒÂƒÃ‚Â¼k harf, 1 rakam, 1 ÃƒÂƒÃ‚Â¶zel karakter olacak ÃƒÂ…Ã…Â¸ekilde
    password = [
        random.choice(string.ascii_uppercase),  # BÃƒÂƒÃ‚Â¼yÃƒÂƒÃ‚Â¼k harf
        random.choice(string.ascii_lowercase),  # KÃƒÂƒÃ‚Â¼ÃƒÂƒÃ‚Â§ÃƒÂƒÃ‚Â¼k harf
        random.choice(string.digits),           # Rakam
        random.choice("!@#$%^&*")              # ÃƒÂƒÃ¢Â€Â“zel karakter
    ]
    # Kalan 4 karakteri rastgele seÃƒÂƒÃ‚Â§
    for _ in range(4):
        password.append(random.choice(characters))

    # KarÃƒÂ„Ã‚Â±ÃƒÂ…Ã…Â¸tÃƒÂ„Ã‚Â±r
    random.shuffle(password)
    return ''.join(password)

def generate_count_password():
    """SayÃƒÂ„Ã‚Â±m iÃƒÂƒÃ‚Â§in parola oluÃƒÂ…Ã…Â¸tur (6 haneli sadece sayÃƒÂ„Ã‚Â±) - Basit ve hÃƒÂ„Ã‚Â±zlÃƒÂ„Ã‚Â± giriÃƒÂ…Ã…Â¸ iÃƒÂƒÃ‚Â§in"""
    return ''.join([str(random.randint(0, 9)) for _ in range(6)])

def save_qr_code_to_file(part_code, qr_id, qr_number):
    """
    QR kodunu fiziksel dosya olarak kaydet
    - Her parÃƒÂƒÃ‚Â§a iÃƒÂƒÃ‚Â§in ayrÃƒÂ„Ã‚Â± klasÃƒÂƒÃ‚Â¶r oluÃƒÂ…Ã…Â¸turur (ÃƒÂƒÃ‚Â¶rn: static/qr_codes/Y129513-14532/)
    - QR kodlarÃƒÂ„Ã‚Â±nÃƒÂ„Ã‚Â± numaralÃƒÂ„Ã‚Â± kaydeder (ÃƒÂƒÃ‚Â¶rn: Y129513-14532_1.png)
    - QR kod altÃƒÂ„Ã‚Â±na kod ve parÃƒÂƒÃ‚Â§a adÃƒÂ„Ã‚Â± yazar
    """
    try:
        import qrcode
        
        # ParÃƒÂƒÃ‚Â§a adÃƒÂ„Ã‚Â±nÃƒÂ„Ã‚Â± database'den al
        conn = get_db()
        cursor = conn.cursor()
        execute_query(cursor, 'SELECT part_name FROM part_codes WHERE part_code = ?', (part_code,))
        result = cursor.fetchone()
        part_name = result[0] if result else ""
        close_db(conn)
        
        # QR klasÃƒÂƒÃ‚Â¶rÃƒÂƒÃ‚Â¼ ana dizini
        qrcodes_base_dir = os.path.join(os.path.dirname(__file__), 'static', 'qr_codes')
        
        # ParÃƒÂƒÃ‚Â§a iÃƒÂƒÃ‚Â§in klasÃƒÂƒÃ‚Â¶r oluÃƒÂ…Ã…Â¸tur (ÃƒÂƒÃ‚Â¶rn: static/qr_codes/Y129513-14532/)
        part_dir = os.path.join(qrcodes_base_dir, part_code)
        os.makedirs(part_dir, exist_ok=True)
        
        # Dosya adÃƒÂ„Ã‚Â±: part_code_number.png (ÃƒÂƒÃ‚Â¶rn: Y129513-14532_1.png)
        filename = f"{part_code}_{qr_number}.png"
        filepath = os.path.join(part_dir, filename)
        
        # QR kod oluÃƒÂ…Ã…Â¸tur - Barkod makinesi iÃƒÂƒÃ‚Â§in optimize (8.7mil ÃƒÂƒÃ‚Â§ÃƒÂƒÃ‚Â¶zÃƒÂƒÃ‚Â¼nÃƒÂƒÃ‚Â¼rlÃƒÂƒÃ‚Â¼k)
        qr = qrcode.QRCode(
            version=1,
            error_correction=qrcode.constants.ERROR_CORRECT_M,  # M seviyesi (15% hata toleransÃƒÂ„Ã‚Â±)
            box_size=8,  # 8px - barkod makinesi iÃƒÂƒÃ‚Â§in ideal (~240x240px)
            border=2,    # 2px quiet zone - tarayÃƒÂ„Ã‚Â±cÃƒÂ„Ã‚Â± minimum gereksinimi
        )
        qr.add_data(qr_id)
        qr.make(fit=True)
        
        # QR kodunu oluÃƒÂ…Ã…Â¸tur
        qr_img = qr.make_image(fill_color="black", back_color="white")
        qr_img = qr_img.convert('RGB')  # PIL Image'a dÃƒÂƒÃ‚Â¶nÃƒÂƒÃ‚Â¼ÃƒÂ…Ã…Â¸tÃƒÂƒÃ‚Â¼r
        
        # QR kod boyutlarÃƒÂ„Ã‚Â±nÃƒÂ„Ã‚Â± al
        qr_width, qr_height = qr_img.size
        
        # AlanlarÃƒÂ„Ã‚Â± hesapla
        logo_height = 40  # Logo iÃƒÂƒÃ‚Â§in ÃƒÂƒÃ‚Â¼st alan
        text_height = 35  # Alt yazÃƒÂ„Ã‚Â± (parÃƒÂƒÃ‚Â§a numarasÃƒÂ„Ã‚Â±) iÃƒÂƒÃ‚Â§in alan
        
        # KÃƒÂ„Ã‚Â±rmÃƒÂ„Ã‚Â±zÃƒÂ„Ã‚Â± ÃƒÂƒÃ‚Â§erÃƒÂƒÃ‚Â§eve iÃƒÂƒÃ‚Â§in padding
        border_width = 3  # 3px kÃƒÂ„Ã‚Â±rmÃƒÂ„Ã‚Â±zÃƒÂ„Ã‚Â± ÃƒÂƒÃ‚Â§erÃƒÂƒÃ‚Â§eve
        
        # Yeni gÃƒÂƒÃ‚Â¶rsel oluÃƒÂ…Ã…Â¸tur (logo + QR + text alanÃƒÂ„Ã‚Â± + ÃƒÂƒÃ‚Â§erÃƒÂƒÃ‚Â§eve)
        final_width = qr_width + (border_width * 2)
        final_height = logo_height + qr_height + text_height + (border_width * 2)
        final_img = Image.new('RGB', (final_width, final_height), '#dc2626')  # KÃƒÂ„Ã‚Â±rmÃƒÂ„Ã‚Â±zÃƒÂ„Ã‚Â± arka plan (ÃƒÂƒÃ‚Â§erÃƒÂƒÃ‚Â§eve)
        
        # Beyaz iÃƒÂƒÃ‚Â§ alan oluÃƒÂ…Ã…Â¸tur (logo + QR + text)
        white_bg = Image.new('RGB', (qr_width, logo_height + qr_height + text_height), 'white')
        
        # Logo ekle (varsa) - ÃƒÂƒÃ‚Â¼st ortasÃƒÂ„Ã‚Â±na
        try:
            logo_path = os.path.join(os.path.dirname(__file__), 'cermak-logo.png')
            if os.path.exists(logo_path):
                logo_img = Image.open(logo_path).convert('RGBA')
                # Logo boyutunu alan yÃƒÂƒÃ‚Â¼ksekliÃƒÂ„Ã…Â¸ine gÃƒÂƒÃ‚Â¶re ayarla
                logo_width = 150
                logo_height_logo = 40
                try:
                    logo_img = logo_img.resize((logo_width, logo_height_logo), Image.Resampling.LANCZOS)
                except AttributeError:
                    logo_img = logo_img.resize((logo_width, logo_height_logo), Image.LANCZOS)
                
                # Logo'yu ortala
                logo_x = (qr_width - logo_width) // 2
                logo_y = 5  # ÃƒÂƒÃ…Â“stten 5px boÃƒÂ…Ã…Â¸luk
                
                # RGBA logo'yu blend et
                if logo_img.mode == 'RGBA':
                    alpha = logo_img.split()[3]
                    logo_img = logo_img.convert('RGB')
                    white_bg.paste(logo_img, (logo_x, logo_y), alpha)
                else:
                    white_bg.paste(logo_img, (logo_x, logo_y))
        except Exception as e:
            print(f"Logo ekleme hatasÃƒÂ„Ã‚Â±: {e}")
        
        # QR kodu beyaz alana yapÃƒÂ„Ã‚Â±ÃƒÂ…Ã…Â¸tÃƒÂ„Ã‚Â±r (logo'nun altÃƒÂ„Ã‚Â±na)
        white_bg.paste(qr_img, (0, logo_height))
        
        # Beyaz alanÃƒÂ„Ã‚Â± kÃƒÂ„Ã‚Â±rmÃƒÂ„Ã‚Â±zÃƒÂ„Ã‚Â± ÃƒÂƒÃ‚Â§erÃƒÂƒÃ‚Â§evenin iÃƒÂƒÃ‚Â§ine yapÃƒÂ„Ã‚Â±ÃƒÂ…Ã…Â¸tÃƒÂ„Ã‚Â±r
        final_img.paste(white_bg, (border_width, border_width))
        
        # Text ekleme iÃƒÂƒÃ‚Â§in draw nesnesi
        draw = ImageDraw.Draw(final_img)
        
        # Font (kalÃƒÂ„Ã‚Â±n ve bÃƒÂƒÃ‚Â¼yÃƒÂƒÃ‚Â¼k)
        try:
            # Windows iÃƒÂƒÃ‚Â§in Arial Black (en kalÃƒÂ„Ã‚Â±n), en bÃƒÂƒÃ‚Â¼yÃƒÂƒÃ‚Â¼k font
            font = ImageFont.truetype("arialblk.ttf", 24)
        except:
            try:
                font = ImageFont.truetype("arialbd.ttf", 24)
            except:
                try:
                    font = ImageFont.truetype("arial.ttf", 24)
                except:
                    font = ImageFont.load_default()
        
        # QR ID yazÃƒÂ„Ã‚Â±sÃƒÂ„Ã‚Â± (Y129150-49811_1) - Sadece bu
        qr_text = f"{part_code}_{qr_number}"
        
        # Text geniÃƒÂ…Ã…Â¸liÃƒÂ„Ã…Â¸ini hesapla (24pt font iÃƒÂƒÃ‚Â§in)
        text_width = len(qr_text) * 14  # 24pt font iÃƒÂƒÃ‚Â§in geniÃƒÂ…Ã…Â¸lik
        
        # QR ID'yi ortala ve yaz (border_width offset ekle)
        x_position = max(border_width, (final_width - text_width) // 2)
        draw.text((x_position, logo_height + qr_height + border_width + 5), qr_text, fill='black', font=font)
        
        # Kaydet
        final_img.save(filepath)
        
        print(f"ÃƒÂ¢Ã…Â“Ã¢Â€Â¦ QR kod kaydedildi: static/qr_codes/{part_code}/{filename}")
        return filepath
        
    except Exception as e:
        print(f"ÃƒÂ¢Ã‚Â�Ã…Â’ QR kod kaydetme hatasÃƒÂ„Ã‚Â±: {e}")
        return None

# Admin sayÃƒÂ„Ã‚Â±m ÃƒÂ…Ã…Â¸ifresi
ADMIN_COUNT_PASSWORD = "@R9t$L7e!xP2w"
print(f"DEBUG: ADMIN_COUNT_PASSWORD = '{ADMIN_COUNT_PASSWORD}'")  # DEBUG

os.makedirs(REPORTS_DIR, exist_ok=True)

# Dosya upload iÃƒÂƒÃ‚Â§in ayarlar
UPLOAD_FOLDER = 'static/part_photos'
ALLOWED_EXTENSIONS = {'png', 'jpg', 'jpeg', 'gif', 'webp'}
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024  # 16MB limit

def allowed_file(filename):
    """Dosya uzantÃƒÂ„Ã‚Â±sÃƒÂ„Ã‚Â±nÃƒÂ„Ã‚Â± kontrol et"""
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

def get_db_placeholder():
    """Database'e gÃƒÂƒÃ‚Â¶re doÃƒÂ„Ã…Â¸ru placeholder dÃƒÂƒÃ‚Â¶ndÃƒÂƒÃ‚Â¼r - SQLite iÃƒÂƒÃ‚Â§in ? kullanÃƒÂ„Ã‚Â±r"""
    return '?'

def execute_query(cursor, query, params=None):
    """
    Execute SQL query - SQLite iÃƒÂƒÃ‚Â§in ? placeholder kullanÃƒÂ„Ã‚Â±r
    """
    # SQLite: ? -> ? dÃƒÂƒÃ‚Â¶nÃƒÂƒÃ‚Â¼ÃƒÂ…Ã…Â¸tÃƒÂƒÃ‚Â¼r (eski kodlardan gelebilecek ?'leri deÃƒÂ„Ã…Â¸iÃƒÂ…Ã…Â¸tir)
    query = query.replace('?', '?')

    if params:
        cursor.execute(query, params)
    else:
        cursor.execute(query)

    return cursor

def get_db():
    """Get database connection - SQLite only (local system)"""
    #  LOKAL: SQLite baÃƒÂ„Ã…Â¸lantÃƒÂ„Ã‚Â±sÃƒÂ„Ã‚Â±
    db_path = os.path.join('instance', 'envanter_local.db')
    os.makedirs('instance', exist_ok=True)
    conn = sqlite3.connect(db_path)
    conn.row_factory = sqlite3.Row
    return conn

def close_db(conn):
    """Close SQLite connection"""
    if conn:
        try:
            conn.close()
        except Exception as e:
            logging.debug(f"Error closing SQLite connection: {e}")

@contextmanager
def db_connection():
    """
    Database connection context manager - Otomatik cleanup
    
    KullanÃƒÂ„Ã‚Â±m:
        with db_connection() as conn:
            cursor = conn.cursor()
            execute_query(cursor, 'SELECT ...')
            conn.commit()
        # conn otomatik kapanÃƒÂ„Ã‚Â±r, hata olsa bile
    """
    conn = get_db()
    try:
        yield conn
    except sqlite3.Error as e:
        logging.error(f"Database error: {e}")
        if conn:
            conn.rollback()
        raise
    except Exception as e:
        logging.error(f"Unexpected error in database operation: {e}")
        if conn:
            conn.rollback()
        raise
    finally:
        close_db(conn)

def create_performance_indexes(cursor):
    """Performans iÃƒÂƒÃ‚Â§in kritik indexleri oluÃƒÂ…Ã…Â¸tur"""
    indexes = [
        # QR Codes indexleri
        "CREATE INDEX IF NOT EXISTS idx_qr_codes_qr_id ON qr_codes(qr_id);",
        "CREATE INDEX IF NOT EXISTS idx_qr_codes_part_code ON qr_codes(part_code_id);",
        "CREATE INDEX IF NOT EXISTS idx_qr_codes_is_used ON qr_codes(is_used);",

        # Count Sessions indexleri
        "CREATE INDEX IF NOT EXISTS idx_count_sessions_is_active ON count_sessions(is_active);",
        "CREATE INDEX IF NOT EXISTS idx_count_sessions_created_by ON count_sessions(created_by);",
        "CREATE INDEX IF NOT EXISTS idx_count_sessions_created_at ON count_sessions(created_at);",

        # Scanned QR indexleri - COMPOSITE INDEX ÃƒÂ„Ã‚Â°ÃƒÂƒÃ¢Â€Â¡ÃƒÂ„Ã‚Â°N ÃƒÂƒÃ¢Â€Â“NEMLÃƒÂ„Ã‚Â°
        "CREATE INDEX IF NOT EXISTS idx_scanned_qr_session_id ON scanned_qr(session_id);",
        "CREATE INDEX IF NOT EXISTS idx_scanned_qr_qr_id ON scanned_qr(qr_id);",
        "CREATE INDEX IF NOT EXISTS idx_scanned_qr_scanned_by ON scanned_qr(scanned_by);",
        "CREATE INDEX IF NOT EXISTS idx_scanned_qr_scanned_at ON scanned_qr(scanned_at);",
        
        # ÃƒÂ¢Ã…Â¡Ã‚Â¡ COMPOSITE INDEX - Duplicate check iÃƒÂƒÃ‚Â§in kritik
        "CREATE INDEX IF NOT EXISTS idx_scanned_qr_session_qr ON scanned_qr(session_id, qr_id);",

        # Part Codes indexleri
        "CREATE INDEX IF NOT EXISTS idx_part_codes_part_code ON part_codes(part_code);",

        # KullanÃƒÂ„Ã‚Â±cÃƒÂ„Ã‚Â± indexleri
        "CREATE INDEX IF NOT EXISTS idx_envanter_users_username ON envanter_users(username);",
        "CREATE INDEX IF NOT EXISTS idx_envanter_users_is_active ON envanter_users(is_active_user);",
    ]

    try:
        for index_sql in indexes:
            execute_query(cursor, index_sql)
        print("ÃƒÂ¢Ã…Â“Ã¢Â€Â¦ Performance indexes created/verified")
    except Exception as e:
        print(f"ÃƒÂ¢Ã…Â¡Ã‚Â ÃƒÂ¯Ã‚Â¸Ã‚Â� Warning: Could not create some indexes: {e}")

def init_db_part_details():
    """Part codes tablosuna detay kolonlarÃƒÂ„Ã‚Â± ekle"""
    try:
        conn = get_db()
        cursor = conn.cursor()
        
        # Yeni kolonlar - Geriye uyumlu (ALTER TABLE)
        new_columns = [
            ('photo_path', 'TEXT'),           # ParÃƒÂƒÃ‚Â§a fotoÃƒÂ„Ã…Â¸rafÃƒÂ„Ã‚Â±
            ('catalog_image', 'TEXT'),        # Katalog gÃƒÂƒÃ‚Â¶rseli
            ('description', 'TEXT'),          # AÃƒÂƒÃ‚Â§ÃƒÂ„Ã‚Â±klama
            ('used_in_machines', 'TEXT'),     # JSON: KullanÃƒÂ„Ã‚Â±ldÃƒÂ„Ã‚Â±ÃƒÂ„Ã…Â¸ÃƒÂ„Ã‚Â± makineler
            ('specifications', 'TEXT'),       # JSON: Teknik ÃƒÂƒÃ‚Â¶zellikler
            ('stock_location', 'TEXT'),       # Stok konumu
            ('supplier', 'TEXT'),             # TedarikÃƒÂƒÃ‚Â§i
            ('unit_price', 'REAL'),          # Birim fiyat
            ('notes', 'TEXT')                # Notlar
        ]
        
        for col_name, col_type in new_columns:
            try:
                cursor.execute(f'ALTER TABLE part_codes ADD COLUMN {col_name} {col_type}')
                print(f"ÃƒÂ¢Ã…Â“Ã¢Â€Â¦ Kolon eklendi: {col_name}")
            except sqlite3.OperationalError as e:
                if 'duplicate column name' in str(e).lower():
                    print(f"ÃƒÂ¢Ã…Â¡Ã‚Â ÃƒÂ¯Ã‚Â¸Ã‚Â� Kolon zaten var: {col_name}")
                else:
                    raise
        
        conn.commit()
        close_db(conn)
        print("ÃƒÂ¢Ã…Â“Ã¢Â€Â¦ Part details schema gÃƒÂƒÃ‚Â¼ncellendi")
        
    except Exception as e:
        print(f"ÃƒÂ¢Ã‚Â�Ã…Â’ Part details schema hatasÃƒÂ„Ã‚Â±: {e}")
        try:
            if 'conn' in locals():
                close_db(conn)
        except:
            pass

def init_db():
    """Initialize database tables using SQLAlchemy ORM

    Lokal: SQLite tables oluÃƒÂ…Ã…Â¸tur
    Production: PostgreSQL tables kontrol et
    """
    try:
        with app.app_context():
            if USE_POSTGRESQL:
                # PRODUCTION: PostgreSQL - SQLAlchemy ile
                print("\n" + "="*70)
                print(" INITIALIZING POSTGRESQL TABLES")
                print("="*70)

                inspector = db.inspect(db.engine)
                existing_tables = inspector.get_table_names()
                print(f"Existing tables: {existing_tables}")

                required_tables = ['envanter_users', 'part_codes', 'qr_codes', 'count_sessions', 'count_passwords', 'scanned_qr']

                missing_tables = [t for t in required_tables if t not in existing_tables]

                if missing_tables:
                    print(f"ÃƒÂ¢Ã‹ÂœÃ‚Â�ÃƒÂ¯Ã‚Â¸Ã‚Â� Creating missing PostgreSQL tables: {', '.join(missing_tables)}")
                    db.create_all()
                    print("ÃƒÂ¢Ã…Â“Ã¢Â€Â¦ PostgreSQL tables created successfully")
                else:
                    print("ÃƒÂ¢Ã…Â“Ã¢Â€Â¦ All PostgreSQL tables already exist")

                #  DATABASE MIGRATION: Add finished_by column to count_sessions
                try:
                    conn = get_db()
                    cursor = conn.cursor()

                    # Check if finished_by column exists
                    execute_query(cursor, """
                        SELECT column_name 
                        FROM information_schema.columns 
                        WHERE table_name='count_sessions' AND column_name='finished_by'
                    """)

                    if not cursor.fetchone():
                        print(" Adding finished_by column to count_sessions table...")
                        execute_query(cursor, """
                            ALTER TABLE count_sessions 
                            ADD COLUMN finished_by INTEGER REFERENCES envanter_users(id)
                        """)
                        conn.commit()
                        print("ÃƒÂ¢Ã…Â“Ã¢Â€Â¦ finished_by column added successfully")
                    else:
                        print("ÃƒÂ¢Ã…Â“Ã¢Â€Â¦ finished_by column already exists")

                    close_db(conn)
                except Exception as e:
                    print(f"ÃƒÂ¢Ã…Â¡Ã‚Â ÃƒÂ¯Ã‚Â¸Ã‚Â� Migration warning: {e}")

                #  DATABASE MIGRATION: Add total_expected and total_scanned columns to count_sessions
                try:
                    conn = get_db()
                    cursor = conn.cursor()

                    # Check total_expected
                    execute_query(cursor, """
                        SELECT column_name FROM information_schema.columns
                        WHERE table_name='count_sessions' AND column_name='total_expected'
                    """)
                    if not cursor.fetchone():
                        print(" Adding total_expected column to count_sessions table...")
                        execute_query(cursor, """
                            ALTER TABLE count_sessions
                            ADD COLUMN total_expected INTEGER DEFAULT 0
                        """)
                        conn.commit()
                        print("ÃƒÂ¢Ã…Â“Ã¢Â€Â¦ total_expected column added successfully")
                    else:
                        print("ÃƒÂ¢Ã…Â“Ã¢Â€Â¦ total_expected column already exists")

                    # Check total_scanned
                    execute_query(cursor, """
                        SELECT column_name FROM information_schema.columns
                        WHERE table_name='count_sessions' AND column_name='total_scanned'
                    """)
                    if not cursor.fetchone():
                        print(" Adding total_scanned column to count_sessions table...")
                        execute_query(cursor, """
                            ALTER TABLE count_sessions
                            ADD COLUMN total_scanned INTEGER DEFAULT 0
                        """)
                        conn.commit()
                        print("ÃƒÂ¢Ã…Â“Ã¢Â€Â¦ total_scanned column added successfully")
                    else:
                        print("ÃƒÂ¢Ã…Â“Ã¢Â€Â¦ total_scanned column already exists")

                    #  DATABASE MIGRATION: Add report_file_path column to count_sessions
                    cursor = conn.cursor()
                    cursor.execute("""
                        SELECT COUNT(*) FROM pragma_table_info('count_sessions') 
                        WHERE name='report_file_path'
                    """)
                    if cursor.fetchone()[0] == 0:
                        print(" Adding report_file_path column to count_sessions table...")
                        cursor.execute("""
                            ALTER TABLE count_sessions
                            ADD COLUMN report_file_path TEXT
                        """)
                        conn.commit()
                        print("ÃƒÂ¢Ã…Â“Ã¢Â€Â¦ report_file_path column added successfully")
                    else:
                        print("ÃƒÂ¢Ã…Â“Ã¢Â€Â¦ report_file_path column already exists")

                    close_db(conn)
                except Exception as e:
                    print(f"ÃƒÂ¢Ã…Â¡Ã‚Â ÃƒÂ¯Ã‚Â¸Ã‚Â� Migration warning (total_*): {e}")

                # Verify scanned_qr table specifically (critical for duplicate detection)
                if 'scanned_qr' in existing_tables or 'scanned_qr' not in missing_tables:
                    print("ÃƒÂ¢Ã…Â“Ã¢Â€Â¦ scanned_qr table verified - duplicate detection will work")
                else:
                    print("ÃƒÂ¢Ã‚Â�Ã…Â’ WARNING: scanned_qr table not found - duplicate detection may fail!")
                    print("   Creating scanned_qr table manually...")
                    try:
                        # Create scanned_qr table manually if SQLAlchemy fails
                        conn = get_db()
                        cursor = conn.cursor()
                        execute_query(cursor, '''
                            CREATE TABLE IF NOT EXISTS scanned_qr (
                                id SERIAL PRIMARY KEY,
                                session_id VARCHAR(255) NOT NULL,
                                qr_id VARCHAR(255) NOT NULL,
                                part_code VARCHAR(255) NOT NULL,
                                scanned_by INTEGER,
                                scanned_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
                            )
                        ''')
                        conn.commit()
                        close_db(conn)
                        print("ÃƒÂ¢Ã…Â“Ã¢Â€Â¦ scanned_qr table created successfully")
                    except Exception as e:
                        print(f"ÃƒÂ¢Ã‚Â�Ã…Â’ Failed to create scanned_qr: {e}")

                # SQLAlchemy User model kullan
                admin_user = User.query.filter_by(username='admin').first()
                if not admin_user:
                    print("ÃƒÂ¢Ã‚ÂžÃ¢Â€Â¢ Creating default PostgreSQL admin user...")
                    from werkzeug.security import generate_password_hash
                    admin_password = generate_password_hash("@R9t$L7e!xP2w")
                    admin = User(
                        username='admin',
                        full_name='Administrator',
                        password_hash=admin_password,
                        role='admin',
                        is_active_user=True
                    )
                    db.session.add(admin)
                    db.session.commit()
                    print("ÃƒÂ¢Ã…Â“Ã¢Â€Â¦ PostgreSQL admin user created (admin/@R9t$L7e!xP2w)")
                else:
                    print("ÃƒÂ¢Ã…Â“Ã¢Â€Â¦ PostgreSQL admin user already exists")

                print("="*70 + "\n")
                # LOCAL: SQLite - Raw SQL ile (basit tablo yapÃƒÂ„Ã‚Â±sÃƒÂ„Ã‚Â±)
                print(" Local SQLite mode - checking simple table structure")

                # SQLite baÃƒÂ„Ã…Â¸lantÃƒÂ„Ã‚Â±sÃƒÂ„Ã‚Â± al
                conn = get_db()
                cursor = conn.cursor()

                # SQLite schema'yÃƒÂ„Ã‚Â± gÃƒÂƒÃ‚Â¼ncelle (full_name column ekle)
                try:
                    execute_query(cursor, "ALTER TABLE envanter_users ADD COLUMN full_name VARCHAR(255)")
                    print("ÃƒÂ¢Ã…Â“Ã¢Â€Â¦ Added full_name column to SQLite")
                except sqlite3.OperationalError:
                    # Column zaten varsa veya baÃƒÂ…Ã…Â¸ka hata
                    pass

                # Admin user var mÃƒÂ„Ã‚Â± kontrol et (SQLite raw SQL)
                execute_query(cursor, "SELECT * FROM envanter_users WHERE username = 'admin'")
                admin_exists = cursor.fetchone()

                if not admin_exists:
                    print("ÃƒÂ¢Ã‚ÂžÃ¢Â€Â¢ Creating default SQLite admin user...")
                    from werkzeug.security import generate_password_hash
                    admin_password_hash = generate_password_hash("admin123")
                    execute_query(cursor, '''
                        INSERT INTO envanter_users (username, password_hash, full_name, role, created_at, is_active)
                        VALUES (?, ?, ?, ?, ?, ?)
                    ''', ('admin', admin_password_hash, 'Administrator', 'admin', datetime.now(), True))
                    conn.commit()
                    print("ÃƒÂ¢Ã…Â“Ã¢Â€Â¦ SQLite admin user created (admin/admin123)")
                else:
                    print("ÃƒÂ¢Ã…Â“Ã¢Â€Â¦ SQLite admin user already exists")

                close_db(conn)
                print("ÃƒÂ¢Ã…Â“Ã¢Â€Â¦ SQLite database initialized successfully")
            
            # Part details kolonlarÃƒÂ„Ã‚Â±nÃƒÂ„Ã‚Â± ekle
            init_db_part_details()

            return True

    except Exception as e:
        print(f"ÃƒÂ¢Ã‚Â�Ã…Â’ Database initialization error: {str(e)}")
        import traceback
        traceback.print_exc()
        return False

def init_db_part_details():
    """Part codes tablosuna detay kolonlarÃƒÂ„Ã‚Â± ekle"""
    try:
        conn = get_db()
        cursor = conn.cursor()
        
        new_columns = [
            ('photo_path', 'TEXT'),
            ('catalog_image', 'TEXT'),
            ('description', 'TEXT'),
            ('used_in_machines', 'TEXT'),  # JSON array
            ('specifications', 'TEXT'),     # JSON object
            ('stock_location', 'TEXT'),
            ('supplier', 'TEXT'),
            ('unit_price', 'REAL'),
            ('critical_stock_level', 'INTEGER'),
            ('notes', 'TEXT'),
            ('last_updated', 'TIMESTAMP'),
            ('updated_by', 'INTEGER')
        ]
        
        for col_name, col_type in new_columns:
            try:
                cursor.execute(f'ALTER TABLE part_codes ADD COLUMN {col_name} {col_type}')
                print(f"ÃƒÂ¢Ã…Â“Ã¢Â€Â¦ Kolon eklendi: {col_name}")
            except sqlite3.OperationalError as e:
                if 'duplicate column name' in str(e).lower():
                    print(f"ÃƒÂ¢Ã…Â¡Ã‚Â ÃƒÂ¯Ã‚Â¸Ã‚Â� Kolon zaten var: {col_name}")
                else:
                    raise
        
        conn.commit()
        close_db(conn)
        print("ÃƒÂ¢Ã…Â“Ã¢Â€Â¦ Part details schema updated")
    except Exception as e:
        print(f"ÃƒÂ¢Ã…Â¡Ã‚Â ÃƒÂ¯Ã‚Â¸Ã‚Â� Part details schema error: {e}")

def login_required(f):
    from functools import wraps
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'user_id' not in session:
            return jsonify({'error': 'GiriÃƒÂ…Ã…Â¸ yapmanÃƒÂ„Ã‚Â±z gerekiyor'}), 401
        return f(*args, **kwargs)
    return decorated_function

def admin_required(f):
    from functools import wraps
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'user_id' not in session or session.get('role') != 'admin':
            return redirect('/admin')
        return f(*args, **kwargs)
    return decorated_function

def admin_required_decorator(f):
    """Sadece admin eriÃƒÂ…Ã…Â¸imi iÃƒÂƒÃ‚Â§in decorator"""
    from functools import wraps
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'user_id' not in session:
            return jsonify({'error': 'GiriÃƒÂ…Ã…Â¸ yapmanÃƒÂ„Ã‚Â±z gerekiyor'}), 401
        if session.get('role') != 'admin':
            return jsonify({'error': 'Bu iÃƒÂ…Ã…Â¸lem iÃƒÂƒÃ‚Â§in admin yetkisi gerekli'}), 403
        return f(*args, **kwargs)
    return decorated_function

@app.route('/favicon.ico')
def favicon():
    return '', 204

# ================================
# PART DETAILS API ENDPOINTS
# ================================

@app.route('/api/upload_part_photo/<part_code>', methods=['POST'])
@login_required
def upload_part_photo(part_code):
    """ParÃƒÂƒÃ‚Â§a fotoÃƒÂ„Ã…Â¸rafÃƒÂ„Ã‚Â± yÃƒÂƒÃ‚Â¼kle"""
    if 'photo' not in request.files:
        return jsonify({'error': 'Dosya bulunamadÃƒÂ„Ã‚Â±'}), 400
    
    file = request.files['photo']
    if file.filename == '':
        return jsonify({'error': 'Dosya seÃƒÂƒÃ‚Â§ilmedi'}), 400
    
    if file and allowed_file(file.filename):
        # KlasÃƒÂƒÃ‚Â¶r yoksa oluÃƒÂ…Ã…Â¸tur
        os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)
        
        # GÃƒÂƒÃ‚Â¼venli dosya adÃƒÂ„Ã‚Â±
        filename = secure_filename(file.filename)
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        new_filename = f"{part_code}_{timestamp}_{filename}"
        filepath = os.path.join(app.config['UPLOAD_FOLDER'], new_filename)
        
        file.save(filepath)
        
        # Database'e kaydet
        conn = get_db()
        cursor = conn.cursor()
        placeholder = get_db_placeholder()
        
        execute_query(cursor, f'''
            UPDATE part_codes 
            SET photo_path = {placeholder}, 
                last_updated = {placeholder},
                updated_by = {placeholder}
            WHERE part_code = {placeholder}
        ''', (f'part_photos/{new_filename}', datetime.now(), session['user_id'], part_code))
        
        conn.commit()
        close_db(conn)
        
        return jsonify({
            'success': True,
            'photo_path': f'part_photos/{new_filename}'
        })
    
    return jsonify({'error': 'GeÃƒÂƒÃ‚Â§ersiz dosya formatÃƒÂ„Ã‚Â± (PNG, JPG, GIF, WEBP)'}), 400

@app.route('/api/upload_catalog_image/<part_code>', methods=['POST'])
@login_required
def upload_catalog_image(part_code):
    """Katalog gÃƒÂƒÃ‚Â¶rÃƒÂƒÃ‚Â¼ntÃƒÂƒÃ‚Â¼sÃƒÂƒÃ‚Â¼ yÃƒÂƒÃ‚Â¼kle"""
    if 'catalog' not in request.files:
        return jsonify({'error': 'Dosya bulunamadÃƒÂ„Ã‚Â±'}), 400
    
    file = request.files['catalog']
    if file.filename == '':
        return jsonify({'error': 'Dosya seÃƒÂƒÃ‚Â§ilmedi'}), 400
    
    if file and allowed_file(file.filename):
        os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)
        
        filename = secure_filename(file.filename)
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        new_filename = f"{part_code}_catalog_{timestamp}_{filename}"
        filepath = os.path.join(app.config['UPLOAD_FOLDER'], new_filename)
        
        file.save(filepath)
        
        conn = get_db()
        cursor = conn.cursor()
        placeholder = get_db_placeholder()
        
        execute_query(cursor, f'''
            UPDATE part_codes 
            SET catalog_image = {placeholder}, 
                last_updated = {placeholder},
                updated_by = {placeholder}
            WHERE part_code = {placeholder}
        ''', (f'part_photos/{new_filename}', datetime.now(), session['user_id'], part_code))
        
        conn.commit()
        close_db(conn)
        
        return jsonify({
            'success': True,
            'catalog_image': f'part_photos/{new_filename}'
        })
    
    return jsonify({'error': 'GeÃƒÂƒÃ‚Â§ersiz dosya formatÃƒÂ„Ã‚Â±'}), 400

@app.route('/api/parts', methods=['GET'])
@login_required
def get_all_parts():
    """TÃƒÂƒÃ‚Â¼m parÃƒÂƒÃ‚Â§alarÃƒÂ„Ã‚Â± listele (admin panel iÃƒÂƒÃ‚Â§in)"""
    conn = get_db()
    cursor = conn.cursor()
    
    execute_query(cursor, '''
        SELECT part_code, description, photo_path, catalog_image
        FROM part_codes
        ORDER BY part_code
    ''')
    
    parts = cursor.fetchall()
    close_db(conn)
    
    result = []
    for part in parts:
        result.append({
            'part_code': part[0],
            'description': part[1],
            'photo_path': part[2],
            'catalog_image': part[3]
        })
    
    return jsonify(result)

@app.route('/api/part_details/<part_code>', methods=['GET'])
def get_part_details(part_code):
    """ParÃƒÂƒÃ‚Â§a detaylarÃƒÂ„Ã‚Â±nÃƒÂ„Ã‚Â± getir (herkes gÃƒÂƒÃ‚Â¶rebilir)"""
    conn = get_db()
    cursor = conn.cursor()
    placeholder = get_db_placeholder()
    
    execute_query(cursor, f'''
        SELECT part_code, description, photo_path, catalog_image,
               used_in_machines, specifications, stock_location,
               supplier, unit_price, critical_stock_level, notes,
               last_updated, updated_by
        FROM part_codes
        WHERE part_code = {placeholder}
    ''', (part_code,))
    
    part = cursor.fetchone()
    close_db(conn)
    
    if not part:
        return jsonify({'error': 'ParÃƒÂƒÃ‚Â§a bulunamadÃƒÂ„Ã‚Â±'}), 404
    
    return jsonify({
        'part_code': part[0],
        'description': part[1],
        'photo_path': part[2],
        'catalog_image': part[3],
        'used_in_machines': json.loads(part[4]) if part[4] else [],
        'specifications': json.loads(part[5]) if part[5] else {},
        'stock_location': part[6],
        'supplier': part[7],
        'unit_price': part[8],
        'critical_stock_level': part[9],
        'notes': part[10],
        'last_updated': part[11],
        'updated_by': part[12]
    })

@app.route('/api/update_part_details/<part_code>', methods=['POST'])
@login_required
def update_part_details(part_code):
    """ParÃƒÂƒÃ‚Â§a detaylarÃƒÂ„Ã‚Â±nÃƒÂ„Ã‚Â± gÃƒÂƒÃ‚Â¼ncelle (admin only)"""
    if session.get('role') != 'admin':
        return jsonify({'error': 'Admin yetkisi gerekli'}), 403
    
    data = request.json
    conn = get_db()
    cursor = conn.cursor()
    placeholder = get_db_placeholder()
    
    # Machines ve specs JSON olarak kaydet
    used_in_machines = json.dumps(data.get('used_in_machines', []))
    specifications = json.dumps(data.get('specifications', {}))
    
    execute_query(cursor, f'''
        UPDATE part_codes
        SET description = {placeholder},
            used_in_machines = {placeholder},
            specifications = {placeholder},
            stock_location = {placeholder},
            supplier = {placeholder},
            unit_price = {placeholder},
            critical_stock_level = {placeholder},
            notes = {placeholder},
            last_updated = {placeholder},
            updated_by = {placeholder}
        WHERE part_code = {placeholder}
    ''', (
        data.get('description'),
        used_in_machines,
        specifications,
        data.get('stock_location'),
        data.get('supplier'),
        data.get('unit_price'),
        data.get('critical_stock_level'),
        data.get('notes'),
        datetime.now(),
        session['user_id'],
        part_code
    ))
    
    conn.commit()
    close_db(conn)
    
    return jsonify({'success': True})

@app.route('/qr/<qr_id>')
def qr_redirect(qr_id):
    """QR Code yÃƒÂƒÃ‚Â¶nlendirme - Aktif sayÃƒÂ„Ã‚Â±m varsa scanner, yoksa part info"""
    conn = get_db()
    cursor = conn.cursor()
    placeholder = get_db_placeholder()
    
    # Aktif sayÃƒÂ„Ã‚Â±m var mÃƒÂ„Ã‚Â±?
    execute_query(cursor, f'SELECT id FROM count_sessions WHERE is_active = {placeholder}', (True,))
    active_session = cursor.fetchone()
    
    if active_session:
        # Aktif sayÃƒÂ„Ã‚Â±m var -> Scanner sayfasÃƒÂ„Ã‚Â±
        close_db(conn)
        return redirect(url_for('scanner'))
    else:
        # Aktif sayÃƒÂ„Ã‚Â±m yok -> Part detay sayfasÃƒÂ„Ã‚Â±
        # QR'dan part_code bul
        execute_query(cursor, f'''
            SELECT pc.part_code
            FROM qr_codes qr
            JOIN part_codes pc ON qr.part_code_id = pc.id
            WHERE qr.qr_id = {placeholder}
        ''', (qr_id,))
        
        result = cursor.fetchone()
        close_db(conn)
        
        if result:
            return redirect(url_for('part_info', part_code=result[0]))
        else:
            return "QR Code bulunamadÃƒÂ„Ã‚Â±", 404

@app.route('/part_info/<part_code>')
def part_info(part_code):
    """ParÃƒÂƒÃ‚Â§a bilgi sayfasÃƒÂ„Ã‚Â±"""
    return render_template('part_info.html', part_code=part_code)

@app.route('/edit_part/<part_code>')
@admin_required
def edit_part(part_code):
    """ParÃƒÂƒÃ‚Â§a dÃƒÂƒÃ‚Â¼zenleme sayfasÃƒÂ„Ã‚Â± (admin)"""
    return render_template('edit_part.html', part_code=part_code)

@app.route('/part_search')
def part_search():
    """ParÃƒÂƒÃ‚Â§a arama sayfasÃƒÂ„Ã‚Â±"""
    return render_template('part_search.html')

@app.route('/api/search_parts')
def search_parts():
    """ParÃƒÂƒÃ‚Â§a arama API'si - TR Klavye uyumlu"""
    query = request.args.get('q', '').strip()
    
    if not query or len(query) < 1:
        return jsonify([])
    
    try:
        # TÃƒÂƒÃ‚Â¼rkÃƒÂƒÃ‚Â§e karakterleri normalize et ve sembol uyuÃƒÂ…Ã…Â¸mazlÃƒÂ„Ã‚Â±ÃƒÂ„Ã…Â¸ÃƒÂ„Ã‚Â±nÃƒÂ„Ã‚Â± dÃƒÂƒÃ‚Â¼zelt
        def normalize_turkish(text):
            """TÃƒÂƒÃ‚Â¼rkÃƒÂƒÃ‚Â§e karakterleri ve sembol hatalarÃƒÂ„Ã‚Â±nÃƒÂ„Ã‚Â± dÃƒÂƒÃ‚Â¼zelt"""
            if not text:
                return text
            
            # SADECE TR klavyede yanlÃƒÂ„Ã‚Â±ÃƒÂ…Ã…Â¸ gelen sembolleri dÃƒÂƒÃ‚Â¼zelt
            # * ve ? -> - veya _ (hangisi varsa onu koru)
            # Son -X veya _X kÃƒÂ„Ã‚Â±smÃƒÂ„Ã‚Â±nÃƒÂ„Ã‚Â± kaldÃƒÂ„Ã‚Â±r (adet sayÃƒÂ„Ã‚Â±sÃƒÂ„Ã‚Â±)
            symbol_map = {
                '*': '-',      # Shift + / gives *
                '?': '-',      # Other combination  
            }
            
            # Sembol haritasÃƒÂ„Ã‚Â±nÃƒÂ„Ã‚Â± uygula
            for wrong_char, correct_char in symbol_map.items():
                text = text.replace(wrong_char, correct_char)
            
            # Son -X veya _X kÃƒÂ„Ã‚Â±smÃƒÂ„Ã‚Â±nÃƒÂ„Ã‚Â± kaldÃƒÂ„Ã‚Â±r (adet sayÃƒÂ„Ã‚Â±sÃƒÂ„Ã‚Â±)
            # ÃƒÂƒÃ¢Â€Â“rnek: Y129150-49811-3 -> Y129150-49811
            import re
            text = re.sub(r'[-_]\d+$', '', text)
            
            # NFD normalizasyonu (decompose) - TÃƒÂƒÃ‚Â¼rkÃƒÂƒÃ‚Â§e karakterler
            text = unicodedata.normalize('NFD', text.lower())
            # Accent iÃƒÂ…Ã…Â¸aretlerini kaldÃƒÂ„Ã‚Â±r
            output = []
            for char in text:
                if unicodedata.category(char) != 'Mn':  # Mn = Mark, Nonspacing
                    output.append(char)
            result = ''.join(output)
            
            # Manuel TR karakterler (eÃƒÂ„Ã…Â¸er normalize etmezse)
            tr_map = {
                'ÃƒÂƒÃ‚Â§': 'c', 'ÃƒÂ…Ã…Â¸': 's', 'ÃƒÂ„Ã…Â¸': 'g', 'ÃƒÂƒÃ‚Â¼': 'u', 'ÃƒÂƒÃ‚Â¶': 'o', 'ÃƒÂ„Ã‚Â±': 'i',
                'ÃƒÂƒÃ¢Â€Â¡': 'c', 'ÃƒÂ…Ã‚Âž': 's', 'ÃƒÂ„Ã‚Âž': 'g', 'ÃƒÂƒÃ…Â“': 'u', 'ÃƒÂƒÃ¢Â€Â“': 'o', 'ÃƒÂ„Ã‚Â°': 'i'
            }
            for tr_char, en_char in tr_map.items():
                result = result.replace(tr_char, en_char)
            
            return result
        
        normalized_query = normalize_turkish(query)
        
        conn = get_db()
        cursor = conn.cursor()
        placeholder = get_db_placeholder()
        
        # ÃƒÂ„Ã‚Â°lk olarak tam kod eÃƒÂ…Ã…Â¸leÃƒÂ…Ã…Â¸mesini ara (QR format: Y129150-49811)
        execute_query(cursor, f'''
            SELECT part_code, description
            FROM part_codes
            WHERE part_code = {placeholder}
            LIMIT 1
        ''', (normalized_query,))
        
        results = cursor.fetchall()
        
        # EÃƒÂ„Ã…Â¸er tam eÃƒÂ…Ã…Â¸leÃƒÂ…Ã…Â¸me yoksa, LIKE ile ara (hem normalized hem normal)
        if not results:
            search_pattern = f'%{normalized_query}%'
            execute_query(cursor, f'''
                SELECT part_code, description
                FROM part_codes
                WHERE part_code LIKE {placeholder} 
                   OR description LIKE {placeholder}
                LIMIT 20
            ''', (search_pattern, search_pattern))
            results = cursor.fetchall()
        
        close_db(conn)
        
        print(f'[Search API] Original: {query} | Normalized: {normalized_query} | Found: {len(results)}')
        
        return jsonify([{
            'part_code': row[0],
            'description': row[1]
        } for row in results])
    
    except Exception as e:
        print(f'Search error: {e}')
        return jsonify([])

# ================================
# EXISTING ROUTES BELOW
# ================================

@app.route('/')
def index():
    """Ana sayfa - Dashboard"""
    if 'user_id' not in session:
        return render_template('login.html')
    
    return render_template('index.html')


@app.route('/dashboard')
@login_required
def dashboard():
    """Dashboard - Ana sayfaya yÃƒÂƒÃ‚Â¶nlendir"""
    return redirect('/')


@app.route('/clear_cache.html')
def clear_cache_page():
    """Cache temizleme yardÃƒÂ„Ã‚Â±mcÃƒÂ„Ã‚Â± sayfasÃƒÂ„Ã‚Â±"""
    with open('clear_cache.html', 'r', encoding='utf-8') as f:
        return f.read()


# Note: a more comprehensive /health endpoint is defined later in this file.
# The detailed health check includes DB and storage checks and will be used by
# load balancers and Render readiness probes.

@app.route('/api/dashboard/stats')
@login_required
def dashboard_stats():
    """Cache'li dashboard istatistikleri"""
    cache_key = 'dashboard_stats'

    # Cache'den kontrol et
    cached_data = cache_get(cache_key)
    if cached_data:
        return jsonify(cached_data)

    # Cache yoksa veritabanÃƒÂ„Ã‚Â±ndan al
    conn = get_db()
    cursor = conn.cursor()

    try:
        # Toplam QR kod sayÃƒÂ„Ã‚Â±sÃƒÂ„Ã‚Â±
        # execute_query(cursor, 'SELECT COUNT(*) FROM envanter')
        total_qr_codes = cursor.fetchone()[0]

        # Toplam sayÃƒÂ„Ã‚Â±m sayÃƒÂ„Ã‚Â±sÃƒÂ„Ã‚Â±
        execute_query(cursor, 'SELECT COUNT(*) FROM scanned_qr')
        total_scans = cursor.fetchone()[0]

        # Aktif oturum sayÃƒÂ„Ã‚Â±sÃƒÂ„Ã‚Â±
        execute_query(cursor, 'SELECT COUNT(*) FROM count_sessions WHERE is_active = ?', (True,))
        active_sessions = cursor.fetchone()[0]

        # Son sayÃƒÂ„Ã‚Â±m tarihi
        execute_query(cursor, 'SELECT MAX(scanned_at) FROM scanned_qr')
        last_scan = cursor.fetchone()[0]

        # BugÃƒÂƒÃ‚Â¼nkÃƒÂƒÃ‚Â¼ sayÃƒÂ„Ã‚Â±mlar
        execute_query(cursor, '''
            SELECT COUNT(*) FROM scanned_qr 
            WHERE DATE(scanned_at) = CURRENT_DATE
        ''')
        today_scans = cursor.fetchone()[0]

        # Bu haftaki sayÃƒÂ„Ã‚Â±mlar
        execute_query(cursor, '''
            SELECT COUNT(*) FROM scanned_qr 
            WHERE scanned_at >= CURRENT_DATE - INTERVAL '7 days'
        ''')
        week_scans = cursor.fetchone()[0]

        stats = {
            'total_qr_codes': total_qr_codes,
            'total_scans': total_scans,
            'active_sessions': active_sessions,
            'last_scan': last_scan.isoformat() if last_scan else None,
            'today_scans': today_scans,
            'week_scans': week_scans,
            'cache_time': datetime.now().isoformat()
        }

        # Cache'e kaydet
        cache_set(cache_key, stats)

        return jsonify(stats)

    finally:
        close_db(conn)

# ========================================
#  QR SCANNER SAYIM SÃƒÂ„Ã‚Â°STEMÃƒÂ„Ã‚Â° (Sadece Desktop Scanner)
# ========================================

@app.route('/scanner')
@login_required
def scanner_page():
    """Desktop QR Scanner SayÃƒÂ„Ã‚Â±m SayfasÃƒÂ„Ã‚Â±"""
    return render_template('scanner.html')

@app.route('/live_dashboard')
@login_required
def live_dashboard_page():
    """CanlÃƒÂ„Ã‚Â± Dashboard SayfasÃƒÂ„Ã‚Â±"""
    return render_template('live_dashboard.html')

@app.route('/api/get_active_count_session')
def get_active_count_session():
    """Aktif sayÃƒÂ„Ã‚Â±m seansÃƒÂ„Ã‚Â±nÃƒÂ„Ã‚Â± kontrol et - Public endpoint (landing page iÃƒÂƒÃ‚Â§in)"""
    try:
        conn = get_db()
        cursor = conn.cursor()
        
        cursor.execute('''
            SELECT id, session_name, created_at, started_at, total_expected, total_scanned, description
            FROM count_sessions 
            WHERE is_active = 1
            ORDER BY created_at DESC
            LIMIT 1
        ''')
        
        result = cursor.fetchone()
        close_db(conn)
        
        if result:
            session_id, name, created_at, started_at, expected, scanned, description = result
            
            # SÃƒÂƒÃ‚Â¼re hesapla
            start_time = datetime.fromisoformat(started_at) if started_at else datetime.fromisoformat(created_at)
            duration_seconds = (datetime.now() - start_time).total_seconds()
            
            # YÃƒÂƒÃ‚Â¼zde hesapla
            percentage = (scanned / expected * 100) if expected > 0 else 0
            
            # Beklenen parÃƒÂƒÃ‚Â§alarÃƒÂ„Ã‚Â± parse et
            expected_parts = []
            if description:
                try:
                    expected_parts = json.loads(description)
                except:
                    pass
            
            return jsonify({
                'success': True,
                'has_active': True,
                'session': {
                    'id': session_id,
                    'name': name,
                    'created_at': created_at,
                    'started_at': started_at,
                    'duration_seconds': int(duration_seconds),
                    'total_expected': expected or 0,
                    'total_scanned': scanned or 0,
                    'percentage': round(percentage, 2),
                    'expected_parts': expected_parts
                }
            })
        else:
            return jsonify({
                'success': True,
                'has_active': False
            })
        
    except Exception as e:
        print(f"ÃƒÂ¢Ã‚Â�Ã…Â’ Get active session error: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/get_session_report/<session_id>')
@login_required
def get_session_report(session_id):
    """SayÃƒÂ„Ã‚Â±m raporu getir - beklenen vs sayÃƒÂ„Ã‚Â±lan karÃƒÂ…Ã…Â¸ÃƒÂ„Ã‚Â±laÃƒÂ…Ã…Â¸tÃƒÂ„Ã‚Â±rmasÃƒÂ„Ã‚Â±"""
    try:
        conn = get_db()
        cursor = conn.cursor()
        
        # Session bilgilerini al
        cursor.execute('''
            SELECT session_name, description, total_expected, total_scanned
            FROM count_sessions 
            WHERE id = ?
        ''', (session_id,))
        
        session_info = cursor.fetchone()
        
        if not session_info:
            close_db(conn)
            return jsonify({'success': False, 'error': 'Session bulunamadÃƒÂ„Ã‚Â±'}), 404
        
        session_name, description, total_expected, total_scanned = session_info
        
        # Beklenen parÃƒÂƒÃ‚Â§alarÃƒÂ„Ã‚Â± parse et
        expected_parts = []
        if description:
            try:
                expected_parts = json.loads(description)
            except:
                pass
        
        # SayÃƒÂ„Ã‚Â±lan parÃƒÂƒÃ‚Â§alarÃƒÂ„Ã‚Â± al - direkt scanned_qr tablosundan
        cursor.execute('''
            SELECT part_code, COUNT(*) as scanned_count
            FROM scanned_qr
            WHERE session_id = ?
            GROUP BY part_code
        ''', (session_id,))
        
        scanned_parts = {}
        for row in cursor.fetchall():
            part_code, count = row
            scanned_parts[part_code] = count
        
        # Part name'leri part_codes tablosundan al
        part_names = {}
        if expected_parts:
            part_codes_list = [item.get('part_code') for item in expected_parts if item.get('part_code')]
            if part_codes_list:
                placeholders = ','.join(['?' for _ in part_codes_list])
                cursor.execute(f'SELECT part_code, part_name FROM part_codes WHERE part_code IN ({placeholders})', part_codes_list)
                for row in cursor.fetchall():
                    part_names[row[0]] = row[1]
        
        close_db(conn)
        
        # Raporu oluÃƒÂ…Ã…Â¸tur
        report_items = []
        for expected in expected_parts:
            part_code = expected.get('part_code')
            part_name = expected.get('part_name') or part_names.get(part_code, '')
            
            # Excel'den gelen 'quantity' veya API'den gelen 'expected_quantity' - ikisini de destekle
            expected_qty = expected.get('expected_quantity') or expected.get('quantity', 0)
            
            scanned_qty = scanned_parts.get(part_code, 0)
            
            difference = scanned_qty - expected_qty
            
            if difference == 0:
                status = 'ÃƒÂ¢Ã…Â“Ã¢Â€Â¦ Tam'
            elif difference > 0:
                status = 'ÃƒÂ¢Ã‚Â¬Ã¢Â€Â ÃƒÂ¯Ã‚Â¸Ã‚Â� Fazla'
            else:
                status = 'ÃƒÂ¢Ã…Â¡Ã‚Â  Eksik'
            
            report_items.append({
                'part_code': part_code,
                'part_name': part_name,
                'expected': expected_qty,
                'scanned': scanned_qty,
                'difference': difference,
                'status': status
            })
        
        return jsonify({
            'success': True,
            'session_name': session_name,
            'total_expected': total_expected,
            'total_scanned': total_scanned,
            'report': report_items
        })
        
    except Exception as e:
        print(f"ÃƒÂ¢Ã‚Â�Ã…Â’ Get session report error: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/start_count_session', methods=['POST'])
@login_required
def start_count_session():
    """Yeni sayÃƒÂ„Ã‚Â±m seansÃƒÂ„Ã‚Â± baÃƒÂ…Ã…Â¸lat"""
    try:
        data = request.get_json()
        expected_parts = data.get('expected_parts', [])  # [{part_code, expected_quantity}]
        
        conn = get_db()
        cursor = conn.cursor()
        
        # ÃƒÂƒÃ¢Â€Â“nce aktif sayÃƒÂ„Ã‚Â±m var mÃƒÂ„Ã‚Â± kontrol et
        cursor.execute('SELECT id FROM count_sessions WHERE is_active = 1')
        if cursor.fetchone():
            close_db(conn)
            return jsonify({
                'success': False,
                'error': 'Zaten aktif bir sayÃƒÂ„Ã‚Â±m var! ÃƒÂƒÃ¢Â€Â“nce onu bitirin.'
            }), 400
        
        # Yeni sayÃƒÂ„Ã‚Â±m seansÃƒÂ„Ã‚Â± oluÃƒÂ…Ã…Â¸tur
        session_name = f"SayÃƒÂ„Ã‚Â±m {datetime.now().strftime('%Y%m%d_%H%M%S')}"
        
        # Excel'den 'quantity' veya API'den 'expected_quantity' - ikisini de destekle
        total_expected = sum(item.get('expected_quantity') or item.get('quantity', 0) for item in expected_parts)
        
        cursor.execute('''
            INSERT INTO count_sessions 
            (session_name, created_by, is_active, created_at, started_at, total_expected, total_scanned, description)
            VALUES (?, ?, 1, ?, ?, ?, 0, ?)
        ''', (
            session_name,
            session.get('user_id', 1),
            datetime.now(),
            datetime.now(),
            total_expected,
            json.dumps(expected_parts) if expected_parts else None
        ))
        
        session_id = cursor.lastrowid
        conn.commit()
        close_db(conn)
        
        return jsonify({
            'success': True,
            'session_id': session_id,
            'session_name': session_name,
            'message': f'ÃƒÂ¢Ã…Â“Ã¢Â€Â¦ SayÃƒÂ„Ã‚Â±m baÃƒÂ…Ã…Â¸latÃƒÂ„Ã‚Â±ldÃƒÂ„Ã‚Â±: {session_name}'
        })
        
    except Exception as e:
        print(f"ÃƒÂ¢Ã‚Â�Ã…Â’ Start count session error: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/finish_count', methods=['POST'])
@login_required
def finish_count():
    """SayÃƒÂ„Ã‚Â±mÃƒÂ„Ã‚Â± bitir, Excel raporunu kaydet ve rapor hazÃƒÂ„Ã‚Â±rla"""
    try:
        data = request.get_json()
        session_id = data.get('session_id', '1')
        
        conn = get_db()
        cursor = conn.cursor()
        
        # Toplam tarama sayÃƒÂ„Ã‚Â±sÃƒÂ„Ã‚Â±nÃƒÂ„Ã‚Â± hesapla
        cursor.execute('''
            SELECT COUNT(*) FROM scanned_qr WHERE session_id = ?
        ''', (str(session_id),))
        total_scanned = cursor.fetchone()[0]
        
        # Session bilgilerini al
        cursor.execute('''
            SELECT session_name, created_at, started_at, description, total_expected
            FROM count_sessions 
            WHERE id = ?
        ''', (str(session_id),))
        
        session_info = cursor.fetchone()
        
        if not session_info:
            close_db(conn)
            return jsonify({'success': False, 'error': 'SayÃƒÂ„Ã‚Â±m bulunamadÃƒÂ„Ã‚Â±'}), 404
        
        session_name, created_at, started_at, description, total_expected = session_info
        
        # Excel raporu oluÃƒÂ…Ã…Â¸tur
        report_filename = None
        try:
            # Beklenen parÃƒÂƒÃ‚Â§alarÃƒÂ„Ã‚Â± parse et
            expected_parts = {}
            if description:
                try:
                    expected_list = json.loads(description)
                    for item in expected_list:
                        part_code = item.get('ParÃƒÂƒÃ‚Â§a Kodu') or item.get('part_code')
                        expected_qty = item.get('Beklenen Adet') or item.get('expected_quantity') or item.get('quantity') or item.get('Adet') or 0
                        if part_code:
                            expected_parts[part_code] = int(expected_qty)
                except:
                    pass
            
            # Taranan parÃƒÂƒÃ‚Â§alarÃƒÂ„Ã‚Â± al
            cursor.execute('''
                SELECT sq.part_code, pc.part_name, COUNT(*) as scanned_count
                FROM scanned_qr sq
                LEFT JOIN part_codes pc ON sq.part_code = pc.part_code
                WHERE sq.session_id = ?
                GROUP BY sq.part_code, pc.part_name
                ORDER BY pc.part_name
            ''', (str(session_id),))
            
            scanned_results = cursor.fetchall()
            
            # Beklenen parÃƒÂƒÃ‚Â§alar iÃƒÂƒÃ‚Â§in part_name'leri ÃƒÂƒÃ‚Â§ek
            part_names = {}
            if expected_parts:
                part_codes_list = list(expected_parts.keys())
                if part_codes_list:
                    placeholders = ','.join(['?' for _ in part_codes_list])
                    cursor.execute(f'SELECT part_code, part_name FROM part_codes WHERE part_code IN ({placeholders})', part_codes_list)
                    for row in cursor.fetchall():
                        part_names[row[0]] = row[1]
            
            # TÃƒÂƒÃ‚Â¼m parÃƒÂƒÃ‚Â§alarÃƒÂ„Ã‚Â± birleÃƒÂ…Ã…Â¸tir (beklenen + taranan)
            all_parts = {}
            
            # Beklenen parÃƒÂƒÃ‚Â§alarÃƒÂ„Ã‚Â± ekle
            for part_code, expected_qty in expected_parts.items():
                all_parts[part_code] = {
                    'part_code': part_code,
                    'part_name': part_names.get(part_code, ''),
                    'expected': expected_qty,
                    'scanned': 0
                }
            
            # Taranan parÃƒÂƒÃ‚Â§alarÃƒÂ„Ã‚Â± ekle/gÃƒÂƒÃ‚Â¼ncelle
            for part_code, part_name, scanned_count in scanned_results:
                if part_code in all_parts:
                    all_parts[part_code]['part_name'] = part_name or ''
                    all_parts[part_code]['scanned'] = scanned_count
                else:
                    all_parts[part_code] = {
                        'part_code': part_code,
                        'part_name': part_name or '',
                        'expected': 0,
                        'scanned': scanned_count
                    }
            
            # DataFrame oluÃƒÂ…Ã…Â¸tur
            rows = []
            for part in all_parts.values():
                difference = part['scanned'] - part['expected']
                
                if difference == 0:
                    status = 'ÃƒÂ¢Ã…Â“Ã¢Â€Â¦ Tam'
                elif difference > 0:
                    status = 'ÃƒÂ¢Ã‚Â¬Ã¢Â€Â ÃƒÂ¯Ã‚Â¸Ã‚Â� Fazla'
                else:
                    status = 'ÃƒÂ¢Ã…Â¡Ã‚Â  Eksik'
                
                rows.append({
                    'ParÃƒÂƒÃ‚Â§a Kodu': part['part_code'],
                    'ParÃƒÂƒÃ‚Â§a AdÃƒÂ„Ã‚Â±': part['part_name'],
                    'Beklenen Adet': part['expected'],
                    'SayÃƒÂ„Ã‚Â±lan Adet': part['scanned'],
                    'Fark': difference,
                    'Durum': status
                })
            
            df = pd.DataFrame(rows)
            
            # Excel dosyasÃƒÂ„Ã‚Â±nÃƒÂ„Ã‚Â± kaydet
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            report_filename = f'sayim_raporu_{session_name}_{timestamp}.xlsx'
            report_path = os.path.join('static', 'reports', report_filename)
            
            # static/reports klasÃƒÂƒÃ‚Â¶rÃƒÂƒÃ‚Â¼nÃƒÂƒÃ‚Â¼ oluÃƒÂ…Ã…Â¸tur
            os.makedirs(os.path.join('static', 'reports'), exist_ok=True)
            
            with pd.ExcelWriter(report_path, engine='openpyxl') as writer:
                df.to_excel(writer, sheet_name='SayÃƒÂ„Ã‚Â±m Raporu', index=False)
                
                worksheet = writer.sheets['SayÃƒÂ„Ã‚Â±m Raporu']
                
                # Kolon geniÃƒÂ…Ã…Â¸liklerini ayarla
                column_widths = {
                    'A': 18,  # ParÃƒÂƒÃ‚Â§a Kodu
                    'B': 30,  # ParÃƒÂƒÃ‚Â§a AdÃƒÂ„Ã‚Â±
                    'C': 15,  # Beklenen Adet
                    'D': 15,  # SayÃƒÂ„Ã‚Â±lan Adet
                    'E': 12,  # Fark
                    'F': 15   # Durum
                }
                
                for col, width in column_widths.items():
                    worksheet.column_dimensions[col].width = width
                
                # BaÃƒÂ…Ã…Â¸lÃƒÂ„Ã‚Â±k satÃƒÂ„Ã‚Â±rÃƒÂ„Ã‚Â±nÃƒÂ„Ã‚Â± formatla
                from openpyxl.styles import Font, Alignment, PatternFill
                header_fill = PatternFill(start_color='DC2626', end_color='DC2626', fill_type='solid')
                header_font = Font(bold=True, color='FFFFFF', size=12)
                
                for cell in worksheet[1]:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                
                # Durum sÃƒÂƒÃ‚Â¼tununu renklendir
                for row_idx in range(2, len(rows) + 2):
                    cell = worksheet[f'F{row_idx}']
                    fark_cell = worksheet[f'E{row_idx}']
                    
                    fark_value = fark_cell.value
                    if fark_value == 0:
                        cell.fill = PatternFill(start_color='D4EDDA', end_color='D4EDDA', fill_type='solid')  # YeÃƒÂ…Ã…Â¸il
                    elif fark_value < 0:
                        cell.fill = PatternFill(start_color='FFF3CD', end_color='FFF3CD', fill_type='solid')  # SarÃƒÂ„Ã‚Â±
                    else:
                        cell.fill = PatternFill(start_color='D1ECF1', end_color='D1ECF1', fill_type='solid')  # Mavi
            
            print(f"ÃƒÂ¢Ã…Â“Ã¢Â€Â¦ Excel raporu kaydedildi: {report_path}")
            
        except Exception as e:
            print(f"ÃƒÂ¢Ã…Â¡Ã‚Â ÃƒÂ¯Ã‚Â¸Ã‚Â� Excel kayÃƒÂ„Ã‚Â±t hatasÃƒÂ„Ã‚Â±: {e}")
            import traceback
            traceback.print_exc()
        
        # Session'ÃƒÂ„Ã‚Â± bitir ve rapor yolunu kaydet
        cursor.execute('''
            UPDATE count_sessions 
            SET is_active = 0, ended_at = ?, total_scanned = ?, report_file_path = ?
            WHERE id = ?
        ''', (datetime.now(), total_scanned, report_filename, session_id))
        
        conn.commit()
        close_db(conn)
        
        return jsonify({
            'success': True,
            'message': f'ÃƒÂ¢Ã…Â“Ã¢Â€Â¦ SayÃƒÂ„Ã‚Â±m tamamlandÃƒÂ„Ã‚Â±! {total_scanned} adet tarama kaydedildi. Rapor oluÃƒÂ…Ã…Â¸turuldu.',
            'report_filename': report_filename
        })
        
    except Exception as e:
        print(f"ÃƒÂ¢Ã‚Â�Ã…Â’ Finish count error: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/reports')
@login_required
def reports_page():
    """TamamlanmÃƒÂ„Ã‚Â±ÃƒÂ…Ã…Â¸ sayÃƒÂ„Ã‚Â±m raporlarÃƒÂ„Ã‚Â±nÃƒÂ„Ã‚Â± listele"""
    return render_template('reports.html')


@app.route('/api/get_saved_reports')
@login_required
def get_saved_reports():
    """KaydedilmiÃƒÂ…Ã…Â¸ raporlarÃƒÂ„Ã‚Â± JSON olarak dÃƒÂƒÃ‚Â¶ndÃƒÂƒÃ‚Â¼r"""
    try:
        conn = get_db()
        cursor = conn.cursor()
        
        # TamamlanmÃƒÂ„Ã‚Â±ÃƒÂ…Ã…Â¸ sayÃƒÂ„Ã‚Â±mlarÃƒÂ„Ã‚Â± al (is_active = 0 ve report_file_path dolu olanlar)
        cursor.execute('''
            SELECT id, session_name, created_at, ended_at, 
                   total_expected, total_scanned, report_file_path
            FROM count_sessions 
            WHERE is_active = 0 AND report_file_path IS NOT NULL
            ORDER BY ended_at DESC
        ''')
        
        sessions = cursor.fetchall()
        close_db(conn)
        
        # Tuple'larÃƒÂ„Ã‚Â± dict'e ÃƒÂƒÃ‚Â§evir
        sessions_list = []
        for s in sessions:
            sessions_list.append({
                'id': s[0],
                'session_name': s[1],
                'created_at': s[2],
                'ended_at': s[3],
                'total_expected': s[4] or 0,
                'total_scanned': s[5] or 0,
                'report_file_path': s[6]
            })
        
        return jsonify(sessions_list)
        
    except Exception as e:
        print(f"ÃƒÂ¢Ã‚Â�Ã…Â’ Get saved reports error: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/download_count_excel/<session_id>')
@login_required
def download_count_excel(session_id):
    """SayÃƒÂ„Ã‚Â±m sonuÃƒÂƒÃ‚Â§larÃƒÂ„Ã‚Â±nÃƒÂ„Ã‚Â± detaylÃƒÂ„Ã‚Â± Excel olarak indir (beklenen vs taranan)"""
    try:
        conn = get_db()
        cursor = conn.cursor()
        
        # SayÃƒÂ„Ã‚Â±m bilgilerini al
        cursor.execute('''
            SELECT session_name, created_at, started_at, ended_at, 
                   description, total_expected, total_scanned
            FROM count_sessions 
            WHERE id = ?
        ''', (str(session_id),))
        
        session_info = cursor.fetchone()
        
        if not session_info:
            close_db(conn)
            return jsonify({'success': False, 'error': 'SayÃƒÂ„Ã‚Â±m bulunamadÃƒÂ„Ã‚Â±'}), 404
        
        session_name, created_at, started_at, ended_at, description, total_expected, total_scanned = session_info
        
        # Beklenen parÃƒÂƒÃ‚Â§alarÃƒÂ„Ã‚Â± parse et
        expected_parts = {}
        if description:
            try:
                expected_list = json.loads(description)
                for item in expected_list:
                    part_code = item.get('ParÃƒÂƒÃ‚Â§a Kodu') or item.get('part_code')
                    # Excel'den 'quantity', API'den 'expected_quantity', eski format 'Beklenen Adet'
                    expected_qty = item.get('Beklenen Adet') or item.get('expected_quantity') or item.get('quantity') or item.get('Adet') or 0
                    if part_code:
                        expected_parts[part_code] = int(expected_qty)
            except:
                pass
        
        # Taranan parÃƒÂƒÃ‚Â§alarÃƒÂ„Ã‚Â± al
        cursor.execute('''
            SELECT sq.part_code, pc.part_name, COUNT(*) as scanned_count
            FROM scanned_qr sq
            LEFT JOIN part_codes pc ON sq.part_code = pc.part_code
            WHERE sq.session_id = ?
            GROUP BY sq.part_code, pc.part_name
            ORDER BY pc.part_name
        ''', (str(session_id),))
        
        scanned_results = cursor.fetchall()
        
        # Beklenen parÃƒÂƒÃ‚Â§alar iÃƒÂƒÃ‚Â§in part_name'leri ÃƒÂƒÃ‚Â§ek
        part_names = {}
        if expected_parts:
            part_codes_list = list(expected_parts.keys())
            if part_codes_list:
                placeholders = ','.join(['?' for _ in part_codes_list])
                cursor.execute(f'SELECT part_code, part_name FROM part_codes WHERE part_code IN ({placeholders})', part_codes_list)
                for row in cursor.fetchall():
                    part_names[row[0]] = row[1]
        
        close_db(conn)
        
        # TÃƒÂƒÃ‚Â¼m parÃƒÂƒÃ‚Â§alarÃƒÂ„Ã‚Â± birleÃƒÂ…Ã…Â¸tir (beklenen + taranan)
        all_parts = {}
        
        # Beklenen parÃƒÂƒÃ‚Â§alarÃƒÂ„Ã‚Â± ekle
        for part_code, expected_qty in expected_parts.items():
            all_parts[part_code] = {
                'part_code': part_code,
                'part_name': part_names.get(part_code, ''),  # DB'den ÃƒÂƒÃ‚Â§ekilen part_name
                'expected': expected_qty,
                'scanned': 0
            }
        
        # Taranan parÃƒÂƒÃ‚Â§alarÃƒÂ„Ã‚Â± ekle/gÃƒÂƒÃ‚Â¼ncelle
        for part_code, part_name, scanned_count in scanned_results:
            if part_code in all_parts:
                all_parts[part_code]['part_name'] = part_name or ''
                all_parts[part_code]['scanned'] = scanned_count
            else:
                all_parts[part_code] = {
                    'part_code': part_code,
                    'part_name': part_name or '',
                    'expected': 0,
                    'scanned': scanned_count
                }
        
        # DataFrame oluÃƒÂ…Ã…Â¸tur - Sadece parÃƒÂƒÃ‚Â§a detaylarÃƒÂ„Ã‚Â±
        rows = []
        for part in all_parts.values():
            difference = part['scanned'] - part['expected']
            
            # Durum emoji
            if difference == 0:
                status = 'ÃƒÂ¢Ã…Â“Ã¢Â€Â¦ Tam'
            elif difference > 0:
                status = 'ÃƒÂ¢Ã‚Â¬Ã¢Â€Â ÃƒÂ¯Ã‚Â¸Ã‚Â� Fazla'
            else:
                status = 'ÃƒÂ¢Ã…Â¡Ã‚Â  Eksik'
            
            rows.append({
                'ParÃƒÂƒÃ‚Â§a Kodu': part['part_code'],
                'ParÃƒÂƒÃ‚Â§a AdÃƒÂ„Ã‚Â±': part['part_name'],
                'Beklenen Adet': part['expected'],
                'SayÃƒÂ„Ã‚Â±lan Adet': part['scanned'],
                'Fark': difference,
                'Durum': status
            })
        
        df = pd.DataFrame(rows)
        
        # Excel buffer
        output = BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            # Tek sayfa - DetaylÃƒÂ„Ã‚Â± Rapor
            df.to_excel(writer, sheet_name='SayÃƒÂ„Ã‚Â±m Raporu', index=False)
            
            # Formatla
            worksheet = writer.sheets['SayÃƒÂ„Ã‚Â±m Raporu']
            
            # Kolon geniÃƒÂ…Ã…Â¸liklerini ayarla
            column_widths = {
                'A': 18,  # ParÃƒÂƒÃ‚Â§a Kodu
                'B': 30,  # ParÃƒÂƒÃ‚Â§a AdÃƒÂ„Ã‚Â±
                'C': 15,  # Beklenen Adet
                'D': 15,  # SayÃƒÂ„Ã‚Â±lan Adet
                'E': 12,  # Fark
                'F': 15   # Durum
            }
            
            for col, width in column_widths.items():
                worksheet.column_dimensions[col].width = width
            
            # BaÃƒÂ…Ã…Â¸lÃƒÂ„Ã‚Â±k satÃƒÂ„Ã‚Â±rÃƒÂ„Ã‚Â±nÃƒÂ„Ã‚Â± kalÃƒÂ„Ã‚Â±nlaÃƒÂ…Ã…Â¸tÃƒÂ„Ã‚Â±r
            from openpyxl.styles import Font, Alignment, PatternFill
            header_fill = PatternFill(start_color='DC2626', end_color='DC2626', fill_type='solid')
            header_font = Font(bold=True, color='FFFFFF', size=12)
            
            for cell in worksheet[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # Durum sÃƒÂƒÃ‚Â¼tununu renklendir
            for row_idx in range(2, len(rows) + 2):
                cell = worksheet[f'F{row_idx}']
                fark_cell = worksheet[f'E{row_idx}']
                
                fark_value = fark_cell.value
                if fark_value == 0:
                    cell.fill = PatternFill(start_color='D4EDDA', end_color='D4EDDA', fill_type='solid')  # YeÃƒÂ…Ã…Â¸il
                elif fark_value < 0:
                    cell.fill = PatternFill(start_color='FFF3CD', end_color='FFF3CD', fill_type='solid')  # SarÃƒÂ„Ã‚Â±
                else:
                    cell.fill = PatternFill(start_color='D1ECF1', end_color='D1ECF1', fill_type='solid')  # Mavi
        
        output.seek(0)
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=f'sayim_raporu_{session_name}_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        )
        
    except Exception as e:
        print(f"ÃƒÂ¢Ã‚Â�Ã…Â’ Download Excel error: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/login', methods=['POST'])
@rate_limit_login
def login():
    from werkzeug.security import check_password_hash
    data = request.get_json()
    username = data.get('username')
    password = data.get('password')

    if not username or not password:
        return jsonify({'error': 'KullanÃƒÂ„Ã‚Â±cÃƒÂ„Ã‚Â± adÃƒÂ„Ã‚Â± ve ÃƒÂ…Ã…Â¸ifre gerekli'}), 400

    conn = get_db()
    cursor = conn.cursor()
    placeholder = get_db_placeholder()
    try:
        # Dual-mode: SQLite vs PostgreSQL table compatibility
        # Get user data first, then verify password
        if USE_POSTGRESQL:
            # PostgreSQL: Full schema with envanter_users table
            execute_query(cursor, f'SELECT id, username, full_name, role, password_hash FROM envanter_users WHERE username = {placeholder}',
                         (username,))
        else:
            # SQLite: envanter_users table with full_name (after column addition)
            execute_query(cursor, f'SELECT id, username, COALESCE(full_name, username) as full_name, role, password_hash FROM envanter_users WHERE username = {placeholder}',
                         (username,))
    except Exception as e:
        #  LOKAL: SQLite hata yÃƒÂƒÃ‚Â¶netimi
        logging.exception(f"Database error during login: {e}")
        try:
            close_db(conn)
        except Exception:
            pass
        return jsonify({'error': 'VeritabanÃƒÂ„Ã‚Â± hatasÃƒÂ„Ã‚Â±'}), 500

    user = cursor.fetchone()
    close_db(conn)

    if user and check_password_hash(user[4], password):  # user[4] is password_hash
        session['user_id'] = user[0]
        session['username'] = user[1]
        session['full_name'] = user[2]
        session['role'] = user[3]
        if user[3] == 'admin':
            session['admin_authenticated'] = True
        return jsonify({'success': True, 'role': user[3]})
    else:
        return jsonify({'error': 'KullanÃƒÂ„Ã‚Â±cÃƒÂ„Ã‚Â± adÃƒÂ„Ã‚Â± veya ÃƒÂ…Ã…Â¸ifre hatalÃƒÂ„Ã‚Â±'}), 401

@app.route('/logout', methods=['POST'])
def logout():
    session.clear()
    return jsonify({'success': True})

@app.route('/check_auth')
def check_auth():
    if 'user_id' in session:
        return jsonify({
            'authenticated': True,
            'username': session.get('username', None),
            'full_name': session.get('full_name', None),
            'role': session.get('role', None)
        })
    return jsonify({'authenticated': False})

# ============================================================================
#  HEALTH CHECK & DEBUG ENDPOINTS
# ============================================================================

@app.route('/health', methods=['GET'])
def health_check():
    """Sistem saÃƒÂ„Ã…Â¸lÃƒÂ„Ã‚Â±k kontrolÃƒÂƒÃ‚Â¼ - monitoring iÃƒÂƒÃ‚Â§in"""
    try:
        # DB baÃƒÂ„Ã…Â¸lantÃƒÂ„Ã‚Â±sÃƒÂ„Ã‚Â± test et
        db.session.execute(db.text('SELECT 1'))
        db_status = 'healthy'
        db_error = None
    except Exception as e:
        db_status = 'unhealthy'
        db_error = str(e)
    
    # Session sayÃƒÂ„Ã‚Â±sÃƒÂ„Ã‚Â±
    try:
        active_sessions = CountSession.query.filter_by(is_active=True).count()
    except:
        active_sessions = 0
    
    # Disk kullanÃƒÂ„Ã‚Â±mÃƒÂ„Ã‚Â±
    try:
        import shutil
        disk_usage = shutil.disk_usage('.')
        disk_free_gb = round(disk_usage.free / (1024**3), 2)
        disk_total_gb = round(disk_usage.total / (1024**3), 2)
        disk_used_percent = round((disk_usage.used / disk_usage.total) * 100, 1)
    except:
        disk_free_gb = 0
        disk_total_gb = 0
        disk_used_percent = 0
    
    # DB dosya boyutu
    try:
        db_size_mb = round(os.path.getsize('instance/envanter_local.db') / (1024**2), 2)
    except:
        db_size_mb = 0
    
    # Uptime
    uptime_seconds = int(time.time() - app.config.get('START_TIME', time.time()))
    uptime_hours = round(uptime_seconds / 3600, 1)
    
    health_data = {
        'status': 'ok' if db_status == 'healthy' else 'degraded',
        'timestamp': datetime.utcnow().isoformat(),
        'database': {
            'status': db_status,
            'error': db_error,
            'size_mb': db_size_mb,
            'type': 'SQLite'
        },
        'environment': {
            'mode': 'LOCAL' if IS_LOCAL else 'PRODUCTION',
            'storage': 'Local Files',
            'python_version': f"{os.sys.version_info.major}.{os.sys.version_info.minor}.{os.sys.version_info.micro}"
        },
        'sessions': {
            'active_count': active_sessions
        },
        'disk': {
            'free_gb': disk_free_gb,
            'total_gb': disk_total_gb,
            'used_percent': disk_used_percent
        },
        'uptime': {
            'seconds': uptime_seconds,
            'hours': uptime_hours
        }
    }
    
    return jsonify(health_data)

@app.route('/log_frontend_error', methods=['POST'])
def log_frontend_error():
    """Frontend hatalarÃƒÂ„Ã‚Â±nÃƒÂ„Ã‚Â± logla"""
    try:
        data = request.json
        
        # Log dosyasÃƒÂ„Ã‚Â±na yaz
        app.logger.error(f"FRONTEND ERROR: {json.dumps(data, ensure_ascii=False)}")
        
        # AyrÃƒÂ„Ã‚Â± dosyaya da yaz
        os.makedirs('logs', exist_ok=True)
        with open('logs/frontend_errors.log', 'a', encoding='utf-8') as f:
            error_line = {
                'timestamp': datetime.utcnow().isoformat(),
                'error': data.get('error', {}),
                'context': data.get('context', {}),
                'user': {
                    'username': session.get('username'),
                    'user_id': session.get('user_id')
                }
            }
            f.write(json.dumps(error_line, ensure_ascii=False) + '\n')
        
        return jsonify({'success': True})
    except Exception as e:
        app.logger.error(f'Frontend error logging hatasÃƒÂ„Ã‚Â±: {e}')
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/admin/frontend_errors_log', methods=['GET'])
@login_required
def get_frontend_errors_log():
    """Frontend error log dosyasÃƒÂ„Ã‚Â±nÃƒÂ„Ã‚Â± oku - sadece admin"""
    # Admin kontrolÃƒÂƒÃ‚Â¼
    if session.get('role') != 'admin':
        return jsonify({'error': 'Admin eriÃƒÂ…Ã…Â¸imi gerekli'}), 403
    
    try:
        errors = []
        log_path = 'logs/frontend_errors.log'
        
        if os.path.exists(log_path):
            with open(log_path, 'r', encoding='utf-8') as f:
                lines = f.readlines()
                # Son 50 satÃƒÂ„Ã‚Â±r
                for line in lines[-50:]:
                    try:
                        errors.append(json.loads(line.strip()))
                    except:
                        pass
        
        return jsonify({
            'success': True,
            'errors': errors,
            'count': len(errors)
        })
    except Exception as e:
        app.logger.error(f'Frontend errors log okuma hatasÃƒÂ„Ã‚Â±: {e}')
        return jsonify({'success': False, 'error': str(e)}), 500

        return jsonify({'success': True, 'message': 'Hata kaydedildi'})
    except Exception as e:
        app.logger.error(f'Frontend error logging failed: {e}')
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/upload_parts', methods=['POST'])
@login_required
def upload_parts():
    """
    Yeni ParÃƒÂƒÃ‚Â§a YÃƒÂƒÃ‚Â¼kleme Sistemi:
    - Excel'den sadece parÃƒÂƒÃ‚Â§a bilgileri (part_code, part_name) yÃƒÂƒÃ‚Â¼klenir
    - QR kod ÃƒÂƒÃ‚Â¼retimi yapÃƒÂ„Ã‚Â±lmaz (manuel olarak parÃƒÂƒÃ‚Â§a detay sayfasÃƒÂ„Ã‚Â±ndan ÃƒÂƒÃ‚Â¼retilir)
    - Mevcut parÃƒÂƒÃ‚Â§alar gÃƒÂƒÃ‚Â¼ncellenir, yeni parÃƒÂƒÃ‚Â§alar eklenir
    """
    if 'file' not in request.files:
        return jsonify({'error': 'Dosya bulunamadÃƒÂ„Ã‚Â±'}), 400

    file = request.files['file']
    if file.filename == '':
        return jsonify({'error': 'Dosya seÃƒÂƒÃ‚Â§ilmedi'}), 400

    if not (file.filename and file.filename.endswith(('.xlsx', '.xls'))):
        return jsonify({'error': 'Sadece Excel dosyalarÃƒÂ„Ã‚Â± yÃƒÂƒÃ‚Â¼klenebilir'}), 400

    try:
        df = pd.read_excel(file)

        # Sadece part_code ve part_name gerekli
        required_columns = ['part_code', 'part_name']
        if not all(col in df.columns for col in required_columns):
            return jsonify({'error': 'Excel dosyasÃƒÂ„Ã‚Â± "part_code" ve "part_name" sÃƒÂƒÃ‚Â¼tunlarÃƒÂ„Ã‚Â±nÃƒÂ„Ã‚Â± iÃƒÂƒÃ‚Â§ermelidir'}), 400

        conn = get_db()
        cursor = conn.cursor()
        placeholder = get_db_placeholder()

        # Mevcut parÃƒÂƒÃ‚Â§alarÃƒÂ„Ã‚Â± al
        execute_query(cursor, 'SELECT id, part_code, part_name FROM part_codes')
        existing_parts = {}  # {part_code: (id, part_name)}
        for row in cursor.fetchall():
            existing_parts[row[1]] = (row[0], row[2])

        new_parts = []
        updated_parts = []
        processing_summary = {
            'new_parts': 0,
            'updated_parts': 0
        }

        print(f"\nÃƒÂ¯Ã‚Â¿Ã‚Â½ PARÃƒÂƒÃ¢Â€Â¡A YÃƒÂƒÃ…Â“KLEME SÃƒÂ„Ã‚Â°STEMÃƒÂ„Ã‚Â°")
        print(f" Excel'den gelen parÃƒÂƒÃ‚Â§a sayÃƒÂ„Ã‚Â±sÃƒÂ„Ã‚Â±: {len(df)}")
        print("="*50)

        for _, row in df.iterrows():
            part_code = str(row['part_code']).strip()
            part_name = str(row['part_name']).strip()

            if not part_code or not part_name:
                continue  # BoÃƒÂ…Ã…Â¸ satÃƒÂ„Ã‚Â±rlarÃƒÂ„Ã‚Â± atla

            if part_code in existing_parts:
                # MEVCUT PARÃƒÂƒÃ¢Â€Â¡A - Sadece ismi gÃƒÂƒÃ‚Â¼ncelle
                part_code_id, old_part_name = existing_parts[part_code]
                if old_part_name != part_name:
                    execute_query(cursor, f'UPDATE part_codes SET part_name = {placeholder} WHERE part_code = {placeholder}',
                                 (part_name, part_code))
                    updated_parts.append(part_code)
                    processing_summary['updated_parts'] += 1
                    print(f" {part_code}: '{old_part_name}' ÃƒÂ¢Ã¢Â€Â Ã¢Â€Â™ '{part_name}'")
                else:
                    print(f"ÃƒÂ¢Ã…Â“Ã¢Â€Â¦ {part_code}: Zaten gÃƒÂƒÃ‚Â¼ncel")
            else:
                # YENÃƒÂ„Ã‚Â° PARÃƒÂƒÃ¢Â€Â¡A - Sadece part_codes'a ekle (QR kod ÃƒÂƒÃ‚Â¼retilmez)
                execute_query(cursor, f'INSERT INTO part_codes (part_code, part_name, description, created_at) VALUES ({placeholder}, {placeholder}, {placeholder}, {placeholder})',
                             (part_code, part_name, '', datetime.now()))
                new_parts.append(part_code)
                processing_summary['new_parts'] += 1
                print(f"Ã„ÂŸÃ…Â¸Ã¢Â€Â Ã¢Â€Â¢ {part_code}: Yeni parÃƒÂƒÃ‚Â§a eklendi - '{part_name}'")

        conn.commit()
        close_db(conn)

        print(f"\nÃƒÂ¢Ã…Â“Ã¢Â€Â¦ ÃƒÂ„Ã‚Â°ÃƒÂ…Ã‚ÂžLEM TAMAMLANDI")
        print(f" Yeni parÃƒÂƒÃ‚Â§alar: {processing_summary['new_parts']}")
        print(f" GÃƒÂƒÃ‚Â¼ncellenen parÃƒÂƒÃ‚Â§alar: {processing_summary['updated_parts']}")
        print(" QR kodlarÃƒÂ„Ã‚Â± parÃƒÂƒÃ‚Â§a detay sayfasÃƒÂ„Ã‚Â±ndan ÃƒÂƒÃ‚Â¼retilebilir")
        print("="*50)

        return jsonify({
            'success': True,
            'message': f'ÃƒÂ„Ã‚Â°ÃƒÂ…Ã…Â¸lem tamamlandÃƒÂ„Ã‚Â±! {processing_summary["new_parts"]} yeni parÃƒÂƒÃ‚Â§a eklendi, {processing_summary["updated_parts"]} parÃƒÂƒÃ‚Â§a gÃƒÂƒÃ‚Â¼ncellendi.',
            'summary': {
                'new_parts': processing_summary['new_parts'],
                'updated_parts': processing_summary['updated_parts']
            }
        })

    except Exception as e:
        logging.exception(f"Error in smart QR upload system: {e}")
        try:
            close_db(conn)
        except Exception:
            pass
        return jsonify({'error': f'ÃƒÂ„Ã‚Â°ÃƒÂ…Ã…Â¸lem sÃƒÂ„Ã‚Â±rasÃƒÂ„Ã‚Â±nda hata oluÃƒÂ…Ã…Â¸tu: {str(e)}'}), 500

@app.route('/get_qr_codes')
@login_required
def get_qr_codes():
    search = request.args.get('search', '').strip()
    page = int(request.args.get('page', 1))
    limit = int(request.args.get('limit', 100))
    offset = (page - 1) * limit

    conn = get_db()
    cursor = conn.cursor()

    if search:
        # ÃƒÂƒÃ¢Â€Â“nce tam eÃƒÂ…Ã…Â¸leÃƒÂ…Ã…Â¸me ara (JOIN ile part_code ve part_name al)
        execute_query(cursor, """
            SELECT qc.qr_id, pc.part_code, pc.part_name, qc.is_used, 
                   CASE WHEN qc.blob_url IS NOT NULL THEN 1 ELSE 0 END as is_downloaded
            FROM qr_codes qc
            JOIN part_codes pc ON qc.part_code_id = pc.id
            WHERE pc.part_code = ? OR pc.part_name = ?
            ORDER BY pc.part_code, qc.qr_id
            LIMIT ? OFFSET ?
        """, (search, search, limit, offset))
        exact_matches = cursor.fetchall()

        if exact_matches:
            qr_codes = []
            for row in exact_matches:
                qr_codes.append({
                    'qr_id': row[0],
                    'part_code': row[1],
                    'part_name': row[2],
                    'is_used': row[3],
                    'is_downloaded': row[4]
                })
        else:
            # Tam eÃƒÂ…Ã…Â¸leÃƒÂ…Ã…Â¸me bulunamazsa kÃƒÂ„Ã‚Â±smi eÃƒÂ…Ã…Â¸leÃƒÂ…Ã…Â¸me ara
            execute_query(cursor, """
                SELECT qc.qr_id, pc.part_code, pc.part_name, qc.is_used,
                       CASE WHEN qc.blob_url IS NOT NULL THEN 1 ELSE 0 END as is_downloaded
                FROM qr_codes qc
                JOIN part_codes pc ON qc.part_code_id = pc.id
                WHERE pc.part_code LIKE ? OR pc.part_name LIKE ?
                ORDER BY pc.part_code, qc.qr_id
                LIMIT ? OFFSET ?
            """, (f'%{search}%', f'%{search}%', limit, offset))
            rows = cursor.fetchall()
            qr_codes = []
            for row in rows:
                qr_codes.append({
                    'qr_id': row[0],
                    'part_code': row[1],
                    'part_name': row[2],
                    'is_used': row[3],
                    'is_downloaded': row[4]
                })
    else:
        # Arama terimi yoksa tÃƒÂƒÃ‚Â¼m QR kodlarÃƒÂ„Ã‚Â± getir (sayfalama ile)
        execute_query(cursor, """
            SELECT qc.qr_id, pc.part_code, pc.part_name, qc.is_used,
                   CASE WHEN qc.blob_url IS NOT NULL THEN 1 ELSE 0 END as is_downloaded
            FROM qr_codes qc
            JOIN part_codes pc ON qc.part_code_id = pc.id
            ORDER BY pc.part_code, qc.qr_id
            LIMIT ? OFFSET ?
        """, (limit, offset))
        rows = cursor.fetchall()
        qr_codes = []
        for row in rows:
            qr_codes.append({
                'qr_id': row[0],
                'part_code': row[1],
                'part_name': row[2],
                'is_used': row[3],
                'is_downloaded': row[4]
            })

    # Toplam sayÃƒÂ„Ã‚Â±yÃƒÂ„Ã‚Â± al (JOIN ile)
    if search:
        execute_query(cursor, """
            SELECT COUNT(*) 
            FROM qr_codes qc
            JOIN part_codes pc ON qc.part_code_id = pc.id
            WHERE pc.part_code LIKE ? OR pc.part_name LIKE ?
        """, (f'%{search}%', f'%{search}%'))
    else:
        execute_query(cursor, "SELECT COUNT(*) FROM qr_codes")

    total_count = cursor.fetchone()[0]

    close_db(conn)

    return jsonify({
        'qr_codes': qr_codes,
        'total_count': total_count,
        'current_page': page,
        'total_pages': (total_count + limit - 1) // limit,
        'has_more': (page * limit) < total_count
    })

@app.route('/clear_all_qrs', methods=['POST'])
@login_required
def clear_all_qrs():
    """TÃƒÂƒÃ‚Â¼m QR kodlarÃƒÂ„Ã‚Â±nÃƒÂ„Ã‚Â± temizle (aktif sayÃƒÂ„Ã‚Â±m oturumu yoksa)"""
    try:
        conn = get_db()
        cursor = conn.cursor()

        # Aktif sayÃƒÂ„Ã‚Â±m oturumu kontrolÃƒÂƒÃ‚Â¼
        execute_query(cursor, 'SELECT COUNT(*) FROM count_sessions WHERE is_active = ?', (True,))
        active_session = cursor.fetchone()[0]

        if active_session > 0:
            close_db(conn)
            return jsonify({'error': 'Aktif bir sayÃƒÂ„Ã‚Â±m oturumu var. QR kodlarÃƒÂ„Ã‚Â± silinemez.'}), 400

        #  LOKAL: QR klasÃƒÂƒÃ‚Â¶rÃƒÂƒÃ‚Â¼nÃƒÂƒÃ‚Â¼ temizle
        qr_dir = os.path.join('static', 'qrcodes')
        if os.path.exists(qr_dir):
            for file in os.listdir(qr_dir):
                try:
                    os.remove(os.path.join(qr_dir, file))
                except Exception as e:
                    logging.error(f"Lokal dosya {file} silinirken hata: {e}")

        # QR kodlarÃƒÂ„Ã‚Â±nÃƒÂ„Ã‚Â± ve parÃƒÂƒÃ‚Â§alarÃƒÂ„Ã‚Â± sil
        execute_query(cursor, 'DELETE FROM qr_codes')
        execute_query(cursor, 'DELETE FROM part_codes')

        conn.commit()
        close_db(conn)

        # Cache'i temizle
        cache_clear()

        logging.info("TÃƒÂƒÃ‚Â¼m QR kodlarÃƒÂ„Ã‚Â± temizlendi")
        return jsonify({
            'success': True,
            'message': 'TÃƒÂƒÃ‚Â¼m QR kodlarÃƒÂ„Ã‚Â± baÃƒÂ…Ã…Â¸arÃƒÂ„Ã‚Â±yla silindi'
        })
    except Exception as e:
        try:
            close_db(conn)
        except:
            pass
        logging.error(f"QR kodlarÃƒÂ„Ã‚Â± silinirken hata: {e}", exc_info=True)
        return jsonify({'error': f'QR kodlarÃƒÂ„Ã‚Â± silinirken hata: {str(e)}'}), 500

@app.route('/generate_qr_image/<qr_id>')
@login_required
def generate_qr_image(qr_id):
    """QR kod oluÃƒÂ…Ã…Â¸turma - QR altÃƒÂ„Ã‚Â±na kod ve parÃƒÂƒÃ‚Â§a adÃƒÂ„Ã‚Â± ekler"""
    try:
        # Cache'den kontrol et
        cache_key = f'qr_image_{qr_id}'
        cached_image = cache_get(cache_key)

        if cached_image:
            buf = BytesIO(cached_image)
            return send_file(buf, mimetype='image/png')

        #  LOKAL: Static klasÃƒÂƒÃ‚Â¶rden kontrol et
        static_path = os.path.join('static', 'qrcodes', f'{qr_id}.png')
        if os.path.exists(static_path):
            with open(static_path, 'rb') as f:
                file_content = f.read()
            cache_set(cache_key, file_content)
            buf = BytesIO(file_content)
            return send_file(buf, mimetype='image/png')

        # QR ID'den parÃƒÂƒÃ‚Â§a kodunu ve numarayÃƒÂ„Ã‚Â± ÃƒÂƒÃ‚Â§ÃƒÂ„Ã‚Â±kar (Y129150-49811_1)
        parts = qr_id.rsplit('_', 1)
        part_code = parts[0] if len(parts) > 0 else qr_id
        qr_number = parts[1] if len(parts) > 1 else "1"
        
        # ParÃƒÂƒÃ‚Â§a adÃƒÂ„Ã‚Â±nÃƒÂ„Ã‚Â± database'den al
        conn = get_db()
        cursor = conn.cursor()
        execute_query(cursor, 'SELECT part_name FROM part_codes WHERE part_code = ?', (part_code,))
        result = cursor.fetchone()
        part_name = result[0] if result else ""
        close_db(conn)

        # QR kod oluÃƒÂ…Ã…Â¸tur - Barkod makinesi iÃƒÂƒÃ‚Â§in optimize
        qr = qrcode.QRCode(
            version=1, 
            box_size=8,  # 8px - barkod makinesi iÃƒÂƒÃ‚Â§in ideal
            border=2,    # 2px quiet zone
            error_correction=qrcode.constants.ERROR_CORRECT_M
        )
        qr.add_data(qr_id)
        qr.make(fit=True)
        qr_img = qr.make_image(fill_color="black", back_color="white")
        qr_img = qr_img.convert('RGB')  # PIL Image'a dÃƒÂƒÃ‚Â¶nÃƒÂƒÃ‚Â¼ÃƒÂ…Ã…Â¸tÃƒÂƒÃ‚Â¼r

        # QR kod boyutlarÃƒÂ„Ã‚Â±nÃƒÂ„Ã‚Â± al
        qr_width, qr_height = qr_img.size
        
        # AlanlarÃƒÂ„Ã‚Â± hesapla
        logo_height = 40  # Logo iÃƒÂƒÃ‚Â§in ÃƒÂƒÃ‚Â¼st alan
        text_height = 35  # Alt yazÃƒÂ„Ã‚Â± (parÃƒÂƒÃ‚Â§a numarasÃƒÂ„Ã‚Â±) iÃƒÂƒÃ‚Â§in alan
        
        # KÃƒÂ„Ã‚Â±rmÃƒÂ„Ã‚Â±zÃƒÂ„Ã‚Â± ÃƒÂƒÃ‚Â§erÃƒÂƒÃ‚Â§eve iÃƒÂƒÃ‚Â§in padding
        border_width = 3  # 3px kÃƒÂ„Ã‚Â±rmÃƒÂ„Ã‚Â±zÃƒÂ„Ã‚Â± ÃƒÂƒÃ‚Â§erÃƒÂƒÃ‚Â§eve
        
        # Yeni gÃƒÂƒÃ‚Â¶rsel oluÃƒÂ…Ã…Â¸tur (logo + QR + text alanÃƒÂ„Ã‚Â± + ÃƒÂƒÃ‚Â§erÃƒÂƒÃ‚Â§eve)
        final_width = qr_width + (border_width * 2)
        final_height = logo_height + qr_height + text_height + (border_width * 2)
        final_img = Image.new('RGB', (final_width, final_height), '#dc2626')  # KÃƒÂ„Ã‚Â±rmÃƒÂ„Ã‚Â±zÃƒÂ„Ã‚Â± arka plan (ÃƒÂƒÃ‚Â§erÃƒÂƒÃ‚Â§eve)
        
        # Beyaz iÃƒÂƒÃ‚Â§ alan oluÃƒÂ…Ã…Â¸tur (logo + QR + text)
        white_bg = Image.new('RGB', (qr_width, logo_height + qr_height + text_height), 'white')
        
        # Logo ekle (varsa) - ÃƒÂƒÃ‚Â¼st ortasÃƒÂ„Ã‚Â±na
        try:
            logo_path = os.path.join(os.path.dirname(__file__), 'cermak-logo.png')
            if os.path.exists(logo_path):
                logo_img = Image.open(logo_path).convert('RGBA')
                # Logo boyutunu alan yÃƒÂƒÃ‚Â¼ksekliÃƒÂ„Ã…Â¸ine gÃƒÂƒÃ‚Â¶re ayarla
                logo_width = 150
                logo_height_logo = 40
                try:
                    logo_img = logo_img.resize((logo_width, logo_height_logo), Image.Resampling.LANCZOS)
                except AttributeError:
                    logo_img = logo_img.resize((logo_width, logo_height_logo), Image.LANCZOS)
                
                # Logo'yu ortala
                logo_x = (qr_width - logo_width) // 2
                logo_y = 5  # ÃƒÂƒÃ…Â“stten 5px boÃƒÂ…Ã…Â¸luk
                
                # RGBA logo'yu blend et
                if logo_img.mode == 'RGBA':
                    alpha = logo_img.split()[3]
                    logo_img = logo_img.convert('RGB')
                    white_bg.paste(logo_img, (logo_x, logo_y), alpha)
                else:
                    white_bg.paste(logo_img, (logo_x, logo_y))
        except Exception as e:
            print(f"Logo ekleme hatasÃƒÂ„Ã‚Â±: {e}")
        
        # QR kodu beyaz alana yapÃƒÂ„Ã‚Â±ÃƒÂ…Ã…Â¸tÃƒÂ„Ã‚Â±r (logo'nun altÃƒÂ„Ã‚Â±na)
        white_bg.paste(qr_img, (0, logo_height))
        
        # Beyaz alanÃƒÂ„Ã‚Â± kÃƒÂ„Ã‚Â±rmÃƒÂ„Ã‚Â±zÃƒÂ„Ã‚Â± ÃƒÂƒÃ‚Â§erÃƒÂƒÃ‚Â§evenin iÃƒÂƒÃ‚Â§ine yapÃƒÂ„Ã‚Â±ÃƒÂ…Ã…Â¸tÃƒÂ„Ã‚Â±r
        final_img.paste(white_bg, (border_width, border_width))
        
        # Text ekleme iÃƒÂƒÃ‚Â§in draw nesnesi
        draw = ImageDraw.Draw(final_img)
        
        # Font (kalÃƒÂ„Ã‚Â±n ve bÃƒÂƒÃ‚Â¼yÃƒÂƒÃ‚Â¼k)
        try:
            font = ImageFont.truetype("arialblk.ttf", 24)
        except:
            try:
                font = ImageFont.truetype("arialbd.ttf", 24)
            except:
                try:
                    font = ImageFont.truetype("arial.ttf", 24)
                except:
                    font = ImageFont.load_default()
        
        # QR ID yazÃƒÂ„Ã‚Â±sÃƒÂ„Ã‚Â± - Sadece bu
        qr_text = qr_id
        
        # Text geniÃƒÂ…Ã…Â¸liÃƒÂ„Ã…Â¸ini hesapla (24pt font iÃƒÂƒÃ‚Â§in)
        text_width = len(qr_text) * 14
        
        # QR ID'yi ortala ve yaz (border_width offset ekle)
        x_position = max(border_width, (final_width - text_width) // 2)
        draw.text((x_position, logo_height + qr_height + border_width + 5), qr_text, fill='black', font=font)

        buf = BytesIO()
        final_img.save(buf, format='PNG', optimize=True)
        buf.seek(0)

        # Cache'e kaydet
        img_data = buf.getvalue()
        cache_set(cache_key, img_data)

        #  LOKAL: static klasÃƒÂƒÃ‚Â¶rÃƒÂƒÃ‚Â¼ne kaydet
        qr_dir = os.path.join('static', 'qrcodes')
        os.makedirs(qr_dir, exist_ok=True)
        with open(static_path, 'wb') as f:
            f.write(img_data)
        
        # DosyayÃƒÂ„Ã‚Â± read-only yap
        os.chmod(static_path, 0o444)

        # DosyayÃƒÂ„Ã‚Â± dÃƒÂƒÃ‚Â¶ndÃƒÂƒÃ‚Â¼r
        buf.seek(0)
        return send_file(buf, mimetype='image/png')

    except Exception as e:
        logging.error(f"Error generating QR image for {qr_id}: {e}")
        # Hata durumunda basit QR oluÃƒÂ…Ã…Â¸tur
        qr = qrcode.QRCode(version=1, box_size=4, border=1)
        qr.add_data(qr_id)
        qr.make(fit=True)
        img = qr.make_image(fill_color="black", back_color="white")

        buf = BytesIO()
        img.save(buf, format='PNG')
        buf.seek(0)

        return send_file(buf, mimetype='image/png')

@app.route('/download_single_qr/<qr_id>')
@login_required
def download_single_qr(qr_id):
    conn = get_db()
    cursor = conn.cursor()
    execute_query(cursor, '''
        SELECT pc.part_code 
        FROM qr_codes qc
        JOIN part_codes pc ON qc.part_code_id = pc.id
        WHERE qc.qr_id = ?
    ''', (qr_id,))
    result = cursor.fetchone()

    if not result:
        close_db(conn)
        return jsonify({'error': 'QR kod bulunamadÃƒÂ„Ã‚Â±'}), 404

    # is_downloaded kolonu yok, sadece blob_url kontrolÃƒÂƒÃ‚Â¼ yeterli
    close_db(conn)

    try:
        if USE_B2_STORAGE and get_b2_service:
            # PRODUCTION: B2'den QR kod'u indir (KALICI)
            b2_service = get_b2_service()
            file_path = f'qr_codes/{qr_id}.png'

            file_content = b2_service.download_file(file_path)

            if file_content:
                # B2'den var olan dosyayÃƒÂ„Ã‚Â± dÃƒÂƒÃ‚Â¶ndÃƒÂƒÃ‚Â¼r
                buf = BytesIO(file_content)
                buf.seek(0)
                return send_file(buf, mimetype='image/png', as_attachment=True, download_name=f'{qr_id}.png')
        else:
            # LOCAL: Static dosyadan kontrol et (GEÃƒÂƒÃ¢Â€Â¡ÃƒÂ„Ã‚Â°CÃƒÂ„Ã‚Â°)
            static_path = os.path.join('static', 'qrcodes', f'{qr_id}.png')
            if os.path.exists(static_path):
                return send_file(static_path, mimetype='image/png', as_attachment=True, download_name=f'{qr_id}.png')

        # QR kod yoksa oluÃƒÂ…Ã…Â¸tur
        qr = qrcode.QRCode(version=1, box_size=10, border=4)
        qr.add_data(qr_id)
        qr.make(fit=True)
        img = qr.make_image(fill_color="black", back_color="white")

        buf = BytesIO()
        img.save(buf, format='PNG')
        buf.seek(0)

        img_data = buf.getvalue()

        if USE_B2_STORAGE and get_b2_service:
            # PRODUCTION: B2'ye yÃƒÂƒÃ‚Â¼kle (KALICI)
            b2_service = get_b2_service()
            file_path = f'qr_codes/{qr_id}.png'
            upload_result = b2_service.upload_file(file_path, img_data, 'image/png')


            if upload_result['success']:
                logging.info(f"QR code uploaded to B2: {file_path}")
        else:
            # LOCAL: Static klasÃƒÂƒÃ‚Â¶rÃƒÂƒÃ‚Â¼ne kaydet (GEÃƒÂƒÃ¢Â€Â¡ÃƒÂ„Ã‚Â°CÃƒÂ„Ã‚Â°)
            qr_dir = os.path.join('static', 'qrcodes')
            os.makedirs(qr_dir, exist_ok=True)
            local_path = os.path.join(qr_dir, f'{qr_id}.png')
            with open(local_path, 'wb') as f:
                f.write(img_data)

        # DosyayÃƒÂ„Ã‚Â± dÃƒÂƒÃ‚Â¶ndÃƒÂƒÃ‚Â¼r
        buf.seek(0)
        return send_file(buf, mimetype='image/png', as_attachment=True, download_name=f'{qr_id}.png')

    except Exception as e:
        logging.error(f"Error downloading QR image for {qr_id}: {e}")
        # Hata durumunda geleneksel yÃƒÂƒÃ‚Â¶ntemle oluÃƒÂ…Ã…Â¸tur
        qr = qrcode.QRCode(version=1, box_size=10, border=4)
        qr.add_data(qr_id)
        qr.make(fit=True)
        img = qr.make_image(fill_color="black", back_color="white")

        buf = BytesIO()
        img.save(buf, format='PNG')
        buf.seek(0)

        return send_file(buf, mimetype='image/png', as_attachment=True, download_name=f'{qr_id}.png')

#  ULTRA QR SCANNER HANDLER
@socketio.on('scan_qr_radical')
def handle_scan_radical(data):
    """ Ultra reliable QR scanning with enhanced features"""
    print("\n" + ""*50)
    print("ULTRA QR SCAN RECEIVED")
    print(""*50)

    try:
        qr_id = data.get('qr_id', '').strip()
        session_id = int(data.get('session_id', 1))  # INT olarak al
        user_id = session.get('user_id', 1)

        logging.debug(f" QR ID: {qr_id}, Session: {session_id}, User: {user_id}")

        if not qr_id:
            logging.warning("QR ID missing in scan request")
            emit('scan_result', {'success': False, 'message': 'ÃƒÂ¢Ã‚Â�Ã…Â’ QR ID eksik!'})
            return

        with db_connection() as conn:
            cursor = conn.cursor()

            # Check QR exists (JOIN ile part bilgisi al)
            execute_query(cursor, '''
                SELECT pc.part_code, pc.part_name 
                FROM qr_codes qc
                JOIN part_codes pc ON qc.part_code_id = pc.id
                WHERE qc.qr_id = ?
            ''', (qr_id,))
            qr_data = cursor.fetchone()

            if not qr_data:
                logging.warning(f"QR not found: {qr_id}")
                emit('scan_result', {'success': False, 'message': f'ÃƒÂ¢Ã‚Â�Ã…Â’ QR kod bulunamadÃƒÂ„Ã‚Â±: {qr_id}'})
                return

            part_code, part_name = qr_data
            logging.info(f"QR found: {part_code} - {part_name}")

            # Ensure session exists
            execute_query(cursor, 'SELECT COUNT(*) FROM count_sessions WHERE id = ?', (session_id,))
            if cursor.fetchone()[0] == 0:
                logging.info(f"Creating new session {session_id}")
                execute_query(cursor, 
                    'INSERT INTO count_sessions (id, session_name, is_active, created_at) VALUES (?, ?, ?, ?)',
                    (session_id, f'Session_{session_id}', 1, datetime.now()))
                conn.commit()

            # ÃƒÂ¢Ã…Â¡Ã‚Â¡ COMPOSITE INDEX kullanÃƒÂ„Ã‚Â±r - ÃƒÂƒÃ‚Â§ok hÃƒÂ„Ã‚Â±zlÃƒÂ„Ã‚Â± + AKILLI DUPLICATE
            execute_query(cursor, '''
                SELECT scanned_by, scanned_at 
                FROM scanned_qr 
                WHERE session_id = ? AND qr_id = ?
            ''', (session_id, qr_id))
            duplicate_record = cursor.fetchone()
            
            if duplicate_record:
                # Duplicate bulundu - AkÃƒÂ„Ã‚Â±llÃƒÂ„Ã‚Â± mesaj oluÃƒÂ…Ã…Â¸tur
                prev_user_id, prev_scanned_at = duplicate_record
                
                # ÃƒÂƒÃ¢Â€Â“nceki kullanÃƒÂ„Ã‚Â±cÃƒÂ„Ã‚Â± bilgisi
                execute_query(cursor, 'SELECT full_name FROM envanter_users WHERE id = ?', (prev_user_id,))
                prev_user_result = cursor.fetchone()
                prev_user_name = prev_user_result[0] if prev_user_result else 'Bilinmeyen'
                
                # Zaman farkÃƒÂ„Ã‚Â± hesapla
                prev_time = datetime.fromisoformat(prev_scanned_at) if isinstance(prev_scanned_at, str) else prev_scanned_at
                time_diff = datetime.now() - prev_time
                
                # Zaman formatÃƒÂ„Ã‚Â±
                if time_diff.total_seconds() < 60:
                    time_str = f"{int(time_diff.total_seconds())} saniye ÃƒÂƒÃ‚Â¶nce"
                    is_suspicious = time_diff.total_seconds() < 30  # 30 saniyeden kÃƒÂ„Ã‚Â±sa ise ÃƒÂ…Ã…Â¸ÃƒÂƒÃ‚Â¼pheli
                elif time_diff.total_seconds() < 3600:
                    time_str = f"{int(time_diff.total_seconds() / 60)} dakika ÃƒÂƒÃ‚Â¶nce"
                    is_suspicious = False
                else:
                    time_str = f"{int(time_diff.total_seconds() / 3600)} saat ÃƒÂƒÃ‚Â¶nce"
                    is_suspicious = False
                
                # AynÃƒÂ„Ã‚Â± kullanÃƒÂ„Ã‚Â±cÃƒÂ„Ã‚Â± mÃƒÂ„Ã‚Â± kontrol et
                is_same_user = (prev_user_id == user_id)
                
                # AkÃƒÂ„Ã‚Â±llÃƒÂ„Ã‚Â± mesaj oluÃƒÂ…Ã…Â¸tur
                if is_suspicious:
                    duplicate_msg = f"ÃƒÂ¢Ã…Â¡Ã‚Â ÃƒÂ¯Ã‚Â¸Ã‚Â� ÃƒÂ…Ã‚ÂžÃƒÂƒÃ…Â“PHELÃƒÂ„Ã‚Â°! {part_name} {time_str} tarandÃƒÂ„Ã‚Â± ({prev_user_name})"
                elif is_same_user:
                    duplicate_msg = f"ÃƒÂ¢Ã…Â¡Ã‚Â ÃƒÂ¯Ã‚Â¸Ã‚Â� {part_name} zaten taradÃƒÂ„Ã‚Â±nÃƒÂ„Ã‚Â±z! ({time_str})"
                else:
                    duplicate_msg = f"ÃƒÂ¢Ã…Â¡Ã‚Â ÃƒÂ¯Ã‚Â¸Ã‚Â� {part_name} zaten sayÃƒÂ„Ã‚Â±ldÃƒÂ„Ã‚Â±! {prev_user_name} tarafÃƒÂ„Ã‚Â±ndan {time_str}"
                
                logging.warning(f"Duplicate scan: {qr_id} - {duplicate_msg}")
                emit('scan_result', {
                    'success': False, 
                    'message': duplicate_msg,
                    'duplicate': True,
                    'previous_user': prev_user_name,
                    'time_ago': time_str,
                    'is_suspicious': is_suspicious,
                    'same_user': is_same_user
                }, broadcast=True)
                return

            # ÃƒÂ¢Ã…Â¡Ã‚Â¡ TRANSACTION - Atomik iÃƒÂ…Ã…Â¸lem
            cursor.execute('BEGIN TRANSACTION')
            try:
                # Insert scan record
                execute_query(cursor, 
                    'INSERT INTO scanned_qr (session_id, qr_id, part_code, scanned_by, scanned_at) VALUES (?, ?, ?, ?, ?)',
                    (session_id, qr_id, part_code, user_id, datetime.now()))

                # Mark QR as used
                execute_query(cursor, 'UPDATE qr_codes SET is_used = ?, used_at = ? WHERE qr_id = ?',
                             (1, datetime.now(), qr_id))

                # Update session stats
                execute_query(cursor, '''
                    UPDATE count_sessions 
                    SET total_scanned = (SELECT COUNT(*) FROM scanned_qr WHERE session_id = ?)
                    WHERE id = ?
                ''', (session_id, session_id))

                conn.commit()
                logging.info(f"SUCCESS: {part_name} scanned")
            except Exception as e:
                conn.rollback()
                logging.error(f"Transaction failed: {e}")
                raise

        # Get user name (ayrÃƒÂ„Ã‚Â± connection)
        with db_connection() as conn2:
            cursor2 = conn2.cursor()
            execute_query(cursor2, 'SELECT full_name FROM envanter_users WHERE id = ?', (user_id,))
            user_result = cursor2.fetchone()
            user_name = user_result[0] if user_result else 'KullanÃƒÂ„Ã‚Â±cÃƒÂ„Ã‚Â±'

        #  TRIPLE BROADCAST
        success_data = {
            'success': True,
            'message': f'ÃƒÂ¢Ã…Â“Ã¢Â€Â¦ {part_name} sayÃƒÂ„Ã‚Â±ldÃƒÂ„Ã‚Â±!',
            'qr_code': qr_id,
            'part_code': part_code,
            'part_name': part_name,
            'session_id': session_id,
            'scanned_by': user_name,
            'scanned_at': datetime.now().strftime('%H:%M:%S')
        }

        # Total scan count (session stats'dan al)
        with db_connection() as conn3:
            cursor3 = conn3.cursor()
            execute_query(cursor3, 'SELECT total_scanned FROM count_sessions WHERE id = ?', (session_id,))
            result = cursor3.fetchone()
            success_data['total_scans'] = result[0] if result else 0

        socketio.emit('scan_result', success_data, broadcast=True)
        socketio.emit('qr_scanned', success_data, broadcast=True)
        socketio.emit('activity_update', success_data, broadcast=True)

        logging.info(f"ULTRA SUCCESS - {part_name} scanned by {user_name}")

    except sqlite3.Error as e:
        logging.error(f"Database error in scan: {e}", exc_info=True)
        emit('scan_result', {'success': False, 'message': f'ÃƒÂ¢Ã‚Â�Ã…Â’ VeritabanÃƒÂ„Ã‚Â± hatasÃƒÂ„Ã‚Â±: {e}'})
    except ValueError as e:
        logging.error(f"Value error in scan: {e}", exc_info=True)
        emit('scan_result', {'success': False, 'message': f'ÃƒÂ¢Ã‚Â�Ã…Â’ GeÃƒÂƒÃ‚Â§ersiz veri: {e}'})
    except Exception as e:
        logging.error(f"Unexpected error in scan: {e}", exc_info=True)
        emit('scan_result', {'success': False, 'message': f'ÃƒÂ¢Ã‚Â�Ã…Â’ Sistem hatasÃƒÂ„Ã‚Â±: {e}'})

#  ULTRA MODERN API ENDPOINTS
@app.route('/api/scan_qr', methods=['POST'])
def api_scan_qr_ultra():
    """ Ultra reliable QR scanning API endpoint with modern tech"""
    try:
        data = request.get_json()
        if not data:
            return jsonify({'success': False, 'message': 'No data provided'}), 400

        qr_id = data.get('qr_id', '').strip()
        session_id = data.get('session_id', '1')

        if not qr_id:
            return jsonify({'success': False, 'message': 'QR ID required'}), 400

        #  SCANNER FIX: BazÃƒÂ„Ã‚Â± QR scanner cihazlarÃƒÂ„Ã‚Â± - ve _ karakterlerini yanlÃƒÂ„Ã‚Â±ÃƒÂ…Ã…Â¸ okuyor
        # * (ASCII 42) -> - (ASCII 45) dÃƒÂƒÃ‚Â¶nÃƒÂƒÃ‚Â¼ÃƒÂ…Ã…Â¸tÃƒÂƒÃ‚Â¼r
        # ? (ASCII 63) -> _ (ASCII 95) dÃƒÂƒÃ‚Â¶nÃƒÂƒÃ‚Â¼ÃƒÂ…Ã…Â¸tÃƒÂƒÃ‚Â¼r
        qr_id_original = qr_id
        qr_id = qr_id.replace('*', '-').replace('?', '_')
        
        if qr_id != qr_id_original:
            print(f" QR FIX: '{qr_id_original}' -> '{qr_id}'")

        print(f" ULTRA API QR Scan: {qr_id} in session {session_id}")

        # Process the scan with ultra reliability
        result = process_qr_scan_ultra(qr_id, session_id)

        # Ultra broadcast - emit to all clients for real-time updates
        socketio.emit('scan_result', result)
        socketio.emit('qr_scanned', result)
        socketio.emit('activity_update', result)

        return jsonify(result)

    except Exception as e:
        print(f"ÃƒÂ¢Ã‚Â�Ã…Â’ ULTRA API scan error: {e}")
        return jsonify({'success': False, 'message': str(e)}), 500

def process_qr_scan_ultra(qr_id, session_id):
    """ Ultra reliable QR processing with enhanced features"""
    conn = None
    try:
        conn = get_db()
        cursor = conn.cursor()

        # Ultra session management - ensure session exists with better naming
        session_name = f"Tarama SeansÃƒÂ„Ã‚Â± {session_id}"

        # Defensive: check if session row exists, then insert using DB-agnostic columns
        # count_sessions kontrolÃƒÂƒÃ‚Â¼ artÃƒÂ„Ã‚Â±k gereksiz - session ID yoksa oluÃƒÂ…Ã…Â¸turma
        # Session ID artÃƒÂ„Ã‚Â±k auto-increment, manuel oluÃƒÂ…Ã…Â¸turmaya gerek yok
        
        # ========================================
        #  PAKET KONTROLÃƒÂƒÃ…Â“ - ÃƒÂƒÃ¢Â€Â“nce paketin olup olmadÃƒÂ„Ã‚Â±ÃƒÂ„Ã…Â¸ÃƒÂ„Ã‚Â±nÃƒÂ„Ã‚Â± kontrol et
        # ========================================
        cursor.execute('''
            SELECT is_package, package_items, part_name 
            FROM part_codes 
            WHERE part_code = ?
        ''', (qr_id,))
        
        package_check = cursor.fetchone()
        
        # EÃƒÂ„Ã…Â¸er bu bir paketse, iÃƒÂƒÃ‚Â§indeki tÃƒÂƒÃ‚Â¼m parÃƒÂƒÃ‚Â§alarÃƒÂ„Ã‚Â± tek tek tara
        if package_check and package_check[0]:  # is_package = True
            try:
                package_items = json.loads(package_check[1]) if package_check[1] else []
                package_name = package_check[2]
                
                if not package_items:
                    return {
                        'success': False,
                        'message': f'ÃƒÂ¢Ã…Â¡Ã‚Â ÃƒÂ¯Ã‚Â¸Ã‚Â� {package_name} paketi boÃƒÂ…Ã…Â¸!',
                        'item_name': package_name
                    }
                
                # Paket taramasÃƒÂ„Ã‚Â± duplicate kontrolÃƒÂƒÃ‚Â¼
                cursor.execute('''
                    SELECT COUNT(*) FROM scanned_qr 
                    WHERE qr_id = ? AND session_id = ?
                ''', (qr_id, str(session_id)))
                
                if cursor.fetchone()[0] > 0:
                    return {
                        'success': False,
                        'message': f'ÃƒÂ¢Ã…Â¡Ã‚Â ÃƒÂ¯Ã‚Â¸Ã‚Â� {package_name} paketi zaten tarandÃƒÂ„Ã‚Â±!',
                        'item_name': package_name,
                        'duplicate': True
                    }
                
                # ÃƒÂ¢Ã…Â¡Ã‚Â ÃƒÂ¯Ã‚Â¸Ã‚Â� PAKET KENDISINI TARA DEÃƒÂ„Ã‚ÂžÃƒÂ„Ã‚Â°L - SADECE ÃƒÂ„Ã‚Â°ÃƒÂƒÃ¢Â€Â¡ERÃƒÂ„Ã‚Â°DEKÃƒÂ„Ã‚Â° PARÃƒÂƒÃ¢Â€Â¡ALARI TARA
                # Paket, sadece tracker olarak kaydediliyor, toplam sayÃƒÂ„Ã‚Â±ya EKLENMÃƒÂ„Ã‚Â°YOR
                
                # Paket iÃƒÂƒÃ‚Â§indeki her parÃƒÂƒÃ‚Â§ayÃƒÂ„Ã‚Â± otomatik tara
                total_items = 0
                for idx, item in enumerate(package_items):
                    part_code = item.get('part_code')
                    quantity = item.get('quantity', 1)
                    
                    # Her bir parÃƒÂƒÃ‚Â§a iÃƒÂƒÃ‚Â§in quantity kadar tarama kaydÃƒÂ„Ã‚Â± oluÃƒÂ…Ã…Â¸tur
                    for qty_idx in range(quantity):
                        virtual_qr = f"{qr_id}_PKG_{part_code}_{idx}_{qty_idx}"
                        cursor.execute('''
                            INSERT INTO scanned_qr (qr_id, session_id, part_code, scanned_by, scanned_at)
                            VALUES (?, ?, ?, ?, ?)
                        ''', (virtual_qr, str(session_id), part_code, session.get('user_id', 1), datetime.now()))
                        total_items += 1
                
                # ÃƒÂ„Ã‚Â°statistikleri gÃƒÂƒÃ‚Â¼ncelle (SADECE ÃƒÂ„Ã‚Â°ÃƒÂƒÃ¢Â€Â¡ERÃƒÂ„Ã‚Â°DEKÃƒÂ„Ã‚Â° PARÃƒÂƒÃ¢Â€Â¡ALARI SAY, PAKETÃƒÂ„Ã‚Â° SAYMA!)
                cursor.execute('''
                    SELECT COUNT(*) as total_scans
                    FROM scanned_qr 
                    WHERE session_id = ?
                ''', (str(session_id),))
                
                total_scans = cursor.fetchone()[0]
                
                try:
                    cursor.execute('UPDATE count_sessions SET total_scanned = ? WHERE id = ?', 
                                 (total_scans, str(session_id)))
                except Exception as e:
                    print(f"ÃƒÂ¢Ã…Â¡Ã‚Â ÃƒÂ¯Ã‚Â¸Ã‚Â� Package total_scanned gÃƒÂƒÃ‚Â¼ncelleme hatasÃƒÂ„Ã‚Â±: {e}")
                    pass
                
                conn.commit()
                
                app.logger.info(f'[PAKET TARAMA] {package_name} - {total_items} parÃƒÂƒÃ‚Â§a otomatik tarandÃƒÂ„Ã‚Â±')
                
                return {
                    'success': True,
                    'message': f'ÃƒÂ¢Ã…Â“Ã¢Â€Â¦  {package_name}\n{total_items} parÃƒÂƒÃ‚Â§a otomatik sayÃƒÂ„Ã‚Â±ldÃƒÂ„Ã‚Â±!',
                    'item_name': package_name,
                    'total_scans': total_scans,
                    'is_package': True,
                    'package_items': total_items
                }
                
            except Exception as pkg_err:
                app.logger.error(f'[PAKET HATASI] {qr_id}: {pkg_err}')
                # Paket hatasÃƒÂ„Ã‚Â± olursa normal QR gibi iÃƒÂ…Ã…Â¸le
                pass
        
        # ========================================
        #  NORMAL QR ÃƒÂ„Ã‚Â°ÃƒÂ…Ã‚ÂžLEME - Standart parÃƒÂƒÃ‚Â§a tarama
        # ========================================
        
        # Check if QR exists with enhanced data retrieval
        qr_data = None
        try:
            # ÃƒÂƒÃ¢Â€Â“nce tam eÃƒÂ…Ã…Â¸leÃƒÂ…Ã…Â¸me dene
            execute_query(cursor, """
                SELECT qc.qr_id, pc.part_code, pc.part_name, qc.is_used, qc.created_at
                FROM qr_codes qc
                LEFT JOIN part_codes pc ON qc.part_code_id = pc.id
                WHERE qc.qr_id = ?
            """, (qr_id,))
            qr_data = cursor.fetchone()
            
            # Hala bulunamadÃƒÂ„Ã‚Â±ysa, part_code olarak ara (QR = part_code durumu)
            if not qr_data:
                execute_query(cursor, """
                    SELECT pc.part_code, pc.part_code, pc.part_name, 0, ?
                    FROM part_codes pc
                    WHERE pc.part_code = ?
                """, (datetime.now(), qr_id))
                qr_data = cursor.fetchone()
                
        except Exception as e:
            # Schema mismatch or missing column in older DBs - fall back to defaults
            print(f"ÃƒÂ¢Ã…Â¡Ã‚Â ÃƒÂ¯Ã‚Â¸Ã‚Â� QR lookup failed (schema mismatch?): {e}")
            qr_data = None

        if not qr_data:
            # Do NOT attempt to modify the schema here. Use a safe fallback QR record in-memory.
            unknown_name = f"Bilinmeyen ÃƒÂƒÃ…Â“rÃƒÂƒÃ‚Â¼n ({qr_id[:15]})"
            print(f"ÃƒÂ¢Ã‚Â�Ã…Â’ QR BULUNAMADI: {qr_id}")  # Debug log
            qr_data = (qr_id, qr_id[:10], unknown_name, False, datetime.now())

        qr_id_db, part_code, part_name, is_used, created_at = qr_data

        #  KALICI DUPLICATE KONTROLÃƒÂƒÃ…Â“ - Bir session'da bir QR sadece 1 kez okunabilir
        execute_query(cursor, """
            SELECT COUNT(*) FROM scanned_qr 
            WHERE qr_id = ? AND session_id = ?
        """, (qr_id, str(session_id)))

        existing_count = cursor.fetchone()[0]
        if existing_count > 0:
            # Bu QR zaten bu session'da taranmÃƒÂ„Ã‚Â±ÃƒÂ…Ã…Â¸ - asla tekrar taranmamalÃƒÂ„Ã‚Â±
            return {
                'success': False,
                'message': f'ÃƒÂ¢Ã…Â¡Ã‚Â ÃƒÂ¯Ã‚Â¸Ã‚Â� {part_name} zaten tarandÃƒÂ„Ã‚Â±!',
                'item_name': part_name,
                'duplicate': True
            }

        # Insert ultra scan record with enhanced data
        try:
            execute_query(cursor, """
                INSERT INTO scanned_qr (qr_id, session_id, part_code, scanned_by, scanned_at)
                VALUES (?, ?, ?, ?, ?)
            """, (qr_id, str(session_id), part_code, session.get('user_id', 1), datetime.now()))
        except Exception:
            # Fallback for older schemas without part_code column
            try:
                execute_query(cursor, """
                    INSERT INTO scanned_qr (qr_id, session_id, scanned_by, scanned_at)
                    VALUES (?, ?, ?, ?)
                """, (qr_id, str(session_id), session.get('user_id', 1), datetime.now()))
            except Exception as ie:
                print(f"ÃƒÂ¢Ã‚Â�Ã…Â’ Failed to insert scanned_qr record: {ie}")
                raise

        # Mark QR as used with timestamp (DB-agnostic) - ignore if schema doesn't match
        try:
            execute_query(cursor, """
                UPDATE qr_codes 
                SET is_used = ?, used_at = ? 
                WHERE qr_id = ?
            """, (True, datetime.now(), qr_id))
        except Exception:
            # Not critical if the QR table lacks these columns in older schemas
            pass

        # Get enhanced session statistics
        execute_query(cursor, """
            SELECT 
                COUNT(*) as total_scans,
                COUNT(DISTINCT qr_id) as unique_items
            FROM scanned_qr 
            WHERE session_id = ?
        """, (str(session_id),))

        stats = cursor.fetchone()
        total_scans = stats[0] if stats else 0
        unique_items = stats[1] if stats else 0

        # Update count_sessions.total_scanned for dashboard
        try:
            execute_query(cursor, 'UPDATE count_sessions SET total_scanned = ? WHERE id = ?', (total_scans, str(session_id)))
        except Exception as e:
            print(f"ÃƒÂ¢Ã…Â¡Ã‚Â ÃƒÂ¯Ã‚Â¸Ã‚Â� total_scanned gÃƒÂƒÃ‚Â¼ncelleme hatasÃƒÂ„Ã‚Â±: {e}")
            pass

        conn.commit()

        # Get user info for enhanced feedback
        user_id = session.get('user_id', 1)
        execute_query(cursor, 'SELECT full_name FROM envanter_users WHERE id = ?', (user_id,))
        user_result = cursor.fetchone()
        user_name = user_result[0] if user_result else 'KullanÃƒÂ„Ã‚Â±cÃƒÂ„Ã‚Â±'

        # Ultra success response
        return {
            'success': True,
            'message': f'ÃƒÂ¢Ã…Â“Ã¢Â€Â¦ {part_name} baÃƒÂ…Ã…Â¸arÃƒÂ„Ã‚Â±yla tarandÃƒÂ„Ã‚Â±! (#{total_scans})',
            'item_name': part_name,
            'part_code': part_code,
            'total_scans': total_scans,
            'unique_items': unique_items,
            'qr_id': qr_id,
            'session_id': session_id,
            'scanned_by': user_name,
            'scan_time': datetime.now().isoformat(),
            'was_used_before': is_used
        }

    except Exception as e:
        if conn:
            conn.rollback()
        print(f"ÃƒÂ¢Ã‚Â�Ã…Â’ ULTRA Process QR error: {e}")
        return {
            'success': False,
            'message': f'ÃƒÂ¢Ã‚Â�Ã…Â’ Sistem hatasÃƒÂ„Ã‚Â±: {str(e)}',
            'error_detail': str(e),
            'qr_id': qr_id
        }
    finally:
        if conn:
            close_db(conn)

@app.route('/api/session/<session_id>/stats', methods=['GET'])
def get_ultra_session_stats(session_id):
    """ Get ultra detailed session statistics"""
    try:
        conn = get_db()
        cursor = conn.cursor()

        # Get session info
        execute_query(cursor, """
            SELECT 
                CASE 
                    WHEN is_active::integer = 1 THEN 'active'
                    WHEN is_active::integer = 0 THEN 'completed'
                    ELSE 'unknown'
                END as status, 
                started_at, 
                ended_at 
            FROM count_sessions 
            WHERE session_id = ?
        """, (str(session_id),))

        session_info = cursor.fetchone()

        if not session_info:
            return jsonify({'error': 'Session not found'}), 404

        # Get ultra scan statistics
        execute_query(cursor, """
            SELECT 
                COUNT(*) as total_scans,
                COUNT(DISTINCT qr_id) as unique_items,
                MIN(scanned_at) as first_scan,
                MAX(scanned_at) as last_scan
            FROM scanned_qr 
            WHERE session_id = ?
        """, (str(session_id),))

        stats = cursor.fetchone()

        # Get recent scans with enhanced data
        execute_query(cursor, """
            SELECT sq.qr_id, qc.part_name, sq.scanned_at, u.full_name
            FROM scanned_qr sq
            LEFT JOIN qr_codes qc ON sq.qr_id = qc.qr_id
            LEFT JOIN envanter_users u ON sq.scanned_by = u.id
            WHERE sq.session_id = ?
            ORDER BY sq.scanned_at DESC
            LIMIT 20
        """, (str(session_id),))

        recent_scans = cursor.fetchall()

        # Calculate session duration
        start_time = session_info[1]
        end_time = session_info[2] or (stats[3] if stats[3] else datetime.now())
        duration_seconds = (end_time - start_time).total_seconds() if start_time else 0

        return jsonify({
            'session_id': session_id,
            'status': session_info[0],
            'started_at': start_time.isoformat() if start_time else None,
            'ended_at': session_info[2].isoformat() if session_info[2] else None,
            'duration_seconds': duration_seconds,
            'total_scans': stats[0] or 0,
            'unique_items': stats[1] or 0,
            'first_scan': stats[2].isoformat() if stats[2] else None,
            'last_scan': stats[3].isoformat() if stats[3] else None,
            'scans_per_minute': round((stats[0] or 0) / max(duration_seconds / 60, 1), 2),
            'recent_scans': [
                {
                    'qr_id': scan[0],
                    'item_name': scan[1] or f'Unknown ({scan[0]})',
                    'scan_time': scan[2].isoformat() if scan[2] else None,
                    'scanned_by': scan[3] or 'Bilinmeyen'
                }
                for scan in recent_scans
            ]
        })

    except Exception as e:
        print(f"ÃƒÂ¢Ã‚Â�Ã…Â’ ULTRA Stats error: {e}")
        return jsonify({'error': str(e)}), 500
    finally:
        if conn:
            close_db(conn)

# ========================================
# ÃƒÂ¯Ã‚Â¿Ã‚Â½ÃƒÂ¯Ã‚Â¸Ã‚Â� TELEFON QR & PARÃƒÂƒÃ¢Â€Â¡A BÃƒÂ„Ã‚Â°LGÃƒÂ„Ã‚Â°LERÃƒÂ„Ã‚Â° SÃƒÂ„Ã‚Â°STEMÃƒÂ„Ã‚Â° KALDIRILDI
# ArtÃƒÂ„Ã‚Â±k telefon sayÃƒÂ„Ã‚Â±mÃƒÂ„Ã‚Â± YOK - sadece desktop QR scanner kullanÃƒÂ„Ã‚Â±lÃƒÂ„Ã‚Â±yor
# KaldÃƒÂ„Ã‚Â±rÃƒÂ„Ã‚Â±lan ÃƒÂƒÃ‚Â¶zellikler:
# - /qr-info (mobil QR bilgi sayfasÃƒÂ„Ã‚Â±)
# - /api/qr_info/<qr_id> (QR bilgi API)
# - /api/part_details/<qr_id> (ParÃƒÂƒÃ‚Â§a detaylarÃƒÂ„Ã‚Â±)
# - /api/update_part_details (ParÃƒÂƒÃ‚Â§a gÃƒÂƒÃ‚Â¼ncelleme)
# - /api/upload_part_photo (FotoÃƒÂ„Ã…Â¸raf yÃƒÂƒÃ‚Â¼kleme)
# - /api/qr/<qr_id>/info (DetaylÃƒÂ„Ã‚Â± QR bilgisi)
# ========================================

#  RAPOR YÃƒÂƒÃ¢Â€Â“NETÃƒÂ„Ã‚Â°MÃƒÂ„Ã‚Â° (Admin Count - ÃƒÂ…Ã‚Âžifre sistemi kaldÃƒÂ„Ã‚Â±rÃƒÂ„Ã‚Â±ldÃƒÂ„Ã‚Â±)
    try:
        #  ULTRA SECURITY: Sadece admin kullanÃƒÂ„Ã‚Â±cÃƒÂ„Ã‚Â±sÃƒÂ„Ã‚Â± sayÃƒÂ„Ã‚Â±mÃƒÂ„Ã‚Â± bitirebilir
        current_user_id = session.get('user_id')

        if not current_user_id:
            return jsonify({'success': False, 'error': 'Oturum bulunamadÃƒÂ„Ã‚Â± - LÃƒÂƒÃ‚Â¼tfen tekrar giriÃƒÂ…Ã…Â¸ yapÃƒÂ„Ã‚Â±n'}), 401

        # Admin kontrolÃƒÂƒÃ‚Â¼ - kullanÃƒÂ„Ã‚Â±cÃƒÂ„Ã‚Â± bilgilerini al
        conn = get_db()
        cursor = conn.cursor()

        execute_query(cursor, "SELECT username, role FROM envanter_users WHERE id = ?", (current_user_id,))
        user_result = cursor.fetchone()

        if not user_result:
            close_db(conn)
            return jsonify({'success': False, 'error': 'KullanÃƒÂ„Ã‚Â±cÃƒÂ„Ã‚Â± bulunamadÃƒÂ„Ã‚Â±'}), 404

        username, user_role = user_result

        # DEBUG LOG
        print(f"[FINISH_COUNT DEBUG] User: {username}, Role: {user_role}, ID: {current_user_id}")
        print(f"[FINISH_COUNT DEBUG] Role type: {type(user_role)}, Value: [{user_role}]")
        print(f"[FINISH_COUNT DEBUG] Role == 'admin': {user_role == 'admin'}")
        print(f"[FINISH_COUNT DEBUG] Role.lower() == 'admin': {user_role.lower() == 'admin' if user_role else 'NULL'}")

        #  SADECE ADMIN YETKÃƒÂ„Ã‚Â°SÃƒÂ„Ã‚Â° - Role tabanlÃƒÂ„Ã‚Â± kontrol (daha gÃƒÂƒÃ‚Â¼venli)
        if not user_role or user_role.lower() != 'admin':
            close_db(conn)
            security_logger.warning(f'USER {username} (Role: {user_role}, ID: {current_user_id}) tried to finish count session - PERMISSION DENIED')
            return jsonify({
                'success': False, 
                'error': f'YETKISIZ ERÃƒÂ„Ã‚Â°ÃƒÂ…Ã‚ÂžÃƒÂ„Ã‚Â°M: Sadece admin yetkisine sahip kullanÃƒÂ„Ã‚Â±cÃƒÂ„Ã‚Â±lar sayÃƒÂ„Ã‚Â±mÃƒÂ„Ã‚Â± bitirebilir (Your role: {user_role})',
                'permission_required': 'admin',
                'current_role': user_role
            }), 403

        # SayÃƒÂ„Ã‚Â±m eriÃƒÂ…Ã…Â¸im kontrolÃƒÂƒÃ‚Â¼ (secondary check) - ARTIK GEREKLÃƒÂ„Ã‚Â° DEÃƒÂ„Ã‚ÂžÃƒÂ„Ã‚Â°L
        # if not session.get('count_access'):
        #     close_db(conn)
        #     return jsonify({'success': False, 'error': 'SayÃƒÂ„Ã‚Â±m eriÃƒÂ…Ã…Â¸imi iÃƒÂƒÃ‚Â§in ÃƒÂ…Ã…Â¸ifre gerekli'}), 403

        # Aktif sayÃƒÂ„Ã‚Â±m oturumunu kontrol et
        execute_query(cursor, "SELECT id, is_active, created_by FROM count_sessions WHERE is_active = ? LIMIT 1", (True,))
        session_result = cursor.fetchone()

        if not session_result:
            close_db(conn)
            return jsonify({'success': False, 'error': 'Aktif sayÃƒÂ„Ã‚Â±m oturumu bulunamadÃƒÂ„Ã‚Â±'}), 400

        session_id, is_active_status, created_by = session_result

        # Log admin action
        security_logger.info(f'ADMIN {username} finishing count session {session_id}')

        # ÃƒÂƒÃ¢Â€Â¡ift iÃƒÂ…Ã…Â¸lem kontrolÃƒÂƒÃ‚Â¼ - eÃƒÂ„Ã…Â¸er bu oturum zaten tamamlandÃƒÂ„Ã‚Â±ysa
        if not is_active_status:
            close_db(conn)
            return jsonify({'success': False, 'error': 'Bu sayÃƒÂ„Ã‚Â±m oturumu zaten tamamlanmÃƒÂ„Ã‚Â±ÃƒÂ…Ã…Â¸'}), 400

        # SayÃƒÂ„Ã‚Â±m oturumunu sonlandÃƒÂ„Ã‚Â±r (admin yetkisiyle)
        execute_query(cursor, "UPDATE count_sessions SET is_active = ?, ended_at = ? WHERE id = ?",
                     (False, datetime.now(), session_id))

        # Rapor verilerini topla - BEKLENEN ADETLERLE KARÃƒÂ…Ã‚ÂžILAÃƒÂ…Ã‚ÂžTIR
        # Her QR tek bir part_code'a ait, o yÃƒÂƒÃ‚Â¼zden JOIN ile part_code ÃƒÂƒÃ‚Â§ekiyoruz
        execute_query(cursor, '''
            SELECT 
                COALESCE(pc.part_code, sq.part_code) as part_code,
                COALESCE(pc.part_name, 'Bilinmeyen ParÃƒÂƒÃ‚Â§a') as part_name,
                COUNT(*) as sayilan_adet
            FROM scanned_qr sq
            LEFT JOIN qr_codes qc ON sq.qr_id = qc.qr_id
            LEFT JOIN part_codes pc ON qc.part_code_id = pc.id
            WHERE sq.session_id = ?
            GROUP BY COALESCE(pc.part_code, sq.part_code), COALESCE(pc.part_name, 'Bilinmeyen ParÃƒÂƒÃ‚Â§a')
            ORDER BY part_code
        ''', (session_id,))

        scanned_parts = {}  # {part_code: (part_name, sayilan_adet)}
        for row in cursor.fetchall():
            part_code = row[0]
            part_name = row[1]
            sayilan_adet = row[2]
            scanned_parts[part_code] = (part_name, sayilan_adet)
            print(f"[RAPOR DEBUG] {part_code}: {part_name} - {sayilan_adet} adet okundu")

        # TÃƒÂƒÃ‚Â¼m part_codes'dan beklenen adetleri al (yÃƒÂƒÃ‚Â¼klenmiÃƒÂ…Ã…Â¸ Excel'den)
        execute_query(cursor, '''
            SELECT 
                pc.part_code,
                pc.part_name,
                COUNT(qc.qr_id) as beklenen_adet
            FROM part_codes pc
            LEFT JOIN qr_codes qc ON qc.part_code_id = pc.id
            GROUP BY pc.part_code, pc.part_name
        ''')

        # Rapor verilerini hazÃƒÂ„Ã‚Â±rla - TÃƒÂƒÃ…Â“M PARÃƒÂƒÃ¢Â€Â¡ALAR (okutulan + okutulmayan)
        report_data = []
        total_scanned = 0
        total_expected = 0

        for row in cursor.fetchall():
            part_code = row[0]
            part_name = row[1]
            beklenen_adet = row[2] or 0
            
            # Bu parÃƒÂƒÃ‚Â§a okutulan parÃƒÂƒÃ‚Â§alar arasÃƒÂ„Ã‚Â±nda var mÃƒÂ„Ã‚Â±?
            if part_code in scanned_parts:
                sayilan_adet = scanned_parts[part_code][1]
            else:
                sayilan_adet = 0  # HiÃƒÂƒÃ‚Â§ okutulmamÃƒÂ„Ã‚Â±ÃƒÂ…Ã…Â¸
            
            fark = sayilan_adet - beklenen_adet
            
            part_data = {
                'ParÃƒÂƒÃ‚Â§a Kodu': part_code or 'Bilinmiyor',
                'ParÃƒÂƒÃ‚Â§a AdÃƒÂ„Ã‚Â±': part_name,
                'Beklenen Adet': beklenen_adet,
                'SayÃƒÂ„Ã‚Â±lan Adet': sayilan_adet,
                'Fark': fark,
                'Durum': 'ÃƒÂ¢Ã…Â“Ã¢Â€Âœ Tamam' if fark == 0 else ('ÃƒÂ¢Ã…Â¡Ã‚Â  Eksik' if fark < 0 else '+ Fazla')
            }
            report_data.append(part_data)
            total_scanned += sayilan_adet
            total_expected += beklenen_adet

        # Excel raporu oluÃƒÂ…Ã…Â¸tur
        df = pd.DataFrame(report_data)

        output = BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='SayÃƒÂ„Ã‚Â±m Raporu')
            
            # Worksheet'i al ve formatla
            workbook = writer.book
            worksheet = writer.sheets['SayÃƒÂ„Ã‚Â±m Raporu']
            
            # SÃƒÂƒÃ‚Â¼tun geniÃƒÂ…Ã…Â¸liklerini ayarla
            worksheet.column_dimensions['A'].width = 15  # ParÃƒÂƒÃ‚Â§a Kodu
            worksheet.column_dimensions['B'].width = 30  # ParÃƒÂƒÃ‚Â§a AdÃƒÂ„Ã‚Â±
            worksheet.column_dimensions['C'].width = 15  # Beklenen Adet
            worksheet.column_dimensions['D'].width = 15  # SayÃƒÂ„Ã‚Â±lan Adet
            worksheet.column_dimensions['E'].width = 10  # Fark
            worksheet.column_dimensions['F'].width = 12  # Durum
            
            # Header stilini ayarla
            from openpyxl.styles import Font, PatternFill, Alignment
            header_fill = PatternFill(start_color='DC3545', end_color='DC3545', fill_type='solid')
            header_font = Font(bold=True, color='FFFFFF')
            
            for cell in worksheet[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
        
        output.seek(0)

        # Rapor dosyasÃƒÂ„Ã‚Â±nÃƒÂ„Ã‚Â± kaydet
        report_filename = f'sayim_raporu_{session_id}_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        report_path = os.path.join(REPORTS_DIR, report_filename)

        with open(report_path, 'wb') as f:
            f.write(output.getvalue())

        # DoÃƒÂ„Ã…Â¸ruluk oranÃƒÂ„Ã‚Â±nÃƒÂ„Ã‚Â± hesapla
        accuracy_rate = (total_scanned / total_expected * 100) if total_expected > 0 else 0.0

        # Raporu count_reports table'ÃƒÂ„Ã‚Â±na kaydet (varsa)
        report_title = f"SayÃƒÂ„Ã‚Â±m Raporu - {datetime.now().strftime('%d.%m.%Y %H:%M')}"
        try:
            execute_query(cursor, '''
                INSERT INTO count_reports (session_id, report_filename, report_title, 
                                         total_expected, total_scanned, accuracy_rate, created_at)
                VALUES (?, ?, ?, ?, ?, ?, ?)
            ''', (session_id, report_filename, report_title, total_expected, total_scanned, accuracy_rate, datetime.now()))
        except Exception as e:
            # count_reports tablosu yoksa sadece dosya kaydet
            logging.warning(f"count_reports tablosu yok, sadece dosya kaydedildi: {e}")

        # Database iÃƒÂ…Ã…Â¸lemini commit et
        conn.commit()
        close_db(conn)

        # WebSocket ile sayÃƒÂ„Ã‚Â±m bittiÃƒÂ„Ã…Â¸i bilgisini gÃƒÂƒÃ‚Â¶nder
        try:
            socketio.emit('count_finished', {'session_id': session_id})
        except Exception as ws_error:
            logging.warning(f"WebSocket notification failed: {str(ws_error)}")

        # Session'dan sayÃƒÂ„Ã‚Â±m bilgilerini temizle
        session.pop('count_access', None)
        session.pop('current_session', None)

        return jsonify({
            'success': True,
            'message': 'SayÃƒÂ„Ã‚Â±m baÃƒÂ…Ã…Â¸arÃƒÂ„Ã‚Â±yla tamamlandÃƒÂ„Ã‚Â±',
            'report_file': report_filename,
            'session_id': session_id,
            'total_expected': total_expected,
            'total_scanned': total_scanned,
            'accuracy_rate': round(accuracy_rate, 2)
        })

    except Exception as e:
        # Hata durumunda database baÃƒÂ„Ã…Â¸lantÃƒÂ„Ã‚Â±sÃƒÂ„Ã‚Â±nÃƒÂ„Ã‚Â± kapatmayÃƒÂ„Ã‚Â± garanti et
        try:
            if 'conn' in locals():
                close_db(conn)
        except:
            pass

        error_msg = f"SayÃƒÂ„Ã‚Â±m tamamlama hatasÃƒÂ„Ã‚Â±: {str(e)}"
        logging.error(error_msg, exc_info=True)

        return jsonify({
            'success': False,
            'error': 'SayÃƒÂ„Ã‚Â±m tamamlanamadÃƒÂ„Ã‚Â± - sistem hatasÃƒÂ„Ã‚Â±',
            'debug': str(e) if app.debug else None
        }), 500

@app.route('/stop_all_counts', methods=['POST'])
def stop_all_counts():
    """TÃƒÂƒÃ‚Â¼m aktif sayÃƒÂ„Ã‚Â±mlarÃƒÂ„Ã‚Â± durdur - ACIL DURUMFONKSÃƒÂ„Ã‚Â°YONU"""
    # Admin authentication check
    admin_password = request.json.get('admin_password')
    if admin_password != ADMIN_COUNT_PASSWORD:
        return jsonify({'success': False, 'error': 'Yetki gerekli - yanlÃƒÂ„Ã‚Â±ÃƒÂ…Ã…Â¸ admin ÃƒÂ…Ã…Â¸ifresi'}), 403

    conn = get_db()
    cursor = conn.cursor()

    try:
        # TÃƒÂƒÃ‚Â¼m aktif sayÃƒÂ„Ã‚Â±mlarÃƒÂ„Ã‚Â± bul
        execute_query(cursor, "SELECT id FROM count_sessions WHERE is_active = TRUE")
        active_sessions = cursor.fetchall()

        if not active_sessions:
            close_db(conn)
            return jsonify({'success': True, 'message': 'Durdurulacak aktif sayÃƒÂ„Ã‚Â±m bulunamadÃƒÂ„Ã‚Â±'})

        # TÃƒÂƒÃ‚Â¼m aktif sayÃƒÂ„Ã‚Â±mlarÃƒÂ„Ã‚Â± "completed" olarak iÃƒÂ…Ã…Â¸aretle
        stopped_count = 0
        for session_tuple in active_sessions:
            session_id = session_tuple[0]
            execute_query(cursor, 'UPDATE count_sessions SET is_active = ?, ended_at = ? WHERE id = ?',
                         (False, datetime.now(), session_id))
            stopped_count += 1

        # Session'larÃƒÂ„Ã‚Â± temizle
        session.pop('count_access', None)
        session.pop('count_authenticated', None) 
        session.pop('current_session', None)

        conn.commit()
        close_db(conn)

        # WebSocket ile tÃƒÂƒÃ‚Â¼m kullanÃƒÂ„Ã‚Â±cÃƒÂ„Ã‚Â±lara sayÃƒÂ„Ã‚Â±mlarÃƒÂ„Ã‚Â±n durdurulduÃƒÂ„Ã…Â¸unu bildir
        socketio.emit('all_counts_stopped', {
            'message': f'{stopped_count} aktif sayÃƒÂ„Ã‚Â±m durduruldu',
            'stopped_sessions': [s[0] for s in active_sessions]
        })

        return jsonify({
            'success': True,
            'message': f'{stopped_count} aktif sayÃƒÂ„Ã‚Â±m baÃƒÂ…Ã…Â¸arÃƒÂ„Ã‚Â±yla durduruldu',
            'stopped_sessions': [s[0] for s in active_sessions]
        })

    except Exception as e:
        conn.rollback()
        close_db(conn)
        return jsonify({'success': False, 'error': f'Sistem hatasÃƒÂ„Ã‚Â±: {str(e)}'}), 500

@app.route('/qr_codes')
def qr_codes_page():
    if 'user_id' not in session:
        return render_template('login.html')
    return render_template('qr_codes.html')

@app.route('/parts')
@login_required
def parts_list():
    """TÃƒÂƒÃ‚Â¼m parÃƒÂƒÃ‚Â§alarÃƒÂ„Ã‚Â± listele"""
    conn = get_db()
    cursor = conn.cursor()
    
    # ParÃƒÂƒÃ‚Â§alarÃƒÂ„Ã‚Â± ve QR kod sayÃƒÂ„Ã‚Â±larÃƒÂ„Ã‚Â±nÃƒÂ„Ã‚Â± getir
    execute_query(cursor, '''
        SELECT 
            pc.id,
            pc.part_code, 
            pc.part_name,
            pc.description,
            pc.created_at,
            COUNT(qc.qr_id) as qr_count
        FROM part_codes pc
        LEFT JOIN qr_codes qc ON pc.id = qc.part_code_id
        GROUP BY pc.id, pc.part_code, pc.part_name, pc.description, pc.created_at
        ORDER BY pc.created_at DESC
    ''')
    
    parts = []
    for row in cursor.fetchall():
        parts.append({
            'id': row[0],
            'part_code': row[1],
            'part_name': row[2],
            'description': row[3] or '',
            'created_at': row[4],
            'qr_count': row[5]
        })
    
    close_db(conn)
    return render_template('parts.html', parts=parts)

@app.route('/parts/<part_code>')
@login_required
def part_detail(part_code):
    """ParÃƒÂƒÃ‚Â§a detay sayfasÃƒÂ„Ã‚Â± - QR kod ÃƒÂƒÃ‚Â¼retme"""
    conn = get_db()
    cursor = conn.cursor()
    placeholder = get_db_placeholder()
    
    # ParÃƒÂƒÃ‚Â§a bilgilerini getir
    execute_query(cursor, f'SELECT id, part_code, part_name, description, created_at FROM part_codes WHERE part_code = {placeholder}', (part_code,))
    part_row = cursor.fetchone()
    
    if not part_row:
        close_db(conn)
        return "ParÃƒÂƒÃ‚Â§a bulunamadÃƒÂ„Ã‚Â±", 404
    
    part = {
        'id': part_row[0],
        'part_code': part_row[1],
        'part_name': part_row[2],
        'description': part_row[3] or '',
        'created_at': part_row[4]
    }
    
    # Bu parÃƒÂƒÃ‚Â§aya ait QR kodlarÃƒÂ„Ã‚Â±nÃƒÂ„Ã‚Â± getir (SADECE KULLANILMAYANLAR)
    execute_query(cursor, f'''
        SELECT qr_id, created_at, is_used,
               CASE WHEN blob_url IS NOT NULL THEN 1 ELSE 0 END as is_downloaded
        FROM qr_codes 
        WHERE part_code_id = {placeholder} AND is_used = 0
        ORDER BY qr_id DESC
    ''', (part['id'],))
    
    qr_codes = []
    for row in cursor.fetchall():
        qr_codes.append({
            'qr_id': row[0],
            'created_at': row[1],
            'is_used': row[2],
            'is_downloaded': row[3]
        })
    
    # Toplam QR sayÃƒÂ„Ã‚Â±sÃƒÂ„Ã‚Â±nÃƒÂ„Ã‚Â± al (kullanÃƒÂ„Ã‚Â±lan + kullanÃƒÂ„Ã‚Â±lmayan)
    execute_query(cursor, f'SELECT COUNT(*) FROM qr_codes WHERE part_code_id = {placeholder}', (part['id'],))
    total_qr_count = cursor.fetchone()[0]
    
    # KullanÃƒÂ„Ã‚Â±lan QR sayÃƒÂ„Ã‚Â±sÃƒÂ„Ã‚Â±nÃƒÂ„Ã‚Â± al
    execute_query(cursor, f'SELECT COUNT(*) FROM qr_codes WHERE part_code_id = {placeholder} AND is_used = 1', (part['id'],))
    used_qr_count = cursor.fetchone()[0]
    
    close_db(conn)
    return render_template('part_detail.html', part=part, qr_codes=qr_codes, total_qr_count=total_qr_count, used_qr_count=used_qr_count)

@app.route('/generate_qr/<part_code>', methods=['POST'])
@login_required
def generate_qr_codes(part_code):
    """Belirtilen parÃƒÂƒÃ‚Â§a iÃƒÂƒÃ‚Â§in birden fazla QR kod ÃƒÂƒÃ‚Â¼ret (quantity parametresi ile)"""
    try:
        req = request.get_json(silent=True) or {}
        quantity = int(req.get('quantity', 1) or 1)
        if quantity < 1:
            quantity = 1
        max_qty = 500
        if quantity > max_qty:
            return jsonify({'success': False, 'error': f'Quantity too large (max {max_qty})'}), 400

        conn = get_db()
        cursor = conn.cursor()
        placeholder = get_db_placeholder()

        # ParÃƒÂƒÃ‚Â§a bilgilerini al
        execute_query(cursor, f'SELECT id, part_name FROM part_codes WHERE part_code = {placeholder}', (part_code,))
        part_row = cursor.fetchone()

        if not part_row:
            close_db(conn)
            return jsonify({'error': 'ParÃƒÂƒÃ‚Â§a bulunamadÃƒÂ„Ã‚Â±'}), 404

        part_code_id = part_row[0]
        part_name = part_row[1]

        # Mevcut QR kod sayÃƒÂ„Ã‚Â±sÃƒÂ„Ã‚Â±nÃƒÂ„Ã‚Â± ÃƒÂƒÃ‚Â¶ÃƒÂ„Ã…Â¸ren (tÃƒÂƒÃ‚Â¼m QR'lar, kullanÃƒÂ„Ã‚Â±lmÃƒÂ„Ã‚Â±ÃƒÂ…Ã…Â¸ olanlar dahil)
        execute_query(cursor, f'SELECT COUNT(*) FROM qr_codes WHERE part_code_id = {placeholder}', (part_code_id,))
        current_count = cursor.fetchone()[0]

        generated = []
        file_paths = []

        for i in range(quantity):
            qr_number = current_count + i + 1
            qr_id = f"{part_code}_{qr_number}"

            execute_query(cursor, f'INSERT INTO qr_codes (qr_id, part_code_id, created_at, is_used) VALUES ({placeholder}, {placeholder}, {placeholder}, {placeholder})',
                         (qr_id, part_code_id, datetime.now(), False))

            save_qr_code_to_file(part_code, qr_id, qr_number)

            generated.append({
                'qr_id': qr_id,
                'qr_number': qr_number,
                'qr_image_url': f'/qr_image/{part_code}/{qr_id}'
            })

            qr_file_path = os.path.join(os.path.dirname(__file__), 'static', 'qr_codes', part_code, f"{qr_id}.png")
            file_paths.append(qr_file_path)

        conn.commit()
        close_db(conn)

        zip_url = None
        try:
            if len(file_paths) > 1:
                zip_dir = os.path.join(os.path.dirname(__file__), 'static', 'qr_codes', part_code)
                os.makedirs(zip_dir, exist_ok=True)
                zip_name = f'bulk_{datetime.now().strftime("%Y%m%d_%H%M%S")}.zip'
                zip_path = os.path.join(zip_dir, zip_name)
                import zipfile
                with zipfile.ZipFile(zip_path, 'w', compression=zipfile.ZIP_DEFLATED) as zf:
                    for fp in file_paths:
                        if os.path.exists(fp):
                            zf.write(fp, arcname=os.path.basename(fp))
                zip_url = f'/static/qr_codes/{part_code}/{zip_name}'
        except Exception:
            logging.exception('ZIP oluÃƒÂ…Ã…Â¸turulamadÃƒÂ„Ã‚Â±')

        print(f"ÃƒÂ¢Ã…Â“Ã¢Â€Â¦ {part_code} iÃƒÂƒÃ‚Â§in {quantity} QR kod ÃƒÂƒÃ‚Â¼retildi. BaÃƒÂ…Ã…Â¸langÃƒÂ„Ã‚Â±ÃƒÂƒÃ‚Â§: {current_count + 1}")

        return jsonify({
            'success': True,
            'message': f'{quantity} adet QR kod ÃƒÂƒÃ‚Â¼retildi',
            'generated': generated,
            'zip_url': zip_url
        })

    except Exception as e:
        logging.exception(f"QR kod ÃƒÂƒÃ‚Â¼retme hatasÃƒÂ„Ã‚Â±: {e}")
        try:
            close_db(conn)
        except:
            pass
        return jsonify({'error': f'Hata: {str(e)}'}), 500

@app.route('/qr_image/<part_code>/<qr_id>')
@login_required
def serve_qr_image(part_code, qr_id):
    """QR kod gÃƒÂƒÃ‚Â¶rselini serve et (text ile birlikte)"""
    try:
        # ÃƒÂƒÃ¢Â€Â“nce statik dosyayÃƒÂ„Ã‚Â± kontrol et
        qr_file = os.path.join(os.path.dirname(__file__), 'static', 'qr_codes', part_code, f"{qr_id}.png")
        
        if os.path.exists(qr_file):
            return send_file(qr_file, mimetype='image/png')
        
        # Dosya yoksa dinamik oluÃƒÂ…Ã…Â¸tur (generate_qr_image gibi)
        # QR ID'den parÃƒÂƒÃ‚Â§a kodunu ve numarayÃƒÂ„Ã‚Â± ÃƒÂƒÃ‚Â§ÃƒÂ„Ã‚Â±kar
        parts = qr_id.rsplit('_', 1)
        qr_number = parts[1] if len(parts) > 1 else "1"
        
        # ParÃƒÂƒÃ‚Â§a adÃƒÂ„Ã‚Â±nÃƒÂ„Ã‚Â± database'den al
        conn = get_db()
        cursor = conn.cursor()
        execute_query(cursor, 'SELECT part_name FROM part_codes WHERE part_code = ?', (part_code,))
        result = cursor.fetchone()
        part_name = result[0] if result else ""
        close_db(conn)

        # QR kod oluÃƒÂ…Ã…Â¸tur - Barkod makinesi iÃƒÂƒÃ‚Â§in optimize
        qr = qrcode.QRCode(
            version=1, 
            box_size=8,  # 8px - barkod makinesi iÃƒÂƒÃ‚Â§in ideal
            border=2,    # 2px quiet zone
            error_correction=qrcode.constants.ERROR_CORRECT_M
        )
        qr.add_data(qr_id)
        qr.make(fit=True)
        qr_img = qr.make_image(fill_color="black", back_color="white")
        
        # PIL Image'a dÃƒÂƒÃ‚Â¶nÃƒÂƒÃ‚Â¼ÃƒÂ…Ã…Â¸tÃƒÂƒÃ‚Â¼r
        qr_img = qr_img.convert('RGB')

        # QR boyutlarÃƒÂ„Ã‚Â±
        qr_width, qr_height = qr_img.size
        
        # AlanlarÃƒÂ„Ã‚Â± hesapla
        logo_height = 40  # Logo iÃƒÂƒÃ‚Â§in ÃƒÂƒÃ‚Â¼st alan
        text_height = 35  # 1 satÃƒÂ„Ã‚Â±r - sadece QR ID
        
        # KÃƒÂ„Ã‚Â±rmÃƒÂ„Ã‚Â±zÃƒÂ„Ã‚Â± ÃƒÂƒÃ‚Â§erÃƒÂƒÃ‚Â§eve iÃƒÂƒÃ‚Â§in padding
        border_width = 3  # 3px kÃƒÂ„Ã‚Â±rmÃƒÂ„Ã‚Â±zÃƒÂ„Ã‚Â± ÃƒÂƒÃ‚Â§erÃƒÂƒÃ‚Â§eve
        
        # Final gÃƒÂƒÃ‚Â¶rsel (ÃƒÂƒÃ‚Â§erÃƒÂƒÃ‚Â§eveli)
        final_width = qr_width + (border_width * 2)
        final_height = logo_height + qr_height + text_height + (border_width * 2)
        final_img = Image.new('RGB', (final_width, final_height), '#dc2626')  # KÃƒÂ„Ã‚Â±rmÃƒÂ„Ã‚Â±zÃƒÂ„Ã‚Â± arka plan (ÃƒÂƒÃ‚Â§erÃƒÂƒÃ‚Â§eve)
        
        # Beyaz iÃƒÂƒÃ‚Â§ alan oluÃƒÂ…Ã…Â¸tur (logo + QR + text)
        white_bg = Image.new('RGB', (qr_width, logo_height + qr_height + text_height), 'white')
        
        # Logo ekle (varsa) - ÃƒÂƒÃ‚Â¼st ortasÃƒÂ„Ã‚Â±na
        try:
            logo_path = os.path.join(os.path.dirname(__file__), 'cermak-logo.png')
            if os.path.exists(logo_path):
                logo_img = Image.open(logo_path).convert('RGBA')
                # Logo boyutunu alan yÃƒÂƒÃ‚Â¼ksekliÃƒÂ„Ã…Â¸ine gÃƒÂƒÃ‚Â¶re ayarla
                logo_width = 150
                logo_height_logo = 40
                try:
                    logo_img = logo_img.resize((logo_width, logo_height_logo), Image.Resampling.LANCZOS)
                except AttributeError:
                    logo_img = logo_img.resize((logo_width, logo_height_logo), Image.LANCZOS)
                
                # Logo'yu ortala
                logo_x = (qr_width - logo_width) // 2
                logo_y = 5  # ÃƒÂƒÃ…Â“stten 5px boÃƒÂ…Ã…Â¸luk
                
                # RGBA logo'yu blend et
                if logo_img.mode == 'RGBA':
                    alpha = logo_img.split()[3]
                    logo_img = logo_img.convert('RGB')
                    white_bg.paste(logo_img, (logo_x, logo_y), alpha)
                else:
                    white_bg.paste(logo_img, (logo_x, logo_y))
        except Exception as e:
            print(f"Logo ekleme hatasÃƒÂ„Ã‚Â±: {e}")
        
        # QR kodu beyaz alana yapÃƒÂ„Ã‚Â±ÃƒÂ…Ã…Â¸tÃƒÂ„Ã‚Â±r (logo'nun altÃƒÂ„Ã‚Â±na)
        white_bg.paste(qr_img, (0, logo_height))
        
        # Beyaz alanÃƒÂ„Ã‚Â± kÃƒÂ„Ã‚Â±rmÃƒÂ„Ã‚Â±zÃƒÂ„Ã‚Â± ÃƒÂƒÃ‚Â§erÃƒÂƒÃ‚Â§evenin iÃƒÂƒÃ‚Â§ine yapÃƒÂ„Ã‚Â±ÃƒÂ…Ã…Â¸tÃƒÂ„Ã‚Â±r
        final_img.paste(white_bg, (border_width, border_width))
        
        # Text ekle
        draw = ImageDraw.Draw(final_img)
        try:
            font = ImageFont.truetype("arialblk.ttf", 24)
        except:
            try:
                font = ImageFont.truetype("arialbd.ttf", 24)
            except:
                try:
                    font = ImageFont.truetype("arial.ttf", 24)
                except:
                    font = ImageFont.load_default()
        
        # Sadece QR ID
        qr_text = qr_id
        text_width = len(qr_text) * 14
        x_position = max(border_width, (final_width - text_width) // 2)
        draw.text((x_position, logo_height + qr_height + border_width + 5), qr_text, fill='black', font=font)

        # BytesIO'ya kaydet
        buf = BytesIO()
        final_img.save(buf, format='PNG', optimize=True)
        buf.seek(0)
        
        return send_file(buf, mimetype='image/png')
        
    except Exception as e:
        logging.exception(f"QR gÃƒÂƒÃ‚Â¶rsel hatasÃƒÂ„Ã‚Â±: {e}")
        return f"Hata: {str(e)}", 500

@app.route('/mark_qr_used/<qr_id>', methods=['POST'])
@login_required
def mark_qr_used(qr_id):
    """QR kodu manuel olarak 'kullanÃƒÂ„Ã‚Â±ldÃƒÂ„Ã‚Â±' iÃƒÂ…Ã…Â¸aretle"""
    try:
        conn = get_db()
        cursor = conn.cursor()
        placeholder = get_db_placeholder()
        
        # QR kodunu kontrol et
        execute_query(cursor, f'SELECT qr_id, is_used FROM qr_codes WHERE qr_id = {placeholder}', (qr_id,))
        qr = cursor.fetchone()
        
        if not qr:
            close_db(conn)
            return jsonify({'error': 'QR kod bulunamadÃƒÂ„Ã‚Â±'}), 404
        
        if qr[1]:  # is_used
            close_db(conn)
            return jsonify({'error': 'Bu QR kod zaten kullanÃƒÂ„Ã‚Â±lmÃƒÂ„Ã‚Â±ÃƒÂ…Ã…Â¸'}), 400
        
        # is_used = True yap
        execute_query(cursor, f'UPDATE qr_codes SET is_used = {placeholder} WHERE qr_id = {placeholder}',
                     (True, qr_id))
        conn.commit()
        close_db(conn)
        
        print(f"ÃƒÂ¢Ã…Â“Ã¢Â€Â¦ QR kod manuel kullanÃƒÂ„Ã‚Â±ldÃƒÂ„Ã‚Â± iÃƒÂ…Ã…Â¸aretlendi: {qr_id}")
        
        return jsonify({
            'success': True,
            'message': f'QR kod kullanÃƒÂ„Ã‚Â±ldÃƒÂ„Ã‚Â± olarak iÃƒÂ…Ã…Â¸aretlendi: {qr_id}'
        })
        
    except Exception as e:
        logging.exception(f"QR iÃƒÂ…Ã…Â¸aretleme hatasÃƒÂ„Ã‚Â±: {e}")
        try:
            close_db(conn)
        except:
            pass
        return jsonify({'error': f'Hata: {str(e)}'}), 500

@app.route('/check_existing_qrs')
@login_required
def check_existing_qrs():
    conn = get_db()
    cursor = conn.cursor()
    execute_query(cursor, 'SELECT COUNT(*) FROM qr_codes')
    count = cursor.fetchone()[0]
    close_db(conn)

    return jsonify({
        'hasQRs': count > 0,
        'count': count
    })

@app.route('/qr_management', methods=['GET'])
@login_required
def qr_management():
    """QR YÃƒÂƒÃ‚Â¶netim Paneli - GÃƒÂƒÃ‚Â¼venli QR iÃƒÂ…Ã…Â¸lemleri"""
    if not session.get('admin_authenticated'):
        return redirect('/admin')
    return render_template('qr_management.html')

@app.route('/get_reports')
@login_required
def get_reports():
    conn = None
    try:
        # VeritabanÃƒÂ„Ã‚Â±ndan raporlarÃƒÂ„Ã‚Â± ÃƒÂƒÃ‚Â§ek
        conn = get_db()
        cursor = conn.cursor()
        
        execute_query(cursor, '''
            SELECT 
                cr.id,
                cr.session_id,
                cr.report_filename,
                cr.report_title,
                cr.total_expected,
                cr.total_scanned,
                cr.accuracy_rate,
                cr.created_at,
                u.username as created_by
            FROM count_reports cr
            LEFT JOIN count_sessions cs ON cr.session_id = cs.id
            LEFT JOIN envanter_users u ON cs.created_by = u.id
            ORDER BY cr.created_at DESC
        ''')
        
        rows = cursor.fetchall()
        reports = []
        
        for row in rows:
            total_difference = (row[5] or 0) - (row[4] or 0)  # scanned - expected
            
            reports.append({
                'id': row[0],
                'session_id': row[1],
                'filename': row[2],
                'title': row[3] or f"SayÃƒÂ„Ã‚Â±m Raporu #{row[1]}",
                'created_at': row[7],
                'total_expected': row[4] or 0,
                'total_scanned': row[5] or 0,
                'accuracy_rate': round(row[6] or 0.0, 2),
                'session_name': f"Oturum #{row[1]}",
                'created_by': row[8] or 'Sistem',
                'total_difference': total_difference
            })
        
        return jsonify(reports)

    except Exception as e:
        logging.exception(f"Error in get_reports: {e}")
        return jsonify({'error': str(e)}), 500
    finally:
        try:
            if conn:
                close_db(conn)
        except Exception:
            pass

@app.route('/download_report/<filename>')
@login_required
def download_report(filename):
    # ÃƒÂ¢Ã…Â“Ã¢Â€Â¦ GÃƒÂƒÃ‚Â¼venlik: Filename formatÃƒÂ„Ã‚Â±nÃƒÂ„Ã‚Â± kontrol et (session_id integer olabilir)
    # Format: sayim_raporu_{session_id}_{timestamp}.xlsx
    if not re.match(r'^sayim_raporu_\d+_\d{8}_\d{6}\.xlsx$', filename):
        print(f"ÃƒÂ¢Ã‚Â�Ã…Â’ Invalid filename format: {filename}")
        return jsonify({'error': 'GeÃƒÂƒÃ‚Â§ersiz dosya adÃƒÂ„Ã‚Â± formatÃƒÂ„Ã‚Â±'}), 400

    safe_filename = secure_filename(filename)
    report_path = os.path.join(REPORTS_DIR, safe_filename)

    if not os.path.exists(report_path):
        print(f"ÃƒÂ¢Ã‚Â�Ã…Â’ Report file not found: {report_path}")
        return jsonify({'error': 'Rapor dosyasÃƒÂ„Ã‚Â± bulunamadÃƒÂ„Ã‚Â±'}), 404

    # ÃƒÂ¢Ã…Â“Ã¢Â€Â¦ Path traversal attack'e karÃƒÂ…Ã…Â¸ÃƒÂ„Ã‚Â± gÃƒÂƒÃ‚Â¼venlik kontrolÃƒÂƒÃ‚Â¼
    real_path = os.path.realpath(report_path)
    reports_real_path = os.path.realpath(REPORTS_DIR)

    if not real_path.startswith(reports_real_path):
        print(f"ÃƒÂ¢Ã‚Â�Ã…Â’ Invalid path (security): {real_path}")
        return jsonify({'error': 'GeÃƒÂƒÃ‚Â§ersiz dosya yolu'}), 403

    print(f"ÃƒÂ¢Ã…Â“Ã¢Â€Â¦ Sending report file: {report_path}")
    return send_file(report_path, as_attachment=True, download_name=filename)

#  SAYIM YÃƒÂƒÃ¢Â€Â“NETÃƒÂ„Ã‚Â°MÃƒÂ„Ã‚Â° (Admin Count - ÃƒÂ…Ã‚Âžifre sistemi kaldÃƒÂ„Ã‚Â±rÃƒÂ„Ã‚Â±ldÃƒÂ„Ã‚Â±)
@app.route('/admin_count')
@login_required
@admin_required_decorator
def admin_count_page():
    """Admin sayÃƒÂ„Ã‚Â±m kontrol sayfasÃƒÂ„Ã‚Â± - Excel yÃƒÂƒÃ‚Â¼kleme ile sayÃƒÂ„Ã‚Â±m baÃƒÂ…Ã…Â¸latma"""
    return render_template('admin_count.html')

@app.route('/admin_count/start_count', methods=['POST'])
@login_required
@admin_required_decorator
def admin_start_count():
    """Admin sayÃƒÂ„Ã‚Â±m baÃƒÂ…Ã…Â¸latma endpoint'i"""
    print("DEBUG: admin_start_count ÃƒÂƒÃ‚Â§aÃƒÂ„Ã…Â¸rÃƒÂ„Ã‚Â±ldÃƒÂ„Ã‚Â±")
    
    try:
        conn = get_db()
        cursor = conn.cursor()
        
        # Aktif sayÃƒÂ„Ã‚Â±m var mÃƒÂ„Ã‚Â± kontrol et
        execute_query(cursor, 'SELECT COUNT(*) FROM count_sessions WHERE is_active = ?', (True,))
        active_count = cursor.fetchone()[0]
        
        if active_count > 0:
            close_db(conn)
            return jsonify({
                'success': False,
                'error': 'Zaten aktif bir sayÃƒÂ„Ã‚Â±m oturumu var!'
            }), 400
        
        # Yeni sayÃƒÂ„Ã‚Â±m oturumu oluÃƒÂ…Ã…Â¸tur (PAROLA YOK ARTIK)
        current_user_id = session.get('user_id')
        
        # Toplam beklenen adet
        execute_query(cursor, 'SELECT COUNT(*) FROM qr_codes')
        total_expected = cursor.fetchone()[0]
        
        execute_query(cursor, '''
            INSERT INTO count_sessions 
            (is_active, started_at, created_by, created_at, total_expected, total_scanned) 
            VALUES (?, ?, ?, ?, ?, ?)
        ''', (True, datetime.now(), current_user_id, datetime.now(), total_expected, 0))
        
        # ID'yi al (SQLite auto-increment)
        session_id = cursor.lastrowid
        
        conn.commit()
        close_db(conn)
        
        print(f"ÃƒÂ¢Ã…Â“Ã¢Â€Â¦ SayÃƒÂ„Ã‚Â±m oturumu baÃƒÂ…Ã…Â¸latÃƒÂ„Ã‚Â±ldÃƒÂ„Ã‚Â±: {session_id}")
        
        #  SOCKET.IO BÃƒÂ„Ã‚Â°LDÃƒÂ„Ã‚Â°RÃƒÂ„Ã‚Â°MÃƒÂ„Ã‚Â°: TÃƒÂƒÃ‚Â¼m clientlara yeni sayÃƒÂ„Ã‚Â±m baÃƒÂ…Ã…Â¸ladÃƒÂ„Ã‚Â±ÃƒÂ„Ã…Â¸ÃƒÂ„Ã‚Â±nÃƒÂ„Ã‚Â± sÃƒÂƒÃ‚Â¶yle
        try:
            socketio.emit('session_reset', {
                'session_id': session_id,
                'total_expected': total_expected,
                'message': ' Yeni sayÃƒÂ„Ã‚Â±m baÃƒÂ…Ã…Â¸latÃƒÂ„Ã‚Â±ldÃƒÂ„Ã‚Â± - sayfa sÃƒÂ„Ã‚Â±fÃƒÂ„Ã‚Â±rlanÃƒÂ„Ã‚Â±yor...'
            }, broadcast=True)
            print(f" Socket bildirimi gÃƒÂƒÃ‚Â¶nderildi: session_reset")
        except Exception as socket_err:
            print(f"ÃƒÂ¢Ã…Â¡Ã‚Â ÃƒÂ¯Ã‚Â¸Ã‚Â� Socket bildirimi gÃƒÂƒÃ‚Â¶nderilemedi: {socket_err}")
        
        return jsonify({
            'success': True,
            'message': 'SayÃƒÂ„Ã‚Â±m oturumu baÃƒÂ…Ã…Â¸latÃƒÂ„Ã‚Â±ldÃƒÂ„Ã‚Â±! PSC Scanner ile QR okutmaya baÃƒÂ…Ã…Â¸layabilirsiniz.',
            'session_id': session_id,
            'total_expected': total_expected,
            'redirect': '/count-scanner'  # PSC Scanner sayfasÃƒÂ„Ã‚Â±na yÃƒÂƒÃ‚Â¶nlendir
        })
        
    except Exception as e:
        print(f"ÃƒÂ¢Ã‚Â�Ã…Â’ SayÃƒÂ„Ã‚Â±m baÃƒÂ…Ã…Â¸latma hatasÃƒÂ„Ã‚Â±: {e}")
        import traceback
        traceback.print_exc()
        
        if conn:
            try:
                conn.rollback()
                close_db(conn)
            except:
                pass
        
        return jsonify({
            'success': False,
            'error': f'SayÃƒÂ„Ã‚Â±m baÃƒÂ…Ã…Â¸latÃƒÂ„Ã‚Â±lamadÃƒÂ„Ã‚Â±: {str(e)}'
        }), 500

# API Endpoints for Dashboard Statistics
@app.route('/api/qr_codes')
@login_required
def api_get_qr_codes():
    """QR kodlarÃƒÂ„Ã‚Â± listesi - istatistik iÃƒÂƒÃ‚Â§in"""
    conn = get_db()
    cursor = conn.cursor()

    execute_query(cursor, '''
        SELECT qc.qr_id, pc.part_code, pc.part_name, qc.is_used, 
               CASE WHEN qc.blob_url IS NOT NULL THEN 1 ELSE 0 END as is_downloaded,
               qc.created_at
        FROM qr_codes qc
        LEFT JOIN part_codes pc ON qc.part_code_id = pc.id
        ORDER BY qc.created_at DESC
    ''')

    qr_codes = []
    for row in cursor.fetchall():
        qr_codes.append({
            'qr_id': row[0],
            'part_code': row[1] or '',
            'part_name': row[2] or 'Bilinmeyen',
            'is_used': bool(row[3]),
            'is_downloaded': bool(row[4]),
            'created_at': row[5]
        })

    close_db(conn)
    return jsonify(qr_codes)

@app.route('/api/reports')
@login_required
def api_get_reports():
    """Raporlar listesi - istatistik iÃƒÂƒÃ‚Â§in"""
    conn = get_db()
    cursor = conn.cursor()

    execute_query(cursor, '''
        SELECT id, session_id, report_name, file_path, created_at, 
               total_expected, total_scanned, accuracy_rate
        FROM count_reports
        ORDER BY created_at DESC
    ''')

    reports = []
    for row in cursor.fetchall():
        reports.append({
            'id': row[0],
            'session_id': row[1],
            'report_name': row[2],
            'file_path': row[3],
            'created_at': row[4],
            'total_expected': row[5],
            'total_scanned': row[6],
            'accuracy_rate': row[7]
        })


    close_db(conn)
    return jsonify(reports)

@app.route('/api/dashboard_stats')
def api_dashboard_stats():

    """Dashboard iÃƒÂƒÃ‚Â§in genel istatistikler"""
    print("DEBUG: /api/dashboard_stats endpoint ÃƒÂƒÃ‚Â§aÃƒÂ„Ã…Â¸rÃƒÂ„Ã‚Â±ldÃƒÂ„Ã‚Â±")  # DEBUG
    conn = get_db()
    cursor = conn.cursor()




    # QR kodlarÃƒÂ„Ã‚Â± sayÃƒÂ„Ã‚Â±sÃƒÂ„Ã‚Â±
    execute_query(cursor, 'SELECT COUNT(*) FROM qr_codes')
    total_qr_codes = cursor.fetchone()[0]

    # Raporlar sayÃƒÂ„Ã‚Â±sÃƒÂ„Ã‚Â±
    execute_query(cursor, 'SELECT COUNT(*) FROM count_sessions')
    total_reports = cursor.fetchone()[0]

    # SayÃƒÂ„Ã‚Â±m bilgileri geÃƒÂƒÃ‚Â§ici olarak sÃƒÂ„Ã‚Â±fÃƒÂ„Ã‚Â±r
    active_counts = 0
    completed_counts = 0
    last_count_date = None

    close_db(conn)

    stats = {
        'total_qr_codes': total_qr_codes,
        'total_reports': total_reports,
        'active_counts': active_counts,
        'completed_counts': completed_counts,
        'last_count_date': last_count_date
    }
    print(f"DEBUG: GÃƒÂƒÃ‚Â¶nderilen stats: {stats}")  # DEBUG
    return jsonify(stats)

# Eksik endpoint'ler
@app.route('/get_session_stats')
@login_required
def get_session_stats():
    """SayÃƒÂ„Ã‚Â±m session istatistikleri"""
    try:
        # URL parametresinden session_id al
        requested_session_id = request.args.get('session_id')
        
        conn = get_db()
        cursor = conn.cursor()
        
        #  EÃƒÂ„Ã…Â¸er session_id verilmiÃƒÂ…Ã…Â¸se, o session'ÃƒÂ„Ã‚Â± kullan
        if requested_session_id:
            session_id = requested_session_id
            
            # Test mode iÃƒÂƒÃ‚Â§in expected=3
            expected = 3 if requested_session_id.startswith('test-') else 0
            
            # count_sessions tablosundan expected deÃƒÂ„Ã…Â¸erini al (varsa)
            try:
                execute_query(cursor, '''
                    SELECT total_expected
                    FROM count_sessions
                    WHERE id = ?
                ''', (session_id,))
                row = cursor.fetchone()
                if row and row[0]:
                    expected = row[0]
            except:
                pass
            
        else:
            # Session ID verilmemiÃƒÂ…Ã…Â¸se, aktif session bul
            execute_query(cursor, '''
                SELECT id, total_expected, total_scanned
                FROM count_sessions
                WHERE is_active = ?
                ORDER BY started_at DESC
                LIMIT 1
            ''', (True,))

            row = cursor.fetchone()

            if not row:
                close_db(conn)
                return jsonify({'success': False, 'message': 'No active session', 'scanned': 0, 'expected': 0, 'scanned_qrs': []})

            session_id = row[0]
            expected = row[1] if row[1] is not None else 0
        
        # scanned_qr tablosundan gerÃƒÂƒÃ‚Â§ek scan sayÃƒÂ„Ã‚Â±sÃƒÂ„Ã‚Â±nÃƒÂ„Ã‚Â± al
        execute_query(cursor, '''
            SELECT COUNT(DISTINCT qr_id)
            FROM scanned_qr
            WHERE session_id = ?
        ''', (session_id,))
        
        count_row = cursor.fetchone()
        scanned_count = count_row[0] if count_row else 0
        
        # Scanned QR listesi (en son taranan en ÃƒÂƒÃ‚Â¼stte)
        execute_query(cursor, '''
            SELECT qr_id, MAX(scanned_at) as last_scan
            FROM scanned_qr
            WHERE session_id = ?
            GROUP BY qr_id
            ORDER BY last_scan DESC
            LIMIT 500
        ''', (session_id,))
        
        scanned_rows = cursor.fetchall()
        scanned_qrs = [r[0] for r in scanned_rows] if scanned_rows else []

        close_db(conn)

        return jsonify({
            'success': True,
            'session_id': session_id,
            'scanned': scanned_count,
            'expected': expected,
            'scanned_qrs': scanned_qrs
        })
        
    except Exception as e:
        logging.error(f"Error in get_session_stats: {e}")
        return jsonify({'error': str(e)}), 500

@app.route('/get_recent_activities')
@login_required
def get_recent_activities():
    """Son QR tarama aktiviteleri - kullanÃƒÂ„Ã‚Â±cÃƒÂ„Ã‚Â± adlarÃƒÂ„Ã‚Â±yla"""
    try:
        conn = get_db()
        cursor = conn.cursor()
        
        # Get recent scanned QRs with user names
        execute_query(cursor, '''
            SELECT 
                sq.qr_id, 
                sq.scanned_by, 
                sq.scanned_at, 
                sq.part_code,
                eu.full_name,
                eu.username,
                pc.part_name
            FROM scanned_qr sq
            LEFT JOIN envanter_users eu ON sq.scanned_by = eu.id
            LEFT JOIN qr_codes qc ON sq.qr_id = qc.qr_id
            LEFT JOIN part_codes pc ON qc.part_code_id = pc.id
            ORDER BY sq.scanned_at DESC
            LIMIT 10
        ''')
        
        activities = []
        for row in cursor.fetchall():
            user_name = row[4] if row[4] else (row[5] if row[5] else f'KullanÃƒÂ„Ã‚Â±cÃƒÂ„Ã‚Â± #{row[1]}')
            part_name = row[6] if row[6] else 'Bilinmeyen ÃƒÂƒÃ…Â“rÃƒÂƒÃ‚Â¼n'
            
            activities.append({
                'qr_code': row[0],
                'scanned_by': user_name,
                'scanned_by_id': row[1],
                'scanned_at': row[2],
                'part_code': row[3] if row[3] else None,
                'part_name': part_name
            })
        
        close_db(conn)
        return jsonify(activities)
    except Exception as e:
        logging.error(f"Error in get_recent_activities: {e}")
        return jsonify([])  # Return empty array instead of error object

@app.route('/get_live_count_status')
@login_required
def get_live_count_status():
    """Aktif sayÃƒÂ„Ã‚Â±m sÃƒÂ„Ã‚Â±rasÃƒÂ„Ã‚Â±nda parÃƒÂƒÃ‚Â§a bazÃƒÂ„Ã‚Â±nda anlÃƒÂ„Ã‚Â±k durum"""
    try:
        conn = get_db()
        cursor = conn.cursor()
        placeholder = get_db_placeholder()
        
        # Aktif sayÃƒÂ„Ã‚Â±m var mÃƒÂ„Ã‚Â±?
        execute_query(cursor, f'SELECT id FROM count_sessions WHERE is_active = {placeholder}', (True,))
        active_session = cursor.fetchone()
        
        if not active_session:
            close_db(conn)
            return jsonify({
                'active': False,
                'message': 'Aktif sayÃƒÂ„Ã‚Â±m bulunamadÃƒÂ„Ã‚Â±'
            })
        
        session_id = active_session[0]
        
        # ParÃƒÂƒÃ‚Â§a bazÃƒÂ„Ã‚Â±nda sayÃƒÂ„Ã‚Â±m durumu
        execute_query(cursor, '''
            SELECT 
                pc.part_code,
                pc.part_name,
                COUNT(DISTINCT qc.qr_id) as beklenen_adet,
                COUNT(DISTINCT sq.qr_id) as sayilan_adet
            FROM part_codes pc
            LEFT JOIN qr_codes qc ON pc.id = qc.part_code_id
            LEFT JOIN scanned_qr sq ON qc.qr_id = sq.qr_id AND sq.session_id = ?
            GROUP BY pc.id, pc.part_code, pc.part_name
            HAVING beklenen_adet > 0
            ORDER BY pc.part_name
        ''', (str(session_id),))
        
        parts = []
        for row in cursor.fetchall():
            part_code = row[0]
            part_name = row[1]
            beklenen = row[2]
            sayilan = row[3]
            kalan = beklenen - sayilan
            durum = 'TamamlandÃƒÂ„Ã‚Â±' if kalan == 0 else f'{kalan} eksik'
            yuzdÃ„ÂžÃ‚Âµ = round((sayilan / beklenen * 100) if beklenen > 0 else 0, 1)
            
            parts.append({
                'part_code': part_code,
                'part_name': part_name,
                'beklenen_adet': beklenen,
                'sayilan_adet': sayilan,
                'kalan_adet': kalan,
                'durum': durum,
                'tamamlanma_yuzdesi': yuzdÃ„ÂžÃ‚Âµ
            })
        
        close_db(conn)
        return jsonify({
            'active': True,
            'session_id': session_id,
            'parts': parts,
            'total_parts': len(parts),
            'completed_parts': len([p for p in parts if p['kalan_adet'] == 0])
        })
        
    except Exception as e:
        logging.error(f"Error in get_live_count_status: {e}")
        return jsonify({'active': False, 'error': str(e)})

@app.route('/export_live_count')
@login_required
def export_live_count():
    """AnlÃƒÂ„Ã‚Â±k sayÃƒÂ„Ã‚Â±m durumunu Excel olarak indir"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from io import BytesIO
        
        conn = get_db()
        cursor = conn.cursor()
        placeholder = get_db_placeholder()
        
        # Aktif sayÃƒÂ„Ã‚Â±m var mÃƒÂ„Ã‚Â±?
        execute_query(cursor, f'SELECT id FROM count_sessions WHERE is_active = {placeholder}', (True,))
        active_session = cursor.fetchone()
        
        if not active_session:
            close_db(conn)
            return jsonify({'error': 'Aktif sayÃƒÂ„Ã‚Â±m bulunamadÃƒÂ„Ã‚Â±'}), 404
        
        session_id = active_session[0]
        
        # ParÃƒÂƒÃ‚Â§a bazÃƒÂ„Ã‚Â±nda sayÃƒÂ„Ã‚Â±m durumu
        execute_query(cursor, '''
            SELECT 
                pc.part_code,
                pc.part_name,
                COUNT(DISTINCT qc.qr_id) as beklenen_adet,
                COUNT(DISTINCT sq.qr_id) as sayilan_adet
            FROM part_codes pc
            LEFT JOIN qr_codes qc ON pc.id = qc.part_code_id
            LEFT JOIN scanned_qr sq ON qc.qr_id = sq.qr_id AND sq.session_id = ?
            GROUP BY pc.id, pc.part_code, pc.part_name
            HAVING beklenen_adet > 0
            ORDER BY pc.part_name
        ''', (str(session_id),))
        
        # Excel oluÃƒÂ…Ã…Â¸tur
        wb = Workbook()
        ws = wb.active
        ws.title = "CanlÃƒÂ„Ã‚Â± SayÃƒÂ„Ã‚Â±m Durumu"
        
        # Header
        headers = ['ParÃƒÂƒÃ‚Â§a Kodu', 'ParÃƒÂƒÃ‚Â§a AdÃƒÂ„Ã‚Â±', 'Beklenen Adet', 'SayÃƒÂ„Ã‚Â±lan Adet', 'Kalan Adet', 'Tamamlanma %', 'Durum']
        ws.append(headers)
        
        # Header stil
        header_fill = PatternFill(start_color="DC3545", end_color="DC3545", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Veri satÃƒÂ„Ã‚Â±rlarÃƒÂ„Ã‚Â±
        for row in cursor.fetchall():
            beklenen = row[2]
            sayilan = row[3]
            kalan = beklenen - sayilan
            yuzde = round((sayilan / beklenen * 100) if beklenen > 0 else 0, 1)
            durum = 'ÃƒÂ¢Ã…Â“Ã¢Â€Â¦ TamamlandÃƒÂ„Ã‚Â±' if kalan == 0 else f'ÃƒÂ¢Ã‚Â�Ã‚Â³ {kalan} eksik'
            
            ws.append([
                row[0],  # part_code
                row[1],  # part_name
                beklenen,
                sayilan,
                kalan,
                f"{yuzde}%",
                durum
            ])
        
        # Kolon geniÃƒÂ…Ã…Â¸likleri
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 35
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 15
        ws.column_dimensions['F'].width = 15
        ws.column_dimensions['G'].width = 20
        
        close_db(conn)
        
        # Excel dosyasÃƒÂ„Ã‚Â±nÃƒÂ„Ã‚Â± BytesIO'ya kaydet
        excel_file = BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)
        
        filename = f"Canli_Sayim_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        return send_file(
            excel_file,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
        
    except Exception as e:
        logging.exception(f"Error in export_live_count: {e}")
        try:
            close_db(conn)
        except:
            pass
        return jsonify({'error': str(e)}), 500

@app.route('/export_qr_activities')
@login_required
def export_qr_activities():
    """QR tarama hareketlerini Excel olarak indir"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill
        from io import BytesIO
        
        conn = get_db()
        cursor = conn.cursor()
        
        # TÃƒÂƒÃ‚Â¼m QR tarama hareketlerini ÃƒÂƒÃ‚Â§ek
        execute_query(cursor, '''
            SELECT 
                sq.qr_id, 
                sq.scanned_by, 
                sq.scanned_at, 
                sq.part_code,
                sq.session_id,
                eu.full_name,
                eu.username,
                pc.part_name,
                pc.part_code as part_code_full
            FROM scanned_qr sq
            LEFT JOIN envanter_users eu ON sq.scanned_by = eu.id
            LEFT JOIN qr_codes qc ON sq.qr_id = qc.qr_id
            LEFT JOIN part_codes pc ON qc.part_code_id = pc.id
            ORDER BY sq.scanned_at DESC
        ''')
        
        activities = cursor.fetchall()
        close_db(conn)
        
        # Excel oluÃƒÂ…Ã…Â¸tur
        wb = Workbook()
        ws = wb.active
        ws.title = "QR Hareketleri"
        
        # BaÃƒÂ…Ã…Â¸lÃƒÂ„Ã‚Â±k satÃƒÂ„Ã‚Â±rÃƒÂ„Ã‚Â±
        headers = ['QR Kodu', 'ParÃƒÂƒÃ‚Â§a Kodu', 'ParÃƒÂƒÃ‚Â§a AdÃƒÂ„Ã‚Â±', 'Okuyan KullanÃƒÂ„Ã‚Â±cÃƒÂ„Ã‚Â±', 
                  'KullanÃƒÂ„Ã‚Â±cÃƒÂ„Ã‚Â± AdÃƒÂ„Ã‚Â±', 'Okuma Tarihi', 'Seans ID']
        ws.append(headers)
        
        # BaÃƒÂ…Ã…Â¸lÃƒÂ„Ã‚Â±k stilini ayarla
        header_fill = PatternFill(start_color="DC3545", end_color="DC3545", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Veri satÃƒÂ„Ã‚Â±rlarÃƒÂ„Ã‚Â±
        for row in activities:
            qr_id = row[0]
            scanned_by_id = row[1]
            scanned_at = row[2]
            part_code = row[3] or row[8] or ''  # sq.part_code veya pc.part_code
            session_id = row[4]
            full_name = row[5] or ''
            username = row[6] or f'user_{scanned_by_id}'
            part_name = row[7] or 'Bilinmeyen ÃƒÂƒÃ…Â“rÃƒÂƒÃ‚Â¼n'
            
            ws.append([
                qr_id,
                part_code,
                part_name,
                full_name,
                username,
                scanned_at,
                session_id
            ])
        
        # Kolon geniÃƒÂ…Ã…Â¸liklerini ayarla
        ws.column_dimensions['A'].width = 25  # QR Kodu
        ws.column_dimensions['B'].width = 15  # ParÃƒÂƒÃ‚Â§a Kodu
        ws.column_dimensions['C'].width = 30  # ParÃƒÂƒÃ‚Â§a AdÃƒÂ„Ã‚Â±
        ws.column_dimensions['D'].width = 20  # Okuyan KullanÃƒÂ„Ã‚Â±cÃƒÂ„Ã‚Â±
        ws.column_dimensions['E'].width = 15  # KullanÃƒÂ„Ã‚Â±cÃƒÂ„Ã‚Â± AdÃƒÂ„Ã‚Â±
        ws.column_dimensions['F'].width = 20  # Okuma Tarihi
        ws.column_dimensions['G'].width = 10  # Seans ID
        
        # Excel'i memory'ye kaydet
        excel_file = BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)
        
        # Dosya adÃƒÂ„Ã‚Â±
        filename = f"QR_Hareketleri_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        return send_file(
            excel_file,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
        
    except Exception as e:
        logging.error(f"Error in export_qr_activities: {e}")
        return jsonify({'error': str(e)}), 500

@app.route('/get_count_status')
@login_required
def get_count_status():
    """Aktif sayÃƒÂ„Ã‚Â±m durumu"""
    try:
        conn = get_db()
        cursor = conn.cursor()
        
        execute_query(cursor, '''
            SELECT id, started_at, total_expected, total_scanned, is_active
            FROM count_sessions
            WHERE is_active = ?
            ORDER BY started_at DESC
            LIMIT 1
        ''', (True,))
        
        row = cursor.fetchone()
        close_db(conn)
        
        if row:
            return jsonify({
                'active': True,
                'session_id': row[0],
                'started_at': row[1],
                'total_expected': row[2],
                'total_scanned': row[3]
            })
        else:
            return jsonify({'active': False})
    except Exception as e:
        logging.error(f"Error in get_count_status: {e}")
        return jsonify({'error': str(e)}), 500

#  KULLANICI YÃƒÂƒÃ¢Â€Â“NETÃƒÂ„Ã‚Â°MÃƒÂ„Ã‚Â° (Sadece Admin)
@app.route('/api/users', methods=['GET'])
@login_required
def get_users():
    """TÃƒÂƒÃ‚Â¼m kullanÃƒÂ„Ã‚Â±cÃƒÂ„Ã‚Â±larÃƒÂ„Ã‚Â± listele - Sadece Admin"""
    if session.get('role') != 'admin':
        return jsonify({'error': 'Admin eriÃƒÂ…Ã…Â¸imi gerekli'}), 403
    
    try:
        conn = get_db()
        cursor = conn.cursor()
        execute_query(cursor, '''
            SELECT id, username, full_name, role, is_active_user, can_mark_used
            FROM envanter_users
            ORDER BY id
        ''')
        users = cursor.fetchall()
        close_db(conn)
        
        user_list = []
        for user in users:
            user_list.append({
                'id': user[0],
                'username': user[1],
                'full_name': user[2],
                'role': user[3],
                'is_active': user[4],
                'can_mark_used': user[5]
            })
        
        return jsonify({'success': True, 'users': user_list})
    except Exception as e:
        logging.error(f"Error in get_users: {e}")
        return jsonify({'error': str(e)}), 500

@app.route('/api/users', methods=['POST'])
@login_required
def create_user():
    """Yeni kullanÃƒÂ„Ã‚Â±cÃƒÂ„Ã‚Â± oluÃƒÂ…Ã…Â¸tur - Sadece Admin"""
    if session.get('role') != 'admin':
        return jsonify({'error': 'Admin eriÃƒÂ…Ã…Â¸imi gerekli'}), 403
    
    from werkzeug.security import generate_password_hash
    
    try:
        data = request.get_json()
        username = data.get('username')
        password = data.get('password')
        full_name = data.get('full_name')
        role = data.get('role', 'user')  # VarsayÃƒÂ„Ã‚Â±lan: user (sadece QR okutabilir)
        
        if not username or not password or not full_name:
            return jsonify({'error': 'KullanÃƒÂ„Ã‚Â±cÃƒÂ„Ã‚Â± adÃƒÂ„Ã‚Â±, ÃƒÂ…Ã…Â¸ifre ve tam ad gerekli'}), 400
        
        # Sadece 'user' ve 'admin' rolleri kabul et
        if role not in ['user', 'admin']:
            role = 'user'
        
        # Admin sayÃƒÂ„Ã‚Â±sÃƒÂ„Ã‚Â±nÃƒÂ„Ã‚Â± kontrol et - Sadece 1 admin olmalÃƒÂ„Ã‚Â±
        conn = get_db()
        cursor = conn.cursor()
        
        if role == 'admin':
            execute_query(cursor, "SELECT COUNT(*) FROM envanter_users WHERE role = 'admin'")
            admin_count = cursor.fetchone()[0]
            if admin_count > 0:
                close_db(conn)
                return jsonify({'error': 'Sistemde zaten bir admin var. Sadece 1 admin olabilir.'}), 400
        
        # KullanÃƒÂ„Ã‚Â±cÃƒÂ„Ã‚Â± adÃƒÂ„Ã‚Â± kontrolÃƒÂƒÃ‚Â¼
        execute_query(cursor, 'SELECT COUNT(*) FROM envanter_users WHERE username = ?', (username,))
        if cursor.fetchone()[0] > 0:
            close_db(conn)
            return jsonify({'error': 'Bu kullanÃƒÂ„Ã‚Â±cÃƒÂ„Ã‚Â± adÃƒÂ„Ã‚Â± zaten kullanÃƒÂ„Ã‚Â±lÃƒÂ„Ã‚Â±yor'}), 400
        
        # ÃƒÂ…Ã‚Âžifreyi hashle
        password_hash = generate_password_hash(password)
        
        # KullanÃƒÂ„Ã‚Â±cÃƒÂ„Ã‚Â± rolÃƒÂƒÃ‚Â¼ne gÃƒÂƒÃ‚Â¶re izinler
        can_mark_used = (role == 'admin')  # Sadece admin QR'larÃƒÂ„Ã‚Â± kullanÃƒÂ„Ã‚Â±lmÃƒÂ„Ã‚Â±ÃƒÂ…Ã…Â¸ iÃƒÂ…Ã…Â¸aretleyebilir
        
        # KullanÃƒÂ„Ã‚Â±cÃƒÂ„Ã‚Â±yÃƒÂ„Ã‚Â± ekle
        execute_query(cursor, '''
            INSERT INTO envanter_users 
            (username, password, password_hash, full_name, role, is_active_user, can_mark_used)
            VALUES (?, ?, ?, ?, ?, ?, ?)
        ''', (username, password, password_hash, full_name, role, True, can_mark_used))
        
        conn.commit()
        close_db(conn)
        
        return jsonify({
            'success': True,
            'message': f'KullanÃƒÂ„Ã‚Â±cÃƒÂ„Ã‚Â± {username} baÃƒÂ…Ã…Â¸arÃƒÂ„Ã‚Â±yla oluÃƒÂ…Ã…Â¸turuldu',
            'user': {
                'username': username,
                'full_name': full_name,
                'role': role,
                'can_mark_used': can_mark_used
            }
        })
    except Exception as e:
        logging.error(f"Error in create_user: {e}")
        return jsonify({'error': str(e)}), 500

@app.route('/api/users/<int:user_id>', methods=['DELETE'])
@login_required
def delete_user(user_id):
    """KullanÃƒÂ„Ã‚Â±cÃƒÂ„Ã‚Â± sil - Sadece Admin"""
    if session.get('role') != 'admin':
        return jsonify({'error': 'Admin eriÃƒÂ…Ã…Â¸imi gerekli'}), 403
    
    try:
        conn = get_db()
        cursor = conn.cursor()
        
        # Admin kendini silemez
        if user_id == session.get('user_id'):
            close_db(conn)
            return jsonify({'error': 'Kendi hesabÃƒÂ„Ã‚Â±nÃƒÂ„Ã‚Â±zÃƒÂ„Ã‚Â± silemezsiniz'}), 400
        
        # KullanÃƒÂ„Ã‚Â±cÃƒÂ„Ã‚Â± kontrolÃƒÂƒÃ‚Â¼
        execute_query(cursor, 'SELECT username, role FROM envanter_users WHERE id = ?', (user_id,))
        user = cursor.fetchone()
        
        if not user:
            close_db(conn)
            return jsonify({'error': 'KullanÃƒÂ„Ã‚Â±cÃƒÂ„Ã‚Â± bulunamadÃƒÂ„Ã‚Â±'}), 404
        
        # Admin silinemez
        if user[1] == 'admin':
            close_db(conn)
            return jsonify({'error': 'Admin kullanÃƒÂ„Ã‚Â±cÃƒÂ„Ã‚Â±sÃƒÂ„Ã‚Â± silinemez'}), 400
        
        # Sil
        execute_query(cursor, 'DELETE FROM envanter_users WHERE id = ?', (user_id,))
        conn.commit()
        close_db(conn)
        
        return jsonify({'success': True, 'message': f'KullanÃƒÂ„Ã‚Â±cÃƒÂ„Ã‚Â± {user[0]} silindi'})
    except Exception as e:
        logging.error(f"Error in delete_user: {e}")
        return jsonify({'error': str(e)}), 500

@app.route('/admin')
@login_required
def admin_panel():
    """Admin panel - ÃƒÂ…Ã‚Âžifre korumalÃƒÂ„Ã‚Â±"""
    # Admin kontrolÃƒÂƒÃ‚Â¼
    if session.get('role') != 'admin':
        return jsonify({'error': 'Admin eriÃƒÂ…Ã…Â¸imi gerekli'}), 403
    
    # Admin panel ÃƒÂ…Ã…Â¸ifresi kontrolÃƒÂƒÃ‚Â¼ (session'da tutulur)
    if not session.get('admin_panel_unlocked'):
        # ÃƒÂ…Ã‚Âžifre girilmemiÃƒÂ…Ã…Â¸, ÃƒÂ…Ã…Â¸ifre sayfasÃƒÂ„Ã‚Â±nÃƒÂ„Ã‚Â± gÃƒÂƒÃ‚Â¶ster
        return render_template('admin_login.html')
    
    # ÃƒÂ…Ã‚Âžifre doÃƒÂ„Ã…Â¸ru, admin paneli gÃƒÂƒÃ‚Â¶ster
    return render_template('admin.html')

@app.route('/admin/verify', methods=['POST'])
@login_required
def verify_admin_password():
    """Admin panel ÃƒÂ…Ã…Â¸ifre doÃƒÂ„Ã…Â¸rulama"""
    if session.get('role') != 'admin':
        return jsonify({'error': 'Admin eriÃƒÂ…Ã…Â¸imi gerekli'}), 403
    
    data = request.get_json()
    password = data.get('password')
    
    # Admin sayÃƒÂ„Ã‚Â±m ÃƒÂ…Ã…Â¸ifresi
    ADMIN_PASSWORD = '@R9t$L7e!xP2w'
    
    if password == ADMIN_PASSWORD:
        # ÃƒÂ…Ã‚Âžifreyi session'a kaydet
        session['admin_panel_unlocked'] = True
        return jsonify({'success': True, 'message': 'Admin paneline hoÃƒÂ…Ã…Â¸geldiniz'})
    else:
        return jsonify({'success': False, 'error': 'HatalÃƒÂ„Ã‚Â± ÃƒÂ…Ã…Â¸ifre'}), 401

@app.route('/admin/logout', methods=['POST'])
@login_required
def admin_panel_logout():
    """Admin panel ÃƒÂƒÃ‚Â§ÃƒÂ„Ã‚Â±kÃƒÂ„Ã‚Â±ÃƒÂ…Ã…Â¸ÃƒÂ„Ã‚Â± (sadece panel kilidini kaldÃƒÂ„Ã‚Â±r)"""
    session.pop('admin_panel_unlocked', None)
    return jsonify({'success': True})

@app.route('/admin/users')
@login_required
def admin_users_page():
    """KullanÃƒÂ„Ã‚Â±cÃƒÂ„Ã‚Â± yÃƒÂƒÃ‚Â¶netim sayfasÃƒÂ„Ã‚Â± veya API - GET request"""
    # Admin kontrolÃƒÂƒÃ‚Â¼
    if session.get('role') != 'admin':
        return jsonify({'error': 'Admin eriÃƒÂ…Ã…Â¸imi gerekli'}), 403
    
    # Accept header kontrolÃƒÂƒÃ‚Â¼ - JSON mu HTML mi isteniyor?
    if request.accept_mimetypes.best_match(['application/json', 'text/html']) == 'application/json':
        # API ÃƒÂƒÃ‚Â§aÃƒÂ„Ã…Â¸rÃƒÂ„Ã‚Â±sÃƒÂ„Ã‚Â± - KullanÃƒÂ„Ã‚Â±cÃƒÂ„Ã‚Â± listesini dÃƒÂƒÃ‚Â¶ndÃƒÂƒÃ‚Â¼r
        try:
            conn = get_db()
            cursor = conn.cursor()
            
            execute_query(cursor, '''
                SELECT id, username, full_name, role, created_at, is_active_user
                FROM envanter_users
                ORDER BY created_at DESC
            ''')
            
            users = []
            for row in cursor.fetchall():
                users.append({
                    'id': row[0],
                    'username': row[1],
                    'fullname': row[2] or '',
                    'role': row[3],
                    'created_at': row[4],
                    'is_active': row[5] if len(row) > 5 else True
                })
            
            close_db(conn)
            return jsonify(users)
            
        except Exception as e:
            logging.error(f"Error in get users API: {e}")
            return jsonify({'error': str(e)}), 500
    else:
        # HTML sayfasÃƒÂ„Ã‚Â± - KullanÃƒÂ„Ã‚Â±cÃƒÂ„Ã‚Â± yÃƒÂƒÃ‚Â¶netim arayÃƒÂƒÃ‚Â¼zÃƒÂƒÃ‚Â¼
        # Admin panel kilidini kontrol et
        if not session.get('admin_panel_unlocked'):
            return redirect('/admin')
        
        # KullanÃƒÂ„Ã‚Â±cÃƒÂ„Ã‚Â±larÃƒÂ„Ã‚Â± veritabanÃƒÂ„Ã‚Â±ndan ÃƒÂƒÃ‚Â§ek
        try:
            conn = get_db()
            cursor = conn.cursor()
            
            execute_query(cursor, '''
                SELECT id, username, full_name, role, created_at, is_active_user
                FROM envanter_users
                ORDER BY created_at DESC
            ''')
            
            users = []
            for row in cursor.fetchall():
                users.append({
                    'id': row[0],
                    'username': row[1],
                    'full_name': row[2] or 'ÃƒÂ„Ã‚Â°simsiz',
                    'role': row[3],
                    'created_at': row[4],
                    'is_active': row[5] if len(row) > 5 else True
                })
            
            close_db(conn)
            
            # KullanÃƒÂ„Ã‚Â±cÃƒÂ„Ã‚Â±larÃƒÂ„Ã‚Â± template'e gÃƒÂƒÃ‚Â¶nder
            return render_template('admin_users.html', users=users)
            
        except Exception as e:
            logging.error(f"Error loading users page: {e}")
            return render_template('admin_users.html', users=[])

@app.route('/admin/users', methods=['POST'])
@login_required
def admin_create_user():
    """Yeni kullanÃƒÂ„Ã‚Â±cÃƒÂ„Ã‚Â± oluÃƒÂ…Ã…Â¸tur - Admin Users sayfasÃƒÂ„Ã‚Â±"""
    from werkzeug.security import generate_password_hash
    
    # Admin kontrolÃƒÂƒÃ‚Â¼
    if session.get('role') != 'admin':
        return jsonify({'error': 'Admin eriÃƒÂ…Ã…Â¸imi gerekli'}), 403
    
    try:
        # Debug: Request bilgilerini logla
        logging.info(f"Content-Type: {request.content_type}")
        logging.info(f"Request data: {request.data}")
        
        # JSON verisini al - force=True ile Content-Type kontrolÃƒÂƒÃ‚Â¼nÃƒÂƒÃ‚Â¼ bypass et
        data = request.get_json(force=True)
        if not data:
            logging.error("JSON data is None or empty")
            return jsonify({'error': 'GeÃƒÂƒÃ‚Â§ersiz JSON verisi'}), 400
        
        logging.info(f"Parsed JSON: {data}")
        
        username = str(data.get('username', '')).strip()
        password = str(data.get('password', '')).strip()
        fullname = str(data.get('fullname', '')).strip()
        role = str(data.get('role', 'user')).strip()
        
        # Validasyonlar
        if not username or not password:
            return jsonify({'error': 'KullanÃƒÂ„Ã‚Â±cÃƒÂ„Ã‚Â± adÃƒÂ„Ã‚Â± ve ÃƒÂ…Ã…Â¸ifre gerekli'}), 400
        
        if len(username) < 3:
            return jsonify({'error': 'KullanÃƒÂ„Ã‚Â±cÃƒÂ„Ã‚Â± adÃƒÂ„Ã‚Â± en az 3 karakter olmalÃƒÂ„Ã‚Â±'}), 400
        
        if len(password) < 4:
            return jsonify({'error': 'ÃƒÂ…Ã‚Âžifre en az 4 karakter olmalÃƒÂ„Ã‚Â±'}), 400
        
        if role not in ['admin', 'user']:
            return jsonify({'error': 'GeÃƒÂƒÃ‚Â§ersiz rol. admin veya user olmalÃƒÂ„Ã‚Â±'}), 400
        
        # KullanÃƒÂ„Ã‚Â±cÃƒÂ„Ã‚Â± zaten var mÃƒÂ„Ã‚Â±?
        conn = get_db()
        cursor = conn.cursor()
        
        execute_query(cursor, 'SELECT id FROM envanter_users WHERE username = ?', (username,))
        if cursor.fetchone():
            close_db(conn)
            return jsonify({'error': 'Bu kullanÃƒÂ„Ã‚Â±cÃƒÂ„Ã‚Â± adÃƒÂ„Ã‚Â± zaten kullanÃƒÂ„Ã‚Â±lÃƒÂ„Ã‚Â±yor'}), 400
        
        # ÃƒÂ…Ã‚Âžifreyi hashle
        hashed_password = generate_password_hash(password)
        
        # Yeni kullanÃƒÂ„Ã‚Â±cÃƒÂ„Ã‚Â± ekle
        execute_query(cursor, '''
            INSERT INTO envanter_users (username, password_hash, full_name, role, created_at, is_active_user)
            VALUES (?, ?, ?, ?, ?, ?)
        ''', (username, hashed_password, fullname or '', role, datetime.now(), True))
        
        conn.commit()
        close_db(conn)
        
        return jsonify({'success': True, 'message': f'KullanÃƒÂ„Ã‚Â±cÃƒÂ„Ã‚Â± {username} oluÃƒÂ…Ã…Â¸turuldu'})
        
    except Exception as e:
        logging.error(f"Error in admin_create_user: {e}")
        return jsonify({'error': str(e)}), 500

@app.route('/admin/users/<int:user_id>', methods=['PUT'])
@login_required
def admin_update_user(user_id):
    """KullanÃƒÂ„Ã‚Â±cÃƒÂ„Ã‚Â± bilgilerini gÃƒÂƒÃ‚Â¼ncelle - Admin Users sayfasÃƒÂ„Ã‚Â±"""
    # Admin kontrolÃƒÂƒÃ‚Â¼
    if session.get('role') != 'admin':
        return jsonify({'error': 'Admin eriÃƒÂ…Ã…Â¸imi gerekli'}), 403
    
    try:
        data = request.get_json()
        if not data:
            return jsonify({'error': 'GeÃƒÂƒÃ‚Â§ersiz JSON verisi'}), 400
        
        fullname = str(data.get('fullname', '')).strip()
        role = str(data.get('role', 'user')).strip()
        
        if role not in ['admin', 'user']:
            return jsonify({'error': 'GeÃƒÂƒÃ‚Â§ersiz rol'}), 400
        
        conn = get_db()
        cursor = conn.cursor()
        
        # KullanÃƒÂ„Ã‚Â±cÃƒÂ„Ã‚Â± var mÃƒÂ„Ã‚Â± kontrol et
        execute_query(cursor, 'SELECT id FROM envanter_users WHERE id = ?', (user_id,))
        if not cursor.fetchone():
            close_db(conn)
            return jsonify({'error': 'KullanÃƒÂ„Ã‚Â±cÃƒÂ„Ã‚Â± bulunamadÃƒÂ„Ã‚Â±'}), 404
        
        # KullanÃƒÂ„Ã‚Â±cÃƒÂ„Ã‚Â±yÃƒÂ„Ã‚Â± gÃƒÂƒÃ‚Â¼ncelle
        execute_query(cursor, '''
            UPDATE envanter_users 
            SET full_name = ?, role = ?, updated_at = ?
            WHERE id = ?
        ''', (fullname, role, datetime.now(), user_id))
        
        conn.commit()
        close_db(conn)
        
        return jsonify({'success': True, 'message': 'KullanÃƒÂ„Ã‚Â±cÃƒÂ„Ã‚Â± gÃƒÂƒÃ‚Â¼ncellendi'})
        
    except Exception as e:
        logging.error(f"Error in admin_update_user: {e}")
        return jsonify({'error': str(e)}), 500

@app.route('/admin/users/<int:user_id>', methods=['DELETE'])
@login_required
def admin_delete_user(user_id):
    """KullanÃƒÂ„Ã‚Â±cÃƒÂ„Ã‚Â±yÃƒÂ„Ã‚Â± sil - Admin Users sayfasÃƒÂ„Ã‚Â±"""
    # Admin kontrolÃƒÂƒÃ‚Â¼
    if session.get('role') != 'admin':
        return jsonify({'error': 'Admin eriÃƒÂ…Ã…Â¸imi gerekli'}), 403
    
    try:
        # Kendi hesabÃƒÂ„Ã‚Â±nÃƒÂ„Ã‚Â± silemesin
        if user_id == session.get('user_id'):
            return jsonify({'error': 'Kendi hesabÃƒÂ„Ã‚Â±nÃƒÂ„Ã‚Â±zÃƒÂ„Ã‚Â± silemezsiniz'}), 400
        
        conn = get_db()
        cursor = conn.cursor()
        
        # KullanÃƒÂ„Ã‚Â±cÃƒÂ„Ã‚Â± bilgilerini al
        execute_query(cursor, 'SELECT username FROM envanter_users WHERE id = ?', (user_id,))
        user = cursor.fetchone()
        
        if not user:
            close_db(conn)
            return jsonify({'error': 'KullanÃƒÂ„Ã‚Â±cÃƒÂ„Ã‚Â± bulunamadÃƒÂ„Ã‚Â±'}), 404
        
        # KullanÃƒÂ„Ã‚Â±cÃƒÂ„Ã‚Â±yÃƒÂ„Ã‚Â± sil
        execute_query(cursor, 'DELETE FROM envanter_users WHERE id = ?', (user_id,))
        conn.commit()
        close_db(conn)
        
        return jsonify({'success': True, 'message': f'KullanÃƒÂ„Ã‚Â±cÃƒÂ„Ã‚Â± {user[0]} silindi'})
        
    except Exception as e:
        logging.error(f"Error in admin_delete_user: {e}")
        return jsonify({'error': str(e)}), 500

@app.route('/admin/users/<int:user_id>/change_password', methods=['POST'])
@login_required
def admin_change_password(user_id):
    """KullanÃƒÂ„Ã‚Â±cÃƒÂ„Ã‚Â± ÃƒÂ…Ã…Â¸ifresini deÃƒÂ„Ã…Â¸iÃƒÂ…Ã…Â¸tir - Admin Users sayfasÃƒÂ„Ã‚Â±"""
    from werkzeug.security import generate_password_hash
    
    # Admin kontrolÃƒÂƒÃ‚Â¼
    if session.get('role') != 'admin':
        return jsonify({'error': 'Admin eriÃƒÂ…Ã…Â¸imi gerekli'}), 403
    
    try:
        data = request.get_json()
        if not data:
            return jsonify({'error': 'GeÃƒÂƒÃ‚Â§ersiz JSON verisi'}), 400
        
        new_password = str(data.get('new_password', '')).strip()
        
        if not new_password:
            return jsonify({'error': 'Yeni ÃƒÂ…Ã…Â¸ifre gerekli'}), 400
        
        if len(new_password) < 4:
            return jsonify({'error': 'ÃƒÂ…Ã‚Âžifre en az 4 karakter olmalÃƒÂ„Ã‚Â±'}), 400
        
        conn = get_db()
        cursor = conn.cursor()
        
        # KullanÃƒÂ„Ã‚Â±cÃƒÂ„Ã‚Â± var mÃƒÂ„Ã‚Â± kontrol et
        execute_query(cursor, 'SELECT username FROM envanter_users WHERE id = ?', (user_id,))
        user = cursor.fetchone()
        if not user:
            close_db(conn)
            return jsonify({'error': 'KullanÃƒÂ„Ã‚Â±cÃƒÂ„Ã‚Â± bulunamadÃƒÂ„Ã‚Â±'}), 404
        
        # ÃƒÂ…Ã‚Âžifreyi hashle ve gÃƒÂƒÃ‚Â¼ncelle
        hashed_password = generate_password_hash(new_password)
        execute_query(cursor, '''
            UPDATE envanter_users 
            SET password_hash = ?, last_password_change = ?
            WHERE id = ?
        ''', (hashed_password, datetime.now(), user_id))
        
        conn.commit()
        close_db(conn)
        
        return jsonify({'success': True, 'message': 'ÃƒÂ…Ã‚Âžifre deÃƒÂ„Ã…Â¸iÃƒÂ…Ã…Â¸tirildi'})
        
    except Exception as e:
        logging.error(f"Error in admin_change_password: {e}")
        return jsonify({'error': str(e)}), 500

@app.route('/unlock_qrcodes', methods=['POST'])
def unlock_qrcodes():
    """QR klasÃƒÂƒÃ‚Â¶rÃƒÂƒÃ‚Â¼ kilidini aÃƒÂƒÃ‚Â§ (admin paneli ÃƒÂ…Ã…Â¸ifresi ile)"""
    password = request.form.get('password', '').strip()
    
    # Admin paneli ÃƒÂ…Ã…Â¸ifresi ile kontrol et
    if password == ADMIN_COUNT_PASSWORD:
        session['qrcodes_unlocked'] = True
        # KullanÃƒÂ„Ã‚Â±cÃƒÂ„Ã‚Â±yÃƒÂ„Ã‚Â± qrcodes klasÃƒÂƒÃ‚Â¶rÃƒÂƒÃ‚Â¼ne yÃƒÂƒÃ‚Â¶nlendir
        return '''
        <html>
        <head>
            <title>EriÃƒÂ…Ã…Â¸im ÃƒÂ„Ã‚Â°zni Verildi</title>
            <meta http-equiv="refresh" content="2;url=/static/qrcodes/">
            <style>
                body { font-family: Arial; text-align: center; padding: 50px; background: #f5f5f5; }
                .success { background: white; padding: 30px; border-radius: 10px; max-width: 400px; margin: auto; box-shadow: 0 2px 10px rgba(0,0,0,0.1); }
                h2 { color: #4caf50; }
            </style>
        </head>
        <body>
            <div class="success">
                <h2>ÃƒÂ¢Ã…Â“Ã¢Â€Â¦ EriÃƒÂ…Ã…Â¸im ÃƒÂ„Ã‚Â°zni Verildi</h2>
                <p>QR klasÃƒÂƒÃ‚Â¶rÃƒÂƒÃ‚Â¼ne yÃƒÂƒÃ‚Â¶nlendiriliyorsunuz...</p>
            </div>
        </body>
        </html>
        '''
    else:
        return '''
        <html>
        <head>
            <title>HatalÃƒÂ„Ã‚Â± ÃƒÂ…Ã‚Âžifre</title>
            <style>
                body { font-family: Arial; text-align: center; padding: 50px; background: #f5f5f5; }
                .error { background: white; padding: 30px; border-radius: 10px; max-width: 400px; margin: auto; box-shadow: 0 2px 10px rgba(0,0,0,0.1); }
                h2 { color: #d32f2f; }
                a { color: #1976d2; text-decoration: none; }
            </style>
        </head>
        <body>
            <div class="error">
                <h2>ÃƒÂ¢Ã‚Â�Ã…Â’ HatalÃƒÂ„Ã‚Â± ÃƒÂ…Ã‚Âžifre</h2>
                <p>Admin paneli ÃƒÂ…Ã…Â¸ifresini doÃƒÂ„Ã…Â¸ru girdiÃƒÂ„Ã…Â¸inizden emin olun.</p>
                <p><a href="/static/qrcodes/">Tekrar Dene</a></p>
            </div>
        </body>
        </html>
        ''', 403

@app.route('/admin/reset_active_sessions', methods=['POST'])
@login_required
@admin_required_decorator
def reset_active_sessions():
    """TÃƒÂƒÃ‚Â¼m aktif sayÃƒÂ„Ã‚Â±m oturumlarÃƒÂ„Ã‚Â±nÃƒÂ„Ã‚Â± sÃƒÂ„Ã‚Â±fÃƒÂ„Ã‚Â±rla"""
    try:
        conn = get_db()
        cursor = conn.cursor()
        
        # TÃƒÂƒÃ‚Â¼m aktif sayÃƒÂ„Ã‚Â±mlarÃƒÂ„Ã‚Â± pasif yap
        execute_query(cursor, 'UPDATE count_sessions SET is_active = ? WHERE is_active = ?', (False, True))
        
        conn.commit()
        affected_rows = cursor.rowcount
        close_db(conn)
        
        return jsonify({
            'success': True,
            'message': f'ÃƒÂ¢Ã…Â“Ã¢Â€Â¦ {affected_rows} aktif sayÃƒÂ„Ã‚Â±m oturumu sÃƒÂ„Ã‚Â±fÃƒÂ„Ã‚Â±rlandÃƒÂ„Ã‚Â±'
        })
        
    except Exception as e:
        logging.exception(f"Reset active sessions error: {e}")
        return jsonify({
            'success': False,
            'message': f'ÃƒÂ¢Ã‚Â�Ã…Â’ Hata: {str(e)}'
        }), 500

@app.route('/metrics')
def metrics():
    """Sistem metrikleri"""
    try:
        conn = get_db()
        cursor = conn.cursor()

        # ÃƒÂ„Ã‚Â°statistikler
        execute_query(cursor, 'SELECT COUNT(*) FROM qr_codes')
        total_qr = cursor.fetchone()[0]

        execute_query(cursor, 'SELECT COUNT(*) FROM qr_codes WHERE is_used = 1')
        used_qr = cursor.fetchone()[0]

        execute_query(cursor, 'SELECT COUNT(*) FROM envanter_users')
        total_users = cursor.fetchone()[0]

        execute_query(cursor, "SELECT COUNT(*) FROM count_sessions WHERE is_active = TRUE")
        active_sessions = cursor.fetchone()[0]

        close_db(conn)

        return jsonify({
            'qr_codes': {
                'total': total_qr,
                'used': used_qr,
                'remaining': total_qr - used_qr
            },
            'users': total_users,
            'active_sessions': active_sessions,
            'timestamp': datetime.now().isoformat()
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500

# Render.com deployment check
def is_render_deployment():
    """Render.com deploy kontrolÃƒÂƒÃ‚Â¼"""
    return os.environ.get('RENDER') is not None

def get_port():
    """Port numarasÃƒÂ„Ã‚Â±nÃƒÂ„Ã‚Â± al"""
    return int(os.environ.get('PORT', 5001))

#  LOKAL SÃƒÂ„Ã‚Â°STEM - QR Admin modÃƒÂƒÃ‚Â¼lÃƒÂƒÃ‚Â¼ kaldÃƒÂ„Ã‚Â±rÃƒÂ„Ã‚Â±ldÃƒÂ„Ã‚Â±

# ==================== CLOUDFLARE TUNNEL YÃƒÂƒÃ¢Â€Â“NETÃƒÂ„Ã‚Â°MÃƒÂ„Ã‚Â° ====================
import subprocess
import threading
import re

# Global deÃƒÂ„Ã…Â¸iÃƒÂ…Ã…Â¸kenler
tunnel_process = None
tunnel_url = None
tunnel_running = False

@app.route('/tunnel/start', methods=['POST'])
@login_required
def start_tunnel():
    """Cloudflare Tunnel baÃƒÂ…Ã…Â¸lat"""
    global tunnel_process, tunnel_url, tunnel_running
    
    try:
        if tunnel_running and tunnel_process:
            return jsonify({
                'success': True,
                'message': 'Tunnel zaten ÃƒÂƒÃ‚Â§alÃƒÂ„Ã‚Â±ÃƒÂ…Ã…Â¸ÃƒÂ„Ã‚Â±yor',
                'url': tunnel_url
            })
        
        # Cloudflared.exe yolu
        cloudflared_path = os.path.join(os.path.dirname(__file__), 'cloudflared.exe')
        
        if not os.path.exists(cloudflared_path):
            return jsonify({
                'success': False,
                'message': f'cloudflared.exe bulunamadÃƒÂ„Ã‚Â±: {cloudflared_path}'
            }), 404
        
        # Tunnel baÃƒÂ…Ã…Â¸lat
        tunnel_process = subprocess.Popen(
            [cloudflared_path, 'tunnel', '--url', 'http://localhost:5002'],
            stdout=subprocess.PIPE,
            stderr=subprocess.PIPE,
            text=True,
            bufsize=1,
            creationflags=subprocess.CREATE_NO_WINDOW if os.name == 'nt' else 0
        )
        
        # URL'yi yakalamak iÃƒÂƒÃ‚Â§in arka planda thread baÃƒÂ…Ã…Â¸lat
        def capture_url():
            global tunnel_url, tunnel_running
            try:
                # Hem stdout hem stderr'i kontrol et
                while True:
                    # stdout'u kontrol et
                    line = tunnel_process.stderr.readline()
                    if not line:
                        break
                    
                    print(f"[Cloudflare] {line.strip()}")
                    
                    # trycloudflare.com URL'sini yakala
                    if 'trycloudflare.com' in line:
                        match = re.search(r'https://[\w-]+\.trycloudflare\.com', line)
                        if match:
                            tunnel_url = match.group(0)
                            tunnel_running = True
                            print(f"ÃƒÂ¢Ã…Â“Ã¢Â€Â¦ Cloudflare Tunnel baÃƒÂ…Ã…Â¸latÃƒÂ„Ã‚Â±ldÃƒÂ„Ã‚Â±: {tunnel_url}")
                            break
            except Exception as e:
                print(f"ÃƒÂ¢Ã‚Â�Ã…Â’ URL yakalama hatasÃƒÂ„Ã‚Â±: {e}")
        
        threading.Thread(target=capture_url, daemon=True).start()
        
        # URL'nin yakalanmasÃƒÂ„Ã‚Â± iÃƒÂƒÃ‚Â§in biraz daha bekle
        import time
        max_wait = 10  # maksimum 10 saniye bekle
        waited = 0
        while not tunnel_url and waited < max_wait:
            time.sleep(0.5)
            waited += 0.5
        
        if tunnel_url:
            return jsonify({
                'success': True,
                'message': 'Tunnel baÃƒÂ…Ã…Â¸arÃƒÂ„Ã‚Â±yla baÃƒÂ…Ã…Â¸latÃƒÂ„Ã‚Â±ldÃƒÂ„Ã‚Â±',
                'url': tunnel_url
            })
        else:
            # Yine de baÃƒÂ…Ã…Â¸arÃƒÂ„Ã‚Â±lÃƒÂ„Ã‚Â± say, arka planda yakalanacak
            tunnel_running = True
            return jsonify({
                'success': True,
                'message': 'Tunnel baÃƒÂ…Ã…Â¸latÃƒÂ„Ã‚Â±ldÃƒÂ„Ã‚Â±, URL yakÃƒÂ„Ã‚Â±nda hazÃƒÂ„Ã‚Â±r olacak',
                'url': None
            })
            
    except Exception as e:
        logging.exception("Tunnel baÃƒÂ…Ã…Â¸latma hatasÃƒÂ„Ã‚Â±")
        tunnel_running = False
        tunnel_url = None
        return jsonify({
            'success': False,
            'message': f'Hata: {str(e)}'
        }), 500

@app.route('/tunnel/stop', methods=['POST'])
@login_required
def stop_tunnel():
    """Cloudflare Tunnel durdur"""
    global tunnel_process, tunnel_url, tunnel_running
    
    try:
        if not tunnel_running or not tunnel_process:
            return jsonify({
                'success': False,
                'message': 'Tunnel zaten durdurulmuÃƒÂ…Ã…Â¸'
            })
        
        # Process'i durdur
        tunnel_process.terminate()
        tunnel_process.wait(timeout=5)
        
        tunnel_process = None
        tunnel_url = None
        tunnel_running = False
        
        print("ÃƒÂ¢Ã…Â“Ã¢Â€Â¦ Cloudflare Tunnel durduruldu")
        
        return jsonify({
            'success': True,
            'message': 'Tunnel baÃƒÂ…Ã…Â¸arÃƒÂ„Ã‚Â±yla durduruldu'
        })
        
    except Exception as e:
        logging.exception("Tunnel durdurma hatasÃƒÂ„Ã‚Â±")
        return jsonify({
            'success': False,
            'message': f'Hata: {str(e)}'
        }), 500

@app.route('/tunnel/status')
@login_required
def tunnel_status():
    """Tunnel durumunu kontrol et"""
    global tunnel_running, tunnel_url, tunnel_process
    
    # Process hala ÃƒÂƒÃ‚Â§alÃƒÂ„Ã‚Â±ÃƒÂ…Ã…Â¸ÃƒÂ„Ã‚Â±yor mu kontrol et
    process_alive = tunnel_process is not None and tunnel_process.poll() is None
    
    return jsonify({
        'running': tunnel_running,
        'url': tunnel_url,
        'process_alive': process_alive,
        'debug': {
            'tunnel_running_flag': tunnel_running,
            'tunnel_url': tunnel_url,
            'process_exists': tunnel_process is not None,
            'process_alive': process_alive
        }
    })

@app.route('/tunnel/logs')
@login_required
def tunnel_logs():
    """Tunnel loglarÃƒÂ„Ã‚Â±nÃƒÂ„Ã‚Â± gÃƒÂƒÃ‚Â¶ster (debug iÃƒÂƒÃ‚Â§in)"""
    global tunnel_process
    
    if not tunnel_process:
        return jsonify({'error': 'Tunnel ÃƒÂƒÃ‚Â§alÃƒÂ„Ã‚Â±ÃƒÂ…Ã…Â¸mÃƒÂ„Ã‚Â±yor'}), 404
    
    # Son birkaÃƒÂƒÃ‚Â§ satÃƒÂ„Ã‚Â±rÃƒÂ„Ã‚Â± oku
    try:
        # stderr'den oku (cloudflared loglarÃƒÂ„Ã‚Â± oraya yazÃƒÂ„Ã‚Â±yor)
        logs = []
        return jsonify({'logs': logs, 'message': 'Log okuma aktif deÃƒÂ„Ã…Â¸il'})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

# ============================================================================
#  DATABASE BACKUP SÃƒÂ„Ã‚Â°STEMÃƒÂ„Ã‚Â°
# ============================================================================

def backup_database():
    """SQLite veritabanini yedekle - Veri kaybi korumasÃƒÂ„Ã‚Â±"""
    try:
        db_path = 'instance/envanter_local.db'
        if not os.path.exists(db_path):
            app.logger.warning(f'[ERROR] Database dosyasÃƒÂ„Ã‚Â± bulunamadÃƒÂ„Ã‚Â±: {db_path}')
            return None
            
        backup_dir = 'backups'
        os.makedirs(backup_dir, exist_ok=True)
        
        # Timestamp ile backup adÃƒÂ„Ã‚Â± (YÃƒÂ„Ã‚Â±l-Ay-GÃƒÂƒÃ‚Â¼n_Saat-Dakika-Saniye)
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        backup_path = f'{backup_dir}/envanter_backup_{timestamp}.db'
        
        # Orijinal dosya boyutu
        original_size = os.path.getsize(db_path)
        
        # Kopyala
        import shutil
        shutil.copy2(db_path, backup_path)
        
        # Backup boyutu kontrol et
        backup_size = os.path.getsize(backup_path)
        
        # BÃƒÂƒÃ‚Â¼tÃƒÂƒÃ‚Â¼nlÃƒÂƒÃ‚Â¼k kontrolÃƒÂƒÃ‚Â¼
        if backup_size != original_size:
            app.logger.error(f'[ERROR] Backup bÃƒÂƒÃ‚Â¼tÃƒÂƒÃ‚Â¼nlÃƒÂƒÃ‚Â¼ÃƒÂ„Ã…Â¸ÃƒÂƒÃ‚Â¼ baÃƒÂ…Ã…Â¸arÃƒÂ„Ã‚Â±sÃƒÂ„Ã‚Â±z! Orijinal: {original_size}, Backup: {backup_size}')
            os.remove(backup_path)
            return None
        
        # BaÃƒÂ…Ã…Â¸arÃƒÂ„Ã‚Â±lÃƒÂ„Ã‚Â±
        size_mb = round(original_size / (1024**2), 2)
        app.logger.info(f'[OK] Database backup oluÃƒÂ…Ã…Â¸turuldu:')
        app.logger.info(f'   Konum: {os.path.abspath(backup_path)}')
        app.logger.info(f'   Boyut: {size_mb} MB')
        app.logger.info(f'   Zaman: {timestamp}')
        
        # Eski backup'larÃƒÂ„Ã‚Â± temizle (son 30'u tut)
        cleanup_old_backups(backup_dir, keep=30)
        
        return backup_path
    except Exception as e:
        app.logger.error(f'[ERROR] Backup hatasÃƒÂ„Ã‚Â±: {e}')
        import traceback
        app.logger.error(traceback.format_exc())
        return None

def verify_backup_integrity():
    """Backup dosyalarÃƒÂ„Ã‚Â±nÃƒÂ„Ã‚Â±n bÃƒÂƒÃ‚Â¼tÃƒÂƒÃ‚Â¼nlÃƒÂƒÃ‚Â¼ÃƒÂ„Ã…Â¸ÃƒÂƒÃ‚Â¼nÃƒÂƒÃ‚Â¼ kontrol et"""
    try:
        backup_dir = 'backups'
        if not os.path.exists(backup_dir):
            return
        
        backups = [
            os.path.join(backup_dir, f) 
            for f in os.listdir(backup_dir) 
            if f.startswith('envanter_backup_') and f.endswith('.db')
        ]
        
        if not backups:
            return
        
        # En son backup'ÃƒÂ„Ã‚Â± kontrol et
        latest_backup = max(backups, key=os.path.getmtime)
        
        try:
            # SQLite database'i aÃƒÂƒÃ‚Â§ ve kontrol et
            import sqlite3
            conn = sqlite3.connect(latest_backup)
            cursor = conn.cursor()
            cursor.execute("SELECT name FROM sqlite_master WHERE type='table';")
            tables = cursor.fetchall()
            conn.close()
            
            if tables:
                app.logger.info(f'[OK] Backup bÃƒÂƒÃ‚Â¼tÃƒÂƒÃ‚Â¼nlÃƒÂƒÃ‚Â¼ÃƒÂ„Ã…Â¸ÃƒÂƒÃ‚Â¼ OK: {os.path.basename(latest_backup)} ({len(tables)} tablo)')
            else:
                app.logger.warning(f'[WARNING] Backup boÃƒÂ…Ã…Â¸ gÃƒÂƒÃ‚Â¶rÃƒÂƒÃ‚Â¼nÃƒÂƒÃ‚Â¼yor: {os.path.basename(latest_backup)}')
        except Exception as e:
            app.logger.warning(f'[WARNING] Backup kontrol hatasÃƒÂ„Ã‚Â±: {os.path.basename(latest_backup)} - {e}')
            
    except Exception as e:
        app.logger.error(f'[ERROR] Backup integrity check hatasÃƒÂ„Ã‚Â±: {e}')

def cleanup_old_backups(backup_dir, keep=30):
    """Eski backup'larÃƒÂ„Ã‚Â± temizle - Disk alanÃƒÂ„Ã‚Â± tasarrufu"""
    try:
        if not os.path.exists(backup_dir):
            return
            
        backups = sorted([
            os.path.join(backup_dir, f) 
            for f in os.listdir(backup_dir) 
            if f.startswith('envanter_backup_') and f.endswith('.db')
        ], key=os.path.getmtime, reverse=True)
        
        deleted_count = 0
        for old_backup in backups[keep:]:
            try:
                size_mb = round(os.path.getsize(old_backup) / (1024**2), 2)
                os.remove(old_backup)
                deleted_count += 1
                app.logger.info(f'[OK] Eski backup silindi: {os.path.basename(old_backup)} ({size_mb} MB)')
            except Exception as e:
                app.logger.error(f'[ERROR] Backup silme hatasÃƒÂ„Ã‚Â±: {old_backup} - {e}')
        
        if deleted_count > 0:
            app.logger.info(f'[OK] {deleted_count} eski backup silindi (Son {keep} tutuldu)')
            
    except Exception as e:
        app.logger.error(f'[ERROR] Backup temizleme hatasÃƒÂ„Ã‚Â±: {e}')

def get_backup_list():
    """Mevcut backup'larÃƒÂ„Ã‚Â± listele"""
    try:
        backup_dir = 'backups'
        if not os.path.exists(backup_dir):
            return []
            
        backups = []
        for f in os.listdir(backup_dir):
            if f.startswith('envanter_backup_') and f.endswith('.db'):
                full_path = os.path.join(backup_dir, f)
                backups.append({
                    'filename': f,
                    'path': full_path,
                    'size_mb': round(os.path.getsize(full_path) / (1024**2), 2),
                    'created_at': datetime.fromtimestamp(os.path.getmtime(full_path)).isoformat()
                })
        
        # En yeniden eskiye sÃƒÂ„Ã‚Â±rala
        backups.sort(key=lambda x: x['created_at'], reverse=True)
        return backups
    except Exception as e:
        app.logger.error(f'[ERROR] Backup listesi hatasÃƒÂ„Ã‚Â±: {e}')
        return []

def restore_database(backup_filename):
    """Backup'tan veritabanÃƒÂ„Ã‚Â±nÃƒÂ„Ã‚Â± geri yÃƒÂƒÃ‚Â¼kle"""
    try:
        backup_path = f'backups/{backup_filename}'
        db_path = 'instance/envanter_local.db'
        
        # Dosya varlÃƒÂ„Ã‚Â±ÃƒÂ„Ã…Â¸ÃƒÂ„Ã‚Â± kontrol et
        if not os.path.exists(backup_path):
            app.logger.error(f'[ERROR] Backup dosyasÃƒÂ„Ã‚Â± bulunamadÃƒÂ„Ã‚Â±: {backup_path}')
            return False, 'Backup dosyasÃƒÂ„Ã‚Â± bulunamadÃƒÂ„Ã‚Â±'
        
        # Mevcut db'nin yedeÃƒÂ„Ã…Â¸ini al
        import shutil
        if os.path.exists(db_path):
            recovery_backup = f'backups/emergency_recovery_{datetime.now().strftime("%Y%m%d_%H%M%S")}.db'
            shutil.copy2(db_path, recovery_backup)
            app.logger.info(f'[RECOVERY] Kurtarma backup oluÃƒÂ…Ã…Â¸turuldu: {recovery_backup}')
        
        # Backup'ÃƒÂ„Ã‚Â± geri yÃƒÂƒÃ‚Â¼kle
        shutil.copy2(backup_path, db_path)
        
        app.logger.warning(f'[RESTORE] Database geri yÃƒÂƒÃ‚Â¼klendi:')
        app.logger.warning(f'   [SRC] Kaynak: {backup_path}')
        app.logger.warning(f'   [DST] Hedef: {db_path}')
        
        return True, f'Database baÃƒÂ…Ã…Â¸arÃƒÂ„Ã‚Â±yla geri yÃƒÂƒÃ‚Â¼klendi: {backup_filename}'
        
    except Exception as e:
        app.logger.error(f'[ERROR] Restore hatasÃƒÂ„Ã‚Â±: {e}')
        import traceback
        app.logger.error(traceback.format_exc())
        return False, f'Restore hatasÃƒÂ„Ã‚Â±: {str(e)}'

@app.route('/admin/backup_now', methods=['POST'])
@login_required
def backup_now():
    """Manuel backup tetikle - sadece admin"""
    # Admin kontrolÃƒÂƒÃ‚Â¼
    if session.get('role') != 'admin':
        return jsonify({'error': 'Admin eriÃƒÂ…Ã…Â¸imi gerekli'}), 403
    
    try:
        app.logger.info('[ADMIN] Admin manuel backup baÃƒÂ…Ã…Â¸lattÃƒÂ„Ã‚Â±')
        backup_path = backup_database()
        if backup_path:
            return jsonify({
                'success': True, 
                'backup_path': backup_path,
                'message': '[OK] Backup baÃƒÂ…Ã…Â¸arÃƒÂ„Ã‚Â±yla oluÃƒÂ…Ã…Â¸turuldu'
            })
        else:
            return jsonify({
                'success': False, 
                'error': '[ERROR] Backup oluÃƒÂ…Ã…Â¸turulamadÃƒÂ„Ã‚Â±'
            }), 500
    except Exception as e:
        app.logger.error(f'[ERROR] Backup endpoint hatasÃƒÂ„Ã‚Â±: {e}')
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/admin/backups', methods=['GET'])
@login_required
def list_backups():
    """Backup listesini getir - sadece admin"""
    # Admin kontrolÃƒÂƒÃ‚Â¼
    if session.get('role') != 'admin':
        return jsonify({'error': 'Admin eriÃƒÂ…Ã…Â¸imi gerekli'}), 403
    
    try:
        backups = get_backup_list()
        backup_dir = os.path.abspath('backups')
        
        return jsonify({
            'success': True,
            'backups': backups,
            'count': len(backups),
            'backup_dir': backup_dir,
            'total_size_mb': sum(b['size_mb'] for b in backups)
        })
    except Exception as e:
        app.logger.error(f'[ERROR] Backup listesi endpoint hatasÃƒÂ„Ã‚Â±: {e}')
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/admin/restore_backup/<filename>', methods=['POST'])
@login_required
def restore_backup(filename):
    """Backup'tan geri yÃƒÂƒÃ‚Â¼kle - sadece admin - ÃƒÂƒÃ¢Â€Â¡OK DÃƒÂ„Ã‚Â°KKATLÃƒÂ„Ã‚Â°!"""
    # Admin kontrolÃƒÂƒÃ‚Â¼
    if session.get('role') != 'admin':
        return jsonify({'error': 'Admin eriÃƒÂ…Ã…Â¸imi gerekli'}), 403
    
    try:
        # GÃƒÂƒÃ‚Â¼venlik: YalnÃƒÂ„Ã‚Â±zca backup dosyalarÃƒÂ„Ã‚Â±
        if not filename.startswith('envanter_backup_') or not filename.endswith('.db'):
            return jsonify({'success': False, 'error': 'GeÃƒÂƒÃ‚Â§ersiz backup dosyasÃƒÂ„Ã‚Â±'}), 400
        
        success, message = restore_database(filename)
        
        if success:
            app.logger.warning(f'[RESTORE-ADMIN] Admin database restore yaptÃƒÂ„Ã‚Â±: {filename}')
            return jsonify({
                'success': True,
                'message': f'[OK] {message}'
            })
        else:
            return jsonify({
                'success': False,
                'error': f'[ERROR] {message}'
            }), 400
            
    except Exception as e:
        app.logger.error(f'[ERROR] Restore endpoint hatasÃƒÂ„Ã‚Â±: {e}')
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/admin/backup_status', methods=['GET'])
@login_required
def backup_status():
    """Backup sistem durumu - sadece admin"""
    if session.get('role') != 'admin':
        return jsonify({'error': 'Admin eriÃƒÂ…Ã…Â¸imi gerekli'}), 403
    
    try:
        backups = get_backup_list()
        db_path = 'instance/envanter_local.db'
        db_size = 0
        
        if os.path.exists(db_path):
            db_size = round(os.path.getsize(db_path) / (1024**2), 2)
        
        last_backup = None
        if backups:
            last_backup = backups[0]
        
        return jsonify({
            'success': True,
            'current_db_size_mb': db_size,
            'total_backups': len(backups),
            'total_backup_size_mb': sum(b['size_mb'] for b in backups),
            'last_backup': last_backup,
            'backup_dir': os.path.abspath('backups'),
            'schedule': {
                'daily_backup': 'Her gÃƒÂƒÃ‚Â¼n 02:00\'de',
                'hourly_check': 'Her saat baÃƒÂ…Ã…Â¸ÃƒÂ„Ã‚Â±nda'
            }
        })
    except Exception as e:
        app.logger.error(f'[ERROR] Backup status endpoint hatasÃƒÂ„Ã‚Â±: {e}')
        return jsonify({'success': False, 'error': str(e)}), 500
        return jsonify({'error': 'Admin eriÃƒÂ…Ã…Â¸imi gerekli'}), 403
    
    try:
        # GÃƒÂƒÃ‚Â¼venlik: sadece backup klasÃƒÂƒÃ‚Â¶rÃƒÂƒÃ‚Â¼ndeki dosyalar
        backup_path = os.path.join('backups', secure_filename(filename))
        
        if not os.path.exists(backup_path):
            return jsonify({'success': False, 'error': 'Backup dosyasÃƒÂ„Ã‚Â± bulunamadÃƒÂ„Ã‚Â±'}), 404
        
        if not filename.startswith('envanter_backup_'):
            return jsonify({'success': False, 'error': 'GeÃƒÂƒÃ‚Â§ersiz backup dosyasÃƒÂ„Ã‚Â±'}), 400
        
        # Mevcut DB'yi yedekle (geri dÃƒÂƒÃ‚Â¶nÃƒÂƒÃ‚Â¼ÃƒÂ…Ã…Â¸ iÃƒÂƒÃ‚Â§in)
        current_backup = backup_database()
        
        # Backup'ÃƒÂ„Ã‚Â± geri yÃƒÂƒÃ‚Â¼kle
        db_path = 'instance/envanter_local.db'
        import shutil
        shutil.copy2(backup_path, db_path)
        
        app.logger.warning(f'[WARNING] DATABASE RESTORE: {filename} by {session.get("username")}')
        
        return jsonify({
            'success': True,
            'message': 'Backup geri yÃƒÂƒÃ‚Â¼klendi. LÃƒÂƒÃ‚Â¼tfen uygulamayÃƒÂ„Ã‚Â± yeniden baÃƒÂ…Ã…Â¸latÃƒÂ„Ã‚Â±n.',
            'restored_from': filename,
            'safety_backup': current_backup
        })
    except Exception as e:
        app.logger.error(f'Restore hatasÃƒÂ„Ã‚Â±: {e}')
        return jsonify({'success': False, 'error': str(e)}), 500

# ========================================
#  PAKET YÃƒÂƒÃ¢Â€Â“NETÃƒÂ„Ã‚Â°MÃƒÂ„Ã‚Â° API ENDPOINT'LERÃƒÂ„Ã‚Â°
# ========================================

@app.route('/api/create_package', methods=['POST'])
def create_package():
    """Yeni paket oluÃƒÂ…Ã…Â¸turma - Tek QR ile birden fazla parÃƒÂƒÃ‚Â§a"""
    try:
        data = request.get_json()
        package_name = data.get('package_name', '').strip()
        package_desc = data.get('package_desc', '').strip()
        items = data.get('items', [])
        
        if not package_name:
            return jsonify({'success': False, 'error': 'Paket adÃƒÂ„Ã‚Â± gerekli'}), 400
        
        if not items or len(items) == 0:
            return jsonify({'success': False, 'error': 'En az 1 parÃƒÂƒÃ‚Â§a ekleyin'}), 400
        
        # Paket kodunun zaten var olup olmadÃƒÂ„Ã‚Â±ÃƒÂ„Ã…Â¸ÃƒÂ„Ã‚Â±nÃƒÂ„Ã‚Â± kontrol et
        conn = get_db()
        cursor = conn.cursor()
        
        cursor.execute('SELECT part_code FROM part_codes WHERE part_code = ?', (package_name,))
        if cursor.fetchone():
            conn.close()
            return jsonify({'success': False, 'error': 'Bu paket adÃƒÂ„Ã‚Â±/QR kodu zaten kullanÃƒÂ„Ã‚Â±mda'}), 400
        
        # JSON formatÃƒÂ„Ã‚Â±nda paket iÃƒÂƒÃ‚Â§eriÃƒÂ„Ã…Â¸ini hazÃƒÂ„Ã‚Â±rla
        package_items_json = json.dumps(items)
        
        # QR kod resmi oluÃƒÂ…Ã…Â¸tur (Base64)
        qr = qrcode.QRCode(version=1, box_size=10, border=5)
        qr.add_data(package_name)
        qr.make(fit=True)
        
        img = qr.make_image(fill_color="black", back_color="white")
        buffer = BytesIO()
        img.save(buffer, format='PNG')
        qr_base64 = base64.b64encode(buffer.getvalue()).decode()
        
        # Paketi veritabanÃƒÂ„Ã‚Â±na ekle
        cursor.execute('''
            INSERT INTO part_codes (part_code, part_name, is_package, package_items)
            VALUES (?, ?, ?, ?)
        ''', (package_name, package_desc if package_desc else f'Paket: {package_name}', True, package_items_json))
        
        package_id = cursor.lastrowid
        
        # QR codes tablosuna da ekle (tek bir QR)
        cursor.execute('''
            INSERT INTO qr_codes (qr_id, part_code_id, is_used, created_at)
            VALUES (?, ?, ?, ?)
        ''', (package_name, package_id, False, datetime.now()))
        
        conn.commit()
        conn.close()
        
        app.logger.info(f'[PAKET] Yeni paket oluÃƒÂ…Ã…Â¸turuldu: {package_name} ({len(items)} parÃƒÂƒÃ‚Â§a)')
        
        return jsonify({
            'success': True,
            'message': f'Paket "{package_name}" oluÃƒÂ…Ã…Â¸turuldu',
            'package_code': package_name,
            'qr_image': qr_base64,  # Base64 QR kodu
            'items_count': len(items),
            'total_quantity': sum(item.get('quantity', 1) for item in items)
        })
        
    except Exception as e:
        app.logger.error(f'[HATA] Paket oluÃƒÂ…Ã…Â¸turma hatasÃƒÂ„Ã‚Â±: {e}')
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/get_packages', methods=['GET'])
def get_packages():
    """TÃƒÂƒÃ‚Â¼m paketleri listele"""
    try:
        conn = get_db()
        cursor = conn.cursor()
        
        cursor.execute('''
            SELECT part_code, part_name, package_items
            FROM part_codes
            WHERE is_package = 1 OR is_package = 'true'
            ORDER BY part_code
        ''')
        
        packages = []
        for row in cursor.fetchall():
            try:
                items = json.loads(row[2]) if row[2] else []
            except:
                items = []
            
            # Her paket iÃƒÂƒÃ‚Â§in QR kod oluÃƒÂ…Ã…Â¸tur
            qr = qrcode.QRCode(version=1, box_size=10, border=5)
            qr.add_data(row[0])  # part_code
            qr.make(fit=True)
            
            img = qr.make_image(fill_color="black", back_color="white")
            buffer = BytesIO()
            img.save(buffer, format='PNG')
            qr_base64 = base64.b64encode(buffer.getvalue()).decode()
            
            packages.append({
                'part_code': row[0],
                'part_name': row[1],
                'items': items,
                'qr_image': qr_base64  # QR kod base64
            })
        
        conn.close()
        
        return jsonify({
            'success': True,
            'packages': packages,
            'count': len(packages)
        })
        
    except Exception as e:
        app.logger.error(f'[HATA] Paket listeleme hatasÃƒÂ„Ã‚Â±: {e}')
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/delete_package', methods=['POST'])
def delete_package():
    """Paket silme"""
    try:
        data = request.get_json()
        part_code = data.get('part_code', '').strip()
        
        if not part_code:
            return jsonify({'success': False, 'error': 'Paket kodu gerekli'}), 400
        
        conn = get_db()
        cursor = conn.cursor()
        
        # Paketin varlÃƒÂ„Ã‚Â±ÃƒÂ„Ã…Â¸ÃƒÂ„Ã‚Â±nÃƒÂ„Ã‚Â± kontrol et
        cursor.execute('SELECT is_package FROM part_codes WHERE part_code = ?', (part_code,))
        row = cursor.fetchone()
        
        if not row:
            return jsonify({'success': False, 'error': 'Paket bulunamadÃƒÂ„Ã‚Â±'}), 404
        
        if not row[0]:
            return jsonify({'success': False, 'error': 'Bu bir paket deÃƒÂ„Ã…Â¸il'}), 400
        
        # Paketi sil
        cursor.execute('DELETE FROM part_codes WHERE part_code = ?', (part_code,))
        conn.commit()
        conn.close()
        
        app.logger.info(f'[PAKET] Paket silindi: {part_code}')
        
        return jsonify({
            'success': True,
            'message': f'Paket "{part_code}" silindi'
        })
        
    except Exception as e:
        app.logger.error(f'[HATA] Paket silme hatasÃƒÂ„Ã‚Â±: {e}')
        return jsonify({'success': False, 'error': str(e)}), 500

#  CANLI DASHBOARD API - GerÃƒÂƒÃ‚Â§ek zamanlÃƒÂ„Ã‚Â± istatistikler
@app.route('/api/live_dashboard/<int:session_id>')
@login_required
def live_dashboard_stats(session_id):
    """
    CanlÃƒÂ„Ã‚Â± dashboard iÃƒÂƒÃ‚Â§in tÃƒÂƒÃ‚Â¼m istatistikler
    
    Returns:
        - Tarama hÃƒÂ„Ã‚Â±zÃƒÂ„Ã‚Â± (son 5dk, son 1 saat)
        - En ÃƒÂƒÃ‚Â§ok taranan 10 parÃƒÂƒÃ‚Â§a
        - KullanÃƒÂ„Ã‚Â±cÃƒÂ„Ã‚Â± performansÃƒÂ„Ã‚Â±
        - SayÃƒÂ„Ã‚Â±m bitiÃƒÂ…Ã…Â¸ tahmini
        - Saatlik tarama daÃƒÂ„Ã…Â¸ÃƒÂ„Ã‚Â±lÃƒÂ„Ã‚Â±mÃƒÂ„Ã‚Â±
    """
    # Session'dan kullanÃƒÂ„Ã‚Â±cÃƒÂ„Ã‚Â± adÃƒÂ„Ã‚Â±nÃƒÂ„Ã‚Â± al
    username = session.get('username', 'Bilinmeyen')
    logging.info(f"Dashboard API called - Session: {session_id}, User: {username}")
    
    try:
        with db_connection() as conn:
            cursor = conn.cursor()
            
            # 1. Tarama HÃƒÂ„Ã‚Â±zÃƒÂ„Ã‚Â± (Scans per Minute)
            execute_query(cursor, '''
                SELECT 
                    COUNT(*) as total_scans,
                    MIN(scanned_at) as first_scan,
                    MAX(scanned_at) as last_scan
                FROM scanned_qr
                WHERE session_id = ?
            ''', (session_id,))
            scan_stats = cursor.fetchone()
            
            total_scans = scan_stats[0] if scan_stats else 0
            first_scan = scan_stats[1] if scan_stats else None
            last_scan = scan_stats[2] if scan_stats else None
            
            logging.debug(f"Total scans: {total_scans}, First: {first_scan}, Last: {last_scan}")
            
            # Tarama hÃƒÂ„Ã‚Â±zÃƒÂ„Ã‚Â± hesapla
            scan_rate_5min = 0
            scan_rate_1hour = 0
            
            if last_scan:
                # Son 5 dakika
                execute_query(cursor, '''
                    SELECT COUNT(*) FROM scanned_qr
                    WHERE session_id = ? AND scanned_at >= datetime('now', '-5 minutes')
                ''', (session_id,))
                scans_5min = cursor.fetchone()[0]
                scan_rate_5min = round(scans_5min / 5, 1)  # tarama/dakika
                
                # Son 1 saat
                execute_query(cursor, '''
                    SELECT COUNT(*) FROM scanned_qr
                    WHERE session_id = ? AND scanned_at >= datetime('now', '-1 hour')
                ''', (session_id,))
                scans_1hour = cursor.fetchone()[0]
                scan_rate_1hour = round(scans_1hour / 60, 1)  # tarama/dakika
            
            # 2. En ÃƒÂƒÃ¢Â€Â¡ok Taranan 10 ParÃƒÂƒÃ‚Â§a
            execute_query(cursor, '''
                SELECT 
                    pc.part_name,
                    pc.part_code,
                    COUNT(*) as scan_count
                FROM scanned_qr sq
                JOIN part_codes pc ON sq.part_code = pc.part_code
                WHERE sq.session_id = ?
                GROUP BY pc.part_code, pc.part_name
                ORDER BY scan_count DESC
                LIMIT 10
            ''', (session_id,))
            top_parts = [{'name': row[0], 'code': row[1], 'count': row[2]} for row in cursor.fetchall()]
            
            # 3. KullanÃƒÂ„Ã‚Â±cÃƒÂ„Ã‚Â± PerformansÃƒÂ„Ã‚Â±
            execute_query(cursor, '''
                SELECT 
                    u.full_name,
                    u.username,
                    COUNT(*) as scan_count,
                    MIN(sq.scanned_at) as first_scan,
                    MAX(sq.scanned_at) as last_scan
                FROM scanned_qr sq
                JOIN envanter_users u ON sq.scanned_by = u.id
                WHERE sq.session_id = ?
                GROUP BY u.id, u.full_name, u.username
                ORDER BY scan_count DESC
            ''', (session_id,))
            
            user_performance = []
            for row in cursor.fetchall():
                user_first = datetime.fromisoformat(row[3]) if row[3] else None
                user_last = datetime.fromisoformat(row[4]) if row[4] else None
                work_time_minutes = 0
                personal_rate = 0
                
                if user_first and user_last:
                    work_time_seconds = (user_last - user_first).total_seconds()
                    work_time_minutes = round(work_time_seconds / 60, 1)
                    if work_time_minutes > 0:
                        personal_rate = round(row[2] / work_time_minutes, 1)
                
                user_performance.append({
                    'name': row[0],
                    'username': row[1],
                    'scan_count': row[2],
                    'work_time_minutes': work_time_minutes,
                    'scans_per_minute': personal_rate
                })
            
            # 4. SayÃƒÂ„Ã‚Â±m BitiÃƒÂ…Ã…Â¸ Tahmini
            execute_query(cursor, '''
                SELECT COUNT(DISTINCT qr_id) FROM qr_codes
            ''')
            total_qr_codes = cursor.fetchone()[0]
            remaining = total_qr_codes - total_scans
            
            eta_minutes = 0
            eta_text = "HesaplanÃƒÂ„Ã‚Â±yor..."
            
            if scan_rate_1hour > 0 and remaining > 0:
                eta_minutes = round(remaining / scan_rate_1hour)
                if eta_minutes < 60:
                    eta_text = f"{eta_minutes} dakika"
                else:
                    eta_hours = round(eta_minutes / 60, 1)
                    eta_text = f"{eta_hours} saat"
            
            # 5. Saatlik Tarama DaÃƒÂ„Ã…Â¸ÃƒÂ„Ã‚Â±lÃƒÂ„Ã‚Â±mÃƒÂ„Ã‚Â± (Son 24 saat)
            execute_query(cursor, '''
                SELECT 
                    strftime('%H', scanned_at) as hour,
                    COUNT(*) as count
                FROM scanned_qr
                WHERE session_id = ? AND scanned_at >= datetime('now', '-24 hours')
                GROUP BY hour
                ORDER BY hour
            ''', (session_id,))
            
            hourly_distribution = {}
            for row in cursor.fetchall():
                hourly_distribution[row[0]] = row[1]
            
            result = {
                'success': True,
                'session_id': session_id,
                'total_scans': total_scans,
                'remaining': remaining,
                'progress_percent': round((total_scans / total_qr_codes * 100) if total_qr_codes > 0 else 0, 1),
                'scan_rate': {
                    'last_5min': scan_rate_5min,
                    'last_1hour': scan_rate_1hour,
                    'unit': 'tarama/dakika'
                },
                'top_parts': top_parts,
                'user_performance': user_performance,
                'eta': {
                    'minutes': eta_minutes,
                    'text': eta_text
                },
                'hourly_distribution': hourly_distribution,
                'first_scan': first_scan,
                'last_scan': last_scan
            }
            
            logging.info(f"Dashboard API success - {total_scans} scans, {len(user_performance)} users")
            return jsonify(result)
            
    except Exception as e:
        logging.error(f"Live dashboard error: {e}", exc_info=True)
        return jsonify({'success': False, 'error': str(e)}), 500

# Otomatik backup scheduler iÃƒÂƒÃ‚Â§in (APScheduler kullanarak)
try:
    from apscheduler.schedulers.background import BackgroundScheduler
    
    backup_scheduler = BackgroundScheduler()
    
    #  GÃƒÂƒÃ…Â“NLÃƒÂƒÃ…Â“K BACKUP: Her gÃƒÂƒÃ‚Â¼n 02:00'de yapÃƒÂ„Ã‚Â±lÃƒÂ„Ã‚Â±r
    backup_scheduler.add_job(
        func=backup_database,
        trigger="cron",
        hour=2,
        minute=0,
        id='daily_auto_backup',
        name='GÃƒÂƒÃ‚Â¼nlÃƒÂƒÃ‚Â¼k Otomatik Database Backup',
        replace_existing=True
    )
    
    #  SAATLÃƒÂ„Ã‚Â°K KONTROL: Her saat baÃƒÂ…Ã…Â¸ÃƒÂ„Ã‚Â±nda backup kontrolÃƒÂƒÃ‚Â¼
    backup_scheduler.add_job(
        func=verify_backup_integrity,
        trigger="cron",
        minute=0,
        id='hourly_backup_check',
        name='Saatlik Backup BÃƒÂƒÃ‚Â¼tÃƒÂƒÃ‚Â¼nlÃƒÂƒÃ‚Â¼ÃƒÂ„Ã…Â¸ÃƒÂƒÃ‚Â¼ Kontrol',
        replace_existing=True
    )
    
    backup_scheduler.start()
    app.logger.info('[OK] Backup Scheduler BaÃƒÂ…Ã…Â¸latÃƒÂ„Ã‚Â±ldÃƒÂ„Ã‚Â±:')
    app.logger.info('   [DAILY] GÃƒÂƒÃ‚Â¼nlÃƒÂƒÃ‚Â¼k Backup: Her gÃƒÂƒÃ‚Â¼n 02:00\'de')
    app.logger.info('   [HOURLY] Saatlik Kontrol: Her saat baÃƒÂ…Ã…Â¸ÃƒÂ„Ã‚Â±nda')
    app.logger.info(f'   [PATH] Backup KlasÃƒÂƒÃ‚Â¶rÃƒÂƒÃ‚Â¼: {os.path.abspath("backups")}')
except ImportError:
    app.logger.warning('[!] APScheduler yÃƒÂƒÃ‚Â¼klÃƒÂƒÃ‚Â¼ deÃƒÂ„Ã…Â¸il. Otomatik backup devre dÃƒÂ„Ã‚Â±ÃƒÂ…Ã…Â¸ÃƒÂ„Ã‚Â±.')
    app.logger.warning('   YÃƒÂƒÃ‚Â¼klemek iÃƒÂƒÃ‚Â§in: pip install apscheduler')
except Exception as e:
    app.logger.error(f'Backup scheduler hatasÃƒÂ„Ã‚Â±: {e}')

# NOTE: This is handled by render_startup_alt.py for Render.com
# DO NOT call socketio.run() here to avoid port binding conflicts
# The render_startup_alt.py script will import this app and call socketio.run()

if __name__ == '__main__':
    print("ÃƒÂ¢Ã…Â¡Ã‚Â ÃƒÂ¯Ã‚Â¸Ã‚Â�  Direct execution detected. Please use render_startup_alt.py")
    print("    Or set RENDER=false for local development with Flask")

    # Initialize database on startup
    try:
        init_db()
    except Exception as e:
        print(f"ÃƒÂ¢Ã‚Â�Ã…Â’ Failed to initialize database: {e}")

    # For local testing only - DO NOT use in production
    if not os.environ.get('RENDER'):
        port = 5002
        print(" Starting EnvanterQR System v2.0 (LOCAL)...")
        print(" Dashboard: http://localhost:5002")
        print(" Admin Panel: http://localhost:5002/admin")
        print(" Health Check: http://localhost:5002/health")
        print(" Metrics: http://localhost:5002/metrics")
        print("ÃƒÂ¢Ã‹ÂœÃ‚Â�ÃƒÂ¯Ã‚Â¸Ã‚Â� Storage: Backblaze B2 Enabled")
        print(" Security: Headers + Rate Limiting Active")
        print()
        
        # Network eriÃƒÂ…Ã…Â¸imi iÃƒÂƒÃ‚Â§in host ayarÃƒÂ„Ã‚Â±
        # '0.0.0.0' = TÃƒÂƒÃ‚Â¼m aÃƒÂ„Ã…Â¸ arayÃƒÂƒÃ‚Â¼zlerinden eriÃƒÂ…Ã…Â¸ime izin ver (WiFi, Ethernet, vb.)
        # '127.0.0.1' = Sadece localhost (kendi PC'nden)
        socketio.run(app, host='0.0.0.0', port=port, debug=True, allow_unsafe_werkzeug=True)
    else:
        print("ERROR: This should not run with RENDER=true")
        print("Use: gunicorn wsgi:app --worker-class eventlet")
        print("Or:  python render_startup_alt.py")
