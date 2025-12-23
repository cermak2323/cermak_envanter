from flask import Flask, render_template, request, jsonify, send_file, session, redirect, url_for

from flask_socketio import SocketIO, emit

from werkzeug.utils import secure_filename

from functools import wraps, lru_cache

import time

from collections import defaultdict

import sqlite3

import pandas as pd

import qrcode

from PIL import Image, ImageDraw, ImageFont

from io import BytesIO

import base64

import os

import uuid

from datetime import datetime, timedelta

import zipfile

import re

import hashlib

import secrets

import random

import string

from dotenv import load_dotenv

from contextlib import contextmanager

import unicodedata

import logging

import threading

import json



# Load environment variables

load_dotenv()



# SQLAlchemy ve Models - Veritaban ORM

from models import db, PartCode, QRCode, CountSession, ScannedQR, User, CountPassword

from db_config import DevelopmentConfig, ProductionConfig



# Logging Configuration

from logging.handlers import RotatingFileHandler

import os



# Log klasr olutur

os.makedirs('logs', exist_ok=True)



# Loglama ayarlar

logging.basicConfig(

    level=logging.INFO,

    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',

    handlers=[

        RotatingFileHandler('logs/app.log', maxBytes=10*1024*1024, backupCount=5),  # 10MB, 5 backup

        RotatingFileHandler('logs/security.log', maxBytes=5*1024*1024, backupCount=3),  # Security events

        logging.StreamHandler()

    ]

)



# Security logger

security_logger = logging.getLogger('security')

security_handler = RotatingFileHandler('logs/security.log', maxBytes=5*1024*1024, backupCount=3)

security_handler.setFormatter(logging.Formatter('%(asctime)s - SECURITY - %(levelname)s - %(message)s'))

security_logger.addHandler(security_handler)

security_logger.setLevel(logging.WARNING)



app = Flask(__name__)

app.config['SECRET_KEY'] = os.environ.get('SESSION_SECRET', 'dev-secret-key-change-in-production')

app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024



#  LOKAL SSTEM - Her zaman SQLite + Local Storage

IS_PRODUCTION = False

IS_LOCAL = True

USE_B2_STORAGE = False

USE_POSTGRESQL = False



print(f"\n LOKAL SISTEM")

print(f" Database: SQLite (instance/envanter_local.db)")

print(f" Storage: Local Files (static/qrcodes/)")

print(f" Data: Yerel dosya sisteminde")



# db_config.py kullan

app.config.from_object(DevelopmentConfig)



print("\n" + "="*60)

print(" LOKAL DEPOLAMA MODU")

print("="*60)

print(" Tm veriler yerel SQLite veritabannda")

print(" QR kodlar static/qrcodes/ klasrnde")

print(" Excel dosyalar static/exports/ klasrnde")

print("="*60)

print()



# SQLAlchemy'yi app'e bala

db.init_app(app)



# Uygulama balang zaman (uptime iin)

app.config['START_TIME'] = time.time()



# Static dosya sktrma iin

app.config['SEND_FILE_MAX_AGE_DEFAULT'] = 31536000  # 1 yl cache



# SocketIO - Lokal sistem iin

socketio = SocketIO(

    app, 

    cors_allowed_origins="*",

    async_mode='eventlet',

    ping_timeout=60,

    ping_interval=25,

    max_http_buffer_size=1000000

)



# ======================

# MOBIL PERFORMANS OPTIMIZASYONLARI

# ======================



@app.before_request

def qr_folder_security():

    """QR klasr gvenlik kontrol"""

    #  QR KLASR GVENL - ifre ile korumal (admin paneli ifresi ile ayn)

    if request.path.startswith('/static/qrcodes/'):

        # Session'da qrcodes_unlocked anahtar yoksa veya False ise eriimi engelle

        if not session.get('qrcodes_unlocked', False):

            return '''

            <html>

            <head>

                <title>QR Klasr - Eriim Engellendi</title>

                <style>

                    body { font-family: Arial; text-align: center; padding: 50px; background: #f5f5f5; }

                    .container { background: white; padding: 30px; border-radius: 10px; max-width: 400px; margin: auto; box-shadow: 0 2px 10px rgba(0,0,0,0.1); }

                    h2 { color: #d32f2f; }

                    input { padding: 10px; width: 100%; margin: 10px 0; border: 1px solid #ddd; border-radius: 5px; }

                    button { padding: 10px 20px; background: #1976d2; color: white; border: none; border-radius: 5px; cursor: pointer; width: 100%; }

                    button:hover { background: #1565c0; }

                </style>

            </head>

            <body>

                <div class="container">

                    <h2> QR Klasr Korumal</h2>

                    <p>Bu klasre erimek iin admin paneli ifresini girin:</p>

                    <form method="POST" action="/unlock_qrcodes">

                        <input type="password" name="password" placeholder="Admin Paneli ifresi" required>

                        <button type="submit">Kilidi A</button>

                    </form>

                </div>

            </body>

            </html>

            ''', 403



@app.after_request

def add_performance_headers(response):

    """Performans iin header'lar ekle"""

    # Static dosyalar iin cache

    if request.endpoint == 'static':

        response.cache_control.max_age = 31536000  # 1 yl

        response.cache_control.public = True



    # Dier dosyalar iin

    else:

        response.cache_control.no_cache = True

        response.cache_control.must_revalidate = True



    # Sktrma header'

    if response.status_code == 200 and response.content_length and response.content_length > 1024:

        response.headers['Vary'] = 'Accept-Encoding'



    return response



# ======================

# PERFORMANS CACHE SISTEMI

# ======================



# Bellek tabanl cache (production'da Redis kullanlmal)

cache_store = {}

cache_lock = threading.Lock()

CACHE_TTL = 300  # 5 dakika cache sresi



def cache_get(key):

    """Cache'den veri al"""

    with cache_lock:

        if key in cache_store:

            data, timestamp = cache_store[key]

            if time.time() - timestamp < CACHE_TTL:

                return data

            else:

                del cache_store[key]

    return None



def cache_set(key, value):

    """Cache'e veri kaydet"""

    with cache_lock:

        cache_store[key] = (value, time.time())



def cache_delete(key):

    """Cache'den veri sil"""

    with cache_lock:

        if key in cache_store:

            del cache_store[key]



def cache_clear():

    """Tm cache'i temizle"""

    with cache_lock:

        cache_store.clear()



#  QR Image Memory Cache - LRU ile disk I/O azaltma

@lru_cache(maxsize=1000)

def generate_qr_pil_image(qr_id, box_size=8, border=2):

    """

    QR kod PIL Image olutur - Barkod makinesi iin optimize

    

    Scanner Specs:

    - Minimum znrlk 2D: 8.7mil (0.22mm)

    - Okuma Mesafesi: 55-350mm

    - Sensr: 640x480 piksel

    

    QR Boyutlar:

    - box_size=8 -> ~240x240px QR (8.7mil znrlk iin ideal)

    - border=2 -> Minimal ereve (tarayc quiet zone iin)

    - Toplam: ~255x275px (ereve + text ile)

    

    Args:

        qr_id: QR kod ID

        box_size: QR kutucuk boyutu (8 = barkod makinesi iin optimize)

        border: Quiet zone (2 = minimal, tarayc iin yeterli)

    

    Returns:

        PIL.Image: QR kod grseli (krmz ereveli)

    """

    qr = qrcode.QRCode(

        version=1,

        box_size=box_size,  # 8px - barkod makinesi iin optimize

        border=border,      # 2px quiet zone - tarayc iin minimum

        error_correction=qrcode.constants.ERROR_CORRECT_M  # M seviyesi (15% hata tolerans)

    )

    qr.add_data(qr_id)

    qr.make(fit=True)

    qr_img = qr.make_image(fill_color="black", back_color="white")

    qr_img = qr_img.convert('RGB')

    

    # QR boyutlar

    qr_width, qr_height = qr_img.size

    text_height = 20  # Text alan (14pt font iin)

    border_width = 3  # Krmz ereve

    

    # Krmz ereveli grsel

    final_width = qr_width + (border_width * 2)

    final_height = qr_height + text_height + (border_width * 2)

    final_img = Image.new('RGB', (final_width, final_height), '#dc2626')

    

    # Beyaz i alan

    white_bg = Image.new('RGB', (qr_width, qr_height + text_height), 'white')

    white_bg.paste(qr_img, (0, 0))

    final_img.paste(white_bg, (border_width, border_width))

    

    # Text ekle

    draw = ImageDraw.Draw(final_img)

    try:

        font = ImageFont.truetype("arialbd.ttf", 14)

    except:

        try:

            font = ImageFont.truetype("arial.ttf", 14)

        except:

            font = ImageFont.load_default()

    

    qr_text = qr_id

    text_width = len(qr_text) * 8

    x_position = max(border_width, (final_width - text_width) // 2)

    draw.text((x_position, qr_height + border_width + 2), qr_text, fill='black', font=font)

    

    return final_img



# Cache temizleme thread'i

def cache_cleanup():

    """Eski cache verilerini temizle"""

    while True:

        time.sleep(60)  # Her dakika kontrol et

        current_time = time.time()

        with cache_lock:

            expired_keys = [

                key for key, (_, timestamp) in cache_store.items()

                if current_time - timestamp > CACHE_TTL

            ]

            for key in expired_keys:

                del cache_store[key]



# Cache temizleme thread'ini balat

cleanup_thread = threading.Thread(target=cache_cleanup, daemon=True)

cleanup_thread.start()



# Rate limiting iin IP tabanl takip

login_attempts = defaultdict(list)



def add_security_headers(response):

    """Gvenlik header'larn ekle"""

    response.headers['X-Content-Type-Options'] = 'nosniff'

    response.headers['X-Frame-Options'] = 'DENY'

    response.headers['X-XSS-Protection'] = '1; mode=block'

    response.headers['Strict-Transport-Security'] = 'max-age=31536000; includeSubDomains'

    response.headers['Content-Security-Policy'] = "default-src 'self'; script-src 'self' 'unsafe-inline' https://cdn.jsdelivr.net https://cdn.socket.io https://unpkg.com; style-src 'self' 'unsafe-inline' https://cdn.jsdelivr.net https://fonts.googleapis.com; font-src 'self' https://cdn.jsdelivr.net https://fonts.gstatic.com; img-src 'self' data:;"

    return response



@app.after_request

def security_headers(response):

    return add_security_headers(response)



def rate_limit_login(f):

    """Login denemelerini snrla"""

    @wraps(f)

    def decorated_function(*args, **kwargs):

        client_ip = request.environ.get('HTTP_X_FORWARDED_FOR', request.environ.get('REMOTE_ADDR', '127.0.0.1'))

        current_time = time.time()



        # Son 15 dakikadaki denemeleri filtrele

        login_attempts[client_ip] = [t for t in login_attempts[client_ip] if current_time - t < 900]



        # 15 dakikada 5'ten fazla deneme varsa engelle

        if len(login_attempts[client_ip]) >= 5:

            return jsonify({'error': 'ok fazla login denemesi. 15 dakika bekleyin.'}), 429



        # Denemeyi kaydet

        login_attempts[client_ip].append(current_time)



        return f(*args, **kwargs)

    return decorated_function



# ----------------------

# Configuration

# ----------------------



# Project root

BASE_DIR = os.path.abspath(os.path.dirname(__file__))



# ======================

# VERTABANI YAPILANDI - SADECE SQLite

# ======================



# Lokal SQLite database

DATABASE_URL = app.config['SQLALCHEMY_DATABASE_URI']

print(f" Local SQLite: {DATABASE_URL}")



REPORTS_DIR = 'reports'



def generate_strong_password():

    """Gl parola olutur (8 karakter: byk harf, kk harf, rakam, zel karakter)"""

    characters = string.ascii_uppercase + string.ascii_lowercase + string.digits + "!@#$%^&*"

    # En az 1 byk harf, 1 kk harf, 1 rakam, 1 zel karakter olacak ekilde

    password = [

        random.choice(string.ascii_uppercase),  # Byk harf

        random.choice(string.ascii_lowercase),  # Kk harf

        random.choice(string.digits),           # Rakam

        random.choice("!@#$%^&*")              # zel karakter

    ]

    # Kalan 4 karakteri rastgele se

    for _ in range(4):

        password.append(random.choice(characters))



    # Kartr

    random.shuffle(password)

    return ''.join(password)



def generate_count_password():

    """Saym iin parola olutur (6 haneli sadece say) - Basit ve hzl giri iin"""

    return ''.join([str(random.randint(0, 9)) for _ in range(6)])



def save_qr_code_to_file(part_code, qr_id, qr_number):

    """

    QR kodunu fiziksel dosya olarak kaydet

    - Her para iin ayr klasr oluturur (rn: static/qr_codes/Y129513-14532/)

    - QR kodlarn numaral kaydeder (rn: Y129513-14532_1.png)

    - QR kod altna kod ve para ad yazar

    """

    try:

        import qrcode

        

        # Para adn database'den al

        conn = get_db()

        cursor = conn.cursor()

        execute_query(cursor, 'SELECT part_name FROM part_codes WHERE part_code = ', (part_code,))

        result = cursor.fetchone()

        part_name = result[0] if result else ""

        close_db(conn)

        

        # QR klasr ana dizini

        qrcodes_base_dir = os.path.join(os.path.dirname(__file__), 'static', 'qr_codes')

        

        # Para iin klasr olutur (rn: static/qr_codes/Y129513-14532/)

        part_dir = os.path.join(qrcodes_base_dir, part_code)

        os.makedirs(part_dir, exist_ok=True)

        

        # Dosya ad: part_code_number.png (rn: Y129513-14532_1.png)

        filename = f"{part_code}_{qr_number}.png"

        filepath = os.path.join(part_dir, filename)

        

        # QR kod olutur - Barkod makinesi iin optimize (8.7mil znrlk)

        qr = qrcode.QRCode(

            version=1,

            error_correction=qrcode.constants.ERROR_CORRECT_M,  # M seviyesi (15% hata tolerans)

            box_size=8,  # 8px - barkod makinesi iin ideal (~240x240px)

            border=2,    # 2px quiet zone - tarayc minimum gereksinimi

        )

        qr.add_data(qr_id)

        qr.make(fit=True)

        

        # QR kodunu olutur

        qr_img = qr.make_image(fill_color="black", back_color="white")

        qr_img = qr_img.convert('RGB')  # PIL Image'a dntr

        

        # QR kod boyutlarn al

        qr_width, qr_height = qr_img.size

        

        # Alanlar hesapla

        logo_height = 40  # Logo iin st alan

        text_height = 35  # Alt yaz (para numaras) iin alan

        

        # Krmz ereve iin padding

        border_width = 3  # 3px krmz ereve

        

        # Yeni grsel olutur (logo + QR + text alan + ereve)

        final_width = qr_width + (border_width * 2)

        final_height = logo_height + qr_height + text_height + (border_width * 2)

        final_img = Image.new('RGB', (final_width, final_height), '#dc2626')  # Krmz arka plan (ereve)

        

        # Beyaz i alan olutur (logo + QR + text)

        white_bg = Image.new('RGB', (qr_width, logo_height + qr_height + text_height), 'white')

        

        # Logo ekle (varsa) - st ortasna

        try:

            logo_path = os.path.join(os.path.dirname(__file__), 'cermak-logo.png')

            if os.path.exists(logo_path):

                logo_img = Image.open(logo_path).convert('RGBA')

                # Logo boyutunu alan yksekliine gre ayarla

                logo_width = 150

                logo_height_logo = 40

                try:

                    logo_img = logo_img.resize((logo_width, logo_height_logo), Image.Resampling.LANCZOS)

                except AttributeError:

                    logo_img = logo_img.resize((logo_width, logo_height_logo), Image.LANCZOS)

                

                # Logo'yu ortala

                logo_x = (qr_width - logo_width) // 2

                logo_y = 5  # stten 5px boluk

                

                # RGBA logo'yu blend et

                if logo_img.mode == 'RGBA':

                    alpha = logo_img.split()[3]

                    logo_img = logo_img.convert('RGB')

                    white_bg.paste(logo_img, (logo_x, logo_y), alpha)

                else:

                    white_bg.paste(logo_img, (logo_x, logo_y))

        except Exception as e:

            print(f"Logo ekleme hatas: {e}")

        

        # QR kodu beyaz alana yaptr (logo'nun altna)

        white_bg.paste(qr_img, (0, logo_height))

        

        # Beyaz alan krmz erevenin iine yaptr

        final_img.paste(white_bg, (border_width, border_width))

        

        # Text ekleme iin draw nesnesi

        draw = ImageDraw.Draw(final_img)

        

        # Font (kaln ve byk)

        try:

            # Windows iin Arial Black (en kaln), en byk font

            font = ImageFont.truetype("arialblk.ttf", 24)

        except:

            try:

                font = ImageFont.truetype("arialbd.ttf", 24)

            except:

                try:

                    font = ImageFont.truetype("arial.ttf", 24)

                except:

                    font = ImageFont.load_default()

        

        # QR ID yazs (Y129150-49811_1) - Sadece bu

        qr_text = f"{part_code}_{qr_number}"

        

        # Text geniliini hesapla (24pt font iin)

        text_width = len(qr_text) * 14  # 24pt font iin genilik

        

        # QR ID'yi ortala ve yaz (border_width offset ekle)

        x_position = max(border_width, (final_width - text_width) // 2)

        draw.text((x_position, logo_height + qr_height + border_width + 5), qr_text, fill='black', font=font)

        

        # Kaydet

        final_img.save(filepath)

        

        print(f" QR kod kaydedildi: static/qr_codes/{part_code}/{filename}")

        return filepath

        

    except Exception as e:

        print(f" QR kod kaydetme hatas: {e}")

        return None



# Admin saym ifresi

ADMIN_COUNT_PASSWORD = "@R9t$L7e!xP2w"

print(f"DEBUG: ADMIN_COUNT_PASSWORD = '{ADMIN_COUNT_PASSWORD}'")  # DEBUG



os.makedirs(REPORTS_DIR, exist_ok=True)



# Dosya upload iin ayarlar

UPLOAD_FOLDER = 'static/part_photos'

ALLOWED_EXTENSIONS = {'png', 'jpg', 'jpeg', 'gif', 'webp'}

app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER

app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024  # 16MB limit



def allowed_file(filename):

    """Dosya uzantsn kontrol et"""

    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS



def get_db_placeholder():

    """Database'e gre doru placeholder dndr - SQLite iin  kullanr"""

    return ''



def execute_query(cursor, query, params=None):

    """

    Execute SQL query - SQLite iin  placeholder kullanr

    """

    # SQLite:  ->  dntr (eski kodlardan gelebilecek 'leri deitir)

    query = query.replace('', '')



    if params:

        cursor.execute(query, params)

    else:

        cursor.execute(query)



    return cursor



def get_db():

    """Get database connection - SQLite only (local system)"""

    #  LOKAL: SQLite balants

    db_path = os.path.join('instance', 'envanter_local.db')

    os.makedirs('instance', exist_ok=True)

    conn = sqlite3.connect(db_path)

    conn.row_factory = sqlite3.Row

    return conn



def close_db(conn):

    """Close SQLite connection"""

    if conn:

        try:

            conn.close()

        except Exception as e:

            logging.debug(f"Error closing SQLite connection: {e}")



@contextmanager

def db_connection():

    """

    Database connection context manager - Otomatik cleanup

    

    Kullanm:

        with db_connection() as conn:

            cursor = conn.cursor()

            execute_query(cursor, 'SELECT ...')

            conn.commit()

        # conn otomatik kapanr, hata olsa bile

    """

    conn = get_db()

    try:

        yield conn

    except sqlite3.Error as e:

        logging.error(f"Database error: {e}")

        if conn:

            conn.rollback()

        raise

    except Exception as e:

        logging.error(f"Unexpected error in database operation: {e}")

        if conn:

            conn.rollback()

        raise

    finally:

        close_db(conn)



def create_performance_indexes(cursor):

    """Performans iin kritik indexleri olutur"""

    indexes = [

        # QR Codes indexleri

        "CREATE INDEX IF NOT EXISTS idx_qr_codes_qr_id ON qr_codes(qr_id);",

        "CREATE INDEX IF NOT EXISTS idx_qr_codes_part_code ON qr_codes(part_code_id);",

        "CREATE INDEX IF NOT EXISTS idx_qr_codes_is_used ON qr_codes(is_used);",



        # Count Sessions indexleri

        "CREATE INDEX IF NOT EXISTS idx_count_sessions_is_active ON count_sessions(is_active);",

        "CREATE INDEX IF NOT EXISTS idx_count_sessions_created_by ON count_sessions(created_by);",

        "CREATE INDEX IF NOT EXISTS idx_count_sessions_created_at ON count_sessions(created_at);",



        # Scanned QR indexleri - COMPOSITE INDEX N NEML

        "CREATE INDEX IF NOT EXISTS idx_scanned_qr_session_id ON scanned_qr(session_id);",

        "CREATE INDEX IF NOT EXISTS idx_scanned_qr_qr_id ON scanned_qr(qr_id);",

        "CREATE INDEX IF NOT EXISTS idx_scanned_qr_scanned_by ON scanned_qr(scanned_by);",

        "CREATE INDEX IF NOT EXISTS idx_scanned_qr_scanned_at ON scanned_qr(scanned_at);",

        

        #  COMPOSITE INDEX - Duplicate check iin kritik

        "CREATE INDEX IF NOT EXISTS idx_scanned_qr_session_qr ON scanned_qr(session_id, qr_id);",



        # Part Codes indexleri

        "CREATE INDEX IF NOT EXISTS idx_part_codes_part_code ON part_codes(part_code);",



        # Kullanc indexleri

        "CREATE INDEX IF NOT EXISTS idx_envanter_users_username ON envanter_users(username);",

        "CREATE INDEX IF NOT EXISTS idx_envanter_users_is_active ON envanter_users(is_active_user);",

    ]



    try:

        for index_sql in indexes:

            execute_query(cursor, index_sql)

        print(" Performance indexes created/verified")

    except Exception as e:

        print(f" Warning: Could not create some indexes: {e}")



def init_db_part_details():

    """Part codes tablosuna detay kolonlar ekle"""

    try:

        conn = get_db()

        cursor = conn.cursor()

        

        # Yeni kolonlar - Geriye uyumlu (ALTER TABLE)

        new_columns = [

            ('photo_path', 'TEXT'),           # Para fotoraf

            ('catalog_image', 'TEXT'),        # Katalog grseli

            ('description', 'TEXT'),          # Aklama

            ('used_in_machines', 'TEXT'),     # JSON: Kullanld makineler

            ('specifications', 'TEXT'),       # JSON: Teknik zellikler

            ('stock_location', 'TEXT'),       # Stok konumu

            ('supplier', 'TEXT'),             # Tedariki

            ('unit_price', 'REAL'),          # Birim fiyat

            ('notes', 'TEXT')                # Notlar

        ]

        

        for col_name, col_type in new_columns:

            try:

                cursor.execute(f'ALTER TABLE part_codes ADD COLUMN {col_name} {col_type}')

                print(f" Kolon eklendi: {col_name}")

            except sqlite3.OperationalError as e:

                if 'duplicate column name' in str(e).lower():

                    print(f" Kolon zaten var: {col_name}")

                else:

                    raise

        

        conn.commit()

        close_db(conn)

        print(" Part details schema gncellendi")

        

    except Exception as e:

        print(f" Part details schema hatas: {e}")

        try:

            if 'conn' in locals():

                close_db(conn)

        except:

            pass



def init_db():

    """Initialize database tables using SQLAlchemy ORM



    Lokal: SQLite tables olutur

    Production: PostgreSQL tables kontrol et

    """

    try:

        with app.app_context():

            if USE_POSTGRESQL:

                # PRODUCTION: PostgreSQL - SQLAlchemy ile

                print("\n" + "="*70)

                print(" INITIALIZING POSTGRESQL TABLES")

                print("="*70)



                inspector = db.inspect(db.engine)

                existing_tables = inspector.get_table_names()

                print(f"Existing tables: {existing_tables}")



                required_tables = ['envanter_users', 'part_codes', 'qr_codes', 'count_sessions', 'count_passwords', 'scanned_qr']



                missing_tables = [t for t in required_tables if t not in existing_tables]



                if missing_tables:

                    print(f" Creating missing PostgreSQL tables: {', '.join(missing_tables)}")

                    db.create_all()

                    print(" PostgreSQL tables created successfully")

                else:

                    print(" All PostgreSQL tables already exist")



                #  DATABASE MIGRATION: Add finished_by column to count_sessions

                try:

                    conn = get_db()

                    cursor = conn.cursor()



                    # Check if finished_by column exists

                    execute_query(cursor, """

                        SELECT column_name 

                        FROM information_schema.columns 

                        WHERE table_name='count_sessions' AND column_name='finished_by'

                    """)



                    if not cursor.fetchone():

                        print(" Adding finished_by column to count_sessions table...")

                        execute_query(cursor, """

                            ALTER TABLE count_sessions 

                            ADD COLUMN finished_by INTEGER REFERENCES envanter_users(id)

                        """)

                        conn.commit()

                        print(" finished_by column added successfully")

                    else:

                        print(" finished_by column already exists")



                    close_db(conn)

                except Exception as e:

                    print(f" Migration warning: {e}")



                #  DATABASE MIGRATION: Add total_expected and total_scanned columns to count_sessions

                try:

                    conn = get_db()

                    cursor = conn.cursor()



                    # Check total_expected

                    execute_query(cursor, """

                        SELECT column_name FROM information_schema.columns

                        WHERE table_name='count_sessions' AND column_name='total_expected'

                    """)

                    if not cursor.fetchone():

                        print(" Adding total_expected column to count_sessions table...")

                        execute_query(cursor, """

                            ALTER TABLE count_sessions

                            ADD COLUMN total_expected INTEGER DEFAULT 0

                        """)

                        conn.commit()

                        print(" total_expected column added successfully")

                    else:

                        print(" total_expected column already exists")



                    # Check total_scanned

                    execute_query(cursor, """

                        SELECT column_name FROM information_schema.columns

                        WHERE table_name='count_sessions' AND column_name='total_scanned'

                    """)

                    if not cursor.fetchone():

                        print(" Adding total_scanned column to count_sessions table...")

                        execute_query(cursor, """

                            ALTER TABLE count_sessions

                            ADD COLUMN total_scanned INTEGER DEFAULT 0

                        """)

                        conn.commit()

                        print(" total_scanned column added successfully")

                    else:

                        print(" total_scanned column already exists")



                    #  DATABASE MIGRATION: Add report_file_path column to count_sessions

                    cursor = conn.cursor()

                    cursor.execute("""

                        SELECT COUNT(*) FROM pragma_table_info('count_sessions') 

                        WHERE name='report_file_path'

                    """)

                    if cursor.fetchone()[0] == 0:

                        print(" Adding report_file_path column to count_sessions table...")

                        cursor.execute("""

                            ALTER TABLE count_sessions

                            ADD COLUMN report_file_path TEXT

                        """)

                        conn.commit()

                        print(" report_file_path column added successfully")

                    else:

                        print(" report_file_path column already exists")



                    close_db(conn)

                except Exception as e:

                    print(f" Migration warning (total_*): {e}")



                # Verify scanned_qr table specifically (critical for duplicate detection)

                if 'scanned_qr' in existing_tables or 'scanned_qr' not in missing_tables:

                    print(" scanned_qr table verified - duplicate detection will work")

                else:

                    print(" WARNING: scanned_qr table not found - duplicate detection may fail!")

                    print("   Creating scanned_qr table manually...")

                    try:

                        # Create scanned_qr table manually if SQLAlchemy fails

                        conn = get_db()

                        cursor = conn.cursor()

                        execute_query(cursor, '''

                            CREATE TABLE IF NOT EXISTS scanned_qr (

                                id SERIAL PRIMARY KEY,

                                session_id VARCHAR(255) NOT NULL,

                                qr_id VARCHAR(255) NOT NULL,

                                part_code VARCHAR(255) NOT NULL,

                                scanned_by INTEGER,

                                scanned_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP

                            )

                        ''')

                        conn.commit()

                        close_db(conn)

                        print(" scanned_qr table created successfully")

                    except Exception as e:

                        print(f" Failed to create scanned_qr: {e}")



                # SQLAlchemy User model kullan

                admin_user = User.query.filter_by(username='admin').first()

                if not admin_user:

                    print(" Creating default PostgreSQL admin user...")

                    from werkzeug.security import generate_password_hash

                    admin_password = generate_password_hash("@R9t$L7e!xP2w")

                    admin = User(

                        username='admin',

                        full_name='Administrator',

                        password_hash=admin_password,

                        role='admin',

                        is_active_user=True

                    )

                    db.session.add(admin)

                    db.session.commit()

                    print(" PostgreSQL admin user created (admin/@R9t$L7e!xP2w)")

                else:

                    print(" PostgreSQL admin user already exists")



                print("="*70 + "\n")

                # LOCAL: SQLite - Raw SQL ile (basit tablo yaps)

                print(" Local SQLite mode - checking simple table structure")



                # SQLite balants al

                conn = get_db()

                cursor = conn.cursor()



                # SQLite schema'y gncelle (full_name column ekle)

                try:

                    execute_query(cursor, "ALTER TABLE envanter_users ADD COLUMN full_name VARCHAR(255)")

                    print(" Added full_name column to SQLite")

                except sqlite3.OperationalError:

                    # Column zaten varsa veya baka hata

                    pass



                # Admin user var m kontrol et (SQLite raw SQL)

                execute_query(cursor, "SELECT * FROM envanter_users WHERE username = 'admin'")

                admin_exists = cursor.fetchone()



                if not admin_exists:

                    print(" Creating default SQLite admin user...")

                    from werkzeug.security import generate_password_hash

                    admin_password_hash = generate_password_hash("admin123")

                    execute_query(cursor, '''

                        INSERT INTO envanter_users (username, password_hash, full_name, role, created_at, is_active)

                        VALUES (, , , , , )

                    ''', ('admin', admin_password_hash, 'Administrator', 'admin', datetime.now(), True))

                    conn.commit()

                    print(" SQLite admin user created (admin/admin123)")

                else:

                    print(" SQLite admin user already exists")



                close_db(conn)

                print(" SQLite database initialized successfully")

            

            # Part details kolonlarn ekle

            init_db_part_details()



            return True



    except Exception as e:

        print(f" Database initialization error: {str(e)}")

        import traceback

        traceback.print_exc()

        return False



def init_db_part_details():

    """Part codes tablosuna detay kolonlar ekle"""

    try:

        conn = get_db()

        cursor = conn.cursor()

        

        new_columns = [

            ('photo_path', 'TEXT'),

            ('catalog_image', 'TEXT'),

            ('description', 'TEXT'),

            ('used_in_machines', 'TEXT'),  # JSON array

            ('specifications', 'TEXT'),     # JSON object

            ('stock_location', 'TEXT'),

            ('supplier', 'TEXT'),

            ('unit_price', 'REAL'),

            ('critical_stock_level', 'INTEGER'),

            ('notes', 'TEXT'),

            ('last_updated', 'TIMESTAMP'),

            ('updated_by', 'INTEGER')

        ]

        

        for col_name, col_type in new_columns:

            try:

                cursor.execute(f'ALTER TABLE part_codes ADD COLUMN {col_name} {col_type}')

                print(f" Kolon eklendi: {col_name}")

            except sqlite3.OperationalError as e:

                if 'duplicate column name' in str(e).lower():

                    print(f" Kolon zaten var: {col_name}")

                else:

                    raise

        

        conn.commit()

        close_db(conn)

        print(" Part details schema updated")

    except Exception as e:

        print(f" Part details schema error: {e}")



def login_required(f):

    from functools import wraps

    @wraps(f)

    def decorated_function(*args, **kwargs):

        if 'user_id' not in session:

            return jsonify({'error': 'Giri yapmanz gerekiyor'}), 401

        return f(*args, **kwargs)

    return decorated_function



def admin_required(f):

    from functools import wraps

    @wraps(f)

    def decorated_function(*args, **kwargs):

        if 'user_id' not in session or session.get('role') != 'admin':

            return redirect('/admin')

        return f(*args, **kwargs)

    return decorated_function



def admin_required_decorator(f):

    """Sadece admin eriimi iin decorator"""

    from functools import wraps

    @wraps(f)

    def decorated_function(*args, **kwargs):

        if 'user_id' not in session:

            return jsonify({'error': 'Giri yapmanz gerekiyor'}), 401

        if session.get('role') != 'admin':

            return jsonify({'error': 'Bu ilem iin admin yetkisi gerekli'}), 403

        return f(*args, **kwargs)

    return decorated_function



@app.route('/favicon.ico')

def favicon():

    return '', 204



# ================================

# PART DETAILS API ENDPOINTS

# ================================



@app.route('/api/upload_part_photo/<part_code>', methods=['POST'])

@login_required

def upload_part_photo(part_code):

    """Para fotoraf ykle"""

    if 'photo' not in request.files:

        return jsonify({'error': 'Dosya bulunamad'}), 400

    

    file = request.files['photo']

    if file.filename == '':

        return jsonify({'error': 'Dosya seilmedi'}), 400

    

    if file and allowed_file(file.filename):

        # Klasr yoksa olutur

        os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)

        

        # Gvenli dosya ad

        filename = secure_filename(file.filename)

        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')

        new_filename = f"{part_code}_{timestamp}_{filename}"

        filepath = os.path.join(app.config['UPLOAD_FOLDER'], new_filename)

        

        file.save(filepath)

        

        # Database'e kaydet

        conn = get_db()

        cursor = conn.cursor()

        placeholder = get_db_placeholder()

        

        execute_query(cursor, f'''

            UPDATE part_codes 

            SET photo_path = {placeholder}, 

                last_updated = {placeholder},

                updated_by = {placeholder}

            WHERE part_code = {placeholder}

        ''', (f'part_photos/{new_filename}', datetime.now(), session['user_id'], part_code))

        

        conn.commit()

        close_db(conn)

        

        return jsonify({

            'success': True,

            'photo_path': f'part_photos/{new_filename}'

        })

    

    return jsonify({'error': 'Geersiz dosya format (PNG, JPG, GIF, WEBP)'}), 400



@app.route('/api/upload_catalog_image/<part_code>', methods=['POST'])

@login_required

def upload_catalog_image(part_code):

    """Katalog grnts ykle"""

    if 'catalog' not in request.files:

        return jsonify({'error': 'Dosya bulunamad'}), 400

    

    file = request.files['catalog']

    if file.filename == '':

        return jsonify({'error': 'Dosya seilmedi'}), 400

    

    if file and allowed_file(file.filename):

        os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)

        

        filename = secure_filename(file.filename)

        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')

        new_filename = f"{part_code}_catalog_{timestamp}_{filename}"

        filepath = os.path.join(app.config['UPLOAD_FOLDER'], new_filename)

        

        file.save(filepath)

        

        conn = get_db()

        cursor = conn.cursor()

        placeholder = get_db_placeholder()

        

        execute_query(cursor, f'''

            UPDATE part_codes 

            SET catalog_image = {placeholder}, 

                last_updated = {placeholder},

                updated_by = {placeholder}

            WHERE part_code = {placeholder}

        ''', (f'part_photos/{new_filename}', datetime.now(), session['user_id'], part_code))

        

        conn.commit()

        close_db(conn)

        

        return jsonify({

            'success': True,

            'catalog_image': f'part_photos/{new_filename}'

        })

    

    return jsonify({'error': 'Geersiz dosya format'}), 400



@app.route('/api/parts', methods=['GET'])

@login_required

def get_all_parts():

    """Tm paralar listele (admin panel iin)"""

    conn = get_db()

    cursor = conn.cursor()

    

    execute_query(cursor, '''

        SELECT part_code, description, photo_path, catalog_image

        FROM part_codes

        ORDER BY part_code

    ''')

    

    parts = cursor.fetchall()

    close_db(conn)

    

    result = []

    for part in parts:

        result.append({

            'part_code': part[0],

            'description': part[1],

            'photo_path': part[2],

            'catalog_image': part[3]

        })

    

    return jsonify(result)



@app.route('/api/part_details/<part_code>', methods=['GET'])

def get_part_details(part_code):

    """Para detaylarn getir (herkes grebilir)"""

    conn = get_db()

    cursor = conn.cursor()

    placeholder = get_db_placeholder()

    

    execute_query(cursor, f'''

        SELECT part_code, description, photo_path, catalog_image,

               used_in_machines, specifications, stock_location,

               supplier, unit_price, critical_stock_level, notes,

               last_updated, updated_by

        FROM part_codes

        WHERE part_code = {placeholder}

    ''', (part_code,))

    

    part = cursor.fetchone()

    close_db(conn)

    

    if not part:

        return jsonify({'error': 'Para bulunamad'}), 404

    

    return jsonify({

        'part_code': part[0],

        'description': part[1],

        'photo_path': part[2],

        'catalog_image': part[3],

        'used_in_machines': json.loads(part[4]) if part[4] else [],

        'specifications': json.loads(part[5]) if part[5] else {},

        'stock_location': part[6],

        'supplier': part[7],

        'unit_price': part[8],

        'critical_stock_level': part[9],

        'notes': part[10],

        'last_updated': part[11],

        'updated_by': part[12]

    })



@app.route('/api/update_part_details/<part_code>', methods=['POST'])

@login_required

def update_part_details(part_code):

    """Para detaylarn gncelle (admin only)"""

    if session.get('role') != 'admin':

        return jsonify({'error': 'Admin yetkisi gerekli'}), 403

    

    data = request.json

    conn = get_db()

    cursor = conn.cursor()

    placeholder = get_db_placeholder()

    

    # Machines ve specs JSON olarak kaydet

    used_in_machines = json.dumps(data.get('used_in_machines', []))

    specifications = json.dumps(data.get('specifications', {}))

    

    execute_query(cursor, f'''

        UPDATE part_codes

        SET description = {placeholder},

            used_in_machines = {placeholder},

            specifications = {placeholder},

            stock_location = {placeholder},

            supplier = {placeholder},

            unit_price = {placeholder},

            critical_stock_level = {placeholder},

            notes = {placeholder},

            last_updated = {placeholder},

            updated_by = {placeholder}

        WHERE part_code = {placeholder}

    ''', (

        data.get('description'),

        used_in_machines,

        specifications,

        data.get('stock_location'),

        data.get('supplier'),

        data.get('unit_price'),

        data.get('critical_stock_level'),

        data.get('notes'),

        datetime.now(),

        session['user_id'],

        part_code

    ))

    

    conn.commit()

    close_db(conn)

    

    return jsonify({'success': True})



@app.route('/qr/<qr_id>')

def qr_redirect(qr_id):

    """QR Code ynlendirme - Aktif saym varsa scanner, yoksa part info"""

    conn = get_db()

    cursor = conn.cursor()

    placeholder = get_db_placeholder()

    

    # Aktif saym var m

    execute_query(cursor, f'SELECT id FROM count_sessions WHERE is_active = {placeholder}', (True,))

    active_session = cursor.fetchone()

    

    if active_session:

        # Aktif saym var -> Scanner sayfas

        close_db(conn)

        return redirect(url_for('scanner'))

    else:

        # Aktif saym yok -> Part detay sayfas

        # QR'dan part_code bul

        execute_query(cursor, f'''

            SELECT pc.part_code

            FROM qr_codes qr

            JOIN part_codes pc ON qr.part_code_id = pc.id

            WHERE qr.qr_id = {placeholder}

        ''', (qr_id,))

        

        result = cursor.fetchone()

        close_db(conn)

        

        if result:

            return redirect(url_for('part_info', part_code=result[0]))

        else:

            return "QR Code bulunamad", 404



@app.route('/part_info/<part_code>')

def part_info(part_code):

    """Para bilgi sayfas"""

    return render_template('part_info.html', part_code=part_code)



@app.route('/edit_part/<part_code>')

@admin_required

def edit_part(part_code):

    """Para dzenleme sayfas (admin)"""

    return render_template('edit_part.html', part_code=part_code)



@app.route('/part_search')

def part_search():

    """Para arama sayfas"""

    return render_template('part_search.html')



@app.route('/api/search_parts')

def search_parts():

    """Para arama API'si - TR Klavye uyumlu"""

    query = request.args.get('q', '').strip()

    

    if not query or len(query) < 1:

        return jsonify([])

    

    try:

        # Trke karakterleri normalize et ve sembol uyumazln dzelt

        def normalize_turkish(text):

            """Trke karakterleri ve sembol hatalarn dzelt"""

            if not text:

                return text

            

            # SADECE TR klavyede yanl gelen sembolleri dzelt

            # * ve  -> - veya _ (hangisi varsa onu koru)

            # Son -X veya _X ksmn kaldr (adet says)

            symbol_map = {

                '*': '-',      # Shift + / gives *

                '': '-',      # Other combination  

            }

            

            # Sembol haritasn uygula

            for wrong_char, correct_char in symbol_map.items():

                text = text.replace(wrong_char, correct_char)

            

            # Son -X veya _X ksmn kaldr (adet says)

            # rnek: Y129150-49811-3 -> Y129150-49811

            import re

            text = re.sub(r'[-_]\d+$', '', text)

            

            # NFD normalizasyonu (decompose) - Trke karakterler

            text = unicodedata.normalize('NFD', text.lower())

            # Accent iaretlerini kaldr

            output = []

            for char in text:

                if unicodedata.category(char) != 'Mn':  # Mn = Mark, Nonspacing

                    output.append(char)

            result = ''.join(output)

            

            # Manuel TR karakterler (eer normalize etmezse)

            tr_map = {

                '': 'c', '': 's', '': 'g', '': 'u', '': 'o', '': 'i',

                '': 'c', '': 's', '': 'g', '': 'u', '': 'o', '': 'i'

            }

            for tr_char, en_char in tr_map.items():

                result = result.replace(tr_char, en_char)

            

            return result

        

        normalized_query = normalize_turkish(query)

        

        conn = get_db()

        cursor = conn.cursor()

        placeholder = get_db_placeholder()

        

        # lk olarak tam kod elemesini ara (QR format: Y129150-49811)

        execute_query(cursor, f'''

            SELECT part_code, description

            FROM part_codes

            WHERE part_code = {placeholder}

            LIMIT 1

        ''', (normalized_query,))

        

        results = cursor.fetchall()

        

        # Eer tam eleme yoksa, LIKE ile ara (hem normalized hem normal)

        if not results:

            search_pattern = f'%{normalized_query}%'

            execute_query(cursor, f'''

                SELECT part_code, description

                FROM part_codes

                WHERE part_code LIKE {placeholder} 

                   OR description LIKE {placeholder}

                LIMIT 20

            ''', (search_pattern, search_pattern))

            results = cursor.fetchall()

        

        close_db(conn)

        

        print(f'[Search API] Original: {query} | Normalized: {normalized_query} | Found: {len(results)}')

        

        return jsonify([{

            'part_code': row[0],

            'description': row[1]

        } for row in results])

    

    except Exception as e:

        print(f'Search error: {e}')

        return jsonify([])



# ================================

# EXISTING ROUTES BELOW

# ================================



@app.route('/')

def index():

    """Ana sayfa - Dashboard"""

    if 'user_id' not in session:

        return render_template('login.html')

    

    return render_template('index.html')





@app.route('/dashboard')

@login_required

def dashboard():

    """Dashboard - Ana sayfaya ynlendir"""

    return redirect('/')





@app.route('/clear_cache.html')

def clear_cache_page():

    """Cache temizleme yardmc sayfas"""

    with open('clear_cache.html', 'r', encoding='utf-8') as f:

        return f.read()





# Note: a more comprehensive /health endpoint is defined later in this file.

# The detailed health check includes DB and storage checks and will be used by

# load balancers and Render readiness probes.



@app.route('/api/dashboard/stats')

@login_required

def dashboard_stats():

    """Cache'li dashboard istatistikleri"""

    cache_key = 'dashboard_stats'



    # Cache'den kontrol et

    cached_data = cache_get(cache_key)

    if cached_data:

        return jsonify(cached_data)



    # Cache yoksa veritabanndan al

    conn = get_db()

    cursor = conn.cursor()



    try:

        # Toplam QR kod says

        # execute_query(cursor, 'SELECT COUNT(*) FROM envanter')

        total_qr_codes = cursor.fetchone()[0]



        # Toplam saym says

        db.session.query(ScannedQR).count())

        total_scans = cursor.fetchone()[0]



        # Aktif oturum says

        execute_query(cursor, 'SELECT COUNT(*) FROM count_sessions WHERE is_active = ', (True,))

        active_sessions = cursor.fetchone()[0]



        # Son saym tarihi

        execute_query(cursor, 'SELECT MAX(scanned_at) FROM scanned_qr')

        last_scan = cursor.fetchone()[0]



        # Bugnk saymlar

        execute_query(cursor, '''

            SELECT COUNT(*) FROM scanned_qr 

            WHERE DATE(scanned_at) = CURRENT_DATE

        ''')

        today_scans = cursor.fetchone()[0]



        # Bu haftaki saymlar

        execute_query(cursor, '''

            SELECT COUNT(*) FROM scanned_qr 

            WHERE scanned_at >= CURRENT_DATE - INTERVAL '7 days'

        ''')

        week_scans = cursor.fetchone()[0]



        stats = {

            'total_qr_codes': total_qr_codes,

            'total_scans': total_scans,

            'active_sessions': active_sessions,

            'last_scan': last_scan.isoformat() if last_scan else None,

            'today_scans': today_scans,

            'week_scans': week_scans,

            'cache_time': datetime.now().isoformat()

        }



        # Cache'e kaydet

        cache_set(cache_key, stats)



        return jsonify(stats)



    finally:

        close_db(conn)



# ========================================

#  QR SCANNER SAYIM SSTEM (Sadece Desktop Scanner)

# ========================================



@app.route('/scanner')

@login_required

def scanner_page():

    """Desktop QR Scanner Saym Sayfas"""

    return render_template('scanner.html')



@app.route('/live_dashboard')

@login_required

def live_dashboard_page():

    """Canl Dashboard Sayfas"""

    return render_template('live_dashboard.html')



@app.route('/api/get_active_count_session')

def get_active_count_session():

    """Aktif saym seansn kontrol et - Public endpoint (landing page iin)"""

    try:

        conn = get_db()

        cursor = conn.cursor()

        

        cursor.execute('''

            SELECT id, session_name, created_at, started_at, total_expected, total_scanned, description

            FROM count_sessions 

            WHERE is_active = 1

            ORDER BY created_at DESC

            LIMIT 1

        ''')

        

        result = cursor.fetchone()

        close_db(conn)

        

        if result:

            session_id, name, created_at, started_at, expected, scanned, description = result

            

            # Sre hesapla

            start_time = datetime.fromisoformat(started_at) if started_at else datetime.fromisoformat(created_at)

            duration_seconds = (datetime.now() - start_time).total_seconds()

            

            # Yzde hesapla

            percentage = (scanned / expected * 100) if expected > 0 else 0

            

            # Beklenen paralar parse et

            expected_parts = []

            if description:

                try:

                    expected_parts = json.loads(description)

                except:

                    pass

            

            return jsonify({

                'success': True,

                'has_active': True,

                'session': {

                    'id': session_id,

                    'name': name,

                    'created_at': created_at,

                    'started_at': started_at,

                    'duration_seconds': int(duration_seconds),

                    'total_expected': expected or 0,

                    'total_scanned': scanned or 0,

                    'percentage': round(percentage, 2),

                    'expected_parts': expected_parts

                }

            })

        else:

            return jsonify({

                'success': True,

                'has_active': False

            })

        

    except Exception as e:

        print(f" Get active session error: {e}")

        return jsonify({'success': False, 'error': str(e)}), 500





@app.route('/api/get_session_report/<session_id>')

@login_required

def get_session_report(session_id):

    """Saym raporu getir - beklenen vs saylan karlatrmas"""

    try:

        conn = get_db()

        cursor = conn.cursor()

        

        # Session bilgilerini al

        cursor.execute('''

            SELECT session_name, description, total_expected, total_scanned

            FROM count_sessions 

            WHERE id = ?

        ''', (session_id,))

        

        session_info = cursor.fetchone()

        

        if not session_info:

            close_db(conn)

            return jsonify({'success': False, 'error': 'Session bulunamad'}), 404

        

        session_name, description, total_expected, total_scanned = session_info

        

        # Beklenen paralar parse et

        expected_parts = []

        if description:

            try:

                expected_parts = json.loads(description)

            except:

                pass

        

        # Saylan paralar al - direkt scanned_qr tablosundan

        cursor.execute('''

            SELECT part_code, COUNT(*) as scanned_count

            FROM scanned_qr

            WHERE session_id = 

            GROUP BY part_code

        ''', (session_id,))

        

        scanned_parts = {}

        for row in cursor.fetchall():

            part_code, count = row

            scanned_parts[part_code] = count

        

        # Part name'leri part_codes tablosundan al

        part_names = {}

        if expected_parts:

            part_codes_list = [item.get('part_code') for item in expected_parts if item.get('part_code')]

            if part_codes_list:

                placeholders = ','.join(['' for _ in part_codes_list])

                cursor.execute(f'SELECT part_code, part_name FROM part_codes WHERE part_code IN ({placeholders})', part_codes_list)

                for row in cursor.fetchall():

                    part_names[row[0]] = row[1]

        

        close_db(conn)

        

        # Raporu olutur

        report_items = []

        for expected in expected_parts:

            part_code = expected.get('part_code')

            part_name = expected.get('part_name') or part_names.get(part_code, '')

            

            # Excel'den gelen 'quantity' veya API'den gelen 'expected_quantity' - ikisini de destekle

            expected_qty = expected.get('expected_quantity') or expected.get('quantity', 0)

            

            scanned_qty = scanned_parts.get(part_code, 0)

            

            difference = scanned_qty - expected_qty

            

            if difference == 0:

                status = ' Tam'

            elif difference > 0:

                status = ' Fazla'

            else:

                status = ' Eksik'

            

            report_items.append({

                'part_code': part_code,

                'part_name': part_name,

                'expected': expected_qty,

                'scanned': scanned_qty,

                'difference': difference,

                'status': status

            })

        

        return jsonify({

            'success': True,

            'session_name': session_name,

            'total_expected': total_expected,

            'total_scanned': total_scanned,

            'report': report_items

        })

        

    except Exception as e:

        print(f" Get session report error: {e}")

        return jsonify({'success': False, 'error': str(e)}), 500





@app.route('/api/start_count_session', methods=['POST'])

@login_required

def start_count_session():

    """Yeni saym seans balat"""

    try:

        data = request.get_json()

        expected_parts = data.get('expected_parts', [])  # [{part_code, expected_quantity}]

        

        conn = get_db()

        cursor = conn.cursor()

        

        # nce aktif saym var m kontrol et

        cursor.execute('SELECT id FROM count_sessions WHERE is_active = 1')

        if cursor.fetchone():

            close_db(conn)

            return jsonify({

                'success': False,

                'error': 'Zaten aktif bir saym var! nce onu bitirin.'

            }), 400

        

        # Yeni saym seans olutur

        session_name = f"Saym {datetime.now().strftime('%Y%m%d_%H%M%S')}"

        

        # Excel'den 'quantity' veya API'den 'expected_quantity' - ikisini de destekle

        total_expected = sum(item.get('expected_quantity') or item.get('quantity', 0) for item in expected_parts)

        

        cursor.execute('''

            INSERT INTO count_sessions 

            (session_name, created_by, is_active, created_at, started_at, total_expected, total_scanned, description)

            VALUES (, , 1, , , , 0, )

        ''', (

            session_name,

            session.get('user_id', 1),

            datetime.now(),

            datetime.now(),

            total_expected,

            json.dumps(expected_parts) if expected_parts else None

        ))

        

        session_id = cursor.lastrowid

        conn.commit()

        close_db(conn)

        

        return jsonify({

            'success': True,

            'session_id': session_id,

            'session_name': session_name,

            'message': f' Saym balatld: {session_name}'

        })

        

    except Exception as e:

        print(f" Start count session error: {e}")

        return jsonify({'success': False, 'error': str(e)}), 500





@app.route('/api/finish_count', methods=['POST'])

@login_required

def finish_count():

    """Saym bitir, Excel raporunu kaydet ve rapor hazrla"""

    try:

        data = request.get_json()

        session_id = data.get('session_id', '1')

        

        conn = get_db()

        cursor = conn.cursor()

        

        # Toplam tarama saysn hesapla

        cursor.execute('''

            SELECT COUNT(*) FROM scanned_qr WHERE session_id = 

        ''', (str(session_id),))

        total_scanned = cursor.fetchone()[0]

        

        # Session bilgilerini al

        cursor.execute('''

            SELECT session_name, created_at, started_at, description, total_expected

            FROM count_sessions 

            WHERE id = ?

        ''', (str(session_id),))

        

        session_info = cursor.fetchone()

        

        if not session_info:

            close_db(conn)

            return jsonify({'success': False, 'error': 'Saym bulunamad'}), 404

        

        session_name, created_at, started_at, description, total_expected = session_info

        

        # Excel raporu olutur

        report_filename = None

        try:

            # Beklenen paralar parse et

            expected_parts = {}

            if description:

                try:

                    expected_list = json.loads(description)

                    for item in expected_list:

                        part_code = item.get('Para Kodu') or item.get('part_code')

                        expected_qty = item.get('Beklenen Adet') or item.get('expected_quantity') or item.get('quantity') or item.get('Adet') or 0

                        if part_code:

                            expected_parts[part_code] = int(expected_qty)

                except:

                    pass

            

            # Taranan paralar al

            cursor.execute('''

                SELECT sq.part_code, pc.part_name, COUNT(*) as scanned_count

                FROM scanned_qr sq

                LEFT JOIN part_codes pc ON sq.part_code = pc.part_code

                WHERE sq.session_id = 

                GROUP BY sq.part_code, pc.part_name

                ORDER BY pc.part_name

            ''', (str(session_id),))

            

            scanned_results = cursor.fetchall()

            

            # Beklenen paralar iin part_name'leri ek

            part_names = {}

            if expected_parts:

                part_codes_list = list(expected_parts.keys())

                if part_codes_list:

                    placeholders = ','.join(['' for _ in part_codes_list])

                    cursor.execute(f'SELECT part_code, part_name FROM part_codes WHERE part_code IN ({placeholders})', part_codes_list)

                    for row in cursor.fetchall():

                        part_names[row[0]] = row[1]

            

            # Tm paralar birletir (beklenen + taranan)

            all_parts = {}

            

            # Beklenen paralar ekle

            for part_code, expected_qty in expected_parts.items():

                all_parts[part_code] = {

                    'part_code': part_code,

                    'part_name': part_names.get(part_code, ''),

                    'expected': expected_qty,

                    'scanned': 0

                }

            

            # Taranan paralar ekle/gncelle

            for part_code, part_name, scanned_count in scanned_results:

                if part_code in all_parts:

                    all_parts[part_code]['part_name'] = part_name or ''

                    all_parts[part_code]['scanned'] = scanned_count

                else:

                    all_parts[part_code] = {

                        'part_code': part_code,

                        'part_name': part_name or '',

                        'expected': 0,

                        'scanned': scanned_count

                    }

            

            # DataFrame olutur

            rows = []

            for part in all_parts.values():

                difference = part['scanned'] - part['expected']

                

                if difference == 0:

                    status = ' Tam'

                elif difference > 0:

                    status = ' Fazla'

                else:

                    status = ' Eksik'

                

                rows.append({

                    'Para Kodu': part['part_code'],

                    'Para Ad': part['part_name'],

                    'Beklenen Adet': part['expected'],

                    'Saylan Adet': part['scanned'],

                    'Fark': difference,

                    'Durum': status

                })

            

            df = pd.DataFrame(rows)

            

            # Excel dosyasn kaydet

            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")

            report_filename = f'sayim_raporu_{session_name}_{timestamp}.xlsx'

            report_path = os.path.join('static', 'reports', report_filename)

            

            # static/reports klasrn olutur

            os.makedirs(os.path.join('static', 'reports'), exist_ok=True)

            

            with pd.ExcelWriter(report_path, engine='openpyxl') as writer:

                df.to_excel(writer, sheet_name='Saym Raporu', index=False)

                

                worksheet = writer.sheets['Saym Raporu']

                

                # Kolon geniliklerini ayarla

                column_widths = {

                    'A': 18,  # Para Kodu

                    'B': 30,  # Para Ad

                    'C': 15,  # Beklenen Adet

                    'D': 15,  # Saylan Adet

                    'E': 12,  # Fark

                    'F': 15   # Durum

                }

                

                for col, width in column_widths.items():

                    worksheet.column_dimensions[col].width = width

                

                # Balk satrn formatla

                from openpyxl.styles import Font, Alignment, PatternFill

                header_fill = PatternFill(start_color='DC2626', end_color='DC2626', fill_type='solid')

                header_font = Font(bold=True, color='FFFFFF', size=12)

                

                for cell in worksheet[1]:

                    cell.fill = header_fill

                    cell.font = header_font

                    cell.alignment = Alignment(horizontal='center', vertical='center')

                

                # Durum stununu renklendir

                for row_idx in range(2, len(rows) + 2):

                    cell = worksheet[f'F{row_idx}']

                    fark_cell = worksheet[f'E{row_idx}']

                    

                    fark_value = fark_cell.value

                    if fark_value == 0:

                        cell.fill = PatternFill(start_color='D4EDDA', end_color='D4EDDA', fill_type='solid')  # Yeil

                    elif fark_value < 0:

                        cell.fill = PatternFill(start_color='FFF3CD', end_color='FFF3CD', fill_type='solid')  # Sar

                    else:

                        cell.fill = PatternFill(start_color='D1ECF1', end_color='D1ECF1', fill_type='solid')  # Mavi

            

            print(f" Excel raporu kaydedildi: {report_path}")

            

        except Exception as e:

            print(f" Excel kayt hatas: {e}")

            import traceback

            traceback.print_exc()

        

        # Session' bitir ve rapor yolunu kaydet

        cursor.execute('''

            UPDATE count_sessions 

            SET is_active = 0, ended_at = , total_scanned = , report_file_path = 

            WHERE id = ?

        ''', (datetime.now(), total_scanned, report_filename, session_id))

        

        conn.commit()

        close_db(conn)

        

        return jsonify({

            'success': True,

            'message': f' Saym tamamland! {total_scanned} adet tarama kaydedildi. Rapor oluturuldu.',

            'report_filename': report_filename

        })

        

    except Exception as e:

        print(f" Finish count error: {e}")

        import traceback

        traceback.print_exc()

        return jsonify({'success': False, 'error': str(e)}), 500





@app.route('/reports')

@login_required

def reports_page():

    """Tamamlanm saym raporlarn listele"""

    return render_template('reports.html')





@app.route('/api/get_saved_reports')

@login_required

def get_saved_reports():

    """Kaydedilmi raporlar JSON olarak dndr"""

    try:

        conn = get_db()

        cursor = conn.cursor()

        

        # Tamamlanm saymlar al (is_active = 0 ve report_file_path dolu olanlar)

        cursor.execute('''

            SELECT id, session_name, created_at, ended_at, 

                   total_expected, total_scanned, report_file_path

            FROM count_sessions 

            WHERE is_active = 0 AND report_file_path IS NOT NULL

            ORDER BY ended_at DESC

        ''')

        

        sessions = cursor.fetchall()

        close_db(conn)

        

        # Tuple'lar dict'e evir

        sessions_list = []

        for s in sessions:

            sessions_list.append({

                'id': s[0],

                'session_name': s[1],

                'created_at': s[2],

                'ended_at': s[3],

                'total_expected': s[4] or 0,

                'total_scanned': s[5] or 0,

                'report_file_path': s[6]

            })

        

        return jsonify(sessions_list)

        

    except Exception as e:

        print(f" Get saved reports error: {e}")

        import traceback

        traceback.print_exc()

        return jsonify({'success': False, 'error': str(e)}), 500





@app.route('/api/download_count_excel/<session_id>')

@login_required

def download_count_excel(session_id):

    """Saym sonularn detayl Excel olarak indir (beklenen vs taranan)"""

    try:

        conn = get_db()

        cursor = conn.cursor()

        

        # Saym bilgilerini al

        cursor.execute('''

            SELECT session_name, created_at, started_at, ended_at, 

                   description, total_expected, total_scanned

            FROM count_sessions 

            WHERE id = ?

        ''', (str(session_id),))

        

        session_info = cursor.fetchone()

        

        if not session_info:

            close_db(conn)

            return jsonify({'success': False, 'error': 'Saym bulunamad'}), 404

        

        session_name, created_at, started_at, ended_at, description, total_expected, total_scanned = session_info

        

        # Beklenen paralar parse et

        expected_parts = {}

        if description:

            try:

                expected_list = json.loads(description)

                for item in expected_list:

                    part_code = item.get('Para Kodu') or item.get('part_code')

                    # Excel'den 'quantity', API'den 'expected_quantity', eski format 'Beklenen Adet'

                    expected_qty = item.get('Beklenen Adet') or item.get('expected_quantity') or item.get('quantity') or item.get('Adet') or 0

                    if part_code:

                        expected_parts[part_code] = int(expected_qty)

            except:

                pass

        

        # Taranan paralar al

        cursor.execute('''

            SELECT sq.part_code, pc.part_name, COUNT(*) as scanned_count

            FROM scanned_qr sq

            LEFT JOIN part_codes pc ON sq.part_code = pc.part_code

            WHERE sq.session_id = 

            GROUP BY sq.part_code, pc.part_name

            ORDER BY pc.part_name

        ''', (str(session_id),))

        

        scanned_results = cursor.fetchall()

        

        # Beklenen paralar iin part_name'leri ek

        part_names = {}

        if expected_parts:

            part_codes_list = list(expected_parts.keys())

            if part_codes_list:

                placeholders = ','.join(['' for _ in part_codes_list])

                cursor.execute(f'SELECT part_code, part_name FROM part_codes WHERE part_code IN ({placeholders})', part_codes_list)

                for row in cursor.fetchall():

                    part_names[row[0]] = row[1]

        

        close_db(conn)

        

        # Tm paralar birletir (beklenen + taranan)

        all_parts = {}

        

        # Beklenen paralar ekle

        for part_code, expected_qty in expected_parts.items():

            all_parts[part_code] = {

                'part_code': part_code,

                'part_name': part_names.get(part_code, ''),  # DB'den ekilen part_name

                'expected': expected_qty,

                'scanned': 0

            }

        

        # Taranan paralar ekle/gncelle

        for part_code, part_name, scanned_count in scanned_results:

            if part_code in all_parts:

                all_parts[part_code]['part_name'] = part_name or ''

                all_parts[part_code]['scanned'] = scanned_count

            else:

                all_parts[part_code] = {

                    'part_code': part_code,

                    'part_name': part_name or '',

                    'expected': 0,

                    'scanned': scanned_count

                }

        

        # DataFrame olutur - Sadece para detaylar

        rows = []

        for part in all_parts.values():

            difference = part['scanned'] - part['expected']

            

            # Durum emoji

            if difference == 0:

                status = ' Tam'

            elif difference > 0:

                status = ' Fazla'

            else:

                status = ' Eksik'

            

            rows.append({

                'Para Kodu': part['part_code'],

                'Para Ad': part['part_name'],

                'Beklenen Adet': part['expected'],

                'Saylan Adet': part['scanned'],

                'Fark': difference,

                'Durum': status

            })

        

        df = pd.DataFrame(rows)

        

        # Excel buffer

        output = BytesIO()

        with pd.ExcelWriter(output, engine='openpyxl') as writer:

            # Tek sayfa - Detayl Rapor

            df.to_excel(writer, sheet_name='Saym Raporu', index=False)

            

            # Formatla

            worksheet = writer.sheets['Saym Raporu']

            

            # Kolon geniliklerini ayarla

            column_widths = {

                'A': 18,  # Para Kodu

                'B': 30,  # Para Ad

                'C': 15,  # Beklenen Adet

                'D': 15,  # Saylan Adet

                'E': 12,  # Fark

                'F': 15   # Durum

            }

            

            for col, width in column_widths.items():

                worksheet.column_dimensions[col].width = width

            

            # Balk satrn kalnlatr

            from openpyxl.styles import Font, Alignment, PatternFill

            header_fill = PatternFill(start_color='DC2626', end_color='DC2626', fill_type='solid')

            header_font = Font(bold=True, color='FFFFFF', size=12)

            

            for cell in worksheet[1]:

                cell.fill = header_fill

                cell.font = header_font

                cell.alignment = Alignment(horizontal='center', vertical='center')

            

            # Durum stununu renklendir

            for row_idx in range(2, len(rows) + 2):

                cell = worksheet[f'F{row_idx}']

                fark_cell = worksheet[f'E{row_idx}']

                

                fark_value = fark_cell.value

                if fark_value == 0:

                    cell.fill = PatternFill(start_color='D4EDDA', end_color='D4EDDA', fill_type='solid')  # Yeil

                elif fark_value < 0:

                    cell.fill = PatternFill(start_color='FFF3CD', end_color='FFF3CD', fill_type='solid')  # Sar

                else:

                    cell.fill = PatternFill(start_color='D1ECF1', end_color='D1ECF1', fill_type='solid')  # Mavi

        

        output.seek(0)

        

        return send_file(

            output,

            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',

            as_attachment=True,

            download_name=f'sayim_raporu_{session_name}_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'

        )

        

    except Exception as e:

        print(f" Download Excel error: {e}")

        import traceback

        traceback.print_exc()

        return jsonify({'success': False, 'error': str(e)}), 500





@app.route('/login', methods=['POST'])

@rate_limit_login

def login():

    from werkzeug.security import check_password_hash

    data = request.get_json()

    username = data.get('username')

    password = data.get('password')



    if not username or not password:

        return jsonify({'error': 'Kullanc ad ve ifre gerekli'}), 400



    conn = get_db()

    cursor = conn.cursor()

    placeholder = get_db_placeholder()

    try:

        # Dual-mode: SQLite vs PostgreSQL table compatibility

        # Get user data first, then verify password

        if USE_POSTGRESQL:

            # PostgreSQL: Full schema with envanter_users table

            execute_query(cursor, f'SELECT id, username, full_name, role, password_hash FROM envanter_users WHERE username = {placeholder}',

                         (username,))

        else:

            # SQLite: envanter_users table with full_name (after column addition)

            execute_query(cursor, f'SELECT id, username, COALESCE(full_name, username) as full_name, role, password_hash FROM envanter_users WHERE username = {placeholder}',

                         (username,))

    except Exception as e:

        #  LOKAL: SQLite hata ynetimi

        logging.exception(f"Database error during login: {e}")

        try:

            close_db(conn)

        except Exception:

            pass

        return jsonify({'error': 'Veritaban hatas'}), 500



    user = cursor.fetchone()

    close_db(conn)



    if user and check_password_hash(user[4], password):  # user[4] is password_hash

        session['user_id'] = user[0]

        session['username'] = user[1]

        session['full_name'] = user[2]

        session['role'] = user[3]

        if user[3] == 'admin':

            session['admin_authenticated'] = True

        return jsonify({'success': True, 'role': user[3]})

    else:

        return jsonify({'error': 'Kullanc ad veya ifre hatal'}), 401



@app.route('/logout', methods=['POST'])

def logout():

    session.clear()

    return jsonify({'success': True})



@app.route('/check_auth')

def check_auth():

    if 'user_id' in session:

        return jsonify({

            'authenticated': True,

            'username': session.get('username', None),

            'full_name': session.get('full_name', None),

            'role': session.get('role', None)

        })

    return jsonify({'authenticated': False})



# ============================================================================

#  HEALTH CHECK & DEBUG ENDPOINTS

# ============================================================================



@app.route('/health', methods=['GET'])

def health_check():

    """Sistem salk kontrol - monitoring iin"""

    try:

        # DB balants test et

        db.session.execute(db.text('SELECT 1'))

        db_status = 'healthy'

        db_error = None

    except Exception as e:

        db_status = 'unhealthy'

        db_error = str(e)

    

    # Session says

    try:

        active_sessions = CountSession.query.filter_by(is_active=True).count()

    except:

        active_sessions = 0

    

    # Disk kullanm

    try:

        import shutil

        disk_usage = shutil.disk_usage('.')

        disk_free_gb = round(disk_usage.free / (1024**3), 2)

        disk_total_gb = round(disk_usage.total / (1024**3), 2)

        disk_used_percent = round((disk_usage.used / disk_usage.total) * 100, 1)

    except:

        disk_free_gb = 0

        disk_total_gb = 0

        disk_used_percent = 0

    

    # DB dosya boyutu

    try:

        db_size_mb = round(os.path.getsize('instance/envanter_local.db') / (1024**2), 2)

    except:

        db_size_mb = 0

    

    # Uptime

    uptime_seconds = int(time.time() - app.config.get('START_TIME', time.time()))

    uptime_hours = round(uptime_seconds / 3600, 1)

    

    health_data = {

        'status': 'ok' if db_status == 'healthy' else 'degraded',

        'timestamp': datetime.utcnow().isoformat(),

        'database': {

            'status': db_status,

            'error': db_error,

            'size_mb': db_size_mb,

            'type': 'SQLite'

        },

        'environment': {

            'mode': 'LOCAL' if IS_LOCAL else 'PRODUCTION',

            'storage': 'Local Files',

            'python_version': f"{os.sys.version_info.major}.{os.sys.version_info.minor}.{os.sys.version_info.micro}"

        },

        'sessions': {

            'active_count': active_sessions

        },

        'disk': {

            'free_gb': disk_free_gb,

            'total_gb': disk_total_gb,

            'used_percent': disk_used_percent

        },

        'uptime': {

            'seconds': uptime_seconds,

            'hours': uptime_hours

        }

    }

    

    return jsonify(health_data)



@app.route('/log_frontend_error', methods=['POST'])

def log_frontend_error():

    """Frontend hatalarn logla"""

    try:

        data = request.json

        

        # Log dosyasna yaz

        app.logger.error(f"FRONTEND ERROR: {json.dumps(data, ensure_ascii=False)}")

        

        # Ayr dosyaya da yaz

        os.makedirs('logs', exist_ok=True)

        with open('logs/frontend_errors.log', 'a', encoding='utf-8') as f:

            error_line = {

                'timestamp': datetime.utcnow().isoformat(),

                'error': data.get('error', {}),

                'context': data.get('context', {}),

                'user': {

                    'username': session.get('username'),

                    'user_id': session.get('user_id')

                }

            }

            f.write(json.dumps(error_line, ensure_ascii=False) + '\n')

        

        return jsonify({'success': True})

    except Exception as e:

        app.logger.error(f'Frontend error logging hatas: {e}')

        return jsonify({'success': False, 'error': str(e)}), 500



@app.route('/admin/frontend_errors_log', methods=['GET'])

@login_required

def get_frontend_errors_log():

    """Frontend error log dosyasn oku - sadece admin"""

    # Admin kontrol

    if session.get('role') != 'admin':

        return jsonify({'error': 'Admin eriimi gerekli'}), 403

    

    try:

        errors = []

        log_path = 'logs/frontend_errors.log'

        

        if os.path.exists(log_path):

            with open(log_path, 'r', encoding='utf-8') as f:

                lines = f.readlines()

                # Son 50 satr

                for line in lines[-50:]:

                    try:

                        errors.append(json.loads(line.strip()))

                    except:

                        pass

        

        return jsonify({

            'success': True,

            'errors': errors,

            'count': len(errors)

        })

    except Exception as e:

        app.logger.error(f'Frontend errors log okuma hatas: {e}')

        return jsonify({'success': False, 'error': str(e)}), 500



        return jsonify({'success': True, 'message': 'Hata kaydedildi'})

    except Exception as e:

        app.logger.error(f'Frontend error logging failed: {e}')

        return jsonify({'success': False, 'error': str(e)}), 500



@app.route('/upload_parts', methods=['POST'])

@login_required

def upload_parts():

    """

    Yeni Para Ykleme Sistemi:

    - Excel'den sadece para bilgileri (part_code, part_name) yklenir

    - QR kod retimi yaplmaz (manuel olarak para detay sayfasndan retilir)

    - Mevcut paralar gncellenir, yeni paralar eklenir

    """

    if 'file' not in request.files:

        return jsonify({'error': 'Dosya bulunamad'}), 400



    file = request.files['file']

    if file.filename == '':

        return jsonify({'error': 'Dosya seilmedi'}), 400



    if not (file.filename and file.filename.endswith(('.xlsx', '.xls'))):

        return jsonify({'error': 'Sadece Excel dosyalar yklenebilir'}), 400



    try:

        df = pd.read_excel(file)



        # Sadece part_code ve part_name gerekli

        required_columns = ['part_code', 'part_name']

        if not all(col in df.columns for col in required_columns):

            return jsonify({'error': 'Excel dosyas "part_code" ve "part_name" stunlarn iermelidir'}), 400



        conn = get_db()

        cursor = conn.cursor()

        placeholder = get_db_placeholder()



        # Mevcut paralar al

        PartCode.query')

        existing_parts = {}  # {part_code: (id, part_name)}

        for row in cursor.fetchall():

            existing_parts[row[1]] = (row[0], row[2])



        new_parts = []

        updated_parts = []

        processing_summary = {

            'new_parts': 0,

            'updated_parts': 0

        }



        print(f"\n PARA YKLEME SSTEM")

        print(f" Excel'den gelen para says: {len(df)}")

        print("="*50)



        for _, row in df.iterrows():

            part_code = str(row['part_code']).strip()

            part_name = str(row['part_name']).strip()



            if not part_code or not part_name:

                continue  # Bo satrlar atla



            if part_code in existing_parts:

                # MEVCUT PARA - Sadece ismi gncelle

                part_code_id, old_part_name = existing_parts[part_code]

                if old_part_name != part_name:

                    execute_query(cursor, f'UPDATE part_codes SET part_name = {placeholder} WHERE part_code = {placeholder}',

                                 (part_name, part_code))

                    updated_parts.append(part_code)

                    processing_summary['updated_parts'] += 1

                    print(f" {part_code}: '{old_part_name}'  '{part_name}'")

                else:

                    print(f" {part_code}: Zaten gncel")

            else:

                # YEN PARA - Sadece part_codes'a ekle (QR kod retilmez)

                execute_query(cursor, f'INSERT INTO part_codes (part_code, part_name, description, created_at) VALUES ({placeholder}, {placeholder}, {placeholder}, {placeholder})',

                             (part_code, part_name, '', datetime.now()))

                new_parts.append(part_code)

                processing_summary['new_parts'] += 1

                print(f" {part_code}: Yeni para eklendi - '{part_name}'")



        conn.commit()

        close_db(conn)



        print(f"\n LEM TAMAMLANDI")

        print(f" Yeni paralar: {processing_summary['new_parts']}")

        print(f" Gncellenen paralar: {processing_summary['updated_parts']}")

        print(" QR kodlar para detay sayfasndan retilebilir")

        print("="*50)



        return jsonify({

            'success': True,

            'message': f'lem tamamland! {processing_summary["new_parts"]} yeni para eklendi, {processing_summary["updated_parts"]} para gncellendi.',

            'summary': {

                'new_parts': processing_summary['new_parts'],

                'updated_parts': processing_summary['updated_parts']

            }

        })



    except Exception as e:

        logging.exception(f"Error in smart QR upload system: {e}")

        try:

            close_db(conn)

        except Exception:

            pass

        return jsonify({'error': f'lem srasnda hata olutu: {str(e)}'}), 500



@app.route('/get_qr_codes')

@login_required

def get_qr_codes():

    search = request.args.get('search', '').strip()

    page = int(request.args.get('page', 1))

    limit = int(request.args.get('limit', 100))

    offset = (page - 1) * limit



    conn = get_db()

    cursor = conn.cursor()



    if search:

        # nce tam eleme ara (JOIN ile part_code ve part_name al)

        execute_query(cursor, """

            SELECT qc.qr_id, pc.part_code, pc.part_name, qc.is_used, 

                   CASE WHEN qc.blob_url IS NOT NULL THEN 1 ELSE 0 END as is_downloaded

            FROM qr_codes qc

            JOIN part_codes pc ON qc.part_code_id = pc.id

            WHERE pc.part_code =  OR pc.part_name = 

            ORDER BY pc.part_code, qc.qr_id

            LIMIT  OFFSET 

        """, (search, search, limit, offset))

        exact_matches = cursor.fetchall()



        if exact_matches:

            qr_codes = []

            for row in exact_matches:

                qr_codes.append({

                    'qr_id': row[0],

                    'part_code': row[1],

                    'part_name': row[2],

                    'is_used': row[3],

                    'is_downloaded': row[4]

                })

        else:

            # Tam eleme bulunamazsa ksmi eleme ara

            execute_query(cursor, """

                SELECT qc.qr_id, pc.part_code, pc.part_name, qc.is_used,

                       CASE WHEN qc.blob_url IS NOT NULL THEN 1 ELSE 0 END as is_downloaded

                FROM qr_codes qc

                JOIN part_codes pc ON qc.part_code_id = pc.id

                WHERE pc.part_code LIKE  OR pc.part_name LIKE 

                ORDER BY pc.part_code, qc.qr_id

                LIMIT  OFFSET 

            """, (f'%{search}%', f'%{search}%', limit, offset))

            rows = cursor.fetchall()

            qr_codes = []

            for row in rows:

                qr_codes.append({

                    'qr_id': row[0],

                    'part_code': row[1],

                    'part_name': row[2],

                    'is_used': row[3],

                    'is_downloaded': row[4]

                })

    else:

        # Arama terimi yoksa tm QR kodlar getir (sayfalama ile)

        execute_query(cursor, """

            SELECT qc.qr_id, pc.part_code, pc.part_name, qc.is_used,

                   CASE WHEN qc.blob_url IS NOT NULL THEN 1 ELSE 0 END as is_downloaded

            FROM qr_codes qc

            JOIN part_codes pc ON qc.part_code_id = pc.id

            ORDER BY pc.part_code, qc.qr_id

            LIMIT  OFFSET 

        """, (limit, offset))

        rows = cursor.fetchall()

        qr_codes = []

        for row in rows:

            qr_codes.append({

                'qr_id': row[0],

                'part_code': row[1],

                'part_name': row[2],

                'is_used': row[3],

                'is_downloaded': row[4]

            })



    # Toplam sayy al (JOIN ile)

    if search:

        execute_query(cursor, """

            SELECT COUNT(*) 

            FROM qr_codes qc

            JOIN part_codes pc ON qc.part_code_id = pc.id

            WHERE pc.part_code LIKE  OR pc.part_name LIKE 

        """, (f'%{search}%', f'%{search}%'))

    else:

        execute_query(cursor, "SELECT COUNT(*) FROM qr_codes")



    total_count = cursor.fetchone()[0]



    close_db(conn)



    return jsonify({

        'qr_codes': qr_codes,

        'total_count': total_count,

        'current_page': page,

        'total_pages': (total_count + limit - 1) // limit,

        'has_more': (page * limit) < total_count

    })



@app.route('/clear_all_qrs', methods=['POST'])

@login_required

def clear_all_qrs():

    """Tm QR kodlarn temizle (aktif saym oturumu yoksa)"""

    try:

        conn = get_db()

        cursor = conn.cursor()



        # Aktif saym oturumu kontrol

        execute_query(cursor, 'SELECT COUNT(*) FROM count_sessions WHERE is_active = ', (True,))

        active_session = cursor.fetchone()[0]



        if active_session > 0:

            close_db(conn)

            return jsonify({'error': 'Aktif bir saym oturumu var. QR kodlar silinemez.'}), 400



        #  LOKAL: QR klasrn temizle

        qr_dir = os.path.join('static', 'qrcodes')

        if os.path.exists(qr_dir):

            for file in os.listdir(qr_dir):

                try:

                    os.remove(os.path.join(qr_dir, file))

                except Exception as e:

                    logging.error(f"Lokal dosya {file} silinirken hata: {e}")



        # QR kodlarn ve paralar sil

        execute_query(cursor, 'DELETE FROM qr_codes')

        execute_query(cursor, 'DELETE FROM part_codes')



        conn.commit()

        close_db(conn)



        # Cache'i temizle

        cache_clear()



        logging.info("Tm QR kodlar temizlendi")

        return jsonify({

            'success': True,

            'message': 'Tm QR kodlar baaryla silindi'

        })

    except Exception as e:

        try:

            close_db(conn)

        except:

            pass

        logging.error(f"QR kodlar silinirken hata: {e}", exc_info=True)

        return jsonify({'error': f'QR kodlar silinirken hata: {str(e)}'}), 500



@app.route('/generate_qr_image/<qr_id>')

@login_required

def generate_qr_image(qr_id):

    """QR kod oluturma - QR altna kod ve para ad ekler"""

    try:

        # Cache'den kontrol et

        cache_key = f'qr_image_{qr_id}'

        cached_image = cache_get(cache_key)



        if cached_image:

            buf = BytesIO(cached_image)

            return send_file(buf, mimetype='image/png')



        #  LOKAL: Static klasrden kontrol et

        static_path = os.path.join('static', 'qrcodes', f'{qr_id}.png')

        if os.path.exists(static_path):

            with open(static_path, 'rb') as f:

                file_content = f.read()

            cache_set(cache_key, file_content)

            buf = BytesIO(file_content)

            return send_file(buf, mimetype='image/png')



        # QR ID'den para kodunu ve numaray kar (Y129150-49811_1)

        parts = qr_id.rsplit('_', 1)

        part_code = parts[0] if len(parts) > 0 else qr_id

        qr_number = parts[1] if len(parts) > 1 else "1"

        

        # Para adn database'den al

        conn = get_db()

        cursor = conn.cursor()

        execute_query(cursor, 'SELECT part_name FROM part_codes WHERE part_code = ', (part_code,))

        result = cursor.fetchone()

        part_name = result[0] if result else ""

        close_db(conn)



        # QR kod olutur - Barkod makinesi iin optimize

        qr = qrcode.QRCode(

            version=1, 

            box_size=8,  # 8px - barkod makinesi iin ideal

            border=2,    # 2px quiet zone

            error_correction=qrcode.constants.ERROR_CORRECT_M

        )

        qr.add_data(qr_id)

        qr.make(fit=True)

        qr_img = qr.make_image(fill_color="black", back_color="white")

        qr_img = qr_img.convert('RGB')  # PIL Image'a dntr



        # QR kod boyutlarn al

        qr_width, qr_height = qr_img.size

        

        # Alanlar hesapla

        logo_height = 40  # Logo iin st alan

        text_height = 35  # Alt yaz (para numaras) iin alan

        

        # Krmz ereve iin padding

        border_width = 3  # 3px krmz ereve

        

        # Yeni grsel olutur (logo + QR + text alan + ereve)

        final_width = qr_width + (border_width * 2)

        final_height = logo_height + qr_height + text_height + (border_width * 2)

        final_img = Image.new('RGB', (final_width, final_height), '#dc2626')  # Krmz arka plan (ereve)

        

        # Beyaz i alan olutur (logo + QR + text)

        white_bg = Image.new('RGB', (qr_width, logo_height + qr_height + text_height), 'white')

        

        # Logo ekle (varsa) - st ortasna

        try:

            logo_path = os.path.join(os.path.dirname(__file__), 'cermak-logo.png')

            if os.path.exists(logo_path):

                logo_img = Image.open(logo_path).convert('RGBA')

                # Logo boyutunu alan yksekliine gre ayarla

                logo_width = 150

                logo_height_logo = 40

                try:

                    logo_img = logo_img.resize((logo_width, logo_height_logo), Image.Resampling.LANCZOS)

                except AttributeError:

                    logo_img = logo_img.resize((logo_width, logo_height_logo), Image.LANCZOS)

                

                # Logo'yu ortala

                logo_x = (qr_width - logo_width) // 2

                logo_y = 5  # stten 5px boluk

                

                # RGBA logo'yu blend et

                if logo_img.mode == 'RGBA':

                    alpha = logo_img.split()[3]

                    logo_img = logo_img.convert('RGB')

                    white_bg.paste(logo_img, (logo_x, logo_y), alpha)

                else:

                    white_bg.paste(logo_img, (logo_x, logo_y))

        except Exception as e:

            print(f"Logo ekleme hatas: {e}")

        

        # QR kodu beyaz alana yaptr (logo'nun altna)

        white_bg.paste(qr_img, (0, logo_height))

        

        # Beyaz alan krmz erevenin iine yaptr

        final_img.paste(white_bg, (border_width, border_width))

        

        # Text ekleme iin draw nesnesi

        draw = ImageDraw.Draw(final_img)

        

        # Font (kaln ve byk)

        try:

            font = ImageFont.truetype("arialblk.ttf", 24)

        except:

            try:

                font = ImageFont.truetype("arialbd.ttf", 24)

            except:

                try:

                    font = ImageFont.truetype("arial.ttf", 24)

                except:

                    font = ImageFont.load_default()

        

        # QR ID yazs - Sadece bu

        qr_text = qr_id

        

        # Text geniliini hesapla (24pt font iin)

        text_width = len(qr_text) * 14

        

        # QR ID'yi ortala ve yaz (border_width offset ekle)

        x_position = max(border_width, (final_width - text_width) // 2)

        draw.text((x_position, logo_height + qr_height + border_width + 5), qr_text, fill='black', font=font)



        buf = BytesIO()

        final_img.save(buf, format='PNG', optimize=True)

        buf.seek(0)



        # Cache'e kaydet

        img_data = buf.getvalue()

        cache_set(cache_key, img_data)



        #  LOKAL: static klasrne kaydet

        qr_dir = os.path.join('static', 'qrcodes')

        os.makedirs(qr_dir, exist_ok=True)

        with open(static_path, 'wb') as f:

            f.write(img_data)

        

        # Dosyay read-only yap

        os.chmod(static_path, 0o444)



        # Dosyay dndr

        buf.seek(0)

        return send_file(buf, mimetype='image/png')



    except Exception as e:

        logging.error(f"Error generating QR image for {qr_id}: {e}")

        # Hata durumunda basit QR olutur

        qr = qrcode.QRCode(version=1, box_size=4, border=1)

        qr.add_data(qr_id)

        qr.make(fit=True)

        img = qr.make_image(fill_color="black", back_color="white")



        buf = BytesIO()

        img.save(buf, format='PNG')

        buf.seek(0)



        return send_file(buf, mimetype='image/png')



@app.route('/download_single_qr/<qr_id>')

@login_required

def download_single_qr(qr_id):

    conn = get_db()

    cursor = conn.cursor()

    execute_query(cursor, '''

        SELECT pc.part_code 

        FROM qr_codes qc

        JOIN part_codes pc ON qc.part_code_id = pc.id

        WHERE qc.qr_id = ??

    ''', (qr_id,))

    result = cursor.fetchone()



    if not result:

        close_db(conn)

        return jsonify({'error': 'QR kod bulunamad'}), 404



    # is_downloaded kolonu yok, sadece blob_url kontrol yeterli

    close_db(conn)



    try:

        if USE_B2_STORAGE and get_b2_service:

            # PRODUCTION: B2'den QR kod'u indir (KALICI)

            b2_service = get_b2_service()

            file_path = f'qr_codes/{qr_id}.png'



            file_content = b2_service.download_file(file_path)



            if file_content:

                # B2'den var olan dosyay dndr

                buf = BytesIO(file_content)

                buf.seek(0)

                return send_file(buf, mimetype='image/png', as_attachment=True, download_name=f'{qr_id}.png')

        else:

            # LOCAL: Static dosyadan kontrol et (GEC)

            static_path = os.path.join('static', 'qrcodes', f'{qr_id}.png')

            if os.path.exists(static_path):

                return send_file(static_path, mimetype='image/png', as_attachment=True, download_name=f'{qr_id}.png')



        # QR kod yoksa olutur

        qr = qrcode.QRCode(version=1, box_size=10, border=4)

        qr.add_data(qr_id)

        qr.make(fit=True)

        img = qr.make_image(fill_color="black", back_color="white")



        buf = BytesIO()

        img.save(buf, format='PNG')

        buf.seek(0)



        img_data = buf.getvalue()



        if USE_B2_STORAGE and get_b2_service:

            # PRODUCTION: B2'ye ykle (KALICI)

            b2_service = get_b2_service()

            file_path = f'qr_codes/{qr_id}.png'

            upload_result = b2_service.upload_file(file_path, img_data, 'image/png')





            if upload_result['success']:

                logging.info(f"QR code uploaded to B2: {file_path}")

        else:

            # LOCAL: Static klasrne kaydet (GEC)

            qr_dir = os.path.join('static', 'qrcodes')

            os.makedirs(qr_dir, exist_ok=True)

            local_path = os.path.join(qr_dir, f'{qr_id}.png')

            with open(local_path, 'wb') as f:

                f.write(img_data)



        # Dosyay dndr

        buf.seek(0)

        return send_file(buf, mimetype='image/png', as_attachment=True, download_name=f'{qr_id}.png')



    except Exception as e:

        logging.error(f"Error downloading QR image for {qr_id}: {e}")

        # Hata durumunda geleneksel yntemle olutur

        qr = qrcode.QRCode(version=1, box_size=10, border=4)

        qr.add_data(qr_id)

        qr.make(fit=True)

        img = qr.make_image(fill_color="black", back_color="white")



        buf = BytesIO()

        img.save(buf, format='PNG')

        buf.seek(0)



        return send_file(buf, mimetype='image/png', as_attachment=True, download_name=f'{qr_id}.png')



#  ULTRA QR SCANNER HANDLER

@socketio.on('scan_qr_radical')

def handle_scan_radical(data):

    """ Ultra reliable QR scanning with enhanced features"""

    print("\n" + ""*50)

    print("ULTRA QR SCAN RECEIVED")

    print(""*50)



    try:

        qr_id = data.get('qr_id', '').strip()

        session_id = int(data.get('session_id', 1))  # INT olarak al

        user_id = session.get('user_id', 1)



        logging.debug(f" QR ID: {qr_id}, Session: {session_id}, User: {user_id}")



        if not qr_id:

            logging.warning("QR ID missing in scan request")

            emit('scan_result', {'success': False, 'message': ' QR ID eksik!'})

            return



        with db_connection() as conn:

            cursor = conn.cursor()



            # Check QR exists (JOIN ile part bilgisi al)

            execute_query(cursor, '''

                SELECT pc.part_code, pc.part_name 

                FROM qr_codes qc

                JOIN part_codes pc ON qc.part_code_id = pc.id

                WHERE qc.qr_id = ??

            ''', (qr_id,))

            qr_data = cursor.fetchone()



            if not qr_data:

                logging.warning(f"QR not found: {qr_id}")

                emit('scan_result', {'success': False, 'message': f' QR kod bulunamad: {qr_id}'})

                return



            part_code, part_name = qr_data

            logging.info(f"QR found: {part_code} - {part_name}")



            # Ensure session exists

            execute_query(cursor, 'SELECT COUNT(*) FROM count_sessions WHERE id = ', (session_id,))

            if cursor.fetchone()[0] == 0:

                logging.info(f"Creating new session {session_id}")

                execute_query(cursor, 

                    'INSERT INTO count_sessions (id, session_name, is_active, created_at) VALUES (?, ?, ?, ?)',

                    (session_id, f'Session_{session_id}', 1, datetime.now()))

                conn.commit()



            #  COMPOSITE INDEX kullanr - ok hzl + AKILLI DUPLICATE

            duplicate_record = ScannedQR.query.filter_by(session_id=session_id, qr_id=qr_id).first()

            

            if duplicate_record:

                # Duplicate bulundu - Akll mesaj olutur

                prev_user_id, prev_scanned_at = duplicate_record

                

                # nceki kullanc bilgisi

                execute_query(cursor, 'SELECT full_name FROM envanter_users WHERE id = ', (prev_user_id,))

                prev_user_result = cursor.fetchone()

                prev_user_name = prev_user_result[0] if prev_user_result else 'Bilinmeyen'

                

                # Zaman fark hesapla

                prev_time = datetime.fromisoformat(prev_scanned_at) if isinstance(prev_scanned_at, str) else prev_scanned_at

                time_diff = datetime.now() - prev_time

                

                # Zaman format

                if time_diff.total_seconds() < 60:

                    time_str = f"{int(time_diff.total_seconds())} saniye nce"

                    is_suspicious = time_diff.total_seconds() < 30  # 30 saniyeden ksa ise pheli

                elif time_diff.total_seconds() < 3600:

                    time_str = f"{int(time_diff.total_seconds() / 60)} dakika nce"

                    is_suspicious = False

                else:

                    time_str = f"{int(time_diff.total_seconds() / 3600)} saat nce"

                    is_suspicious = False

                

                # Ayn kullanc m kontrol et

                is_same_user = (prev_user_id == user_id)

                

                # Akll mesaj olutur

                if is_suspicious:

                    duplicate_msg = f" PHEL! {part_name} {time_str} tarand ({prev_user_name})"

                elif is_same_user:

                    duplicate_msg = f" {part_name} zaten taradnz! ({time_str})"

                else:

                    duplicate_msg = f" {part_name} zaten sayld! {prev_user_name} tarafndan {time_str}"

                

                logging.warning(f"Duplicate scan: {qr_id} - {duplicate_msg}")

                emit('scan_result', {

                    'success': False, 

                    'message': duplicate_msg,

                    'duplicate': True,

                    'previous_user': prev_user_name,

                    'time_ago': time_str,

                    'is_suspicious': is_suspicious,

                    'same_user': is_same_user

                }, broadcast=True)

                return



            #  TRANSACTION - Atomik ilem

            cursor.execute('BEGIN TRANSACTION')

            try:

                # Insert scan record

                execute_query(cursor, 

                    'INSERT INTO scanned_qr (session_id, qr_id, part_code, scanned_by, scanned_at) VALUES (, , , , )',

                    (session_id, qr_id, part_code, user_id, datetime.now()))



                # Mark QR as used

                execute_query(cursor, 'UPDATE qr_codes SET is_used = , used_at =  WHERE qr_id = ',

                             (1, datetime.now(), qr_id))



                # Update session stats

                execute_query(cursor, '''

                    UPDATE count_sessions 

                    SET total_scanned = (SELECT COUNT(*) FROM scanned_qr WHERE session_id = )

                    WHERE id = ?

                ''', (session_id, session_id))



                conn.commit()

                logging.info(f"SUCCESS: {part_name} scanned")

            except Exception as e:

                conn.rollback()

                logging.error(f"Transaction failed: {e}")

                raise



        # Get user name (ayr connection)

        with db_connection() as conn2:

            cursor2 = conn2.cursor()

            execute_query(cursor2, 'SELECT full_name FROM envanter_users WHERE id = ', (user_id,))

            user_result = cursor2.fetchone()

            user_name = user_result[0] if user_result else 'Kullanc'



        #  TRIPLE BROADCAST

        success_data = {

            'success': True,

            'message': f' {part_name} sayld!',

            'qr_code': qr_id,

            'part_code': part_code,

            'part_name': part_name,

            'session_id': session_id,

            'scanned_by': user_name,

            'scanned_at': datetime.now().strftime('%H:%M:%S')

        }



        # Total scan count (session stats'dan al)

        with db_connection() as conn3:

            cursor3 = conn3.cursor()

            execute_query(cursor3, 'SELECT total_scanned FROM count_sessions WHERE id = ', (session_id,))

            result = cursor3.fetchone()

            success_data['total_scans'] = result[0] if result else 0



        socketio.emit('scan_result', success_data, broadcast=True)

        socketio.emit('qr_scanned', success_data, broadcast=True)

        socketio.emit('activity_update', success_data, broadcast=True)



        logging.info(f"ULTRA SUCCESS - {part_name} scanned by {user_name}")



    except sqlite3.Error as e:

        logging.error(f"Database error in scan: {e}", exc_info=True)

        emit('scan_result', {'success': False, 'message': f' Veritaban hatas: {e}'})

    except ValueError as e:

        logging.error(f"Value error in scan: {e}", exc_info=True)

        emit('scan_result', {'success': False, 'message': f' Geersiz veri: {e}'})

    except Exception as e:

        logging.error(f"Unexpected error in scan: {e}", exc_info=True)

        emit('scan_result', {'success': False, 'message': f' Sistem hatas: {e}'})



#  ULTRA MODERN API ENDPOINTS

@app.route('/api/scan_qr', methods=['POST'])

def api_scan_qr_ultra():

    """ Ultra reliable QR scanning API endpoint with modern tech"""

    try:

        data = request.get_json()

        if not data:

            return jsonify({'success': False, 'message': 'No data provided'}), 400



        qr_id = data.get('qr_id', '').strip()

        session_id = data.get('session_id', '1')



        if not qr_id:

            return jsonify({'success': False, 'message': 'QR ID required'}), 400



        #  SCANNER FIX: Baz QR scanner cihazlar - ve _ karakterlerini yanl okuyor

        # * (ASCII 42) -> - (ASCII 45) dntr

        #  (ASCII 63) -> _ (ASCII 95) dntr

        qr_id_original = qr_id

        qr_id = qr_id.replace('*', '-').replace('', '_')

        

        if qr_id != qr_id_original:

            print(f" QR FIX: '{qr_id_original}' -> '{qr_id}'")



        print(f" ULTRA API QR Scan: {qr_id} in session {session_id}")



        # Process the scan with ultra reliability

        result = process_qr_scan_ultra(qr_id, session_id)



        # Ultra broadcast - emit to all clients for real-time updates

        socketio.emit('scan_result', result)

        socketio.emit('qr_scanned', result)

        socketio.emit('activity_update', result)



        return jsonify(result)



    except Exception as e:

        print(f" ULTRA API scan error: {e}")

        return jsonify({'success': False, 'message': str(e)}), 500



def process_qr_scan_ultra(qr_id, session_id):

    """ Ultra reliable QR processing with enhanced features"""

    conn = None

    try:

        conn = get_db()

        cursor = conn.cursor()



        # Ultra session management - ensure session exists with better naming

        session_name = f"Tarama Seans {session_id}"



        # Defensive: check if session row exists, then insert using DB-agnostic columns

        # count_sessions kontrol artk gereksiz - session ID yoksa oluturma

        # Session ID artk auto-increment, manuel oluturmaya gerek yok

        

        # ========================================

        #  PAKET KONTROL - nce paketin olup olmadn kontrol et

        # ========================================

        cursor.execute('''

            SELECT is_package, package_items, part_name 

            FROM part_codes 

            WHERE part_code = 

        ''', (qr_id,))

        

        package_check = cursor.fetchone()

        

        # Eer bu bir paketse, iindeki tm paralar tek tek tara

        if package_check and package_check[0]:  # is_package = True

            try:

                package_items = json.loads(package_check[1]) if package_check[1] else []

                package_name = package_check[2]

                

                if not package_items:

                    return {

                        'success': False,

                        'message': f' {package_name} paketi bo!',

                        'item_name': package_name

                    }

                

                # Paket taramas duplicate kontrol

                cursor.execute('''

                    SELECT COUNT(*) FROM scanned_qr 

                    WHERE qr_id =  AND session_id = 

                ''', (qr_id, str(session_id)))

                

                if cursor.fetchone()[0] > 0:

                    return {

                        'success': False,

                        'message': f' {package_name} paketi zaten tarand!',

                        'item_name': package_name,

                        'duplicate': True

                    }

                

                #  PAKET KENDISINI TARA DEL - SADECE ERDEK PARALARI TARA

                # Paket, sadece tracker olarak kaydediliyor, toplam sayya EKLENMYOR

                

                # Paket iindeki her paray otomatik tara

                total_items = 0

                for idx, item in enumerate(package_items):

                    part_code = item.get('part_code')

                    quantity = item.get('quantity', 1)

                    

                    # Her bir para iin quantity kadar tarama kayd olutur

                    for qty_idx in range(quantity):

                        virtual_qr = f"{qr_id}_PKG_{part_code}_{idx}_{qty_idx}"

                        cursor.execute('''

                            INSERT INTO scanned_qr (qr_id, session_id, part_code, scanned_by, scanned_at)

                            VALUES (, , , , )

                        ''', (virtual_qr, str(session_id), part_code, session.get('user_id', 1), datetime.now()))

                        total_items += 1

                

                # statistikleri gncelle (SADECE ERDEK PARALARI SAY, PAKET SAYMA!)

                cursor.execute('''

                    SELECT COUNT(*) as total_scans

                    FROM scanned_qr 

                    WHERE session_id = 

                ''', (str(session_id),))

                

                total_scans = cursor.fetchone()[0]

                

                try:

                    cursor.execute('UPDATE count_sessions SET total_scanned =  WHERE id = ', 

                                 (total_scans, str(session_id)))

                except Exception as e:

                    print(f" Package total_scanned gncelleme hatas: {e}")

                    pass

                

                conn.commit()

                

                app.logger.info(f'[PAKET TARAMA] {package_name} - {total_items} para otomatik tarand')

                

                return {

                    'success': True,

                    'message': f'  {package_name}\n{total_items} para otomatik sayld!',

                    'item_name': package_name,

                    'total_scans': total_scans,

                    'is_package': True,

                    'package_items': total_items

                }

                

            except Exception as pkg_err:

                app.logger.error(f'[PAKET HATASI] {qr_id}: {pkg_err}')

                # Paket hatas olursa normal QR gibi ile

                pass

        

        # ========================================

        #  NORMAL QR LEME - Standart para tarama

        # ========================================

        

        # Check if QR exists with enhanced data retrieval

        qr_data = None

        try:

            # nce tam eleme dene

            execute_query(cursor, """

                SELECT qc.qr_id, pc.part_code, pc.part_name, qc.is_used, qc.created_at

                FROM qr_codes qc

                LEFT JOIN part_codes pc ON qc.part_code_id = pc.id

                WHERE qc.qr_id = ??

            """, (qr_id,))

            qr_data = cursor.fetchone()

            

            # Hala bulunamadysa, part_code olarak ara (QR = part_code durumu)

            if not qr_data:

                execute_query(cursor, """

                    SELECT pc.part_code, pc.part_code, pc.part_name, 0, 

                    FROM part_codes pc

                    WHERE pc.part_code = 

                """, (datetime.now(), qr_id))

                qr_data = cursor.fetchone()

                

        except Exception as e:

            # Schema mismatch or missing column in older DBs - fall back to defaults

            print(f" QR lookup failed (schema mismatch): {e}")

            qr_data = None



        if not qr_data:

            # Do NOT attempt to modify the schema here. Use a safe fallback QR record in-memory.

            unknown_name = f"Bilinmeyen rn ({qr_id[:15]})"

            print(f" QR BULUNAMADI: {qr_id}")  # Debug log

            qr_data = (qr_id, qr_id[:10], unknown_name, False, datetime.now())



        qr_id_db, part_code, part_name, is_used, created_at = qr_data



        #  KALICI DUPLICATE KONTROL - Bir session'da bir QR sadece 1 kez okunabilir

        execute_query(cursor, """

            SELECT COUNT(*) FROM scanned_qr 

            WHERE qr_id =  AND session_id = 

        """, (qr_id, str(session_id)))



        existing_count = cursor.fetchone()[0]

        if existing_count > 0:

            # Bu QR zaten bu session'da taranm - asla tekrar taranmamal

            return {

                'success': False,

                'message': f' {part_name} zaten tarand!',

                'item_name': part_name,

                'duplicate': True

            }



        # Insert ultra scan record with enhanced data

        try:

            execute_query(cursor, """

                INSERT INTO scanned_qr (qr_id, session_id, part_code, scanned_by, scanned_at)

                VALUES (, , , , )

            """, (qr_id, str(session_id), part_code, session.get('user_id', 1), datetime.now()))

        except Exception:

            # Fallback for older schemas without part_code column

            try:

                execute_query(cursor, """

                    INSERT INTO scanned_qr (qr_id, session_id, scanned_by, scanned_at)

                    VALUES (?, ?, ?, ?)

                """, (qr_id, str(session_id), session.get('user_id', 1), datetime.now()))

            except Exception as ie:

                print(f" Failed to insert scanned_qr record: {ie}")

                raise



        # Mark QR as used with timestamp (DB-agnostic) - ignore if schema doesn't match

        try:

            execute_query(cursor, """

                UPDATE qr_codes 

                SET is_used = , used_at =  

                WHERE qr_id = 

            """, (True, datetime.now(), qr_id))

        except Exception:

            # Not critical if the QR table lacks these columns in older schemas

            pass



        # Get enhanced session statistics

        execute_query(cursor, """

            SELECT 

                COUNT(*) as total_scans,

                COUNT(DISTINCT qr_id) as unique_items

            FROM scanned_qr 

            WHERE session_id = 

        """, (str(session_id),))



        stats = cursor.fetchone()

        total_scans = stats[0] if stats else 0

        unique_items = stats[1] if stats else 0



        # Update count_sessions.total_scanned for dashboard

        try:

            execute_query(cursor, 'UPDATE count_sessions SET total_scanned =  WHERE id = ', (total_scans, str(session_id)))

        except Exception as e:

            print(f" total_scanned gncelleme hatas: {e}")

            pass



        conn.commit()



        # Get user info for enhanced feedback

        user_id = session.get('user_id', 1)

        execute_query(cursor, 'SELECT full_name FROM envanter_users WHERE id = ', (user_id,))

        user_result = cursor.fetchone()

        user_name = user_result[0] if user_result else 'Kullanc'



        # Ultra success response

        return {

            'success': True,

            'message': f' {part_name} baaryla tarand! (#{total_scans})',

            'item_name': part_name,

            'part_code': part_code,

            'total_scans': total_scans,

            'unique_items': unique_items,

            'qr_id': qr_id,

            'session_id': session_id,

            'scanned_by': user_name,

            'scan_time': datetime.now().isoformat(),

            'was_used_before': is_used

        }



    except Exception as e:

        if conn:

            conn.rollback()

        print(f" ULTRA Process QR error: {e}")

        return {

            'success': False,

            'message': f' Sistem hatas: {str(e)}',

            'error_detail': str(e),

            'qr_id': qr_id

        }

    finally:

        if conn:

            close_db(conn)



@app.route('/api/session/<session_id>/stats', methods=['GET'])

def get_ultra_session_stats(session_id):

    """ Get ultra detailed session statistics"""

    try:

        conn = get_db()

        cursor = conn.cursor()



        # Get session info

        execute_query(cursor, """

            SELECT 

                CASE 

                    WHEN is_active::integer = 1 THEN 'active'

                    WHEN is_active::integer = 0 THEN 'completed'

                    ELSE 'unknown'

                END as status, 

                started_at, 

                ended_at 

            FROM count_sessions 

            WHERE session_id = 

        """, (str(session_id),))



        session_info = cursor.fetchone()



        if not session_info:

            return jsonify({'error': 'Session not found'}), 404



        # Get ultra scan statistics

        execute_query(cursor, """

            SELECT 

                COUNT(*) as total_scans,

                COUNT(DISTINCT qr_id) as unique_items,

                MIN(scanned_at) as first_scan,

                MAX(scanned_at) as last_scan

            FROM scanned_qr 

            WHERE session_id = 

        """, (str(session_id),))



        stats = cursor.fetchone()



        # Get recent scans with enhanced data

        execute_query(cursor, """

            SELECT sq.qr_id, qc.part_name, sq.scanned_at, u.full_name

            FROM scanned_qr sq

            LEFT JOIN qr_codes qc ON sq.qr_id = qc.qr_id

            LEFT JOIN envanter_users u ON sq.scanned_by = u.id

            WHERE sq.session_id = 

            ORDER BY sq.scanned_at DESC

            LIMIT 20

        """, (str(session_id),))



        recent_scans = cursor.fetchall()



        # Calculate session duration

        start_time = session_info[1]

        end_time = session_info[2] or (stats[3] if stats[3] else datetime.now())

        duration_seconds = (end_time - start_time).total_seconds() if start_time else 0



        return jsonify({

            'session_id': session_id,

            'status': session_info[0],

            'started_at': start_time.isoformat() if start_time else None,

            'ended_at': session_info[2].isoformat() if session_info[2] else None,

            'duration_seconds': duration_seconds,

            'total_scans': stats[0] or 0,

            'unique_items': stats[1] or 0,

            'first_scan': stats[2].isoformat() if stats[2] else None,

            'last_scan': stats[3].isoformat() if stats[3] else None,

            'scans_per_minute': round((stats[0] or 0) / max(duration_seconds / 60, 1), 2),

            'recent_scans': [

                {

                    'qr_id': scan[0],

                    'item_name': scan[1] or f'Unknown ({scan[0]})',

                    'scan_time': scan[2].isoformat() if scan[2] else None,

                    'scanned_by': scan[3] or 'Bilinmeyen'

                }

                for scan in recent_scans

            ]

        })



    except Exception as e:

        print(f" ULTRA Stats error: {e}")

        return jsonify({'error': str(e)}), 500

    finally:

        if conn:

            close_db(conn)



# ========================================

#  TELEFON QR & PARA BLGLER SSTEM KALDIRILDI

# Artk telefon saym YOK - sadece desktop QR scanner kullanlyor

# Kaldrlan zellikler:

# - /qr-info (mobil QR bilgi sayfas)

# - /api/qr_info/<qr_id> (QR bilgi API)

# - /api/part_details/<qr_id> (Para detaylar)

# - /api/update_part_details (Para gncelleme)

# - /api/upload_part_photo (Fotoraf ykleme)

# - /api/qr/<qr_id>/info (Detayl QR bilgisi)

# ========================================



#  RAPOR YNETM (Admin Count - ifre sistemi kaldrld)

    try:

        #  ULTRA SECURITY: Sadece admin kullancs saym bitirebilir

        current_user_id = session.get('user_id')



        if not current_user_id:

            return jsonify({'success': False, 'error': 'Oturum bulunamad - Ltfen tekrar giri yapn'}), 401



        # Admin kontrol - kullanc bilgilerini al

        conn = get_db()

        cursor = conn.cursor()



        execute_query(cursor, "SELECT username, role FROM envanter_users WHERE id = ", (current_user_id,))

        user_result = cursor.fetchone()



        if not user_result:

            close_db(conn)

            return jsonify({'success': False, 'error': 'Kullanc bulunamad'}), 404



        username, user_role = user_result



        # DEBUG LOG

        print(f"[FINISH_COUNT DEBUG] User: {username}, Role: {user_role}, ID: {current_user_id}")

        print(f"[FINISH_COUNT DEBUG] Role type: {type(user_role)}, Value: [{user_role}]")

        print(f"[FINISH_COUNT DEBUG] Role == 'admin': {user_role == 'admin'}")

        print(f"[FINISH_COUNT DEBUG] Role.lower() == 'admin': {user_role.lower() == 'admin' if user_role else 'NULL'}")



        #  SADECE ADMIN YETKS - Role tabanl kontrol (daha gvenli)

        if not user_role or user_role.lower() != 'admin':

            close_db(conn)

            security_logger.warning(f'USER {username} (Role: {user_role}, ID: {current_user_id}) tried to finish count session - PERMISSION DENIED')

            return jsonify({

                'success': False, 

                'error': f'YETKISIZ ERM: Sadece admin yetkisine sahip kullanclar saym bitirebilir (Your role: {user_role})',

                'permission_required': 'admin',

                'current_role': user_role

            }), 403



        # Saym eriim kontrol (secondary check) - ARTIK GEREKL DEL

        # if not session.get('count_access'):

        #     close_db(conn)

        #     return jsonify({'success': False, 'error': 'Saym eriimi iin ifre gerekli'}), 403



        # Aktif saym oturumunu kontrol et

        execute_query(cursor, "SELECT id, is_active, created_by FROM count_sessions WHERE is_active =  LIMIT 1", (True,))

        session_result = cursor.fetchone()



        if not session_result:

            close_db(conn)

            return jsonify({'success': False, 'error': 'Aktif saym oturumu bulunamad'}), 400



        session_id, is_active_status, created_by = session_result



        # Log admin action

        security_logger.info(f'ADMIN {username} finishing count session {session_id}')



        # ift ilem kontrol - eer bu oturum zaten tamamlandysa

        if not is_active_status:

            close_db(conn)

            return jsonify({'success': False, 'error': 'Bu saym oturumu zaten tamamlanm'}), 400



        # Saym oturumunu sonlandr (admin yetkisiyle)

        execute_query(cursor, "UPDATE count_sessions SET is_active = , ended_at =  WHERE id = ",

                     (False, datetime.now(), session_id))



        # Rapor verilerini topla - BEKLENEN ADETLERLE KARILATIR

        # Her QR tek bir part_code'a ait, o yzden JOIN ile part_code ekiyoruz

        execute_query(cursor, '''

            SELECT 

                COALESCE(pc.part_code, sq.part_code) as part_code,

                COALESCE(pc.part_name, 'Bilinmeyen Para') as part_name,

                COUNT(*) as sayilan_adet

            FROM scanned_qr sq

            LEFT JOIN qr_codes qc ON sq.qr_id = qc.qr_id

            LEFT JOIN part_codes pc ON qc.part_code_id = pc.id

            WHERE sq.session_id = 

            GROUP BY COALESCE(pc.part_code, sq.part_code), COALESCE(pc.part_name, 'Bilinmeyen Para')

            ORDER BY part_code

        ''', (session_id,))



        scanned_parts = {}  # {part_code: (part_name, sayilan_adet)}

        for row in cursor.fetchall():

            part_code = row[0]

            part_name = row[1]

            sayilan_adet = row[2]

            scanned_parts[part_code] = (part_name, sayilan_adet)

            print(f"[RAPOR DEBUG] {part_code}: {part_name} - {sayilan_adet} adet okundu")



        # Tm part_codes'dan beklenen adetleri al (yklenmi Excel'den)

        execute_query(cursor, '''

            SELECT 

                pc.part_code,

                pc.part_name,

                COUNT(qc.qr_id) as beklenen_adet

            FROM part_codes pc

            LEFT JOIN qr_codes qc ON qc.part_code_id = pc.id

            GROUP BY pc.part_code, pc.part_name

        ''')



        # Rapor verilerini hazrla - TM PARALAR (okutulan + okutulmayan)

        report_data = []

        total_scanned = 0

        total_expected = 0



        for row in cursor.fetchall():

            part_code = row[0]

            part_name = row[1]

            beklenen_adet = row[2] or 0

            

            # Bu para okutulan paralar arasnda var m

            if part_code in scanned_parts:

                sayilan_adet = scanned_parts[part_code][1]

            else:

                sayilan_adet = 0  # Hi okutulmam

            

            fark = sayilan_adet - beklenen_adet

            

            part_data = {

                'Para Kodu': part_code or 'Bilinmiyor',

                'Para Ad': part_name,

                'Beklenen Adet': beklenen_adet,

                'Saylan Adet': sayilan_adet,

                'Fark': fark,

                'Durum': ' Tamam' if fark == 0 else (' Eksik' if fark < 0 else '+ Fazla')

            }

            report_data.append(part_data)

            total_scanned += sayilan_adet

            total_expected += beklenen_adet



        # Excel raporu olutur

        df = pd.DataFrame(report_data)



        output = BytesIO()

        with pd.ExcelWriter(output, engine='openpyxl') as writer:

            df.to_excel(writer, index=False, sheet_name='Saym Raporu')

            

            # Worksheet'i al ve formatla

            workbook = writer.book

            worksheet = writer.sheets['Saym Raporu']

            

            # Stun geniliklerini ayarla

            worksheet.column_dimensions['A'].width = 15  # Para Kodu

            worksheet.column_dimensions['B'].width = 30  # Para Ad

            worksheet.column_dimensions['C'].width = 15  # Beklenen Adet

            worksheet.column_dimensions['D'].width = 15  # Saylan Adet

            worksheet.column_dimensions['E'].width = 10  # Fark

            worksheet.column_dimensions['F'].width = 12  # Durum

            

            # Header stilini ayarla

            from openpyxl.styles import Font, PatternFill, Alignment

            header_fill = PatternFill(start_color='DC3545', end_color='DC3545', fill_type='solid')

            header_font = Font(bold=True, color='FFFFFF')

            

            for cell in worksheet[1]:

                cell.fill = header_fill

                cell.font = header_font

                cell.alignment = Alignment(horizontal='center', vertical='center')

        

        output.seek(0)



        # Rapor dosyasn kaydet

        report_filename = f'sayim_raporu_{session_id}_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'

        report_path = os.path.join(REPORTS_DIR, report_filename)



        with open(report_path, 'wb') as f:

            f.write(output.getvalue())



        # Doruluk orann hesapla

        accuracy_rate = (total_scanned / total_expected * 100) if total_expected > 0 else 0.0



        # Raporu count_reports table'na kaydet (varsa)

        report_title = f"Saym Raporu - {datetime.now().strftime('%d.%m.%Y %H:%M')}"

        try:

            execute_query(cursor, '''

                INSERT INTO count_reports (session_id, report_filename, report_title, 

                                         total_expected, total_scanned, accuracy_rate, created_at)

                VALUES (, , , , , , )

            ''', (session_id, report_filename, report_title, total_expected, total_scanned, accuracy_rate, datetime.now()))

        except Exception as e:

            # count_reports tablosu yoksa sadece dosya kaydet

            logging.warning(f"count_reports tablosu yok, sadece dosya kaydedildi: {e}")



        # Database ilemini commit et

        conn.commit()

        close_db(conn)



        # WebSocket ile saym bittii bilgisini gnder

        try:

            socketio.emit('count_finished', {'session_id': session_id})

        except Exception as ws_error:

            logging.warning(f"WebSocket notification failed: {str(ws_error)}")



        # Session'dan saym bilgilerini temizle

        session.pop('count_access', None)

        session.pop('current_session', None)



        return jsonify({

            'success': True,

            'message': 'Saym baaryla tamamland',

            'report_file': report_filename,

            'session_id': session_id,

            'total_expected': total_expected,

            'total_scanned': total_scanned,

            'accuracy_rate': round(accuracy_rate, 2)

        })



    except Exception as e:

        # Hata durumunda database balantsn kapatmay garanti et

        try:

            if 'conn' in locals():

                close_db(conn)

        except:

            pass



        error_msg = f"Saym tamamlama hatas: {str(e)}"

        logging.error(error_msg, exc_info=True)



        return jsonify({

            'success': False,

            'error': 'Saym tamamlanamad - sistem hatas',

            'debug': str(e) if app.debug else None

        }), 500



@app.route('/stop_all_counts', methods=['POST'])

def stop_all_counts():

    """Tm aktif saymlar durdur - ACIL DURUMFONKSYONU"""

    # Admin authentication check

    admin_password = request.json.get('admin_password')

    if admin_password != ADMIN_COUNT_PASSWORD:

        return jsonify({'success': False, 'error': 'Yetki gerekli - yanl admin ifresi'}), 403



    conn = get_db()

    cursor = conn.cursor()



    try:

        # Tm aktif saymlar bul

        execute_query(cursor, "SELECT id FROM count_sessions WHERE is_active = TRUE")

        active_sessions = cursor.fetchall()



        if not active_sessions:

            close_db(conn)

            return jsonify({'success': True, 'message': 'Durdurulacak aktif saym bulunamad'})



        # Tm aktif saymlar "completed" olarak iaretle

        stopped_count = 0

        for session_tuple in active_sessions:

            session_id = session_tuple[0]

            execute_query(cursor, 'UPDATE count_sessions SET is_active = , ended_at =  WHERE id = ',

                         (False, datetime.now(), session_id))

            stopped_count += 1



        # Session'lar temizle

        session.pop('count_access', None)

        session.pop('count_authenticated', None) 

        session.pop('current_session', None)



        conn.commit()

        close_db(conn)



        # WebSocket ile tm kullanclara saymlarn durdurulduunu bildir

        socketio.emit('all_counts_stopped', {

            'message': f'{stopped_count} aktif saym durduruldu',

            'stopped_sessions': [s[0] for s in active_sessions]

        })



        return jsonify({

            'success': True,

            'message': f'{stopped_count} aktif saym baaryla durduruldu',

            'stopped_sessions': [s[0] for s in active_sessions]

        })



    except Exception as e:

        conn.rollback()

        close_db(conn)

        return jsonify({'success': False, 'error': f'Sistem hatas: {str(e)}'}), 500



@app.route('/qr_codes')

def qr_codes_page():

    if 'user_id' not in session:

        return render_template('login.html')

    return render_template('qr_codes.html')



@app.route('/parts')

@login_required

def parts_list():

    """Tm paralar listele"""

    conn = get_db()

    cursor = conn.cursor()

    

    # Paralar ve QR kod saylarn getir

    execute_query(cursor, '''

        SELECT 

            pc.id,

            pc.part_code, 

            pc.part_name,

            pc.description,

            pc.created_at,

            COUNT(qc.qr_id) as qr_count

        FROM part_codes pc

        LEFT JOIN qr_codes qc ON pc.id = qc.part_code_id

        GROUP BY pc.id, pc.part_code, pc.part_name, pc.description, pc.created_at

        ORDER BY pc.created_at DESC

    ''')

    

    parts = []

    for row in cursor.fetchall():

        parts.append({

            'id': row[0],

            'part_code': row[1],

            'part_name': row[2],

            'description': row[3] or '',

            'created_at': row[4],

            'qr_count': row[5]

        })

    

    close_db(conn)

    return render_template('parts.html', parts=parts)



@app.route('/parts/<part_code>')

@login_required

def part_detail(part_code):

    """Para detay sayfas - QR kod retme"""

    conn = get_db()

    cursor = conn.cursor()

    placeholder = get_db_placeholder()

    

    # Para bilgilerini getir

    execute_query(cursor, f'SELECT id, part_code, part_name, description, created_at FROM part_codes WHERE part_code = {placeholder}', (part_code,))

    part_row = cursor.fetchone()

    

    if not part_row:

        close_db(conn)

        return "Para bulunamad", 404

    

    part = {

        'id': part_row[0],

        'part_code': part_row[1],

        'part_name': part_row[2],

        'description': part_row[3] or '',

        'created_at': part_row[4]

    }

    

    # Bu paraya ait QR kodlarn getir (SADECE KULLANILMAYANLAR)

    execute_query(cursor, f'''

        SELECT qr_id, created_at, is_used,

               CASE WHEN blob_url IS NOT NULL THEN 1 ELSE 0 END as is_downloaded

        FROM qr_codes 

        WHERE part_code_id = {placeholder} AND is_used = 0

        ORDER BY qr_id DESC

    ''', (part['id'],))

    

    qr_codes = []

    for row in cursor.fetchall():

        qr_codes.append({

            'qr_id': row[0],

            'created_at': row[1],

            'is_used': row[2],

            'is_downloaded': row[3]

        })

    

    # Toplam QR saysn al (kullanlan + kullanlmayan)

    execute_query(cursor, f'SELECT COUNT(*) FROM qr_codes WHERE part_code_id = {placeholder}', (part['id'],))

    total_qr_count = cursor.fetchone()[0]

    

    # Kullanlan QR saysn al

    execute_query(cursor, f'SELECT COUNT(*) FROM qr_codes WHERE part_code_id = {placeholder} AND is_used = 1', (part['id'],))

    used_qr_count = cursor.fetchone()[0]

    

    close_db(conn)

    return render_template('part_detail.html', part=part, qr_codes=qr_codes, total_qr_count=total_qr_count, used_qr_count=used_qr_count)



@app.route('/generate_qr/<part_code>', methods=['POST'])

@login_required

def generate_qr_codes(part_code):

    """Belirtilen para iin birden fazla QR kod ret (quantity parametresi ile)"""

    try:

        req = request.get_json(silent=True) or {}

        quantity = int(req.get('quantity', 1) or 1)

        if quantity < 1:

            quantity = 1

        max_qty = 500

        if quantity > max_qty:

            return jsonify({'success': False, 'error': f'Quantity too large (max {max_qty})'}), 400



        conn = get_db()

        cursor = conn.cursor()

        placeholder = get_db_placeholder()



        # Para bilgilerini al

        execute_query(cursor, f'SELECT id, part_name FROM part_codes WHERE part_code = {placeholder}', (part_code,))

        part_row = cursor.fetchone()



        if not part_row:

            close_db(conn)

            return jsonify({'error': 'Para bulunamad'}), 404



        part_code_id = part_row[0]

        part_name = part_row[1]



        # Mevcut QR kod saysn ren (tm QR'lar, kullanlm olanlar dahil)

        execute_query(cursor, f'SELECT COUNT(*) FROM qr_codes WHERE part_code_id = {placeholder}', (part_code_id,))

        current_count = cursor.fetchone()[0]



        generated = []

        file_paths = []



        for i in range(quantity):

            qr_number = current_count + i + 1

            qr_id = f"{part_code}_{qr_number}"



            execute_query(cursor, f'INSERT INTO qr_codes (qr_id, part_code_id, created_at, is_used) VALUES ({placeholder}, {placeholder}, {placeholder}, {placeholder})',

                         (qr_id, part_code_id, datetime.now(), False))



            save_qr_code_to_file(part_code, qr_id, qr_number)



            generated.append({

                'qr_id': qr_id,

                'qr_number': qr_number,

                'qr_image_url': f'/qr_image/{part_code}/{qr_id}'

            })



            qr_file_path = os.path.join(os.path.dirname(__file__), 'static', 'qr_codes', part_code, f"{qr_id}.png")

            file_paths.append(qr_file_path)



        conn.commit()

        close_db(conn)



        zip_url = None

        try:

            if len(file_paths) > 1:

                zip_dir = os.path.join(os.path.dirname(__file__), 'static', 'qr_codes', part_code)

                os.makedirs(zip_dir, exist_ok=True)

                zip_name = f'bulk_{datetime.now().strftime("%Y%m%d_%H%M%S")}.zip'

                zip_path = os.path.join(zip_dir, zip_name)

                import zipfile

                with zipfile.ZipFile(zip_path, 'w', compression=zipfile.ZIP_DEFLATED) as zf:

                    for fp in file_paths:

                        if os.path.exists(fp):

                            zf.write(fp, arcname=os.path.basename(fp))

                zip_url = f'/static/qr_codes/{part_code}/{zip_name}'

        except Exception:

            logging.exception('ZIP oluturulamad')



        print(f" {part_code} iin {quantity} QR kod retildi. Balang: {current_count + 1}")



        return jsonify({

            'success': True,

            'message': f'{quantity} adet QR kod retildi',

            'generated': generated,

            'zip_url': zip_url

        })



    except Exception as e:

        logging.exception(f"QR kod retme hatas: {e}")

        try:

            close_db(conn)

        except:

            pass

        return jsonify({'error': f'Hata: {str(e)}'}), 500



@app.route('/qr_image/<part_code>/<qr_id>')

@login_required

def serve_qr_image(part_code, qr_id):

    """QR kod grselini serve et (text ile birlikte)"""

    try:

        # nce statik dosyay kontrol et

        qr_file = os.path.join(os.path.dirname(__file__), 'static', 'qr_codes', part_code, f"{qr_id}.png")

        

        if os.path.exists(qr_file):

            return send_file(qr_file, mimetype='image/png')

        

        # Dosya yoksa dinamik olutur (generate_qr_image gibi)

        # QR ID'den para kodunu ve numaray kar

        parts = qr_id.rsplit('_', 1)

        qr_number = parts[1] if len(parts) > 1 else "1"

        

        # Para adn database'den al

        conn = get_db()

        cursor = conn.cursor()

        execute_query(cursor, 'SELECT part_name FROM part_codes WHERE part_code = ', (part_code,))

        result = cursor.fetchone()

        part_name = result[0] if result else ""

        close_db(conn)



        # QR kod olutur - Barkod makinesi iin optimize

        qr = qrcode.QRCode(

            version=1, 

            box_size=8,  # 8px - barkod makinesi iin ideal

            border=2,    # 2px quiet zone

            error_correction=qrcode.constants.ERROR_CORRECT_M

        )

        qr.add_data(qr_id)

        qr.make(fit=True)

        qr_img = qr.make_image(fill_color="black", back_color="white")

        

        # PIL Image'a dntr

        qr_img = qr_img.convert('RGB')



        # QR boyutlar

        qr_width, qr_height = qr_img.size

        

        # Alanlar hesapla

        logo_height = 40  # Logo iin st alan

        text_height = 35  # 1 satr - sadece QR ID

        

        # Krmz ereve iin padding

        border_width = 3  # 3px krmz ereve

        

        # Final grsel (ereveli)

        final_width = qr_width + (border_width * 2)

        final_height = logo_height + qr_height + text_height + (border_width * 2)

        final_img = Image.new('RGB', (final_width, final_height), '#dc2626')  # Krmz arka plan (ereve)

        

        # Beyaz i alan olutur (logo + QR + text)

        white_bg = Image.new('RGB', (qr_width, logo_height + qr_height + text_height), 'white')

        

        # Logo ekle (varsa) - st ortasna

        try:

            logo_path = os.path.join(os.path.dirname(__file__), 'cermak-logo.png')

            if os.path.exists(logo_path):

                logo_img = Image.open(logo_path).convert('RGBA')

                # Logo boyutunu alan yksekliine gre ayarla

                logo_width = 150

                logo_height_logo = 40

                try:

                    logo_img = logo_img.resize((logo_width, logo_height_logo), Image.Resampling.LANCZOS)

                except AttributeError:

                    logo_img = logo_img.resize((logo_width, logo_height_logo), Image.LANCZOS)

                

                # Logo'yu ortala

                logo_x = (qr_width - logo_width) // 2

                logo_y = 5  # stten 5px boluk

                

                # RGBA logo'yu blend et

                if logo_img.mode == 'RGBA':

                    alpha = logo_img.split()[3]

                    logo_img = logo_img.convert('RGB')

                    white_bg.paste(logo_img, (logo_x, logo_y), alpha)

                else:

                    white_bg.paste(logo_img, (logo_x, logo_y))

        except Exception as e:

            print(f"Logo ekleme hatas: {e}")

        

        # QR kodu beyaz alana yaptr (logo'nun altna)

        white_bg.paste(qr_img, (0, logo_height))

        

        # Beyaz alan krmz erevenin iine yaptr

        final_img.paste(white_bg, (border_width, border_width))

        

        # Text ekle

        draw = ImageDraw.Draw(final_img)

        try:

            font = ImageFont.truetype("arialblk.ttf", 24)

        except:

            try:

                font = ImageFont.truetype("arialbd.ttf", 24)

            except:

                try:

                    font = ImageFont.truetype("arial.ttf", 24)

                except:

                    font = ImageFont.load_default()

        

        # Sadece QR ID

        qr_text = qr_id

        text_width = len(qr_text) * 14

        x_position = max(border_width, (final_width - text_width) // 2)

        draw.text((x_position, logo_height + qr_height + border_width + 5), qr_text, fill='black', font=font)



        # BytesIO'ya kaydet

        buf = BytesIO()

        final_img.save(buf, format='PNG', optimize=True)

        buf.seek(0)

        

        return send_file(buf, mimetype='image/png')

        

    except Exception as e:

        logging.exception(f"QR grsel hatas: {e}")

        return f"Hata: {str(e)}", 500



@app.route('/mark_qr_used/<qr_id>', methods=['POST'])

@login_required

def mark_qr_used(qr_id):

    """QR kodu manuel olarak 'kullanld' iaretle"""

    try:

        conn = get_db()

        cursor = conn.cursor()

        placeholder = get_db_placeholder()

        

        # QR kodunu kontrol et

        execute_query(cursor, f'SELECT qr_id, is_used FROM qr_codes WHERE qr_id = {placeholder}', (qr_id,))

        qr = cursor.fetchone()

        

        if not qr:

            close_db(conn)

            return jsonify({'error': 'QR kod bulunamad'}), 404

        

        if qr[1]:  # is_used

            close_db(conn)

            return jsonify({'error': 'Bu QR kod zaten kullanlm'}), 400

        

        # is_used = True yap

        execute_query(cursor, f'UPDATE qr_codes SET is_used = {placeholder} WHERE qr_id = {placeholder}',

                     (True, qr_id))

        conn.commit()

        close_db(conn)

        

        print(f" QR kod manuel kullanld iaretlendi: {qr_id}")

        

        return jsonify({

            'success': True,

            'message': f'QR kod kullanld olarak iaretlendi: {qr_id}'

        })

        

    except Exception as e:

        logging.exception(f"QR iaretleme hatas: {e}")

        try:

            close_db(conn)

        except:

            pass

        return jsonify({'error': f'Hata: {str(e)}'}), 500



@app.route('/check_existing_qrs')

@login_required

def check_existing_qrs():

    conn = get_db()

    cursor = conn.cursor()

    db.session.query(QRCode).count())

    count = cursor.fetchone()[0]

    close_db(conn)



    return jsonify({

        'hasQRs': count > 0,

        'count': count

    })



@app.route('/qr_management', methods=['GET'])

@login_required

def qr_management():

    """QR Ynetim Paneli - Gvenli QR ilemleri"""

    if not session.get('admin_authenticated'):

        return redirect('/admin')

    return render_template('qr_management.html')



@app.route('/get_reports')

@login_required

def get_reports():

    conn = None

    try:

        # Veritabanndan raporlar ek

        conn = get_db()

        cursor = conn.cursor()

        

        execute_query(cursor, '''

            SELECT 

                cr.id,

                cr.session_id,

                cr.report_filename,

                cr.report_title,

                cr.total_expected,

                cr.total_scanned,

                cr.accuracy_rate,

                cr.created_at,

                u.username as created_by

            FROM count_reports cr

            LEFT JOIN count_sessions cs ON cr.session_id = cs.id

            LEFT JOIN envanter_users u ON cs.created_by = u.id

            ORDER BY cr.created_at DESC

        ''')

        

        rows = cursor.fetchall()

        reports = []

        

        for row in rows:

            total_difference = (row[5] or 0) - (row[4] or 0)  # scanned - expected

            

            reports.append({

                'id': row[0],

                'session_id': row[1],

                'filename': row[2],

                'title': row[3] or f"Saym Raporu #{row[1]}",

                'created_at': row[7],

                'total_expected': row[4] or 0,

                'total_scanned': row[5] or 0,

                'accuracy_rate': round(row[6] or 0.0, 2),

                'session_name': f"Oturum #{row[1]}",

                'created_by': row[8] or 'Sistem',

                'total_difference': total_difference

            })

        

        return jsonify(reports)



    except Exception as e:

        logging.exception(f"Error in get_reports: {e}")

        return jsonify({'error': str(e)}), 500

    finally:

        try:

            if conn:

                close_db(conn)

        except Exception:

            pass



@app.route('/download_report/<filename>')

@login_required

def download_report(filename):

    #  Gvenlik: Filename formatn kontrol et (session_id integer olabilir)

    # Format: sayim_raporu_{session_id}_{timestamp}.xlsx

    if not re.match(r'^sayim_raporu_\d+_\d{8}_\d{6}\.xlsx$', filename):

        print(f" Invalid filename format: {filename}")

        return jsonify({'error': 'Geersiz dosya ad format'}), 400



    safe_filename = secure_filename(filename)

    report_path = os.path.join(REPORTS_DIR, safe_filename)



    if not os.path.exists(report_path):

        print(f" Report file not found: {report_path}")

        return jsonify({'error': 'Rapor dosyas bulunamad'}), 404



    #  Path traversal attack'e kar gvenlik kontrol

    real_path = os.path.realpath(report_path)

    reports_real_path = os.path.realpath(REPORTS_DIR)



    if not real_path.startswith(reports_real_path):

        print(f" Invalid path (security): {real_path}")

        return jsonify({'error': 'Geersiz dosya yolu'}), 403



    print(f" Sending report file: {report_path}")

    return send_file(report_path, as_attachment=True, download_name=filename)



#  SAYIM YNETM (Admin Count - ifre sistemi kaldrld)

@app.route('/admin_count')

@login_required

@admin_required_decorator

def admin_count_page():

    """Admin saym kontrol sayfas - Excel ykleme ile saym balatma"""

    return render_template('admin_count.html')



@app.route('/admin_count/start_count', methods=['POST'])

@login_required

@admin_required_decorator

def admin_start_count():

    """Admin saym balatma endpoint'i"""

    print("DEBUG: admin_start_count arld")

    

    try:

        conn = get_db()

        cursor = conn.cursor()

        

        # Aktif saym var m kontrol et

        execute_query(cursor, 'SELECT COUNT(*) FROM count_sessions WHERE is_active = ', (True,))

        active_count = cursor.fetchone()[0]

        

        if active_count > 0:

            close_db(conn)

            return jsonify({

                'success': False,

                'error': 'Zaten aktif bir saym oturumu var!'

            }), 400

        

        # Yeni saym oturumu olutur (PAROLA YOK ARTIK)

        current_user_id = session.get('user_id')

        

        # Toplam beklenen adet

        db.session.query(QRCode).count())

        total_expected = cursor.fetchone()[0]

        

        execute_query(cursor, '''

            INSERT INTO count_sessions 

            (is_active, started_at, created_by, created_at, total_expected, total_scanned) 

            VALUES (, , , , , )

        ''', (True, datetime.now(), current_user_id, datetime.now(), total_expected, 0))

        

        # ID'yi al (SQLite auto-increment)

        session_id = cursor.lastrowid

        

        conn.commit()

        close_db(conn)

        

        print(f" Saym oturumu balatld: {session_id}")

        

        #  SOCKET.IO BLDRM: Tm clientlara yeni saym baladn syle

        try:

            socketio.emit('session_reset', {

                'session_id': session_id,

                'total_expected': total_expected,

                'message': ' Yeni saym balatld - sayfa sfrlanyor...'

            }, broadcast=True)

            print(f" Socket bildirimi gnderildi: session_reset")

        except Exception as socket_err:

            print(f" Socket bildirimi gnderilemedi: {socket_err}")

        

        return jsonify({

            'success': True,

            'message': 'Saym oturumu balatld! PSC Scanner ile QR okutmaya balayabilirsiniz.',

            'session_id': session_id,

            'total_expected': total_expected,

            'redirect': '/count-scanner'  # PSC Scanner sayfasna ynlendir

        })

        

    except Exception as e:

        print(f" Saym balatma hatas: {e}")

        import traceback

        traceback.print_exc()

        

        if conn:

            try:

                conn.rollback()

                close_db(conn)

            except:

                pass

        

        return jsonify({

            'success': False,

            'error': f'Saym balatlamad: {str(e)}'

        }), 500



# API Endpoints for Dashboard Statistics

@app.route('/api/qr_codes')

@login_required

def api_get_qr_codes():

    """QR kodlar listesi - istatistik iin"""

    conn = get_db()

    cursor = conn.cursor()



    execute_query(cursor, '''

        SELECT qc.qr_id, pc.part_code, pc.part_name, qc.is_used, 

               CASE WHEN qc.blob_url IS NOT NULL THEN 1 ELSE 0 END as is_downloaded,

               qc.created_at

        FROM qr_codes qc

        LEFT JOIN part_codes pc ON qc.part_code_id = pc.id

        ORDER BY qc.created_at DESC

    ''')



    qr_codes = []

    for row in cursor.fetchall():

        qr_codes.append({

            'qr_id': row[0],

            'part_code': row[1] or '',

            'part_name': row[2] or 'Bilinmeyen',

            'is_used': bool(row[3]),

            'is_downloaded': bool(row[4]),

            'created_at': row[5]

        })



    close_db(conn)

    return jsonify(qr_codes)



@app.route('/api/reports')

@login_required

def api_get_reports():

    """Raporlar listesi - istatistik iin"""

    conn = get_db()

    cursor = conn.cursor()



    execute_query(cursor, '''

        SELECT id, session_id, report_name, file_path, created_at, 

               total_expected, total_scanned, accuracy_rate

        FROM count_reports

        ORDER BY created_at DESC

    ''')



    reports = []

    for row in cursor.fetchall():

        reports.append({

            'id': row[0],

            'session_id': row[1],

            'report_name': row[2],

            'file_path': row[3],

            'created_at': row[4],

            'total_expected': row[5],

            'total_scanned': row[6],

            'accuracy_rate': row[7]

        })





    close_db(conn)

    return jsonify(reports)



@app.route('/api/dashboard_stats')

def api_dashboard_stats():



    """Dashboard iin genel istatistikler"""

    print("DEBUG: /api/dashboard_stats endpoint arld")  # DEBUG

    conn = get_db()

    cursor = conn.cursor()









    # QR kodlar says

    db.session.query(QRCode).count())

    total_qr_codes = cursor.fetchone()[0]



    # Raporlar says

    db.session.query(CountSession).count())

    total_reports = cursor.fetchone()[0]



    # Saym bilgileri geici olarak sfr

    active_counts = 0

    completed_counts = 0

    last_count_date = None



    close_db(conn)



    stats = {

        'total_qr_codes': total_qr_codes,

        'total_reports': total_reports,

        'active_counts': active_counts,

        'completed_counts': completed_counts,

        'last_count_date': last_count_date

    }

    print(f"DEBUG: Gnderilen stats: {stats}")  # DEBUG

    return jsonify(stats)



# Eksik endpoint'ler

@app.route('/get_session_stats')

@login_required

def get_session_stats():

    """Saym session istatistikleri"""

    try:

        # URL parametresinden session_id al

        requested_session_id = request.args.get('session_id')

        

        conn = get_db()

        cursor = conn.cursor()

        

        #  Eer session_id verilmise, o session' kullan

        if requested_session_id:

            session_id = requested_session_id

            

            # Test mode iin expected=3

            expected = 3 if requested_session_id.startswith('test-') else 0

            

            # count_sessions tablosundan expected deerini al (varsa)

            try:

                execute_query(cursor, '''

                    SELECT total_expected

                    FROM count_sessions

                    WHERE id = ?

                ''', (session_id,))

                row = cursor.fetchone()

                if row and row[0]:

                    expected = row[0]

            except:

                pass

            

        else:

            # Session ID verilmemise, aktif session bul

            execute_query(cursor, '''

                SELECT id, total_expected, total_scanned

                FROM count_sessions

                WHERE is_active = 

                ORDER BY started_at DESC

                LIMIT 1

            ''', (True,))



            row = cursor.fetchone()



            if not row:

                close_db(conn)

                return jsonify({'success': False, 'message': 'No active session', 'scanned': 0, 'expected': 0, 'scanned_qrs': []})



            session_id = row[0]

            expected = row[1] if row[1] is not None else 0

        

        # scanned_qr tablosundan gerek scan saysn al

        execute_query(cursor, '''

            SELECT COUNT(DISTINCT qr_id)

            FROM scanned_qr

            WHERE session_id = 

        ''', (session_id,))

        

        count_row = cursor.fetchone()

        scanned_count = count_row[0] if count_row else 0

        

        # Scanned QR listesi (en son taranan en stte)

        execute_query(cursor, '''

            SELECT qr_id, MAX(scanned_at) as last_scan

            FROM scanned_qr

            WHERE session_id = 

            GROUP BY qr_id

            ORDER BY last_scan DESC

            LIMIT 500

        ''', (session_id,))

        

        scanned_rows = cursor.fetchall()

        scanned_qrs = [r[0] for r in scanned_rows] if scanned_rows else []



        close_db(conn)



        return jsonify({

            'success': True,

            'session_id': session_id,

            'scanned': scanned_count,

            'expected': expected,

            'scanned_qrs': scanned_qrs

        })

        

    except Exception as e:

        logging.error(f"Error in get_session_stats: {e}")

        return jsonify({'error': str(e)}), 500



@app.route('/get_recent_activities')

@login_required

def get_recent_activities():

    """Son QR tarama aktiviteleri - kullanc adlaryla"""

    try:

        conn = get_db()

        cursor = conn.cursor()

        

        # Get recent scanned QRs with user names

        execute_query(cursor, '''

            SELECT 

                sq.qr_id, 

                sq.scanned_by, 

                sq.scanned_at, 

                sq.part_code,

                eu.full_name,

                eu.username,

                pc.part_name

            FROM scanned_qr sq

            LEFT JOIN envanter_users eu ON sq.scanned_by = eu.id

            LEFT JOIN qr_codes qc ON sq.qr_id = qc.qr_id

            LEFT JOIN part_codes pc ON qc.part_code_id = pc.id

            ORDER BY sq.scanned_at DESC

            LIMIT 10

        ''')

        

        activities = []

        for row in cursor.fetchall():

            user_name = row[4] if row[4] else (row[5] if row[5] else f'Kullanc #{row[1]}')

            part_name = row[6] if row[6] else 'Bilinmeyen rn'

            

            activities.append({

                'qr_code': row[0],

                'scanned_by': user_name,

                'scanned_by_id': row[1],

                'scanned_at': row[2],

                'part_code': row[3] if row[3] else None,

                'part_name': part_name

            })

        

        close_db(conn)

        return jsonify(activities)

    except Exception as e:

        logging.error(f"Error in get_recent_activities: {e}")

        return jsonify([])  # Return empty array instead of error object



@app.route('/get_live_count_status')

@login_required

def get_live_count_status():

    """Aktif saym srasnda para baznda anlk durum"""

    try:

        conn = get_db()

        cursor = conn.cursor()

        placeholder = get_db_placeholder()

        

        # Aktif saym var m

        execute_query(cursor, f'SELECT id FROM count_sessions WHERE is_active = {placeholder}', (True,))

        active_session = cursor.fetchone()

        

        if not active_session:

            close_db(conn)

            return jsonify({

                'active': False,

                'message': 'Aktif saym bulunamad'

            })

        

        session_id = active_session[0]

        

        # Para baznda saym durumu

        execute_query(cursor, '''

            SELECT 

                pc.part_code,

                pc.part_name,

                COUNT(DISTINCT qc.qr_id) as beklenen_adet,

                COUNT(DISTINCT sq.qr_id) as sayilan_adet

            FROM part_codes pc

            LEFT JOIN qr_codes qc ON pc.id = qc.part_code_id

            LEFT JOIN scanned_qr sq ON qc.qr_id = ?sq.qr_id AND sq.session_id = 

            GROUP BY pc.id, pc.part_code, pc.part_name

            HAVING beklenen_adet > 0

            ORDER BY pc.part_name

        ''', (str(session_id),))

        

        parts = []

        for row in cursor.fetchall():

            part_code = row[0]

            part_name = row[1]

            beklenen = row[2]

            sayilan = row[3]

            kalan = beklenen - sayilan

            durum = 'Tamamland' if kalan == 0 else f'{kalan} eksik'

            yuzd = round((sayilan / beklenen * 100) if beklenen > 0 else 0, 1)

            

            parts.append({

                'part_code': part_code,

                'part_name': part_name,

                'beklenen_adet': beklenen,

                'sayilan_adet': sayilan,

                'kalan_adet': kalan,

                'durum': durum,

                'tamamlanma_yuzdesi': yuzd

            })

        

        close_db(conn)

        return jsonify({

            'active': True,

            'session_id': session_id,

            'parts': parts,

            'total_parts': len(parts),

            'completed_parts': len([p for p in parts if p['kalan_adet'] == 0])

        })

        

    except Exception as e:

        logging.error(f"Error in get_live_count_status: {e}")

        return jsonify({'active': False, 'error': str(e)})



@app.route('/export_live_count')

@login_required

def export_live_count():

    """Anlk saym durumunu Excel olarak indir"""

    try:

        from openpyxl import Workbook

        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

        from io import BytesIO

        

        conn = get_db()

        cursor = conn.cursor()

        placeholder = get_db_placeholder()

        

        # Aktif saym var m

        execute_query(cursor, f'SELECT id FROM count_sessions WHERE is_active = {placeholder}', (True,))

        active_session = cursor.fetchone()

        

        if not active_session:

            close_db(conn)

            return jsonify({'error': 'Aktif saym bulunamad'}), 404

        

        session_id = active_session[0]

        

        # Para baznda saym durumu

        execute_query(cursor, '''

            SELECT 

                pc.part_code,

                pc.part_name,

                COUNT(DISTINCT qc.qr_id) as beklenen_adet,

                COUNT(DISTINCT sq.qr_id) as sayilan_adet

            FROM part_codes pc

            LEFT JOIN qr_codes qc ON pc.id = qc.part_code_id

            LEFT JOIN scanned_qr sq ON qc.qr_id = ?sq.qr_id AND sq.session_id = 

            GROUP BY pc.id, pc.part_code, pc.part_name

            HAVING beklenen_adet > 0

            ORDER BY pc.part_name

        ''', (str(session_id),))

        

        # Excel olutur

        wb = Workbook()

        ws = wb.active

        ws.title = "Canl Saym Durumu"

        

        # Header

        headers = ['Para Kodu', 'Para Ad', 'Beklenen Adet', 'Saylan Adet', 'Kalan Adet', 'Tamamlanma %', 'Durum']

        ws.append(headers)

        

        # Header stil

        header_fill = PatternFill(start_color="DC3545", end_color="DC3545", fill_type="solid")

        header_font = Font(bold=True, color="FFFFFF", size=12)

        

        for cell in ws[1]:

            cell.fill = header_fill

            cell.font = header_font

            cell.alignment = Alignment(horizontal='center', vertical='center')

        

        # Veri satrlar

        for row in cursor.fetchall():

            beklenen = row[2]

            sayilan = row[3]

            kalan = beklenen - sayilan

            yuzde = round((sayilan / beklenen * 100) if beklenen > 0 else 0, 1)

            durum = ' Tamamland' if kalan == 0 else f' {kalan} eksik'

            

            ws.append([

                row[0],  # part_code

                row[1],  # part_name

                beklenen,

                sayilan,

                kalan,

                f"{yuzde}%",

                durum

            ])

        

        # Kolon genilikleri

        ws.column_dimensions['A'].width = 20

        ws.column_dimensions['B'].width = 35

        ws.column_dimensions['C'].width = 15

        ws.column_dimensions['D'].width = 15

        ws.column_dimensions['E'].width = 15

        ws.column_dimensions['F'].width = 15

        ws.column_dimensions['G'].width = 20

        

        close_db(conn)

        

        # Excel dosyasn BytesIO'ya kaydet

        excel_file = BytesIO()

        wb.save(excel_file)

        excel_file.seek(0)

        

        filename = f"Canli_Sayim_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

        

        return send_file(

            excel_file,

            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',

            as_attachment=True,

            download_name=filename

        )

        

    except Exception as e:

        logging.exception(f"Error in export_live_count: {e}")

        try:

            close_db(conn)

        except:

            pass

        return jsonify({'error': str(e)}), 500



@app.route('/export_qr_activities')

@login_required

def export_qr_activities():

    """QR tarama hareketlerini Excel olarak indir"""

    try:

        from openpyxl import Workbook

        from openpyxl.styles import Font, Alignment, PatternFill

        from io import BytesIO

        

        conn = get_db()

        cursor = conn.cursor()

        

        # Tm QR tarama hareketlerini ek

        execute_query(cursor, '''

            SELECT 

                sq.qr_id, 

                sq.scanned_by, 

                sq.scanned_at, 

                sq.part_code,

                sq.session_id,

                eu.full_name,

                eu.username,

                pc.part_name,

                pc.part_code as part_code_full

            FROM scanned_qr sq

            LEFT JOIN envanter_users eu ON sq.scanned_by = eu.id

            LEFT JOIN qr_codes qc ON sq.qr_id = qc.qr_id

            LEFT JOIN part_codes pc ON qc.part_code_id = pc.id

            ORDER BY sq.scanned_at DESC

        ''')

        

        activities = cursor.fetchall()

        close_db(conn)

        

        # Excel olutur

        wb = Workbook()

        ws = wb.active

        ws.title = "QR Hareketleri"

        

        # Balk satr

        headers = ['QR Kodu', 'Para Kodu', 'Para Ad', 'Okuyan Kullanc', 

                  'Kullanc Ad', 'Okuma Tarihi', 'Seans ID']

        ws.append(headers)

        

        # Balk stilini ayarla

        header_fill = PatternFill(start_color="DC3545", end_color="DC3545", fill_type="solid")

        header_font = Font(bold=True, color="FFFFFF", size=12)

        

        for cell in ws[1]:

            cell.fill = header_fill

            cell.font = header_font

            cell.alignment = Alignment(horizontal='center', vertical='center')

        

        # Veri satrlar

        for row in activities:

            qr_id = row[0]

            scanned_by_id = row[1]

            scanned_at = row[2]

            part_code = row[3] or row[8] or ''  # sq.part_code veya pc.part_code

            session_id = row[4]

            full_name = row[5] or ''

            username = row[6] or f'user_{scanned_by_id}'

            part_name = row[7] or 'Bilinmeyen rn'

            

            ws.append([

                qr_id,

                part_code,

                part_name,

                full_name,

                username,

                scanned_at,

                session_id

            ])

        

        # Kolon geniliklerini ayarla

        ws.column_dimensions['A'].width = 25  # QR Kodu

        ws.column_dimensions['B'].width = 15  # Para Kodu

        ws.column_dimensions['C'].width = 30  # Para Ad

        ws.column_dimensions['D'].width = 20  # Okuyan Kullanc

        ws.column_dimensions['E'].width = 15  # Kullanc Ad

        ws.column_dimensions['F'].width = 20  # Okuma Tarihi

        ws.column_dimensions['G'].width = 10  # Seans ID

        

        # Excel'i memory'ye kaydet

        excel_file = BytesIO()

        wb.save(excel_file)

        excel_file.seek(0)

        

        # Dosya ad

        filename = f"QR_Hareketleri_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

        

        return send_file(

            excel_file,

            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',

            as_attachment=True,

            download_name=filename

        )

        

    except Exception as e:

        logging.error(f"Error in export_qr_activities: {e}")

        return jsonify({'error': str(e)}), 500



@app.route('/get_count_status')

@login_required

def get_count_status():

    """Aktif saym durumu"""

    try:

        conn = get_db()

        cursor = conn.cursor()

        

        execute_query(cursor, '''

            SELECT id, started_at, total_expected, total_scanned, is_active

            FROM count_sessions

            WHERE is_active = 

            ORDER BY started_at DESC

            LIMIT 1

        ''', (True,))

        

        row = cursor.fetchone()

        close_db(conn)

        

        if row:

            return jsonify({

                'active': True,

                'session_id': row[0],

                'started_at': row[1],

                'total_expected': row[2],

                'total_scanned': row[3]

            })

        else:

            return jsonify({'active': False})

    except Exception as e:

        logging.error(f"Error in get_count_status: {e}")

        return jsonify({'error': str(e)}), 500



#  KULLANICI YNETM (Sadece Admin)

@app.route('/api/users', methods=['GET'])

@login_required

def get_users():

    """Tm kullanclar listele - Sadece Admin"""

    if session.get('role') != 'admin':

        return jsonify({'error': 'Admin eriimi gerekli'}), 403

    

    try:

        conn = get_db()

        cursor = conn.cursor()

        execute_query(cursor, '''

            SELECT id, username, full_name, role, is_active_user, can_mark_used

            FROM envanter_users

            ORDER BY id

        ''')

        users = cursor.fetchall()

        close_db(conn)

        

        user_list = []

        for user in users:

            user_list.append({

                'id': user[0],

                'username': user[1],

                'full_name': user[2],

                'role': user[3],

                'is_active': user[4],

                'can_mark_used': user[5]

            })

        

        return jsonify({'success': True, 'users': user_list})

    except Exception as e:

        logging.error(f"Error in get_users: {e}")

        return jsonify({'error': str(e)}), 500



@app.route('/api/users', methods=['POST'])

@login_required

def create_user():

    """Yeni kullanc olutur - Sadece Admin"""

    if session.get('role') != 'admin':

        return jsonify({'error': 'Admin eriimi gerekli'}), 403

    

    from werkzeug.security import generate_password_hash

    

    try:

        data = request.get_json()

        username = data.get('username')

        password = data.get('password')

        full_name = data.get('full_name')

        role = data.get('role', 'user')  # Varsaylan: user (sadece QR okutabilir)

        

        if not username or not password or not full_name:

            return jsonify({'error': 'Kullanc ad, ifre ve tam ad gerekli'}), 400

        

        # Sadece 'user' ve 'admin' rolleri kabul et

        if role not in ['user', 'admin']:

            role = 'user'

        

        # Admin saysn kontrol et - Sadece 1 admin olmal

        conn = get_db()

        cursor = conn.cursor()

        

        if role == 'admin':

            execute_query(cursor, "SELECT COUNT(*) FROM envanter_users WHERE role = 'admin'")

            admin_count = cursor.fetchone()[0]

            if admin_count > 0:

                close_db(conn)

                return jsonify({'error': 'Sistemde zaten bir admin var. Sadece 1 admin olabilir.'}), 400

        

        # Kullanc ad kontrol

        execute_query(cursor, 'SELECT COUNT(*) FROM envanter_users WHERE username = ', (username,))

        if cursor.fetchone()[0] > 0:

            close_db(conn)

            return jsonify({'error': 'Bu kullanc ad zaten kullanlyor'}), 400

        

        # ifreyi hashle

        password_hash = generate_password_hash(password)

        

        # Kullanc rolne gre izinler

        can_mark_used = (role == 'admin')  # Sadece admin QR'lar kullanlm iaretleyebilir

        

        # Kullancy ekle

        execute_query(cursor, '''

            INSERT INTO envanter_users 

            (username, password, password_hash, full_name, role, is_active_user, can_mark_used)

            VALUES (, , , , , , )

        ''', (username, password, password_hash, full_name, role, True, can_mark_used))

        

        conn.commit()

        close_db(conn)

        

        return jsonify({

            'success': True,

            'message': f'Kullanc {username} baaryla oluturuldu',

            'user': {

                'username': username,

                'full_name': full_name,

                'role': role,

                'can_mark_used': can_mark_used

            }

        })

    except Exception as e:

        logging.error(f"Error in create_user: {e}")

        return jsonify({'error': str(e)}), 500



@app.route('/api/users/<int:user_id>', methods=['DELETE'])

@login_required

def delete_user(user_id):

    """Kullanc sil - Sadece Admin"""

    if session.get('role') != 'admin':

        return jsonify({'error': 'Admin eriimi gerekli'}), 403

    

    try:

        conn = get_db()

        cursor = conn.cursor()

        

        # Admin kendini silemez

        if user_id == session.get('user_id'):

            close_db(conn)

            return jsonify({'error': 'Kendi hesabnz silemezsiniz'}), 400

        

        # Kullanc kontrol

        execute_query(cursor, 'SELECT username, role FROM envanter_users WHERE id = ', (user_id,))

        user = cursor.fetchone()

        

        if not user:

            close_db(conn)

            return jsonify({'error': 'Kullanc bulunamad'}), 404

        

        # Admin silinemez

        if user[1] == 'admin':

            close_db(conn)

            return jsonify({'error': 'Admin kullancs silinemez'}), 400

        

        # Sil

        execute_query(cursor, 'DELETE FROM envanter_users WHERE id = ', (user_id,))

        conn.commit()

        close_db(conn)

        

        return jsonify({'success': True, 'message': f'Kullanc {user[0]} silindi'})

    except Exception as e:

        logging.error(f"Error in delete_user: {e}")

        return jsonify({'error': str(e)}), 500



@app.route('/admin')

@login_required

def admin_panel():

    """Admin panel - ifre korumal"""

    # Admin kontrol

    if session.get('role') != 'admin':

        return jsonify({'error': 'Admin eriimi gerekli'}), 403

    

    # Admin panel ifresi kontrol (session'da tutulur)

    if not session.get('admin_panel_unlocked'):

        # ifre girilmemi, ifre sayfasn gster

        return render_template('admin_login.html')

    

    # ifre doru, admin paneli gster

    return render_template('admin.html')



@app.route('/admin/verify', methods=['POST'])

@login_required

def verify_admin_password():

    """Admin panel ifre dorulama"""

    if session.get('role') != 'admin':

        return jsonify({'error': 'Admin eriimi gerekli'}), 403

    

    data = request.get_json()

    password = data.get('password')

    

    # Admin saym ifresi

    ADMIN_PASSWORD = '@R9t$L7e!xP2w'

    

    if password == ADMIN_PASSWORD:

        # ifreyi session'a kaydet

        session['admin_panel_unlocked'] = True

        return jsonify({'success': True, 'message': 'Admin paneline hogeldiniz'})

    else:

        return jsonify({'success': False, 'error': 'Hatal ifre'}), 401



@app.route('/admin/logout', methods=['POST'])

@login_required

def admin_panel_logout():

    """Admin panel k (sadece panel kilidini kaldr)"""

    session.pop('admin_panel_unlocked', None)

    return jsonify({'success': True})



@app.route('/admin/users')

@login_required

def admin_users_page():

    """Kullanc ynetim sayfas veya API - GET request"""

    # Admin kontrol

    if session.get('role') != 'admin':

        return jsonify({'error': 'Admin eriimi gerekli'}), 403

    

    # Accept header kontrol - JSON mu HTML mi isteniyor

    if request.accept_mimetypes.best_match(['application/json', 'text/html']) == 'application/json':

        # API ars - Kullanc listesini dndr

        try:

            conn = get_db()

            cursor = conn.cursor()

            

            execute_query(cursor, '''

                SELECT id, username, full_name, role, created_at, is_active_user

                FROM envanter_users

                ORDER BY created_at DESC

            ''')

            

            users = []

            for row in cursor.fetchall():

                users.append({

                    'id': row[0],

                    'username': row[1],

                    'fullname': row[2] or '',

                    'role': row[3],

                    'created_at': row[4],

                    'is_active': row[5] if len(row) > 5 else True

                })

            

            close_db(conn)

            return jsonify(users)

            

        except Exception as e:

            logging.error(f"Error in get users API: {e}")

            return jsonify({'error': str(e)}), 500

    else:

        # HTML sayfas - Kullanc ynetim arayz

        # Admin panel kilidini kontrol et

        if not session.get('admin_panel_unlocked'):

            return redirect('/admin')

        

        # Kullanclar veritabanndan ek

        try:

            conn = get_db()

            cursor = conn.cursor()

            

            execute_query(cursor, '''

                SELECT id, username, full_name, role, created_at, is_active_user

                FROM envanter_users

                ORDER BY created_at DESC

            ''')

            

            users = []

            for row in cursor.fetchall():

                users.append({

                    'id': row[0],

                    'username': row[1],

                    'full_name': row[2] or 'simsiz',

                    'role': row[3],

                    'created_at': row[4],

                    'is_active': row[5] if len(row) > 5 else True

                })

            

            close_db(conn)

            

            # Kullanclar template'e gnder

            return render_template('admin_users.html', users=users)

            

        except Exception as e:

            logging.error(f"Error loading users page: {e}")

            return render_template('admin_users.html', users=[])



@app.route('/admin/users', methods=['POST'])

@login_required

def admin_create_user():

    """Yeni kullanc olutur - Admin Users sayfas"""

    from werkzeug.security import generate_password_hash

    

    # Admin kontrol

    if session.get('role') != 'admin':

        return jsonify({'error': 'Admin eriimi gerekli'}), 403

    

    try:

        # Debug: Request bilgilerini logla

        logging.info(f"Content-Type: {request.content_type}")

        logging.info(f"Request data: {request.data}")

        

        # JSON verisini al - force=True ile Content-Type kontroln bypass et

        data = request.get_json(force=True)

        if not data:

            logging.error("JSON data is None or empty")

            return jsonify({'error': 'Geersiz JSON verisi'}), 400

        

        logging.info(f"Parsed JSON: {data}")

        

        username = str(data.get('username', '')).strip()

        password = str(data.get('password', '')).strip()

        fullname = str(data.get('fullname', '')).strip()

        role = str(data.get('role', 'user')).strip()

        

        # Validasyonlar

        if not username or not password:

            return jsonify({'error': 'Kullanc ad ve ifre gerekli'}), 400

        

        if len(username) < 3:

            return jsonify({'error': 'Kullanc ad en az 3 karakter olmal'}), 400

        

        if len(password) < 4:

            return jsonify({'error': 'ifre en az 4 karakter olmal'}), 400

        

        if role not in ['admin', 'user']:

            return jsonify({'error': 'Geersiz rol. admin veya user olmal'}), 400

        

        # Kullanc zaten var m

        conn = get_db()

        cursor = conn.cursor()

        

        execute_query(cursor, 'SELECT id FROM envanter_users WHERE username = ', (username,))

        if cursor.fetchone():

            close_db(conn)

            return jsonify({'error': 'Bu kullanc ad zaten kullanlyor'}), 400

        

        # ifreyi hashle

        hashed_password = generate_password_hash(password)

        

        # Yeni kullanc ekle

        execute_query(cursor, '''

            INSERT INTO envanter_users (username, password_hash, full_name, role, created_at, is_active_user)

            VALUES (, , , , , )

        ''', (username, hashed_password, fullname or '', role, datetime.now(), True))

        

        conn.commit()

        close_db(conn)

        

        return jsonify({'success': True, 'message': f'Kullanc {username} oluturuldu'})

        

    except Exception as e:

        logging.error(f"Error in admin_create_user: {e}")

        return jsonify({'error': str(e)}), 500



@app.route('/admin/users/<int:user_id>', methods=['PUT'])

@login_required

def admin_update_user(user_id):

    """Kullanc bilgilerini gncelle - Admin Users sayfas"""

    # Admin kontrol

    if session.get('role') != 'admin':

        return jsonify({'error': 'Admin eriimi gerekli'}), 403

    

    try:

        data = request.get_json()

        if not data:

            return jsonify({'error': 'Geersiz JSON verisi'}), 400

        

        fullname = str(data.get('fullname', '')).strip()

        role = str(data.get('role', 'user')).strip()

        

        if role not in ['admin', 'user']:

            return jsonify({'error': 'Geersiz rol'}), 400

        

        conn = get_db()

        cursor = conn.cursor()

        

        # Kullanc var m kontrol et

        execute_query(cursor, 'SELECT id FROM envanter_users WHERE id = ', (user_id,))

        if not cursor.fetchone():

            close_db(conn)

            return jsonify({'error': 'Kullanc bulunamad'}), 404

        

        # Kullancy gncelle

        execute_query(cursor, '''

            UPDATE envanter_users 

            SET full_name = , role = , updated_at = 

            WHERE id = ?

        ''', (fullname, role, datetime.now(), user_id))

        

        conn.commit()

        close_db(conn)

        

        return jsonify({'success': True, 'message': 'Kullanc gncellendi'})

        

    except Exception as e:

        logging.error(f"Error in admin_update_user: {e}")

        return jsonify({'error': str(e)}), 500



@app.route('/admin/users/<int:user_id>', methods=['DELETE'])

@login_required

def admin_delete_user(user_id):

    """Kullancy sil - Admin Users sayfas"""

    # Admin kontrol

    if session.get('role') != 'admin':

        return jsonify({'error': 'Admin eriimi gerekli'}), 403

    

    try:

        # Kendi hesabn silemesin

        if user_id == session.get('user_id'):

            return jsonify({'error': 'Kendi hesabnz silemezsiniz'}), 400

        

        conn = get_db()

        cursor = conn.cursor()

        

        # Kullanc bilgilerini al

        execute_query(cursor, 'SELECT username FROM envanter_users WHERE id = ', (user_id,))

        user = cursor.fetchone()

        

        if not user:

            close_db(conn)

            return jsonify({'error': 'Kullanc bulunamad'}), 404

        

        # Kullancy sil

        execute_query(cursor, 'DELETE FROM envanter_users WHERE id = ', (user_id,))

        conn.commit()

        close_db(conn)

        

        return jsonify({'success': True, 'message': f'Kullanc {user[0]} silindi'})

        

    except Exception as e:

        logging.error(f"Error in admin_delete_user: {e}")

        return jsonify({'error': str(e)}), 500



@app.route('/admin/users/<int:user_id>/change_password', methods=['POST'])

@login_required

def admin_change_password(user_id):

    """Kullanc ifresini deitir - Admin Users sayfas"""

    from werkzeug.security import generate_password_hash

    

    # Admin kontrol

    if session.get('role') != 'admin':

        return jsonify({'error': 'Admin eriimi gerekli'}), 403

    

    try:

        data = request.get_json()

        if not data:

            return jsonify({'error': 'Geersiz JSON verisi'}), 400

        

        new_password = str(data.get('new_password', '')).strip()

        

        if not new_password:

            return jsonify({'error': 'Yeni ifre gerekli'}), 400

        

        if len(new_password) < 4:

            return jsonify({'error': 'ifre en az 4 karakter olmal'}), 400

        

        conn = get_db()

        cursor = conn.cursor()

        

        # Kullanc var m kontrol et

        execute_query(cursor, 'SELECT username FROM envanter_users WHERE id = ', (user_id,))

        user = cursor.fetchone()

        if not user:

            close_db(conn)

            return jsonify({'error': 'Kullanc bulunamad'}), 404

        

        # ifreyi hashle ve gncelle

        hashed_password = generate_password_hash(new_password)

        execute_query(cursor, '''

            UPDATE envanter_users 

            SET password_hash = , last_password_change = 

            WHERE id = ?

        ''', (hashed_password, datetime.now(), user_id))

        

        conn.commit()

        close_db(conn)

        

        return jsonify({'success': True, 'message': 'ifre deitirildi'})

        

    except Exception as e:

        logging.error(f"Error in admin_change_password: {e}")

        return jsonify({'error': str(e)}), 500



@app.route('/unlock_qrcodes', methods=['POST'])

def unlock_qrcodes():

    """QR klasr kilidini a (admin paneli ifresi ile)"""

    password = request.form.get('password', '').strip()

    

    # Admin paneli ifresi ile kontrol et

    if password == ADMIN_COUNT_PASSWORD:

        session['qrcodes_unlocked'] = True

        # Kullancy qrcodes klasrne ynlendir

        return '''

        <html>

        <head>

            <title>Eriim zni Verildi</title>

            <meta http-equiv="refresh" content="2;url=/static/qrcodes/">

            <style>

                body { font-family: Arial; text-align: center; padding: 50px; background: #f5f5f5; }

                .success { background: white; padding: 30px; border-radius: 10px; max-width: 400px; margin: auto; box-shadow: 0 2px 10px rgba(0,0,0,0.1); }

                h2 { color: #4caf50; }

            </style>

        </head>

        <body>

            <div class="success">

                <h2> Eriim zni Verildi</h2>

                <p>QR klasrne ynlendiriliyorsunuz...</p>

            </div>

        </body>

        </html>

        '''

    else:

        return '''

        <html>

        <head>

            <title>Hatal ifre</title>

            <style>

                body { font-family: Arial; text-align: center; padding: 50px; background: #f5f5f5; }

                .error { background: white; padding: 30px; border-radius: 10px; max-width: 400px; margin: auto; box-shadow: 0 2px 10px rgba(0,0,0,0.1); }

                h2 { color: #d32f2f; }

                a { color: #1976d2; text-decoration: none; }

            </style>

        </head>

        <body>

            <div class="error">

                <h2> Hatal ifre</h2>

                <p>Admin paneli ifresini doru girdiinizden emin olun.</p>

                <p><a href="/static/qrcodes/">Tekrar Dene</a></p>

            </div>

        </body>

        </html>

        ''', 403



@app.route('/admin/reset_active_sessions', methods=['POST'])

@login_required

@admin_required_decorator

def reset_active_sessions():

    """Tm aktif saym oturumlarn sfrla"""

    try:

        conn = get_db()

        cursor = conn.cursor()

        

        # Tm aktif saymlar pasif yap

        execute_query(cursor, 'UPDATE count_sessions SET is_active =  WHERE is_active = ', (False, True))

        

        conn.commit()

        affected_rows = cursor.rowcount

        close_db(conn)

        

        return jsonify({

            'success': True,

            'message': f' {affected_rows} aktif saym oturumu sfrland'

        })

        

    except Exception as e:

        logging.exception(f"Reset active sessions error: {e}")

        return jsonify({

            'success': False,

            'message': f' Hata: {str(e)}'

        }), 500



@app.route('/metrics')

def metrics():

    """Sistem metrikleri"""

    try:

        conn = get_db()

        cursor = conn.cursor()



        # statistikler

        db.session.query(QRCode).count())

        total_qr = cursor.fetchone()[0]



        execute_query(cursor, 'SELECT COUNT(*) FROM qr_codes WHERE is_used = 1')

        used_qr = cursor.fetchone()[0]



        db.session.query(User).count())

        total_users = cursor.fetchone()[0]



        execute_query(cursor, "SELECT COUNT(*) FROM count_sessions WHERE is_active = TRUE")

        active_sessions = cursor.fetchone()[0]



        close_db(conn)



        return jsonify({

            'qr_codes': {

                'total': total_qr,

                'used': used_qr,

                'remaining': total_qr - used_qr

            },

            'users': total_users,

            'active_sessions': active_sessions,

            'timestamp': datetime.now().isoformat()

        })

    except Exception as e:

        return jsonify({'error': str(e)}), 500



# Render.com deployment check

def is_render_deployment():

    """Render.com deploy kontrol"""

    return os.environ.get('RENDER') is not None



def get_port():

    """Port numarasn al"""

    return int(os.environ.get('PORT', 5001))



#  LOKAL SSTEM - QR Admin modl kaldrld



# ==================== CLOUDFLARE TUNNEL YNETM ====================

import subprocess

import threading

import re



# Global deikenler

tunnel_process = None

tunnel_url = None

tunnel_running = False



@app.route('/tunnel/start', methods=['POST'])

@login_required

def start_tunnel():

    """Cloudflare Tunnel balat"""

    global tunnel_process, tunnel_url, tunnel_running

    

    try:

        if tunnel_running and tunnel_process:

            return jsonify({

                'success': True,

                'message': 'Tunnel zaten alyor',

                'url': tunnel_url

            })

        

        # Cloudflared.exe yolu

        cloudflared_path = os.path.join(os.path.dirname(__file__), 'cloudflared.exe')

        

        if not os.path.exists(cloudflared_path):

            return jsonify({

                'success': False,

                'message': f'cloudflared.exe bulunamad: {cloudflared_path}'

            }), 404

        

        # Tunnel balat

        tunnel_process = subprocess.Popen(

            [cloudflared_path, 'tunnel', '--url', 'http://localhost:5002'],

            stdout=subprocess.PIPE,

            stderr=subprocess.PIPE,

            text=True,

            bufsize=1,

            creationflags=subprocess.CREATE_NO_WINDOW if os.name == 'nt' else 0

        )

        

        # URL'yi yakalamak iin arka planda thread balat

        def capture_url():

            global tunnel_url, tunnel_running

            try:

                # Hem stdout hem stderr'i kontrol et

                while True:

                    # stdout'u kontrol et

                    line = tunnel_process.stderr.readline()

                    if not line:

                        break

                    

                    print(f"[Cloudflare] {line.strip()}")

                    

                    # trycloudflare.com URL'sini yakala

                    if 'trycloudflare.com' in line:

                        match = re.search(r'https://[\w-]+\.trycloudflare\.com', line)

                        if match:

                            tunnel_url = match.group(0)

                            tunnel_running = True

                            print(f" Cloudflare Tunnel balatld: {tunnel_url}")

                            break

            except Exception as e:

                print(f" URL yakalama hatas: {e}")

        

        threading.Thread(target=capture_url, daemon=True).start()

        

        # URL'nin yakalanmas iin biraz daha bekle

        import time

        max_wait = 10  # maksimum 10 saniye bekle

        waited = 0

        while not tunnel_url and waited < max_wait:

            time.sleep(0.5)

            waited += 0.5

        

        if tunnel_url:

            return jsonify({

                'success': True,

                'message': 'Tunnel baaryla balatld',

                'url': tunnel_url

            })

        else:

            # Yine de baarl say, arka planda yakalanacak

            tunnel_running = True

            return jsonify({

                'success': True,

                'message': 'Tunnel balatld, URL yaknda hazr olacak',

                'url': None

            })

            

    except Exception as e:

        logging.exception("Tunnel balatma hatas")

        tunnel_running = False

        tunnel_url = None

        return jsonify({

            'success': False,

            'message': f'Hata: {str(e)}'

        }), 500



@app.route('/tunnel/stop', methods=['POST'])

@login_required

def stop_tunnel():

    """Cloudflare Tunnel durdur"""

    global tunnel_process, tunnel_url, tunnel_running

    

    try:

        if not tunnel_running or not tunnel_process:

            return jsonify({

                'success': False,

                'message': 'Tunnel zaten durdurulmu'

            })

        

        # Process'i durdur

        tunnel_process.terminate()

        tunnel_process.wait(timeout=5)

        

        tunnel_process = None

        tunnel_url = None

        tunnel_running = False

        

        print(" Cloudflare Tunnel durduruldu")

        

        return jsonify({

            'success': True,

            'message': 'Tunnel baaryla durduruldu'

        })

        

    except Exception as e:

        logging.exception("Tunnel durdurma hatas")

        return jsonify({

            'success': False,

            'message': f'Hata: {str(e)}'

        }), 500



@app.route('/tunnel/status')

@login_required

def tunnel_status():

    """Tunnel durumunu kontrol et"""

    global tunnel_running, tunnel_url, tunnel_process

    

    # Process hala alyor mu kontrol et

    process_alive = tunnel_process is not None and tunnel_process.poll() is None

    

    return jsonify({

        'running': tunnel_running,

        'url': tunnel_url,

        'process_alive': process_alive,

        'debug': {

            'tunnel_running_flag': tunnel_running,

            'tunnel_url': tunnel_url,

            'process_exists': tunnel_process is not None,

            'process_alive': process_alive

        }

    })



@app.route('/tunnel/logs')

@login_required

def tunnel_logs():

    """Tunnel loglarn gster (debug iin)"""

    global tunnel_process

    

    if not tunnel_process:

        return jsonify({'error': 'Tunnel almyor'}), 404

    

    # Son birka satr oku

    try:

        # stderr'den oku (cloudflared loglar oraya yazyor)

        logs = []

        return jsonify({'logs': logs, 'message': 'Log okuma aktif deil'})

    except Exception as e:

        return jsonify({'error': str(e)}), 500



# ============================================================================

#  DATABASE BACKUP SSTEM

# ============================================================================



def backup_database():

    """SQLite veritabanini yedekle - Veri kaybi korumas"""

    try:

        db_path = 'instance/envanter_local.db'

        if not os.path.exists(db_path):

            app.logger.warning(f'[ERROR] Database dosyas bulunamad: {db_path}')

            return None

            

        backup_dir = 'backups'

        os.makedirs(backup_dir, exist_ok=True)

        

        # Timestamp ile backup ad (Yl-Ay-Gn_Saat-Dakika-Saniye)

        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')

        backup_path = f'{backup_dir}/envanter_backup_{timestamp}.db'

        

        # Orijinal dosya boyutu

        original_size = os.path.getsize(db_path)

        

        # Kopyala

        import shutil

        shutil.copy2(db_path, backup_path)

        

        # Backup boyutu kontrol et

        backup_size = os.path.getsize(backup_path)

        

        # Btnlk kontrol

        if backup_size != original_size:

            app.logger.error(f'[ERROR] Backup btnl baarsz! Orijinal: {original_size}, Backup: {backup_size}')

            os.remove(backup_path)

            return None

        

        # Baarl

        size_mb = round(original_size / (1024**2), 2)

        app.logger.info(f'[OK] Database backup oluturuldu:')

        app.logger.info(f'   Konum: {os.path.abspath(backup_path)}')

        app.logger.info(f'   Boyut: {size_mb} MB')

        app.logger.info(f'   Zaman: {timestamp}')

        

        # Eski backup'lar temizle (son 30'u tut)

        cleanup_old_backups(backup_dir, keep=30)

        

        return backup_path

    except Exception as e:

        app.logger.error(f'[ERROR] Backup hatas: {e}')

        import traceback

        app.logger.error(traceback.format_exc())

        return None



def verify_backup_integrity():

    """Backup dosyalarnn btnln kontrol et"""

    try:

        backup_dir = 'backups'

        if not os.path.exists(backup_dir):

            return

        

        backups = [

            os.path.join(backup_dir, f) 

            for f in os.listdir(backup_dir) 

            if f.startswith('envanter_backup_') and f.endswith('.db')

        ]

        

        if not backups:

            return

        

        # En son backup' kontrol et

        latest_backup = max(backups, key=os.path.getmtime)

        

        try:

            # SQLite database'i a ve kontrol et

            import sqlite3

            conn = sqlite3.connect(latest_backup)

            cursor = conn.cursor()

            cursor.execute("SELECT name FROM sqlite_master WHERE type='table';")

            tables = cursor.fetchall()

            conn.close()

            

            if tables:

                app.logger.info(f'[OK] Backup btnl OK: {os.path.basename(latest_backup)} ({len(tables)} tablo)')

            else:

                app.logger.warning(f'[WARNING] Backup bo grnyor: {os.path.basename(latest_backup)}')

        except Exception as e:

            app.logger.warning(f'[WARNING] Backup kontrol hatas: {os.path.basename(latest_backup)} - {e}')

            

    except Exception as e:

        app.logger.error(f'[ERROR] Backup integrity check hatas: {e}')



def cleanup_old_backups(backup_dir, keep=30):

    """Eski backup'lar temizle - Disk alan tasarrufu"""

    try:

        if not os.path.exists(backup_dir):

            return

            

        backups = sorted([

            os.path.join(backup_dir, f) 

            for f in os.listdir(backup_dir) 

            if f.startswith('envanter_backup_') and f.endswith('.db')

        ], key=os.path.getmtime, reverse=True)

        

        deleted_count = 0

        for old_backup in backups[keep:]:

            try:

                size_mb = round(os.path.getsize(old_backup) / (1024**2), 2)

                os.remove(old_backup)

                deleted_count += 1

                app.logger.info(f'[OK] Eski backup silindi: {os.path.basename(old_backup)} ({size_mb} MB)')

            except Exception as e:

                app.logger.error(f'[ERROR] Backup silme hatas: {old_backup} - {e}')

        

        if deleted_count > 0:

            app.logger.info(f'[OK] {deleted_count} eski backup silindi (Son {keep} tutuldu)')

            

    except Exception as e:

        app.logger.error(f'[ERROR] Backup temizleme hatas: {e}')



def get_backup_list():

    """Mevcut backup'lar listele"""

    try:

        backup_dir = 'backups'

        if not os.path.exists(backup_dir):

            return []

            

        backups = []

        for f in os.listdir(backup_dir):

            if f.startswith('envanter_backup_') and f.endswith('.db'):

                full_path = os.path.join(backup_dir, f)

                backups.append({

                    'filename': f,

                    'path': full_path,

                    'size_mb': round(os.path.getsize(full_path) / (1024**2), 2),

                    'created_at': datetime.fromtimestamp(os.path.getmtime(full_path)).isoformat()

                })

        

        # En yeniden eskiye srala

        backups.sort(key=lambda x: x['created_at'], reverse=True)

        return backups

    except Exception as e:

        app.logger.error(f'[ERROR] Backup listesi hatas: {e}')

        return []



def restore_database(backup_filename):

    """Backup'tan veritabann geri ykle"""

    try:

        backup_path = f'backups/{backup_filename}'

        db_path = 'instance/envanter_local.db'

        

        # Dosya varl kontrol et

        if not os.path.exists(backup_path):

            app.logger.error(f'[ERROR] Backup dosyas bulunamad: {backup_path}')

            return False, 'Backup dosyas bulunamad'

        

        # Mevcut db'nin yedeini al

        import shutil

        if os.path.exists(db_path):

            recovery_backup = f'backups/emergency_recovery_{datetime.now().strftime("%Y%m%d_%H%M%S")}.db'

            shutil.copy2(db_path, recovery_backup)

            app.logger.info(f'[RECOVERY] Kurtarma backup oluturuldu: {recovery_backup}')

        

        # Backup' geri ykle

        shutil.copy2(backup_path, db_path)

        

        app.logger.warning(f'[RESTORE] Database geri yklendi:')

        app.logger.warning(f'   [SRC] Kaynak: {backup_path}')

        app.logger.warning(f'   [DST] Hedef: {db_path}')

        

        return True, f'Database baaryla geri yklendi: {backup_filename}'

        

    except Exception as e:

        app.logger.error(f'[ERROR] Restore hatas: {e}')

        import traceback

        app.logger.error(traceback.format_exc())

        return False, f'Restore hatas: {str(e)}'



@app.route('/admin/backup_now', methods=['POST'])

@login_required

def backup_now():

    """Manuel backup tetikle - sadece admin"""

    # Admin kontrol

    if session.get('role') != 'admin':

        return jsonify({'error': 'Admin eriimi gerekli'}), 403

    

    try:

        app.logger.info('[ADMIN] Admin manuel backup balatt')

        backup_path = backup_database()

        if backup_path:

            return jsonify({

                'success': True, 

                'backup_path': backup_path,

                'message': '[OK] Backup baaryla oluturuldu'

            })

        else:

            return jsonify({

                'success': False, 

                'error': '[ERROR] Backup oluturulamad'

            }), 500

    except Exception as e:

        app.logger.error(f'[ERROR] Backup endpoint hatas: {e}')

        return jsonify({'success': False, 'error': str(e)}), 500



@app.route('/admin/backups', methods=['GET'])

@login_required

def list_backups():

    """Backup listesini getir - sadece admin"""

    # Admin kontrol

    if session.get('role') != 'admin':

        return jsonify({'error': 'Admin eriimi gerekli'}), 403

    

    try:

        backups = get_backup_list()

        backup_dir = os.path.abspath('backups')

        

        return jsonify({

            'success': True,

            'backups': backups,

            'count': len(backups),

            'backup_dir': backup_dir,

            'total_size_mb': sum(b['size_mb'] for b in backups)

        })

    except Exception as e:

        app.logger.error(f'[ERROR] Backup listesi endpoint hatas: {e}')

        return jsonify({'success': False, 'error': str(e)}), 500



@app.route('/admin/restore_backup/<filename>', methods=['POST'])

@login_required

def restore_backup(filename):

    """Backup'tan geri ykle - sadece admin - OK DKKATL!"""

    # Admin kontrol

    if session.get('role') != 'admin':

        return jsonify({'error': 'Admin eriimi gerekli'}), 403

    

    try:

        # Gvenlik: Yalnzca backup dosyalar

        if not filename.startswith('envanter_backup_') or not filename.endswith('.db'):

            return jsonify({'success': False, 'error': 'Geersiz backup dosyas'}), 400

        

        success, message = restore_database(filename)

        

        if success:

            app.logger.warning(f'[RESTORE-ADMIN] Admin database restore yapt: {filename}')

            return jsonify({

                'success': True,

                'message': f'[OK] {message}'

            })

        else:

            return jsonify({

                'success': False,

                'error': f'[ERROR] {message}'

            }), 400

            

    except Exception as e:

        app.logger.error(f'[ERROR] Restore endpoint hatas: {e}')

        return jsonify({'success': False, 'error': str(e)}), 500



@app.route('/admin/backup_status', methods=['GET'])

@login_required

def backup_status():

    """Backup sistem durumu - sadece admin"""

    if session.get('role') != 'admin':

        return jsonify({'error': 'Admin eriimi gerekli'}), 403

    

    try:

        backups = get_backup_list()

        db_path = 'instance/envanter_local.db'

        db_size = 0

        

        if os.path.exists(db_path):

            db_size = round(os.path.getsize(db_path) / (1024**2), 2)

        

        last_backup = None

        if backups:

            last_backup = backups[0]

        

        return jsonify({

            'success': True,

            'current_db_size_mb': db_size,

            'total_backups': len(backups),

            'total_backup_size_mb': sum(b['size_mb'] for b in backups),

            'last_backup': last_backup,

            'backup_dir': os.path.abspath('backups'),

            'schedule': {

                'daily_backup': 'Her gn 02:00\'de',

                'hourly_check': 'Her saat banda'

            }

        })

    except Exception as e:

        app.logger.error(f'[ERROR] Backup status endpoint hatas: {e}')

        return jsonify({'success': False, 'error': str(e)}), 500

        return jsonify({'error': 'Admin eriimi gerekli'}), 403

    

    try:

        # Gvenlik: sadece backup klasrndeki dosyalar

        backup_path = os.path.join('backups', secure_filename(filename))

        

        if not os.path.exists(backup_path):

            return jsonify({'success': False, 'error': 'Backup dosyas bulunamad'}), 404

        

        if not filename.startswith('envanter_backup_'):

            return jsonify({'success': False, 'error': 'Geersiz backup dosyas'}), 400

        

        # Mevcut DB'yi yedekle (geri dn iin)

        current_backup = backup_database()

        

        # Backup' geri ykle

        db_path = 'instance/envanter_local.db'

        import shutil

        shutil.copy2(backup_path, db_path)

        

        app.logger.warning(f'[WARNING] DATABASE RESTORE: {filename} by {session.get("username")}')

        

        return jsonify({

            'success': True,

            'message': 'Backup geri yklendi. Ltfen uygulamay yeniden balatn.',

            'restored_from': filename,

            'safety_backup': current_backup

        })

    except Exception as e:

        app.logger.error(f'Restore hatas: {e}')

        return jsonify({'success': False, 'error': str(e)}), 500



# ========================================

#  PAKET YNETM API ENDPOINT'LER

# ========================================



@app.route('/api/create_package', methods=['POST'])

def create_package():

    """Yeni paket oluturma - Tek QR ile birden fazla para"""

    try:

        data = request.get_json()

        package_name = data.get('package_name', '').strip()

        package_desc = data.get('package_desc', '').strip()

        items = data.get('items', [])

        

        if not package_name:

            return jsonify({'success': False, 'error': 'Paket ad gerekli'}), 400

        

        if not items or len(items) == 0:

            return jsonify({'success': False, 'error': 'En az 1 para ekleyin'}), 400

        

        # Paket kodunun zaten var olup olmadn kontrol et

        conn = get_db()

        cursor = conn.cursor()

        

        cursor.execute('SELECT part_code FROM part_codes WHERE part_code = ', (package_name,))

        if cursor.fetchone():

            conn.close()

            return jsonify({'success': False, 'error': 'Bu paket ad/QR kodu zaten kullanmda'}), 400

        

        # JSON formatnda paket ieriini hazrla

        package_items_json = json.dumps(items)

        

        # QR kod resmi olutur (Base64)

        qr = qrcode.QRCode(version=1, box_size=10, border=5)

        qr.add_data(package_name)

        qr.make(fit=True)

        

        img = qr.make_image(fill_color="black", back_color="white")

        buffer = BytesIO()

        img.save(buffer, format='PNG')

        qr_base64 = base64.b64encode(buffer.getvalue()).decode()

        

        # Paketi veritabanna ekle

        cursor.execute('''

            INSERT INTO part_codes (part_code, part_name, is_package, package_items)

            VALUES (?, ?, ?, ?)

        ''', (package_name, package_desc if package_desc else f'Paket: {package_name}', True, package_items_json))

        

        package_id = cursor.lastrowid

        

        # QR codes tablosuna da ekle (tek bir QR)

        cursor.execute('''

            INSERT INTO qr_codes (qr_id, part_code_id, is_used, created_at)

            VALUES (?, ?, ?, ?)

        ''', (package_name, package_id, False, datetime.now()))

        

        conn.commit()

        conn.close()

        

        app.logger.info(f'[PAKET] Yeni paket oluturuldu: {package_name} ({len(items)} para)')

        

        return jsonify({

            'success': True,

            'message': f'Paket "{package_name}" oluturuldu',

            'package_code': package_name,

            'qr_image': qr_base64,  # Base64 QR kodu

            'items_count': len(items),

            'total_quantity': sum(item.get('quantity', 1) for item in items)

        })

        

    except Exception as e:

        app.logger.error(f'[HATA] Paket oluturma hatas: {e}')

        return jsonify({'success': False, 'error': str(e)}), 500



@app.route('/api/get_packages', methods=['GET'])

def get_packages():

    """Tm paketleri listele"""

    try:

        conn = get_db()

        cursor = conn.cursor()

        

        cursor.execute('''

            SELECT part_code, part_name, package_items

            FROM part_codes

            WHERE is_package = 1 OR is_package = 'true'

            ORDER BY part_code

        ''')

        

        packages = []

        for row in cursor.fetchall():

            try:

                items = json.loads(row[2]) if row[2] else []

            except:

                items = []

            

            # Her paket iin QR kod olutur

            qr = qrcode.QRCode(version=1, box_size=10, border=5)

            qr.add_data(row[0])  # part_code

            qr.make(fit=True)

            

            img = qr.make_image(fill_color="black", back_color="white")

            buffer = BytesIO()

            img.save(buffer, format='PNG')

            qr_base64 = base64.b64encode(buffer.getvalue()).decode()

            

            packages.append({

                'part_code': row[0],

                'part_name': row[1],

                'items': items,

                'qr_image': qr_base64  # QR kod base64

            })

        

        conn.close()

        

        return jsonify({

            'success': True,

            'packages': packages,

            'count': len(packages)

        })

        

    except Exception as e:

        app.logger.error(f'[HATA] Paket listeleme hatas: {e}')

        return jsonify({'success': False, 'error': str(e)}), 500



@app.route('/api/delete_package', methods=['POST'])

def delete_package():

    """Paket silme"""

    try:

        data = request.get_json()

        part_code = data.get('part_code', '').strip()

        

        if not part_code:

            return jsonify({'success': False, 'error': 'Paket kodu gerekli'}), 400

        

        conn = get_db()

        cursor = conn.cursor()

        

        # Paketin varln kontrol et

        cursor.execute('SELECT is_package FROM part_codes WHERE part_code = ', (part_code,))

        row = cursor.fetchone()

        

        if not row:

            return jsonify({'success': False, 'error': 'Paket bulunamad'}), 404

        

        if not row[0]:

            return jsonify({'success': False, 'error': 'Bu bir paket deil'}), 400

        

        # Paketi sil

        cursor.execute('DELETE FROM part_codes WHERE part_code = ', (part_code,))

        conn.commit()

        conn.close()

        

        app.logger.info(f'[PAKET] Paket silindi: {part_code}')

        

        return jsonify({

            'success': True,

            'message': f'Paket "{part_code}" silindi'

        })

        

    except Exception as e:

        app.logger.error(f'[HATA] Paket silme hatas: {e}')

        return jsonify({'success': False, 'error': str(e)}), 500



#  CANLI DASHBOARD API - Gerek zamanl istatistikler

@app.route('/api/live_dashboard/<int:session_id>')

@login_required

def live_dashboard_stats(session_id):

    """

    Canl dashboard iin tm istatistikler

    

    Returns:

        - Tarama hz (son 5dk, son 1 saat)

        - En ok taranan 10 para

        - Kullanc performans

        - Saym biti tahmini

        - Saatlik tarama dalm

    """

    # Session'dan kullanc adn al

    username = session.get('username', 'Bilinmeyen')

    logging.info(f"Dashboard API called - Session: {session_id}, User: {username}")

    

    try:

        with db_connection() as conn:

            cursor = conn.cursor()

            

            # 1. Tarama Hz (Scans per Minute)

            execute_query(cursor, '''

                SELECT 

                    COUNT(*) as total_scans,

                    MIN(scanned_at) as first_scan,

                    MAX(scanned_at) as last_scan

                FROM scanned_qr

                WHERE session_id = 

            ''', (session_id,))

            scan_stats = cursor.fetchone()

            

            total_scans = scan_stats[0] if scan_stats else 0

            first_scan = scan_stats[1] if scan_stats else None

            last_scan = scan_stats[2] if scan_stats else None

            

            logging.debug(f"Total scans: {total_scans}, First: {first_scan}, Last: {last_scan}")

            

            # Tarama hz hesapla

            scan_rate_5min = 0

            scan_rate_1hour = 0

            

            if last_scan:

                # Son 5 dakika

                execute_query(cursor, '''

                    SELECT COUNT(*) FROM scanned_qr

                    WHERE session_id =  AND scanned_at >= datetime('now', '-5 minutes')

                ''', (session_id,))

                scans_5min = cursor.fetchone()[0]

                scan_rate_5min = round(scans_5min / 5, 1)  # tarama/dakika

                

                # Son 1 saat

                execute_query(cursor, '''

                    SELECT COUNT(*) FROM scanned_qr

                    WHERE session_id =  AND scanned_at >= datetime('now', '-1 hour')

                ''', (session_id,))

                scans_1hour = cursor.fetchone()[0]

                scan_rate_1hour = round(scans_1hour / 60, 1)  # tarama/dakika

            

            # 2. En ok Taranan 10 Para

            execute_query(cursor, '''

                SELECT 

                    pc.part_name,

                    pc.part_code,

                    COUNT(*) as scan_count

                FROM scanned_qr sq

                JOIN part_codes pc ON sq.part_code = pc.part_code

                WHERE sq.session_id = 

                GROUP BY pc.part_code, pc.part_name

                ORDER BY scan_count DESC

                LIMIT 10

            ''', (session_id,))

            top_parts = [{'name': row[0], 'code': row[1], 'count': row[2]} for row in cursor.fetchall()]

            

            # 3. Kullanc Performans

            execute_query(cursor, '''

                SELECT 

                    u.full_name,

                    u.username,

                    COUNT(*) as scan_count,

                    MIN(sq.scanned_at) as first_scan,

                    MAX(sq.scanned_at) as last_scan

                FROM scanned_qr sq

                JOIN envanter_users u ON sq.scanned_by = u.id

                WHERE sq.session_id = 

                GROUP BY u.id, u.full_name, u.username

                ORDER BY scan_count DESC

            ''', (session_id,))

            

            user_performance = []

            for row in cursor.fetchall():

                user_first = datetime.fromisoformat(row[3]) if row[3] else None

                user_last = datetime.fromisoformat(row[4]) if row[4] else None

                work_time_minutes = 0

                personal_rate = 0

                

                if user_first and user_last:

                    work_time_seconds = (user_last - user_first).total_seconds()

                    work_time_minutes = round(work_time_seconds / 60, 1)

                    if work_time_minutes > 0:

                        personal_rate = round(row[2] / work_time_minutes, 1)

                

                user_performance.append({

                    'name': row[0],

                    'username': row[1],

                    'scan_count': row[2],

                    'work_time_minutes': work_time_minutes,

                    'scans_per_minute': personal_rate

                })

            

            # 4. Saym Biti Tahmini

            execute_query(cursor, '''

                SELECT COUNT(DISTINCT qr_id) FROM qr_codes

            ''')

            total_qr_codes = cursor.fetchone()[0]

            remaining = total_qr_codes - total_scans

            

            eta_minutes = 0

            eta_text = "Hesaplanyor..."

            

            if scan_rate_1hour > 0 and remaining > 0:

                eta_minutes = round(remaining / scan_rate_1hour)

                if eta_minutes < 60:

                    eta_text = f"{eta_minutes} dakika"

                else:

                    eta_hours = round(eta_minutes / 60, 1)

                    eta_text = f"{eta_hours} saat"

            

            # 5. Saatlik Tarama Dalm (Son 24 saat)

            execute_query(cursor, '''

                SELECT 

                    strftime('%H', scanned_at) as hour,

                    COUNT(*) as count

                FROM scanned_qr

                WHERE session_id =  AND scanned_at >= datetime('now', '-24 hours')

                GROUP BY hour

                ORDER BY hour

            ''', (session_id,))

            

            hourly_distribution = {}

            for row in cursor.fetchall():

                hourly_distribution[row[0]] = row[1]

            

            result = {

                'success': True,

                'session_id': session_id,

                'total_scans': total_scans,

                'remaining': remaining,

                'progress_percent': round((total_scans / total_qr_codes * 100) if total_qr_codes > 0 else 0, 1),

                'scan_rate': {

                    'last_5min': scan_rate_5min,

                    'last_1hour': scan_rate_1hour,

                    'unit': 'tarama/dakika'

                },

                'top_parts': top_parts,

                'user_performance': user_performance,

                'eta': {

                    'minutes': eta_minutes,

                    'text': eta_text

                },

                'hourly_distribution': hourly_distribution,

                'first_scan': first_scan,

                'last_scan': last_scan

            }

            

            logging.info(f"Dashboard API success - {total_scans} scans, {len(user_performance)} users")

            return jsonify(result)

            

    except Exception as e:

        logging.error(f"Live dashboard error: {e}", exc_info=True)

        return jsonify({'success': False, 'error': str(e)}), 500



# Otomatik backup scheduler iin (APScheduler kullanarak)

try:

    from apscheduler.schedulers.background import BackgroundScheduler

    

    backup_scheduler = BackgroundScheduler()

    

    #  GNLK BACKUP: Her gn 02:00'de yaplr

    backup_scheduler.add_job(

        func=backup_database,

        trigger="cron",

        hour=2,

        minute=0,

        id='daily_auto_backup',

        name='Gnlk Otomatik Database Backup',

        replace_existing=True

    )

    

    #  SAATLK KONTROL: Her saat banda backup kontrol

    backup_scheduler.add_job(

        func=verify_backup_integrity,

        trigger="cron",

        minute=0,

        id='hourly_backup_check',

        name='Saatlik Backup Btnl Kontrol',

        replace_existing=True

    )

    

    backup_scheduler.start()

    app.logger.info('[OK] Backup Scheduler Balatld:')

    app.logger.info('   [DAILY] Gnlk Backup: Her gn 02:00\'de')

    app.logger.info('   [HOURLY] Saatlik Kontrol: Her saat banda')

    app.logger.info(f'   [PATH] Backup Klasr: {os.path.abspath("backups")}')

except ImportError:

    app.logger.warning('[!] APScheduler ykl deil. Otomatik backup devre d.')

    app.logger.warning('   Yklemek iin: pip install apscheduler')

except Exception as e:

    app.logger.error(f'Backup scheduler hatas: {e}')



# NOTE: This is handled by render_startup_alt.py for Render.com

# DO NOT call socketio.run() here to avoid port binding conflicts

# The render_startup_alt.py script will import this app and call socketio.run()



if __name__ == '__main__':

    print("  Direct execution detected. Please use render_startup_alt.py")

    print("    Or set RENDER=false for local development with Flask")



    # Initialize database on startup

    try:

        init_db()

    except Exception as e:

        print(f" Failed to initialize database: {e}")



    # For local testing only - DO NOT use in production

    if not os.environ.get('RENDER'):

        port = 5002

        print(" Starting EnvanterQR System v2.0 (LOCAL)...")

        print(" Dashboard: http://localhost:5002")

        print(" Admin Panel: http://localhost:5002/admin")

        print(" Health Check: http://localhost:5002/health")

        print(" Metrics: http://localhost:5002/metrics")

        print(" Storage: Backblaze B2 Enabled")

        print(" Security: Headers + Rate Limiting Active")

        print()

        

        # Network eriimi iin host ayar

        # '0.0.0.0' = Tm a arayzlerinden eriime izin ver (WiFi, Ethernet, vb.)

        # '127.0.0.1' = Sadece localhost (kendi PC'nden)

        socketio.run(app, host='0.0.0.0', port=port, debug=True, allow_unsafe_werkzeug=True)

    else:

        print("ERROR: This should not run with RENDER=true")

        print("Use: gunicorn wsgi:app --worker-class eventlet")

        print("Or:  python render_startup_alt.py")

