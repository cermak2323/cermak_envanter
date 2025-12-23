#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
PARÇA SİPARİŞ SİSTEMİ MODÜLÜ
Envanter sisteminden tamamen bağımsız çalışan sipariş yönetimi
"""

from flask import Blueprint, render_template, request, jsonify, session, redirect, url_for
from functools import wraps
import pymysql
from openpyxl import load_workbook
import io
from datetime import datetime
import logging

# Blueprint oluştur
order_bp = Blueprint('order_system', __name__, url_prefix='/order_system')

# Database helper fonksiyonları (app.py'den import edilecek)
DB_CONFIG = {
    'host': '192.168.0.57',
    'port': 3306,
    'user': 'flaskuser',
    'password': 'FlaskSifre123!',
    'database': 'flaskdb',
    'charset': 'utf8mb4'
}

def get_order_db():
    """Sipariş sistemi için veritabanı bağlantısı"""
    return pymysql.connect(**DB_CONFIG, cursorclass=pymysql.cursors.DictCursor)

def execute_order_query(cursor, query, params=None):
    """Güvenli SQL sorgusu"""
    cursor.execute(query, params or ())
    return cursor

def login_required_order(f):
    """Sipariş sistemi için login kontrolü"""
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'username' not in session:
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return decorated_function


# ============================================================================
# 1. ANA MENÜ - SİPARİŞ SİSTEMİ ANA SAYFASI
# ============================================================================

@order_bp.route('/')
@login_required_order
def index():
    """Sipariş sistemi ana sayfası - modül menüsü"""
    return render_template('order_system/index.html', username=session.get('username'))


# ============================================================================
# 2. EXCEL STOK GÜNCELLEME MODÜLÜ
# ============================================================================

@order_bp.route('/upload_stock')
@login_required_order
def upload_stock_page():
    """Excel stok yükleme sayfası"""
    return render_template('order_system/upload_stock.html', username=session.get('username'))


@order_bp.route('/api/upload_stock', methods=['POST'])
@login_required_order
def upload_stock_api():
    """
    Excel ile stok güncelleme ve sipariş kontrolü
    Beklenen Excel başlıkları (ilk satır):
    | Parça Kodu | Parça Adı | Stok | Kritik stok | Beklenen stok | Tedarikçi | Geliş (Euro) | BUILD OUT | Değişen Parça Kodu |
    
    Mantık:
    - BUILD OUT varsa: Sipariş etmeme, "BUILD OUT" yazısı göster
    - Değişen Parça Kodu varsa: Değişen kodu kullanarak sipariş et
    """
    try:
        if 'file' not in request.files:
            return jsonify({'success': False, 'error': 'Dosya bulunamadı'}), 400
        
        file = request.files['file']
        if file.filename == '':
            return jsonify({'success': False, 'error': 'Dosya seçilmedi'}), 400
        
        if not file.filename.endswith(('.xlsx', '.xls')):
            return jsonify({'success': False, 'error': 'Sadece Excel dosyaları kabul edilir'}), 400
        
        # Excel'i oku
        wb = load_workbook(io.BytesIO(file.read()), data_only=True)
        ws = wb.active
        
        conn = get_order_db()
        cursor = conn.cursor()
        
        # Şema güvencesi: expected_stock, build_out, replacement_code kolonları yoksa ekleyelim
        try:
            for col_name in ['expected_stock', 'build_out', 'replacement_code']:
                cursor.execute("""
                    SELECT COUNT(*) AS cnt FROM information_schema.COLUMNS
                    WHERE TABLE_SCHEMA = %s AND TABLE_NAME = 'order_system_stock' AND COLUMN_NAME = %s
                """, (DB_CONFIG['database'], col_name))
                res = cursor.fetchone()
                if not res or res.get('cnt', 0) == 0:
                    if col_name == 'expected_stock':
                        cursor.execute("ALTER TABLE order_system_stock ADD COLUMN expected_stock INT DEFAULT 0 AFTER critical_stock_level")
                    elif col_name == 'build_out':
                        cursor.execute("ALTER TABLE order_system_stock ADD COLUMN build_out BOOLEAN DEFAULT FALSE AFTER unit_price")
                    elif col_name == 'replacement_code':
                        cursor.execute("ALTER TABLE order_system_stock ADD COLUMN replacement_code VARCHAR(100) DEFAULT NULL AFTER build_out")
                    conn.commit()
        except Exception as e:
            print(f"Schema check error: {e}")
            pass

        updated_parts = []
        build_out_parts = []
        replacement_orders = []
        errors = []
        
        # Excel satırlarını işle (başlık satırını atla)
        for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if not row or not row[0]:  # Boş satır
                continue
            
            try:
                def to_str(v):
                    return str(v).strip() if v is not None else None

                def to_int(v):
                    if v is None or v == '':
                        return None
                    try:
                        return int(float(v))
                    except Exception:
                        return None

                def to_float(v):
                    if v is None or v == '':
                        return None
                    try:
                        return float(v)
                    except Exception:
                        return None

                def to_bool(v):
                    if v is None or v == '':
                        return False
                    if isinstance(v, bool):
                        return v
                    s = str(v).strip().lower()
                    return s in ['evet', 'yes', 'true', '1', '✓', 'v', 'x']

                part_code = to_str(row[0])
                part_name = to_str(row[1])
                stock_quantity = to_int(row[2])
                critical_level = to_int(row[3])
                expected_stock = to_int(row[4])
                supplier = to_str(row[5])
                unit_price = to_float(row[6])
                build_out = to_bool(row[7]) if len(row) > 7 else False
                replacement_code = to_str(row[8]) if len(row) > 8 else None

                if not part_code:
                    raise ValueError('Parça Kodu boş olamaz')

                # Stoku ve diğer alanları güncelle/ekle
                cursor.execute(
                    """
                    INSERT INTO order_system_stock
                        (part_code, part_name, stock_quantity, critical_stock_level, expected_stock, supplier, unit_price, build_out, replacement_code, last_updated)
                    VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, NOW())
                    ON DUPLICATE KEY UPDATE
                        part_name = IF(%s IS NOT NULL, %s, part_name),
                        stock_quantity = IF(%s IS NOT NULL, %s, stock_quantity),
                        critical_stock_level = IF(%s IS NOT NULL, %s, critical_stock_level),
                        expected_stock = IF(%s IS NOT NULL, %s, expected_stock),
                        supplier = IF(%s IS NOT NULL, %s, supplier),
                        unit_price = IF(%s IS NOT NULL, %s, unit_price),
                        build_out = %s,
                        replacement_code = %s,
                        last_updated = NOW()
                    """,
                    (
                        part_code,
                        part_name,
                        stock_quantity,
                        critical_level,
                        expected_stock,
                        supplier,
                        unit_price,
                        build_out,
                        replacement_code,
                        part_name, part_name,
                        stock_quantity, stock_quantity,
                        critical_level, critical_level,
                        expected_stock, expected_stock,
                        supplier, supplier,
                        unit_price, unit_price,
                        build_out,
                        replacement_code
                    ),
                )
                
                # BUILD OUT kontrolü
                if build_out:
                    build_out_parts.append({
                        'part_code': part_code,
                        'part_name': part_name,
                        'status': '🔴 BUILD OUT - SİPARİŞ ETMEYİN'
                    })
                # Değişen parça kodu kontrolü
                elif replacement_code:
                    replacement_orders.append({
                        'original_code': part_code,
                        'part_name': part_name,
                        'replacement_code': replacement_code,
                        'quantity': expected_stock if expected_stock and expected_stock > stock_quantity else 0
                    })
                
                updated_parts.append({
                    'part_code': part_code,
                    'part_name': part_name,
                    'new_stock': stock_quantity if stock_quantity is not None else '-',
                    'build_out': build_out,
                    'replacement_code': replacement_code
                })
                
            except Exception as e:
                errors.append(f"Satır {row_num}: {str(e)}")
        
        conn.commit()
        
        # Yüklenen parçaların tam bilgilerini al (tablo için)
        uploaded_parts_details = []
        if updated_parts:
            part_codes = [p['part_code'] for p in updated_parts]
            placeholders = ','.join(['%s'] * len(part_codes))
            cursor.execute(f"""
                SELECT part_code, part_name, stock_quantity, critical_stock_level, 
                       expected_stock, supplier, unit_price, build_out, replacement_code
                FROM order_system_stock
                WHERE part_code IN ({placeholders})
            """, part_codes)
            uploaded_parts_details = cursor.fetchall()
        
        # Sipariş listesindeki 'Geldi' durumundaki parçaları stoka ekle
        cursor.execute("""
            SELECT part_code, ordered_quantity
            FROM order_list
            WHERE status = 'Geldi'
        """)
        arrived_orders = cursor.fetchall()
        
        for order in arrived_orders:
            cursor.execute("""
                UPDATE order_system_stock
                SET stock_quantity = stock_quantity + %s,
                    last_updated = NOW()
                WHERE part_code = %s
            """, (order['ordered_quantity'], order['part_code']))
        
        # 'Geldi' olan siparişleri listeden sil
        cursor.execute("DELETE FROM order_list WHERE status = 'Geldi'")
        
        # Protected parts tablosunu temizle
        cursor.execute("TRUNCATE TABLE protected_parts")
        
        # Eski 'Gelmedi' siparişlerini temizle
        cursor.execute("DELETE FROM order_list WHERE status = 'Gelmedi' AND order_type = 'Otomatik'")
        
        # ✅ YENI: Değişen parça kodlarıyla sipariş oluştur
        auto_created_orders = 0
        if replacement_orders:
            for repl in replacement_orders:
                # Değişen kod parçasının bilgisini al
                cursor.execute("""
                    SELECT part_name, supplier, unit_price, stock_quantity 
                    FROM order_system_stock
                    WHERE part_code = %s
                """, (repl['replacement_code'],))
                repl_part = cursor.fetchone()
                
                if repl_part:
                    # Sipariş oluştur (değişen kodla)
                    quantity_to_order = max(0, repl['quantity'])
                    if quantity_to_order > 0:
                        cursor.execute("""
                            INSERT INTO order_list
                            (part_code, part_name, ordered_quantity, supplier, unit_price, status, order_type, created_at, notes)
                            VALUES (%s, %s, %s, %s, %s, 'Beklemede', 'Otomatik', NOW(), %s)
                        """, (
                            repl['replacement_code'],
                            repl_part['part_name'],
                            quantity_to_order,
                            repl_part['supplier'],
                            repl_part['unit_price'],
                            f"Değişen kod: {repl['original_code']} → {repl['replacement_code']}"
                        ))
                        auto_created_orders += 1
        
        conn.commit()
        
        cursor.close()
        conn.close()
        
        # Log kaydı
        log_order_action(
            order_id=None,
            part_code='BULK_UPLOAD',
            action='EXCEL_UPLOAD',
            notes=f"{len(updated_parts)} parça güncellendi, {auto_created_orders} otomatik sipariş oluşturuldu",
            action_by=session.get('username')
        )
        
        return jsonify({
            'success': True,
            'updated_count': len(updated_parts),
            'error_count': len(errors),
            'arrived_orders_processed': len(arrived_orders),
            'build_out_count': len(build_out_parts),
            'replacement_orders_created': auto_created_orders,
            'uploaded_parts': uploaded_parts_details,
            'build_out_parts': build_out_parts,
            'replacement_orders': replacement_orders,
            'errors': errors
        })
        
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


# ============================================================================
# 2B. SADECE STOK GÜNCELLEME MODÜLÜ (Basit Excel - 2 Sütun)
# ============================================================================

@order_bp.route('/update_stock_only')
@login_required_order
def update_stock_only_page():
    """Sadece stok güncelleme sayfası - Basit Excel (Parça Kodu | Stok)"""
    return render_template('order_system/update_stock_only.html', username=session.get('username'))


@order_bp.route('/api/update_stock_only', methods=['POST'])
@login_required_order
def update_stock_only_api():
    """
    Sadece stok güncelleme - Basit Excel
    Format: | Parça Kodu | Stok |
    """
    try:
        if 'file' not in request.files:
            return jsonify({'success': False, 'error': 'Dosya bulunamadı'}), 400
        
        file = request.files['file']
        if file.filename == '':
            return jsonify({'success': False, 'error': 'Dosya seçilmedi'}), 400
        
        if not file.filename.endswith(('.xlsx', '.xls')):
            return jsonify({'success': False, 'error': 'Sadece Excel dosyaları kabul edilir'}), 400
        
        # Excel'i oku
        wb = load_workbook(io.BytesIO(file.read()), data_only=True)
        ws = wb.active
        
        conn = get_order_db()
        cursor = conn.cursor()
        
        updated_parts = []
        errors = []
        
        # Excel satırlarını işle (başlık satırını atla)
        for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if not row or not row[0]:  # Boş satır
                continue
            
            try:
                part_code = str(row[0]).strip() if row[0] is not None else None
                
                # Stok miktarını oku
                stock_quantity = None
                if row[1] is not None and row[1] != '':
                    try:
                        stock_quantity = int(float(row[1]))
                    except Exception:
                        raise ValueError('Stok miktarı sayı olmalı')

                if not part_code:
                    raise ValueError('Parça Kodu boş olamaz')
                
                if stock_quantity is None:
                    raise ValueError('Stok miktarı boş olamaz')

                # Sadece stoku güncelle (parça zaten kayıtlıysa)
                cursor.execute("""
                    UPDATE order_system_stock
                    SET stock_quantity = %s,
                        last_updated = NOW()
                    WHERE part_code = %s
                """, (stock_quantity, part_code))
                
                if cursor.rowcount > 0:
                    updated_parts.append({
                        'part_code': part_code,
                        'new_stock': stock_quantity
                    })
                else:
                    errors.append(f"Satır {row_num}: Parça '{part_code}' sistemde bulunamadı")
                
            except Exception as e:
                errors.append(f"Satır {row_num}: {str(e)}")
        
        conn.commit()
        cursor.close()
        conn.close()
        
        # Log
        log_order_action(
            order_id=None,
            part_code='STOCK_ONLY_UPDATE',
            action='STOCK_ONLY_UPLOAD',
            notes=f"{len(updated_parts)} parça stoku güncellendi",
            action_by=session.get('username')
        )
        
        return jsonify({
            'success': True,
            'updated_count': len(updated_parts),
            'error_count': len(errors),
            'errors': errors
        })
        
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


# ============================================================================
# 3. SİPARİŞ OLUŞTURMA MODÜLÜ
# ============================================================================

@order_bp.route('/select_supplier')
@login_required_order
def select_supplier_page():
    """Tedarikçi seçim sayfası"""
    return render_template('order_system/select_supplier.html', username=session.get('username'))


@order_bp.route('/api/get_suppliers', methods=['GET'])
@login_required_order
def get_suppliers():
    """Tüm tedarikçileri ve istatistiklerini getir"""
    try:
        conn = get_order_db()
        cursor = conn.cursor()
        
        # Benzersiz tedarikçiler
        cursor.execute("""
            SELECT DISTINCT supplier FROM order_system_stock 
            WHERE supplier IS NOT NULL AND supplier != ''
            ORDER BY supplier
        """)
        suppliers_list = cursor.fetchall()
        
        suppliers = []
        for row in suppliers_list:
            supplier_name = row['supplier'] if isinstance(row, dict) else row[0]
            
            # Her tedarikçi için istatistikler
            cursor.execute("""
                SELECT 
                    COUNT(*) as part_count,
                    COALESCE(SUM(stock_quantity), 0) as total_stock
                FROM order_system_stock
                WHERE supplier = %s
            """, (supplier_name,))
            stock_stats = cursor.fetchone()
            
            cursor.execute("""
                SELECT 
                    COUNT(*) as order_count,
                    SUM(CASE WHEN status = 'Beklemede' THEN 1 ELSE 0 END) as pending_count
                FROM order_list
                WHERE supplier = %s
            """, (supplier_name,))
            order_stats = cursor.fetchone()
            
            suppliers.append({
                'name': supplier_name,
                'part_count': stock_stats['part_count'] if isinstance(stock_stats, dict) else stock_stats[0],
                'order_count': order_stats['order_count'] if isinstance(order_stats, dict) else order_stats[0],
                'pending_count': order_stats['pending_count'] if isinstance(order_stats, dict) else order_stats[1]
            })
        
        conn.close()
        
        return jsonify({
            'success': True,
            'suppliers': suppliers
        })
    
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@order_bp.route('/create_orders')
@login_required_order
def create_orders_page():
    """Tedarikçiye göre sipariş oluşturma sayfası"""
    supplier = request.args.get('supplier')
    
    if not supplier:
        # Tedarikçi seçim sayfasına yönlendir
        return redirect('/order_system/select_supplier')
    
    return render_template('order_system/create_orders.html', 
                          username=session.get('username'),
                          selected_supplier=supplier)


@order_bp.route('/create_orders_fast')
@login_required_order
def create_orders_fast_page():
    """Mevcut siparişleri hızlıca kopyalama sayfası"""
    return render_template('order_system/create_orders_fast.html', username=session.get('username'))


@order_bp.route('/api/get_orders', methods=['GET'])
@login_required_order
def get_orders():
    """Sipariş listesini filtrelerek getir (status, supplier vs)"""
    try:
        status = request.args.get('status', 'Gelmedi')  # Varsayılan: Gelmedi
        supplier = request.args.get('supplier')
        part_code = request.args.get('part_code')
        
        conn = get_order_db()
        cursor = conn.cursor()
        
        # Temel sorgu
        query = "SELECT * FROM order_list WHERE 1=1"
        params = []
        
        # Filtreler
        if status:
            query += " AND status = %s"
            params.append(status)
        
        if supplier:
            query += " AND supplier = %s"
            params.append(supplier)
        
        if part_code:
            query += " AND part_code LIKE %s"
            params.append(f"%{part_code}%")
        
        query += " ORDER BY order_date DESC"
        
        cursor.execute(query, params)
        orders = cursor.fetchall()
        
        conn.close()
        
        # Sonuçları JSON'a çevir
        results = []
        for order in orders:
            results.append({
                'id': order['id'],
                'part_code': order['part_code'],
                'part_name': order['part_name'],
                'supplier': order['supplier'],
                'ordered_quantity': order['ordered_quantity'],
                'received_quantity': order.get('received_quantity', 0),
                'unit_price': order.get('unit_price', 0),
                'total_price': order.get('total_price', 0),
                'currency': order.get('currency', 'EUR'),
                'status': order['status'],
                'order_date': order['order_date'].strftime('%d.%m.%Y') if order.get('order_date') else '-',
                'order_type': order.get('order_type', 'Manuel')
            })
        
        return jsonify({
            'success': True,
            'orders': results,
            'count': len(results)
        })
        
    except Exception as e:
        logging.error(f"Get orders error: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500


@order_bp.route('/api/check_critical_stock', methods=['GET'])
@login_required_order
def check_critical_stock():
    """Kritik stok altındaki parçaları listele (isteğe bağlı olarak tedarikçiye göre filtrele)"""
    try:
        # Query parametrelerinden supplier al
        supplier = request.args.get('supplier')
        
        conn = get_order_db()
        cursor = conn.cursor()
        
        # Debug: Kaç parça var?
        cursor.execute("SELECT COUNT(*) as cnt FROM order_system_stock")
        total = cursor.fetchone()['cnt']
        print(f"[DEBUG] order_system_stock'ta {total} parça var")
        
        # Debug: Kaç tanesi kritik stok tanımlı?
        cursor.execute("SELECT COUNT(*) as cnt FROM order_system_stock WHERE critical_stock_level IS NOT NULL AND critical_stock_level > 0")
        with_critical = cursor.fetchone()['cnt']
        print(f"[DEBUG] Kritik stok tanımlı: {with_critical} parça")
        
        # Base query
        query = """
            SELECT s.part_code, s.part_name, s.stock_quantity, s.critical_stock_level,
                   s.expected_stock, s.supplier, s.unit_price,
                   (s.expected_stock - s.stock_quantity) as needed_quantity,
                   p.part_code IS NOT NULL as is_protected
            FROM order_system_stock s
            LEFT JOIN protected_parts p ON s.part_code = p.part_code
            WHERE s.critical_stock_level IS NOT NULL
              AND s.critical_stock_level > 0
              AND s.stock_quantity <= s.critical_stock_level
        """
        
        params = []
        
        # Eğer tedarikçi filtresi varsa ekle
        if supplier:
            query += " AND s.supplier = %s"
            params.append(supplier)
            print(f"[DEBUG] Supplier filtresi uygulanıyor: {supplier}")
        
        query += """
            ORDER BY 
                CASE WHEN p.part_code IS NULL THEN 0 ELSE 1 END,
                (s.expected_stock - s.stock_quantity) DESC
        """
        
        cursor.execute(query, params)
        
        critical_parts = cursor.fetchall()
        print(f"[DEBUG] Kritik stok altında: {len(critical_parts)} parça (supplier: {supplier})")
        
        cursor.close()
        conn.close()
        
        return jsonify({
            'success': True,
            'critical_parts': critical_parts,
            'total_count': len(critical_parts)
        })
        
    except Exception as e:
        print(f"[ERROR] check_critical_stock: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500


@order_bp.route('/api/create_automatic_orders', methods=['POST'])
@login_required_order
def create_automatic_orders():
    """Kritik stok altındaki parçalar için otomatik sipariş oluştur"""
    try:
        # JSON body'den sipariş listesi adı ve seçili parçaları al
        data = request.get_json() or {}
        order_list_name = data.get('order_list_name', '').strip()
        selected_parts = data.get('selected_parts', [])
        
        if not order_list_name:
            return jsonify({'success': False, 'error': 'Sipariş listesi adı gereklidir'}), 400
        
        conn = get_order_db()
        cursor = conn.cursor()
        
        # Seçili parçaları veritabanından al
        if selected_parts:
            placeholders = ','.join(['%s'] * len(selected_parts))
            cursor.execute(f"""
                SELECT s.part_code, s.part_name, s.stock_quantity, s.critical_stock_level,
                       s.expected_stock, s.supplier, s.unit_price
                FROM order_system_stock s
                LEFT JOIN protected_parts p ON s.part_code = p.part_code
                WHERE s.part_code IN ({placeholders})
                  AND p.part_code IS NULL
                  AND s.critical_stock_level > 0
                  AND s.expected_stock > 0
            """, selected_parts)
        else:
            # Hiç seçili yoksa tüm kritik parçaları al
            cursor.execute("""
                SELECT s.part_code, s.part_name, s.stock_quantity, s.critical_stock_level,
                       s.expected_stock, s.supplier, s.unit_price
                FROM order_system_stock s
                LEFT JOIN protected_parts p ON s.part_code = p.part_code
                WHERE s.stock_quantity <= s.critical_stock_level
                  AND p.part_code IS NULL
                  AND s.critical_stock_level > 0
                  AND s.expected_stock > 0
            """)
        
        critical_parts = cursor.fetchall()
        new_orders = []
        
        for part in critical_parts:
            expected = part.get('expected_stock') or 0
            current_stock = part.get('stock_quantity') or 0
            order_quantity = expected - current_stock
            
            # Sipariş miktarı pozitif olmalı
            if order_quantity <= 0:
                continue
            
            total_price = order_quantity * float(part['unit_price'] or 0)
            
            # Sipariş listesi adını notes alanına ekle
            cursor.execute("""
                INSERT INTO order_list 
                (part_code, part_name, supplier, ordered_quantity, unit_price, total_price, 
                 order_type, created_by, status, notes)
                VALUES (%s, %s, %s, %s, %s, %s, 'Otomatik', %s, 'Gelmedi', %s)
            """, (part['part_code'], part['part_name'], part['supplier'], 
                  order_quantity, part['unit_price'], total_price, 
                  session.get('username'), order_list_name))
            
            order_id = cursor.lastrowid
            
            # Protected parts'a ekle
            cursor.execute("""
                INSERT INTO protected_parts (part_code, order_id, reason)
                VALUES (%s, %s, %s)
            """, (part['part_code'], order_id, f'Sipariş listesi: {order_list_name}'))
            
            new_orders.append({
                'part_code': part['part_code'],
                'part_name': part['part_name'],
                'order_quantity': order_quantity,
                'total_price': total_price
            })
        
        conn.commit()
        
        # Log
        log_order_action(
            order_id=None,
            part_code='AUTO_CREATE',
            action='AUTO_ORDERS_CREATED',
            notes=f"'{order_list_name}' - {len(new_orders)} sipariş oluşturuldu",
            action_by=session.get('username')
        )
        
        cursor.close()
        conn.close()
        
        return jsonify({
            'success': True,
            'orders_created': len(new_orders),
            'order_list_name': order_list_name,
            'orders': new_orders
        })
        
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@order_bp.route('/api/get_all_parts')
@login_required_order
def get_all_parts():
    """Tüm parçaları getir (manuel ekleme için)"""
    try:
        conn = get_order_db()
        cursor = conn.cursor()
        
        cursor.execute("""
            SELECT part_code, part_name, supplier, unit_price
            FROM order_system_stock
            ORDER BY part_code
        """)
        
        parts = cursor.fetchall()
        cursor.close()
        conn.close()
        
        return jsonify({
            'success': True,
            'parts': parts
        })
        
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@order_bp.route('/api/add_manual_orders', methods=['POST'])
@login_required_order
def add_manual_orders():
    """Manuel olarak eklenen siparişleri kaydet (mevcut listeye veya yeni liste)"""
    try:
        data = request.get_json()
        list_name = data.get('list_name', '').strip()
        supplier = data.get('supplier', '').strip()
        items = data.get('items', [])
        
        if not list_name or not items:
            return jsonify({'success': False, 'error': 'Liste adı ve parçalar gereklidir'}), 400
        
        conn = get_order_db()
        cursor = conn.cursor()
        
        orders_created = 0
        
        for item in items:
            part_code = item.get('part_code')
            part_name = item.get('part_name')
            supplier = item.get('supplier', supplier)
            quantity = int(item.get('quantity', 0))
            unit_price = float(item.get('unit_price', 0))
            total_price = float(item.get('total_price', 0))
            
            if not part_code or quantity <= 0:
                continue
            
            # Siparişi ekle
            cursor.execute("""
                INSERT INTO order_list 
                (part_code, part_name, supplier, ordered_quantity, unit_price, total_price,
                 order_type, created_by, status, notes)
                VALUES (%s, %s, %s, %s, %s, %s, 'Manuel', %s, 'Gelmedi', %s)
            """, (part_code, part_name, supplier, quantity, unit_price, total_price,
                  session.get('username'), list_name))
            
            order_id = cursor.lastrowid
            
            # Protected parts'a ekle
            cursor.execute("""
                INSERT INTO protected_parts (part_code, order_id, reason)
                VALUES (%s, %s, %s)
                ON DUPLICATE KEY UPDATE order_id = VALUES(order_id), reason = VALUES(reason)
            """, (part_code, order_id, f'Sipariş listesi: {list_name}'))
            
            orders_created += 1
        
        conn.commit()
        
        # Log
        log_order_action(
            order_id=None,
            part_code='MANUAL_BULK',
            action='MANUAL_ORDERS_ADDED',
            notes=f"'{list_name}' - {orders_created} sipariş eklendi",
            action_by=session.get('username')
        )
        
        cursor.close()
        conn.close()
        
        return jsonify({
            'success': True,
            'orders_created': orders_created,
            'list_name': list_name
        })
        
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


# ============================================================================
# 4. MANUEL SİPARİŞ EKLEME MODÜLÜ
# ============================================================================

@order_bp.route('/manual_order')
@login_required_order
def manual_order_page():
    """Manuel sipariş ekleme sayfası"""
    return render_template('order_system/manual_order.html', username=session.get('username'))


@order_bp.route('/api/get_part_info/<part_code>', methods=['GET'])
@login_required_order
def get_part_info(part_code):
    """Parça bilgilerini getir (manuel sipariş için)"""
    try:
        conn = get_order_db()
        cursor = conn.cursor()
        
        cursor.execute("""
            SELECT part_code, part_name, supplier, unit_price, stock_quantity, critical_stock_level
            FROM order_system_stock
            WHERE part_code = %s
        """, (part_code,))
        
        part = cursor.fetchone()
        
        cursor.close()
        conn.close()
        
        if part:
            return jsonify({'success': True, 'part': part})
        else:
            return jsonify({'success': False, 'error': 'Parça bulunamadı'}), 404
            
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@order_bp.route('/api/add_manual_order', methods=['POST'])
@login_required_order
def add_manual_order():
    """Manuel sipariş ekle"""
    try:
        data = request.get_json()
        part_code = data.get('part_code')
        quantity = int(data.get('quantity', 0))
        list_name = data.get('list_name', 'Manuel')  # ✅ Liste adını al
        notes = data.get('notes', '')
        
        if not part_code or quantity <= 0:
            return jsonify({'success': False, 'error': 'Geçersiz veri'}), 400
        
        conn = get_order_db()
        cursor = conn.cursor()
        
        # Parça bilgilerini çek
        cursor.execute("""
            SELECT part_name, supplier, unit_price
            FROM order_system_stock
            WHERE part_code = %s
        """, (part_code,))
        
        part = cursor.fetchone()
        
        if not part:
            cursor.close()
            conn.close()
            return jsonify({'success': False, 'error': 'Parça bulunamadı'}), 404
        
        total_price = quantity * float(part['unit_price'] or 0)
        
        # Sipariş ekle
        cursor.execute("""
            INSERT INTO order_list 
            (part_code, part_name, supplier, ordered_quantity, unit_price, total_price,
             order_type, created_by, status, notes)
            VALUES (%s, %s, %s, %s, %s, %s, 'Manuel', %s, 'Gelmedi', %s)
        """, (part_code, part['part_name'], part['supplier'], quantity,
              part['unit_price'], total_price, session.get('username'), list_name))  # ✅ list_name kullan
        
        order_id = cursor.lastrowid
        
        # Protected parts'a ekle (manuel sipariş de bekleyen listesinde)
        cursor.execute("""
            INSERT INTO protected_parts (part_code, order_id, reason)
            VALUES (%s, %s, 'Manuel sipariş - Beklemede')
        """, (part_code, order_id))
        
        conn.commit()
        
        # Log
        log_order_action(
            order_id=order_id,
            part_code=part_code,
            action='MANUAL_ORDER_CREATED',
            notes=f"Manuel sipariş: {quantity} adet",
            action_by=session.get('username')
        )
        
        cursor.close()
        conn.close()
        
        return jsonify({
            'success': True,
            'order_id': order_id,
            'part_name': part['part_name'],
            'total_price': total_price
        })
        
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


# ============================================================================
# 5. SİPARİŞ TAKİP MODÜLÜ
# ============================================================================

@order_bp.route('/track_orders')
@login_required_order
def track_orders_page():
    """Sipariş takip sayfası"""
    return render_template('order_system/track_orders.html', username=session.get('username'))


@order_bp.route('/api/search_orders', methods=['GET'])
@login_required_order
def search_orders():
    """Parça kodundan sipariş arama"""
    try:
        part_code = request.args.get('part_code', '').strip()
        
        if not part_code:
            return jsonify({'success': False, 'error': 'Parça kodu gereklidir'}), 400
        
        conn = get_order_db()
        cursor = conn.cursor()
        
        # Sipariş listesinde ara
        cursor.execute("""
            SELECT id, part_code, part_name, ordered_quantity, status, supplier, unit_price,
                   received_quantity, order_date, notes
            FROM order_list
            WHERE part_code LIKE %s OR part_name LIKE %s
            ORDER BY order_date DESC
            LIMIT 50
        """, (f'%{part_code}%', f'%{part_code}%'))
        
        orders = cursor.fetchall()
        
        # Sipariş sisteminde yoksa stok tablosundan bilgi al
        if not orders:
            cursor.execute("""
                SELECT id, part_code, part_name, status, supplier, unit_price
                FROM order_system_stock
                WHERE part_code LIKE %s OR part_name LIKE %s
                LIMIT 20
            """, (f'%{part_code}%', f'%{part_code}%'))
            
            stock_parts = cursor.fetchall()
            orders = stock_parts
        
        cursor.close()
        conn.close()
        
        # Sonuçları formatla
        formatted_orders = []
        for order in orders:
            if isinstance(order, dict):
                formatted_orders.append({
                    'id': order.get('id'),
                    'part_code': order.get('part_code'),
                    'part_name': order.get('part_name'),
                    'ordered_quantity': order.get('ordered_quantity', 0),
                    'arrived_quantity': order.get('received_quantity', 0),  # received_quantity -> arrived_quantity
                    'status': order.get('status', 'Beklemede'),
                    'supplier': order.get('supplier', '-'),
                    'unit_price': order.get('unit_price')
                })
            else:  # Tuple response from cursor
                formatted_orders.append({
                    'id': order[0],
                    'part_code': order[1],
                    'part_name': order[2],
                    'ordered_quantity': order[3],
                    'status': order[4],
                    'supplier': order[5],
                    'unit_price': float(order[6]) if order[6] else 0,
                    'arrived_quantity': order[7] or 0,  # received_quantity (index 7)
                })
        
        return jsonify({
            'success': True,
            'orders': formatted_orders
        })
    
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@order_bp.route('/api/get_order_lists', methods=['GET'])
@login_required_order
def get_order_lists():
    """Sipariş listelerini grupla ve getir (notes alanına göre)"""
    try:
        conn = get_order_db()
        cursor = conn.cursor()
        
        # received_quantity kolonunu ekle (yoksa)
        try:
            cursor.execute("""
                SELECT COUNT(*) AS cnt FROM information_schema.COLUMNS
                WHERE TABLE_SCHEMA = %s AND TABLE_NAME = 'order_list' AND COLUMN_NAME = 'received_quantity'
            """, (DB_CONFIG['database'],))
            res = cursor.fetchone()
            if not res or res.get('cnt', 0) == 0:
                cursor.execute("ALTER TABLE order_list ADD COLUMN received_quantity INT DEFAULT 0 AFTER ordered_quantity")
                conn.commit()
                print("[ORDER SYSTEM] received_quantity kolonu eklendi")
        except Exception as col_error:
            print(f"[ORDER SYSTEM] Kolon kontrolü hatası (normal olabilir): {col_error}")
        
        # Teslimat geçmişi tablosu oluştur (yoksa)
        try:
            cursor.execute("""
                CREATE TABLE IF NOT EXISTS delivery_history (
                    id INT AUTO_INCREMENT PRIMARY KEY,
                    order_id INT NOT NULL,
                    part_code VARCHAR(100) NOT NULL,
                    part_name VARCHAR(255),
                    received_quantity INT NOT NULL,
                    delivery_date TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                    received_by VARCHAR(100),
                    notes TEXT,
                    INDEX idx_order_id (order_id),
                    INDEX idx_part_code (part_code),
                    INDEX idx_delivery_date (delivery_date)
                ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4 COLLATE=utf8mb4_unicode_ci
            """)
            conn.commit()
            print("[ORDER SYSTEM] delivery_history tablosu kontrol edildi")
        except Exception as table_error:
            print(f"[ORDER SYSTEM] Tablo kontrolü hatası: {table_error}")
        
        # Tüm siparişleri çek
        cursor.execute("""
            SELECT id, part_code, part_name, supplier, ordered_quantity, 
                   COALESCE(received_quantity, 0) as received_quantity,
                   unit_price, total_price, currency, order_date, status, order_type, 
                   created_by, notes
            FROM order_list
            WHERE status != 'Geldi'
            ORDER BY supplier, COALESCE(notes, 'ZZZ'), order_date DESC
        """)
        
        orders = cursor.fetchall()
        print(f"[ORDER SYSTEM] {len(orders)} sipariş bulundu")
        
        # Tedarikçi ve sipariş listesi adına göre iç içe grupla
        order_lists = {}
        for order in orders:
            supplier = order.get('supplier') or 'Tedarikçi Belirtilmemiş'
            list_name = order.get('notes') or 'İsimsiz Liste'
            
            if supplier not in order_lists:
                order_lists[supplier] = {}
            
            if list_name not in order_lists[supplier]:
                order_lists[supplier][list_name] = {
                    'orders': [],
                    'order_date': None,
                    'total_amount': 0
                }
            
            # İlk siparişin tarihini al (en yeni)
            if order_lists[supplier][list_name]['order_date'] is None and order.get('order_date'):
                order_lists[supplier][list_name]['order_date'] = order.get('order_date').strftime('%d.%m.%Y') if order.get('order_date') else None
            
            # Toplam tutarı hesapla
            order_lists[supplier][list_name]['total_amount'] += float(order.get('total_price') or 0)
            
            order_lists[supplier][list_name]['orders'].append(order)
        
        print(f"[ORDER SYSTEM] {len(order_lists)} tedarikçi, {sum(len(v) for v in order_lists.values())} liste grubu oluşturuldu")
        
        cursor.close()
        conn.close()
        
        return jsonify({
            'success': True,
            'order_lists': order_lists,
            'user_role': session.get('role')
        })
        
    except Exception as e:
        print(f"[ORDER SYSTEM] get_order_lists HATA: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500


@order_bp.route('/api/partial_receive', methods=['POST'])
@login_required_order
def partial_receive():
    """Kısmi teslimat işle"""
    try:
        data = request.get_json()
        order_id = data.get('order_id')
        received_qty = data.get('received_quantity')
        
        if not order_id or not received_qty or received_qty <= 0:
            return jsonify({'success': False, 'error': 'Geçersiz parametreler'}), 400
        
        conn = get_order_db()
        cursor = conn.cursor()
        
        # Sipariş bilgilerini çek
        cursor.execute("""
            SELECT part_code, part_name, ordered_quantity, received_quantity
            FROM order_list
            WHERE id = %s
        """, (order_id,))
        
        order = cursor.fetchone()
        
        if not order:
            cursor.close()
            conn.close()
            return jsonify({'success': False, 'error': 'Sipariş bulunamadı'}), 404
        
        current_received = order['received_quantity'] or 0
        new_received_total = current_received + received_qty
        remaining = order['ordered_quantity'] - new_received_total
        
        if new_received_total > order['ordered_quantity']:
            cursor.close()
            conn.close()
            return jsonify({'success': False, 'error': 'Gelen miktar sipariş miktarını aşıyor'}), 400
        
        # Stoku artır
        cursor.execute("""
            UPDATE order_system_stock
            SET stock_quantity = stock_quantity + %s,
                last_updated = NOW()
            WHERE part_code = %s
        """, (received_qty, order['part_code']))
        
        # Siparişteki alınan miktarı güncelle
        cursor.execute("""
            UPDATE order_list
            SET received_quantity = %s,
                status = IF(%s >= ordered_quantity, 'Geldi', 'Gelmedi'),
                status_updated_date = NOW()
            WHERE id = %s
        """, (new_received_total, new_received_total, order_id))
        
        # Teslimat geçmişine kaydet
        cursor.execute("""
            INSERT INTO delivery_history 
            (order_id, part_code, part_name, received_quantity, received_by, notes)
            VALUES (%s, %s, %s, %s, %s, %s)
        """, (order_id, order['part_code'], order['part_name'], 
              received_qty, session.get('username'), 
              f"Kısmi teslimat ({new_received_total}/{order['ordered_quantity']})"))
        
        # Eğer tam teslim olduysa protected_parts'tan kaldır
        if remaining <= 0:
            cursor.execute("DELETE FROM protected_parts WHERE order_id = %s", (order_id,))
        
        conn.commit()
        
        # Log
        log_order_action(
            order_id=order_id,
            part_code=order['part_code'],
            action='PARTIAL_RECEIVE',
            notes=f"Kısmi teslimat: {received_qty} adet ({new_received_total}/{order['ordered_quantity']}) - {session.get('username')}",
            action_by=session.get('username')
        )
        
        cursor.close()
        conn.close()
        
        return jsonify({
            'success': True,
            'received_total': new_received_total,
            'remaining': remaining,
            'completed': remaining <= 0
        })
        
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@order_bp.route('/api/get_delivery_history/<int:order_id>', methods=['GET'])
@login_required_order
def get_delivery_history(order_id):
    """Sipariş için teslimat geçmişini getir"""
    try:
        conn = get_order_db()
        cursor = conn.cursor()
        
        cursor.execute("""
            SELECT id, received_quantity, delivery_date, received_by, notes
            FROM delivery_history
            WHERE order_id = %s
            ORDER BY delivery_date DESC
        """, (order_id,))
        
        history = cursor.fetchall()
        
        cursor.close()
        conn.close()
        
        return jsonify({
            'success': True,
            'history': history
        })
        
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@order_bp.route('/api/update_order_status', methods=['POST'])
@login_required_order
def update_order_status():
    """Sipariş durumunu güncelle - durum, gelen miktar ve notlar"""
    try:
        data = request.get_json() or {}
        order_id = data.get('order_id')
        new_status = data.get('status')  # 'Geldi', 'Beklemede', 'Kısmi'
        arrived_quantity = data.get('arrived_quantity', 0)
        notes = data.get('notes', '')
        
        if not order_id or not new_status:
            return jsonify({'success': False, 'error': 'Eksik parametreler'}), 400
        
        conn = get_order_db()
        cursor = conn.cursor()
        
        # Sipariş bilgilerini çek
        cursor.execute("""
            SELECT part_code, part_name, ordered_quantity, status, received_quantity
            FROM order_list
            WHERE id = %s
        """, (order_id,))
        
        order = cursor.fetchone()
        
        if not order:
            cursor.close()
            conn.close()
            return jsonify({'success': False, 'error': 'Sipariş bulunamadı'}), 404
        
        old_status = order['status'] if isinstance(order, dict) else order[3]
        part_code = order['part_code'] if isinstance(order, dict) else order[0]
        part_name = order['part_name'] if isinstance(order, dict) else order[1]
        ordered_qty = order['ordered_quantity'] if isinstance(order, dict) else order[2]
        
        # Gelen miktar kontrol
        if arrived_quantity > ordered_qty:
            arrived_quantity = ordered_qty
        
        # Sipariş güncelle
        cursor.execute("""
            UPDATE order_list
            SET status = %s,
                received_quantity = %s,
                notes = %s,
                status_updated_date = NOW()
            WHERE id = %s
        """, (new_status, arrived_quantity, notes, order_id))
        
        # Eğer "Geldi" ise stoku artır
        if new_status == 'Geldi':
            cursor.execute("""
                UPDATE order_system_stock
                SET stock_quantity = stock_quantity + %s,
                    last_updated = NOW()
                WHERE part_code = %s
            """, (arrived_quantity, part_code))
            
            # Protected parts'tan kaldır
            cursor.execute("DELETE FROM protected_parts WHERE order_id = %s", (order_id,))
        
        elif new_status == 'Kısmi':
            # Kısmi gelen miktarı stoka ekle
            if arrived_quantity > 0:
                cursor.execute("""
                    UPDATE order_system_stock
                    SET stock_quantity = stock_quantity + %s,
                        last_updated = NOW()
                    WHERE part_code = %s
                """, (arrived_quantity, part_code))
        
        conn.commit()
        
        # Log
        log_order_action(
            order_id=order_id,
            part_code=part_code,
            action='STATUS_UPDATE',
            old_status=old_status,
            new_status=new_status,
            notes=f"Durum: {old_status} → {new_status}, Gelen: {arrived_quantity}/{ordered_qty}, Not: {notes}",
            action_by=session.get('username')
        )
        
        cursor.close()
        conn.close()
        
        return jsonify({
            'success': True,
            'message': f'Sipariş başarıyla güncellendi: {new_status}'
        })
        
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500
        
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


# ============================================================================
# 6. SİPARİŞ GEÇMİŞİ VE FİLTRELEME
# ============================================================================

@order_bp.route('/api/search_parts', methods=['GET'])
@login_required_order
def search_parts():
    """Parça arama (autocomplete için)"""
    try:
        search_term = request.args.get('q', '').strip()
        
        if len(search_term) < 2:
            return jsonify({'success': True, 'parts': []})
        
        conn = get_order_db()
        cursor = conn.cursor()
        
        cursor.execute("""
            SELECT part_code, part_name, stock_quantity, critical_stock_level, supplier
            FROM order_system_stock
            WHERE part_code LIKE %s OR part_name LIKE %s
            LIMIT 20
        """, (f'%{search_term}%', f'%{search_term}%'))
        
        parts = cursor.fetchall()
        
        cursor.close()
        conn.close()
        
        return jsonify({'success': True, 'parts': parts})
        
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@order_bp.route('/api/export_order_to_excel', methods=['POST'])
@login_required_order
def export_order_to_excel():
    """Sipariş listesini Excel'e aktar ve sunucuya kaydet"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        import os
        
        data = request.get_json()
        supplier = data.get('supplier')
        list_name = data.get('list_name')
        
        if not supplier or not list_name:
            return jsonify({'success': False, 'error': 'Tedarikçi ve liste adı gerekli'}), 400
        
        conn = get_order_db()
        cursor = conn.cursor()
        
        # Siparişleri çek
        cursor.execute("""
            SELECT part_code, part_name, ordered_quantity, 
                   COALESCE(received_quantity, 0) as received_quantity,
                   unit_price, total_price, currency, order_date
            FROM order_list
            WHERE supplier = %s AND notes = %s AND status != 'Geldi'
            ORDER BY part_code
        """, (supplier, list_name))
        
        orders = cursor.fetchall()
        cursor.close()
        conn.close()
        
        if not orders:
            return jsonify({'success': False, 'error': 'Sipariş bulunamadı'}), 404
        
        # Excel oluştur
        wb = Workbook()
        ws = wb.active
        ws.title = "Sipariş Listesi"
        
        # Başlık stilleri
        header_fill = PatternFill(start_color="DC2626", end_color="DC2626", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Başlıklar
        headers = ['Parça Kodu', 'Parça Adı', 'Sipariş Miktarı', 'Gelen', 'Kalan', 'Birim Fiyat', 'Toplam', 'Para Birimi', 'Sipariş Tarihi']
        ws.append(headers)
        
        # Başlık stili uygula
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
            cell.border = border
        
        # Verileri ekle
        for order in orders:
            remaining = order['ordered_quantity'] - order['received_quantity']
            ws.append([
                order['part_code'],
                order['part_name'] or '-',
                order['ordered_quantity'],
                order['received_quantity'],
                remaining,
                order['unit_price'] or 0,
                order['total_price'] or 0,
                order['currency'] or 'EUR',
                order['order_date'].strftime('%d.%m.%Y') if order['order_date'] else '-'
            ])
        
        # Kolon genişlikleri
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 30
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 10
        ws.column_dimensions['E'].width = 10
        ws.column_dimensions['F'].width = 12
        ws.column_dimensions['G'].width = 12
        ws.column_dimensions['H'].width = 12
        ws.column_dimensions['I'].width = 15
        
        # Sunucu yoluna kaydet
        server_folder = r"\\DCSRV\tahsinortak\CermakDepo\CermakEnvanter\static\excel"
        safe_supplier = "".join(c for c in supplier if c.isalnum() or c in (' ', '-', '_')).strip()
        safe_list_name = "".join(c for c in list_name if c.isalnum() or c in (' ', '-', '_')).strip()
        filename = f"{safe_supplier}_{safe_list_name}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        server_path = os.path.join(server_folder, filename)
        
        # Klasörü oluştur (yoksa)
        os.makedirs(server_folder, exist_ok=True)
        
        # Kaydet
        wb.save(server_path)
        
        # Kullanıcıya indir
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        from flask import send_file
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
        
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500


@order_bp.route('/api/delete_order_list', methods=['POST'])
@login_required_order
def delete_order_list():
    """Sipariş listesini sil (sadece admin)"""
    try:
        print(f"[DELETE] User role: {session.get('role')}")
        print(f"[DELETE] User: {session.get('username')}")
        
        # Yetki kontrolü
        if session.get('role') != 'admin':
            print(f"[DELETE] YETKİ HATASI: Role = {session.get('role')}")
            return jsonify({'success': False, 'error': 'Bu işlem için yetkiniz yok'}), 403
        
        data = request.get_json()
        supplier = data.get('supplier')
        list_name = data.get('list_name')
        
        print(f"[DELETE] Supplier: {supplier}, List: {list_name}")
        
        if not supplier:
            return jsonify({'success': False, 'error': 'Tedarikçi gerekli'}), 400
        
        # "İsimsiz Liste" frontend placeholder'ı, veritabanında NULL olarak saklanıyor
        if list_name == 'İsimsiz Liste':
            list_name = None
        
        conn = get_order_db()
        cursor = conn.cursor()
        
        # Önce kaç sipariş var kontrol et
        if list_name is None:
            cursor.execute("""
                SELECT COUNT(*) as cnt FROM order_list
                WHERE supplier = %s AND (notes IS NULL OR notes = '')
            """, (supplier,))
        else:
            cursor.execute("""
                SELECT COUNT(*) as cnt FROM order_list
                WHERE supplier = %s AND notes = %s
            """, (supplier, list_name))
        count_before = cursor.fetchone()['cnt']
        print(f"[DELETE] Silinecek sipariş sayısı: {count_before}")
        
        # Siparişleri sil
        if list_name is None:
            cursor.execute("""
                DELETE FROM order_list
                WHERE supplier = %s AND (notes IS NULL OR notes = '')
            """, (supplier,))
        else:
            cursor.execute("""
                DELETE FROM order_list
                WHERE supplier = %s AND notes = %s
            """, (supplier, list_name))
        
        deleted_count = cursor.rowcount
        conn.commit()
        
        print(f"[DELETE] Silindi: {deleted_count} sipariş")
        
        cursor.close()
        conn.close()
        
        return jsonify({
            'success': True,
            'message': f'{deleted_count} sipariş silindi',
            'deleted_count': deleted_count
        })
        
    except Exception as e:
        import traceback
        print(f"[DELETE] HATA: {str(e)}")
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500


# ============================================================================
# HELPER FONKSİYONLARI
# ============================================================================

def log_order_action(order_id=None, part_code=None, action=None, old_status=None, 
                      new_status=None, stock_before=None, stock_after=None, 
                      notes=None, action_by=None):
    """Sipariş işlemlerini logla"""
    try:
        conn = get_order_db()
        cursor = conn.cursor()
        
        cursor.execute("""
            INSERT INTO order_history_log 
            (order_id, part_code, action, old_status, new_status, stock_before, 
             stock_after, action_by, notes)
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)
        """, (order_id, part_code, action, old_status, new_status, stock_before,
              stock_after, action_by, notes))
        
        conn.commit()
        cursor.close()
        conn.close()
        
    except Exception as e:
        print(f"Log error: {e}")


# Blueprint'i export et
def register_order_system(app):
    """App'e sipariş sistemi blueprint'ini kaydet"""
    app.register_blueprint(order_bp)
    print("[ORDER SYSTEM] Parça Sipariş Sistemi modülü yüklendi (OK)")
