# -*- coding: utf-8 -*-
"""
TAKEUCHI PARÇA SİPARİŞ MODÜLÜ
Ayrı, izole, basitleştirilmiş sipariş ve kontrol sistemi
Mevcut envanter sistemine DOKUNMAYAN tamamen bağımsız modül
"""

from flask import Flask, jsonify, request
from models import (
    db, 
    TakeuchiPartOrder, 
    TakeuchiOrderItem, 
    TakeuchiTempOrder, 
    TakeuchiTempOrderItem,
    TakeuchiPart,
    User,
    PartCode
)
from datetime import datetime
import uuid
import logging

logger = logging.getLogger(__name__)


class TakeuchiOrderManager:
    """Takeuchi Sipariş Yönetim Sistemi"""
    
    @staticmethod
    def create_temp_order_session(user_id):
        """
        Kullanıcı için yeni geçici sipariş oturumu oluştur
        Aynı anda bir kullanıcının sadece bir aktif temp order'ı olabilir
        """
        try:
            # Önceki açık temp order'ı kontrol et
            existing = TakeuchiTempOrder.query.filter_by(created_by=user_id).first()
            if existing:
                session_id = existing.session_id
                logger.info(f"[TAKEUCHI] User {user_id} existing temp order session: {session_id}")
                return {'success': True, 'session_id': session_id, 'is_new': False}
            
            # Yeni oturum oluştur
            session_id = str(uuid.uuid4())
            temp_order = TakeuchiTempOrder(
                session_id=session_id,
                created_by=user_id
            )
            db.session.add(temp_order)
            db.session.commit()
            
            logger.info(f"[TAKEUCHI] New temp order session created: {session_id}")
            return {'success': True, 'session_id': session_id, 'is_new': True}
            
        except Exception as e:
            logger.error(f"[TAKEUCHI] Error creating temp order session: {e}")
            db.session.rollback()
            return {'success': False, 'error': str(e)}
    
    @staticmethod
    def add_part_to_temp_order(session_id, part_code, quantity, user_id):
        """
        Geçici siparişe parça ekle
        
        Kurallar:
        1. Parça kodu valid mi?
        2. Bu parça için aktif sipariş var mı?
        3. Varsa, tamamen teslim alınmamış mı?
        """
        try:
            # Parça bilgisini al
            part = PartCode.query.filter_by(part_code=part_code).first()
            if not part:
                return {
                    'success': False, 
                    'error': f'Parça kodu bulunamadı: {part_code}'
                }
            
            # Aktif sipariş var mı kontrol et
            active_order = TakeuchiOrderItem.query.join(
                TakeuchiPartOrder,
                TakeuchiOrderItem.order_id == TakeuchiPartOrder.id
            ).filter(
                TakeuchiOrderItem.part_code == part_code,
                TakeuchiPartOrder.status == 'pending'
            ).first()
            
            if active_order:
                return {
                    'success': False,
                    'warning': True,
                    'message': f'Bu parça için henüz tamamlanmamış bir sipariş bulunmaktadır. (Sipariş: {active_order.order.order_code})',
                    'active_order_code': active_order.order.order_code
                }
            
            # Geçici siparişe ekle
            temp_order = TakeuchiTempOrder.query.filter_by(session_id=session_id).first()
            if not temp_order:
                return {'success': False, 'error': 'Geçici sipariş oturumu bulunamadı'}
            
            # Aynı parça zaten var mı?
            existing_item = TakeuchiTempOrderItem.query.filter_by(
                temp_order_id=temp_order.id,
                part_code=part_code
            ).first()
            
            if existing_item:
                # Miktarı güncelle
                existing_item.quantity += quantity
                logger.info(f"[TAKEUCHI] Updated quantity for {part_code} in session {session_id}")
            else:
                # Yeni item ekle
                item = TakeuchiTempOrderItem(
                    temp_order_id=temp_order.id,
                    part_code=part_code,
                    part_name=part.part_name,
                    quantity=quantity
                )
                db.session.add(item)
                logger.info(f"[TAKEUCHI] Added {part_code} (qty: {quantity}) to session {session_id}")
            
            db.session.commit()
            
            return {
                'success': True,
                'part_code': part_code,
                'part_name': part.part_name,
                'quantity': quantity,
                'message': f'{part.part_name} ({quantity} adet) listeye eklendi'
            }
            
        except Exception as e:
            logger.error(f"[TAKEUCHI] Error adding part: {e}")
            db.session.rollback()
            return {'success': False, 'error': str(e)}
    
    @staticmethod
    def get_part_history(part_code):
        """Parçanın sipariş geçmişini al"""
        try:
            orders = db.session.query(TakeuchiOrderItem, TakeuchiPartOrder).join(
                TakeuchiPartOrder,
                TakeuchiOrderItem.order_id == TakeuchiPartOrder.id
            ).filter(
                TakeuchiOrderItem.part_code == part_code
            ).order_by(
                TakeuchiPartOrder.created_at.desc()
            ).all()
            
            history = []
            for item, order in orders:
                history.append({
                    'order_code': order.order_code,
                    'ordered_quantity': item.ordered_quantity,
                    'received_quantity': item.received_quantity,
                    'status': item.status,
                    'created_at': order.created_at.isoformat(),
                    'completed_at': order.completed_at.isoformat() if order.completed_at else None
                })
            
            return {'success': True, 'history': history}
            
        except Exception as e:
            logger.error(f"[TAKEUCHI] Error getting part history: {e}")
            return {'success': False, 'error': str(e)}
    
    @staticmethod
    def get_temp_order_items(session_id):
        """Geçici siparişin kalemleri al"""
        try:
            temp_order = TakeuchiTempOrder.query.filter_by(session_id=session_id).first()
            if not temp_order:
                return {'success': False, 'error': 'Geçici sipariş bulunamadı'}
            
            items = []
            for item in temp_order.items:
                items.append({
                    'id': item.id,
                    'part_code': item.part_code,
                    'part_name': item.part_name,
                    'quantity': item.quantity,
                    'added_at': item.added_at.isoformat()
                })
            
            return {
                'success': True,
                'session_id': session_id,
                'items': items,
                'item_count': len(items),
                'created_at': temp_order.created_at.isoformat()
            }
            
        except Exception as e:
            logger.error(f"[TAKEUCHI] Error getting temp order items: {e}")
            return {'success': False, 'error': str(e)}
    
    @staticmethod
    def remove_temp_order_item(item_id):
        """Geçici siparişten parça kaldır"""
        try:
            item = TakeuchiTempOrderItem.query.get(item_id)
            if not item:
                return {'success': False, 'error': 'Parça bulunamadı'}
            
            part_code = item.part_code
            db.session.delete(item)
            db.session.commit()
            
            logger.info(f"[TAKEUCHI] Removed {part_code} from temp order")
            return {'success': True, 'message': f'{part_code} listeden çıkarıldı'}
            
        except Exception as e:
            logger.error(f"[TAKEUCHI] Error removing item: {e}")
            db.session.rollback()
            return {'success': False, 'error': str(e)}
    
    @staticmethod
    def create_official_order(temp_order_id, order_name, user_id):
        """
        Admin: Geçici siparişi resmi sipariş haline dönüştür
        Sipariş kodu otomatik oluştur: CER2025001, CER2025002, vb.
        """
        try:
            temp_order = TakeuchiTempOrder.query.get(temp_order_id)
            if not temp_order:
                return {'success': False, 'error': 'Geçici sipariş bulunamadı'}
            
            if not temp_order.items:
                return {'success': False, 'error': 'Geçici siparişte parça yok'}
            
            # Sipariş kodu oluştur: CER + YIL + Sıra
            year = datetime.utcnow().year
            max_order = TakeuchiPartOrder.query.filter(
                TakeuchiPartOrder.order_code.like(f'CER{year}%')
            ).count()
            order_code = f'CER{year}{str(max_order + 1).zfill(3)}'
            
            # Resmi sipariş oluştur
            official_order = TakeuchiPartOrder(
                order_code=order_code,
                order_name=order_name,
                status='pending',
                created_by=user_id
            )
            db.session.add(official_order)
            db.session.flush()  # ID'yi almak için
            
            # Geçici siparişin kalemleri resmi sipariş kalemleri olarak kopyala
            for temp_item in temp_order.items:
                order_item = TakeuchiOrderItem(
                    order_id=official_order.id,
                    part_code=temp_item.part_code,
                    part_name=temp_item.part_name,
                    ordered_quantity=temp_item.quantity,
                    received_quantity=0,
                    status='pending'
                )
                db.session.add(order_item)
            
            # Geçici siparişi sil
            db.session.delete(temp_order)
            db.session.commit()
            
            logger.info(f"[TAKEUCHI] Official order created: {order_code}")
            
            return {
                'success': True,
                'order_code': order_code,
                'order_name': order_name,
                'item_count': len(temp_order.items),
                'message': f'Sipariş oluşturuldu: {order_code}'
            }
            
        except Exception as e:
            logger.error(f"[TAKEUCHI] Error creating official order: {e}")
            db.session.rollback()
            return {'success': False, 'error': str(e)}
    
    @staticmethod
    def get_all_orders():
        """Tüm Takeuchi siparişlerini listele"""
        try:
            orders = TakeuchiPartOrder.query.order_by(
                TakeuchiPartOrder.created_at.desc()
            ).all()
            
            result = []
            for order in orders:
                items = []
                for item in order.items:
                    items.append({
                        'id': item.id,
                        'part_code': item.part_code,
                        'part_name': item.part_name,
                        'ordered_quantity': item.ordered_quantity,
                        'received_quantity': item.received_quantity,
                        'status': item.status
                    })
                
                result.append({
                    'id': order.id,
                    'order_code': order.order_code,
                    'order_name': order.order_name,
                    'status': order.status,
                    'items': items,
                    'item_count': len(items),
                    'created_at': order.created_at.isoformat(),
                    'completed_at': order.completed_at.isoformat() if order.completed_at else None
                })
            
            return {'success': True, 'orders': result}
            
        except Exception as e:
            logger.error(f"[TAKEUCHI] Error getting orders: {e}")
            return {'success': False, 'error': str(e)}
    
    @staticmethod
    def mark_item_received(item_id, received_quantity):
        """
        Sipariş kalemini teslim alındı olarak işaretle
        
        Kurallar:
        - received_quantity <= ordered_quantity
        - Eğer received_quantity == ordered_quantity ise status = 'completed'
        - Eğer 0 < received_quantity < ordered_quantity ise status = 'partial'
        - Eğer tüm kalemleri tamamlandı ise order.status = 'completed'
        """
        try:
            item = TakeuchiOrderItem.query.get(item_id)
            if not item:
                return {'success': False, 'error': 'Kalem bulunamadı'}
            
            if received_quantity > item.ordered_quantity:
                return {
                    'success': False,
                    'error': f'Teslim miktarı siparişli miktardan fazla (Sipariş: {item.ordered_quantity})'
                }
            
            # Eski durumdan değişim
            old_status = item.status
            item.received_quantity = received_quantity
            
            # Yeni durumu belirle
            if received_quantity == item.ordered_quantity:
                item.status = 'completed'
                if not item.fully_received_at:
                    item.fully_received_at = datetime.utcnow()
            elif received_quantity > 0:
                item.status = 'partial'
                if not item.first_received_at:
                    item.first_received_at = datetime.utcnow()
            else:
                item.status = 'pending'
            
            # İlk teslim tarihini ayarla
            if received_quantity > 0 and not item.first_received_at:
                item.first_received_at = datetime.utcnow()
            
            db.session.commit()
            
            # Sipariş tüm kalemleri tamamlandı mı kontrol et
            order = item.order
            all_items = TakeuchiOrderItem.query.filter_by(order_id=order.id).all()
            
            if all(itm.status == 'completed' for itm in all_items):
                order.status = 'completed'
                order.completed_at = datetime.utcnow()
                db.session.commit()
                logger.info(f"[TAKEUCHI] Order {order.order_code} completed")
            
            logger.info(f"[TAKEUCHI] Item {item.part_code} marked received: {received_quantity}/{item.ordered_quantity}")
            
            return {
                'success': True,
                'part_code': item.part_code,
                'received_quantity': item.received_quantity,
                'ordered_quantity': item.ordered_quantity,
                'status': item.status,
                'order_status': order.status
            }
            
        except Exception as e:
            logger.error(f"[TAKEUCHI] Error marking item received: {e}")
            db.session.rollback()
            return {'success': False, 'error': str(e)}
    
    @staticmethod
    def get_temp_orders_for_admin():
        """Admin için geçici siparişleri listele"""
        try:
            temp_orders = TakeuchiTempOrder.query.all()
            
            result = []
            for temp_order in temp_orders:
                items = []
                for item in temp_order.items:
                    items.append({
                        'id': item.id,
                        'part_code': item.part_code,
                        'part_name': item.part_name,
                        'quantity': item.quantity
                    })
                
                creator = User.query.get(temp_order.created_by)
                creator_name = creator.full_name if creator else 'Bilinmiyor'
                
                result.append({
                    'id': temp_order.id,
                    'session_id': temp_order.session_id,
                    'items': items,
                    'item_count': len(items),
                    'created_by': creator_name,
                    'created_at': temp_order.created_at.isoformat(),
                    'updated_at': temp_order.updated_at.isoformat()
                })
            
            return {'success': True, 'temp_orders': result}
            
        except Exception as e:
            logger.error(f"[TAKEUCHI] Error getting temp orders: {e}")
            return {'success': False, 'error': str(e)}

    @staticmethod
    def import_parts_from_excel(file_content, user_id):
        """
        Excel dosyasından Takeuchi parçalarını içeri aktar
        Sütunlar: Parça Kodu, Parça Adı, Değişen Parça Kodu, Build Out, Geliş Fiyatı
        """
        try:
            import openpyxl
            from io import BytesIO
            
            # Excel dosyasını oku
            workbook = openpyxl.load_workbook(BytesIO(file_content))
            worksheet = workbook.active
            
            imported_count = 0
            error_rows = []
            
            # Header satırını atla (sıra 1)
            for row_idx, row in enumerate(worksheet.iter_rows(min_row=2, values_only=True), start=2):
                try:
                    # Sütunları eşle
                    part_code = str(row[0]).strip() if row[0] else None
                    part_name = str(row[1]).strip() if row[1] else None
                    alternative_code = str(row[2]).strip() if row[2] else None
                    build_out = str(row[3]).strip() if row[3] else None
                    
                    # Geliş fiyatını float'a çevir
                    cost_price = 0.0
                    if row[4]:
                        try:
                            cost_price = float(row[4])
                        except:
                            cost_price = 0.0
                    
                    # Validasyon
                    if not part_code or not part_name:
                        error_rows.append(f"Satır {row_idx}: Parça Kodu ve Adı gerekli")
                        continue
                    
                    # Zaten var mı kontrol et
                    existing = TakeuchiPart.query.filter_by(part_code=part_code).first()
                    if existing:
                        # Güncelle
                        existing.part_name = part_name
                        existing.alternative_code = alternative_code if alternative_code else None
                        existing.build_out = build_out if build_out else None
                        existing.cost_price = cost_price
                        existing.updated_at = datetime.utcnow()
                        logger.info(f"[TAKEUCHI] Part {part_code} güncellendi")
                    else:
                        # Yeni parça ekle
                        new_part = TakeuchiPart(
                            part_code=part_code,
                            part_name=part_name,
                            alternative_code=alternative_code if alternative_code else None,
                            build_out=build_out if build_out else None,
                            cost_price=cost_price,
                            uploaded_by=user_id,
                            is_active=True
                        )
                        db.session.add(new_part)
                        logger.info(f"[TAKEUCHI] Part {part_code} eklendi")
                    
                    imported_count += 1
                    
                except Exception as row_error:
                    error_rows.append(f"Satır {row_idx}: {str(row_error)}")
                    logger.error(f"[TAKEUCHI] Row {row_idx} error: {row_error}")
            
            # Commit
            db.session.commit()
            
            result = {
                'success': True,
                'imported_count': imported_count,
                'error_rows': error_rows,
                'total_rows': worksheet.max_row - 1  # Header hariç
            }
            
            logger.info(f"[TAKEUCHI] Excel import tamamlandı: {imported_count} parça")
            return result
            
        except Exception as e:
            db.session.rollback()
            logger.error(f"[TAKEUCHI] Excel import hatası: {e}")
            return {'success': False, 'error': str(e)}
    
    @staticmethod
    def get_all_takeuchi_parts():
        """Tüm Takeuchi parçalarını getir"""
        try:
            parts = TakeuchiPart.query.filter_by(is_active=True).order_by(TakeuchiPart.part_code).all()
            
            result = []
            for part in parts:
                result.append({
                    'id': part.id,
                    'part_code': part.part_code,
                    'part_name': part.part_name,
                    'alternative_code': part.alternative_code,
                    'build_out': part.build_out,
                    'cost_price': part.cost_price,
                    'created_at': part.created_at.isoformat()
                })
            
            return {'success': True, 'parts': result, 'total': len(result)}
        
        except Exception as e:
            logger.error(f"[TAKEUCHI] Error getting parts: {e}")
            return {'success': False, 'error': str(e)}
